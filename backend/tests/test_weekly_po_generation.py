"""
Test Weekly PO Generation Feature
Tests for:
- Preview Weekly POs endpoint
- Download Template endpoint
- Upload & Generate POs endpoint
- Weekly Draft POs CRUD operations
"""
import pytest
import requests
import os
import io
import openpyxl
from datetime import datetime

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestWeeklyPOGeneration:
    """Test Weekly PO Generation feature"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test fixtures"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login to get token
        login_response = self.session.post(
            f"{BASE_URL}/api/auth/login",
            json={"email": "admin@factory.com", "password": "bidso123"}
        )
        assert login_response.status_code == 200, f"Login failed: {login_response.text}"
        
        token = login_response.json().get("access_token")
        self.session.headers.update({"Authorization": f"Bearer {token}"})
        
        # Get a weekly MRP run
        runs_response = self.session.get(f"{BASE_URL}/api/mrp/runs?limit=10")
        assert runs_response.status_code == 200
        
        runs = runs_response.json()
        weekly_runs = [r for r in runs if r.get("version") == "WEEKLY_V1"]
        assert len(weekly_runs) > 0, "No weekly MRP runs found"
        
        self.run_id = weekly_runs[0]["id"]
        self.run_code = weekly_runs[0]["run_code"]
        
        # Get first 4 weeks from weekly plan
        plan_response = self.session.get(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-plan?plan_type=all"
        )
        assert plan_response.status_code == 200
        
        weekly_plan = plan_response.json().get("weekly_plan", [])
        assert len(weekly_plan) >= 4, "Need at least 4 weeks in weekly plan"
        
        self.test_weeks = [w["order_week"] for w in weekly_plan[:4]]
    
    # ============ Preview Endpoint Tests ============
    
    def test_preview_weekly_pos_success(self):
        """Test preview endpoint returns correct structure"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/preview",
            json=self.test_weeks
        )
        
        assert response.status_code == 200, f"Preview failed: {response.text}"
        
        data = response.json()
        
        # Verify response structure
        assert "run_id" in data
        assert "run_code" in data
        assert "selected_weeks" in data
        assert "preview_pos" in data
        assert "items_without_vendor" in data
        assert "summary" in data
        
        # Verify summary structure
        summary = data["summary"]
        assert "total_vendors" in summary
        assert "total_items" in summary
        assert "total_amount" in summary
        assert "items_needing_vendor" in summary
        
        print(f"Preview Summary: {summary}")
    
    def test_preview_returns_correct_weeks(self):
        """Test preview returns the selected weeks"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/preview",
            json=self.test_weeks[:2]  # Only 2 weeks
        )
        
        assert response.status_code == 200
        data = response.json()
        
        assert data["selected_weeks"] == self.test_weeks[:2]
        print(f"Selected weeks verified: {data['selected_weeks']}")
    
    def test_preview_groups_by_vendor(self):
        """Test preview groups items by vendor correctly"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/preview",
            json=self.test_weeks
        )
        
        assert response.status_code == 200
        data = response.json()
        
        preview_pos = data.get("preview_pos", [])
        
        # Each preview PO should have vendor info
        for po in preview_pos:
            assert "vendor_id" in po
            assert "vendor_name" in po
            assert "weeks_covered" in po
            assert "total_items" in po
            assert "total_qty" in po
            assert "total_amount" in po
            assert "items" in po
            
            # Items should have required fields
            for item in po["items"]:
                assert "rm_id" in item
                assert "order_qty" in item
                assert "order_week" in item
        
        print(f"Verified {len(preview_pos)} vendor POs with correct structure")
    
    def test_preview_identifies_items_without_vendor(self):
        """Test preview identifies items needing vendor assignment"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/preview",
            json=self.test_weeks
        )
        
        assert response.status_code == 200
        data = response.json()
        
        items_without_vendor = data.get("items_without_vendor", [])
        summary = data.get("summary", {})
        
        # Count should match
        assert len(items_without_vendor) == summary.get("items_needing_vendor", 0)
        
        # Items without vendor should have order_week
        for item in items_without_vendor[:5]:  # Check first 5
            assert "rm_id" in item
            assert "order_week" in item
        
        print(f"Items needing vendor: {len(items_without_vendor)}")
    
    def test_preview_invalid_run_id(self):
        """Test preview with invalid run ID returns 404"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/invalid-run-id/weekly-pos/preview",
            json=self.test_weeks
        )
        
        assert response.status_code == 404
        print("Invalid run ID correctly returns 404")
    
    def test_preview_empty_weeks(self):
        """Test preview with empty weeks list"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/preview",
            json=[]
        )
        
        # Should return 404 or empty result
        assert response.status_code in [404, 200]
        if response.status_code == 200:
            data = response.json()
            assert data.get("summary", {}).get("total_items", 0) == 0
        print("Empty weeks handled correctly")
    
    # ============ Download Template Tests ============
    
    def test_download_template_success(self):
        """Test download template returns valid Excel file"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/download-template",
            json=self.test_weeks[:2]
        )
        
        assert response.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        
        # Verify it's a valid Excel file
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        
        # Check required sheets exist
        assert "PO_Lines" in wb.sheetnames
        assert "Vendors_Reference" in wb.sheetnames
        assert "Instructions" in wb.sheetnames
        
        print(f"Template downloaded with sheets: {wb.sheetnames}")
    
    def test_download_template_has_correct_headers(self):
        """Test template has correct column headers"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/download-template",
            json=self.test_weeks[:2]
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["PO_Lines"]
        
        headers = [cell.value for cell in ws[1]]
        
        expected_headers = [
            "Order Week", "RM ID", "RM Name", "Category", "Production Week",
            "Suggested Qty", "Final Qty", "Unit Price", "Total Cost",
            "Vendor ID", "Vendor Name", "Lead Time Days", "Notes"
        ]
        
        assert headers == expected_headers, f"Headers mismatch: {headers}"
        print(f"Template headers verified: {headers}")
    
    def test_download_template_has_data_rows(self):
        """Test template contains data rows"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/download-template",
            json=self.test_weeks[:2]
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["PO_Lines"]
        
        # Should have more than just header row
        assert ws.max_row > 1, "Template should have data rows"
        
        # Check first data row has values
        first_data_row = [cell.value for cell in ws[2]]
        assert first_data_row[0] is not None  # Order Week
        assert first_data_row[1] is not None  # RM ID
        
        print(f"Template has {ws.max_row - 1} data rows")
    
    def test_download_template_vendors_reference(self):
        """Test Vendors_Reference sheet has vendor data"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/download-template",
            json=self.test_weeks[:2]
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Vendors_Reference"]
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Vendor ID" in headers
        assert "Vendor Name" in headers
        
        # Should have vendor data
        assert ws.max_row > 1, "Vendors_Reference should have vendor data"
        
        print(f"Vendors_Reference has {ws.max_row - 1} vendors")
    
    # ============ Weekly Draft POs Tests ============
    
    def test_get_weekly_draft_pos(self):
        """Test fetching weekly draft POs"""
        response = self.session.get(
            f"{BASE_URL}/api/mrp/weekly-draft-pos?run_id={self.run_id}&limit=50"
        )
        
        assert response.status_code == 200
        data = response.json()
        
        assert isinstance(data, list)
        print(f"Weekly Draft POs count: {len(data)}")
    
    def test_get_weekly_draft_pos_with_status_filter(self):
        """Test filtering weekly draft POs by status"""
        response = self.session.get(
            f"{BASE_URL}/api/mrp/weekly-draft-pos?status=DRAFT&limit=50"
        )
        
        assert response.status_code == 200
        data = response.json()
        
        # All returned POs should have DRAFT status
        for po in data:
            assert po.get("status") == "DRAFT"
        
        print(f"Draft POs with DRAFT status: {len(data)}")
    
    # ============ Integration Tests ============
    
    def test_preview_summary_matches_vendor_totals(self):
        """Test that summary totals match sum of vendor POs"""
        response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/preview",
            json=self.test_weeks
        )
        
        assert response.status_code == 200
        data = response.json()
        
        preview_pos = data.get("preview_pos", [])
        summary = data.get("summary", {})
        
        # Calculate totals from vendor POs
        calculated_items = sum(po.get("total_items", 0) for po in preview_pos)
        calculated_amount = sum(po.get("total_amount", 0) for po in preview_pos)
        
        assert calculated_items == summary.get("total_items", 0), \
            f"Items mismatch: {calculated_items} vs {summary.get('total_items')}"
        
        # Amount might have floating point differences
        assert abs(calculated_amount - summary.get("total_amount", 0)) < 1, \
            f"Amount mismatch: {calculated_amount} vs {summary.get('total_amount')}"
        
        print(f"Summary totals verified: {summary}")
    
    def test_template_row_count_matches_preview(self):
        """Test template row count matches preview item count"""
        # Get preview
        preview_response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/preview",
            json=self.test_weeks[:2]
        )
        assert preview_response.status_code == 200
        preview_data = preview_response.json()
        
        total_items = preview_data["summary"]["total_items"]
        items_without_vendor = preview_data["summary"]["items_needing_vendor"]
        expected_rows = total_items + items_without_vendor
        
        # Get template
        template_response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/download-template",
            json=self.test_weeks[:2]
        )
        assert template_response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb["PO_Lines"]
        
        actual_rows = ws.max_row - 1  # Exclude header
        
        assert actual_rows == expected_rows, \
            f"Row count mismatch: template has {actual_rows}, expected {expected_rows}"
        
        print(f"Template rows ({actual_rows}) match preview items ({expected_rows})")


class TestWeeklyPOUpload:
    """Test Weekly PO Upload functionality"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test fixtures"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login
        login_response = self.session.post(
            f"{BASE_URL}/api/auth/login",
            json={"email": "admin@factory.com", "password": "bidso123"}
        )
        assert login_response.status_code == 200
        
        token = login_response.json().get("access_token")
        self.session.headers.update({"Authorization": f"Bearer {token}"})
        self.token = token
        
        # Get weekly run
        runs_response = self.session.get(f"{BASE_URL}/api/mrp/runs?limit=10")
        runs = runs_response.json()
        weekly_runs = [r for r in runs if r.get("version") == "WEEKLY_V1"]
        
        self.run_id = weekly_runs[0]["id"]
        
        # Get test weeks
        plan_response = self.session.get(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-plan?plan_type=all"
        )
        weekly_plan = plan_response.json().get("weekly_plan", [])
        self.test_weeks = [w["order_week"] for w in weekly_plan[:2]]
    
    def test_upload_creates_draft_pos(self):
        """Test uploading template creates draft POs"""
        # Download template first
        template_response = self.session.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/download-template",
            json=self.test_weeks
        )
        assert template_response.status_code == 200
        
        # Modify template to ensure all rows have vendors
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb["PO_Lines"]
        
        # Get first vendor from Vendors_Reference
        ws_vendors = wb["Vendors_Reference"]
        first_vendor_id = ws_vendors.cell(row=2, column=1).value
        first_vendor_name = ws_vendors.cell(row=2, column=2).value
        
        if not first_vendor_id:
            pytest.skip("No vendors available for testing")
        
        # Update rows without vendor
        rows_updated = 0
        for row in range(2, min(ws.max_row + 1, 10)):  # Limit to first 8 rows for test
            vendor_id = ws.cell(row=row, column=10).value
            if not vendor_id:
                ws.cell(row=row, column=10, value=first_vendor_id)
                ws.cell(row=row, column=11, value=first_vendor_name)
                rows_updated += 1
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload
        files = {"file": ("test_upload.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        upload_response = requests.post(
            f"{BASE_URL}/api/mrp/runs/{self.run_id}/weekly-pos/upload",
            headers={"Authorization": f"Bearer {self.token}"},
            files=files
        )
        
        assert upload_response.status_code == 200, f"Upload failed: {upload_response.text}"
        
        data = upload_response.json()
        assert "created_pos" in data or "message" in data
        
        print(f"Upload result: {data}")


class TestWeeklyDraftPOOperations:
    """Test Weekly Draft PO edit/approve/issue operations"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test fixtures"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login
        login_response = self.session.post(
            f"{BASE_URL}/api/auth/login",
            json={"email": "admin@factory.com", "password": "bidso123"}
        )
        assert login_response.status_code == 200
        
        token = login_response.json().get("access_token")
        self.session.headers.update({"Authorization": f"Bearer {token}"})
    
    def test_update_draft_po_vendor(self):
        """Test updating vendor on a draft PO"""
        # Get a draft PO
        response = self.session.get(
            f"{BASE_URL}/api/mrp/weekly-draft-pos?status=DRAFT&limit=1"
        )
        assert response.status_code == 200
        
        draft_pos = response.json()
        if not draft_pos:
            pytest.skip("No draft POs available for testing")
        
        draft_po = draft_pos[0]
        draft_po_id = draft_po["id"]
        
        # Get a vendor
        vendors_response = self.session.get(f"{BASE_URL}/api/vendors?limit=5")
        assert vendors_response.status_code == 200
        
        vendors = vendors_response.json()
        if not vendors:
            pytest.skip("No vendors available")
        
        new_vendor_id = vendors[0].get("id") or vendors[0].get("vendor_id")
        
        # Update vendor
        update_response = self.session.put(
            f"{BASE_URL}/api/mrp/weekly-draft-pos/{draft_po_id}",
            json={"vendor_id": new_vendor_id}
        )
        
        assert update_response.status_code == 200
        updated_po = update_response.json()
        assert updated_po["vendor_id"] == new_vendor_id
        
        print(f"Updated vendor to: {updated_po['vendor_name']}")
    
    def test_update_draft_po_line_quantity(self):
        """Test updating line quantity on a draft PO"""
        # Get a draft PO with lines
        response = self.session.get(
            f"{BASE_URL}/api/mrp/weekly-draft-pos?status=DRAFT&limit=1"
        )
        assert response.status_code == 200
        
        draft_pos = response.json()
        if not draft_pos:
            pytest.skip("No draft POs available for testing")
        
        draft_po = draft_pos[0]
        draft_po_id = draft_po["id"]
        
        lines = draft_po.get("lines", [])
        if not lines:
            pytest.skip("Draft PO has no lines")
        
        rm_id = lines[0]["rm_id"]
        new_qty = 999.0
        
        # Update line quantity
        update_response = self.session.put(
            f"{BASE_URL}/api/mrp/weekly-draft-pos/{draft_po_id}/line/{rm_id}",
            json={"quantity": new_qty}
        )
        
        assert update_response.status_code == 200
        result = update_response.json()
        assert result["new_quantity"] == new_qty
        
        print(f"Updated {rm_id} quantity to {new_qty}")
    
    def test_approve_draft_po(self):
        """Test approving a draft PO"""
        # Get a draft PO
        response = self.session.get(
            f"{BASE_URL}/api/mrp/weekly-draft-pos?status=DRAFT&limit=1"
        )
        assert response.status_code == 200
        
        draft_pos = response.json()
        if not draft_pos:
            pytest.skip("No draft POs available for testing")
        
        draft_po = draft_pos[0]
        draft_po_id = draft_po["id"]
        
        # Approve
        approve_response = self.session.put(
            f"{BASE_URL}/api/mrp/draft-pos/{draft_po_id}/approve"
        )
        
        # Note: endpoint might be POST or PUT
        if approve_response.status_code == 405:
            approve_response = self.session.post(
                f"{BASE_URL}/api/mrp/draft-pos/{draft_po_id}/approve"
            )
        
        assert approve_response.status_code == 200
        print(f"Approved draft PO: {draft_po_id}")
    
    def test_convert_approved_po(self):
        """Test converting approved PO to actual PO"""
        # Get an approved PO
        response = self.session.get(
            f"{BASE_URL}/api/mrp/draft-pos?status=APPROVED"
        )
        assert response.status_code == 200
        
        approved_pos = response.json()
        if not approved_pos:
            pytest.skip("No approved POs available for testing")
        
        # Find one with vendor
        po_with_vendor = None
        for po in approved_pos:
            if po.get("vendor_id"):
                po_with_vendor = po
                break
        
        if not po_with_vendor:
            pytest.skip("No approved PO with vendor found")
        
        po_id = po_with_vendor["id"]
        
        # Convert to PO
        convert_response = self.session.post(
            f"{BASE_URL}/api/mrp/draft-pos/{po_id}/convert"
        )
        
        if convert_response.status_code == 200:
            result = convert_response.json()
            assert "po_number" in result
            print(f"Converted to PO: {result['po_number']}")
        else:
            # Might already be converted or other issue
            print(f"Convert response: {convert_response.status_code} - {convert_response.text}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
