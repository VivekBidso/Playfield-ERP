"""Backend tests for RM Price Upload + Margin feature.

Endpoints under test (router prefix /api/rm-prices):
    GET    /template
    POST   /upload
    GET    /stats
    GET    /avg-prices
    GET    /history
    GET    /bom-cost/{buyer_sku_id}
    POST   /bom-cost-bulk
    GET    /margin-report
    DELETE /history
"""
import os
import io
from datetime import datetime, timedelta
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
API = f"{BASE_URL}/api/rm-prices"


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    return s


@pytest.fixture(scope="module")
def known_vendor_and_rm(session):
    """Pick first existing vendor + RM for valid upload tests."""
    v = session.get(f"{BASE_URL}/api/vendors").json()
    vendors = v if isinstance(v, list) else v.get("items", v.get("vendors", []))
    assert vendors, "No vendors in DB"
    vendor_id = vendors[0]["vendor_id"]

    r = session.get(f"{BASE_URL}/api/raw-materials").json()
    rms = r if isinstance(r, list) else r.get("items", r.get("raw_materials", []))
    assert rms, "No raw materials in DB"
    rm_id = rms[0]["rm_id"]
    return vendor_id, rm_id


def _build_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Invoice No", "Vendor ID", "RM ID", "Price"])
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ---------------- Template ----------------
class TestTemplate:
    def test_template_download_xlsx(self, session):
        r = session.get(f"{API}/template")
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "octet-stream" in ct
        # Valid xlsx magic bytes (zip header)
        assert r.content[:2] == b"PK"
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Date" in headers and "Invoice No" in headers
        assert "Vendor ID" in headers and "RM ID" in headers
        assert "Price" in headers


# ---------------- Upload ----------------
class TestUpload:
    def test_upload_invalid_vendor_and_rm(self, session, known_vendor_and_rm):
        vendor_id, rm_id = known_vendor_and_rm
        today = datetime.utcnow().strftime("%Y-%m-%d")
        bio = _build_xlsx([
            [today, "INV-BAD-1", "VND_DOES_NOT_EXIST", rm_id, 100.00],
            [today, "INV-BAD-2", vendor_id, "RM_DOES_NOT_EXIST", 50.00],
        ])
        r = session.post(
            f"{API}/upload",
            files={"file": ("bad.xlsx", bio, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["inserted"] == 0
        assert data["error_count"] >= 2
        joined = " ".join(data.get("errors", []))
        assert "Vendor" in joined or "RM" in joined

    def test_upload_valid_rows_appended(self, session, known_vendor_and_rm):
        vendor_id, rm_id = known_vendor_and_rm
        today = datetime.utcnow()
        rows = [
            [(today - timedelta(days=10)).strftime("%Y-%m-%d"), "TEST_INV_001", vendor_id, rm_id, 111.11],
            [(today - timedelta(days=20)).strftime("%Y-%m-%d"), "TEST_INV_002", vendor_id, rm_id, 222.22],
        ]
        bio = _build_xlsx(rows)
        r = session.post(
            f"{API}/upload",
            files={"file": ("valid.xlsx", bio, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["inserted"] == 2
        assert data["error_count"] == 0


# ---------------- Stats ----------------
class TestStats:
    def test_stats_shape(self, session):
        r = session.get(f"{API}/stats")
        assert r.status_code == 200
        d = r.json()
        for k in ["total_records", "unique_rms", "unique_vendors"]:
            assert k in d
        assert d["total_records"] >= 1
        assert isinstance(d["unique_rms"], int)
        assert isinstance(d["unique_vendors"], int)


# ---------------- Avg Prices ----------------
class TestAvgPrices:
    def test_avg_prices_enrichment(self, session):
        r = session.get(f"{API}/avg-prices")
        assert r.status_code == 200
        d = r.json()
        assert "items" in d and "window_months" in d
        assert d["window_months"] == 3
        if d["items"]:
            sample = d["items"][0]
            for k in ["rm_id", "avg_price", "record_count"]:
                assert k in sample
            # rm_name enrichment field present (may be None if not found)
            assert "rm_name" in sample
            assert "category" in sample


# ---------------- History ----------------
class TestHistory:
    def test_history_pagination_and_filters(self, session, known_vendor_and_rm):
        vendor_id, rm_id = known_vendor_and_rm
        r = session.get(f"{API}/history", params={"page": 1, "page_size": 10})
        assert r.status_code == 200
        d = r.json()
        for k in ["items", "total", "page", "page_size", "total_pages"]:
            assert k in d
        assert d["page"] == 1 and d["page_size"] == 10

        r2 = session.get(f"{API}/history", params={"rm_id": rm_id})
        assert r2.status_code == 200
        for it in r2.json()["items"]:
            assert it["rm_id"] == rm_id

        r3 = session.get(f"{API}/history", params={"vendor_id": vendor_id})
        assert r3.status_code == 200
        for it in r3.json()["items"]:
            assert it["vendor_id"] == vendor_id


# ---------------- BOM Cost ----------------
class TestBomCost:
    def test_bom_cost_for_known_sku(self, session):
        # Use known SKU from agent context
        r = session.get(f"{API}/bom-cost/AD_KS_BE_010")
        assert r.status_code == 200
        d = r.json()
        for k in ["buyer_sku_id", "total_cost", "items", "rm_count", "missing_price_count"]:
            assert k in d
        assert d["buyer_sku_id"] == "AD_KS_BE_010"
        assert isinstance(d["items"], list)
        if d["items"]:
            it = d["items"][0]
            for k in ["rm_id", "quantity", "avg_price", "line_cost", "has_price"]:
                assert k in it
            assert "rm_name" in it

    def test_bom_cost_404_for_unknown(self, session):
        r = session.get(f"{API}/bom-cost/UNKNOWN_SKU_XYZ_999")
        assert r.status_code == 404

    def test_bom_cost_bulk(self, session):
        r = session.post(
            f"{API}/bom-cost-bulk",
            json={"buyer_sku_ids": ["AD_KS_BE_010", "UNKNOWN_SKU_XYZ_999"]},
        )
        assert r.status_code == 200
        d = r.json()
        assert "costs" in d and "count" in d
        assert "AD_KS_BE_010" in d["costs"]
        assert "UNKNOWN_SKU_XYZ_999" not in d["costs"]
        entry = d["costs"]["AD_KS_BE_010"]
        for k in ["total_cost", "rm_count", "missing_price_count"]:
            assert k in entry


# ---------------- Margin Report ----------------
class TestMarginReport:
    def test_margin_report_shape(self, session):
        r = session.get(f"{API}/margin-report")
        assert r.status_code == 200
        d = r.json()
        assert "items" in d and "totals" in d and "count" in d
        for k in ["total_qty", "total_revenue", "total_cogs", "gross_profit", "overall_margin_pct"]:
            assert k in d["totals"]
        if d["items"]:
            row = d["items"][0]
            for k in ["buyer_sku_id", "avg_asp", "bom_cost", "margin_pct", "margin_value"]:
                assert k in row

    def test_margin_report_month_filter(self, session):
        # Past wide window — just ensure no 500
        r = session.get(f"{API}/margin-report", params={"from_month": "2020-01", "to_month": "2030-12"})
        assert r.status_code == 200
        assert "items" in r.json()
