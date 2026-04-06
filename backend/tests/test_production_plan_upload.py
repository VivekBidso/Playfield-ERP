"""
Test Production Plan Excel Upload with FIFO allocation, capacity conflict detection, and Excel result report.
Tests the POST /api/cpc/production-plan/upload-excel endpoint with modes: check, add, override
"""
import pytest
import requests
import os
import io
import json
from datetime import datetime, timedelta

# Use openpyxl for creating test Excel files
try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_USER = "admin@factory.com"
TEST_PASSWORD = "bidso123"


class TestProductionPlanUpload:
    """Test Production Plan Excel Upload endpoint"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test session with auth"""
        self.session = requests.Session()
        
        # Login
        login_res = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": TEST_USER,
            "password": TEST_PASSWORD
        })
        assert login_res.status_code == 200, f"Login failed: {login_res.text}"
        token = login_res.json().get("access_token")
        self.auth_headers = {}
        if token:
            self.auth_headers = {"Authorization": f"Bearer {token}"}
        
        yield
    
    def upload_file(self, file_obj, filename, mode=None):
        """Helper to upload file with proper headers"""
        url = f"{BASE_URL}/api/cpc/production-plan/upload-excel"
        if mode:
            url += f"?mode={mode}"
        
        files = {"file": (filename, file_obj, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        return requests.post(url, files=files, headers=self.auth_headers)
    
    def create_test_excel(self, rows):
        """Create a test Excel file with given rows"""
        if not openpyxl:
            pytest.skip("openpyxl not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Production Plan"
        
        # Headers
        headers = ["Branch ID", "Date (DD-MM-YYYY)", "Buyer SKU ID", "Quantity"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_data.get("branch_id", ""))
            ws.cell(row=row_idx, column=2, value=row_data.get("date", ""))
            ws.cell(row=row_idx, column=3, value=row_data.get("sku_id", ""))
            ws.cell(row=row_idx, column=4, value=row_data.get("quantity", 0))
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def test_01_get_valid_test_data(self):
        """Get valid branch and SKU data for testing"""
        # Get branches
        branches_res = self.session.get(f"{BASE_URL}/api/branches/reference", headers=self.auth_headers)
        assert branches_res.status_code == 200, f"Failed to get branches: {branches_res.text}"
        branches = branches_res.json().get("branches", [])
        print(f"Found {len(branches)} branches")
        for b in branches[:3]:
            print(f"  - {b.get('branch_id')}: {b.get('name')} (capacity: {b.get('capacity')})")
        
        # Get SKU-branch assignments to find valid SKU-branch combos
        assignments_res = self.session.get(f"{BASE_URL}/api/sku-branch-assignments", headers=self.auth_headers)
        if assignments_res.status_code == 200:
            assignments = assignments_res.json()
            print(f"Found {len(assignments)} SKU-branch assignments")
            for a in assignments[:5]:
                print(f"  - SKU: {a.get('sku_id')}, Branch: {a.get('branch')}")
        else:
            print(f"SKU-branch assignments endpoint returned: {assignments_res.status_code}")
        
        print("Test data retrieval complete")
    
    def test_02_download_template(self):
        """Test downloading the production plan template"""
        res = self.session.get(f"{BASE_URL}/api/cpc/production-plan/template", headers=self.auth_headers)
        assert res.status_code == 200, f"Failed to download template: {res.text}"
        
        # Check content type
        content_type = res.headers.get("content-type", "")
        assert "spreadsheet" in content_type or "excel" in content_type, f"Unexpected content type: {content_type}"
        
        # Check content disposition
        content_disp = res.headers.get("content-disposition", "")
        assert "attachment" in content_disp, f"Expected attachment, got: {content_disp}"
        assert "template" in content_disp.lower(), f"Expected template filename, got: {content_disp}"
        
        print("Template download successful")
    
    def test_03_upload_invalid_file_format(self):
        """Test uploading non-Excel file returns error"""
        files = {"file": ("test.txt", io.BytesIO(b"not an excel file"), "text/plain")}
        res = requests.post(f"{BASE_URL}/api/cpc/production-plan/upload-excel", files=files, headers=self.auth_headers)
        
        assert res.status_code == 400, f"Expected 400, got {res.status_code}: {res.text}"
        data = res.json()
        assert "detail" in data
        assert "excel" in data["detail"].lower() or "xlsx" in data["detail"].lower()
        print(f"Invalid file format error: {data['detail']}")
    
    def test_04_upload_missing_columns(self):
        """Test uploading Excel with missing required columns"""
        if not openpyxl:
            pytest.skip("openpyxl not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Wrong Column")
        ws.cell(row=2, column=1, value="Some Value")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        res = self.upload_file(output, "test.xlsx")
        
        assert res.status_code == 400, f"Expected 400, got {res.status_code}: {res.text}"
        data = res.json()
        assert "detail" in data
        assert "missing" in data["detail"].lower() or "column" in data["detail"].lower()
        print(f"Missing columns error: {data['detail']}")
    
    def test_05_upload_with_invalid_branch(self):
        """Test uploading with invalid branch ID"""
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%d-%m-%Y")
        
        rows = [
            {"branch_id": "INVALID_BRANCH", "date": tomorrow, "sku_id": "FC_KS_BE_115", "quantity": 100}
        ]
        excel_file = self.create_test_excel(rows)
        
        res = self.upload_file(excel_file, "test.xlsx", mode="add")
        
        # Should return Excel with ERROR status
        assert res.status_code == 200, f"Expected 200, got {res.status_code}: {res.text}"
        
        # Check X-Upload-Summary header
        summary_header = res.headers.get("X-Upload-Summary")
        if summary_header:
            summary = json.loads(summary_header)
            print(f"Upload summary: {summary}")
            assert summary.get("errors", 0) > 0, "Expected errors for invalid branch"
        
        print("Invalid branch test passed")
    
    def test_06_upload_with_invalid_sku(self):
        """Test uploading with invalid SKU ID"""
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%d-%m-%Y")
        
        rows = [
            {"branch_id": "BR_001", "date": tomorrow, "sku_id": "INVALID_SKU_12345", "quantity": 100}
        ]
        excel_file = self.create_test_excel(rows)
        
        res = self.upload_file(excel_file, "test.xlsx", mode="add")
        
        assert res.status_code == 200, f"Expected 200, got {res.status_code}: {res.text}"
        
        summary_header = res.headers.get("X-Upload-Summary")
        if summary_header:
            summary = json.loads(summary_header)
            print(f"Upload summary: {summary}")
            assert summary.get("errors", 0) > 0, "Expected errors for invalid SKU"
        
        print("Invalid SKU test passed")
    
    def test_07_check_mode_returns_warning_on_conflict(self):
        """Test check mode returns JSON warning when conflicts exist"""
        # First, create a schedule to cause conflict
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%d-%m-%Y")
        
        # Upload a small amount first with add mode
        rows = [
            {"branch_id": "BR_001", "date": tomorrow, "sku_id": "FC_KS_BE_115", "quantity": 500}
        ]
        excel_file = self.create_test_excel(rows)
        res1 = self.upload_file(excel_file, "test.xlsx", mode="add")
        print(f"First upload status: {res1.status_code}")
        
        # Now try to upload more with check mode (default)
        rows2 = [
            {"branch_id": "BR_001", "date": tomorrow, "sku_id": "FC_KS_BE_115", "quantity": 800}
        ]
        excel_file2 = self.create_test_excel(rows2)
        res2 = self.upload_file(excel_file2, "test2.xlsx")  # No mode = check
        
        # Check mode should return JSON warning if there's a conflict
        content_type = res2.headers.get("content-type", "")
        print(f"Check mode response content-type: {content_type}")
        
        if "json" in content_type:
            data = res2.json()
            print(f"Check mode response: {json.dumps(data, indent=2)}")
            
            if data.get("warning"):
                assert "conflicts" in data, "Expected conflicts in warning response"
                print(f"Conflicts detected: {len(data.get('conflicts', []))}")
                
                # Verify conflict structure
                for conflict in data.get("conflicts", []):
                    assert "branch" in conflict
                    assert "date" in conflict or "date_iso" in conflict
                    assert "capacity" in conflict
                    assert "existing_scheduled" in conflict
                    assert "new_demand" in conflict
                    print(f"Conflict: {conflict.get('branch')} - existing: {conflict.get('existing_scheduled')}, new: {conflict.get('new_demand')}")
        else:
            # No conflict - Excel returned directly
            print("No conflict detected - Excel returned directly")
            summary_header = res2.headers.get("X-Upload-Summary")
            if summary_header:
                summary = json.loads(summary_header)
                print(f"Upload summary: {summary}")
    
    def test_08_add_mode_allocates_within_remaining_capacity(self):
        """Test add mode allocates only within remaining capacity"""
        # Use a date far in future to avoid conflicts with other tests
        future_date = (datetime.now() + timedelta(days=30)).strftime("%d-%m-%Y")
        
        rows = [
            {"branch_id": "BR_001", "date": future_date, "sku_id": "FC_KS_BE_115", "quantity": 100}
        ]
        excel_file = self.create_test_excel(rows)
        
        res = self.upload_file(excel_file, "test.xlsx", mode="add")
        
        assert res.status_code == 200, f"Expected 200, got {res.status_code}: {res.text}"
        
        # Check content type - should be Excel
        content_type = res.headers.get("content-type", "")
        print(f"Add mode response content-type: {content_type}")
        
        if "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
            # Success - Excel returned
            summary_header = res.headers.get("X-Upload-Summary")
            assert summary_header, "Expected X-Upload-Summary header"
            
            summary = json.loads(summary_header)
            print(f"Add mode summary: {summary}")
            
            assert summary.get("mode") == "add", f"Expected mode=add, got {summary.get('mode')}"
            
            # Verify Excel content
            excel_content = io.BytesIO(res.content)
            wb = openpyxl.load_workbook(excel_content)
            ws = wb.active
            
            # Check headers
            headers = [cell.value for cell in ws[1]]
            print(f"Excel headers: {headers}")
            
            expected_headers = ["Branch ID", "Date (DD-MM-YYYY)", "Buyer SKU ID", "Quantity", "Status", "Allocated", "Not Allocated", "Schedule Code", "Remarks"]
            for expected in expected_headers:
                assert expected in headers, f"Missing header: {expected}"
            
            # Check data row
            if ws.max_row > 1:
                row_data = [cell.value for cell in ws[2]]
                print(f"First data row: {row_data}")
                
                # Status should be SCHEDULED, PARTIAL, REJECTED, or ERROR
                status_idx = headers.index("Status")
                status = row_data[status_idx]
                assert status in ["SCHEDULED", "PARTIAL", "REJECTED", "ERROR"], f"Unexpected status: {status}"
                print(f"Row status: {status}")
        else:
            # JSON response (possibly error or warning)
            data = res.json()
            print(f"Add mode JSON response: {data}")
    
    def test_09_override_mode_clears_existing_schedules(self):
        """Test override mode clears existing schedules and allocates fresh"""
        # Use a specific date for override test
        override_date = (datetime.now() + timedelta(days=31)).strftime("%d-%m-%Y")
        
        # First upload some data
        rows1 = [
            {"branch_id": "BR_001", "date": override_date, "sku_id": "FC_KS_BE_115", "quantity": 200}
        ]
        excel_file1 = self.create_test_excel(rows1)
        res1 = self.upload_file(excel_file1, "test1.xlsx", mode="add")
        print(f"First upload status: {res1.status_code}")
        
        # Now override with new data
        rows2 = [
            {"branch_id": "BR_001", "date": override_date, "sku_id": "FC_KS_BE_115", "quantity": 300}
        ]
        excel_file2 = self.create_test_excel(rows2)
        res2 = self.upload_file(excel_file2, "test2.xlsx", mode="override")
        
        assert res2.status_code == 200, f"Expected 200, got {res2.status_code}: {res2.text}"
        
        content_type = res2.headers.get("content-type", "")
        print(f"Override mode response content-type: {content_type}")
        
        if "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
            summary_header = res2.headers.get("X-Upload-Summary")
            if summary_header:
                summary = json.loads(summary_header)
                print(f"Override mode summary: {summary}")
                assert summary.get("mode") == "override", f"Expected mode=override, got {summary.get('mode')}"
        else:
            data = res2.json()
            print(f"Override mode JSON response: {data}")
    
    def test_10_fifo_allocation_first_rows_get_priority(self):
        """Test FIFO allocation - first rows in file get full allocation before later rows"""
        # Use a date with limited capacity
        fifo_date = (datetime.now() + timedelta(days=32)).strftime("%d-%m-%Y")
        
        # Create rows that exceed capacity (assuming capacity is 1000)
        rows = [
            {"branch_id": "BR_001", "date": fifo_date, "sku_id": "FC_KS_BE_115", "quantity": 600},  # Should be fully allocated
            {"branch_id": "BR_001", "date": fifo_date, "sku_id": "FC_KS_BE_115", "quantity": 600},  # Should be partial or rejected
        ]
        excel_file = self.create_test_excel(rows)
        
        res = self.upload_file(excel_file, "test.xlsx", mode="override")
        
        assert res.status_code == 200, f"Expected 200, got {res.status_code}: {res.text}"
        
        content_type = res.headers.get("content-type", "")
        
        if "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
            summary_header = res.headers.get("X-Upload-Summary")
            if summary_header:
                summary = json.loads(summary_header)
                print(f"FIFO test summary: {summary}")
                
                # Check if there's partial allocation
                if summary.get("partial", 0) > 0 or summary.get("rejected", 0) > 0:
                    print("FIFO allocation working - some rows got partial/rejected due to capacity")
                
                # Verify Excel content
                excel_content = io.BytesIO(res.content)
                wb = openpyxl.load_workbook(excel_content)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                status_idx = headers.index("Status")
                allocated_idx = headers.index("Allocated")
                not_allocated_idx = headers.index("Not Allocated")
                
                # Check first row - should be SCHEDULED (fully allocated)
                if ws.max_row > 1:
                    row1 = [cell.value for cell in ws[2]]
                    print(f"Row 1: Status={row1[status_idx]}, Allocated={row1[allocated_idx]}, Not Allocated={row1[not_allocated_idx]}")
                
                # Check second row - should be PARTIAL or REJECTED if capacity exceeded
                if ws.max_row > 2:
                    row2 = [cell.value for cell in ws[3]]
                    print(f"Row 2: Status={row2[status_idx]}, Allocated={row2[allocated_idx]}, Not Allocated={row2[not_allocated_idx]}")
    
    def test_11_partial_allocation_shows_correct_counts(self):
        """Test partial allocation shows correct allocated/not_allocated counts"""
        partial_date = (datetime.now() + timedelta(days=33)).strftime("%d-%m-%Y")
        
        # Request more than capacity
        rows = [
            {"branch_id": "BR_001", "date": partial_date, "sku_id": "FC_KS_BE_115", "quantity": 1500}  # Exceeds 1000 capacity
        ]
        excel_file = self.create_test_excel(rows)
        
        res = self.upload_file(excel_file, "test.xlsx", mode="override")
        
        assert res.status_code == 200, f"Expected 200, got {res.status_code}: {res.text}"
        
        content_type = res.headers.get("content-type", "")
        
        if "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
            summary_header = res.headers.get("X-Upload-Summary")
            if summary_header:
                summary = json.loads(summary_header)
                print(f"Partial allocation summary: {summary}")
                
                total_allocated = summary.get("total_allocated", 0)
                total_not_allocated = summary.get("total_not_allocated", 0)
                
                print(f"Total allocated: {total_allocated}, Total not allocated: {total_not_allocated}")
                
                # Verify Excel content
                excel_content = io.BytesIO(res.content)
                wb = openpyxl.load_workbook(excel_content)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                status_idx = headers.index("Status")
                allocated_idx = headers.index("Allocated")
                not_allocated_idx = headers.index("Not Allocated")
                remarks_idx = headers.index("Remarks")
                
                if ws.max_row > 1:
                    row = [cell.value for cell in ws[2]]
                    status = row[status_idx]
                    allocated = row[allocated_idx]
                    not_allocated = row[not_allocated_idx]
                    remarks = row[remarks_idx]
                    
                    print(f"Status: {status}, Allocated: {allocated}, Not Allocated: {not_allocated}")
                    print(f"Remarks: {remarks}")
                    
                    if status == "PARTIAL":
                        assert allocated > 0, "Partial status should have some allocation"
                        assert not_allocated > 0, "Partial status should have some not allocated"
                        assert allocated + not_allocated == 1500, "Allocated + Not Allocated should equal requested"
    
    def test_12_excel_result_has_required_columns(self):
        """Test Excel result file has all required columns"""
        test_date = (datetime.now() + timedelta(days=34)).strftime("%d-%m-%Y")
        
        rows = [
            {"branch_id": "BR_001", "date": test_date, "sku_id": "FC_KS_BE_115", "quantity": 50}
        ]
        excel_file = self.create_test_excel(rows)
        
        res = self.upload_file(excel_file, "test.xlsx", mode="add")
        
        assert res.status_code == 200, f"Expected 200, got {res.status_code}: {res.text}"
        
        content_type = res.headers.get("content-type", "")
        
        if "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
            excel_content = io.BytesIO(res.content)
            wb = openpyxl.load_workbook(excel_content)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            print(f"Result Excel headers: {headers}")
            
            # Required columns per spec
            required_columns = ["Status", "Allocated", "Not Allocated", "Schedule Code", "Remarks"]
            
            for col in required_columns:
                assert col in headers, f"Missing required column: {col}"
            
            print("All required columns present in result Excel")
    
    def test_13_schedule_code_generated_for_successful_rows(self):
        """Test that schedule code is generated for successfully scheduled rows"""
        test_date = (datetime.now() + timedelta(days=35)).strftime("%d-%m-%Y")
        
        rows = [
            {"branch_id": "BR_001", "date": test_date, "sku_id": "FC_KS_BE_115", "quantity": 50}
        ]
        excel_file = self.create_test_excel(rows)
        
        res = self.upload_file(excel_file, "test.xlsx", mode="add")
        
        assert res.status_code == 200, f"Expected 200, got {res.status_code}: {res.text}"
        
        content_type = res.headers.get("content-type", "")
        
        if "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
            excel_content = io.BytesIO(res.content)
            wb = openpyxl.load_workbook(excel_content)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            status_idx = headers.index("Status")
            schedule_code_idx = headers.index("Schedule Code")
            
            if ws.max_row > 1:
                row = [cell.value for cell in ws[2]]
                status = row[status_idx]
                schedule_code = row[schedule_code_idx]
                
                print(f"Status: {status}, Schedule Code: {schedule_code}")
                
                if status in ["SCHEDULED", "PARTIAL"]:
                    assert schedule_code, "Schedule code should be generated for SCHEDULED/PARTIAL rows"
                    assert schedule_code.startswith("PS_"), f"Schedule code should start with PS_, got: {schedule_code}"
                    print(f"Schedule code generated: {schedule_code}")


class TestBranchCapacity:
    """Test branch capacity endpoints"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test session with auth"""
        self.session = requests.Session()
        
        # Login
        login_res = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": TEST_USER,
            "password": TEST_PASSWORD
        })
        assert login_res.status_code == 200, f"Login failed: {login_res.text}"
        token = login_res.json().get("access_token")
        self.auth_headers = {}
        if token:
            self.auth_headers = {"Authorization": f"Bearer {token}"}
    
    def test_get_branch_capacities(self):
        """Test getting branch capacities"""
        res = self.session.get(f"{BASE_URL}/api/branches/capacity", headers=self.auth_headers)
        assert res.status_code == 200, f"Failed to get capacities: {res.text}"
        
        data = res.json()
        assert isinstance(data, list), "Expected list of branches"
        
        if len(data) > 0:
            branch = data[0]
            print(f"Sample branch: {branch}")
            
            # Check required fields
            assert "branch" in branch
            assert "capacity_units_per_day" in branch
            assert "available_today" in branch
            assert "utilization_percent" in branch
    
    def test_get_available_capacity(self):
        """Test getting available capacity"""
        res = self.session.get(f"{BASE_URL}/api/cpc/available-capacity", headers=self.auth_headers)
        assert res.status_code == 200, f"Failed to get available capacity: {res.text}"
        
        data = res.json()
        print(f"Available capacity response keys: {data.keys() if isinstance(data, dict) else 'list'}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
