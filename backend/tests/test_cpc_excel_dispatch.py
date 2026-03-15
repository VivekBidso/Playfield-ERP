"""
CPC Excel Download and Dispatch Lots with forecast_id Tests
Tests for: 
- GET /api/cpc/demand-forecasts/download - Excel file download
- POST /api/dispatch-lots/multi with forecast_id - linking dispatch lots to forecasts
"""
import pytest
import requests
import os
from datetime import datetime, timedelta
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestCPCExcelDownload:
    """CPC Demand Forecasts Excel Download tests"""
    
    def test_download_demand_forecasts_excel(self):
        """Test GET /api/cpc/demand-forecasts/download returns valid Excel file"""
        response = requests.get(f"{BASE_URL}/api/cpc/demand-forecasts/download")
        assert response.status_code == 200
        
        # Verify content type is Excel
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheetml' in content_type or 'application/vnd' in content_type
        
        # Verify content disposition header
        content_disposition = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disposition
        assert 'demand_forecasts' in content_disposition
        assert '.xlsx' in content_disposition
        
        # Verify file is not empty
        assert len(response.content) > 0
        
        # Verify it's a valid Excel file (starts with PK - ZIP signature)
        assert response.content[:2] == b'PK'
    
    def test_download_excel_file_structure(self):
        """Test downloaded Excel file has correct structure"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(f"{BASE_URL}/api/cpc/demand-forecasts/download")
        assert response.status_code == 200
        
        # Load workbook from response content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify sheet name
        assert ws.title == "Demand Forecasts"
        
        # Verify headers
        expected_headers = [
            "Forecast ID", "Buyer Code", "Buyer Name", "Vertical", "Brand", "Model",
            "SKU ID", "SKU Description", "Forecast Month", "Forecast Qty",
            "Scheduled Qty", "Remaining Qty", "Priority", "Status"
        ]
        
        actual_headers = [cell.value for cell in ws[1]]
        assert actual_headers == expected_headers
        
        # Verify there's at least header row
        assert ws.max_row >= 1
        assert ws.max_column == 14


class TestDispatchLotsWithForecastId:
    """Dispatch Lots Multi endpoint with forecast_id tests"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Get valid buyer and forecast for testing"""
        # Get a buyer with forecasts
        buyers_response = requests.get(f"{BASE_URL}/api/dispatch-lots/buyers-with-forecasts")
        if buyers_response.status_code == 200 and len(buyers_response.json()) > 0:
            self.test_buyer_id = buyers_response.json()[0]["id"]
        else:
            self.test_buyer_id = None
        
        # Get a confirmed forecast
        forecasts_response = requests.get(f"{BASE_URL}/api/cpc/demand-forecasts")
        if forecasts_response.status_code == 200 and len(forecasts_response.json()) > 0:
            # Find a forecast with remaining qty
            for f in forecasts_response.json():
                if f.get("remaining_qty", 0) > 0 and f.get("sku_id"):
                    self.test_forecast = f
                    break
            else:
                self.test_forecast = forecasts_response.json()[0]
        else:
            self.test_forecast = None
        
        # Get forecasted SKUs for the buyer
        if self.test_buyer_id:
            skus_response = requests.get(f"{BASE_URL}/api/dispatch-lots/forecasted-skus?buyer_id={self.test_buyer_id}")
            if skus_response.status_code == 200 and len(skus_response.json()) > 0:
                self.test_sku = skus_response.json()[0]
            else:
                self.test_sku = None
        else:
            self.test_sku = None
    
    def test_create_dispatch_lot_with_forecast_id(self):
        """Test POST /api/dispatch-lots/multi saves forecast_id at lot level"""
        if not self.test_buyer_id or not self.test_sku:
            pytest.skip("No buyer or SKU available for testing")
        
        target_date = (datetime.now() + timedelta(days=30)).isoformat()
        
        payload = {
            "buyer_id": self.test_buyer_id,
            "forecast_id": self.test_forecast["id"] if self.test_forecast else None,
            "target_date": target_date,
            "priority": "MEDIUM",
            "notes": "TEST_DISPATCH_LOT_FORECAST_LINK",
            "lines": [
                {
                    "sku_id": self.test_sku["sku_id"],
                    "quantity": 25
                }
            ]
        }
        
        response = requests.post(f"{BASE_URL}/api/dispatch-lots/multi", json=payload)
        assert response.status_code == 200
        
        data = response.json()
        
        # Verify lot has forecast_id
        assert "id" in data
        assert "lot_code" in data
        assert data["buyer_id"] == self.test_buyer_id
        
        if self.test_forecast:
            assert data["forecast_id"] == self.test_forecast["id"]
        
        # Store for cleanup
        self.created_lot_id = data["id"]
    
    def test_create_dispatch_lot_with_line_level_forecast_id(self):
        """Test POST /api/dispatch-lots/multi saves forecast_id at line item level"""
        if not self.test_buyer_id or not self.test_sku:
            pytest.skip("No buyer or SKU available for testing")
        
        target_date = (datetime.now() + timedelta(days=30)).isoformat()
        
        payload = {
            "buyer_id": self.test_buyer_id,
            "target_date": target_date,
            "priority": "HIGH",
            "notes": "TEST_LINE_LEVEL_FORECAST_ID",
            "lines": [
                {
                    "sku_id": self.test_sku["sku_id"],
                    "quantity": 15,
                    "forecast_id": self.test_forecast["id"] if self.test_forecast else None
                }
            ]
        }
        
        response = requests.post(f"{BASE_URL}/api/dispatch-lots/multi", json=payload)
        assert response.status_code == 200
        
        data = response.json()
        
        # Verify lot created
        assert "id" in data
        assert "lines" in data
        assert len(data["lines"]) == 1
        
        # Verify line has forecast_id
        line = data["lines"][0]
        if self.test_forecast:
            assert line["forecast_id"] == self.test_forecast["id"]
    
    def test_dispatch_lot_details_shows_forecast_id(self):
        """Test GET /api/dispatch-lots/{id}/details returns forecast_id"""
        if not self.test_buyer_id or not self.test_sku:
            pytest.skip("No buyer or SKU available for testing")
        
        # First create a lot with forecast_id
        target_date = (datetime.now() + timedelta(days=30)).isoformat()
        
        payload = {
            "buyer_id": self.test_buyer_id,
            "forecast_id": self.test_forecast["id"] if self.test_forecast else None,
            "target_date": target_date,
            "priority": "LOW",
            "notes": "TEST_DETAILS_FORECAST_ID",
            "lines": [
                {
                    "sku_id": self.test_sku["sku_id"],
                    "quantity": 10,
                    "forecast_id": self.test_forecast["id"] if self.test_forecast else None
                }
            ]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/dispatch-lots/multi", json=payload)
        assert create_response.status_code == 200
        
        lot_id = create_response.json()["id"]
        
        # Get lot details
        details_response = requests.get(f"{BASE_URL}/api/dispatch-lots/{lot_id}/details")
        assert details_response.status_code == 200
        
        details = details_response.json()
        
        # Verify forecast_id is present in lot
        if self.test_forecast:
            assert details["forecast_id"] == self.test_forecast["id"]
        
        # Verify forecast_id is present in lines
        assert "lines" in details
        if len(details["lines"]) > 0 and self.test_forecast:
            assert details["lines"][0]["forecast_id"] == self.test_forecast["id"]
    
    def test_dispatch_lot_without_forecast_id(self):
        """Test POST /api/dispatch-lots/multi works without forecast_id"""
        if not self.test_buyer_id or not self.test_sku:
            pytest.skip("No buyer or SKU available for testing")
        
        target_date = (datetime.now() + timedelta(days=30)).isoformat()
        
        payload = {
            "buyer_id": self.test_buyer_id,
            "target_date": target_date,
            "priority": "MEDIUM",
            "notes": "TEST_NO_FORECAST_ID",
            "lines": [
                {
                    "sku_id": self.test_sku["sku_id"],
                    "quantity": 20
                }
            ]
        }
        
        response = requests.post(f"{BASE_URL}/api/dispatch-lots/multi", json=payload)
        assert response.status_code == 200
        
        data = response.json()
        
        # Verify lot created without forecast_id
        assert "id" in data
        # forecast_id should be None or not present
        assert data.get("forecast_id") is None


class TestCPCDemandForecastsAPI:
    """CPC Demand Forecasts API tests"""
    
    def test_get_demand_forecasts(self):
        """Test GET /api/cpc/demand-forecasts returns forecasts with dispatch lots"""
        response = requests.get(f"{BASE_URL}/api/cpc/demand-forecasts")
        assert response.status_code == 200
        
        data = response.json()
        assert isinstance(data, list)
        
        # Verify forecast structure
        if len(data) > 0:
            forecast = data[0]
            assert "id" in forecast
            assert "forecast_code" in forecast
            assert "buyer_id" in forecast
            assert "forecast_qty" in forecast
            assert "scheduled_qty" in forecast
            assert "remaining_qty" in forecast
            assert "dispatch_lots" in forecast
            assert "is_fully_scheduled" in forecast
    
    def test_get_demand_forecasts_summary(self):
        """Test GET /api/cpc/demand-forecasts/summary returns summary stats"""
        response = requests.get(f"{BASE_URL}/api/cpc/demand-forecasts/summary")
        assert response.status_code == 200
        
        data = response.json()
        assert "total_forecasts" in data
        assert "total_forecast_qty" in data
        assert "total_scheduled_qty" in data
        assert "remaining_to_schedule" in data
        assert "scheduling_percent" in data


class TestDispatchLotsLinkedToForecasts:
    """Test dispatch lots appear linked to forecasts in CPC view"""
    
    def test_forecast_shows_linked_dispatch_lots(self):
        """Test forecasts show their linked dispatch lots"""
        response = requests.get(f"{BASE_URL}/api/cpc/demand-forecasts")
        assert response.status_code == 200
        
        data = response.json()
        
        # Find a forecast with dispatch lots
        forecast_with_lots = None
        for f in data:
            if f.get("dispatch_lots") and len(f["dispatch_lots"]) > 0:
                forecast_with_lots = f
                break
        
        if forecast_with_lots:
            # Verify dispatch lot structure
            lot = forecast_with_lots["dispatch_lots"][0]
            assert "lot_code" in lot
            assert "quantity" in lot
            assert "status" in lot
            
            # Verify dispatch_qty is calculated
            assert "dispatch_qty" in forecast_with_lots
            assert forecast_with_lots["dispatch_qty"] >= 0


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
