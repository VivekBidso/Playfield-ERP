"""Branch Operations routes - Production schedules view for branch users"""
from fastapi import APIRouter, HTTPException, Depends, Query
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from database import db
from services.auth_service import get_current_user
from models.auth import User
from routes.sku_management_routes import generate_rm_description
from services.utils import update_branch_rm_inventory, get_branch_rm_stock
from services import sku_service
import uuid

router = APIRouter(tags=["Branch Operations"])


def serialize_doc(doc):
    """Serialize MongoDB document for JSON response"""
    if doc and 'created_at' in doc and isinstance(doc['created_at'], str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    if doc and 'target_date' in doc and isinstance(doc['target_date'], str):
        doc['target_date'] = datetime.fromisoformat(doc['target_date'])
    if doc and 'completed_at' in doc and isinstance(doc['completed_at'], str):
        doc['completed_at'] = datetime.fromisoformat(doc['completed_at'])
    return doc


async def get_merged_bom_for_sku(buyer_sku_id: str) -> dict:
    """
    Get merged BOM (common_bom + brand_bom) for a buyer SKU.
    Returns: {rm_id: quantity_per_unit, ...}
    """
    # Get buyer SKU to find bidso_sku_id
    buyer_sku = await db.buyer_skus.find_one(
        {"buyer_sku_id": buyer_sku_id},
        {"_id": 0, "bidso_sku_id": 1}
    )
    
    if not buyer_sku:
        return {}
    
    bidso_sku_id = buyer_sku.get("bidso_sku_id")
    
    # Get common BOM
    common_bom = await db.common_bom.find_one(
        {"bidso_sku_id": bidso_sku_id},
        {"_id": 0, "items": 1}
    ) if bidso_sku_id else None
    
    # Get brand BOM
    brand_bom = await db.brand_bom.find_one(
        {"buyer_sku_id": buyer_sku_id},
        {"_id": 0, "items": 1}
    )
    
    # Merge BOMs
    merged = {}
    
    if common_bom and common_bom.get("items"):
        for item in common_bom["items"]:
            rm_id = item.get("rm_id")
            if rm_id:
                merged[rm_id] = item.get("quantity", 0)
    
    if brand_bom and brand_bom.get("items"):
        for item in brand_bom["items"]:
            rm_id = item.get("rm_id")
            if rm_id:
                # Add to existing (brand-specific additions)
                merged[rm_id] = merged.get(rm_id, 0) + item.get("quantity", 0)
    
    return merged


async def check_rm_availability_for_production(
    branch: str, 
    buyer_sku_id: str, 
    quantity: int
) -> dict:
    """
    Check if branch has sufficient RM stock for production.
    
    Returns:
        {
            "sufficient": True/False,
            "shortages": [
                {"rm_id": "SP_001", "description": "Screw 8x13", "required": 100, "available": 50, "shortage": 50},
                ...
            ],
            "bom": {rm_id: qty_per_unit, ...}
        }
    """
    bom = await get_merged_bom_for_sku(buyer_sku_id)
    
    if not bom:
        return {"sufficient": True, "shortages": [], "bom": {}, "message": "No BOM defined for this SKU"}
    
    shortages = []
    
    # Get RM details for descriptions
    rm_ids = list(bom.keys())
    rms = await db.raw_materials.find(
        {"rm_id": {"$in": rm_ids}},
        {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1, "unit": 1}
    ).to_list(1000)
    rm_details = {rm["rm_id"]: rm for rm in rms}
    
    for rm_id, qty_per_unit in bom.items():
        required = qty_per_unit * quantity
        available = await get_branch_rm_stock(branch, rm_id)
        
        if available < required:
            rm_info = rm_details.get(rm_id, {})
            description = await generate_rm_description(
                rm_info.get("category", ""),
                rm_info.get("category_data", {}),
                rm_info.get("name", "")
            )
            shortages.append({
                "rm_id": rm_id,
                "description": description,
                "unit": rm_info.get("unit", ""),
                "required": round(required, 2),
                "available": round(available, 2),
                "shortage": round(required - available, 2)
            })
    
    return {
        "sufficient": len(shortages) == 0,
        "shortages": shortages,
        "bom": bom
    }


async def consume_rm_for_production(
    branch: str,
    buyer_sku_id: str,
    quantity: int,
    schedule_code: str
) -> dict:
    """
    Deduct RM from branch inventory based on BOM.
    Should be called AFTER availability check passes.
    
    Returns:
        {"consumed": [{rm_id, qty_consumed}, ...], "total_items": N}
    """
    bom = await get_merged_bom_for_sku(buyer_sku_id)
    consumed = []
    
    for rm_id, qty_per_unit in bom.items():
        qty_to_consume = qty_per_unit * quantity
        
        # Deduct from inventory (negative change)
        await update_branch_rm_inventory(branch, rm_id, -qty_to_consume)
        
        # Log the consumption
        await db.rm_consumption_log.insert_one({
            "id": str(uuid.uuid4()),
            "branch": branch,
            "rm_id": rm_id,
            "quantity": qty_to_consume,
            "buyer_sku_id": buyer_sku_id,
            "schedule_code": schedule_code,
            "consumed_at": datetime.now(timezone.utc).isoformat()
        })
        
        consumed.append({"rm_id": rm_id, "quantity": qty_to_consume})
    
    return {"consumed": consumed, "total_items": len(consumed)}


async def add_fg_inventory(
    branch: str,
    buyer_sku_id: str,
    quantity: int,
    schedule_code: str
) -> dict:
    """
    Add finished goods (Buyer SKU) to branch inventory after production completion.
    Writes to both branch_sku_inventory (for branch ops) and fg_inventory (for inventory tab).
    """
    # Get branch_id for fg_inventory (matches inventory_routes query structure)
    branch_doc = await db.branches.find_one({"name": branch}, {"_id": 0, "branch_id": 1})
    branch_id = branch_doc.get("branch_id") if branch_doc else None
    
    # ========== Update branch_sku_inventory (single source of truth for FG) ==========
    await db.branch_sku_inventory.update_one(
        {"branch": branch, "buyer_sku_id": buyer_sku_id},
        {
            "$inc": {"current_stock": quantity},
            "$set": {"is_active": True, "updated_at": datetime.now(timezone.utc).isoformat()},
            "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc).isoformat()}
        },
        upsert=True
    )
    
    # Log the production entry
    await db.fg_production_log.insert_one({
        "id": str(uuid.uuid4()),
        "branch": branch,
        "buyer_sku_id": buyer_sku_id,
        "quantity": quantity,
        "schedule_code": schedule_code,
        "produced_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"buyer_sku_id": buyer_sku_id, "quantity_added": quantity, "branch": branch, "branch_id": branch_id}


@router.get("/branch-ops/my-branches")
async def get_my_branches(current_user: User = Depends(get_current_user)):
    """Get branches assigned to the logged-in user"""
    # Master admin can see all branches
    if current_user.role == "master_admin" or current_user.role == "MASTER_ADMIN":
        branches = await db.branches.find({"is_active": True}, {"_id": 0}).to_list(100)
        return {
            "user": current_user.name,
            "role": current_user.role,
            "branches": [b["name"] for b in branches],
            "is_all_access": True
        }
    
    # Branch ops user sees only assigned branches
    return {
        "user": current_user.name,
        "role": current_user.role,
        "branches": current_user.assigned_branches,
        "is_all_access": False
    }


@router.get("/branch-ops/schedules")
async def get_branch_schedules(
    date_filter: str = Query("today", description="Filter: today, week, month, custom"),
    start_date: Optional[str] = Query(None, description="Start date for custom filter (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date for custom filter (YYYY-MM-DD)"),
    branch: Optional[str] = Query(None, description="Specific branch filter (optional)"),
    status: Optional[str] = Query(None, description="Status filter: SCHEDULED, COMPLETED, CANCELLED"),
    current_user: User = Depends(get_current_user)
):
    """
    Get production schedules for branch ops user.
    - Master admin sees all branches (or can filter by specific branch)
    - Branch ops user sees ONLY their assigned branches
    """
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Determine date range based on filter
    if date_filter == "today":
        start = today
        end = today + timedelta(days=1)
    elif date_filter == "week":
        # Start from Monday of current week
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=7)
    elif date_filter == "month":
        start = today.replace(day=1)
        # End of month
        if today.month == 12:
            end = today.replace(year=today.year + 1, month=1, day=1)
        else:
            end = today.replace(month=today.month + 1, day=1)
    elif date_filter == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date and end_date required for custom filter")
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        raise HTTPException(status_code=400, detail="Invalid date_filter. Use: today, week, month, custom")
    
    # Determine which branches user can access
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    
    if is_master:
        # Master admin can see all or filter by specific branch
        if branch:
            user_branches = [branch]
        else:
            branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1}).to_list(100)
            user_branches = [b["name"] for b in branches]
    else:
        # Branch ops user - only their assigned branches
        user_branches = current_user.assigned_branches
        if not user_branches:
            return {
                "message": "No branches assigned to your account. Contact admin.",
                "schedules": [],
                "summary": {"total": 0, "scheduled": 0, "completed": 0}
            }
        # If specific branch requested, verify user has access
        if branch:
            if branch not in user_branches:
                raise HTTPException(status_code=403, detail=f"Access denied to branch: {branch}")
            user_branches = [branch]
    
    # Build query - always exclude DELETED schedules
    query = {
        "branch": {"$in": user_branches},
        "target_date": {"$gte": start, "$lt": end},
        "status": {"$nin": ["DELETED"]}  # Never show deleted schedules
    }
    
    if status:
        # Override with specific status filter (still excludes DELETED)
        query["status"] = status
    
    # Fetch schedules
    schedules = await db.production_schedules.find(query, {"_id": 0}).sort("target_date", 1).to_list(1000)
    
    # Enrich with SKU details
    sku_ids = list(set(s.get("sku_id") for s in schedules if s.get("sku_id")))
    skus = await sku_service.get_skus_by_sku_ids(sku_ids)
    sku_map = {s["sku_id"]: s for s in skus}
    
    enriched_schedules = []
    for schedule in schedules:
        sku = sku_map.get(schedule.get("sku_id"), {})
        schedule["sku_details"] = {
            "description": sku.get("description", ""),
            "model": sku.get("model", ""),
            "brand": sku.get("brand", ""),
            "vertical": sku.get("vertical", "")
        }
        enriched_schedules.append(serialize_doc(schedule))
    
    # Calculate summary
    summary = {
        "total": len(schedules),
        "scheduled": len([s for s in schedules if s.get("status") == "SCHEDULED"]),
        "completed": len([s for s in schedules if s.get("status") == "COMPLETED"]),
        "cancelled": len([s for s in schedules if s.get("status") == "CANCELLED"]),
        "total_target_qty": sum(s.get("target_quantity", 0) for s in schedules),
        "total_completed_qty": sum(s.get("completed_quantity", 0) for s in schedules)
    }
    
    return {
        "user": current_user.name,
        "branches": user_branches,
        "date_filter": date_filter,
        "date_range": {
            "start": start.strftime("%Y-%m-%d"),
            "end": (end - timedelta(days=1)).strftime("%Y-%m-%d")
        },
        "schedules": enriched_schedules,
        "summary": summary
    }


@router.put("/branch-ops/schedules/{schedule_id}/complete")
async def complete_schedule(
    schedule_id: str,
    completed_quantity: int = Query(..., description="Actual quantity produced"),
    notes: Optional[str] = Query(None, description="Completion notes"),
    current_user: User = Depends(get_current_user)
):
    """
    Mark a production schedule as completed.
    
    This endpoint:
    1. Checks RM availability for the completed quantity
    2. If sufficient: Deducts RM from branch_rm_inventory, adds FG to branch_sku_inventory
    3. If insufficient: Returns error with list of shortage RMs
    
    Branch ops user can only complete schedules for their assigned branches.
    """
    schedule = await db.production_schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    branch = schedule.get("branch")
    buyer_sku_id = schedule.get("sku_id")
    schedule_code = schedule.get("schedule_code")
    
    # Check branch access
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    if not is_master:
        if branch not in current_user.assigned_branches:
            raise HTTPException(status_code=403, detail="Access denied. Schedule belongs to different branch.")
    
    if schedule.get("status") == "COMPLETED":
        raise HTTPException(status_code=400, detail="Schedule already completed")
    
    if schedule.get("status") == "CANCELLED":
        raise HTTPException(status_code=400, detail="Cannot complete a cancelled schedule")
    
    # ========== STEP 1: Check RM Availability ==========
    availability = await check_rm_availability_for_production(branch, buyer_sku_id, completed_quantity)
    
    if not availability["sufficient"]:
        # Return detailed shortage information
        shortage_details = []
        for s in availability["shortages"]:
            shortage_details.append({
                "rm_id": s["rm_id"],
                "description": s["description"],
                "unit": s["unit"],
                "required": s["required"],
                "available": s["available"],
                "shortage": s["shortage"]
            })
        
        raise HTTPException(
            status_code=400,
            detail={
                "error": "INSUFFICIENT_RM_STOCK",
                "message": f"Cannot complete production. {len(shortage_details)} RM(s) have insufficient stock.",
                "shortages": shortage_details
            }
        )
    
    # ========== STEP 2: Consume RM from Inventory ==========
    consumption_result = await consume_rm_for_production(
        branch=branch,
        buyer_sku_id=buyer_sku_id,
        quantity=completed_quantity,
        schedule_code=schedule_code
    )
    
    # ========== STEP 3: Add FG to Branch SKU Inventory ==========
    await add_fg_inventory(
        branch=branch,
        buyer_sku_id=buyer_sku_id,
        quantity=completed_quantity,
        schedule_code=schedule_code
    )
    
    # ========== STEP 4: Update Schedule Status ==========
    update_data = {
        "status": "COMPLETED",
        "completed_quantity": completed_quantity,
        "completed_at": datetime.now(timezone.utc),
        "completed_by": current_user.id,
        "completed_by_name": current_user.name,
        "rm_consumed": consumption_result["consumed"],
        "fg_added": True
    }
    
    if notes:
        update_data["completion_notes"] = notes
    
    await db.production_schedules.update_one(
        {"id": schedule_id},
        {"$set": update_data}
    )
    
    # Update dispatch lot if linked
    if schedule.get("dispatch_lot_id"):
        await db.dispatch_lots.update_one(
            {"id": schedule["dispatch_lot_id"]},
            {"$set": {"status": "FULLY_PRODUCED"}}
        )
    
    return {
        "message": f"Schedule {schedule_code} completed successfully",
        "schedule_id": schedule_id,
        "completed_quantity": completed_quantity,
        "completed_by": current_user.name,
        "rm_consumed": {
            "total_items": consumption_result["total_items"],
            "items": consumption_result["consumed"][:10]  # Show first 10 items
        },
        "fg_added": {
            "buyer_sku_id": buyer_sku_id,
            "quantity": completed_quantity,
            "branch": branch
        }
    }


@router.get("/branch-ops/schedules/{schedule_id}/check-rm")
async def check_rm_for_schedule(
    schedule_id: str,
    quantity: Optional[int] = Query(None, description="Quantity to check (defaults to target quantity)"),
    current_user: User = Depends(get_current_user)
):
    """
    Pre-check RM availability for a production schedule before marking complete.
    Returns availability status and any shortages.
    """
    schedule = await db.production_schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    branch = schedule.get("branch")
    buyer_sku_id = schedule.get("sku_id")
    check_qty = quantity or schedule.get("target_quantity", 0)
    
    # Check branch access
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    if not is_master:
        if branch not in current_user.assigned_branches:
            raise HTTPException(status_code=403, detail="Access denied")
    
    availability = await check_rm_availability_for_production(branch, buyer_sku_id, check_qty)
    
    return {
        "schedule_id": schedule_id,
        "schedule_code": schedule.get("schedule_code"),
        "branch": branch,
        "buyer_sku_id": buyer_sku_id,
        "quantity_checked": check_qty,
        "sufficient": availability["sufficient"],
        "shortages": availability["shortages"],
        "bom_items": len(availability["bom"]),
        "message": availability.get("message", "")
    }


@router.get("/branch-ops/dashboard")
async def get_branch_ops_dashboard(
    current_user: User = Depends(get_current_user)
):
    """
    Dashboard summary for branch ops user.
    Shows today's work, pending schedules, and completion stats.
    """
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=7)
    
    # Determine user's branches
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    if is_master:
        branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1}).to_list(100)
        user_branches = [b["name"] for b in branches]
    else:
        user_branches = current_user.assigned_branches
    
    if not user_branches:
        return {
            "message": "No branches assigned",
            "today": {"scheduled": 0, "completed": 0, "target_qty": 0, "completed_qty": 0},
            "week": {"scheduled": 0, "completed": 0},
            "branches": []
        }
    
    # Today's stats
    today_schedules = await db.production_schedules.find({
        "branch": {"$in": user_branches},
        "target_date": {"$gte": today, "$lt": tomorrow},
        "status": {"$ne": "CANCELLED"}
    }, {"_id": 0}).to_list(1000)
    
    today_stats = {
        "total": len(today_schedules),
        "scheduled": len([s for s in today_schedules if s.get("status") == "SCHEDULED"]),
        "completed": len([s for s in today_schedules if s.get("status") == "COMPLETED"]),
        "target_qty": sum(s.get("target_quantity", 0) for s in today_schedules),
        "completed_qty": sum(s.get("completed_quantity", 0) for s in today_schedules if s.get("status") == "COMPLETED")
    }
    
    # Week stats
    week_schedules = await db.production_schedules.find({
        "branch": {"$in": user_branches},
        "target_date": {"$gte": week_start, "$lt": week_end},
        "status": {"$ne": "CANCELLED"}
    }, {"_id": 0, "status": 1, "target_quantity": 1, "completed_quantity": 1}).to_list(1000)
    
    week_stats = {
        "total": len(week_schedules),
        "scheduled": len([s for s in week_schedules if s.get("status") == "SCHEDULED"]),
        "completed": len([s for s in week_schedules if s.get("status") == "COMPLETED"]),
        "target_qty": sum(s.get("target_quantity", 0) for s in week_schedules),
        "completed_qty": sum(s.get("completed_quantity", 0) for s in week_schedules if s.get("status") == "COMPLETED")
    }
    
    # Per-branch breakdown for today
    branch_breakdown = []
    for branch in user_branches:
        branch_today = [s for s in today_schedules if s.get("branch") == branch]
        branch_breakdown.append({
            "branch": branch,
            "total": len(branch_today),
            "pending": len([s for s in branch_today if s.get("status") == "SCHEDULED"]),
            "completed": len([s for s in branch_today if s.get("status") == "COMPLETED"]),
            "target_qty": sum(s.get("target_quantity", 0) for s in branch_today),
            "completed_qty": sum(s.get("completed_quantity", 0) for s in branch_today if s.get("status") == "COMPLETED")
        })
    
    return {
        "user": current_user.name,
        "branches": user_branches,
        "today": today_stats,
        "week": week_stats,
        "branch_breakdown": branch_breakdown,
        "date": today.strftime("%Y-%m-%d")
    }



@router.get("/rm-shortage-report")
async def get_rm_shortage_report(
    branch: Optional[str] = Query(None, description="Branch name (required for branch ops, optional for procurement)"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD), defaults to today"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD), defaults to today + 7 days"),
    current_user: User = Depends(get_current_user)
):
    """
    Calculate RM shortage report based on production schedules and BOM.
    
    Logic:
    1. Get current RM stock for the branch
    2. Calculate RM consumption for INTERIM period (today → start_date - 1) 
    3. Calculate RM requirement for SELECTED period (start_date → end_date)
    4. Shortage = Current Stock - Interim Consumption - Period Requirement
    
    BOM = common_bom (via bidso_sku_id) + brand_bom (via buyer_sku_id) merged
    """
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Parse dates with defaults (next 7 days)
    if not start_date:
        start_dt = today
    else:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    if not end_date:
        end_dt = today + timedelta(days=7)
    else:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    # Determine branches to query
    is_admin = current_user.role in ["master_admin", "MASTER_ADMIN"]
    is_procurement = "PROCUREMENT" in current_user.role.upper() if current_user.role else False
    
    if branch:
        branches_to_query = [branch]
    elif is_admin or is_procurement:
        # Get all active branches
        all_branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1}).to_list(100)
        branches_to_query = [b["name"] for b in all_branches]
    else:
        # Branch ops user - use assigned branches
        branches_to_query = current_user.assigned_branches or []
    
    if not branches_to_query:
        return {"error": "No branches available", "data": []}
    
    # Build BOM lookup: buyer_sku_id -> list of {rm_id, quantity}
    # Step 1: Get all buyer SKUs with their bidso_sku_id mapping
    buyer_skus = await db.buyer_skus.find(
        {"status": "ACTIVE"},
        {"_id": 0, "buyer_sku_id": 1, "bidso_sku_id": 1}
    ).to_list(10000)
    buyer_to_bidso = {s["buyer_sku_id"]: s.get("bidso_sku_id") for s in buyer_skus}
    
    # Step 2: Get all common BOMs
    common_boms = await db.common_bom.find({}, {"_id": 0, "bidso_sku_id": 1, "items": 1}).to_list(10000)
    bidso_to_bom = {b["bidso_sku_id"]: b.get("items", []) for b in common_boms}
    
    # Step 3: Get all brand BOMs
    brand_boms = await db.brand_bom.find({}, {"_id": 0, "buyer_sku_id": 1, "items": 1}).to_list(10000)
    buyer_to_brand_bom = {b["buyer_sku_id"]: b.get("items", []) for b in brand_boms}
    
    def get_merged_bom(buyer_sku_id):
        """Get merged BOM: common + brand-specific"""
        bidso_sku_id = buyer_to_bidso.get(buyer_sku_id)
        common_items = bidso_to_bom.get(bidso_sku_id, []) if bidso_sku_id else []
        brand_items = buyer_to_brand_bom.get(buyer_sku_id, [])
        
        # Merge: brand items override common items with same rm_id
        merged = {}
        for item in common_items:
            rm_id = item.get("rm_id")
            if rm_id:
                merged[rm_id] = item.get("quantity", 0)
        for item in brand_items:
            rm_id = item.get("rm_id")
            if rm_id:
                merged[rm_id] = merged.get(rm_id, 0) + item.get("quantity", 0)
        
        return merged  # {rm_id: quantity}
    
    # Calculate RM requirements for each branch
    result_by_branch = {}
    
    for branch_name in branches_to_query:
        # Get current RM stock for this branch
        branch_inventory = await db.branch_rm_inventory.find(
            {"branch": branch_name},
            {"_id": 0, "rm_id": 1, "current_stock": 1}
        ).to_list(15000)
        stock_map = {inv["rm_id"]: inv.get("current_stock", 0) for inv in branch_inventory}
        
        # Get production schedules for INTERIM period (today to start_date - 1)
        interim_end = start_dt - timedelta(days=1)
        interim_schedules = []
        if interim_end >= today:
            interim_schedules = await db.production_schedules.find(
                {
                    "branch": branch_name,
                    "target_date": {"$gte": today, "$lte": interim_end + timedelta(days=1)},
                    "status": {"$nin": ["CANCELLED", "COMPLETED"]}
                },
                {"_id": 0, "sku_id": 1, "target_quantity": 1}
            ).to_list(5000)
        
        # Get production schedules for SELECTED period (start_date to end_date)
        period_schedules = await db.production_schedules.find(
            {
                "branch": branch_name,
                "target_date": {"$gte": start_dt, "$lte": end_dt + timedelta(days=1)},
                "status": {"$nin": ["CANCELLED", "COMPLETED"]}
            },
            {"_id": 0, "sku_id": 1, "target_quantity": 1}
        ).to_list(5000)
        
        # Calculate RM consumption for interim period
        interim_consumption = {}
        for sched in interim_schedules:
            sku_id = sched.get("sku_id")
            qty = sched.get("target_quantity", 0)
            bom = get_merged_bom(sku_id)
            for rm_id, rm_qty in bom.items():
                interim_consumption[rm_id] = interim_consumption.get(rm_id, 0) + (rm_qty * qty)
        
        # Calculate RM requirement for selected period
        period_requirement = {}
        for sched in period_schedules:
            sku_id = sched.get("sku_id")
            qty = sched.get("target_quantity", 0)
            bom = get_merged_bom(sku_id)
            for rm_id, rm_qty in bom.items():
                period_requirement[rm_id] = period_requirement.get(rm_id, 0) + (rm_qty * qty)
        
        # Get all unique RM IDs
        all_rm_ids = set(interim_consumption.keys()) | set(period_requirement.keys())
        
        # Get RM details
        rm_details = {}
        if all_rm_ids:
            rms = await db.raw_materials.find(
                {"rm_id": {"$in": list(all_rm_ids)}},
                {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1, "unit": 1}
            ).to_list(5000)
            for rm in rms:
                rm_details[rm["rm_id"]] = {
                    "description": await generate_rm_description(
                        rm.get("category", ""), 
                        rm.get("category_data", {}), 
                        rm.get("name", "")
                    ),
                    "unit": rm.get("unit", ""),
                    "category": rm.get("category", "")
                }
        
        # Build result for this branch
        branch_data = []
        for rm_id in sorted(all_rm_ids):
            current = stock_map.get(rm_id, 0)
            interim = interim_consumption.get(rm_id, 0)
            required = period_requirement.get(rm_id, 0)
            projected = current - interim
            shortage = projected - required
            
            rm_info = rm_details.get(rm_id, {})
            branch_data.append({
                "rm_id": rm_id,
                "description": rm_info.get("description", ""),
                "unit": rm_info.get("unit", ""),
                "category": rm_info.get("category", ""),
                "current_stock": round(current, 2),
                "interim_consumption": round(interim, 2),
                "projected_stock": round(projected, 2),
                "period_requirement": round(required, 2),
                "shortage": round(shortage, 2),
                "is_shortage": shortage < 0
            })
        
        result_by_branch[branch_name] = {
            "branch": branch_name,
            "data": branch_data,
            "summary": {
                "total_rms": len(branch_data),
                "rms_in_shortage": len([d for d in branch_data if d["is_shortage"]]),
                "total_shortage_value": sum(abs(d["shortage"]) for d in branch_data if d["is_shortage"])
            }
        }
    
    # If single branch requested, return flat structure
    if branch and branch in result_by_branch:
        return {
            "branch": branch,
            "start_date": start_dt.strftime("%Y-%m-%d"),
            "end_date": end_dt.strftime("%Y-%m-%d"),
            "interim_period": f"{today.strftime('%Y-%m-%d')} to {(start_dt - timedelta(days=1)).strftime('%Y-%m-%d')}" if start_dt > today else "N/A",
            **result_by_branch[branch]
        }
    
    # Multi-branch: return all branches
    return {
        "start_date": start_dt.strftime("%Y-%m-%d"),
        "end_date": end_dt.strftime("%Y-%m-%d"),
        "interim_period": f"{today.strftime('%Y-%m-%d')} to {(start_dt - timedelta(days=1)).strftime('%Y-%m-%d')}" if start_dt > today else "N/A",
        "branches": list(result_by_branch.values()),
        "overall_summary": {
            "total_branches": len(result_by_branch),
            "branches_with_shortage": len([b for b in result_by_branch.values() if b["summary"]["rms_in_shortage"] > 0])
        }
    }


@router.get("/rm-shortage-report/export")
async def export_rm_shortage_report(
    branch: Optional[str] = Query(None, description="Branch name"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user)
):
    """
    Export RM shortage report as Excel with consumption pattern.
    Shows: Buyer SKU ID | RM ID | RM Description | Total Qty Required | Qty Shortage
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io
    
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Parse dates with defaults (next 7 days)
    if not start_date:
        start_dt = today
    else:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    if not end_date:
        end_dt = today + timedelta(days=7)
    else:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    # Determine branches to query
    is_admin = current_user.role in ["master_admin", "MASTER_ADMIN"]
    is_procurement = "PROCUREMENT" in current_user.role.upper() if current_user.role else False
    
    if branch:
        branches_to_query = [branch]
    elif is_admin or is_procurement:
        all_branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1}).to_list(100)
        branches_to_query = [b["name"] for b in all_branches]
    else:
        branches_to_query = current_user.assigned_branches or []
    
    if not branches_to_query:
        raise HTTPException(status_code=400, detail="No branches available")
    
    # Build BOM lookup
    buyer_skus = await db.buyer_skus.find(
        {"status": "ACTIVE"},
        {"_id": 0, "buyer_sku_id": 1, "bidso_sku_id": 1, "name": 1}
    ).to_list(10000)
    buyer_to_bidso = {s["buyer_sku_id"]: s.get("bidso_sku_id") for s in buyer_skus}
    buyer_sku_names = {s["buyer_sku_id"]: s.get("name", "") for s in buyer_skus}
    
    common_boms = await db.common_bom.find({}, {"_id": 0, "bidso_sku_id": 1, "items": 1}).to_list(10000)
    bidso_to_bom = {b["bidso_sku_id"]: b.get("items", []) for b in common_boms}
    
    brand_boms = await db.brand_bom.find({}, {"_id": 0, "buyer_sku_id": 1, "items": 1}).to_list(10000)
    buyer_to_brand_bom = {b["buyer_sku_id"]: b.get("items", []) for b in brand_boms}
    
    def get_merged_bom(buyer_sku_id):
        bidso_sku_id = buyer_to_bidso.get(buyer_sku_id)
        common_items = bidso_to_bom.get(bidso_sku_id, []) if bidso_sku_id else []
        brand_items = buyer_to_brand_bom.get(buyer_sku_id, [])
        
        merged = {}
        for item in common_items:
            rm_id = item.get("rm_id")
            if rm_id:
                merged[rm_id] = item.get("quantity", 0)
        for item in brand_items:
            rm_id = item.get("rm_id")
            if rm_id:
                merged[rm_id] = merged.get(rm_id, 0) + item.get("quantity", 0)
        return merged
    
    # Get all RM details with category_data for proper description
    all_rms = await db.raw_materials.find({}, {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1, "unit": 1}).to_list(15000)
    rm_details = {}
    for rm in all_rms:
        rm_details[rm["rm_id"]] = {
            "rm_id": rm["rm_id"],
            "description": await generate_rm_description(
                rm.get("category", ""),
                rm.get("category_data", {}),
                rm.get("name", "")
            ),
            "unit": rm.get("unit", "")
        }
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RM Consumption Pattern"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    shortage_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Headers for consumption pattern
    headers = ["Branch", "Buyer SKU ID", "SKU Name", "RM ID", "RM Description", "Unit", 
               "RM Qty per Unit", "SKU Production Qty", "Total RM Required", "Current Stock", "Shortage"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    row_num = 2
    
    for branch_name in branches_to_query:
        # Get branch RM stock
        branch_inventory = await db.branch_rm_inventory.find(
            {"branch": branch_name},
            {"_id": 0, "rm_id": 1, "current_stock": 1}
        ).to_list(15000)
        stock_map = {inv["rm_id"]: inv.get("current_stock", 0) for inv in branch_inventory}
        
        # Get production schedules for the period
        period_schedules = await db.production_schedules.find(
            {
                "branch": branch_name,
                "target_date": {"$gte": start_dt, "$lte": end_dt + timedelta(days=1)},
                "status": {"$nin": ["CANCELLED", "COMPLETED"]}
            },
            {"_id": 0, "sku_id": 1, "target_quantity": 1}
        ).to_list(5000)
        
        # Aggregate by SKU
        sku_qty_map = {}
        for sched in period_schedules:
            sku_id = sched.get("sku_id")
            qty = sched.get("target_quantity", 0)
            sku_qty_map[sku_id] = sku_qty_map.get(sku_id, 0) + qty
        
        # Calculate RM requirements per SKU and track totals
        rm_totals = {}  # rm_id -> total required
        
        for sku_id, sku_qty in sku_qty_map.items():
            bom = get_merged_bom(sku_id)
            sku_name = buyer_sku_names.get(sku_id, "")
            
            for rm_id, rm_per_unit in bom.items():
                total_rm_needed = rm_per_unit * sku_qty
                rm_totals[rm_id] = rm_totals.get(rm_id, 0) + total_rm_needed
                
                rm_info = rm_details.get(rm_id, {})
                current_stock = stock_map.get(rm_id, 0)
                
                ws.cell(row=row_num, column=1, value=branch_name)
                ws.cell(row=row_num, column=2, value=sku_id)
                ws.cell(row=row_num, column=3, value=sku_name)
                ws.cell(row=row_num, column=4, value=rm_id)
                ws.cell(row=row_num, column=5, value=rm_info.get("description", ""))
                ws.cell(row=row_num, column=6, value=rm_info.get("unit", ""))
                ws.cell(row=row_num, column=7, value=rm_per_unit)
                ws.cell(row=row_num, column=8, value=sku_qty)
                ws.cell(row=row_num, column=9, value=total_rm_needed)
                ws.cell(row=row_num, column=10, value=current_stock)
                
                # Calculate shortage (will be computed after aggregation in summary sheet)
                # For this row, we show per-SKU contribution
                ws.cell(row=row_num, column=11, value="")  # Will be in summary sheet
                
                row_num += 1
    
    # Create Summary sheet with aggregated shortages
    ws_summary = wb.create_sheet("RM Shortage Summary")
    summary_headers = ["Branch", "RM ID", "RM Description", "Unit", "Total Qty Required", "Current Stock", "Qty Shortage"]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    summary_row = 2
    
    for branch_name in branches_to_query:
        # Get branch RM stock
        branch_inventory = await db.branch_rm_inventory.find(
            {"branch": branch_name},
            {"_id": 0, "rm_id": 1, "current_stock": 1}
        ).to_list(15000)
        stock_map = {inv["rm_id"]: inv.get("current_stock", 0) for inv in branch_inventory}
        
        # Get production schedules
        period_schedules = await db.production_schedules.find(
            {
                "branch": branch_name,
                "target_date": {"$gte": start_dt, "$lte": end_dt + timedelta(days=1)},
                "status": {"$nin": ["CANCELLED", "COMPLETED"]}
            },
            {"_id": 0, "sku_id": 1, "target_quantity": 1}
        ).to_list(5000)
        
        # Aggregate RM requirements
        rm_totals = {}
        for sched in period_schedules:
            sku_id = sched.get("sku_id")
            qty = sched.get("target_quantity", 0)
            bom = get_merged_bom(sku_id)
            for rm_id, rm_per_unit in bom.items():
                rm_totals[rm_id] = rm_totals.get(rm_id, 0) + (rm_per_unit * qty)
        
        # Write summary rows
        for rm_id in sorted(rm_totals.keys()):
            total_required = rm_totals[rm_id]
            current_stock = stock_map.get(rm_id, 0)
            shortage = current_stock - total_required
            
            rm_info = rm_details.get(rm_id, {})
            
            ws_summary.cell(row=summary_row, column=1, value=branch_name)
            ws_summary.cell(row=summary_row, column=2, value=rm_id)
            ws_summary.cell(row=summary_row, column=3, value=rm_info.get("description", ""))
            ws_summary.cell(row=summary_row, column=4, value=rm_info.get("unit", ""))
            ws_summary.cell(row=summary_row, column=5, value=round(total_required, 2))
            ws_summary.cell(row=summary_row, column=6, value=round(current_stock, 2))
            ws_summary.cell(row=summary_row, column=7, value=round(shortage, 2))
            
            # Highlight shortage rows
            if shortage < 0:
                for c in range(1, 8):
                    ws_summary.cell(row=summary_row, column=c).fill = shortage_fill
            
            summary_row += 1
    
    # Auto-width columns
    col_widths = [18, 18, 30, 12, 30, 8, 15, 18, 18, 15, 12]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    summary_widths = [18, 12, 35, 8, 18, 15, 15]
    for idx, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rm_shortage_report_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ============ OVERDUE SCHEDULE HANDLING ============

@router.get("/branch-ops/overdue-schedules")
async def get_overdue_schedules(
    branch: Optional[str] = Query(None, description="Filter by branch"),
    current_user: User = Depends(get_current_user)
):
    """
    Get all overdue production schedules (target_date < today AND status = SCHEDULED).
    Returns list of overdue schedules with days overdue.
    """
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Determine user's branches
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    if is_master:
        if branch:
            user_branches = [branch]
        else:
            branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1}).to_list(100)
            user_branches = [b["name"] for b in branches]
    else:
        user_branches = current_user.assigned_branches or []
        if branch and branch in user_branches:
            user_branches = [branch]
    
    if not user_branches:
        return {"overdue": [], "count": 0}
    
    # Find overdue schedules
    overdue_schedules = await db.production_schedules.find(
        {
            "branch": {"$in": user_branches},
            "target_date": {"$lt": today},
            "status": "SCHEDULED"
        },
        {"_id": 0}
    ).sort("target_date", 1).to_list(500)
    
    # Enrich with days overdue and SKU details
    sku_ids = list(set(s.get("sku_id") for s in overdue_schedules if s.get("sku_id")))
    skus = await db.buyer_skus.find({"buyer_sku_id": {"$in": sku_ids}}, {"_id": 0, "buyer_sku_id": 1, "name": 1}).to_list(1000)
    sku_map = {s["buyer_sku_id"]: s.get("name", "") for s in skus}
    
    enriched = []
    for schedule in overdue_schedules:
        target_date = schedule.get("target_date")
        if isinstance(target_date, str):
            target_date = datetime.fromisoformat(target_date)
        if target_date.tzinfo is None:
            target_date = target_date.replace(tzinfo=timezone.utc)
        
        days_overdue = (today - target_date).days
        
        enriched.append({
            **serialize_doc(schedule),
            "days_overdue": days_overdue,
            "sku_name": sku_map.get(schedule.get("sku_id"), ""),
            "is_critical": days_overdue >= 3  # Flag if 3+ days overdue
        })
    
    return {
        "overdue": enriched,
        "count": len(enriched),
        "by_branch": {
            branch: len([s for s in enriched if s.get("branch") == branch])
            for branch in set(s.get("branch") for s in enriched)
        }
    }


@router.get("/branch-ops/overdue-count")
async def get_overdue_count(
    current_user: User = Depends(get_current_user)
):
    """Quick count of overdue schedules for dashboard badge"""
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    if is_master:
        branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1}).to_list(100)
        user_branches = [b["name"] for b in branches]
    else:
        user_branches = current_user.assigned_branches or []
    
    if not user_branches:
        return {"count": 0, "critical": 0}
    
    # Count overdue
    count = await db.production_schedules.count_documents({
        "branch": {"$in": user_branches},
        "target_date": {"$lt": today},
        "status": "SCHEDULED"
    })
    
    # Count critical (3+ days overdue)
    critical_date = today - timedelta(days=3)
    critical = await db.production_schedules.count_documents({
        "branch": {"$in": user_branches},
        "target_date": {"$lt": critical_date},
        "status": "SCHEDULED"
    })
    
    return {"count": count, "critical": critical}


from pydantic import BaseModel

class RescheduleRequest(BaseModel):
    schedule_ids: List[str]
    new_date: str  # YYYY-MM-DD format
    notes: Optional[str] = None


@router.post("/branch-ops/reschedule")
async def reschedule_schedules(
    request: RescheduleRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Reschedule one or more production schedules to a new date.
    Used for handling overdue schedules.
    """
    try:
        new_date = datetime.strptime(request.new_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    if new_date < today:
        raise HTTPException(status_code=400, detail="Cannot reschedule to a past date")
    
    # Get schedules
    schedules = await db.production_schedules.find(
        {"id": {"$in": request.schedule_ids}, "status": "SCHEDULED"},
        {"_id": 0}
    ).to_list(100)
    
    if not schedules:
        raise HTTPException(status_code=404, detail="No valid schedules found")
    
    # Verify user has access to all branches
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    if not is_master:
        for schedule in schedules:
            if schedule.get("branch") not in current_user.assigned_branches:
                raise HTTPException(
                    status_code=403, 
                    detail=f"Access denied to branch: {schedule.get('branch')}"
                )
    
    # Update schedules
    rescheduled = []
    for schedule in schedules:
        old_date = schedule.get("target_date")
        if isinstance(old_date, datetime):
            old_date_str = old_date.strftime("%Y-%m-%d")
        else:
            old_date_str = str(old_date)[:10]
        
        update_data = {
            "target_date": new_date,
            "rescheduled_from": old_date_str,
            "rescheduled_at": datetime.now(timezone.utc).isoformat(),
            "rescheduled_by": current_user.name
        }
        
        if request.notes:
            update_data["reschedule_notes"] = request.notes
        
        await db.production_schedules.update_one(
            {"id": schedule["id"]},
            {"$set": update_data}
        )
        
        rescheduled.append({
            "schedule_code": schedule.get("schedule_code"),
            "old_date": old_date_str,
            "new_date": request.new_date
        })
    
    return {
        "message": f"Rescheduled {len(rescheduled)} schedule(s) to {request.new_date}",
        "rescheduled": rescheduled
    }


# ============ PARTIAL COMPLETION / SPILLOVER HANDLING ============

class SpilloverRequest(BaseModel):
    parent_schedule_id: str
    spillover_quantity: int
    target_date: str  # YYYY-MM-DD format
    notes: Optional[str] = None


@router.post("/branch-ops/create-spillover")
async def create_spillover_schedule(
    request: SpilloverRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Create a spillover schedule for remaining quantity from partial completion.
    Links back to the parent schedule.
    """
    # Get parent schedule
    parent = await db.production_schedules.find_one(
        {"id": request.parent_schedule_id},
        {"_id": 0}
    )
    
    if not parent:
        raise HTTPException(status_code=404, detail="Parent schedule not found")
    
    # Verify access
    is_master = current_user.role in ["master_admin", "MASTER_ADMIN"]
    if not is_master:
        if parent.get("branch") not in current_user.assigned_branches:
            raise HTTPException(status_code=403, detail="Access denied to this branch")
    
    # Parse target date
    try:
        target_date = datetime.strptime(request.target_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    if target_date < today:
        raise HTTPException(status_code=400, detail="Cannot create spillover for past date")
    
    if request.spillover_quantity <= 0:
        raise HTTPException(status_code=400, detail="Spillover quantity must be greater than 0")
    
    # Generate new schedule code
    now = datetime.now(timezone.utc)
    prefix = f"PS_{now.strftime('%Y%m')}"
    last_schedule = await db.production_schedules.find_one(
        {"schedule_code": {"$regex": f"^{prefix}"}},
        sort=[("schedule_code", -1)]
    )
    
    if last_schedule:
        try:
            last_num = int(last_schedule["schedule_code"].split("_")[-1])
            new_num = last_num + 1
        except:
            new_num = 1
    else:
        new_num = 1
    
    schedule_code = f"{prefix}_{new_num:04d}"
    
    # Create spillover schedule
    spillover = {
        "id": str(uuid.uuid4()),
        "schedule_code": schedule_code,
        "branch": parent.get("branch"),
        "sku_id": parent.get("sku_id"),
        "sku_description": parent.get("sku_description", ""),
        "target_quantity": request.spillover_quantity,
        "allocated_quantity": request.spillover_quantity,
        "completed_quantity": 0,
        "target_date": target_date,
        "status": "SCHEDULED",
        "priority": parent.get("priority", "MEDIUM"),
        "parent_schedule_id": request.parent_schedule_id,
        "parent_schedule_code": parent.get("schedule_code"),
        "is_spillover": True,
        "notes": request.notes or f"Spillover from {parent.get('schedule_code')}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id,
        "created_by_name": current_user.name
    }
    
    await db.production_schedules.insert_one(spillover)
    
    # Update parent to track spillover
    await db.production_schedules.update_one(
        {"id": request.parent_schedule_id},
        {"$set": {
            "has_spillover": True,
            "spillover_schedule_id": spillover["id"],
            "spillover_quantity": request.spillover_quantity
        }}
    )
    
    return {
        "message": f"Spillover schedule {schedule_code} created",
        "schedule": {
            "id": spillover["id"],
            "schedule_code": schedule_code,
            "branch": spillover["branch"],
            "sku_id": spillover["sku_id"],
            "quantity": request.spillover_quantity,
            "target_date": request.target_date,
            "parent_schedule_code": parent.get("schedule_code")
        }
    }


@router.get("/branch-ops/schedules/{schedule_id}/spillovers")
async def get_schedule_spillovers(
    schedule_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all spillover schedules linked to a parent schedule"""
    spillovers = await db.production_schedules.find(
        {"parent_schedule_id": schedule_id},
        {"_id": 0}
    ).to_list(100)
    
    return {
        "parent_schedule_id": schedule_id,
        "spillovers": [serialize_doc(s) for s in spillovers],
        "count": len(spillovers)
    }
