"""RM Price History Routes.

Upload historical RM purchase prices, compute rolling 3-month average per RM,
and derive BOM cost for Buyer SKUs along with margin reporting.

Upload format (Excel):
    Date | Invoice No | Vendor ID | RM ID | Price (per unit)
"""
from fastapi import APIRouter, HTTPException, File, UploadFile, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from typing import Optional, List
import uuid
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel

from database import db

router = APIRouter(prefix="/rm-prices", tags=["RM Prices"])


# ==================== Helpers ====================

def _parse_date(raw) -> Optional[datetime]:
    """Accept Excel date cell, datetime, or string in common formats."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.replace(tzinfo=timezone.utc) if raw.tzinfo is None else raw
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _rolling_window_start(months: int = 3) -> datetime:
    """Return UTC datetime representing 'months' months before now (approx 30 days each)."""
    return datetime.now(timezone.utc) - timedelta(days=30 * months)


async def compute_avg_prices(window_months: int = 3) -> dict:
    """Return { rm_id: {avg_price, record_count, latest_date, latest_price} } using a rolling window."""
    start = _rolling_window_start(window_months)
    pipeline = [
        {"$match": {"date": {"$gte": start}}},
        {"$group": {
            "_id": "$rm_id",
            "avg_price": {"$avg": "$price_per_unit"},
            "record_count": {"$sum": 1},
            "latest_date": {"$max": "$date"},
            "latest_price": {"$last": "$price_per_unit"},
        }},
    ]
    result = {}
    async for row in db.rm_prices_history.aggregate(pipeline):
        rm_id = row["_id"]
        if not rm_id:
            continue
        result[rm_id] = {
            "rm_id": rm_id,
            "avg_price": round(row["avg_price"] or 0, 4),
            "record_count": row["record_count"],
            "latest_date": row["latest_date"].isoformat() if row.get("latest_date") else None,
            "latest_price": round(row.get("latest_price") or 0, 4),
        }
    return result


async def compute_avg_with_fallback(rm_ids: list, window_months: int = 3) -> dict:
    """For each RM, return avg price with source attribution.

    Priority:
      1. rm_prices_history (last `window_months` months) -> source='invoice'
      2. vendor_rm_prices (lowest tagged vendor price) -> source='vendor_map'
      3. None -> source=None

    Returns { rm_id: {avg_price, source, record_count, vendor_count} }
    """
    if not rm_ids:
        return {}

    # 1. Invoice history (preferred)
    start = _rolling_window_start(window_months)
    invoice_pipeline = [
        {"$match": {"date": {"$gte": start}, "rm_id": {"$in": list(rm_ids)}}},
        {"$group": {
            "_id": "$rm_id",
            "avg_price": {"$avg": "$price_per_unit"},
            "record_count": {"$sum": 1},
        }},
    ]
    result = {}
    async for row in db.rm_prices_history.aggregate(invoice_pipeline):
        rm_id = row["_id"]
        if not rm_id:
            continue
        result[rm_id] = {
            "rm_id": rm_id,
            "avg_price": round(row["avg_price"] or 0, 4),
            "source": "invoice",
            "record_count": row["record_count"],
            "vendor_count": 0,
        }

    # 2. Fallback to vendor_rm_prices for RMs not yet covered
    missing = [r for r in rm_ids if r not in result]
    if missing:
        vendor_pipeline = [
            {"$match": {"rm_id": {"$in": missing}, "is_active": {"$ne": False}}},
            {"$group": {
                "_id": "$rm_id",
                "min_price": {"$min": "$price"},
                "vendor_count": {"$sum": 1},
            }},
        ]
        async for row in db.vendor_rm_prices.aggregate(vendor_pipeline):
            rm_id = row["_id"]
            if not rm_id:
                continue
            result[rm_id] = {
                "rm_id": rm_id,
                "avg_price": round(row["min_price"] or 0, 4),
                "source": "vendor_map",
                "record_count": 0,
                "vendor_count": row["vendor_count"],
            }

    # 3. Mark remaining as no data
    for rm_id in rm_ids:
        if rm_id not in result:
            result[rm_id] = {
                "rm_id": rm_id,
                "avg_price": 0,
                "source": None,
                "record_count": 0,
                "vendor_count": 0,
            }

    return result


async def get_bom_cost_for_buyer_sku(buyer_sku_id: str, avg_price_map: dict) -> dict:
    """Compute derived BOM cost for a Buyer SKU using supplied avg_price map."""
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        return None

    bidso_sku_id = buyer_sku.get("bidso_sku_id")
    brand_id = buyer_sku.get("brand_id")

    common_bom = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0}) if bidso_sku_id else None
    brand_bom = await db.brand_specific_bom.find_one(
        {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id}, {"_id": 0}
    ) if (bidso_sku_id and brand_id) else None

    items = []
    total = 0.0
    missing = 0

    def _add_items(src_items, source_label):
        nonlocal total, missing
        for it in src_items or []:
            rm_id = it.get("rm_id")
            qty = float(it.get("quantity") or 0)
            price_info = avg_price_map.get(rm_id)
            avg_price = price_info["avg_price"] if price_info else 0
            line_cost = round(qty * avg_price, 4)
            if not price_info:
                missing += 1
            total += line_cost
            items.append({
                "rm_id": rm_id,
                "rm_name": it.get("rm_name"),
                "quantity": qty,
                "unit": it.get("unit", "PCS"),
                "avg_price": avg_price,
                "line_cost": line_cost,
                "source": source_label,
                "has_price": price_info is not None,
            })

    _add_items(common_bom.get("items", []) if common_bom else [], "common")
    _add_items(brand_bom.get("items", []) if brand_bom else [], "brand_specific")

    return {
        "buyer_sku_id": buyer_sku_id,
        "bidso_sku_id": bidso_sku_id,
        "total_cost": round(total, 4),
        "items": items,
        "rm_count": len(items),
        "missing_price_count": missing,
    }


# ==================== Upload ====================

@router.post("/upload")
async def upload_rm_prices(
    file: UploadFile = File(...),
    mode: str = Query("append", description="'append' or 'overwrite' (deletes all existing before insert)"),
):
    """Upload RM price invoices.

    Columns: Date | Invoice No | Vendor ID | RM ID | Price
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel: {e}")

    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers_raw = [str(c.value or "").strip().lower() for c in header_row]

    col = {}
    for idx, h in enumerate(headers_raw):
        if h in ("date", "invoice date", "purchase date"):
            col["date"] = idx
        elif h in ("invoice no", "invoice_no", "invoice number", "invoice#", "invoice"):
            col["invoice_no"] = idx
        elif h in ("vendor id", "vendor_id", "vendor code", "supplier id"):
            col["vendor_id"] = idx
        elif h in ("rm id", "rm_id", "material id", "raw material id"):
            col["rm_id"] = idx
        elif h in ("price", "price per unit", "unit price", "rate"):
            col["price"] = idx

    required = ["date", "invoice_no", "vendor_id", "rm_id", "price"]
    missing = [k for k in required if k not in col]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing columns: {missing}. Found: {headers_raw}",
        )

    # Pre-load vendor and RM maps for validation + enrichment
    vendors = await db.vendors.find({}, {"_id": 0, "vendor_id": 1, "name": 1}).to_list(5000)
    vendor_map = {v.get("vendor_id"): v.get("name") for v in vendors if v.get("vendor_id")}
    rm_ids_known = set(await db.raw_materials.distinct("rm_id"))

    rows_to_insert = []
    errors = []
    now = datetime.now(timezone.utc)

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        date_raw = row[col["date"]]
        invoice_no = str(row[col["invoice_no"]] or "").strip()
        vendor_id = str(row[col["vendor_id"]] or "").strip()
        rm_id = str(row[col["rm_id"]] or "").strip()
        price_raw = row[col["price"]]

        if not invoice_no or not vendor_id or not rm_id:
            errors.append(f"Row {idx}: Missing invoice/vendor/rm_id")
            continue

        parsed_date = _parse_date(date_raw)
        if not parsed_date:
            errors.append(f"Row {idx}: Invalid date '{date_raw}'")
            continue

        try:
            price = float(price_raw)
        except (TypeError, ValueError):
            errors.append(f"Row {idx}: Invalid price '{price_raw}'")
            continue
        if price <= 0:
            errors.append(f"Row {idx}: Price must be > 0 (got {price})")
            continue

        if rm_id not in rm_ids_known:
            errors.append(f"Row {idx}: RM ID '{rm_id}' not found in system")
            continue

        vendor_name = vendor_map.get(vendor_id)
        if not vendor_name:
            errors.append(f"Row {idx}: Vendor ID '{vendor_id}' not found in system")
            continue

        rows_to_insert.append({
            "id": str(uuid.uuid4()),
            "date": parsed_date,
            "invoice_no": invoice_no,
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "rm_id": rm_id,
            "price_per_unit": round(price, 4),
            "month_key": f"{parsed_date.year}-{parsed_date.month:02d}",
            "uploaded_at": now,
        })

    deleted = 0
    if mode == "overwrite":
        res = await db.rm_prices_history.delete_many({})
        deleted = res.deleted_count

    if rows_to_insert:
        await db.rm_prices_history.insert_many(rows_to_insert)

    return {
        "message": f"Inserted {len(rows_to_insert)} price records",
        "inserted": len(rows_to_insert),
        "deleted_previous": deleted,
        "errors": errors[:50],
        "error_count": len(errors),
    }


@router.get("/template")
async def download_rm_price_template():
    """Download Excel template for RM price upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RM Prices"

    headers = ["Date", "Invoice No", "Vendor ID", "RM ID", "Price"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    samples = [
        ["2026-01-15", "INV-2026-001", "VND001", "INP_1001", 125.50],
        ["2026-01-28", "INV-2026-007", "VND002", "INP_1001", 128.00],
        ["2026-02-05", "INV-2026-015", "VND001", "ACC_001", 42.75],
    ]
    for r, row in enumerate(samples, 2):
        for cidx, val in enumerate(row, 1):
            ws.cell(row=r, column=cidx, value=val)

    for col_letter, width in zip("ABCDE", [14, 20, 14, 14, 12]):
        ws.column_dimensions[col_letter].width = width

    ws_info = wb.create_sheet("Instructions")
    lines = [
        "RM Price Upload Template",
        "",
        "Required columns (first row is header):",
        "  - Date: Invoice/Purchase date (YYYY-MM-DD, DD-MM-YYYY, or Excel date)",
        "  - Invoice No: Unique invoice number from vendor",
        "  - Vendor ID: Existing Vendor ID in the system (see 'Vendors' tab)",
        "  - RM ID: Existing Raw Material ID (e.g., INP_1001)",
        "  - Price: Price PER UNIT (numeric, > 0)",
        "",
        "Notes:",
        "  - Upload the last 3 months of invoices.",
        "  - System computes a simple 3-month rolling average price per RM.",
        "  - Average price is then multiplied by BOM quantity to derive Buyer SKU BOM cost.",
        "  - Rows with unknown Vendor ID or RM ID are rejected.",
    ]
    for i, line in enumerate(lines, 1):
        ws_info.cell(row=i, column=1, value=line)
    ws_info.column_dimensions["A"].width = 90

    # Vendors reference tab
    ws_vendors = wb.create_sheet("Vendors")
    ws_vendors.cell(row=1, column=1, value="Vendor ID").font = Font(bold=True)
    ws_vendors.cell(row=1, column=2, value="Vendor Name").font = Font(bold=True)
    ws_vendors.cell(row=1, column=1).fill = header_fill
    ws_vendors.cell(row=1, column=2).fill = header_fill
    ws_vendors.cell(row=1, column=1).font = header_font
    ws_vendors.cell(row=1, column=2).font = header_font
    vendors = await db.vendors.find({}, {"_id": 0, "vendor_id": 1, "name": 1}).sort("vendor_id", 1).to_list(5000)
    for idx, v in enumerate(vendors, 2):
        ws_vendors.cell(row=idx, column=1, value=v.get("vendor_id", ""))
        ws_vendors.cell(row=idx, column=2, value=v.get("name", ""))
    ws_vendors.column_dimensions["A"].width = 16
    ws_vendors.column_dimensions["B"].width = 50

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rm_price_upload_template.xlsx"},
    )


# ==================== Stats & History ====================

@router.get("/stats")
async def rm_price_stats():
    total = await db.rm_prices_history.count_documents({})
    if total == 0:
        return {
            "total_records": 0,
            "unique_rms": 0,
            "unique_vendors": 0,
            "date_range": None,
            "window_start": _rolling_window_start().isoformat(),
            "rms_with_avg_price": 0,
        }
    unique_rms = len(await db.rm_prices_history.distinct("rm_id"))
    unique_vendors = len(await db.rm_prices_history.distinct("vendor_id"))

    pipeline = [{"$group": {"_id": None, "min_date": {"$min": "$date"}, "max_date": {"$max": "$date"}}}]
    agg = await db.rm_prices_history.aggregate(pipeline).to_list(1)
    min_d = agg[0]["min_date"] if agg else None
    max_d = agg[0]["max_date"] if agg else None

    avg_map = await compute_avg_prices()

    return {
        "total_records": total,
        "unique_rms": unique_rms,
        "unique_vendors": unique_vendors,
        "date_range": {
            "min": min_d.isoformat() if min_d else None,
            "max": max_d.isoformat() if max_d else None,
        },
        "window_start": _rolling_window_start().isoformat(),
        "rms_with_avg_price": len(avg_map),
    }


@router.get("/avg-prices")
async def get_avg_prices(window_months: int = Query(3, ge=1, le=24)):
    """Return avg price per RM id (last N months rolling window)."""
    avg_map = await compute_avg_prices(window_months)

    # Enrich with RM name/category
    rm_ids = list(avg_map.keys())
    rm_docs = await db.raw_materials.find(
        {"rm_id": {"$in": rm_ids}},
        {"_id": 0, "rm_id": 1, "name": 1, "category": 1}
    ).to_list(len(rm_ids) or 1) if rm_ids else []
    rm_info = {r["rm_id"]: r for r in rm_docs}

    rows = []
    for rm_id, info in avg_map.items():
        rm = rm_info.get(rm_id, {})
        rows.append({
            **info,
            "rm_name": rm.get("name"),
            "category": rm.get("category"),
        })
    rows.sort(key=lambda r: r["rm_id"])

    return {
        "window_months": window_months,
        "window_start": _rolling_window_start(window_months).isoformat(),
        "count": len(rows),
        "items": rows,
    }


@router.get("/history")
async def get_price_history(
    rm_id: Optional[str] = None,
    vendor_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=5000),
):
    """Paginated browse of RM price history."""
    q = {}
    if rm_id:
        q["rm_id"] = rm_id
    if vendor_id:
        q["vendor_id"] = vendor_id

    total = await db.rm_prices_history.count_documents(q)
    skip = (page - 1) * page_size
    cursor = db.rm_prices_history.find(q, {"_id": 0}).sort("date", -1).skip(skip).limit(page_size)
    items = []
    async for row in cursor:
        if isinstance(row.get("date"), datetime):
            row["date"] = row["date"].isoformat()
        if isinstance(row.get("uploaded_at"), datetime):
            row["uploaded_at"] = row["uploaded_at"].isoformat()
        items.append(row)

    total_pages = (total + page_size - 1) // page_size if total else 1
    return {"items": items, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}


@router.delete("/history")
async def clear_price_history():
    """Admin utility — delete all records."""
    res = await db.rm_prices_history.delete_many({})
    return {"deleted": res.deleted_count}


# ==================== BOM Cost & Margin ====================

@router.get("/bom-cost/{buyer_sku_id}")
async def bom_cost_for_sku(buyer_sku_id: str):
    """Derived BOM cost for a single Buyer SKU with per-RM breakdown."""
    avg_map = await compute_avg_prices()
    result = await get_bom_cost_for_buyer_sku(buyer_sku_id, avg_map)
    if result is None:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")

    # Enrich each line with RM name for display
    rm_ids = [it["rm_id"] for it in result["items"] if it.get("rm_id")]
    if rm_ids:
        rm_docs = await db.raw_materials.find(
            {"rm_id": {"$in": rm_ids}},
            {"_id": 0, "rm_id": 1, "name": 1, "category": 1}
        ).to_list(len(rm_ids))
        name_map = {r["rm_id"]: r for r in rm_docs}
        for it in result["items"]:
            rm = name_map.get(it["rm_id"], {})
            if not it.get("rm_name"):
                it["rm_name"] = rm.get("name")
            it["category"] = rm.get("category")
    return result


class BulkBomRequest(BaseModel):
    buyer_sku_ids: List[str]


@router.post("/bom-cost-bulk")
async def bulk_bom_cost(payload: BulkBomRequest):
    """Returns { buyer_sku_id: {total_cost, rm_count, missing_price_count} } — compact form."""
    avg_map = await compute_avg_prices()
    out = {}
    for sku_id in payload.buyer_sku_ids[:2000]:
        full = await get_bom_cost_for_buyer_sku(sku_id, avg_map)
        if full is None:
            continue
        out[sku_id] = {
            "total_cost": full["total_cost"],
            "rm_count": full["rm_count"],
            "missing_price_count": full["missing_price_count"],
        }
    return {"count": len(out), "costs": out}


@router.get("/margin-report")
async def margin_report(
    from_month: Optional[str] = Query(None, description="YYYY-MM"),
    to_month: Optional[str] = Query(None, description="YYYY-MM"),
    limit: int = Query(500, ge=1, le=5000),
):
    """
    Margin report joining historical_sales avg ASP with derived BOM cost per Buyer SKU.
    Margin % = (ASP - BOM Cost) / ASP * 100
    """
    match = {"asp": {"$gt": 0}}
    if from_month:
        match["month_key"] = {"$gte": from_month}
    if to_month:
        match.setdefault("month_key", {})["$lte"] = to_month

    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": "$buyer_sku_id",
            "avg_asp": {"$avg": "$asp"},
            "total_qty": {"$sum": "$qty"},
            "total_revenue": {"$sum": "$revenue"},
            "buyer_sku_name": {"$first": "$buyer_sku_name"},
            "bidso_sku_id": {"$first": "$bidso_sku_id"},
            "brand_code": {"$first": "$brand_code"},
            "vertical_code": {"$first": "$vertical_code"},
            "model_code": {"$first": "$model_code"},
            "records": {"$sum": 1},
        }},
        {"$sort": {"total_revenue": -1}},
        {"$limit": limit},
    ]

    rows = await db.historical_sales.aggregate(pipeline).to_list(limit)
    avg_map = await compute_avg_prices()

    report = []
    totals = {"revenue": 0.0, "cogs": 0.0, "qty": 0}

    for r in rows:
        buyer_sku_id = r["_id"]
        if not buyer_sku_id:
            continue
        bom = await get_bom_cost_for_buyer_sku(buyer_sku_id, avg_map)
        bom_cost = bom["total_cost"] if bom else 0
        avg_asp = round(r["avg_asp"] or 0, 2)
        margin_value = round(avg_asp - bom_cost, 2)
        margin_pct = round(((avg_asp - bom_cost) / avg_asp) * 100, 2) if avg_asp > 0 else 0

        total_qty = r["total_qty"] or 0
        total_revenue = round(r["total_revenue"] or 0, 2)
        total_cogs = round(bom_cost * total_qty, 2)

        totals["revenue"] += total_revenue
        totals["cogs"] += total_cogs
        totals["qty"] += total_qty

        report.append({
            "buyer_sku_id": buyer_sku_id,
            "buyer_sku_name": r.get("buyer_sku_name"),
            "bidso_sku_id": r.get("bidso_sku_id"),
            "brand_code": r.get("brand_code"),
            "vertical_code": r.get("vertical_code"),
            "model_code": r.get("model_code"),
            "avg_asp": avg_asp,
            "bom_cost": bom_cost,
            "margin_value": margin_value,
            "margin_pct": margin_pct,
            "total_qty": total_qty,
            "total_revenue": total_revenue,
            "total_cogs": total_cogs,
            "gross_profit": round(total_revenue - total_cogs, 2),
            "rm_count": bom["rm_count"] if bom else 0,
            "missing_price_count": bom["missing_price_count"] if bom else 0,
            "sales_records": r["records"],
        })

    overall_margin_pct = 0
    if totals["revenue"] > 0:
        overall_margin_pct = round(((totals["revenue"] - totals["cogs"]) / totals["revenue"]) * 100, 2)

    return {
        "items": report,
        "totals": {
            "total_qty": totals["qty"],
            "total_revenue": round(totals["revenue"], 2),
            "total_cogs": round(totals["cogs"], 2),
            "gross_profit": round(totals["revenue"] - totals["cogs"], 2),
            "overall_margin_pct": overall_margin_pct,
        },
        "count": len(report),
    }


# ==================== Buyer SKU BOM Cost (with fallback pricing) ====================

@router.get("/buyer-sku-cost-detail/{buyer_sku_id}")
async def buyer_sku_cost_detail(buyer_sku_id: str):
    """Return BOM with avg-price (3-mo rolling, vendor_rm_prices fallback) + ASP + Margin %."""
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        raise HTTPException(status_code=404, detail=f"Buyer SKU {buyer_sku_id} not found")

    bidso_sku_id = buyer_sku.get("bidso_sku_id")
    brand_id = buyer_sku.get("brand_id")

    common_bom = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0}) if bidso_sku_id else None
    brand_bom = await db.brand_specific_bom.find_one(
        {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id}, {"_id": 0}
    ) if (bidso_sku_id and brand_id) else None

    bom_items = []
    seen = set()
    for src_label, bom_doc in (("common", common_bom), ("brand_specific", brand_bom)):
        if not bom_doc:
            continue
        for it in bom_doc.get("items", []):
            if not isinstance(it, dict):
                continue
            rid = it.get("rm_id")
            if not rid or rid in seen:
                continue
            seen.add(rid)
            bom_items.append({
                "rm_id": rid,
                "quantity": float(it.get("quantity") or 0),
                "unit": it.get("unit", "PCS"),
                "bom_source": src_label,
            })

    # Fetch RM descriptions
    rm_ids = [b["rm_id"] for b in bom_items]
    rm_info = {}
    if rm_ids:
        async for rm in db.raw_materials.find(
            {"rm_id": {"$in": rm_ids}},
            {"_id": 0, "rm_id": 1, "name": 1, "description": 1, "category": 1, "category_data": 1}
        ):
            rm_info[rm["rm_id"]] = rm

    # Compute avg with fallback
    price_map = await compute_avg_with_fallback(rm_ids)

    items = []
    total_cost = 0.0
    missing = 0
    invoice_count = 0
    vendor_map_count = 0
    for b in bom_items:
        rid = b["rm_id"]
        rm = rm_info.get(rid, {})
        rm_name = rm.get("description") or rm.get("name") or rid
        pinfo = price_map.get(rid, {})
        avg_price = pinfo.get("avg_price", 0)
        source = pinfo.get("source")
        line_cost = round(b["quantity"] * avg_price, 4)
        total_cost += line_cost
        if source is None:
            missing += 1
        elif source == "invoice":
            invoice_count += 1
        elif source == "vendor_map":
            vendor_map_count += 1
        items.append({
            "rm_id": rid,
            "rm_name": rm_name,
            "rm_category": rm.get("category"),
            "quantity": b["quantity"],
            "unit": b["unit"],
            "avg_price": avg_price,
            "line_cost": line_cost,
            "price_source": source,
            "bom_source": b["bom_source"],
        })

    total_cost = round(total_cost, 2)

    # ASP from historical_sales
    asp_pipeline = [
        {"$match": {"buyer_sku_id": buyer_sku_id, "asp": {"$gt": 0}}},
        {"$group": {"_id": None, "avg_asp": {"$avg": "$asp"}, "qty": {"$sum": "$qty"}, "records": {"$sum": 1}}},
    ]
    asp_data = await db.historical_sales.aggregate(asp_pipeline).to_list(1)
    avg_asp = None
    margin_value = None
    margin_pct = None
    if asp_data:
        avg_asp = round(asp_data[0]["avg_asp"] or 0, 2)
        if avg_asp > 0:
            margin_value = round(avg_asp - total_cost, 2)
            margin_pct = round(((avg_asp - total_cost) / avg_asp) * 100, 2)

    return {
        "buyer_sku_id": buyer_sku_id,
        "buyer_sku_name": buyer_sku.get("name"),
        "bidso_sku_id": bidso_sku_id,
        "brand_id": brand_id,
        "brand_code": buyer_sku.get("brand_code"),
        "vertical_code": buyer_sku.get("vertical_code"),
        "model_code": buyer_sku.get("model_code"),
        "items": items,
        "total_cost": total_cost,
        "avg_asp": avg_asp,
        "margin_value": margin_value,
        "margin_pct": margin_pct,
        "rm_count": len(items),
        "invoice_count": invoice_count,
        "vendor_map_count": vendor_map_count,
        "missing_price_count": missing,
    }


@router.get("/buyer-sku-cost-export")
async def buyer_sku_cost_export(
    vertical_id: Optional[str] = None,
    model_id: Optional[str] = None,
    vertical_code: Optional[str] = None,
    model_code: Optional[str] = None,
    brand_id: Optional[str] = None,
    buyer_sku_id: Optional[str] = None,
):
    """Export Buyer SKU BOM with fallback prices.

    Excel columns: Buyer SKU ID | RM ID | RM Description | Qty | Price
    Filters apply progressively (vertical → model → brand → SKU).

    Vertical / Model filtering uses UUIDs (vertical_id / model_id) resolved against
    `db.bidso_skus` → same source of truth as Tech Ops. Falls back to code-regex
    on bidso_sku_id for backward compat.
    """
    # vertical_code / model_code are encoded as the first two parts of bidso_sku_id
    q = {"status": {"$ne": "INACTIVE"}}
    if vertical_id or model_id:
        bidso_q = {}
        if vertical_id:
            bidso_q["vertical_id"] = vertical_id
        if model_id:
            bidso_q["model_id"] = model_id
        bidso_ids = [b["bidso_sku_id"] async for b in db.bidso_skus.find(bidso_q, {"_id": 0, "bidso_sku_id": 1})]
        if not bidso_ids:
            raise HTTPException(status_code=404, detail="No Buyer SKUs match the filter")
        q["bidso_sku_id"] = {"$in": bidso_ids}
    elif vertical_code and model_code:
        q["bidso_sku_id"] = {"$regex": f"^{vertical_code}_{model_code}_"}
    elif vertical_code:
        q["bidso_sku_id"] = {"$regex": f"^{vertical_code}_"}
    elif model_code:
        q["bidso_sku_id"] = {"$regex": f"^[^_]+_{model_code}_"}
    if brand_id:
        q["brand_id"] = brand_id
    if buyer_sku_id:
        q["buyer_sku_id"] = buyer_sku_id

    buyer_skus = await db.buyer_skus.find(q, {"_id": 0}).sort("buyer_sku_id", 1).to_list(20000)
    if not buyer_skus:
        raise HTTPException(status_code=404, detail="No Buyer SKUs match the filter")

    # Collect bidsoid+brand pairs to load BOMs efficiently
    bidso_ids = list({s.get("bidso_sku_id") for s in buyer_skus if s.get("bidso_sku_id")})
    brand_ids = list({s.get("brand_id") for s in buyer_skus if s.get("brand_id")})

    common_boms = {}
    if bidso_ids:
        async for cb in db.common_bom.find({"bidso_sku_id": {"$in": bidso_ids}}, {"_id": 0}):
            common_boms[cb["bidso_sku_id"]] = cb.get("items", [])
    brand_boms = {}
    if bidso_ids and brand_ids:
        async for bb in db.brand_specific_bom.find(
            {"bidso_sku_id": {"$in": bidso_ids}, "brand_id": {"$in": brand_ids}},
            {"_id": 0}
        ):
            brand_boms[(bb["bidso_sku_id"], bb["brand_id"])] = bb.get("items", [])

    # Collect all unique RM IDs for price + description prefetch
    all_rm_ids = set()
    for items in common_boms.values():
        for it in items:
            if isinstance(it, dict) and it.get("rm_id"):
                all_rm_ids.add(it["rm_id"])
    for items in brand_boms.values():
        for it in items:
            if isinstance(it, dict) and it.get("rm_id"):
                all_rm_ids.add(it["rm_id"])

    rm_info = {}
    if all_rm_ids:
        async for rm in db.raw_materials.find(
            {"rm_id": {"$in": list(all_rm_ids)}},
            {"_id": 0, "rm_id": 1, "name": 1, "description": 1}
        ):
            rm_info[rm["rm_id"]] = rm.get("description") or rm.get("name") or rm["rm_id"]

    price_map = await compute_avg_with_fallback(list(all_rm_ids))

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buyer SKU BOM Cost"

    headers = ["Buyer SKU ID", "RM ID", "RM Description", "Qty", "Price"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    row = 2
    for sku in buyer_skus:
        seen = set()
        bsku_id = sku.get("buyer_sku_id")
        bidso = sku.get("bidso_sku_id")
        bid = sku.get("brand_id")
        items_list = []
        items_list.extend(common_boms.get(bidso, []))
        items_list.extend(brand_boms.get((bidso, bid), []))
        for it in items_list:
            if not isinstance(it, dict):
                continue
            rid = it.get("rm_id")
            if not rid or rid in seen:
                continue
            seen.add(rid)
            qty = float(it.get("quantity") or 0)
            price = price_map.get(rid, {}).get("avg_price", 0)
            ws.cell(row=row, column=1, value=bsku_id)
            ws.cell(row=row, column=2, value=rid)
            ws.cell(row=row, column=3, value=rm_info.get(rid, ""))
            ws.cell(row=row, column=4, value=qty)
            ws.cell(row=row, column=5, value=price)
            row += 1

    for col_letter, width in zip("ABCDE", [16, 14, 50, 10, 12]):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"

    suffix_parts = []
    if vertical_code:
        suffix_parts.append(vertical_code)
    elif vertical_id:
        suffix_parts.append(vertical_id[:8])
    if brand_id:
        suffix_parts.append(brand_id[:8])
    if buyer_sku_id:
        suffix_parts.append(buyer_sku_id)
    suffix = "_".join(suffix_parts) if suffix_parts else "all"
    filename = f"buyer_sku_bom_cost_{suffix}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
