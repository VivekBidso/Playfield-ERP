"""CPC (Central Production Control) routes - Scheduling and Branch Allocation"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from datetime import datetime, timezone, date, timedelta
from typing import Optional, List
import uuid
import io
import json

from database import db, BRANCHES
from services.stock_origin_service import create_origin_entry
from services import sku_service

router = APIRouter(tags=["CPC - Central Production Control"])

# Import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

# Import pandas for Excel parsing
try:
    import pandas as pd
except ImportError:
    pd = None

def serialize_doc(doc):
    if doc and 'created_at' in doc and isinstance(doc['created_at'], str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    if doc and 'plan_date' in doc and isinstance(doc['plan_date'], str):
        doc['plan_date'] = datetime.fromisoformat(doc['plan_date'])
    if doc and 'target_date' in doc and isinstance(doc['target_date'], str):
        doc['target_date'] = datetime.fromisoformat(doc['target_date'])
    return doc


async def get_effective_branch_capacity(branch_name: str, date_str: str, base_capacity: int = None) -> int:
    """
    Get effective capacity for a branch on a specific date.
    Priority: Daily override > Base capacity
    
    Args:
        branch_name: Name of the branch
        date_str: Date in YYYY-MM-DD format
        base_capacity: Optional base capacity (if already fetched, to avoid extra DB call)
    
    Returns:
        Effective capacity for the branch on that date
    """
    # Check for daily override first
    daily_override = await db.branch_daily_capacity.find_one({
        "branch": branch_name,
        "date": date_str
    }, {"_id": 0, "capacity": 1})
    
    if daily_override:
        return daily_override.get("capacity", 0)
    
    # Fall back to base capacity
    if base_capacity is not None:
        return base_capacity
    
    # Fetch base capacity from branch if not provided
    branch = await db.branches.find_one({"name": branch_name}, {"_id": 0, "capacity_units_per_day": 1})
    return branch.get("capacity_units_per_day", 0) if branch else 0

# ===== Models =====
class BranchCapacityUpdate(BaseModel):
    capacity_units_per_day: int
    effective_from: Optional[datetime] = None

class ProductionScheduleCreate(BaseModel):
    dispatch_lot_id: Optional[str] = None
    sku_id: str
    branch: str  # REQUIRED - no production without branch assignment
    target_quantity: int
    target_date: datetime
    priority: str = "MEDIUM"
    notes: str = ""

class BranchAllocationCreate(BaseModel):
    schedule_id: str
    branch: str
    allocated_quantity: int
    planned_date: datetime

class AutoAllocateRequest(BaseModel):
    schedule_id: str
    preferred_branches: Optional[List[str]] = None

# ===== Branch Capacity Management =====
@router.get("/branches/capacity")
async def get_branch_capacities():
    """Get all branch capacities, including daily overrides for today"""
    branches = await db.branches.find({"is_active": True}, {"_id": 0}).to_list(100)
    result = []
    
    # Get today's date info
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    today_str = today.strftime("%Y-%m-%d")
    
    for b in branches:
        branch_name = b["name"]
        base_capacity = b.get("capacity_units_per_day", 0)
        
        # Check for day-specific capacity override for today
        daily_capacity = await db.branch_daily_capacity.find_one({
            "branch": branch_name,
            "date": today_str
        }, {"_id": 0})
        
        # Use daily capacity if exists, otherwise base capacity
        if daily_capacity:
            effective_capacity = daily_capacity.get("capacity", base_capacity)
            capacity_source = "daily_override"
        else:
            effective_capacity = base_capacity
            capacity_source = "base"
        
        # Get current utilization for today
        allocations = await db.branch_allocations.find({
            "branch": branch_name,
            "planned_date": {"$gte": today, "$lt": tomorrow},
            "status": {"$in": ["PENDING", "IN_PROGRESS"]}
        }).to_list(1000)
        
        # Also count production schedules assigned to this branch for today
        schedules = await db.production_schedules.find({
            "branch": branch_name,
            "target_date": {"$gte": today, "$lt": tomorrow},
            "status": {"$nin": ["CANCELLED", "COMPLETED"]}
        }, {"_id": 0, "target_quantity": 1}).to_list(1000)
        
        allocated_from_allocations = sum(a.get("allocated_quantity", 0) for a in allocations)
        allocated_from_schedules = sum(s.get("target_quantity", 0) for s in schedules)
        allocated_today = allocated_from_allocations + allocated_from_schedules
        
        result.append({
            "branch_id": b.get("id"),
            "branch_code": b.get("branch_id"),  # The new BR_XXX format
            "branch": branch_name,
            "capacity_units_per_day": effective_capacity,
            "base_capacity": base_capacity,
            "capacity_source": capacity_source,
            "allocated_today": allocated_today,
            "available_today": max(0, effective_capacity - allocated_today),
            "utilization_percent": round((allocated_today / effective_capacity * 100), 1) if effective_capacity > 0 else 0
        })
    return result


@router.get("/branches/reference")
async def get_branches_reference():
    """Get all branches with their IDs for lookup"""
    branches = await db.branches.find(
        {"is_active": True},
        {"_id": 0, "branch_id": 1, "name": 1, "code": 1, "capacity_units_per_day": 1, "branch_type": 1}
    ).to_list(100)
    
    return {
        "branches": [
            {
                "branch_id": b.get("branch_id", ""),
                "name": b.get("name", ""),
                "code": b.get("code", ""),
                "type": b.get("branch_type", ""),
                "capacity": b.get("capacity_units_per_day", 0)
            }
            for b in branches
        ],
        "total": len(branches)
    }


@router.put("/branches/{branch_name}/capacity")
async def update_branch_capacity(branch_name: str, data: BranchCapacityUpdate):
    """Update branch production capacity"""
    result = await db.branches.update_one(
        {"name": branch_name},
        {"$set": {
            "capacity_units_per_day": data.capacity_units_per_day,
            "capacity_updated_at": datetime.now(timezone.utc)
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    # Log capacity change
    await db.capacity_history.insert_one({
        "id": str(uuid.uuid4()),
        "branch": branch_name,
        "capacity_units_per_day": data.capacity_units_per_day,
        "effective_from": data.effective_from or datetime.now(timezone.utc),
        "created_at": datetime.now(timezone.utc)
    })
    
    return {"message": f"Capacity updated to {data.capacity_units_per_day} units/day"}

@router.get("/branches/{branch_name}/capacity-forecast")
async def get_branch_capacity_forecast(branch_name: str, days: int = 7):
    """Get capacity utilization forecast for next N days.
    Uses daily override capacity if exists, otherwise falls back to base capacity.
    """
    branch = await db.branches.find_one({"name": branch_name}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    base_capacity = branch.get("capacity_units_per_day", 0)
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Get all daily capacity overrides for this branch in the date range
    date_strings = [(today + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(days)]
    daily_overrides = await db.branch_daily_capacity.find({
        "branch": branch_name,
        "date": {"$in": date_strings}
    }, {"_id": 0}).to_list(100)
    
    # Create lookup map for daily overrides
    override_map = {d["date"]: d["capacity"] for d in daily_overrides}
    
    forecast = []
    for i in range(days):
        day_start = today + timedelta(days=i)
        day_end = day_start + timedelta(days=1)
        date_str = day_start.strftime("%Y-%m-%d")
        
        # Use daily override if exists, otherwise base capacity
        effective_capacity = override_map.get(date_str, base_capacity)
        is_override = date_str in override_map
        
        allocations = await db.branch_allocations.find({
            "branch": branch_name,
            "planned_date": {"$gte": day_start, "$lt": day_end},
            "status": {"$in": ["PENDING", "IN_PROGRESS"]}
        }).to_list(1000)
        
        # Also get production schedules (exclude DELETED)
        schedules = await db.production_schedules.find({
            "branch": branch_name,
            "target_date": {"$gte": day_start, "$lt": day_end},
            "status": {"$nin": ["CANCELLED", "COMPLETED", "DELETED"]}
        }, {"_id": 0, "target_quantity": 1}).to_list(1000)
        
        allocated_from_allocations = sum(a.get("allocated_quantity", 0) for a in allocations)
        allocated_from_schedules = sum(s.get("target_quantity", 0) for s in schedules)
        allocated = allocated_from_allocations + allocated_from_schedules
        
        forecast.append({
            "date": date_str,
            "day": day_start.strftime("%A"),
            "capacity": effective_capacity,
            "base_capacity": base_capacity,
            "is_override": is_override,
            "allocated": allocated,
            "available": max(0, effective_capacity - allocated),
            "utilization_percent": round((allocated / effective_capacity * 100), 1) if effective_capacity > 0 else 0
        })
    
    return {
        "branch": branch_name,
        "base_capacity_units_per_day": base_capacity,
        "forecast": forecast
    }


# ===== Day-wise Branch Capacity =====
class DailyCapacityEntry(BaseModel):
    branch: str
    date: str  # Format: YYYY-MM-DD
    capacity: int


class DailyCapacityBulkUpload(BaseModel):
    entries: List[DailyCapacityEntry]


@router.post("/branches/daily-capacity")
async def upload_daily_capacity(data: DailyCapacityBulkUpload):
    """Upload day-wise capacity for branches. Overwrites existing data for same branch+date."""
    if not data.entries:
        raise HTTPException(status_code=400, detail="No entries provided")
    
    # Validate branches
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1}).to_list(100)
    valid_branches = {b["name"] for b in branches}
    
    inserted = 0
    updated = 0
    errors = []
    
    for entry in data.entries:
        # Validate branch
        if entry.branch not in valid_branches:
            errors.append(f"Invalid branch: {entry.branch}")
            continue
        
        # Validate date format
        try:
            datetime.strptime(entry.date, "%Y-%m-%d")
        except ValueError:
            errors.append(f"Invalid date format: {entry.date}. Use YYYY-MM-DD")
            continue
        
        # Validate capacity
        if entry.capacity < 0:
            errors.append(f"Invalid capacity for {entry.branch} on {entry.date}: {entry.capacity}")
            continue
        
        # Upsert (overwrite if exists)
        result = await db.branch_daily_capacity.update_one(
            {"branch": entry.branch, "date": entry.date},
            {
                "$set": {
                    "branch": entry.branch,
                    "date": entry.date,
                    "capacity": entry.capacity,
                    "updated_at": datetime.now(timezone.utc)
                },
                "$setOnInsert": {
                    "id": str(uuid.uuid4()),
                    "created_at": datetime.now(timezone.utc)
                }
            },
            upsert=True
        )
        
        if result.upserted_id:
            inserted += 1
        else:
            updated += 1
    
    return {
        "message": "Daily capacity uploaded",
        "inserted": inserted,
        "updated": updated,
        "total": inserted + updated,
        "errors": errors[:10] if errors else []
    }


@router.get("/branches/daily-capacity")
async def get_daily_capacities(branch: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Get day-wise capacity overrides"""
    query = {}
    if branch:
        query["branch"] = branch
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    capacities = await db.branch_daily_capacity.find(query, {"_id": 0}).sort("date", 1).to_list(5000)
    return capacities


@router.get("/branches/daily-capacity/template")
async def download_daily_capacity_template():
    """Download Excel template for daily capacity upload"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Capacity"
    
    # Headers
    headers = ["Branch ID", "Date (DD-MM-YYYY)", "Capacity"]
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Get branches for reference
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1, "capacity_units_per_day": 1}).to_list(100)
    
    # Generate branch_id if missing (BR_001, BR_002, etc.)
    for idx, b in enumerate(branches, 1):
        if not b.get("branch_id"):
            b["branch_id"] = f"BR_{idx:03d}"
    
    # Add sample rows for each branch for next 7 days
    row = 2
    today = datetime.now(timezone.utc)
    for b in branches[:3]:  # First 3 branches as samples
        for i in range(7):  # Next 7 days
            date_str = (today + timedelta(days=i)).strftime("%d-%m-%Y")
            ws.cell(row=row, column=1, value=b.get("branch_id", "BR_001"))
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=3, value=b.get("capacity_units_per_day", 100))
            row += 1
    
    # Branches reference sheet
    ws_ref = wb.create_sheet("Branches Reference")
    ws_ref.cell(row=1, column=1, value="Branch ID")
    ws_ref.cell(row=1, column=2, value="Branch Name")
    ws_ref.cell(row=1, column=3, value="Base Capacity")
    ws_ref["A1"].font = Font(bold=True)
    ws_ref["B1"].font = Font(bold=True)
    ws_ref["C1"].font = Font(bold=True)
    for i, b in enumerate(branches, 2):
        ws_ref.cell(row=i, column=1, value=b.get("branch_id", ""))
        ws_ref.cell(row=i, column=2, value=b["name"])
        ws_ref.cell(row=i, column=3, value=b.get("capacity_units_per_day", 0))
    ws_ref.column_dimensions["A"].width = 12
    ws_ref.column_dimensions["B"].width = 25
    ws_ref.column_dimensions["C"].width = 15
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daily_capacity_template.xlsx"}
    )


@router.post("/branches/daily-capacity/upload-excel")
async def upload_daily_capacity_excel(file: UploadFile = File(...)):
    """Upload daily capacity from Excel file. Overwrites existing data for same branch+date."""
    if not pd:
        raise HTTPException(status_code=500, detail="pandas not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    
    # Normalize column names
    df.columns = [str(c).strip().lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_") for c in df.columns]
    
    # Map columns
    col_map = {}
    branch_opts = ["branch_id", "branch", "branch_name", "unit", "unit_name"]
    date_opts = ["date", "date_dd_mm_yyyy", "date_yyyy_mm_dd", "yyyy_mm_dd"]
    capacity_opts = ["capacity", "capacity_qty", "qty", "units"]
    
    for opt in branch_opts:
        if opt in df.columns:
            col_map["branch"] = opt
            break
    for opt in date_opts:
        if opt in df.columns:
            col_map["date"] = opt
            break
    for opt in capacity_opts:
        if opt in df.columns:
            col_map["capacity"] = opt
            break
    
    if len(col_map) < 3:
        raise HTTPException(status_code=400, detail=f"Missing required columns. Found: {list(df.columns)}")
    
    # Validate branches - support both branch_id and branch_name
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    branch_id_to_name = {b.get("branch_id", ""): b["name"] for b in branches if b.get("branch_id")}
    branch_name_to_name = {b["name"]: b["name"] for b in branches}
    
    inserted = 0
    updated = 0
    errors = []
    
    for idx, row in df.iterrows():
        try:
            branch_value = str(row[col_map["branch"]]).strip()
            date_val = row[col_map["date"]]
            capacity = int(row[col_map["capacity"]])
            
            # Resolve branch ID to name
            branch_name = branch_id_to_name.get(branch_value) or branch_name_to_name.get(branch_value)
            
            # Parse date - supports DD-MM-YYYY (primary) and other formats
            if isinstance(date_val, str):
                date_str = date_val.strip()
                try:
                    # Try DD-MM-YYYY first (primary format)
                    parsed = datetime.strptime(date_str, "%d-%m-%Y")
                    date_str = parsed.strftime("%Y-%m-%d")
                except:
                    try:
                        # Fallback to YYYY-MM-DD
                        datetime.strptime(date_str, "%Y-%m-%d")
                        # Already in correct format
                    except:
                        # Try pandas auto-detection with dayfirst
                        parsed = pd.to_datetime(date_val, dayfirst=True)
                        date_str = parsed.strftime("%Y-%m-%d")
            else:
                # It's a datetime object from pandas
                date_str = pd.to_datetime(date_val, dayfirst=True).strftime("%Y-%m-%d")
            
            # Validate branch
            if not branch_name:
                errors.append(f"Row {idx+2}: Invalid branch '{branch_value}'. Use Branch ID (e.g., BR_001)")
                continue
            
            # Validate capacity
            if capacity < 0:
                errors.append(f"Row {idx+2}: Invalid capacity {capacity}")
                continue
            
            # Upsert
            result = await db.branch_daily_capacity.update_one(
                {"branch": branch_name, "date": date_str},
                {
                    "$set": {
                        "branch": branch_name,
                        "date": date_str,
                        "capacity": capacity,
                        "updated_at": datetime.now(timezone.utc)
                    },
                    "$setOnInsert": {
                        "id": str(uuid.uuid4()),
                        "created_at": datetime.now(timezone.utc)
                    }
                },
                upsert=True
            )
            
            if result.upserted_id:
                inserted += 1
            else:
                updated += 1
                
        except Exception as e:
            errors.append(f"Row {idx+2}: {str(e)}")
    
    return {
        "message": "Excel upload complete",
        "inserted": inserted,
        "updated": updated,
        "total": inserted + updated,
        "errors": errors[:20] if errors else [],
        "total_errors": len(errors)
    }


# ===== Production Scheduling =====
@router.get("/production-schedules")
async def get_production_schedules(
    status: Optional[str] = None,
    sku_id: Optional[str] = None,
    priority: Optional[str] = None
):
    """Get all production schedules (excludes DELETED)"""
    query = {"status": {"$ne": "DELETED"}}  # Always exclude deleted
    if status:
        query["status"] = status
    if sku_id:
        query["sku_id"] = sku_id
    if priority:
        query["priority"] = priority
    
    schedules = await db.production_schedules.find(query, {"_id": 0}).sort("target_date", 1).to_list(1000)
    
    # Enrich with allocation info
    for schedule in schedules:
        allocations = await db.branch_allocations.find(
            {"schedule_id": schedule["id"]},
            {"_id": 0}
        ).to_list(100)
        schedule["allocations"] = allocations
        schedule["total_allocated"] = sum(a.get("allocated_quantity", 0) for a in allocations)
        schedule["total_completed"] = sum(a.get("completed_quantity", 0) for a in allocations)
    
    return [serialize_doc(s) for s in schedules]


# ===== Soft Delete Production Schedules (MUST be before /{schedule_id} routes) =====

class BulkSoftDeleteRequest(BaseModel):
    month: str  # Format: YYYY-MM
    branch: str


@router.get("/production-schedules/preview-delete")
async def preview_delete_schedules(month: str, branch: str):
    """
    Preview schedules that will be soft-deleted for a given month and branch.
    Returns list of schedules with their completed quantities.
    """
    # Validate month format
    try:
        year, month_num = month.split("-")
        year = int(year)
        month_num = int(month_num)
        month_start = datetime(year, month_num, 1, 0, 0, 0, tzinfo=timezone.utc)
        
        # Calculate month end
        if month_num == 12:
            month_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
        else:
            month_end = datetime(year, month_num + 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    
    # Check if month is in the past
    today = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if month_start < today:
        raise HTTPException(status_code=400, detail="Cannot delete schedules for past months")
    
    # Get schedules for this month and branch (excluding already deleted)
    schedules = await db.production_schedules.find({
        "branch": branch,
        "target_date": {"$gte": month_start, "$lt": month_end},
        "status": {"$ne": "DELETED"}
    }, {"_id": 0}).sort("target_date", 1).to_list(5000)
    
    # Calculate summary
    total_count = len(schedules)
    completed_count = sum(1 for s in schedules if s.get("completed_quantity", 0) > 0)
    total_target = sum(s.get("target_quantity", 0) for s in schedules)
    total_completed = sum(s.get("completed_quantity", 0) for s in schedules)
    
    return {
        "month": month,
        "branch": branch,
        "schedules": [serialize_doc(s) for s in schedules],
        "summary": {
            "total_count": total_count,
            "schedules_with_completion": completed_count,
            "total_target_quantity": total_target,
            "total_completed_quantity": total_completed
        }
    }


@router.post("/production-schedules/bulk-soft-delete")
async def bulk_soft_delete_schedules(data: BulkSoftDeleteRequest):
    """
    Soft-delete all production schedules for a given month and branch.
    Sets status to 'DELETED' and records deletion metadata.
    Completed quantities are preserved for auto-population when new schedules are created.
    """
    # Validate month format
    try:
        year, month_num = data.month.split("-")
        year = int(year)
        month_num = int(month_num)
        month_start = datetime(year, month_num, 1, 0, 0, 0, tzinfo=timezone.utc)
        
        if month_num == 12:
            month_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
        else:
            month_end = datetime(year, month_num + 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    
    # Check if month is in the past
    today = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if month_start < today:
        raise HTTPException(status_code=400, detail="Cannot delete schedules for past months")
    
    # Get count before deletion
    count = await db.production_schedules.count_documents({
        "branch": data.branch,
        "target_date": {"$gte": month_start, "$lt": month_end},
        "status": {"$ne": "DELETED"}
    })
    
    if count == 0:
        return {
            "message": "No schedules found to delete",
            "deleted_count": 0
        }
    
    # Soft delete - mark as DELETED
    result = await db.production_schedules.update_many(
        {
            "branch": data.branch,
            "target_date": {"$gte": month_start, "$lt": month_end},
            "status": {"$ne": "DELETED"}
        },
        {
            "$set": {
                "status": "DELETED",
                "deleted_at": datetime.now(timezone.utc),
                "previous_status": "$status"  # Store previous status for reference
            }
        }
    )
    
    return {
        "message": f"Successfully soft-deleted {result.modified_count} schedules for {data.branch} in {data.month}",
        "deleted_count": result.modified_count,
        "month": data.month,
        "branch": data.branch
    }


@router.get("/production-schedules/deleted-completions")
async def get_deleted_completions(month: str, branch: str):
    """
    Get completed quantities from deleted schedules for auto-population.
    Used when creating new schedules to recover completed work.
    Returns a map of (target_date, sku_id) -> {completed_quantity, schedule_id}
    """
    try:
        year, month_num = month.split("-")
        year = int(year)
        month_num = int(month_num)
        month_start = datetime(year, month_num, 1, 0, 0, 0, tzinfo=timezone.utc)
        
        if month_num == 12:
            month_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
        else:
            month_end = datetime(year, month_num + 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    
    # Get deleted schedules with completed quantities > 0
    deleted_schedules = await db.production_schedules.find({
        "branch": branch,
        "target_date": {"$gte": month_start, "$lt": month_end},
        "status": "DELETED",
        "completed_quantity": {"$gt": 0}
    }, {"_id": 0}).to_list(5000)
    
    # Create a map for easy lookup
    completions = {}
    for s in deleted_schedules:
        target_date = s.get("target_date")
        if isinstance(target_date, datetime):
            date_key = target_date.strftime("%Y-%m-%d")
        else:
            date_key = str(target_date)[:10]
        
        sku_id = s.get("sku_id")
        key = f"{date_key}_{sku_id}"
        
        completions[key] = {
            "completed_quantity": s.get("completed_quantity", 0),
            "deleted_schedule_id": s.get("id"),
            "deleted_schedule_code": s.get("schedule_code"),
            "target_date": date_key,
            "sku_id": sku_id,
            "original_target_quantity": s.get("target_quantity", 0)
        }
    
    return {
        "month": month,
        "branch": branch,
        "completions": completions,
        "count": len(completions)
    }


@router.post("/production-schedules")
async def create_production_schedule(data: ProductionScheduleCreate):
    """Create a new production schedule. Branch is required - no DRAFT status."""
    count = await db.production_schedules.count_documents({})
    schedule_code = f"PS_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    # Verify SKU exists
    sku = await sku_service.get_sku_by_sku_id(data.sku_id)
    if not sku:
        raise HTTPException(status_code=404, detail="SKU not found")
    
    # Verify branch exists and is active
    branch = await db.branches.find_one({"name": data.branch, "is_active": True}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail=f"Branch '{data.branch}' not found or inactive")
    
    # Check capacity for the target date
    date_str = data.target_date.strftime("%Y-%m-%d")
    effective_capacity = await get_effective_branch_capacity(data.branch, date_str, branch.get("capacity_units_per_day", 0))
    
    # Get existing schedules for this date/branch (exclude DELETED)
    day_start = data.target_date.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)
    existing_schedules = await db.production_schedules.find({
        "branch": data.branch,
        "target_date": {"$gte": day_start, "$lt": day_end},
        "status": {"$nin": ["CANCELLED", "DELETED"]}
    }, {"_id": 0, "target_quantity": 1}).to_list(1000)
    
    existing_allocated = sum(s.get("target_quantity", 0) for s in existing_schedules)
    available = effective_capacity - existing_allocated
    
    if data.target_quantity > available:
        raise HTTPException(
            status_code=400, 
            detail=f"Insufficient capacity. Available: {available}, Requested: {data.target_quantity}"
        )
    
    schedule = {
        "id": str(uuid.uuid4()),
        "schedule_code": schedule_code,
        "dispatch_lot_id": data.dispatch_lot_id,
        "branch": data.branch,  # Branch is always assigned
        "sku_id": data.sku_id,
        "sku_description": sku.get("description", ""),
        "target_quantity": data.target_quantity,
        "allocated_quantity": data.target_quantity,  # Fully allocated since branch assigned
        "completed_quantity": 0,
        "target_date": data.target_date,
        "priority": data.priority,
        "status": "SCHEDULED",  # Only statuses: SCHEDULED, COMPLETED, CANCELLED
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc)
    }
    await db.production_schedules.insert_one(schedule)
    del schedule["_id"]
    
    # Update dispatch lot if linked
    if data.dispatch_lot_id:
        await db.dispatch_lots.update_one(
            {"id": data.dispatch_lot_id},
            {"$set": {"status": "PRODUCTION_ASSIGNED"}}
        )
    
    return serialize_doc(schedule)

@router.get("/production-schedules/{schedule_id}")
async def get_production_schedule(schedule_id: str):
    """Get schedule with all allocations and progress"""
    schedule = await db.production_schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    allocations = await db.branch_allocations.find(
        {"schedule_id": schedule_id},
        {"_id": 0}
    ).to_list(100)
    
    schedule["allocations"] = [serialize_doc(a) for a in allocations]
    schedule["total_allocated"] = sum(a.get("allocated_quantity", 0) for a in allocations)
    schedule["total_completed"] = sum(a.get("completed_quantity", 0) for a in allocations)
    
    return serialize_doc(schedule)

# ===== Branch Allocation =====
@router.post("/branch-allocations")
async def create_branch_allocation(data: BranchAllocationCreate):
    """Manually allocate production to a branch"""
    # Verify schedule exists
    schedule = await db.production_schedules.find_one({"id": data.schedule_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    # Check branch capacity
    branch = await db.branches.find_one({"name": data.branch})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    # Check existing allocations for that day
    day_start = data.planned_date.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)
    date_str = day_start.strftime("%Y-%m-%d")
    
    # Get effective capacity (daily override or base)
    base_capacity = branch.get("capacity_units_per_day", 0)
    capacity = await get_effective_branch_capacity(data.branch, date_str, base_capacity)
    
    existing = await db.branch_allocations.find({
        "branch": data.branch,
        "planned_date": {"$gte": day_start, "$lt": day_end},
        "status": {"$in": ["PENDING", "IN_PROGRESS"]}
    }).to_list(1000)
    
    allocated_today = sum(a.get("allocated_quantity", 0) for a in existing)
    available = capacity - allocated_today
    
    if data.allocated_quantity > available:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient capacity. Available: {available}, Requested: {data.allocated_quantity}"
        )
    
    allocation = {
        "id": str(uuid.uuid4()),
        "schedule_id": data.schedule_id,
        "sku_id": schedule["sku_id"],
        "branch": data.branch,
        "allocated_quantity": data.allocated_quantity,
        "completed_quantity": 0,
        "planned_date": data.planned_date,
        "status": "PENDING",  # PENDING, IN_PROGRESS, COMPLETED, CANCELLED
        "created_at": datetime.now(timezone.utc)
    }
    await db.branch_allocations.insert_one(allocation)
    
    # Update schedule totals
    await db.production_schedules.update_one(
        {"id": data.schedule_id},
        {
            "$inc": {"allocated_quantity": data.allocated_quantity},
            "$set": {"status": "SCHEDULED"}
        }
    )
    
    del allocation["_id"]
    return serialize_doc(allocation)

@router.post("/branch-allocations/auto-allocate")
async def auto_allocate_production(data: AutoAllocateRequest):
    """Automatically allocate production across branches based on capacity"""
    schedule = await db.production_schedules.find_one({"id": data.schedule_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    remaining = schedule["target_quantity"] - schedule.get("allocated_quantity", 0)
    if remaining <= 0:
        return {"message": "Schedule already fully allocated", "allocations": []}
    
    # Get branches with capacity
    branches = await db.branches.find({"is_active": True}, {"_id": 0}).to_list(100)
    if data.preferred_branches:
        branches = [b for b in branches if b["name"] in data.preferred_branches]
    
    # Sort by available capacity (descending)
    target_date = schedule["target_date"]
    day_start = target_date.replace(hour=0, minute=0, second=0, microsecond=0) if isinstance(target_date, datetime) else datetime.fromisoformat(str(target_date))
    day_end = day_start + timedelta(days=1)
    date_str = day_start.strftime("%Y-%m-%d")
    
    branch_availability = []
    for b in branches:
        base_capacity = b.get("capacity_units_per_day", 0)
        # Get effective capacity (daily override or base)
        capacity = await get_effective_branch_capacity(b["name"], date_str, base_capacity)
        if capacity == 0:
            continue
        
        existing = await db.branch_allocations.find({
            "branch": b["name"],
            "planned_date": {"$gte": day_start, "$lt": day_end},
            "status": {"$in": ["PENDING", "IN_PROGRESS"]}
        }).to_list(1000)
        
        allocated = sum(a.get("allocated_quantity", 0) for a in existing)
        available = capacity - allocated
        
        if available > 0:
            branch_availability.append({
                "branch": b["name"],
                "capacity": capacity,
                "available": available
            })
    
    # Sort by availability
    branch_availability.sort(key=lambda x: x["available"], reverse=True)
    
    allocations_created = []
    for ba in branch_availability:
        if remaining <= 0:
            break
        
        allocate_qty = min(remaining, ba["available"])
        
        allocation = {
            "id": str(uuid.uuid4()),
            "schedule_id": data.schedule_id,
            "sku_id": schedule["sku_id"],
            "branch": ba["branch"],
            "allocated_quantity": allocate_qty,
            "completed_quantity": 0,
            "planned_date": day_start,
            "status": "PENDING",
            "created_at": datetime.now(timezone.utc)
        }
        await db.branch_allocations.insert_one(allocation)
        
        remaining -= allocate_qty
        allocations_created.append({
            "branch": ba["branch"],
            "quantity": allocate_qty
        })
    
    # Update schedule
    total_allocated = schedule.get("allocated_quantity", 0) + sum(a["quantity"] for a in allocations_created)
    
    await db.production_schedules.update_one(
        {"id": data.schedule_id},
        {"$set": {"allocated_quantity": total_allocated, "status": "SCHEDULED"}}
    )
    
    return {
        "message": f"Allocated {sum(a['quantity'] for a in allocations_created)} units across {len(allocations_created)} branches",
        "remaining_unallocated": remaining,
        "allocations": allocations_created
    }

@router.put("/production-schedules/{schedule_id}/complete")
async def complete_production_schedule(schedule_id: str, completed_quantity: int):
    """Mark production schedule as completed. Called by branch ops team."""
    import logging
    logger = logging.getLogger(__name__)
    
    schedule = await db.production_schedules.find_one({"id": schedule_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    if schedule.get("status") == "COMPLETED":
        raise HTTPException(status_code=400, detail="Schedule already completed")
    
    if schedule.get("status") == "CANCELLED":
        raise HTTPException(status_code=400, detail="Cannot complete a cancelled schedule")
    
    completion_time = datetime.now(timezone.utc)
    
    await db.production_schedules.update_one(
        {"id": schedule_id},
        {"$set": {
            "status": "COMPLETED",
            "completed_quantity": completed_quantity,
            "completed_at": completion_time
        }}
    )
    
    # Get SKU and branch info
    sku_id = schedule.get("sku_id") or schedule.get("bidso_sku_id")
    branch = schedule.get("branch")
    
    logger.info(f"Completing schedule {schedule.get('schedule_code')}: SKU={sku_id}, Branch={branch}, Qty={completed_quantity}")
    
    # Get branch_id for FG inventory (matches inventory_routes query structure)
    branch_doc = await db.branches.find_one({"name": branch}, {"_id": 0, "branch_id": 1})
    branch_id = branch_doc.get("branch_id") if branch_doc else None
    
    logger.info(f"Branch lookup: name={branch} -> branch_id={branch_id}")
    
    fg_updated = False
    fg_error = None
    
    # Update FG inventory
    if sku_id and branch and completed_quantity > 0:
        try:
            fg_existing = await db.fg_inventory.find_one(
                {"buyer_sku_id": sku_id, "branch_id": branch_id}
            )
            
            if fg_existing:
                result = await db.fg_inventory.update_one(
                    {"buyer_sku_id": sku_id, "branch_id": branch_id},
                    {"$inc": {"quantity": completed_quantity}}
                )
                fg_updated = result.modified_count > 0
                logger.info(f"FG inventory updated: modified_count={result.modified_count}")
            else:
                result = await db.fg_inventory.insert_one({
                    "id": str(uuid.uuid4()),
                    "buyer_sku_id": sku_id,
                    "sku_id": sku_id,
                    "branch_id": branch_id,
                    "branch": branch,
                    "quantity": completed_quantity,
                    "created_at": completion_time.isoformat()
                })
                fg_updated = result.inserted_id is not None
                logger.info(f"FG inventory inserted: inserted_id={result.inserted_id}")
            
            # Create stock origin entry for manufacturing origin tracking
            await create_origin_entry(
                sku_id=sku_id,
                branch=branch,
                quantity=completed_quantity,
                manufacturing_unit=branch,
                production_date=completion_time,
                production_schedule_id=schedule_id
            )
        except Exception as e:
            fg_error = str(e)
            logger.error(f"FG inventory update failed: {e}")
    else:
        logger.warning(f"Skipping FG update: sku_id={sku_id}, branch={branch}, qty={completed_quantity}")
    
    # Update dispatch lot if linked
    if schedule.get("dispatch_lot_id"):
        await db.dispatch_lots.update_one(
            {"id": schedule["dispatch_lot_id"]},
            {"$set": {"status": "FULLY_PRODUCED"}}
        )
    
    return {
        "message": f"Production schedule {schedule.get('schedule_code')} completed with {completed_quantity} units",
        "schedule_id": schedule_id,
        "sku_id": sku_id,
        "branch": branch,
        "branch_id": branch_id,
        "fg_inventory_updated": fg_updated,
        "fg_error": fg_error,
        "manufacturing_origin": branch
    }

@router.put("/production-schedules/{schedule_id}/cancel")
async def cancel_production_schedule(schedule_id: str):
    """Cancel a production schedule."""
    schedule = await db.production_schedules.find_one({"id": schedule_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    if schedule.get("status") == "COMPLETED":
        raise HTTPException(status_code=400, detail="Cannot cancel a completed schedule")
    
    await db.production_schedules.update_one(
        {"id": schedule_id},
        {"$set": {"status": "CANCELLED", "cancelled_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": f"Production schedule {schedule.get('schedule_code')} cancelled"}


# Legacy endpoint - kept for backward compatibility but simplified
@router.put("/branch-allocations/{allocation_id}/start")
async def start_production(allocation_id: str):
    """Legacy: Mark allocation as in progress - now just returns success"""
    return {"message": "Production tracking simplified - use schedule complete endpoint"}

@router.put("/branch-allocations/{allocation_id}/complete")
async def complete_allocation(allocation_id: str, completed_quantity: int):
    """Legacy: Complete allocation - redirects to schedule completion"""
    allocation = await db.branch_allocations.find_one({"id": allocation_id})
    if not allocation:
        raise HTTPException(status_code=404, detail="Allocation not found")
    
    # Mark allocation complete
    await db.branch_allocations.update_one(
        {"id": allocation_id},
        {"$set": {
            "status": "COMPLETED",
            "completed_quantity": completed_quantity,
            "completed_at": datetime.now(timezone.utc)
        }}
    )
    
    # Update schedule completed quantity
    await db.production_schedules.update_one(
        {"id": allocation["schedule_id"]},
        {"$inc": {"completed_quantity": completed_quantity}}
    )
    
    return {"message": f"Completed {completed_quantity} units"}

# ===== Dashboard & Reports =====
@router.get("/cpc/dashboard")
async def get_cpc_dashboard():
    """Get CPC overview dashboard"""
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    
    # Schedule counts - only SCHEDULED and COMPLETED statuses exist (exclude DELETED)
    scheduled_count = await db.production_schedules.count_documents({"status": "SCHEDULED"})
    completed_count = await db.production_schedules.count_documents({"status": "COMPLETED"})
    
    # Today's schedules (exclude DELETED)
    todays_schedules = await db.production_schedules.find({
        "target_date": {"$gte": today, "$lt": tomorrow},
        "status": {"$nin": ["CANCELLED", "DELETED"]}
    }).to_list(1000)
    
    # Branch utilization
    branches = await db.branches.find({"is_active": True}, {"_id": 0}).to_list(100)
    branch_stats = []
    for b in branches:
        branch_schedules = [s for s in todays_schedules if s.get("branch") == b["name"]]
        allocated = sum(s.get("target_quantity", 0) for s in branch_schedules)
        completed = sum(s.get("completed_quantity", 0) for s in branch_schedules if s.get("status") == "COMPLETED")
        
        # Get effective capacity for today
        date_str = today.strftime("%Y-%m-%d")
        capacity = await get_effective_branch_capacity(b["name"], date_str, b.get("capacity_units_per_day", 0))
        
        branch_stats.append({
            "branch": b["name"],
            "capacity": capacity,
            "allocated": allocated,
            "completed": completed,
            "utilization": round((allocated / capacity * 100), 1) if capacity > 0 else 0
        })
    
    return {
        "scheduled_count": scheduled_count,
        "completed_count": completed_count,
        "todays_planned_quantity": sum(s.get("target_quantity", 0) for s in todays_schedules),
        "todays_completed_quantity": sum(s.get("completed_quantity", 0) for s in todays_schedules if s.get("status") == "COMPLETED"),
        "branch_utilization": branch_stats
    }


# ===== Data Cleanup Endpoints =====

@router.delete("/cpc/cleanup/unassigned-schedules")
async def cleanup_unassigned_schedules():
    """
    Delete all production schedules that have no branch assigned.
    Rule: All schedules must have a branch - unassigned schedules are invalid.
    """
    # Find unassigned schedules
    unassigned = await db.production_schedules.find(
        {"$or": [{"branch": None}, {"branch": ""}, {"branch": {"$exists": False}}]},
        {"_id": 0, "schedule_code": 1, "id": 1}
    ).to_list(1000)
    
    if not unassigned:
        return {"message": "No unassigned schedules found", "deleted": 0}
    
    # Delete them
    result = await db.production_schedules.delete_many(
        {"$or": [{"branch": None}, {"branch": ""}, {"branch": {"$exists": False}}]}
    )
    
    return {
        "message": f"Deleted {result.deleted_count} unassigned schedules",
        "deleted": result.deleted_count,
        "schedule_codes": [s.get("schedule_code") for s in unassigned[:20]]
    }


@router.get("/cpc/unassigned-schedules")
async def get_unassigned_schedules():
    """List all production schedules without branch assignment (should be 0)"""
    unassigned = await db.production_schedules.find(
        {"$or": [{"branch": None}, {"branch": ""}, {"branch": {"$exists": False}}]},
        {"_id": 0}
    ).to_list(1000)
    
    return {
        "count": len(unassigned),
        "schedules": [serialize_doc(s) for s in unassigned],
        "message": "All schedules must have a branch assigned. Use cleanup endpoint to remove invalid schedules."
    }


@router.post("/cpc/fix-draft-schedules")
async def fix_draft_schedules_with_branch():
    """
    Legacy migration: Fix any old DRAFT schedules by updating to SCHEDULED.
    New workflow only has: SCHEDULED, COMPLETED, CANCELLED
    """
    # Find any remaining DRAFT schedules and update to SCHEDULED
    result = await db.production_schedules.update_many(
        {"status": "DRAFT"},
        {"$set": {"status": "SCHEDULED"}}
    )
    
    # Also fix any IN_PROGRESS to SCHEDULED (simplified workflow)
    result2 = await db.production_schedules.update_many(
        {"status": "IN_PROGRESS"},
        {"$set": {"status": "SCHEDULED"}}
    )
    
    return {
        "message": f"Migration complete. Updated {result.modified_count} DRAFT and {result2.modified_count} IN_PROGRESS to SCHEDULED",
        "draft_updated": result.modified_count,
        "in_progress_updated": result2.modified_count
    }


# ===== RM Shortage Report =====

@router.get("/cpc/rm-shortage-report")
async def get_rm_shortage_report(branch: Optional[str] = None):
    """
    Get RM shortage report based on branch-level production schedules and branch RM stock.
    Shows what raw materials are needed vs what's available at each branch.
    """
    # Build schedule query
    query = {"status": {"$in": ["SCHEDULED", "IN_PROGRESS"]}}
    if branch:
        query["branch"] = branch
    
    # Get all scheduled production
    schedules = await db.production_schedules.find(query, {"_id": 0}).to_list(5000)
    
    if not schedules:
        return {
            "message": "No scheduled production found",
            "branch_filter": branch,
            "branches": []
        }
    
    # Group schedules by branch
    branch_schedules = {}
    for s in schedules:
        b = s.get("branch") or "Unassigned"
        if b not in branch_schedules:
            branch_schedules[b] = []
        branch_schedules[b].append(s)
    
    # Get all SKU IDs from schedules
    sku_ids = list(set(s.get("sku_id") for s in schedules if s.get("sku_id")))
    
    # Get BOM mappings for these SKUs
    bom_mappings = await db.bill_of_materials.find(
        {"sku_id": {"$in": sku_ids}},
        {"_id": 0, "sku_id": 1, "rm_id": 1, "quantity": 1}
    ).to_list(50000)
    
    # Also check sku_rm_mapping collection
    sku_rm_mappings = await db.sku_rm_mapping.find(
        {"sku_id": {"$in": sku_ids}},
        {"_id": 0, "sku_id": 1, "rm_id": 1, "quantity": 1}
    ).to_list(50000)
    
    # Combine BOM data
    bom_by_sku = {}
    for m in bom_mappings + sku_rm_mappings:
        sku = m.get("sku_id")
        if sku not in bom_by_sku:
            bom_by_sku[sku] = []
        bom_by_sku[sku].append({
            "rm_id": m.get("rm_id"),
            "qty_per_unit": m.get("quantity", 1)
        })
    
    # Get all RM IDs needed
    all_rm_ids = set()
    for sku, mappings in bom_by_sku.items():
        for m in mappings:
            all_rm_ids.add(m.get("rm_id"))
    
    # Get RM details
    rms = await db.raw_materials.find(
        {"rm_id": {"$in": list(all_rm_ids)}},
        {"_id": 0, "rm_id": 1, "description": 1, "category": 1}
    ).to_list(5000)
    rm_map = {r["rm_id"]: r for r in rms}
    
    # Build report per branch
    branch_reports = []
    
    for branch_name, branch_sched in branch_schedules.items():
        if branch_name == "Unassigned":
            continue  # Skip unassigned (shouldn't exist)
        
        # Calculate RM requirements for this branch
        rm_requirements = {}
        for s in branch_sched:
            sku_id = s.get("sku_id")
            target_qty = s.get("target_quantity", 0)
            completed_qty = s.get("completed_quantity", 0)
            pending_qty = target_qty - completed_qty
            
            if pending_qty <= 0:
                continue
            
            bom = bom_by_sku.get(sku_id, [])
            for m in bom:
                rm_id = m.get("rm_id")
                qty_per_unit = m.get("qty_per_unit", 1)
                required = pending_qty * qty_per_unit
                
                if rm_id not in rm_requirements:
                    rm_requirements[rm_id] = 0
                rm_requirements[rm_id] += required
        
        # Get branch RM inventory
        branch_rm_inv = await db.branch_rm_inventory.find(
            {"branch": branch_name, "rm_id": {"$in": list(rm_requirements.keys())}},
            {"_id": 0, "rm_id": 1, "quantity": 1}
        ).to_list(5000)
        
        inv_by_rm = {inv["rm_id"]: inv.get("quantity", 0) for inv in branch_rm_inv}
        
        # Calculate shortages
        shortages = []
        for rm_id, required in rm_requirements.items():
            available = inv_by_rm.get(rm_id, 0)
            shortage = max(0, required - available)
            
            rm_info = rm_map.get(rm_id, {})
            shortages.append({
                "rm_id": rm_id,
                "description": rm_info.get("description", ""),
                "category": rm_info.get("category", ""),
                "required": required,
                "available": available,
                "shortage": shortage,
                "status": "CRITICAL" if shortage > 0 and available == 0 else ("SHORT" if shortage > 0 else "OK")
            })
        
        # Sort by shortage (highest first)
        shortages.sort(key=lambda x: -x["shortage"])
        
        branch_reports.append({
            "branch": branch_name,
            "scheduled_qty": sum(s.get("target_quantity", 0) for s in branch_sched),
            "pending_qty": sum(max(0, s.get("target_quantity", 0) - s.get("completed_quantity", 0)) for s in branch_sched),
            "total_rms_needed": len(rm_requirements),
            "rms_with_shortage": len([s for s in shortages if s["shortage"] > 0]),
            "critical_shortages": len([s for s in shortages if s["status"] == "CRITICAL"]),
            "shortages": shortages[:50]  # Top 50 shortages
        })
    
    # Sort branches by critical shortages
    branch_reports.sort(key=lambda x: -x["critical_shortages"])
    
    return {
        "branch_filter": branch,
        "total_branches": len(branch_reports),
        "total_shortages": sum(b["rms_with_shortage"] for b in branch_reports),
        "total_critical": sum(b["critical_shortages"] for b in branch_reports),
        "branches": branch_reports
    }


@router.get("/cpc/rm-shortage-report/download")
async def download_rm_shortage_report(branch: Optional[str] = None):
    """Download RM shortage report as Excel"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Get the report data
    report = await get_rm_shortage_report(branch)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RM Shortage Report"
    
    # Headers
    headers = ["Branch", "RM ID", "Description", "Category", "Required", "Available", "Shortage", "Status"]
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row = 2
    for branch_data in report.get("branches", []):
        branch_name = branch_data.get("branch", "")
        for shortage in branch_data.get("shortages", []):
            ws.cell(row=row, column=1, value=branch_name)
            ws.cell(row=row, column=2, value=shortage.get("rm_id", ""))
            ws.cell(row=row, column=3, value=shortage.get("description", ""))
            ws.cell(row=row, column=4, value=shortage.get("category", ""))
            ws.cell(row=row, column=5, value=shortage.get("required", 0))
            ws.cell(row=row, column=6, value=shortage.get("available", 0))
            ws.cell(row=row, column=7, value=shortage.get("shortage", 0))
            ws.cell(row=row, column=8, value=shortage.get("status", ""))
            
            # Highlight shortages
            if shortage.get("status") == "CRITICAL":
                for c in range(1, 9):
                    ws.cell(row=row, column=c).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif shortage.get("status") == "SHORT":
                for c in range(1, 9):
                    ws.cell(row=row, column=c).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
            row += 1
    
    # Set column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rm_shortage_report_{branch or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/cpc/schedule-suggestions")
async def get_schedule_suggestions():
    """Get dispatch lots that need scheduling"""
    # Find dispatch lots without production schedules
    lots = await db.dispatch_lots.find({
        "status": {"$in": ["CREATED", "PRODUCTION_ASSIGNED"]},
        "required_quantity": {"$gt": 0}
    }, {"_id": 0}).sort("target_date", 1).to_list(100)
    
    suggestions = []
    for lot in lots:
        # Check if already scheduled
        existing_schedule = await db.production_schedules.find_one({"dispatch_lot_id": lot["id"]})
        if existing_schedule:
            continue
        
        # Get SKU info
        sku = await sku_service.get_sku_by_sku_id(lot["sku_id"])
        
        suggestions.append({
            "dispatch_lot_id": lot["id"],
            "lot_code": lot.get("lot_code"),
            "sku_id": lot["sku_id"],
            "sku_description": sku.get("description", "") if sku else "",
            "required_quantity": lot["required_quantity"],
            "target_date": lot.get("target_date"),
            "priority": lot.get("priority", "MEDIUM")
        })
    
    return suggestions


# ===== Branch-wise Production Schedule View =====

@router.get("/cpc/branch-schedules")
async def get_branch_wise_schedules(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch: Optional[str] = None
):
    """
    Get branch-wise per-day production schedules.
    Shows all schedules grouped by branch and date.
    """
    # Default to next 14 days if no date range provided
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            start = today
    else:
        start = today
    
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
        except ValueError:
            end = today + timedelta(days=14)
    else:
        end = today + timedelta(days=14)
    
    # Build query - exclude CANCELLED and DELETED schedules
    query = {
        "target_date": {"$gte": start, "$lt": end},
        "status": {"$nin": ["CANCELLED", "DELETED"]}
    }
    if branch:
        query["branch"] = branch
    
    # Get schedules
    schedules = await db.production_schedules.find(
        query,
        {"_id": 0}
    ).sort("target_date", 1).to_list(5000)
    
    # Get branch info
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1, "capacity_units_per_day": 1}).to_list(100)
    branch_capacity = {b["name"]: b.get("capacity_units_per_day", 0) for b in branches}
    
    # Get daily capacity overrides for date range
    date_strs = []
    current = start
    while current < end:
        date_strs.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)
    
    daily_overrides = await db.branch_daily_capacity.find(
        {"date": {"$in": date_strs}},
        {"_id": 0}
    ).to_list(5000)
    
    # Map daily overrides: {branch: {date: capacity}}
    override_map = {}
    for o in daily_overrides:
        b = o.get("branch")
        d = o.get("date")
        if b not in override_map:
            override_map[b] = {}
        override_map[b][d] = o.get("capacity", 0)
    
    # Get SKU info
    sku_ids = list(set(s.get("sku_id") for s in schedules if s.get("sku_id")))
    skus = await sku_service.get_skus_by_sku_ids(sku_ids)
    sku_map = {s["sku_id"]: s.get("description", "") for s in skus}
    
    # Get forecast info
    forecast_ids = list(set(s.get("forecast_id") for s in schedules if s.get("forecast_id")))
    forecasts = await db.forecasts.find({"id": {"$in": forecast_ids}}, {"_id": 0, "id": 1, "forecast_code": 1}).to_list(1000)
    forecast_map = {f["id"]: f.get("forecast_code", "") for f in forecasts}
    
    # Group schedules by branch and date
    branch_date_schedules = {}
    for s in schedules:
        b = s.get("branch") or "Unassigned"
        target_date = s.get("target_date")
        if isinstance(target_date, str):
            date_str = target_date[:10]
        else:
            date_str = target_date.strftime("%Y-%m-%d") if target_date else "No Date"
        
        key = f"{b}|{date_str}"
        if key not in branch_date_schedules:
            branch_date_schedules[key] = {
                "branch": b,
                "date": date_str,
                "schedules": [],
                "total_qty": 0
            }
        
        branch_date_schedules[key]["schedules"].append({
            "schedule_code": s.get("schedule_code"),
            "sku_id": s.get("sku_id"),
            "sku_description": sku_map.get(s.get("sku_id"), s.get("sku_description", "")),
            "forecast_code": forecast_map.get(s.get("forecast_id"), ""),
            "target_quantity": s.get("target_quantity", 0),
            "completed_quantity": s.get("completed_quantity", 0),
            "status": s.get("status"),
            "priority": s.get("priority", "MEDIUM")
        })
        branch_date_schedules[key]["total_qty"] += s.get("target_quantity", 0)
    
    # Build result with capacity info
    result = []
    for key, data in branch_date_schedules.items():
        b = data["branch"]
        date_str = data["date"]
        
        # Get effective capacity (daily override or base)
        if b in override_map and date_str in override_map[b]:
            capacity = override_map[b][date_str]
            capacity_source = "daily_override"
        else:
            capacity = branch_capacity.get(b, 0)
            capacity_source = "base"
        
        result.append({
            "branch": b,
            "date": date_str,
            "capacity": capacity,
            "capacity_source": capacity_source,
            "total_scheduled": data["total_qty"],
            "available": max(0, capacity - data["total_qty"]),
            "utilization_percent": round(data["total_qty"] / capacity * 100, 1) if capacity > 0 else 0,
            "schedules": data["schedules"]
        })
    
    # Sort by date then branch
    result.sort(key=lambda x: (x["date"], x["branch"]))
    
    return result



# ===== Demand Forecasts Visibility for CPC =====

@router.get("/cpc/demand-forecasts")
async def get_demand_forecasts_for_cpc(
    status: Optional[str] = None,
    buyer_id: Optional[str] = None,
    include_draft: bool = False
):
    """
    Get demand forecasts for CPC to view and schedule production.
    Shows forecasts with scheduled production qty, inventory available, and schedule pending.
    Rule: Schedule Pending = Forecast Qty - Inventory - Already Scheduled
    """
    if include_draft:
        query = {"status": {"$in": ["DRAFT", "CONFIRMED", "CONVERTED"]}}
    else:
        query = {"status": {"$in": ["CONFIRMED", "CONVERTED"]}}
    
    if status:
        query["status"] = status
    if buyer_id:
        query["buyer_id"] = buyer_id
    
    forecasts = await db.forecasts.find(query, {"_id": 0}).sort("forecast_month", 1).to_list(1000)
    
    # Get all production SCHEDULES to calculate scheduled quantities
    # ONLY count schedules that are linked to forecasts
    all_schedules = await db.production_schedules.find(
        {"forecast_id": {"$exists": True, "$ne": None}, "status": {"$ne": "CANCELLED"}},
        {"_id": 0, "forecast_id": 1, "target_quantity": 1}
    ).to_list(10000)
    
    # Sum scheduled qty by forecast_id
    scheduled_by_forecast = {}
    for s in all_schedules:
        fid = s.get("forecast_id")
        if fid:
            scheduled_by_forecast[fid] = scheduled_by_forecast.get(fid, 0) + s.get("target_quantity", 0)
    
    # Get all dispatch lots linked to forecasts
    all_lots = await db.dispatch_lots.find(
        {"forecast_id": {"$exists": True, "$ne": None}},
        {"_id": 0, "forecast_id": 1, "lot_code": 1, "total_quantity": 1, "required_quantity": 1, "status": 1}
    ).to_list(5000)
    
    lots_by_forecast = {}
    for lot in all_lots:
        fid = lot.get("forecast_id")
        if fid:
            if fid not in lots_by_forecast:
                lots_by_forecast[fid] = []
            lots_by_forecast[fid].append({
                "lot_code": lot.get("lot_code"),
                "quantity": lot.get("total_quantity") or lot.get("required_quantity", 0),
                "status": lot.get("status")
            })
    
    # Get buyer names
    buyer_ids = list(set(f.get("buyer_id") for f in forecasts if f.get("buyer_id")))
    buyers = await db.buyers.find({"id": {"$in": buyer_ids}}, {"_id": 0, "id": 1, "name": 1, "code": 1}).to_list(100)
    buyer_map = {b["id"]: b for b in buyers}
    
    # Get SKU details including current_stock
    sku_ids = list(set(f.get("sku_id") for f in forecasts if f.get("sku_id")))
    skus = await sku_service.get_skus_by_sku_ids(sku_ids)
    sku_map = {s["sku_id"]: s for s in skus}
    
    # Get FG inventory by SKU
    fg_inventory = await db.fg_inventory.find(
        {"sku_id": {"$in": sku_ids}},
        {"_id": 0, "sku_id": 1, "quantity": 1}
    ).to_list(5000)
    
    # Sum inventory by SKU
    inventory_by_sku = {}
    for inv in fg_inventory:
        sku = inv.get("sku_id")
        inventory_by_sku[sku] = inventory_by_sku.get(sku, 0) + inv.get("quantity", 0)
    
    # Add SKU current_stock if no FG inventory
    for sku_id in sku_ids:
        if sku_id not in inventory_by_sku or inventory_by_sku[sku_id] == 0:
            sku_data = sku_map.get(sku_id, {})
            inventory_by_sku[sku_id] = sku_data.get("current_stock", 0)
    
    # Get vertical details
    vertical_ids = list(set(f.get("vertical_id") for f in forecasts if f.get("vertical_id")))
    verticals = await db.verticals.find({"id": {"$in": vertical_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    vertical_map = {v["id"]: v["name"] for v in verticals}
    
    result = []
    for f in forecasts:
        forecast_qty = f.get("quantity", 0)
        forecast_id = f.get("id")
        sku_id = f.get("sku_id")
        
        # Get scheduled production qty (from production_schedules)
        scheduled_qty = scheduled_by_forecast.get(forecast_id, 0)
        
        # Cap scheduled_qty at forecast_qty (rule: scheduled cannot exceed forecast)
        scheduled_qty = min(scheduled_qty, forecast_qty)
        
        # Get inventory available for this SKU
        inventory_qty = inventory_by_sku.get(sku_id, 0) if sku_id else 0
        
        # Calculate schedule pending: Forecast - Inventory - Scheduled
        schedule_pending = max(0, forecast_qty - inventory_qty - scheduled_qty)
        
        # Remaining qty (for backward compatibility, same as schedule_pending)
        remaining_qty = schedule_pending
        
        # Get linked dispatch lots
        dispatch_lots = lots_by_forecast.get(forecast_id, [])
        dispatch_qty = sum(l.get("quantity", 0) for l in dispatch_lots)
        
        buyer = buyer_map.get(f.get("buyer_id"), {})
        sku = sku_map.get(sku_id, {})
        
        result.append({
            "id": forecast_id,
            "forecast_code": f.get("forecast_code"),
            "buyer_id": f.get("buyer_id"),
            "buyer_name": buyer.get("name"),
            "buyer_code": buyer.get("code"),
            "vertical_id": f.get("vertical_id"),
            "vertical_name": vertical_map.get(f.get("vertical_id"), f.get("vertical_id")),
            "sku_id": sku_id,
            "sku_description": sku.get("description", ""),
            "brand": sku.get("brand", ""),
            "forecast_month": f.get("forecast_month"),
            "forecast_qty": forecast_qty,
            "inventory_qty": inventory_qty,
            "scheduled_qty": scheduled_qty,
            "schedule_pending": schedule_pending,
            "remaining_qty": remaining_qty,
            "dispatch_qty": dispatch_qty,
            "dispatch_lots": dispatch_lots,
            "is_fully_scheduled": schedule_pending == 0,
            "priority": f.get("priority", "MEDIUM"),
            "status": f.get("status"),
            "notes": f.get("notes", "")
        })
    
    # Sort by schedule_pending (highest first) then by forecast month
    result.sort(key=lambda x: (-x["schedule_pending"], x.get("forecast_month", "")))
    
    return result


@router.get("/cpc/demand-forecasts/summary")
async def get_demand_forecasts_summary():
    """Get summary of demand forecasts for CPC dashboard"""
    forecasts = await db.forecasts.find(
        {"status": {"$in": ["CONFIRMED", "CONVERTED"]}},
        {"_id": 0, "quantity": 1, "status": 1, "id": 1, "sku_id": 1}
    ).to_list(5000)
    
    forecast_ids = [f["id"] for f in forecasts]
    total_forecast_qty = sum(f.get("quantity", 0) for f in forecasts)
    
    # Get scheduled quantities - ONLY count schedules linked to forecasts
    schedules = await db.production_schedules.find(
        {"forecast_id": {"$in": forecast_ids}, "status": {"$ne": "CANCELLED"}},
        {"_id": 0, "target_quantity": 1}
    ).to_list(5000)
    
    total_scheduled_qty = sum(s.get("target_quantity", 0) for s in schedules)
    
    # Ensure scheduled qty doesn't exceed forecast qty in the summary
    total_scheduled_qty = min(total_scheduled_qty, total_forecast_qty)
    
    # Get FG inventory for SKUs in forecasts
    sku_ids = list(set(f.get("sku_id") for f in forecasts if f.get("sku_id")))
    fg_inventory = await db.fg_inventory.find(
        {"sku_id": {"$in": sku_ids}},
        {"_id": 0, "sku_id": 1, "quantity": 1}
    ).to_list(5000)
    total_inventory = sum(inv.get("quantity", 0) for inv in fg_inventory)
    
    # Also check SKU current_stock
    skus = await sku_service.get_skus_by_sku_ids(sku_ids)
    
    # Add current_stock if not already in fg_inventory
    inventory_skus = set(inv.get("sku_id") for inv in fg_inventory)
    for sku in skus:
        if sku.get("sku_id") not in inventory_skus:
            total_inventory += sku.get("current_stock", 0)
    
    # Schedule Pending = Forecast Qty - Inventory - Already Scheduled
    remaining_to_schedule = max(0, total_forecast_qty - total_inventory - total_scheduled_qty)
    
    return {
        "total_forecasts": len(forecasts),
        "total_forecast_qty": total_forecast_qty,
        "total_inventory": total_inventory,
        "total_scheduled_qty": total_scheduled_qty,
        "remaining_to_schedule": remaining_to_schedule,
        "scheduling_percent": round((total_inventory + total_scheduled_qty) / total_forecast_qty * 100, 1) if total_forecast_qty > 0 else 0
    }


@router.get("/cpc/demand-forecasts/download")
async def download_demand_forecasts():
    """Download confirmed demand forecasts as Excel file for CPC"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Get confirmed forecasts
    forecasts = await db.forecasts.find(
        {"status": {"$in": ["CONFIRMED", "CONVERTED"]}},
        {"_id": 0}
    ).sort("forecast_month", 1).to_list(5000)
    
    # Get all production schedules for remaining qty calculation
    all_schedules = await db.production_schedules.find(
        {"forecast_id": {"$exists": True, "$ne": None}, "status": {"$ne": "CANCELLED"}},
        {"_id": 0, "forecast_id": 1, "target_quantity": 1}
    ).to_list(10000)
    
    scheduled_by_forecast = {}
    for s in all_schedules:
        fid = s.get("forecast_id")
        if fid:
            scheduled_by_forecast[fid] = scheduled_by_forecast.get(fid, 0) + s.get("target_quantity", 0)
    
    # Get buyer, vertical, SKU details
    buyer_ids = list(set(f.get("buyer_id") for f in forecasts if f.get("buyer_id")))
    buyers = await db.buyers.find({"id": {"$in": buyer_ids}}, {"_id": 0, "id": 1, "name": 1, "code": 1}).to_list(100)
    buyer_map = {b["id"]: b for b in buyers}
    
    sku_ids = list(set(f.get("sku_id") for f in forecasts if f.get("sku_id")))
    skus = await sku_service.get_skus_by_sku_ids(sku_ids)
    sku_map = {s["sku_id"]: s for s in skus}
    
    vertical_ids = list(set(f.get("vertical_id") for f in forecasts if f.get("vertical_id")))
    verticals = await db.verticals.find({"id": {"$in": vertical_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    vertical_map = {v["id"]: v["name"] for v in verticals}
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demand Forecasts"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Forecast ID", "Buyer Code", "Buyer Name", "Vertical", "Brand", "Model",
        "SKU ID", "SKU Description", "Forecast Month", "Forecast Qty",
        "Scheduled Qty", "Remaining Qty", "Priority", "Status"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, f in enumerate(forecasts, 2):
        buyer = buyer_map.get(f.get("buyer_id"), {})
        sku = sku_map.get(f.get("sku_id"), {})
        forecast_qty = f.get("quantity", 0)
        scheduled_qty = scheduled_by_forecast.get(f.get("id"), 0)
        remaining_qty = max(0, forecast_qty - scheduled_qty)
        
        row_data = [
            f.get("forecast_code", ""),
            buyer.get("code", ""),
            buyer.get("name", ""),
            vertical_map.get(f.get("vertical_id"), sku.get("vertical", "")),
            sku.get("brand", ""),
            sku.get("model", ""),
            f.get("sku_id", ""),
            sku.get("description", ""),
            f.get("forecast_month").strftime("%Y-%m") if isinstance(f.get("forecast_month"), datetime) else str(f.get("forecast_month", ""))[:7],
            forecast_qty,
            scheduled_qty,
            remaining_qty,
            f.get("priority", "MEDIUM"),
            f.get("status", "")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in [10, 11, 12]:  # Numeric columns
                cell.alignment = Alignment(horizontal="right")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"demand_forecasts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/cpc/production-plan/template")
async def download_production_plan_template():
    """Download Excel template for bulk production plan upload.
    Format: Branch ID | Date | Buyer SKU ID | Quantity
    """
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Plan"
    
    # Headers - Simplified: Branch ID | Date | Buyer SKU ID | Quantity
    headers = ["Branch ID", "Date (DD-MM-YYYY)", "Buyer SKU ID", "Quantity"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    
    # Get sample SKUs and branches for reference
    buyer_skus = await db.buyer_skus.find(
        {"status": "ACTIVE"},
        {"_id": 0, "buyer_sku_id": 1, "name": 1}
    ).to_list(20)
    
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1, "branch_id": 1}).to_list(100)
    
    # Generate branch_id if missing (BR_001, BR_002, etc.)
    for idx, b in enumerate(branches, 1):
        if not b.get("branch_id"):
            b["branch_id"] = f"BR_{idx:03d}"
    
    # Add sample rows
    row = 2
    sample_date = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%d-%m-%Y")
    for sku in buyer_skus[:5]:
        ws.cell(row=row, column=1, value=branches[0].get("branch_id", "BR_001") if branches else "BR_001")
        ws.cell(row=row, column=2, value=sample_date)
        ws.cell(row=row, column=3, value=sku.get("buyer_sku_id", ""))
        ws.cell(row=row, column=4, value=100)
        row += 1
    
    # Branches reference sheet
    ws_branches = wb.create_sheet("Branches Reference")
    ws_branches.cell(row=1, column=1, value="Branch ID")
    ws_branches.cell(row=1, column=2, value="Branch Name")
    ws_branches.cell(row=1, column=3, value="Capacity/Day")
    ws_branches["A1"].font = Font(bold=True)
    ws_branches["B1"].font = Font(bold=True)
    ws_branches["C1"].font = Font(bold=True)
    
    all_branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1, "capacity_units_per_day": 1}).to_list(100)
    
    # Generate branch_id if missing for all branches
    for idx, b in enumerate(all_branches, 1):
        if not b.get("branch_id"):
            b["branch_id"] = f"BR_{idx:03d}"
    
    for i, b in enumerate(all_branches, 2):
        ws_branches.cell(row=i, column=1, value=b.get("branch_id", ""))
        ws_branches.cell(row=i, column=2, value=b.get("name", ""))
        ws_branches.cell(row=i, column=3, value=b.get("capacity_units_per_day", 0))
    
    ws_branches.column_dimensions["A"].width = 12
    ws_branches.column_dimensions["B"].width = 25
    ws_branches.column_dimensions["C"].width = 15
    
    # Buyer SKUs reference sheet
    ws_skus = wb.create_sheet("Buyer SKUs Reference")
    ws_skus.cell(row=1, column=1, value="Buyer SKU ID")
    ws_skus.cell(row=1, column=2, value="Name")
    ws_skus["A1"].font = Font(bold=True)
    ws_skus["B1"].font = Font(bold=True)
    
    all_skus = await db.buyer_skus.find(
        {"status": "ACTIVE"},
        {"_id": 0, "buyer_sku_id": 1, "name": 1}
    ).to_list(500)
    for i, s in enumerate(all_skus, 2):
        ws_skus.cell(row=i, column=1, value=s.get("buyer_sku_id", ""))
        ws_skus.cell(row=i, column=2, value=s.get("name", ""))
    
    ws_skus.column_dimensions["A"].width = 25
    ws_skus.column_dimensions["B"].width = 40
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=production_plan_template.xlsx"}
    )


@router.post("/cpc/production-plan/upload-excel")
async def upload_production_plan_excel(
    file: UploadFile = File(...),
    mode: str = Query("check", description="Mode: 'check' (default) to check for conflicts, 'override' to clear existing schedules, 'add' to add to remaining capacity")
):
    """
    Bulk upload production plans from Excel with FIFO allocation.
    
    Format: Branch ID | Date | Buyer SKU ID | Quantity
    
    Modes:
    - 'check': Validate and check for conflicts (default). Returns warning if existing schedules + capacity breach.
    - 'override': Clear existing schedules for conflicting dates/branches, then allocate fresh.
    - 'add': Keep existing schedules, allocate only within remaining capacity.
    
    Returns downloadable Excel with status for each row.
    """
    if not pd:
        raise HTTPException(status_code=500, detail="pandas not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    
    # Normalize column names
    df.columns = [str(c).strip().lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_") for c in df.columns]
    
    # Map columns
    col_map = {}
    branch_opts = ["branch_id", "branch", "branch_name", "unit"]
    date_opts = ["date", "target_date", "date_dd_mm_yyyy", "date_yyyy_mm_dd"]
    sku_opts = ["buyer_sku_id", "sku_id", "sku", "buyer_sku"]
    qty_opts = ["quantity", "qty", "plan_qty"]
    
    for opt in branch_opts:
        if opt in df.columns:
            col_map["branch"] = opt
            break
    for opt in date_opts:
        if opt in df.columns:
            col_map["date"] = opt
            break
    for opt in sku_opts:
        if opt in df.columns:
            col_map["sku"] = opt
            break
    for opt in qty_opts:
        if opt in df.columns:
            col_map["qty"] = opt
            break
    
    required = ["branch", "date", "sku", "qty"]
    missing = [r for r in required if r not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}. Found: {list(df.columns)}. Expected: Branch ID, Date, Buyer SKU ID, Quantity")
    
    # Load lookup data
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1, "capacity_units_per_day": 1}).to_list(100)
    branch_id_to_name = {b.get("branch_id", ""): b["name"] for b in branches if b.get("branch_id")}
    branch_name_to_name = {b["name"]: b["name"] for b in branches}
    branch_capacity_map = {b["name"]: b.get("capacity_units_per_day", 0) for b in branches}
    
    # Get all Buyer SKUs
    buyer_skus = await db.buyer_skus.find(
        {"status": "ACTIVE"},
        {"_id": 0, "buyer_sku_id": 1, "name": 1, "description": 1}
    ).to_list(10000)
    sku_map = {s["buyer_sku_id"]: s for s in buyer_skus}
    
    # Get SKU-Branch subscriptions
    sku_assignments = await db.sku_branch_assignments.find({}, {"_id": 0, "sku_id": 1, "branch": 1}).to_list(20000)
    sku_branch_subscribed = {(a["sku_id"], a["branch"]): True for a in sku_assignments}
    
    # Parse and validate all rows first, preserving order
    parsed_rows = []
    row_results = []  # Store results for each row
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)
        result = {
            "row": row_num,
            "branch_id": str(row[col_map["branch"]]).strip() if pd.notna(row[col_map["branch"]]) else "",
            "date": "",
            "buyer_sku_id": str(row[col_map["sku"]]).strip() if pd.notna(row[col_map["sku"]]) else "",
            "quantity": 0,
            "status": "PENDING",
            "allocated": 0,
            "not_allocated": 0,
            "schedule_code": "",
            "remarks": ""
        }
        
        try:
            branch_value = result["branch_id"]
            sku_id = result["buyer_sku_id"]
            qty = int(row[col_map["qty"]]) if pd.notna(row[col_map["qty"]]) else 0
            result["quantity"] = qty
            
            # Parse date
            date_val = row[col_map["date"]]
            if pd.isna(date_val):
                result["status"] = "ERROR"
                result["remarks"] = "Date is required"
                row_results.append(result)
                continue
                
            if isinstance(date_val, str):
                date_str = date_val.strip()
                try:
                    target_date = datetime.strptime(date_str, "%d-%m-%Y").replace(tzinfo=timezone.utc)
                except:
                    try:
                        target_date = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    except:
                        target_date = pd.to_datetime(date_val, dayfirst=True).to_pydatetime().replace(tzinfo=timezone.utc)
            else:
                target_date = pd.to_datetime(date_val, dayfirst=True).to_pydatetime().replace(tzinfo=timezone.utc)
            
            result["date"] = target_date.strftime("%d-%m-%Y")
            date_str_iso = target_date.strftime("%Y-%m-%d")
            
            # Resolve branch
            branch_name = branch_id_to_name.get(branch_value) or branch_name_to_name.get(branch_value)
            if not branch_name or branch_name not in branch_capacity_map:
                result["status"] = "ERROR"
                result["remarks"] = f"Invalid branch '{branch_value}'"
                row_results.append(result)
                continue
            
            # Validate SKU
            if sku_id not in sku_map:
                result["status"] = "ERROR"
                result["remarks"] = f"Buyer SKU '{sku_id}' not found"
                row_results.append(result)
                continue
            
            # Validate SKU-Branch subscription
            if (sku_id, branch_name) not in sku_branch_subscribed:
                result["status"] = "ERROR"
                result["remarks"] = f"SKU not subscribed to branch '{branch_name}'"
                row_results.append(result)
                continue
            
            # Validate quantity
            if qty <= 0:
                result["status"] = "ERROR"
                result["remarks"] = f"Invalid quantity {qty}"
                row_results.append(result)
                continue
            
            # Valid row - add to parsed rows for processing
            parsed_rows.append({
                "row_idx": len(row_results),  # Index in row_results
                "row_num": row_num,
                "branch_name": branch_name,
                "branch_value": branch_value,
                "date_iso": date_str_iso,
                "target_date": target_date,
                "sku_id": sku_id,
                "sku": sku_map[sku_id],
                "qty": qty
            })
            row_results.append(result)
            
        except Exception as e:
            result["status"] = "ERROR"
            result["remarks"] = str(e)
            row_results.append(result)
    
    # Group rows by (date, branch) for capacity analysis
    date_branch_groups = {}
    for pr in parsed_rows:
        key = f"{pr['date_iso']}|{pr['branch_name']}"
        if key not in date_branch_groups:
            date_branch_groups[key] = {
                "date_iso": pr["date_iso"],
                "branch_name": pr["branch_name"],
                "rows": [],
                "total_demand": 0
            }
        date_branch_groups[key]["rows"].append(pr)
        date_branch_groups[key]["total_demand"] += pr["qty"]
    
    # Check for existing schedules and calculate conflicts
    conflicts = []
    for key, group in date_branch_groups.items():
        date_iso = group["date_iso"]
        branch_name = group["branch_name"]
        
        # Get branch capacity (with daily override)
        base_capacity = branch_capacity_map.get(branch_name, 0)
        daily_cap = await db.branch_daily_capacity.find_one(
            {"branch": branch_name, "date": date_iso},
            {"_id": 0, "capacity": 1}
        )
        capacity = daily_cap.get("capacity", base_capacity) if daily_cap else base_capacity
        
        # Get existing schedules
        day_start = datetime.strptime(date_iso, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        day_end = day_start + timedelta(days=1)
        existing_schedules = await db.production_schedules.find(
            {
                "branch": branch_name,
                "target_date": {"$gte": day_start, "$lt": day_end},
                "status": {"$ne": "CANCELLED"}
            },
            {"_id": 0, "id": 1, "schedule_code": 1, "sku_id": 1, "target_quantity": 1}
        ).to_list(1000)
        existing_scheduled = sum(s.get("target_quantity", 0) for s in existing_schedules)
        
        group["capacity"] = capacity
        group["existing_scheduled"] = existing_scheduled
        group["existing_schedules"] = existing_schedules
        group["available"] = capacity - existing_scheduled
        
        # Check for conflict: existing schedules + capacity breach
        if existing_scheduled > 0 and group["total_demand"] > group["available"]:
            conflicts.append({
                "date": day_start.strftime("%d-%m-%Y"),
                "date_iso": date_iso,
                "branch": branch_name,
                "capacity": capacity,
                "existing_scheduled": existing_scheduled,
                "new_demand": group["total_demand"],
                "total_demand": existing_scheduled + group["total_demand"],
                "available": group["available"],
                "overflow": group["total_demand"] - group["available"]
            })
    
    # If mode is 'check' and there are conflicts, return warning
    if mode == "check" and conflicts:
        return {
            "success": False,
            "warning": True,
            "message": "Capacity breach detected with existing schedules. Please confirm action.",
            "conflicts": conflicts,
            "total_rows": len(df),
            "valid_rows": len(parsed_rows),
            "error_rows": len(row_results) - len(parsed_rows),
            "options": {
                "override": "Clear existing schedules and allocate new demand up to capacity",
                "add": "Keep existing schedules, allocate only within remaining capacity"
            },
            "instruction": "Re-upload with mode='override' or mode='add' query parameter"
        }
    
    # Process with FIFO allocation
    created_count = 0
    
    # Pre-load deleted completions for auto-population
    # Group all months/branches we need
    months_branches_needed = set()
    for key, group in date_branch_groups.items():
        date_obj = datetime.strptime(group["date_iso"], "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")
        months_branches_needed.add((month_key, group["branch_name"]))
    
    # Load all deleted completions
    deleted_completions_map = {}  # key: "{date}_{sku_id}" -> {completed_quantity, deleted_schedule_id}
    for month, branch in months_branches_needed:
        try:
            year, month_num = month.split("-")
            year = int(year)
            month_num = int(month_num)
            month_start = datetime(year, month_num, 1, 0, 0, 0, tzinfo=timezone.utc)
            
            if month_num == 12:
                month_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
            else:
                month_end = datetime(year, month_num + 1, 1, 0, 0, 0, tzinfo=timezone.utc)
            
            deleted_schedules = await db.production_schedules.find({
                "branch": branch,
                "target_date": {"$gte": month_start, "$lt": month_end},
                "status": "DELETED",
                "completed_quantity": {"$gt": 0}
            }, {"_id": 0}).to_list(5000)
            
            for s in deleted_schedules:
                target_date = s.get("target_date")
                if isinstance(target_date, datetime):
                    date_key = target_date.strftime("%Y-%m-%d")
                else:
                    date_key = str(target_date)[:10]
                
                sku_id = s.get("sku_id")
                lookup_key = f"{date_key}_{sku_id}_{branch}"
                
                deleted_completions_map[lookup_key] = {
                    "completed_quantity": s.get("completed_quantity", 0),
                    "deleted_schedule_id": s.get("id"),
                    "deleted_schedule_code": s.get("schedule_code"),
                    "target_date": date_key,
                    "sku_id": sku_id,
                    "branch": branch
                }
        except Exception as e:
            logger.warning(f"Failed to load deleted completions for {month}/{branch}: {e}")
    
    logger.info(f"Loaded {len(deleted_completions_map)} deleted completions for auto-population")
    
    for key, group in date_branch_groups.items():
        capacity = group["capacity"]
        existing_scheduled = group["existing_scheduled"]
        
        # Determine available capacity based on mode
        if mode == "override" and existing_scheduled > 0:
            # Clear existing schedules for this date/branch
            day_start = datetime.strptime(group["date_iso"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
            day_end = day_start + timedelta(days=1)
            
            deleted = await db.production_schedules.delete_many({
                "branch": group["branch_name"],
                "target_date": {"$gte": day_start, "$lt": day_end},
                "status": {"$ne": "CANCELLED"}
            })
            
            available = capacity  # Full capacity available after clearing
        else:
            available = capacity - existing_scheduled
        
        # FIFO allocation - process rows in order
        for pr in group["rows"]:
            result = row_results[pr["row_idx"]]
            qty = pr["qty"]
            
            if available <= 0:
                # No capacity - reject
                result["status"] = "REJECTED"
                result["allocated"] = 0
                result["not_allocated"] = qty
                result["remarks"] = f"No capacity available. Branch capacity: {capacity}, Already scheduled: {capacity - available}"
                continue
            
            if qty <= available:
                # Full allocation
                allocated_qty = qty
                overflow_qty = 0
                result["status"] = "SCHEDULED"
                result["remarks"] = "Fully allocated"
            else:
                # Partial allocation
                allocated_qty = available
                overflow_qty = qty - available
                result["status"] = "PARTIAL"
                result["remarks"] = f"Capacity limit reached. {overflow_qty} units overflow."
            
            result["allocated"] = allocated_qty
            result["not_allocated"] = overflow_qty
            
            # Create schedule
            count = await db.production_schedules.count_documents({})
            schedule_code = f"PS_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
            
            # Check for deleted completion to auto-populate completed_quantity
            date_key = pr["target_date"].strftime("%Y-%m-%d") if isinstance(pr["target_date"], datetime) else str(pr["target_date"])[:10]
            completion_lookup_key = f"{date_key}_{pr['sku_id']}_{pr['branch_name']}"
            deleted_completion = deleted_completions_map.get(completion_lookup_key)
            
            completed_qty = 0
            replaced_schedule_id = None
            replaced_schedule_code = None
            
            if deleted_completion:
                completed_qty = deleted_completion.get("completed_quantity", 0)
                replaced_schedule_id = deleted_completion.get("deleted_schedule_id")
                replaced_schedule_code = deleted_completion.get("deleted_schedule_code")
                logger.info(f"Auto-populating completed_quantity={completed_qty} from deleted schedule {replaced_schedule_code} for {pr['sku_id']} on {date_key}")
            
            schedule = {
                "id": str(uuid.uuid4()),
                "schedule_code": schedule_code,
                "forecast_id": None,
                "dispatch_lot_id": None,
                "branch": pr["branch_name"],
                "sku_id": pr["sku_id"],
                "sku_description": pr["sku"].get("name", "") or pr["sku"].get("description", ""),
                "target_quantity": allocated_qty,
                "allocated_quantity": allocated_qty,
                "completed_quantity": completed_qty,
                "target_date": pr["target_date"],
                "priority": "MEDIUM",
                "status": "COMPLETED" if completed_qty >= allocated_qty else "SCHEDULED",
                "replaced_schedule_id": replaced_schedule_id,
                "replaced_schedule_code": replaced_schedule_code,
                "notes": f"Bulk upload Row {pr['row_num']}" + (f" (partial: {qty} requested, {allocated_qty} allocated)" if overflow_qty > 0 else "") + (f" | Recovered {completed_qty} completed from deleted schedule {replaced_schedule_code}" if replaced_schedule_code else ""),
                "created_at": datetime.now(timezone.utc)
            }
            
            await db.production_schedules.insert_one(schedule)
            result["schedule_code"] = schedule_code
            
            # Add completion info to result
            if completed_qty > 0:
                result["remarks"] += f" | Auto-populated {completed_qty} completed units from deleted schedule"
            
            created_count += 1
            
            # Update available capacity
            available -= allocated_qty
    
    # AUTO-RECOVER ORPHANED COMPLETIONS
    # Find deleted SKUs with completions that were NOT in the upload and create schedules for them
    used_completion_keys = set()
    for key, group in date_branch_groups.items():
        for pr in group["rows"]:
            date_key = pr["target_date"].strftime("%Y-%m-%d") if isinstance(pr["target_date"], datetime) else str(pr["target_date"])[:10]
            used_completion_keys.add(f"{date_key}_{pr['sku_id']}_{pr['branch_name']}")
    
    orphaned_completions = {k: v for k, v in deleted_completions_map.items() if k not in used_completion_keys}
    
    if orphaned_completions:
        logger.info(f"Found {len(orphaned_completions)} orphaned completions to auto-recover")
        
        for lookup_key, completion in orphaned_completions.items():
            # Parse the key: "date_skuid_branch"
            parts = lookup_key.rsplit("_", 2)  # Split from right to handle SKU IDs with underscores
            if len(parts) < 3:
                continue
            
            # Key format is "{date}_{sku_id}_{branch}" but sku_id might have underscores
            # So we need to be smarter - use the completion data instead
            date_key = completion.get("target_date") or parts[0]
            sku_id = completion.get("sku_id")
            branch = completion.get("branch")
            
            if not sku_id or not branch:
                # Try to extract from the original deleted schedule
                deleted_schedule = await db.production_schedules.find_one(
                    {"id": completion.get("deleted_schedule_id")},
                    {"_id": 0, "sku_id": 1, "branch": 1, "target_date": 1, "sku_description": 1}
                )
                if deleted_schedule:
                    sku_id = deleted_schedule.get("sku_id")
                    branch = deleted_schedule.get("branch")
                    date_key = deleted_schedule.get("target_date")
                    if isinstance(date_key, datetime):
                        date_key = date_key.strftime("%Y-%m-%d")
            
            if not sku_id or not branch:
                logger.warning(f"Cannot recover orphaned completion: {lookup_key}")
                continue
            
            completed_qty = completion.get("completed_quantity", 0)
            
            # Create auto-recovered schedule
            count = await db.production_schedules.count_documents({})
            schedule_code = f"PS_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
            
            # Parse date
            try:
                if isinstance(date_key, str):
                    target_date = datetime.strptime(date_key, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                else:
                    target_date = date_key
            except:
                continue
            
            # Get SKU description
            sku_doc = await db.buyer_skus.find_one({"buyer_sku_id": sku_id}, {"_id": 0, "name": 1, "description": 1})
            sku_description = ""
            if sku_doc:
                sku_description = sku_doc.get("name") or sku_doc.get("description") or ""
            
            schedule = {
                "id": str(uuid.uuid4()),
                "schedule_code": schedule_code,
                "forecast_id": None,
                "dispatch_lot_id": None,
                "branch": branch,
                "sku_id": sku_id,
                "sku_description": sku_description,
                "target_quantity": completed_qty,  # Target = Completed (fully done)
                "allocated_quantity": completed_qty,
                "completed_quantity": completed_qty,
                "target_date": target_date,
                "priority": "MEDIUM",
                "status": "COMPLETED",
                "replaced_schedule_id": completion.get("deleted_schedule_id"),
                "replaced_schedule_code": completion.get("deleted_schedule_code"),
                "notes": f"Auto-recovered from deleted schedule {completion.get('deleted_schedule_code')} (SKU not in new upload)",
                "created_at": datetime.now(timezone.utc)
            }
            
            await db.production_schedules.insert_one(schedule)
            created_count += 1
            
            # Add to result data
            row_results.append({
                "row": "AUTO",
                "branch_id": branch,
                "date": target_date.strftime("%d-%m-%Y"),
                "buyer_sku_id": sku_id,
                "quantity": completed_qty,
                "status": "RECOVERED",
                "allocated": completed_qty,
                "not_allocated": 0,
                "schedule_code": schedule_code,
                "remarks": f"Auto-recovered completed SKU from deleted schedule {completion.get('deleted_schedule_code')}"
            })
            
            logger.info(f"Auto-recovered orphaned completion: {sku_id} on {date_key} at {branch} with {completed_qty} units")

    # Generate result Excel file
    result_data = []
    for r in row_results:
        result_data.append({
            "Branch ID": r["branch_id"],
            "Date (DD-MM-YYYY)": r["date"],
            "Buyer SKU ID": r["buyer_sku_id"],
            "Quantity": r["quantity"],
            "Status": r["status"],
            "Allocated": r["allocated"],
            "Not Allocated": r["not_allocated"],
            "Schedule Code": r["schedule_code"],
            "Remarks": r["remarks"]
        })
    
    result_df = pd.DataFrame(result_data)
    
    # Create Excel with formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='Upload Result')
        
        # Apply formatting
        ws = writer.sheets['Upload Result']
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 50
        
        # Header styling
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Status column color coding
        status_colors = {
            "SCHEDULED": "C6EFCE",  # Green
            "PARTIAL": "FFEB9C",    # Yellow
            "REJECTED": "FFC7CE",   # Red
            "ERROR": "D9D9D9",      # Gray
            "PENDING": "FFFFFF"     # White
        }
        
        for row_num in range(2, len(result_data) + 2):
            status_cell = ws.cell(row=row_num, column=5)
            status = status_cell.value
            if status in status_colors:
                status_cell.fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type="solid")
    
    output.seek(0)
    
    # Generate upload ID for download reference
    upload_id = str(uuid.uuid4())[:8]
    
    # Store result in memory/cache for download (simplified - in production use Redis/DB)
    # For now, return the file directly in base64 or as download
    
    # Calculate summary stats
    scheduled_count = sum(1 for r in row_results if r["status"] == "SCHEDULED")
    partial_count = sum(1 for r in row_results if r["status"] == "PARTIAL")
    rejected_count = sum(1 for r in row_results if r["status"] == "REJECTED")
    error_count = sum(1 for r in row_results if r["status"] == "ERROR")
    total_allocated = sum(r["allocated"] for r in row_results)
    total_not_allocated = sum(r["not_allocated"] for r in row_results)
    
    # Return file as StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=production_plan_result_{upload_id}.xlsx",
            "X-Upload-Summary": json.dumps({
                "success": True,
                "upload_id": upload_id,
                "mode": mode,
                "total_rows": len(df),
                "scheduled": scheduled_count,
                "partial": partial_count,
                "rejected": rejected_count,
                "errors": error_count,
                "total_allocated": total_allocated,
                "total_not_allocated": total_not_allocated,
                "schedules_created": created_count
            })
        }
    )


@router.get("/cpc/available-capacity")
async def get_available_capacity(
    start_date: str = None,
    end_date: str = None,
    branch: str = None
):
    """
    Get available capacity day-wise for branches.
    Returns: For each branch/date combo: total capacity, scheduled, available.
    """
    # Default date range: today + 30 days
    if not start_date:
        start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not end_date:
        end_dt = datetime.now(timezone.utc) + timedelta(days=30)
        end_date = end_dt.strftime("%Y-%m-%d")
    
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    # Get branches
    branch_query = {"is_active": True}
    if branch:
        branch_query["name"] = branch
    branches = await db.branches.find(branch_query, {"_id": 0, "branch_id": 1, "name": 1, "capacity_units_per_day": 1}).to_list(100)
    
    # Get daily capacity overrides
    daily_overrides = await db.branch_daily_capacity.find(
        {"date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0}
    ).to_list(5000)
    override_map = {f"{d['branch']}|{d['date']}": d.get("capacity", 0) for d in daily_overrides}
    
    # Get scheduled quantities
    schedules = await db.production_schedules.find(
        {
            "target_date": {"$gte": start_dt, "$lte": end_dt + timedelta(days=1)},
            "status": {"$ne": "CANCELLED"}
        },
        {"_id": 0, "branch": 1, "target_date": 1, "target_quantity": 1}
    ).to_list(10000)
    
    # Build scheduled map
    scheduled_map = {}
    for s in schedules:
        dt = s["target_date"]
        if isinstance(dt, datetime):
            date_key = dt.strftime("%Y-%m-%d")
        else:
            date_key = str(dt)[:10]
        key = f"{s['branch']}|{date_key}"
        scheduled_map[key] = scheduled_map.get(key, 0) + s.get("target_quantity", 0)
    
    # Build result
    result = []
    current = start_dt
    while current <= end_dt:
        date_str = current.strftime("%Y-%m-%d")
        for b in branches:
            branch_name = b["name"]
            key = f"{branch_name}|{date_str}"
            
            # Get capacity (override or default)
            capacity = override_map.get(key, b.get("capacity_units_per_day", 0))
            scheduled = scheduled_map.get(key, 0)
            available = max(0, capacity - scheduled)
            
            result.append({
                "branch_id": b.get("branch_id", ""),
                "branch": branch_name,
                "date": date_str,
                "capacity": capacity,
                "scheduled": scheduled,
                "available": available,
                "utilization_pct": round((scheduled / capacity * 100) if capacity > 0 else 0, 1)
            })
        current += timedelta(days=1)
    
    return {
        "start_date": start_date,
        "end_date": end_date,
        "data": result,
        "total_rows": len(result)
    }


@router.get("/cpc/available-capacity/download")
async def download_available_capacity(
    start_date: str = None,
    end_date: str = None,
    branch: str = None
):
    """Download available capacity day-wise as Excel file."""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Default date range: today + 30 days
    if not start_date:
        start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not end_date:
        end_dt = datetime.now(timezone.utc) + timedelta(days=30)
        end_date = end_dt.strftime("%Y-%m-%d")
    
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    # Get branches
    branch_query = {"is_active": True}
    if branch:
        branch_query["name"] = branch
    branches = await db.branches.find(branch_query, {"_id": 0, "branch_id": 1, "name": 1, "capacity_units_per_day": 1}).to_list(100)
    
    # Get daily capacity overrides
    daily_overrides = await db.branch_daily_capacity.find(
        {"date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0}
    ).to_list(5000)
    override_map = {f"{d['branch']}|{d['date']}": d.get("capacity", 0) for d in daily_overrides}
    
    # Get scheduled quantities
    schedules = await db.production_schedules.find(
        {
            "target_date": {"$gte": start_dt, "$lte": end_dt + timedelta(days=1)},
            "status": {"$ne": "CANCELLED"}
        },
        {"_id": 0, "branch": 1, "target_date": 1, "target_quantity": 1}
    ).to_list(10000)
    
    # Build scheduled map
    scheduled_map = {}
    for s in schedules:
        dt = s["target_date"]
        if isinstance(dt, datetime):
            date_key = dt.strftime("%Y-%m-%d")
        else:
            date_key = str(dt)[:10]
        key = f"{s['branch']}|{date_key}"
        scheduled_map[key] = scheduled_map.get(key, 0) + s.get("target_quantity", 0)
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Available Capacity"
    
    # Headers
    headers = ["Branch ID", "Branch", "Date", "Capacity", "Scheduled", "Available", "Utilization %"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row_num = 2
    current = start_dt
    while current <= end_dt:
        date_str = current.strftime("%Y-%m-%d")
        for b in branches:
            branch_name = b["name"]
            key = f"{branch_name}|{date_str}"
            
            capacity = override_map.get(key, b.get("capacity_units_per_day", 0))
            scheduled = scheduled_map.get(key, 0)
            available = max(0, capacity - scheduled)
            utilization = round((scheduled / capacity * 100) if capacity > 0 else 0, 1)
            
            ws.cell(row=row_num, column=1, value=b.get("branch_id", ""))
            ws.cell(row=row_num, column=2, value=branch_name)
            ws.cell(row=row_num, column=3, value=date_str)
            ws.cell(row=row_num, column=4, value=capacity)
            ws.cell(row=row_num, column=5, value=scheduled)
            ws.cell(row=row_num, column=6, value=available)
            ws.cell(row=row_num, column=7, value=utilization)
            
            # Highlight low availability
            if available < capacity * 0.2:  # Less than 20% available
                for c in range(1, 8):
                    ws.cell(row=row_num, column=c).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            
            row_num += 1
        current += timedelta(days=1)
    
    # Auto-width columns
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"available_capacity_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/cpc/allocate-overflow")
async def allocate_overflow_schedule(data: dict):
    """
    Allocate overflow quantity to a different date.
    Input: {sku_id, branch, date, quantity}
    """
    sku_id = data.get("sku_id")
    branch = data.get("branch")
    date_str = data.get("date")
    qty = int(data.get("quantity", 0))
    
    if not all([sku_id, branch, date_str, qty > 0]):
        raise HTTPException(status_code=400, detail="Missing required fields: sku_id, branch, date, quantity")
    
    # Validate branch
    branch_doc = await db.branches.find_one({"name": branch, "is_active": True}, {"_id": 0, "capacity_units_per_day": 1})
    if not branch_doc:
        raise HTTPException(status_code=400, detail=f"Branch '{branch}' not found")
    
    # Validate SKU
    sku = await db.buyer_skus.find_one({"buyer_sku_id": sku_id, "status": "ACTIVE"}, {"_id": 0, "name": 1, "description": 1})
    if not sku:
        raise HTTPException(status_code=400, detail=f"Buyer SKU '{sku_id}' not found")
    
    # Parse date
    target_date = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    # Check capacity
    branch_capacity = branch_doc.get("capacity_units_per_day", 0)
    daily_cap = await db.branch_daily_capacity.find_one(
        {"branch": branch, "date": date_str},
        {"_id": 0, "capacity": 1}
    )
    if daily_cap:
        branch_capacity = daily_cap.get("capacity", branch_capacity)
    
    # Get existing schedules
    day_start = datetime(target_date.year, target_date.month, target_date.day, tzinfo=timezone.utc)
    day_end = day_start + timedelta(days=1)
    existing = await db.production_schedules.find(
        {
            "branch": branch,
            "target_date": {"$gte": day_start, "$lt": day_end},
            "status": {"$ne": "CANCELLED"}
        },
        {"_id": 0, "target_quantity": 1}
    ).to_list(1000)
    existing_allocated = sum(s.get("target_quantity", 0) for s in existing)
    
    available = branch_capacity - existing_allocated
    if qty > available:
        raise HTTPException(
            status_code=400, 
            detail=f"Requested qty {qty} exceeds available capacity {available} for {branch} on {date_str}"
        )
    
    # Create schedule
    count = await db.production_schedules.count_documents({})
    schedule_code = f"PS_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    schedule = {
        "id": str(uuid.uuid4()),
        "schedule_code": schedule_code,
        "forecast_id": None,
        "dispatch_lot_id": None,
        "branch": branch,
        "sku_id": sku_id,
        "sku_description": sku.get("name", "") or sku.get("description", ""),
        "target_quantity": qty,
        "allocated_quantity": qty,
        "completed_quantity": 0,
        "target_date": target_date,
        "priority": "MEDIUM",
        "status": "SCHEDULED",
        "notes": "Overflow allocation",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.production_schedules.insert_one(schedule)
    
    return {
        "success": True,
        "schedule_code": schedule_code,
        "message": f"Created schedule {schedule_code} for {qty} units on {date_str}"
    }


class ScheduleFromForecastRequest(BaseModel):
    forecast_id: str
    quantity: int
    target_date: datetime
    branch: str  # REQUIRED: Branch assignment is mandatory
    priority: Optional[str] = "MEDIUM"
    notes: Optional[str] = ""


# NEW: Model capacity upload model
class BranchModelCapacityUpload(BaseModel):
    month: str  # Format: "2026-03"
    day: int    # 1-31
    model_id: str
    capacity_qty: int


class BranchModelCapacityBulkUpload(BaseModel):
    branch: str
    capacities: List[BranchModelCapacityUpload]


# ============ BRANCH MODEL CAPACITY ENDPOINTS ============

@router.post("/branches/model-capacity/upload")
async def upload_branch_model_capacity(data: BranchModelCapacityBulkUpload):
    """Upload model-specific capacity for a branch (Month, Day, Model, Qty)"""
    if not data.capacities:
        raise HTTPException(status_code=400, detail="No capacity data provided")
    
    # Validate branch exists
    branch = await db.branches.find_one({"name": data.branch, "is_active": True})
    if not branch:
        raise HTTPException(status_code=404, detail=f"Branch '{data.branch}' not found")
    
    # Validate models exist
    model_ids = list(set(c.model_id for c in data.capacities))
    existing_models = await db.models.find(
        {"id": {"$in": model_ids}},
        {"_id": 0, "id": 1}
    ).to_list(1000)
    existing_model_ids = {m["id"] for m in existing_models}
    
    missing_models = set(model_ids) - existing_model_ids
    if missing_models:
        raise HTTPException(status_code=400, detail=f"Models not found: {', '.join(missing_models)}")
    
    # Insert or update capacities
    inserted = 0
    updated = 0
    
    for cap in data.capacities:
        record = {
            "branch": data.branch,
            "month": cap.month,
            "day": cap.day,
            "model_id": cap.model_id,
            "capacity_qty": cap.capacity_qty,
            "updated_at": datetime.now(timezone.utc)
        }
        
        result = await db.branch_model_capacity.update_one(
            {
                "branch": data.branch,
                "month": cap.month,
                "day": cap.day,
                "model_id": cap.model_id
            },
            {"$set": record, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc)}},
            upsert=True
        )
        
        if result.upserted_id:
            inserted += 1
        else:
            updated += 1
    
    return {
        "message": f"Capacity uploaded for branch {data.branch}",
        "inserted": inserted,
        "updated": updated,
        "total": len(data.capacities)
    }


@router.get("/branches/{branch}/model-capacity")
async def get_branch_model_capacity(branch: str, month: Optional[str] = None):
    """Get model-specific capacity for a branch"""
    query = {"branch": branch}
    if month:
        query["month"] = month
    
    capacities = await db.branch_model_capacity.find(
        query,
        {"_id": 0}
    ).to_list(5000)
    
    return capacities


@router.get("/branches/{branch}/capacity-for-date")
async def get_branch_capacity_for_date(branch: str, date_str: str, model_id: Optional[str] = None):
    """Get available capacity for a branch on a specific date, optionally filtered by model.
    Priority: 1) Model-specific capacity, 2) Daily override, 3) Base capacity
    """
    # Parse date
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_str = date_obj.strftime("%Y-%m")
        day = date_obj.day
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Get branch base capacity
    branch_cap = await db.branches.find_one({"name": branch, "is_active": True}, {"_id": 0})
    if not branch_cap:
        raise HTTPException(status_code=404, detail=f"Branch '{branch}' not found")
    
    base_capacity = branch_cap.get("capacity_units_per_day", 0)
    
    # Check for daily capacity override
    daily_override = await db.branch_daily_capacity.find_one({
        "branch": branch,
        "date": date_str
    }, {"_id": 0})
    
    daily_override_capacity = daily_override.get("capacity") if daily_override else None
    
    # Check if there's model-specific capacity
    model_query = {"branch": branch, "month": month_str, "day": day}
    if model_id:
        model_query["model_id"] = model_id
    
    model_capacities = await db.branch_model_capacity.find(
        model_query,
        {"_id": 0}
    ).to_list(100)
    
    # Get already allocated for this date
    schedules_on_date = await db.production_schedules.find(
        {
            "target_date": {
                "$gte": datetime(date_obj.year, date_obj.month, date_obj.day, tzinfo=timezone.utc),
                "$lt": datetime(date_obj.year, date_obj.month, date_obj.day, 23, 59, 59, tzinfo=timezone.utc)
            },
            "status": {"$ne": "CANCELLED"}
        },
        {"_id": 0, "target_quantity": 1, "sku_id": 1, "branch": 1}
    ).to_list(1000)
    
    # Filter schedules by branch if branch field exists
    branch_schedules = [s for s in schedules_on_date if s.get("branch") == branch]
    total_allocated = sum(s.get("target_quantity", 0) for s in branch_schedules)
    
    # Determine effective capacity (priority: model > daily override > base)
    if model_capacities:
        # Use model-specific capacity
        total_model_capacity = sum(m.get("capacity_qty", 0) for m in model_capacities)
        effective_capacity = total_model_capacity
        capacity_type = "model_specific"
    elif daily_override_capacity is not None:
        # Use daily override capacity
        effective_capacity = daily_override_capacity
        capacity_type = "daily_override"
    else:
        # Use base capacity (default for all days)
        effective_capacity = base_capacity
        capacity_type = "base"
    
    available = max(0, effective_capacity - total_allocated)
    
    return {
        "branch": branch,
        "date": date_str,
        "base_capacity": base_capacity,
        "daily_override_capacity": daily_override_capacity,
        "effective_capacity": effective_capacity,
        "model_capacities": model_capacities,
        "allocated": total_allocated,
        "available": available,
        "capacity_type": capacity_type
    }


@router.get("/skus/{sku_id}/assigned-branches")
async def get_sku_assigned_branches(sku_id: str):
    """Get branches where a SKU is assigned/subscribed"""
    # Check if SKU exists
    sku = await sku_service.get_sku_by_sku_id(sku_id)
    if not sku:
        raise HTTPException(status_code=404, detail=f"SKU {sku_id} not found")
    
    # Get SKU-branch assignments
    assignments = await db.sku_branch_assignments.find(
        {"sku_id": sku_id, "is_active": True},
        {"_id": 0, "branch": 1}
    ).to_list(100)
    
    assigned_branches = [a["branch"] for a in assignments]
    
    # If no specific assignments, return all branches with capacity (for flexibility)
    if not assigned_branches:
        all_branches = await db.branches.find(
            {"is_active": True},
            {"_id": 0, "name": 1, "capacity_units_per_day": 1}
        ).to_list(100)
        return {
            "sku_id": sku_id,
            "assignment_type": "all",
            "branches": [b["name"] for b in all_branches if b.get("capacity_units_per_day", 0) > 0]
        }
    
    return {
        "sku_id": sku_id,
        "assignment_type": "specific",
        "branches": assigned_branches
    }


@router.get("/branches/model-capacity/template")
async def download_model_capacity_template():
    """Download Excel template for model capacity upload"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model Capacity"
    
    # Headers - use Branch ID instead of Branch name
    headers = ["Branch ID", "Month (YYYY-MM)", "Day", "Model Name", "Capacity Qty"]
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # Get branches and models for reference - include branch_id
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    models = await db.models.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    
    # Generate branch_id if missing (BR_001, BR_002, etc.)
    for idx, b in enumerate(branches, 1):
        if not b.get("branch_id"):
            b["branch_id"] = f"BR_{idx:03d}"
    
    # Add sample rows - use branch_id instead of name
    sample_branch_id = branches[0].get("branch_id", "BR_001") if branches else "BR_001"
    sample_model = models[0]["name"] if models else "Model A"
    sample_month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    ws.cell(row=2, column=1, value=sample_branch_id)
    ws.cell(row=2, column=2, value=sample_month)
    ws.cell(row=2, column=3, value=1)
    ws.cell(row=2, column=4, value=sample_model)
    ws.cell(row=2, column=5, value=100)
    
    # Add reference sheets
    # Branches reference - show both Branch ID and Name
    ws_branches = wb.create_sheet("Branches Reference")
    ws_branches.cell(row=1, column=1, value="Branch ID")
    ws_branches.cell(row=1, column=2, value="Branch Name")
    ws_branches["A1"].font = Font(bold=True)
    ws_branches["B1"].font = Font(bold=True)
    for i, b in enumerate(branches, 2):
        ws_branches.cell(row=i, column=1, value=b.get("branch_id", ""))
        ws_branches.cell(row=i, column=2, value=b.get("name", ""))
    
    ws_branches.column_dimensions["A"].width = 12
    ws_branches.column_dimensions["B"].width = 25
    
    # Models reference
    ws_models = wb.create_sheet("Models Reference")
    ws_models.cell(row=1, column=1, value="Model ID")
    ws_models.cell(row=1, column=2, value="Model Name")
    ws_models["A1"].font = Font(bold=True)
    ws_models["B1"].font = Font(bold=True)
    for i, m in enumerate(models, 2):
        ws_models.cell(row=i, column=1, value=m["id"])
        ws_models.cell(row=i, column=2, value=m["name"])
    ws_models.column_dimensions["A"].width = 40
    ws_models.column_dimensions["B"].width = 30
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=model_capacity_template.xlsx"}
    )


@router.post("/branches/model-capacity/upload-excel")
async def upload_model_capacity_excel(file: UploadFile = File(...)):
    """Upload model capacity from Excel file (Branch, Month, Day, Model Name, Capacity)"""
    if not pd:
        raise HTTPException(status_code=500, detail="pandas not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    
    # Normalize column names
    df.columns = [str(c).strip().lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_") for c in df.columns]
    
    # Expected columns (flexible matching) - support branch_id
    required_cols = {
        "branch": ["branch_id", "branch", "branch_name", "unit"],
        "month": ["month", "month_yyyy_mm", "yyyy_mm"],
        "day": ["day", "day_of_month"],
        "model": ["model", "model_name", "model_id"],
        "capacity": ["capacity", "capacity_qty", "qty", "quantity"]
    }
    
    # Map columns
    col_map = {}
    for key, options in required_cols.items():
        found = False
        for opt in options:
            if opt in df.columns:
                col_map[key] = opt
                found = True
                break
        if not found:
            raise HTTPException(status_code=400, detail=f"Missing required column: {key}. Expected one of: {options}")
    
    # Validate and build model name to ID map
    models = await db.models.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    model_name_to_id = {m["name"].lower(): m["id"] for m in models}
    model_id_set = {m["id"] for m in models}
    
    # Validate branches - support both branch_id and branch_name
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    branch_id_to_name = {b.get("branch_id", ""): b["name"] for b in branches if b.get("branch_id")}
    valid_branch_names = {b["name"] for b in branches}
    
    # Process rows
    results = {"inserted": 0, "updated": 0, "errors": []}
    
    for idx, row in df.iterrows():
        try:
            branch_value = str(row[col_map["branch"]]).strip()
            month = str(row[col_map["month"]]).strip()
            day = int(row[col_map["day"]])
            model_value = str(row[col_map["model"]]).strip()
            capacity = int(row[col_map["capacity"]])
            
            # Resolve branch - support both branch_id and branch_name
            branch_name = branch_id_to_name.get(branch_value) or (branch_value if branch_value in valid_branch_names else None)
            if not branch_name:
                results["errors"].append(f"Row {idx+2}: Invalid branch '{branch_value}'")
                continue
            
            # Validate month format (YYYY-MM)
            if len(month) < 7 or month[4] != '-':
                # Try to parse as date and extract month
                try:
                    parsed_date = pd.to_datetime(month)
                    month = parsed_date.strftime("%Y-%m")
                except:
                    results["errors"].append(f"Row {idx+2}: Invalid month format '{month}'. Use YYYY-MM")
                    continue
            
            # Validate day
            if day < 1 or day > 31:
                results["errors"].append(f"Row {idx+2}: Invalid day {day}")
                continue
            
            # Resolve model ID
            if model_value in model_id_set:
                model_id = model_value
            elif model_value.lower() in model_name_to_id:
                model_id = model_name_to_id[model_value.lower()]
            else:
                results["errors"].append(f"Row {idx+2}: Model '{model_value}' not found")
                continue
            
            # Upsert capacity record
            record = {
                "branch": branch_name,
                "month": month,
                "day": day,
                "model_id": model_id,
                "capacity_qty": capacity,
                "updated_at": datetime.now(timezone.utc)
            }
            
            result = await db.branch_model_capacity.update_one(
                {
                    "branch": branch_name,
                    "month": month,
                    "day": day,
                    "model_id": model_id
                },
                {"$set": record, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc)}},
                upsert=True
            )
            
            if result.upserted_id:
                results["inserted"] += 1
            else:
                results["updated"] += 1
                
        except Exception as e:
            results["errors"].append(f"Row {idx+2}: {str(e)}")
    
    return {
        "message": f"Excel upload complete",
        "inserted": results["inserted"],
        "updated": results["updated"],
        "total_processed": results["inserted"] + results["updated"],
        "errors": results["errors"][:20],  # Limit error messages
        "total_errors": len(results["errors"])
    }


# ============ END BRANCH MODEL CAPACITY ============


@router.post("/cpc/schedule-from-forecast")
async def create_schedule_from_forecast(data: ScheduleFromForecastRequest):
    """Create a production schedule directly from a demand forecast.
    Branch is REQUIRED - no production schedule can be created without a branch."""
    
    # VALIDATE BRANCH IS PROVIDED
    if not data.branch or not data.branch.strip():
        raise HTTPException(status_code=400, detail="Branch is required. All production schedules must be assigned to a branch.")
    
    # Get the forecast
    forecast = await db.forecasts.find_one({"id": data.forecast_id}, {"_id": 0})
    if not forecast:
        raise HTTPException(status_code=404, detail="Forecast not found")
    
    if forecast.get("status") not in ["CONFIRMED", "CONVERTED"]:
        raise HTTPException(status_code=400, detail="Can only schedule from confirmed forecasts")
    
    # Get SKU from forecast
    sku_id = forecast.get("sku_id")
    if not sku_id:
        raise HTTPException(status_code=400, detail="Forecast has no SKU linked")
    
    # Verify SKU exists
    sku = await sku_service.get_sku_by_sku_id(sku_id)
    if not sku:
        raise HTTPException(status_code=404, detail=f"SKU {sku_id} not found")
    
    # VALIDATE BRANCH EXISTS
    branch_cap = await db.branches.find_one({"name": data.branch, "is_active": True}, {"_id": 0})
    if not branch_cap:
        raise HTTPException(status_code=404, detail=f"Branch '{data.branch}' not found")
    
    # Check if SKU is assigned to this branch
    sku_assignments = await db.sku_branch_assignments.find(
        {"sku_id": sku_id, "is_active": True},
        {"_id": 0, "branch": 1}
    ).to_list(100)
    
    assigned_branches = [a["branch"] for a in sku_assignments]
    
    # If there are specific assignments, verify the branch is in the list
    if assigned_branches and data.branch not in assigned_branches:
        raise HTTPException(
            status_code=400, 
            detail=f"SKU {sku_id} is not assigned to branch '{data.branch}'. Assigned branches: {', '.join(assigned_branches)}"
        )
    
    # Check branch capacity for the target date
    target_date_obj = data.target_date
    date_str = target_date_obj.strftime("%Y-%m-%d")
    month_str = target_date_obj.strftime("%Y-%m")
    day = target_date_obj.day
    
    # Get model-specific capacity if available
    model_id = sku.get("model_id")
    model_capacity = None
    if model_id:
        model_capacity = await db.branch_model_capacity.find_one(
            {"branch": data.branch, "month": month_str, "day": day, "model_id": model_id},
            {"_id": 0}
        )
    
    # Calculate already allocated for this branch on target date
    day_start = datetime(target_date_obj.year, target_date_obj.month, target_date_obj.day, tzinfo=timezone.utc)
    day_end = datetime(target_date_obj.year, target_date_obj.month, target_date_obj.day, 23, 59, 59, tzinfo=timezone.utc)
    
    existing_on_date = await db.production_schedules.find(
        {
            "branch": data.branch,
            "target_date": {"$gte": day_start, "$lte": day_end},
            "status": {"$ne": "CANCELLED"}
        },
        {"_id": 0, "target_quantity": 1}
    ).to_list(1000)
    
    already_allocated = sum(s.get("target_quantity", 0) for s in existing_on_date)
    
    # Determine capacity limit - check daily override first
    daily_override = await db.branch_daily_capacity.find_one(
        {"branch": data.branch, "date": date_str},
        {"_id": 0, "capacity": 1}
    )
    
    if daily_override:
        capacity_limit = daily_override.get("capacity", 0)
        capacity_type = "daily-override"
    elif model_capacity:
        capacity_limit = model_capacity.get("capacity_qty", 0)
        capacity_type = "model-specific"
    else:
        capacity_limit = branch_cap.get("capacity_units_per_day", 0)
        capacity_type = "base"
    
    available_capacity = capacity_limit - already_allocated
    
    # Check if quantity exceeds available capacity
    if data.quantity > available_capacity:
        raise HTTPException(
            status_code=400,
            detail=f"Quantity ({data.quantity}) exceeds available capacity for branch '{data.branch}' on {date_str}. "
                   f"Capacity ({capacity_type}): {capacity_limit}, Already allocated: {already_allocated}, Available: {available_capacity}"
        )
    
    # Check remaining quantity
    existing_schedules = await db.production_schedules.find(
        {"forecast_id": data.forecast_id, "status": {"$ne": "CANCELLED"}},
        {"_id": 0, "target_quantity": 1}
    ).to_list(100)
    
    already_scheduled = sum(s.get("target_quantity", 0) for s in existing_schedules)
    forecast_qty = forecast.get("quantity", 0)
    remaining = forecast_qty - already_scheduled
    
    if data.quantity > remaining:
        raise HTTPException(
            status_code=400, 
            detail=f"Quantity exceeds remaining. Forecast: {forecast_qty}, Already scheduled: {already_scheduled}, Remaining: {remaining}"
        )
    
    # Create production schedule - status SCHEDULED (not DRAFT since branch is assigned)
    count = await db.production_schedules.count_documents({})
    schedule_code = f"PS_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    schedule = {
        "id": str(uuid.uuid4()),
        "schedule_code": schedule_code,
        "forecast_id": data.forecast_id,
        "dispatch_lot_id": None,
        "branch": data.branch,  # Branch is REQUIRED
        "sku_id": sku_id,
        "sku_description": sku.get("description", ""),
        "target_quantity": data.quantity,
        "allocated_quantity": 0,
        "completed_quantity": 0,
        "target_date": data.target_date,
        "priority": data.priority or forecast.get("priority", "MEDIUM"),
        "status": "SCHEDULED",  # SCHEDULED since branch is assigned
        "notes": data.notes or f"Scheduled from forecast {forecast.get('forecast_code', '')}",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.production_schedules.insert_one(schedule)
    
    # Update forecast status if fully scheduled
    new_scheduled_total = already_scheduled + data.quantity
    if new_scheduled_total >= forecast_qty:
        await db.forecasts.update_one(
            {"id": data.forecast_id},
            {"$set": {"status": "CONVERTED"}}
        )
    
    del schedule["_id"]
    return {
        "message": "Production schedule created from forecast",
        "schedule": serialize_doc(schedule),
        "remaining_forecast_qty": remaining - data.quantity
    }
