"""
Inventory Management Routes
- View and bulk import RM and Finished Goods inventory
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from database import db
from datetime import datetime, timezone
from typing import Optional
import io
import uuid
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

router = APIRouter(prefix="/inventory", tags=["Inventory"])


# ============ RM INVENTORY ============
# NOTE: Uses branch_rm_inventory collection (same as RM Stock View, Production, etc.)

@router.get("/rm")
async def get_rm_inventory(
    branch: Optional[str] = None,
    branch_id: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """
    Get RM inventory with filters.
    Uses branch_rm_inventory collection - same source as RM Stock View and Production.
    
    Supports both branch name (e.g., "Unit 1 Vedica") and branch_id (e.g., "BR_001").
    """
    query = {}
    
    # Resolve branch filter - support both branch name and branch_id
    branch_filter = branch or branch_id
    if branch_filter:
        # Check if it's a branch_id (BR_xxx format)
        if branch_filter.startswith("BR_"):
            # Look up branch name from branch_id
            branch_doc = await db.branches.find_one({"branch_id": branch_filter}, {"_id": 0, "name": 1})
            if branch_doc:
                query["branch"] = branch_doc["name"]
            else:
                query["branch"] = branch_filter
        else:
            query["branch"] = branch_filter
    
    # Get inventory records from branch_rm_inventory
    cursor = db.branch_rm_inventory.find(query, {"_id": 0}).skip(skip).limit(limit).sort("rm_id", 1)
    items = await cursor.to_list(limit)
    
    # Get total count
    total = await db.branch_rm_inventory.count_documents(query)
    
    # Enrich with RM details
    for item in items:
        rm = await db.raw_materials.find_one({"rm_id": item["rm_id"]}, {"_id": 0, "name": 1, "category": 1, "default_unit": 1})
        if rm:
            item["rm_name"] = rm.get("name", "")
            item["category"] = rm.get("category", "")
            item["unit"] = rm.get("default_unit", "PCS")
        
        # Add quantity alias for backward compatibility
        item["quantity"] = item.get("current_stock", 0)
        
        # Add branch_id for display
        if item.get("branch"):
            branch_doc = await db.branches.find_one({"name": item["branch"]}, {"_id": 0, "branch_id": 1})
            if branch_doc:
                item["branch_id"] = branch_doc.get("branch_id", "")
    
    # Apply category filter (post-fetch since it requires RM lookup)
    if category:
        items = [item for item in items if item.get("category") == category]
        total = len(items)
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        items = [item for item in items if 
                 search_lower in item.get("rm_id", "").lower() or 
                 search_lower in item.get("rm_name", "").lower()]
        total = len(items)
    
    return {"items": items, "total": total}


@router.get("/rm/template")
async def download_rm_inventory_template():
    """
    Download Excel template for RM inventory import.
    Uses branch names (same as branch_rm_inventory collection).
    """
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RM Inventory"
    
    # Headers - use BRANCH (name) to match branch_rm_inventory
    headers = ["RM_ID", "BRANCH", "QUANTITY"]
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Get sample data
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    rms = await db.raw_materials.find({}, {"_id": 0, "rm_id": 1}).to_list(10)
    
    # Add sample rows using branch NAME
    sample_branch = branches[0].get("name", "Unit 1 Vedica") if branches else "Unit 1 Vedica"
    for i, rm in enumerate(rms[:5], 2):
        ws.cell(row=i, column=1, value=rm.get("rm_id", ""))
        ws.cell(row=i, column=2, value=sample_branch)
        ws.cell(row=i, column=3, value=100)
    
    # Branches Reference sheet - show both name and ID
    ws_branches = wb.create_sheet("Branches Reference")
    ws_branches.cell(row=1, column=1, value="Branch Name (use this)")
    ws_branches.cell(row=1, column=2, value="Branch ID")
    ws_branches["A1"].font = Font(bold=True)
    ws_branches["B1"].font = Font(bold=True)
    for i, b in enumerate(branches, 2):
        ws_branches.cell(row=i, column=1, value=b.get("name", ""))
        ws_branches.cell(row=i, column=2, value=b.get("branch_id", ""))
    ws_branches.column_dimensions["A"].width = 25
    ws_branches.column_dimensions["B"].width = 12
    
    # RM Reference sheet
    ws_rm = wb.create_sheet("RM Reference")
    ws_rm.cell(row=1, column=1, value="RM ID")
    ws_rm.cell(row=1, column=2, value="RM Name")
    ws_rm.cell(row=1, column=3, value="Category")
    ws_rm["A1"].font = Font(bold=True)
    ws_rm["B1"].font = Font(bold=True)
    ws_rm["C1"].font = Font(bold=True)
    
    all_rms = await db.raw_materials.find({}, {"_id": 0, "rm_id": 1, "name": 1, "category": 1}).to_list(1000)
    for i, rm in enumerate(all_rms, 2):
        ws_rm.cell(row=i, column=1, value=rm.get("rm_id", ""))
        ws_rm.cell(row=i, column=2, value=rm.get("name", ""))
        ws_rm.cell(row=i, column=3, value=rm.get("category", ""))
    ws_rm.column_dimensions["A"].width = 15
    ws_rm.column_dimensions["B"].width = 35
    ws_rm.column_dimensions["C"].width = 15
    
    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "RM Inventory Import Instructions",
        "",
        "1. Use the 'RM Inventory' sheet for your data",
        "2. Required columns: RM_ID, BRANCH, QUANTITY",
        "3. BRANCH: Use branch NAME (e.g., 'Unit 1 Vedica'), not branch ID",
        "4. See 'Branches Reference' sheet for valid branch names",
        "5. See 'RM Reference' sheet for valid RM IDs",
        "",
        "Import Modes:",
        "- ADD: Adds uploaded quantity to existing stock",
        "- REPLACE: Sets stock to uploaded quantity (overwrites)",
        "",
        "Note: This updates the same inventory used by RM Stock View and Production."
    ]
    for i, line in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=line)
    ws_inst.column_dimensions["A"].width = 70
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rm_inventory_template.xlsx"}
    )


@router.post("/rm/bulk-import")
async def bulk_import_rm_inventory(
    file: UploadFile = File(...),
    mode: str = Query("add", description="Import mode: 'add' to add to existing, 'replace' to replace stock")
):
    """
    Bulk import RM inventory from Excel.
    
    IMPORTANT: This writes to branch_rm_inventory collection - same source used by:
    - RM Stock View page
    - Production shortage calculations
    - Branch Operations
    
    Mode:
    - 'add': Add uploaded quantities to existing stock
    - 'replace': Replace existing stock with uploaded quantities
    """
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    ws = wb.active
    
    results = {
        "processed": 0,
        "added": 0,
        "updated": 0,
        "errors": [],
        "mode": mode,
        "collection": "branch_rm_inventory",
        "success": True,
        "branch_mappings_used": {},  # Track which branches were mapped
        "branches_imported_to": []  # List of branches data was imported to
    }
    
    # Build branch lookup (supports both branch_id and branch name)
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1, "code": 1}).to_list(100)
    branch_lookup = {}  # Maps various inputs to branch NAME (what branch_rm_inventory uses)
    
    for b in branches:
        branch_name = b.get("name", "")
        bid = b.get("branch_id", "")
        code = b.get("code", "")
        
        if not branch_name:
            continue
            
        # Map branch_id to branch name (e.g., BR_001 -> Unit 2 Trikes)
        if bid:
            branch_lookup[bid] = branch_name
            branch_lookup[bid.lower()] = branch_name
            branch_lookup[bid.upper()] = branch_name
        
        # Map branch code to branch name (e.g., UNIT_2_TRIKES -> Unit 2 Trikes)
        if code:
            branch_lookup[code] = branch_name
            branch_lookup[code.lower()] = branch_name
            branch_lookup[code.upper()] = branch_name
        
        # Map branch name to itself (case-insensitive)
        branch_lookup[branch_name] = branch_name
        branch_lookup[branch_name.lower()] = branch_name
        branch_lookup[branch_name.upper()] = branch_name
    
    # Log available branches for debugging
    results["available_branches"] = [{"id": b.get("branch_id"), "name": b.get("name")} for b in branches]
    
    # Detect column headers
    headers = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]
    rm_id_col = None
    branch_col = None
    qty_col = None
    
    for idx, h in enumerate(headers):
        if h in ["RM_ID", "RM ID", "RMID"]:
            rm_id_col = idx
        elif h in ["BRANCH", "BRANCH_ID", "BRANCH ID", "BRANCHID"]:
            branch_col = idx
        elif h in ["QUANTITY", "QTY", "STOCK", "CURRENT_STOCK"]:
            qty_col = idx
    
    if rm_id_col is None or branch_col is None or qty_col is None:
        raise HTTPException(
            status_code=400, 
            detail=f"Missing required columns. Found: {headers}. Required: RM_ID, BRANCH, QUANTITY"
        )
    
    # Process rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[rm_id_col]:
            continue
        
        try:
            rm_id = str(row[rm_id_col]).strip()
            branch_value = str(row[branch_col]).strip() if row[branch_col] else None
            quantity = float(row[qty_col]) if row[qty_col] else 0
            
            if not branch_value:
                results["errors"].append(f"Row {row_idx}: BRANCH is required")
                continue
            
            # Resolve branch to NAME (what branch_rm_inventory uses)
            branch_name = branch_lookup.get(branch_value) or branch_lookup.get(branch_value.lower()) or branch_lookup.get(branch_value.upper())
            if not branch_name:
                # List available branch IDs for user reference
                available_ids = [b.get("branch_id") for b in branches if b.get("branch_id")]
                results["errors"].append(f"Row {row_idx}: Invalid branch '{branch_value}'. Available: {available_ids}")
                continue
            
            # Track the mapping used
            if branch_value != branch_name:
                results["branch_mappings_used"][branch_value] = branch_name
            
            if branch_name not in results["branches_imported_to"]:
                results["branches_imported_to"].append(branch_name)
            
            # Verify RM exists
            rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0, "rm_id": 1})
            if not rm:
                # Case-insensitive search
                rm = await db.raw_materials.find_one(
                    {"rm_id": {"$regex": f"^{re.escape(rm_id)}$", "$options": "i"}},
                    {"_id": 0, "rm_id": 1}
                )
            if not rm:
                results["errors"].append(f"Row {row_idx}: RM '{rm_id}' not found")
                continue
            
            actual_rm_id = rm.get("rm_id", rm_id)
            
            # Check existing inventory in branch_rm_inventory
            existing = await db.branch_rm_inventory.find_one({
                "rm_id": actual_rm_id,
                "branch": branch_name
            })
            
            if mode == "replace":
                new_quantity = quantity
            else:  # add
                new_quantity = (existing.get("current_stock", 0) if existing else 0) + quantity
            
            if existing:
                await db.branch_rm_inventory.update_one(
                    {"rm_id": actual_rm_id, "branch": branch_name},
                    {"$set": {
                        "current_stock": new_quantity,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                results["updated"] += 1
            else:
                await db.branch_rm_inventory.insert_one({
                    "id": str(uuid.uuid4()),
                    "rm_id": actual_rm_id,
                    "branch": branch_name,
                    "current_stock": new_quantity,
                    "is_active": True,
                    "activated_at": datetime.now(timezone.utc).isoformat()
                })
                results["added"] += 1
            
            results["processed"] += 1
            
        except Exception as e:
            results["errors"].append(f"Row {row_idx}: {str(e)}")
    
    if results["errors"] and results["processed"] == 0:
        results["success"] = False
    
    return results


# ============ FINISHED GOODS INVENTORY ============

@router.get("/fg")
async def get_fg_inventory(
    branch_id: Optional[str] = None,
    model_id: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get Finished Goods inventory with filters"""
    query = {}
    
    if branch_id:
        query["branch_id"] = branch_id
    if model_id:
        query["model_id"] = model_id
    if search:
        query["$or"] = [
            {"buyer_sku_id": {"$regex": search, "$options": "i"}},
            {"sku_name": {"$regex": search, "$options": "i"}}
        ]
    
    # Get inventory records
    cursor = db.fg_inventory.find(query, {"_id": 0}).skip(skip).limit(limit).sort("buyer_sku_id", 1)
    items = await cursor.to_list(limit)
    
    # Get total count
    total = await db.fg_inventory.count_documents(query)
    
    # Enrich with SKU details if not present
    for item in items:
        if not item.get("sku_name"):
            sku = await db.buyer_skus.find_one({"buyer_sku_id": item["buyer_sku_id"]}, {"_id": 0, "name": 1, "model_id": 1})
            if sku:
                item["sku_name"] = sku.get("name", "")
                item["model_id"] = sku.get("model_id", "")
    
    return {"items": items, "total": total}


@router.get("/fg/template")
async def download_fg_inventory_template():
    """Download Excel template for FG inventory import"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FG Inventory"
    
    # Headers
    headers = ["BUYER_SKU_ID", "BRANCH_ID", "QUANTITY"]
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Get sample data
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    skus = await db.buyer_skus.find({"status": "ACTIVE"}, {"_id": 0, "buyer_sku_id": 1}).to_list(10)
    
    # Generate branch_id if missing
    for idx, b in enumerate(branches, 1):
        if not b.get("branch_id"):
            b["branch_id"] = f"BR_{idx:03d}"
    
    # Add sample rows
    sample_branch = branches[0].get("branch_id", "BR_001") if branches else "BR_001"
    for i, sku in enumerate(skus[:5], 2):
        ws.cell(row=i, column=1, value=sku.get("buyer_sku_id", ""))
        ws.cell(row=i, column=2, value=sample_branch)
        ws.cell(row=i, column=3, value=50)
    
    # Branches Reference sheet
    ws_branches = wb.create_sheet("Branches Reference")
    ws_branches.cell(row=1, column=1, value="Branch ID")
    ws_branches.cell(row=1, column=2, value="Branch Name")
    ws_branches["A1"].font = Font(bold=True)
    ws_branches["B1"].font = Font(bold=True)
    for i, b in enumerate(branches, 2):
        ws_branches.cell(row=i, column=1, value=b.get("branch_id", ""))
        ws_branches.cell(row=i, column=2, value=b.get("name", ""))
    ws_branches.column_dimensions["A"].width = 12
    ws_branches.column_dimensions["B"].width = 25
    
    # SKU Reference sheet
    ws_sku = wb.create_sheet("Buyer SKU Reference")
    ws_sku.cell(row=1, column=1, value="Buyer SKU ID")
    ws_sku.cell(row=1, column=2, value="SKU Name")
    ws_sku.cell(row=1, column=3, value="Brand")
    ws_sku["A1"].font = Font(bold=True)
    ws_sku["B1"].font = Font(bold=True)
    ws_sku["C1"].font = Font(bold=True)
    
    all_skus = await db.buyer_skus.find(
        {"status": "ACTIVE"}, 
        {"_id": 0, "buyer_sku_id": 1, "name": 1, "brand_code": 1}
    ).to_list(2000)
    for i, sku in enumerate(all_skus, 2):
        ws_sku.cell(row=i, column=1, value=sku.get("buyer_sku_id", ""))
        ws_sku.cell(row=i, column=2, value=sku.get("name", ""))
        ws_sku.cell(row=i, column=3, value=sku.get("brand_code", ""))
    ws_sku.column_dimensions["A"].width = 20
    ws_sku.column_dimensions["B"].width = 40
    ws_sku.column_dimensions["C"].width = 15
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fg_inventory_template.xlsx"}
    )


@router.post("/fg/bulk-import")
async def bulk_import_fg_inventory(
    file: UploadFile = File(...),
    mode: str = Query("add", description="Import mode: 'add' to add to existing, 'replace' to replace stock")
):
    """
    Bulk import Finished Goods inventory from Excel.
    
    Mode:
    - 'add': Add uploaded quantities to existing stock
    - 'replace': Replace existing stock with uploaded quantities
    """
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    ws = wb.active
    
    results = {
        "processed": 0,
        "added": 0,
        "updated": 0,
        "errors": [],
        "mode": mode,
        "success": True
    }
    
    # Build branch lookup
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    branch_lookup = {}
    for idx, b in enumerate(branches, 1):
        bid = b.get("branch_id") or f"BR_{idx:03d}"
        branch_lookup[bid] = bid
        branch_lookup[bid.lower()] = bid
        branch_lookup[b["name"]] = bid
        branch_lookup[b["name"].lower()] = bid
    
    # Process rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        try:
            buyer_sku_id = str(row[0]).strip()
            branch_value = str(row[1]).strip() if row[1] else None
            quantity = float(row[2]) if row[2] else 0
            
            if not branch_value:
                results["errors"].append(f"Row {row_idx}: BRANCH_ID is required")
                continue
            
            # Resolve branch
            branch_id = branch_lookup.get(branch_value) or branch_lookup.get(branch_value.lower())
            if not branch_id:
                results["errors"].append(f"Row {row_idx}: Invalid branch '{branch_value}'")
                continue
            
            # Verify SKU exists
            sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0, "buyer_sku_id": 1, "name": 1, "model_id": 1})
            if not sku:
                # Case-insensitive search
                sku = await db.buyer_skus.find_one(
                    {"buyer_sku_id": {"$regex": f"^{re.escape(buyer_sku_id)}$", "$options": "i"}},
                    {"_id": 0, "buyer_sku_id": 1, "name": 1, "model_id": 1}
                )
            if not sku:
                results["errors"].append(f"Row {row_idx}: Buyer SKU '{buyer_sku_id}' not found")
                continue
            
            actual_sku_id = sku.get("buyer_sku_id", buyer_sku_id)
            
            # Check existing inventory
            existing = await db.fg_inventory.find_one({
                "buyer_sku_id": actual_sku_id,
                "branch_id": branch_id
            })
            
            if mode == "replace":
                new_quantity = quantity
            else:  # add
                new_quantity = (existing.get("quantity", 0) if existing else 0) + quantity
            
            if existing:
                await db.fg_inventory.update_one(
                    {"buyer_sku_id": actual_sku_id, "branch_id": branch_id},
                    {"$set": {
                        "quantity": new_quantity,
                        "sku_name": sku.get("name", ""),
                        "model_id": sku.get("model_id", ""),
                        "updated_at": datetime.now(timezone.utc)
                    }}
                )
                results["updated"] += 1
            else:
                await db.fg_inventory.insert_one({
                    "id": str(uuid.uuid4()),
                    "buyer_sku_id": actual_sku_id,
                    "sku_name": sku.get("name", ""),
                    "model_id": sku.get("model_id", ""),
                    "branch_id": branch_id,
                    "quantity": new_quantity,
                    "created_at": datetime.now(timezone.utc)
                })
                results["added"] += 1
            
            results["processed"] += 1
            
        except Exception as e:
            results["errors"].append(f"Row {row_idx}: {str(e)}")
    
    if results["errors"] and results["processed"] == 0:
        results["success"] = False
    
    return results


# ============ SUMMARY STATS ============

@router.get("/summary")
async def get_inventory_summary():
    """Get inventory summary stats - reads from branch_rm_inventory (unified source)"""
    # Use branch_rm_inventory - the unified collection for all RM inventory
    rm_total = await db.branch_rm_inventory.count_documents({})
    fg_total = await db.fg_inventory.count_documents({})
    
    # Get unique RMs and SKUs in inventory
    rm_unique = len(await db.branch_rm_inventory.distinct("rm_id"))
    fg_unique = len(await db.fg_inventory.distinct("buyer_sku_id"))
    
    # Get branch count
    branches_with_rm = len(await db.branch_rm_inventory.distinct("branch"))
    branches_with_fg = len(await db.fg_inventory.distinct("branch_id"))
    
    return {
        "rm": {
            "total_records": rm_total,
            "unique_items": rm_unique,
            "branches": branches_with_rm
        },
        "fg": {
            "total_records": fg_total,
            "unique_items": fg_unique,
            "branches": branches_with_fg
        }
    }


# ============ FILTERS DATA ============

@router.get("/filters")
async def get_inventory_filters():
    """Get filter options for inventory pages"""
    # Branches
    branches = await db.branches.find({"is_active": True}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    for idx, b in enumerate(branches, 1):
        if not b.get("branch_id"):
            b["branch_id"] = f"BR_{idx:03d}"
    
    # RM Categories
    categories = await db.raw_materials.distinct("category")
    
    # Models
    models = await db.models.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    
    return {
        "branches": branches,
        "categories": [c for c in categories if c],
        "models": models
    }


@router.get("/fg/debug")
async def debug_fg_inventory():
    """Debug endpoint to check FG inventory state"""
    total = await db.fg_inventory.count_documents({})
    
    # Get all records
    all_records = await db.fg_inventory.find({}, {"_id": 0}).to_list(100)
    
    # Check field usage
    with_buyer_sku_id = await db.fg_inventory.count_documents({"buyer_sku_id": {"$exists": True}})
    with_sku_id = await db.fg_inventory.count_documents({"sku_id": {"$exists": True}})
    with_branch_id = await db.fg_inventory.count_documents({"branch_id": {"$exists": True}})
    with_branch = await db.fg_inventory.count_documents({"branch": {"$exists": True}})
    
    return {
        "total_records": total,
        "field_usage": {
            "with_buyer_sku_id": with_buyer_sku_id,
            "with_sku_id": with_sku_id,
            "with_branch_id": with_branch_id,
            "with_branch": with_branch
        },
        "records": all_records[:20]  # First 20 records
    }

