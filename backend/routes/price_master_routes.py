"""
Price Master Routes - Customer-specific pricing for Buyer SKUs

Maintained by: Demand Team
Used by: Finance Team (auto-populate invoice rates)

Structure:
- Customer ID (Buyer) + Buyer SKU ID = Unique Price
- Supports effective dates for price versioning
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query
from pydantic import BaseModel
from datetime import datetime, timezone
from typing import Optional, List
import uuid
import io
import openpyxl

from database import db
from services.utils import get_current_user, serialize_doc

router = APIRouter(prefix="/price-master", tags=["Price Master"])


# ============ Pydantic Models ============

class PriceEntry(BaseModel):
    """Price entry for a Customer + Buyer SKU combination"""
    customer_id: str
    buyer_sku_id: str
    unit_price: float
    currency: str = "INR"
    effective_from: Optional[str] = None  # ISO date string
    effective_to: Optional[str] = None  # ISO date string, null = active
    notes: Optional[str] = None


class PriceEntryCreate(BaseModel):
    """Create price entry request"""
    customer_id: str
    buyer_sku_id: str
    unit_price: float
    currency: str = "INR"
    effective_from: Optional[str] = None
    notes: Optional[str] = None


class PriceEntryUpdate(BaseModel):
    """Update price entry request"""
    unit_price: Optional[float] = None
    currency: Optional[str] = None
    effective_to: Optional[str] = None
    notes: Optional[str] = None


# ============ Helper Functions ============

async def get_customer_name(customer_id: str) -> str:
    """Get customer name from buyers collection"""
    buyer = await db.buyers.find_one({"customer_code": customer_id})
    if not buyer:
        buyer = await db.buyers.find_one({"id": customer_id})
    return buyer.get("name", "Unknown") if buyer else "Unknown"


async def get_sku_name(buyer_sku_id: str) -> str:
    """Get SKU name from buyer_skus collection"""
    sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
    return sku.get("name", buyer_sku_id) if sku else buyer_sku_id


# ============ CRUD Endpoints ============

@router.get("")
async def list_prices(
    customer_id: Optional[str] = None,
    buyer_sku_id: Optional[str] = None,
    active_only: bool = True,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user = Depends(get_current_user)
):
    """List price entries with optional filters"""
    query = {}
    
    if customer_id:
        query["customer_id"] = customer_id
    
    if buyer_sku_id:
        query["buyer_sku_id"] = buyer_sku_id
    
    if active_only:
        query["$or"] = [
            {"effective_to": None},
            {"effective_to": {"$gte": datetime.now(timezone.utc).isoformat()}}
        ]
    
    total = await db.price_master.count_documents(query)
    skip = (page - 1) * page_size
    
    cursor = db.price_master.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size)
    prices = await cursor.to_list(length=page_size)
    
    return {
        "prices": prices,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/lookup")
async def lookup_price(
    customer_id: str,
    buyer_sku_id: str,
    current_user = Depends(get_current_user)
):
    """Get active price for a specific Customer + SKU combination"""
    now = datetime.now(timezone.utc).isoformat()
    
    # Find active price
    price = await db.price_master.find_one({
        "customer_id": customer_id,
        "buyer_sku_id": buyer_sku_id,
        "$or": [
            {"effective_to": None},
            {"effective_to": {"$gte": now}}
        ],
        "$or": [
            {"effective_from": None},
            {"effective_from": {"$lte": now}}
        ]
    }, {"_id": 0})
    
    if not price:
        return {"found": False, "price": None}
    
    return {"found": True, "price": price}


@router.post("")
async def create_price(
    data: PriceEntryCreate,
    current_user = Depends(get_current_user)
):
    """Create a new price entry"""
    # Validate customer exists
    customer = await db.buyers.find_one({
        "$or": [
            {"customer_code": data.customer_id},
            {"id": data.customer_id}
        ]
    })
    if not customer:
        raise HTTPException(status_code=400, detail=f"Customer {data.customer_id} not found")
    
    # Validate SKU exists
    sku = await db.buyer_skus.find_one({"buyer_sku_id": data.buyer_sku_id})
    if not sku:
        raise HTTPException(status_code=400, detail=f"Buyer SKU {data.buyer_sku_id} not found")
    
    # Check for existing active price - deactivate it
    now = datetime.now(timezone.utc).isoformat()
    await db.price_master.update_many(
        {
            "customer_id": data.customer_id,
            "buyer_sku_id": data.buyer_sku_id,
            "effective_to": None
        },
        {"$set": {"effective_to": now}}
    )
    
    # Create new price entry
    price_id = str(uuid.uuid4())
    price_entry = {
        "id": price_id,
        "customer_id": data.customer_id,
        "customer_name": customer.get("name", ""),
        "buyer_sku_id": data.buyer_sku_id,
        "sku_name": sku.get("name", data.buyer_sku_id),
        "unit_price": data.unit_price,
        "currency": data.currency,
        "effective_from": data.effective_from or now,
        "effective_to": None,  # Active
        "notes": data.notes,
        "created_by": current_user.id,
        "created_at": now
    }
    
    await db.price_master.insert_one(price_entry)
    
    # Remove MongoDB _id before returning
    price_entry.pop("_id", None)
    
    return {"message": "Price created successfully", "id": price_id, "price": price_entry}


@router.put("/{price_id}")
async def update_price(
    price_id: str,
    data: PriceEntryUpdate,
    current_user = Depends(get_current_user)
):
    """Update a price entry"""
    existing = await db.price_master.find_one({"id": price_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Price entry not found")
    
    updates = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.id}
    
    if data.unit_price is not None:
        updates["unit_price"] = data.unit_price
    if data.currency is not None:
        updates["currency"] = data.currency
    if data.effective_to is not None:
        updates["effective_to"] = data.effective_to
    if data.notes is not None:
        updates["notes"] = data.notes
    
    await db.price_master.update_one({"id": price_id}, {"$set": updates})
    
    return {"message": "Price updated successfully"}


@router.delete("/{price_id}")
async def delete_price(
    price_id: str,
    current_user = Depends(get_current_user)
):
    """Delete (deactivate) a price entry"""
    result = await db.price_master.update_one(
        {"id": price_id},
        {"$set": {
            "effective_to": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.id
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Price entry not found")
    
    return {"message": "Price deactivated successfully"}


# ============ Bulk Operations ============

@router.post("/bulk-upload")
async def bulk_upload_prices(
    file: UploadFile = File(...),
    current_user = Depends(get_current_user)
):
    """
    Bulk upload prices from Excel.
    
    Expected columns: customer_id, buyer_sku_id, unit_price, currency (optional), notes (optional)
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    # Get headers
    headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
    
    required = ["customer_id", "buyer_sku_id", "unit_price"]
    for req in required:
        if req not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required column: {req}")
    
    created = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        row_data = dict(zip(headers, row))
        customer_id = str(row_data.get("customer_id", "")).strip()
        buyer_sku_id = str(row_data.get("buyer_sku_id", "")).strip()
        unit_price = row_data.get("unit_price")
        
        if not customer_id or not buyer_sku_id or unit_price is None:
            errors.append({"row": row_idx, "error": "Missing required fields"})
            continue
        
        try:
            unit_price = float(unit_price)
        except:
            errors.append({"row": row_idx, "error": "Invalid unit_price"})
            continue
        
        # Validate customer
        customer = await db.buyers.find_one({
            "$or": [
                {"customer_code": customer_id},
                {"id": customer_id}
            ]
        })
        if not customer:
            errors.append({"row": row_idx, "error": f"Customer {customer_id} not found"})
            continue
        
        # Validate SKU
        sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
        if not sku:
            errors.append({"row": row_idx, "error": f"Buyer SKU {buyer_sku_id} not found"})
            continue
        
        # Deactivate existing
        await db.price_master.update_many(
            {"customer_id": customer_id, "buyer_sku_id": buyer_sku_id, "effective_to": None},
            {"$set": {"effective_to": now}}
        )
        
        # Create new
        price_entry = {
            "id": str(uuid.uuid4()),
            "customer_id": customer_id,
            "customer_name": customer.get("name", ""),
            "buyer_sku_id": buyer_sku_id,
            "sku_name": sku.get("name", buyer_sku_id),
            "unit_price": unit_price,
            "currency": str(row_data.get("currency", "INR")).strip() or "INR",
            "effective_from": now,
            "effective_to": None,
            "notes": str(row_data.get("notes", "")).strip() if row_data.get("notes") else None,
            "created_by": current_user.id,
            "created_at": now
        }
        
        await db.price_master.insert_one(price_entry)
        created += 1
    
    return {
        "message": f"Bulk upload complete. {created} prices created.",
        "created": created,
        "errors": errors
    }


@router.get("/template")
async def download_template(current_user = Depends(get_current_user)):
    """Get template columns for bulk upload"""
    return {
        "columns": ["customer_id", "buyer_sku_id", "unit_price", "currency", "notes"],
        "required": ["customer_id", "buyer_sku_id", "unit_price"],
        "example": {
            "customer_id": "CUST_0001",
            "buyer_sku_id": "ERW001_TVS",
            "unit_price": 1500.00,
            "currency": "INR",
            "notes": "FY 2026-27 pricing"
        }
    }


@router.get("/by-customer/{customer_id}")
async def get_prices_by_customer(
    customer_id: str,
    current_user = Depends(get_current_user)
):
    """Get all active prices for a customer"""
    now = datetime.now(timezone.utc).isoformat()
    
    cursor = db.price_master.find({
        "customer_id": customer_id,
        "$or": [
            {"effective_to": None},
            {"effective_to": {"$gte": now}}
        ]
    }, {"_id": 0}).sort("buyer_sku_id", 1)
    
    prices = await cursor.to_list(length=1000)
    
    return {"customer_id": customer_id, "prices": prices, "count": len(prices)}
