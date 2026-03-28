"""
MRP (Material Requisition Planning) Routes

API endpoints for:
- Model-level forecasts management
- RM procurement parameters
- MRP run calculation and management
- Draft PO generation and approval
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Body, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from dateutil.relativedelta import relativedelta
import uuid
import logging
import io

from database import db
from services.auth_service import get_current_user
from models.auth import User
from models.mrp_models import (
    ModelLevelForecastCreate,
    ModelLevelForecastUpdate,
    RMProcurementParametersCreate,
    RMProcurementParametersUpdate,
)
from services.mrp_service import mrp_service
from services.mrp_weekly_service import weekly_mrp_service

# Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

router = APIRouter(prefix="/mrp", tags=["MRP"])
logger = logging.getLogger(__name__)


def serialize_doc(doc):
    """Remove MongoDB _id and convert dates"""
    if doc and "_id" in doc:
        del doc["_id"]
    return doc


# ============ Dashboard ============

@router.get("/dashboard")
async def get_mrp_dashboard(current_user: User = Depends(get_current_user)):
    """Get MRP dashboard statistics"""
    # Count MRP runs
    total_runs = await db.mrp_runs.count_documents({})
    pending_approval = await db.mrp_runs.count_documents({"status": "CALCULATED"})
    
    # Count Draft POs
    total_draft_pos = await db.mrp_draft_pos.count_documents({})
    pending_po_approval = await db.mrp_draft_pos.count_documents({"status": "DRAFT"})
    
    # Get last run
    last_run = await db.mrp_runs.find_one(
        {}, {"_id": 0, "run_date": 1},
        sort=[("created_at", -1)]
    )
    
    # Count RMs with shortage (net_requirement > 0)
    total_rm_shortage = 0
    total_order_value_pending = 0
    
    latest_run = await db.mrp_runs.find_one(
        {"status": {"$in": ["CALCULATED", "APPROVED"]}},
        {"_id": 0, "rm_requirements": 1, "total_order_value": 1},
        sort=[("created_at", -1)]
    )
    
    if latest_run:
        rm_reqs = latest_run.get("rm_requirements", [])
        total_rm_shortage = sum(1 for r in rm_reqs if r.get("net_requirement", 0) > 0)
        total_order_value_pending = latest_run.get("total_order_value", 0)
    
    # Get model forecast counts
    total_model_forecasts = await db.model_level_forecasts.count_documents({})
    
    # Get RM params count
    total_rm_params = await db.rm_procurement_parameters.count_documents({})
    
    return {
        "total_runs": total_runs,
        "pending_approval": pending_approval,
        "total_draft_pos": total_draft_pos,
        "pending_po_approval": pending_po_approval,
        "last_run_date": last_run.get("run_date") if last_run else None,
        "total_rm_shortage": total_rm_shortage,
        "total_order_value_pending": total_order_value_pending,
        "total_model_forecasts": total_model_forecasts,
        "total_rm_params": total_rm_params
    }


# ============ Model Level Forecasts ============

@router.get("/model-forecasts")
async def get_model_forecasts(
    model_id: Optional[str] = None,
    vertical_id: Optional[str] = None,
    month_year: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get model-level forecasts with optional filters"""
    query = {}
    if model_id:
        query["model_id"] = model_id
    if vertical_id:
        query["vertical_id"] = vertical_id
    if month_year:
        query["month_year"] = month_year
    
    forecasts = await db.model_level_forecasts.find(
        query, {"_id": 0}
    ).sort("month_year", 1).to_list(10000)
    
    return forecasts


@router.post("/model-forecasts")
async def create_model_forecast(
    data: ModelLevelForecastCreate,
    current_user: User = Depends(get_current_user)
):
    """Create or update a model-level forecast"""
    # Get model info
    model = await db.models.find_one({"id": data.model_id}, {"_id": 0})
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    
    # Get vertical info
    vertical = await db.verticals.find_one({"id": model.get("vertical_id")}, {"_id": 0})
    
    # Check for existing forecast
    existing = await db.model_level_forecasts.find_one({
        "model_id": data.model_id,
        "month_year": data.month_year
    })
    
    if existing:
        # Update existing
        await db.model_level_forecasts.update_one(
            {"id": existing["id"]},
            {"$set": {
                "forecast_qty": data.forecast_qty,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user.id
            }}
        )
        return {"message": "Forecast updated", "id": existing["id"]}
    
    # Create new
    forecast = {
        "id": str(uuid.uuid4()),
        "model_id": data.model_id,
        "model_code": model.get("code", ""),
        "model_name": model.get("name", ""),
        "vertical_id": model.get("vertical_id", ""),
        "vertical_code": vertical.get("code", "") if vertical else "",
        "month_year": data.month_year,
        "forecast_qty": data.forecast_qty,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    
    await db.model_level_forecasts.insert_one(forecast)
    return {"message": "Forecast created", "id": forecast["id"]}


@router.post("/model-forecasts/bulk")
async def bulk_create_model_forecasts(
    forecasts: List[dict] = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Bulk create/update model-level forecasts"""
    created = 0
    updated = 0
    errors = []
    
    for f in forecasts:
        try:
            model_id = f.get("model_id")
            month_year = f.get("month_year")
            forecast_qty = f.get("forecast_qty", 0)
            
            if not model_id or not month_year:
                errors.append({"data": f, "error": "Missing model_id or month_year"})
                continue
            
            # Get model info
            model = await db.models.find_one({"id": model_id}, {"_id": 0})
            if not model:
                # Try by code
                model = await db.models.find_one({"code": model_id}, {"_id": 0})
            
            if not model:
                errors.append({"data": f, "error": f"Model not found: {model_id}"})
                continue
            
            actual_model_id = model.get("id")
            
            # Check existing
            existing = await db.model_level_forecasts.find_one({
                "model_id": actual_model_id,
                "month_year": month_year
            })
            
            if existing:
                await db.model_level_forecasts.update_one(
                    {"id": existing["id"]},
                    {"$set": {
                        "forecast_qty": forecast_qty,
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                        "updated_by": current_user.id
                    }}
                )
                updated += 1
            else:
                vertical = await db.verticals.find_one(
                    {"id": model.get("vertical_id")}, {"_id": 0}
                )
                
                forecast = {
                    "id": str(uuid.uuid4()),
                    "model_id": actual_model_id,
                    "model_code": model.get("code", ""),
                    "model_name": model.get("name", ""),
                    "vertical_id": model.get("vertical_id", ""),
                    "vertical_code": vertical.get("code", "") if vertical else "",
                    "month_year": month_year,
                    "forecast_qty": forecast_qty,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": current_user.id
                }
                await db.model_level_forecasts.insert_one(forecast)
                created += 1
                
        except Exception as e:
            errors.append({"data": f, "error": str(e)})
    
    return {
        "created": created,
        "updated": updated,
        "errors": errors
    }


@router.delete("/model-forecasts/{forecast_id}")
async def delete_model_forecast(
    forecast_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a model-level forecast"""
    result = await db.model_level_forecasts.delete_one({"id": forecast_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Forecast not found")
    return {"message": "Forecast deleted"}


@router.get("/model-forecasts/pivot")
async def get_model_forecasts_pivot(
    vertical_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get model forecasts in pivot format:
    Rows = Models, Columns = Months (mm-yy)
    """
    # Get next 12 months
    now = datetime.now(timezone.utc)
    months = []
    for i in range(1, 13):
        month_date = now.replace(day=1) + relativedelta(months=i)
        months.append({
            "month_year": month_date.strftime("%Y-%m"),
            "display": month_date.strftime("%b-%y")
        })
    
    # Get all models
    model_query = {"status": "ACTIVE"}
    if vertical_id:
        model_query["vertical_id"] = vertical_id
    
    models = await db.models.find(
        model_query,
        {"_id": 0, "id": 1, "code": 1, "name": 1, "vertical_id": 1}
    ).sort("code", 1).to_list(1000)
    
    # Get all forecasts for the period
    month_years = [m["month_year"] for m in months]
    forecasts = await db.model_level_forecasts.find(
        {"month_year": {"$in": month_years}},
        {"_id": 0}
    ).to_list(50000)
    
    # Build forecast lookup
    forecast_map = {}
    for f in forecasts:
        key = f"{f['model_id']}_{f['month_year']}"
        forecast_map[key] = f.get("forecast_qty", 0)
    
    # Get vertical info
    vertical_ids = list(set(m.get("vertical_id") for m in models if m.get("vertical_id")))
    verticals = await db.verticals.find(
        {"id": {"$in": vertical_ids}},
        {"_id": 0, "id": 1, "code": 1, "name": 1}
    ).to_list(100)
    vertical_map = {v["id"]: v for v in verticals}
    
    # Build pivot data
    pivot_data = []
    for model in models:
        vertical = vertical_map.get(model.get("vertical_id"), {})
        row = {
            "model_id": model["id"],
            "model_code": model.get("code", ""),
            "model_name": model.get("name", ""),
            "vertical_code": vertical.get("code", ""),
            "vertical_name": vertical.get("name", ""),
            "forecasts": {}
        }
        
        total = 0
        for month in months:
            key = f"{model['id']}_{month['month_year']}"
            qty = forecast_map.get(key, 0)
            row["forecasts"][month["month_year"]] = qty
            total += qty
        
        row["total"] = total
        pivot_data.append(row)
    
    return {
        "months": months,
        "models": pivot_data
    }


@router.get("/model-forecasts/template")
async def download_forecast_template(
    current_user: User = Depends(get_current_user)
):
    """Download Excel template for model forecasts in pivot format"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel export not available")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model Forecasts"
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Generate next 12 months
    now = datetime.now(timezone.utc)
    months = []
    for i in range(1, 13):
        month_date = now.replace(day=1) + relativedelta(months=i)
        months.append({
            "month_year": month_date.strftime("%Y-%m"),
            "display": month_date.strftime("%b-%y")
        })
    
    # Headers
    headers = ["Model Code", "Model Name", "Vertical"] + [m["display"] for m in months]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Get all active models
    models = await db.models.find(
        {"status": "ACTIVE"},
        {"_id": 0, "id": 1, "code": 1, "name": 1, "vertical_id": 1}
    ).sort("code", 1).to_list(1000)
    
    # Get vertical info
    vertical_ids = list(set(m.get("vertical_id") for m in models if m.get("vertical_id")))
    verticals = await db.verticals.find(
        {"id": {"$in": vertical_ids}},
        {"_id": 0, "id": 1, "code": 1}
    ).to_list(100)
    vertical_map = {v["id"]: v.get("code", "") for v in verticals}
    
    # Get existing forecasts
    month_years = [m["month_year"] for m in months]
    forecasts = await db.model_level_forecasts.find(
        {"month_year": {"$in": month_years}},
        {"_id": 0}
    ).to_list(50000)
    
    forecast_map = {}
    for f in forecasts:
        key = f"{f['model_id']}_{f['month_year']}"
        forecast_map[key] = f.get("forecast_qty", 0)
    
    # Data rows
    for row_num, model in enumerate(models, 2):
        ws.cell(row=row_num, column=1, value=model.get("code", "")).border = thin_border
        ws.cell(row=row_num, column=2, value=model.get("name", "")).border = thin_border
        ws.cell(row=row_num, column=3, value=vertical_map.get(model.get("vertical_id"), "")).border = thin_border
        
        for col_num, month in enumerate(months, 4):
            key = f"{model['id']}_{month['month_year']}"
            qty = forecast_map.get(key, "")
            cell = ws.cell(row=row_num, column=col_num, value=qty if qty else "")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    for col in range(4, 4 + len(months)):
        ws.column_dimensions[get_column_letter(col)].width = 10
    
    # Add instruction sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "Model Forecast Template Instructions:",
        "",
        "1. Fill in forecast quantities for each model and month",
        "2. Leave cells empty for zero forecast",
        "3. Do NOT modify Model Code, Model Name, or Vertical columns",
        "4. Save as .xlsx and upload using 'Upload Forecasts' button",
        "",
        "Column format: Model Code | Model Name | Vertical | Jan-26 | Feb-26 | ...",
        "",
        "Note: Existing forecasts will be updated, new ones will be created."
    ]
    for i, text in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=text)
    ws_inst.column_dimensions['A'].width = 60
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"model_forecast_template_{now.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/model-forecasts/upload")
async def upload_model_forecasts(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload model forecasts from Excel in pivot format"""
    if not pd:
        raise HTTPException(status_code=500, detail="Excel import not available")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
        
        # Get model lookup
        models = await db.models.find({}, {"_id": 0, "id": 1, "code": 1}).to_list(1000)
        model_map = {m["code"]: m["id"] for m in models}
        
        created = 0
        updated = 0
        errors = []
        
        # Process each row
        for idx, row in df.iterrows():
            model_code = str(row.get("Model Code", "")).strip()
            if not model_code or model_code not in model_map:
                continue
            
            model_id = model_map[model_code]
            
            # Process each month column (skip first 3 columns: Model Code, Model Name, Vertical)
            for col in df.columns[3:]:
                try:
                    col_str = str(col).strip()
                    # Parse mm-yy format
                    if '-' in col_str:
                        parts = col_str.split('-')
                        if len(parts) == 2:
                            month_abbr = parts[0]
                            year_short = parts[1]
                            # Convert to YYYY-MM
                            month_map = {
                                'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
                                'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
                                'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
                            }
                            month_num = month_map.get(month_abbr.lower()[:3], None)
                            if month_num and len(year_short) == 2:
                                year_full = f"20{year_short}"
                                month_year = f"{year_full}-{month_num}"
                                
                                qty_val = row[col]
                                if pd.isna(qty_val) or qty_val == "" or qty_val == 0:
                                    continue
                                
                                qty = int(float(qty_val))
                                if qty <= 0:
                                    continue
                                
                                # Check existing
                                existing = await db.model_level_forecasts.find_one({
                                    "model_id": model_id,
                                    "month_year": month_year
                                })
                                
                                if existing:
                                    await db.model_level_forecasts.update_one(
                                        {"id": existing["id"]},
                                        {"$set": {
                                            "forecast_qty": qty,
                                            "updated_at": datetime.now(timezone.utc).isoformat(),
                                            "updated_by": current_user.id
                                        }}
                                    )
                                    updated += 1
                                else:
                                    # Get model details
                                    model = await db.models.find_one({"id": model_id}, {"_id": 0})
                                    vertical = await db.verticals.find_one(
                                        {"id": model.get("vertical_id")}, {"_id": 0}
                                    ) if model else None
                                    
                                    forecast = {
                                        "id": str(uuid.uuid4()),
                                        "model_id": model_id,
                                        "model_code": model.get("code", "") if model else "",
                                        "model_name": model.get("name", "") if model else "",
                                        "vertical_id": model.get("vertical_id", "") if model else "",
                                        "vertical_code": vertical.get("code", "") if vertical else "",
                                        "month_year": month_year,
                                        "forecast_qty": qty,
                                        "created_at": datetime.now(timezone.utc).isoformat(),
                                        "created_by": current_user.id
                                    }
                                    await db.model_level_forecasts.insert_one(forecast)
                                    created += 1
                except Exception as e:
                    errors.append({"row": idx + 2, "column": col, "error": str(e)})
        
        return {
            "message": f"Upload complete: {created} created, {updated} updated",
            "created": created,
            "updated": updated,
            "errors": errors[:20]  # Limit error list
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


# ============ RM Procurement Parameters ============

@router.get("/rm-params")
async def get_rm_procurement_params(
    rm_id: Optional[str] = None,
    category: Optional[str] = None,
    has_vendor: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """Get RM procurement parameters"""
    query = {}
    if rm_id:
        query["rm_id"] = rm_id
    if category:
        query["category"] = category
    if has_vendor is not None:
        if has_vendor:
            query["preferred_vendor_id"] = {"$ne": None}
        else:
            query["preferred_vendor_id"] = None
    
    params = await db.rm_procurement_parameters.find(
        query, {"_id": 0}
    ).to_list(10000)
    
    return params


@router.post("/rm-params")
async def create_rm_procurement_params(
    data: RMProcurementParametersCreate,
    current_user: User = Depends(get_current_user)
):
    """Create or update RM procurement parameters"""
    # Get RM info
    rm = await db.raw_materials.find_one({"rm_id": data.rm_id}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail="RM not found")
    
    # Get vendor info if provided
    vendor_name = None
    if data.preferred_vendor_id:
        vendor = await db.vendors.find_one(
            {"$or": [{"id": data.preferred_vendor_id}, {"vendor_id": data.preferred_vendor_id}]},
            {"_id": 0, "name": 1}
        )
        vendor_name = vendor.get("name") if vendor else None
    
    # Check existing
    existing = await db.rm_procurement_parameters.find_one({"rm_id": data.rm_id})
    
    if existing:
        # Update
        update_data = {
            "safety_stock": data.safety_stock,
            "reorder_point": data.reorder_point,
            "moq": data.moq,
            "batch_size": data.batch_size,
            "lead_time_days": data.lead_time_days,
            "preferred_vendor_id": data.preferred_vendor_id,
            "preferred_vendor_name": vendor_name,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.rm_procurement_parameters.update_one(
            {"rm_id": data.rm_id},
            {"$set": update_data}
        )
        return {"message": "Parameters updated", "rm_id": data.rm_id}
    
    # Create new
    params = {
        "id": str(uuid.uuid4()),
        "rm_id": data.rm_id,
        "rm_name": rm.get("name", data.rm_id),
        "category": rm.get("category", ""),
        "safety_stock": data.safety_stock,
        "reorder_point": data.reorder_point,
        "moq": data.moq,
        "batch_size": data.batch_size,
        "lead_time_days": data.lead_time_days,
        "preferred_vendor_id": data.preferred_vendor_id,
        "preferred_vendor_name": vendor_name,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.rm_procurement_parameters.insert_one(params)
    return {"message": "Parameters created", "rm_id": data.rm_id}


@router.put("/rm-params/{rm_id}")
async def update_rm_procurement_params(
    rm_id: str,
    data: RMProcurementParametersUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update RM procurement parameters"""
    existing = await db.rm_procurement_parameters.find_one({"rm_id": rm_id})
    if not existing:
        raise HTTPException(status_code=404, detail="RM parameters not found")
    
    update_data = {}
    if data.safety_stock is not None:
        update_data["safety_stock"] = data.safety_stock
    if data.reorder_point is not None:
        update_data["reorder_point"] = data.reorder_point
    if data.moq is not None:
        update_data["moq"] = data.moq
    if data.batch_size is not None:
        update_data["batch_size"] = data.batch_size
    if data.lead_time_days is not None:
        update_data["lead_time_days"] = data.lead_time_days
    if data.preferred_vendor_id is not None:
        update_data["preferred_vendor_id"] = data.preferred_vendor_id
        # Get vendor name
        vendor = await db.vendors.find_one(
            {"$or": [{"id": data.preferred_vendor_id}, {"vendor_id": data.preferred_vendor_id}]},
            {"_id": 0, "name": 1}
        )
        update_data["preferred_vendor_name"] = vendor.get("name") if vendor else None
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.rm_procurement_parameters.update_one(
            {"rm_id": rm_id},
            {"$set": update_data}
        )
    
    return {"message": "Parameters updated", "rm_id": rm_id}


@router.post("/rm-params/bulk")
async def bulk_create_rm_params(
    params_list: List[dict] = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Bulk create/update RM procurement parameters"""
    created = 0
    updated = 0
    errors = []
    
    for p in params_list:
        try:
            rm_id = p.get("rm_id")
            if not rm_id:
                errors.append({"data": p, "error": "Missing rm_id"})
                continue
            
            # Check RM exists
            rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
            if not rm:
                errors.append({"data": p, "error": f"RM not found: {rm_id}"})
                continue
            
            existing = await db.rm_procurement_parameters.find_one({"rm_id": rm_id})
            
            vendor_name = None
            if p.get("preferred_vendor_id"):
                vendor = await db.vendors.find_one(
                    {"$or": [{"id": p["preferred_vendor_id"]}, {"vendor_id": p["preferred_vendor_id"]}]},
                    {"_id": 0, "name": 1}
                )
                vendor_name = vendor.get("name") if vendor else None
            
            if existing:
                update_data = {
                    "safety_stock": p.get("safety_stock", existing.get("safety_stock", 0)),
                    "moq": p.get("moq", existing.get("moq", 1)),
                    "batch_size": p.get("batch_size", existing.get("batch_size", 1)),
                    "lead_time_days": p.get("lead_time_days", existing.get("lead_time_days", 7)),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                if p.get("preferred_vendor_id"):
                    update_data["preferred_vendor_id"] = p["preferred_vendor_id"]
                    update_data["preferred_vendor_name"] = vendor_name
                
                await db.rm_procurement_parameters.update_one(
                    {"rm_id": rm_id},
                    {"$set": update_data}
                )
                updated += 1
            else:
                params = {
                    "id": str(uuid.uuid4()),
                    "rm_id": rm_id,
                    "rm_name": rm.get("name", rm_id),
                    "category": rm.get("category", ""),
                    "safety_stock": p.get("safety_stock", 0),
                    "reorder_point": p.get("reorder_point", 0),
                    "moq": p.get("moq", 1),
                    "batch_size": p.get("batch_size", 1),
                    "lead_time_days": p.get("lead_time_days", 7),
                    "preferred_vendor_id": p.get("preferred_vendor_id"),
                    "preferred_vendor_name": vendor_name,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.rm_procurement_parameters.insert_one(params)
                created += 1
                
        except Exception as e:
            errors.append({"data": p, "error": str(e)})
    
    return {"created": created, "updated": updated, "errors": errors}


@router.get("/rm-params/template")
async def download_rm_params_template(
    current_user: User = Depends(get_current_user)
):
    """Download Excel template for RM procurement parameters with all RMs"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel export not available")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RM Parameters"
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["RM ID", "RM Name", "Category", "MOQ", "Batch Size", "Lead Time (days)", 
               "Safety Stock", "Preferred Vendor ID", "Preferred Vendor Name"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Get all RMs that are in BOMs
    boms = await db.common_bom.find({}, {"_id": 0, "items": 1}).to_list(10000)
    rm_ids_in_bom = set()
    for bom in boms:
        for item in bom.get("items", []):
            if item.get("rm_id"):
                rm_ids_in_bom.add(item["rm_id"])
    
    # Get all raw materials
    raw_materials = await db.raw_materials.find(
        {"rm_id": {"$in": list(rm_ids_in_bom)}},
        {"_id": 0, "rm_id": 1, "name": 1, "category": 1}
    ).sort("rm_id", 1).to_list(10000)
    
    # Get existing parameters
    existing_params = await db.rm_procurement_parameters.find(
        {}, {"_id": 0}
    ).to_list(10000)
    params_map = {p["rm_id"]: p for p in existing_params}
    
    # Data rows
    for row_num, rm in enumerate(raw_materials, 2):
        rm_id = rm.get("rm_id", "")
        param = params_map.get(rm_id, {})
        
        ws.cell(row=row_num, column=1, value=rm_id).border = thin_border
        ws.cell(row=row_num, column=2, value=rm.get("name", "")).border = thin_border
        ws.cell(row=row_num, column=3, value=rm.get("category", "")).border = thin_border
        ws.cell(row=row_num, column=4, value=param.get("moq", 1)).border = thin_border
        ws.cell(row=row_num, column=5, value=param.get("batch_size", 1)).border = thin_border
        ws.cell(row=row_num, column=6, value=param.get("lead_time_days", 7)).border = thin_border
        ws.cell(row=row_num, column=7, value=param.get("safety_stock", 0)).border = thin_border
        ws.cell(row=row_num, column=8, value=param.get("preferred_vendor_id", "")).border = thin_border
        ws.cell(row=row_num, column=9, value=param.get("preferred_vendor_name", "")).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 30
    
    # Add vendor reference sheet
    ws_vendors = wb.create_sheet("Vendors")
    vendors = await db.vendors.find({}, {"_id": 0, "id": 1, "vendor_id": 1, "name": 1}).to_list(1000)
    
    ws_vendors.cell(row=1, column=1, value="Vendor ID").font = Font(bold=True)
    ws_vendors.cell(row=1, column=2, value="Vendor Name").font = Font(bold=True)
    
    for row_num, vendor in enumerate(vendors, 2):
        ws_vendors.cell(row=row_num, column=1, value=vendor.get("id") or vendor.get("vendor_id", ""))
        ws_vendors.cell(row=row_num, column=2, value=vendor.get("name", ""))
    
    ws_vendors.column_dimensions['A'].width = 40
    ws_vendors.column_dimensions['B'].width = 50
    
    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "RM Parameters Template Instructions:",
        "",
        "1. Update MOQ, Batch Size, Lead Time, and Safety Stock values as needed",
        "2. For Preferred Vendor ID, copy the vendor ID from the 'Vendors' sheet",
        "3. Do NOT modify RM ID, RM Name, or Category columns",
        "4. Save as .xlsx and upload using 'Upload Parameters' button",
        "",
        "Column descriptions:",
        "- MOQ: Minimum Order Quantity (default: 1)",
        "- Batch Size: Order in multiples of this value (default: 1)",
        "- Lead Time: Days from PO to delivery (default: 7)",
        "- Safety Stock: Minimum stock to maintain (default: 0)",
        "",
        "Note: Existing parameters will be updated, new ones will be created."
    ]
    for i, text in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=text)
    ws_inst.column_dimensions['A'].width = 70
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    now = datetime.now(timezone.utc)
    filename = f"rm_parameters_template_{now.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/rm-params/upload")
async def upload_rm_params(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload RM procurement parameters from Excel"""
    if not pd:
        raise HTTPException(status_code=500, detail="Excel import not available")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
        
        created = 0
        updated = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                rm_id = str(row.get("RM ID", "")).strip()
                if not rm_id:
                    continue
                
                # Check RM exists
                rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
                if not rm:
                    errors.append({"row": idx + 2, "rm_id": rm_id, "error": "RM not found"})
                    continue
                
                # Parse values
                moq = float(row.get("MOQ", 1)) if not pd.isna(row.get("MOQ")) else 1
                batch_size = float(row.get("Batch Size", 1)) if not pd.isna(row.get("Batch Size")) else 1
                lead_time = int(row.get("Lead Time (days)", 7)) if not pd.isna(row.get("Lead Time (days)")) else 7
                safety_stock = float(row.get("Safety Stock", 0)) if not pd.isna(row.get("Safety Stock")) else 0
                
                vendor_id = str(row.get("Preferred Vendor ID", "")).strip() if not pd.isna(row.get("Preferred Vendor ID")) else None
                vendor_name = None
                
                if vendor_id:
                    vendor = await db.vendors.find_one(
                        {"$or": [{"id": vendor_id}, {"vendor_id": vendor_id}]},
                        {"_id": 0, "name": 1}
                    )
                    vendor_name = vendor.get("name") if vendor else None
                
                # Check existing
                existing = await db.rm_procurement_parameters.find_one({"rm_id": rm_id})
                
                if existing:
                    await db.rm_procurement_parameters.update_one(
                        {"rm_id": rm_id},
                        {"$set": {
                            "moq": moq,
                            "batch_size": batch_size,
                            "lead_time_days": lead_time,
                            "safety_stock": safety_stock,
                            "preferred_vendor_id": vendor_id if vendor_id else None,
                            "preferred_vendor_name": vendor_name,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    updated += 1
                else:
                    params = {
                        "id": str(uuid.uuid4()),
                        "rm_id": rm_id,
                        "rm_name": rm.get("name", rm_id),
                        "category": rm.get("category", ""),
                        "safety_stock": safety_stock,
                        "reorder_point": 0,
                        "moq": moq,
                        "batch_size": batch_size,
                        "lead_time_days": lead_time,
                        "preferred_vendor_id": vendor_id if vendor_id else None,
                        "preferred_vendor_name": vendor_name,
                        "is_active": True,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.rm_procurement_parameters.insert_one(params)
                    created += 1
                    
            except Exception as e:
                errors.append({"row": idx + 2, "error": str(e)})
        
        return {
            "message": f"Upload complete: {created} created, {updated} updated",
            "created": created,
            "updated": updated,
            "errors": errors[:20]
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


# ============ MRP Runs ============

@router.get("/runs")
async def get_mrp_runs(
    status: Optional[str] = None,
    limit: int = Query(default=50, le=100),
    current_user: User = Depends(get_current_user)
):
    """Get MRP runs list"""
    query = {}
    if status:
        query["status"] = status
    
    runs = await db.mrp_runs.find(
        query,
        {
            "_id": 0,
            "id": 1,
            "run_code": 1,
            "run_date": 1,
            "status": 1,
            "planning_horizon_months": 1,
            "total_skus": 1,
            "total_rms": 1,
            "total_order_value": 1,
            "created_at": 1,
            "created_by": 1,
            "version": 1,
            "common_weeks_count": 1,
            "brand_specific_weeks_count": 1,
            "summary": 1
        }
    ).sort("created_at", -1).to_list(limit)
    
    return runs


@router.get("/runs/{run_id}")
async def get_mrp_run_detail(
    run_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get detailed MRP run data"""
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    return run


@router.get("/runs/{run_id}/weekly-requirements")
async def get_mrp_weekly_requirements(
    run_id: str,
    weeks: int = Query(default=12, ge=1, le=24),
    current_user: User = Depends(get_current_user)
):
    """
    Get weekly RM requirements breakdown for an MRP run.
    Distributes total requirements across weeks based on production schedule.
    
    Args:
        run_id: MRP run ID
        weeks: Number of weeks to show (12 or 24)
    """
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    
    rm_requirements = run.get("rm_requirements", [])
    
    # Calculate weekly breakdown
    # For simplicity, we'll distribute requirements evenly across weeks
    # with lead time considerations
    now = datetime.now(timezone.utc)
    week_start = now - timedelta(days=now.weekday())  # Monday
    
    # Generate week labels
    week_labels = []
    for i in range(weeks):
        week_date = week_start + timedelta(weeks=i)
        week_labels.append({
            "week_num": i + 1,
            "start_date": week_date.strftime("%Y-%m-%d"),
            "label": f"W{i+1} ({week_date.strftime('%d-%b')})"
        })
    
    # Build weekly requirements for each RM
    weekly_data = []
    for rm in rm_requirements:
        if rm.get("order_qty", 0) <= 0:
            continue
        
        lead_time_days = rm.get("lead_time_days", 7)
        total_qty = rm.get("order_qty", 0)
        
        # Distribute requirements:
        # - Higher quantity in earlier weeks (front-loaded)
        # - Consider lead time for delivery
        weekly_qty = {}
        
        # Simple distribution: split across first few weeks based on lead time
        distribution_weeks = min(4, weeks)  # Distribute over first 4 weeks
        
        # Calculate distribution (60% in week 1, decreasing)
        weights = [0.4, 0.3, 0.2, 0.1][:distribution_weeks]
        total_weight = sum(weights)
        
        remaining = total_qty
        for i, weight in enumerate(weights):
            if i < len(weights) - 1:
                qty = round(total_qty * weight / total_weight)
            else:
                qty = remaining
            weekly_qty[f"week_{i+1}"] = qty
            remaining -= qty
        
        # Fill remaining weeks with 0
        for i in range(distribution_weeks, weeks):
            weekly_qty[f"week_{i+1}"] = 0
        
        weekly_data.append({
            "rm_id": rm["rm_id"],
            "rm_name": rm.get("rm_name", ""),
            "category": rm.get("category", ""),
            "total_required": rm.get("total_required", 0),
            "order_qty": total_qty,
            "lead_time_days": lead_time_days,
            "vendor_name": rm.get("vendor_name", ""),
            "unit_price": rm.get("unit_price", 0),
            "weekly_qty": weekly_qty
        })
    
    return {
        "run_id": run_id,
        "run_code": run.get("run_code", ""),
        "weeks": weeks,
        "week_labels": week_labels,
        "requirements": weekly_data,
        "summary": {
            "total_rms": len(weekly_data),
            "total_order_value": sum(r.get("order_qty", 0) * r.get("unit_price", 0) for r in weekly_data)
        }
    }


@router.get("/runs/{run_id}/download")
async def download_mrp_run_results(
    run_id: str,
    current_user: User = Depends(get_current_user)
):
    """Download MRP run results as Excel"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel export not available")
    
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    
    wb = openpyxl.Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_data = [
        ["MRP Run Summary", ""],
        ["Run Code", run.get("run_code", "")],
        ["Run Date", run.get("run_date", "")],
        ["Status", run.get("status", "")],
        ["Total SKUs", run.get("total_skus", 0)],
        ["Total RMs", run.get("total_rms", 0)],
        ["Total Order Value", run.get("total_order_value", 0)],
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=2, value=value)
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40
    
    # RM Requirements sheet
    ws_rm = wb.create_sheet("RM Requirements")
    
    rm_headers = ["RM ID", "RM Name", "Category", "Total Required", "Current Stock", 
                  "Safety Stock", "Net Requirement", "MOQ", "Order Qty", 
                  "Lead Time (days)", "Vendor", "Unit Price", "Total Cost"]
    
    for col, header in enumerate(rm_headers, 1):
        cell = ws_rm.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    rm_requirements = run.get("rm_requirements", [])
    for row_num, rm in enumerate(rm_requirements, 2):
        ws_rm.cell(row=row_num, column=1, value=rm.get("rm_id", "")).border = thin_border
        ws_rm.cell(row=row_num, column=2, value=rm.get("rm_name", "")).border = thin_border
        ws_rm.cell(row=row_num, column=3, value=rm.get("category", "")).border = thin_border
        ws_rm.cell(row=row_num, column=4, value=rm.get("total_required", 0)).border = thin_border
        ws_rm.cell(row=row_num, column=5, value=rm.get("current_stock", 0)).border = thin_border
        ws_rm.cell(row=row_num, column=6, value=rm.get("safety_stock", 0)).border = thin_border
        ws_rm.cell(row=row_num, column=7, value=rm.get("net_requirement", 0)).border = thin_border
        ws_rm.cell(row=row_num, column=8, value=rm.get("moq", 1)).border = thin_border
        ws_rm.cell(row=row_num, column=9, value=rm.get("order_qty", 0)).border = thin_border
        ws_rm.cell(row=row_num, column=10, value=rm.get("lead_time_days", 7)).border = thin_border
        ws_rm.cell(row=row_num, column=11, value=rm.get("vendor_name", "")).border = thin_border
        ws_rm.cell(row=row_num, column=12, value=rm.get("unit_price", 0)).border = thin_border
        ws_rm.cell(row=row_num, column=13, value=rm.get("total_cost", 0)).border = thin_border
    
    # Set column widths
    for col_num in range(1, 14):
        ws_rm.column_dimensions[get_column_letter(col_num)].width = 15
    ws_rm.column_dimensions['B'].width = 40
    
    # Weekly breakdown sheet (12 weeks)
    ws_weekly = wb.create_sheet("Weekly Breakdown")
    
    # Generate week headers
    now = datetime.now(timezone.utc)
    week_start = now - timedelta(days=now.weekday())
    
    weekly_headers = ["RM ID", "RM Name", "Total Qty"]
    for i in range(12):
        week_date = week_start + timedelta(weeks=i)
        weekly_headers.append(f"W{i+1} ({week_date.strftime('%d-%b')})")
    
    for col, header in enumerate(weekly_headers, 1):
        cell = ws_weekly.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Distribute requirements across weeks
    for row_num, rm in enumerate(rm_requirements, 2):
        if rm.get("order_qty", 0) <= 0:
            continue
        
        ws_weekly.cell(row=row_num, column=1, value=rm.get("rm_id", "")).border = thin_border
        ws_weekly.cell(row=row_num, column=2, value=rm.get("rm_name", "")).border = thin_border
        ws_weekly.cell(row=row_num, column=3, value=rm.get("order_qty", 0)).border = thin_border
        
        # Simple distribution
        total_qty = rm.get("order_qty", 0)
        weights = [0.4, 0.3, 0.2, 0.1]
        remaining = total_qty
        
        for i in range(12):
            if i < 4:
                qty = round(total_qty * weights[i])
                if i == 3:
                    qty = remaining
                remaining -= qty
            else:
                qty = 0
            ws_weekly.cell(row=row_num, column=i+4, value=qty).border = thin_border
    
    # Set widths
    ws_weekly.column_dimensions['A'].width = 15
    ws_weekly.column_dimensions['B'].width = 30
    for i in range(3, 16):
        ws_weekly.column_dimensions[get_column_letter(i)].width = 12
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"mrp_run_{run.get('run_code', run_id)}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/runs/calculate")
async def calculate_mrp(
    planning_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Run MRP calculation (Legacy - non-weekly).
    
    This calculates:
    - Month 1: From production_plans
    - Months 2-12: From model_level_forecasts split by rolling ratios
    - BOM Explosion: SKU -> RM requirements
    - Net requirements with safety stock
    - Order quantities with MOQ/batch size
    - Vendor assignment
    """
    try:
        if planning_date:
            plan_dt = datetime.fromisoformat(planning_date.replace("Z", "+00:00"))
        else:
            plan_dt = None
        
        result = await mrp_service.calculate_mrp(current_user.id, plan_dt)
        
        return {
            "message": "MRP calculation completed",
            "run_id": result["id"],
            "run_code": result["run_code"],
            "total_skus": result["total_skus"],
            "total_rms": result["total_rms"],
            "total_order_value": result["total_order_value"]
        }
        
    except Exception as e:
        logger.error(f"MRP calculation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/runs/calculate-weekly")
async def calculate_weekly_mrp(
    planning_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Run Weekly Time-Phased MRP calculation.
    
    Features:
    - Weekly breakdown of forecasts
    - M1: Buyer SKU level (common + brand-specific BOMs)
    - M2-M12: Bidso SKU level (common BOMs only)
    - Order timing with 7-day site buffer
    - Open PO / scheduled receipts consideration
    - Separate common vs brand-specific ordering plan
    """
    try:
        if planning_date:
            plan_dt = datetime.fromisoformat(planning_date.replace("Z", "+00:00"))
        else:
            plan_dt = None
        
        result = await weekly_mrp_service.calculate_weekly_mrp(current_user.id, plan_dt)
        
        return {
            "message": "Weekly MRP calculation completed",
            "run_id": result["id"],
            "run_code": result["run_code"],
            "version": result.get("version", "WEEKLY_V1"),
            "summary": result.get("summary", {}),
            "alerts_count": len(result.get("alerts", []))
        }
        
    except Exception as e:
        logger.error(f"Weekly MRP calculation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/runs/{run_id}/weekly-plan")
async def get_weekly_order_plan(
    run_id: str,
    plan_type: str = Query("all", description="all, common, or brand_specific"),
    order_week: Optional[str] = Query(None, description="Filter by specific order week (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user)
):
    """
    Get weekly order plan for an MRP run.
    
    Args:
        run_id: MRP run ID
        plan_type: Filter by plan type (all, common, brand_specific)
        order_week: Optional filter for specific week
    """
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    
    # Check if this is a weekly MRP run
    if run.get("version") != "WEEKLY_V1":
        raise HTTPException(
            status_code=400, 
            detail="This MRP run does not have weekly plan. Use calculate-weekly endpoint."
        )
    
    result = {
        "run_id": run_id,
        "run_code": run.get("run_code"),
        "run_date": run.get("run_date"),
        "config": run.get("config", {}),
        "summary": run.get("summary", {})
    }
    
    # Build query
    query = {"run_id": run_id}
    
    if plan_type == "common":
        query["plan_type"] = "COMMON"
    elif plan_type == "brand_specific":
        query["plan_type"] = "BRAND_SPECIFIC"
    
    if order_week:
        query["order_week"] = order_week
    
    # Fetch weekly plans from separate collection
    weekly_plan = []
    async for doc in db.mrp_weekly_plans.find(query, {"_id": 0}).sort("order_week", 1):
        weekly_plan.append(doc.get("week_data", {}))
    
    # If getting all, merge by order week
    if plan_type == "all" and not order_week:
        combined = {}
        for week in weekly_plan:
            ow = week.get("order_week", "")
            if ow not in combined:
                combined[ow] = {**week, "items": list(week.get("items", []))}
            else:
                combined[ow]["items"].extend(week.get("items", []))
                items = combined[ow]["items"]
                combined[ow]["week_summary"] = {
                    "total_items": len(items),
                    "total_cost": round(sum(i.get("total_cost", 0) for i in items), 2)
                }
        weekly_plan = list(sorted(combined.values(), key=lambda x: x.get("order_week", "")))
    
    result["weekly_plan"] = weekly_plan
    result["alerts"] = run.get("alerts", [])
    
    return result


@router.get("/runs/{run_id}/weekly-plan/export")
async def export_weekly_order_plan(
    run_id: str,
    plan_type: str = Query("all", description="all, common, or brand_specific"),
    current_user: User = Depends(get_current_user)
):
    """Export weekly order plan to Excel"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    
    if run.get("version") != "WEEKLY_V1":
        raise HTTPException(status_code=400, detail="Not a weekly MRP run")
    
    # Build query
    query = {"run_id": run_id}
    if plan_type == "common":
        query["plan_type"] = "COMMON"
    elif plan_type == "brand_specific":
        query["plan_type"] = "BRAND_SPECIFIC"
    
    # Fetch weekly plan data from separate collection
    weekly_plan = []
    async for doc in db.mrp_weekly_plans.find(query, {"_id": 0}).sort("order_week", 1):
        weekly_plan.append(doc.get("week_data", {}))
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Order Plan"
    
    # Headers
    headers = [
        "Order Week", "Place By", "RM ID", "RM Name", "Category", "Type",
        "Production Week", "Arrival Date", "Lead Time",
        "Gross Qty", "Safety Stock", "Current Stock", "Scheduled Receipts",
        "Net Qty", "Order Qty", "Vendor", "Unit Price", "Total Cost"
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    row = 2
    for week in weekly_plan:
        order_week = week.get("order_week", "")
        place_by = week.get("place_order_by", order_week)
        
        for item in week.get("items", []):
            ws.cell(row=row, column=1, value=order_week)
            ws.cell(row=row, column=2, value=place_by)
            ws.cell(row=row, column=3, value=item.get("rm_id", ""))
            ws.cell(row=row, column=4, value=item.get("rm_name", ""))
            ws.cell(row=row, column=5, value=item.get("category", ""))
            ws.cell(row=row, column=6, value=item.get("rm_type", ""))
            ws.cell(row=row, column=7, value=item.get("production_week", ""))
            ws.cell(row=row, column=8, value=item.get("arrival_date", ""))
            ws.cell(row=row, column=9, value=item.get("lead_time_days", 0))
            ws.cell(row=row, column=10, value=item.get("gross_qty", 0))
            ws.cell(row=row, column=11, value=item.get("safety_stock", 0))
            ws.cell(row=row, column=12, value=item.get("current_stock", 0))
            ws.cell(row=row, column=13, value=item.get("scheduled_receipts", 0))
            ws.cell(row=row, column=14, value=item.get("net_qty", 0))
            ws.cell(row=row, column=15, value=item.get("order_qty", 0))
            ws.cell(row=row, column=16, value=item.get("vendor_name", ""))
            ws.cell(row=row, column=17, value=item.get("unit_price", 0))
            ws.cell(row=row, column=18, value=item.get("total_cost", 0))
            row += 1
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"weekly_order_plan_{run.get('run_code', run_id)}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/runs/{run_id}/approve")
async def approve_mrp_run(
    run_id: str,
    current_user: User = Depends(get_current_user)
):
    """Approve an MRP run"""
    result = await db.mrp_runs.update_one(
        {"id": run_id, "status": "CALCULATED"},
        {"$set": {
            "status": "APPROVED",
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "approved_by": current_user.id
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(
            status_code=400, 
            detail="Run not found or not in CALCULATED status"
        )
    
    return {"message": "MRP run approved", "run_id": run_id}


@router.post("/runs/{run_id}/generate-pos")
async def generate_draft_pos(
    run_id: str,
    current_user: User = Depends(get_current_user)
):
    """Generate Draft POs from an MRP run"""
    try:
        draft_pos = await mrp_service.generate_draft_pos(run_id, current_user.id)
        
        return {
            "message": f"Generated {len(draft_pos)} draft POs",
            "draft_pos": [
                {
                    "id": po["id"],
                    "draft_po_code": po["draft_po_code"],
                    "vendor_name": po["vendor_name"],
                    "total_items": po["total_items"],
                    "total_amount": po["total_amount"]
                }
                for po in draft_pos
            ]
        }
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Draft PO generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Draft POs ============

@router.get("/draft-pos")
async def get_draft_pos(
    mrp_run_id: Optional[str] = None,
    vendor_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get Draft POs"""
    query = {}
    if mrp_run_id:
        query["mrp_run_id"] = mrp_run_id
    if vendor_id:
        query["vendor_id"] = vendor_id
    if status:
        query["status"] = status
    
    draft_pos = await db.mrp_draft_pos.find(
        query, {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    return draft_pos


@router.get("/draft-pos/{draft_po_id}")
async def get_draft_po_detail(
    draft_po_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get Draft PO details"""
    draft_po = await db.mrp_draft_pos.find_one({"id": draft_po_id}, {"_id": 0})
    if not draft_po:
        raise HTTPException(status_code=404, detail="Draft PO not found")
    return draft_po


@router.put("/draft-pos/{draft_po_id}/vendor")
async def update_draft_po_vendor(
    draft_po_id: str,
    vendor_id: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user)
):
    """Update vendor assignment for a draft PO"""
    # Get vendor info
    vendor = await db.vendors.find_one(
        {"$or": [{"id": vendor_id}, {"vendor_id": vendor_id}]},
        {"_id": 0, "name": 1}
    )
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    result = await db.mrp_draft_pos.update_one(
        {"id": draft_po_id, "status": "DRAFT"},
        {"$set": {
            "vendor_id": vendor_id,
            "vendor_name": vendor.get("name", "")
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(
            status_code=400,
            detail="Draft PO not found or not in DRAFT status"
        )
    
    return {"message": "Vendor updated"}


@router.post("/draft-pos/{draft_po_id}/approve")
async def approve_draft_po(
    draft_po_id: str,
    current_user: User = Depends(get_current_user)
):
    """Approve a draft PO"""
    result = await db.mrp_draft_pos.update_one(
        {"id": draft_po_id, "status": "DRAFT"},
        {"$set": {
            "status": "APPROVED",
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "approved_by": current_user.id
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(
            status_code=400,
            detail="Draft PO not found or not in DRAFT status"
        )
    
    return {"message": "Draft PO approved"}


@router.post("/draft-pos/{draft_po_id}/convert-to-po")
async def convert_draft_to_po(
    draft_po_id: str,
    current_user: User = Depends(get_current_user)
):
    """Convert an approved draft PO to an actual PO"""
    draft_po = await db.mrp_draft_pos.find_one({"id": draft_po_id}, {"_id": 0})
    if not draft_po:
        raise HTTPException(status_code=404, detail="Draft PO not found")
    
    if draft_po.get("status") != "APPROVED":
        raise HTTPException(
            status_code=400,
            detail="Draft PO must be APPROVED before conversion"
        )
    
    if not draft_po.get("vendor_id"):
        raise HTTPException(
            status_code=400,
            detail="Cannot convert: No vendor assigned"
        )
    
    # Generate PO number
    po_count = await db.purchase_orders.count_documents({})
    po_number = f"PO-{datetime.now(timezone.utc).strftime('%Y%m')}-{po_count + 1:04d}"
    
    # Create PO
    po = {
        "id": str(uuid.uuid4()),
        "po_number": po_number,
        "vendor_id": draft_po["vendor_id"],
        "branch_id": "",  # Can be set later
        "branch": "",
        "production_plan_id": None,
        "order_date": datetime.now(timezone.utc).isoformat(),
        "expected_delivery_date": draft_po.get("expected_delivery_date"),
        "total_amount": draft_po["total_amount"],
        "currency": draft_po["currency"],
        "status": "DRAFT",
        "payment_status": "PENDING",
        "notes": f"Generated from MRP - {draft_po['mrp_run_code']}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id,
        "source_draft_po_id": draft_po_id
    }
    
    await db.purchase_orders.insert_one(po)
    
    # Create PO lines
    for line in draft_po.get("lines", []):
        po_line = {
            "id": str(uuid.uuid4()),
            "po_id": po["id"],
            "rm_id": line["rm_id"],
            "quantity_ordered": line["quantity"],
            "quantity_received": 0,
            "unit_price": line["unit_price"],
            "unit_of_measure": "nos",
            "line_total": line["line_total"],
            "status": "PENDING"
        }
        await db.purchase_order_lines.insert_one(po_line)
    
    # Update draft PO
    await db.mrp_draft_pos.update_one(
        {"id": draft_po_id},
        {"$set": {
            "status": "SENT",
            "converted_po_id": po["id"],
            "converted_po_number": po_number
        }}
    )
    
    return {
        "message": "PO created successfully",
        "po_id": po["id"],
        "po_number": po_number
    }


# ============ Weekly PO Generation ============

@router.post("/runs/{run_id}/weekly-pos/preview")
async def preview_weekly_pos(
    run_id: str,
    weeks: List[str] = Body(..., description="List of order_week dates to generate POs for"),
    current_user: User = Depends(get_current_user)
):
    """
    Preview Draft POs for selected weeks - groups items by vendor.
    Returns preview without creating any records.
    """
    # Verify run exists and is weekly
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    
    if run.get("version") != "WEEKLY_V1":
        raise HTTPException(status_code=400, detail="Only weekly MRP runs support weekly PO generation")
    
    # Get weekly plans for selected weeks
    weekly_plans = await db.mrp_weekly_plans.find(
        {"run_id": run_id, "order_week": {"$in": weeks}},
        {"_id": 0}
    ).to_list(100)
    
    if not weekly_plans:
        raise HTTPException(status_code=404, detail="No weekly plan data found for selected weeks")
    
    # Aggregate items by vendor
    vendor_items = {}  # vendor_id -> { vendor_info, items: [], weeks: set }
    items_without_vendor = []
    
    for plan in weekly_plans:
        order_week = plan.get("order_week")
        # Items can be in plan["items"] or plan["week_data"]["items"]
        week_data = plan.get("week_data", {})
        items = plan.get("items", []) or week_data.get("items", [])
        
        for item in items:
            vendor_id = item.get("vendor_id")
            if not vendor_id:
                items_without_vendor.append({
                    **item,
                    "order_week": order_week
                })
                continue
            
            if vendor_id not in vendor_items:
                vendor_items[vendor_id] = {
                    "vendor_id": vendor_id,
                    "vendor_name": item.get("vendor_name", "Unknown Vendor"),
                    "items": [],
                    "weeks": set(),
                    "total_qty": 0,
                    "total_amount": 0
                }
            
            vendor_items[vendor_id]["items"].append({
                "rm_id": item.get("rm_id"),
                "rm_name": item.get("rm_name"),
                "category": item.get("category"),
                "order_week": order_week,
                "production_week": item.get("production_week"),
                "order_qty": item.get("order_qty", 0),
                "unit_price": item.get("unit_price", 0),
                "total_cost": item.get("total_cost", 0),
                "lead_time_days": item.get("lead_time_days", 7)
            })
            vendor_items[vendor_id]["weeks"].add(order_week)
            vendor_items[vendor_id]["total_qty"] += item.get("order_qty", 0)
            vendor_items[vendor_id]["total_amount"] += item.get("total_cost", 0)
    
    # Convert sets to lists for JSON serialization
    preview_pos = []
    for vendor_id, data in vendor_items.items():
        preview_pos.append({
            "vendor_id": vendor_id,
            "vendor_name": data["vendor_name"],
            "weeks_covered": sorted(list(data["weeks"])),
            "total_items": len(data["items"]),
            "total_qty": data["total_qty"],
            "total_amount": data["total_amount"],
            "items": data["items"]
        })
    
    # Sort by vendor name
    preview_pos.sort(key=lambda x: x["vendor_name"])
    
    return {
        "run_id": run_id,
        "run_code": run.get("run_code"),
        "selected_weeks": weeks,
        "preview_pos": preview_pos,
        "items_without_vendor": items_without_vendor,
        "summary": {
            "total_vendors": len(preview_pos),
            "total_items": sum(p["total_items"] for p in preview_pos),
            "total_amount": sum(p["total_amount"] for p in preview_pos),
            "items_needing_vendor": len(items_without_vendor)
        }
    }


@router.post("/runs/{run_id}/weekly-pos/download-template")
async def download_weekly_po_template(
    run_id: str,
    weeks: List[str] = Body(..., description="List of order_week dates"),
    current_user: User = Depends(get_current_user)
):
    """
    Download Excel template with suggested POs for selected weeks.
    User can modify qty, vendor, and upload back.
    """
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel support not available")
    
    # Get preview data
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    
    weekly_plans = await db.mrp_weekly_plans.find(
        {"run_id": run_id, "order_week": {"$in": weeks}},
        {"_id": 0}
    ).to_list(100)
    
    # Get all vendors for dropdown reference
    vendors = await db.vendors.find({}, {"_id": 0, "id": 1, "name": 1, "code": 1}).to_list(1000)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: PO Lines (editable)
    ws_lines = wb.active
    ws_lines.title = "PO_Lines"
    
    headers = ["Order Week", "RM ID", "RM Name", "Category", "Production Week", 
               "Suggested Qty", "Final Qty", "Unit Price", "Total Cost",
               "Vendor ID", "Vendor Name", "Lead Time Days", "Notes"]
    
    # Style headers
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws_lines.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    row_num = 2
    for plan in weekly_plans:
        order_week = plan.get("order_week")
        # Items can be in plan["items"] or plan["week_data"]["items"]
        week_data = plan.get("week_data", {})
        items = plan.get("items", []) or week_data.get("items", [])
        
        for item in items:
            vendor_id = item.get("vendor_id", "")
            vendor_name = item.get("vendor_name", "")
            
            ws_lines.cell(row=row_num, column=1, value=order_week)
            ws_lines.cell(row=row_num, column=2, value=item.get("rm_id"))
            ws_lines.cell(row=row_num, column=3, value=item.get("rm_name"))
            ws_lines.cell(row=row_num, column=4, value=item.get("category"))
            ws_lines.cell(row=row_num, column=5, value=item.get("production_week"))
            ws_lines.cell(row=row_num, column=6, value=item.get("order_qty", 0))
            ws_lines.cell(row=row_num, column=7, value=item.get("order_qty", 0))  # Final Qty (editable)
            ws_lines.cell(row=row_num, column=8, value=item.get("unit_price", 0))
            ws_lines.cell(row=row_num, column=9, value=item.get("total_cost", 0))
            ws_lines.cell(row=row_num, column=10, value=vendor_id)
            ws_lines.cell(row=row_num, column=11, value=vendor_name)
            ws_lines.cell(row=row_num, column=12, value=item.get("lead_time_days", 7))
            ws_lines.cell(row=row_num, column=13, value="")  # Notes
            
            # Highlight rows without vendor
            if not vendor_id:
                warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                for col in range(1, 14):
                    ws_lines.cell(row=row_num, column=col).fill = warning_fill
            
            row_num += 1
    
    # Set column widths
    column_widths = [12, 12, 30, 10, 12, 12, 12, 12, 14, 40, 35, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws_lines.column_dimensions[get_column_letter(col)].width = width
    
    # Sheet 2: Vendor Reference
    ws_vendors = wb.create_sheet("Vendors_Reference")
    ws_vendors.cell(row=1, column=1, value="Vendor ID").font = Font(bold=True)
    ws_vendors.cell(row=1, column=2, value="Vendor Name").font = Font(bold=True)
    ws_vendors.cell(row=1, column=3, value="Vendor Code").font = Font(bold=True)
    
    for idx, vendor in enumerate(vendors, 2):
        ws_vendors.cell(row=idx, column=1, value=vendor.get("id"))
        ws_vendors.cell(row=idx, column=2, value=vendor.get("name"))
        ws_vendors.cell(row=idx, column=3, value=vendor.get("code", ""))
    
    ws_vendors.column_dimensions["A"].width = 40
    ws_vendors.column_dimensions["B"].width = 50
    ws_vendors.column_dimensions["C"].width = 15
    
    # Sheet 3: Instructions
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Weekly PO Template Instructions"],
        [""],
        ["1. Review the PO_Lines sheet"],
        ["2. Modify 'Final Qty' column to adjust order quantities"],
        ["3. To change vendor, update 'Vendor ID' column (refer to Vendors_Reference sheet)"],
        ["4. Yellow highlighted rows need vendor assignment"],
        ["5. Add notes in the 'Notes' column if needed"],
        ["6. Save and upload the file to generate Draft POs"],
        [""],
        [f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC"],
        [f"MRP Run: {run.get('run_code')}"],
        [f"Weeks: {', '.join(weeks)}"],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws_instructions.cell(row=row_idx, column=col_idx, value=value)
    
    ws_instructions.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_instructions.column_dimensions["A"].width = 80
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Weekly_PO_Template_{run.get('run_code')}_{'-'.join(weeks[:2])}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/runs/{run_id}/weekly-pos/upload")
async def upload_weekly_pos(
    run_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Upload modified Excel and generate Draft POs.
    Groups items by vendor and creates separate Draft POs.
    """
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel support not available")
    
    # Verify run
    run = await db.mrp_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="MRP run not found")
    
    # Read Excel
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb["PO_Lines"]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    # Get vendor lookup
    vendors = await db.vendors.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    vendor_lookup = {v["id"]: v["name"] for v in vendors}
    
    # Parse rows and group by vendor
    vendor_items = {}  # vendor_id -> { items, weeks, total }
    errors = []
    
    for row_num in range(2, ws.max_row + 1):
        order_week = ws.cell(row=row_num, column=1).value
        rm_id = ws.cell(row=row_num, column=2).value
        rm_name = ws.cell(row=row_num, column=3).value
        category = ws.cell(row=row_num, column=4).value
        production_week = ws.cell(row=row_num, column=5).value
        final_qty = ws.cell(row=row_num, column=7).value or 0
        unit_price = ws.cell(row=row_num, column=8).value or 0
        vendor_id = ws.cell(row=row_num, column=10).value
        lead_time = ws.cell(row=row_num, column=12).value or 7
        notes = ws.cell(row=row_num, column=13).value or ""
        
        # Skip empty rows
        if not rm_id:
            continue
        
        # Skip zero quantity
        if final_qty <= 0:
            continue
        
        # Validate vendor
        if not vendor_id:
            errors.append(f"Row {row_num}: {rm_id} - No vendor assigned")
            continue
        
        if vendor_id not in vendor_lookup:
            errors.append(f"Row {row_num}: {rm_id} - Invalid vendor ID: {vendor_id}")
            continue
        
        # Calculate line total
        line_total = float(final_qty) * float(unit_price)
        
        # Group by vendor
        if vendor_id not in vendor_items:
            vendor_items[vendor_id] = {
                "vendor_name": vendor_lookup[vendor_id],
                "items": [],
                "weeks": set(),
                "total_amount": 0
            }
        
        vendor_items[vendor_id]["items"].append({
            "rm_id": rm_id,
            "rm_name": rm_name,
            "category": category,
            "order_week": order_week,
            "production_week": production_week,
            "quantity": float(final_qty),
            "unit_price": float(unit_price),
            "line_total": line_total,
            "lead_time_days": lead_time,
            "notes": notes
        })
        vendor_items[vendor_id]["weeks"].add(str(order_week))
        vendor_items[vendor_id]["total_amount"] += line_total
    
    if errors and not vendor_items:
        raise HTTPException(status_code=400, detail={"message": "All rows have errors", "errors": errors})
    
    # Generate Draft POs
    created_pos = []
    now = datetime.now(timezone.utc)
    
    for vendor_id, data in vendor_items.items():
        weeks_list = sorted(list(data["weeks"]))
        
        # Calculate expected delivery (max lead time + 7 day buffer)
        max_lead_time = max(item["lead_time_days"] for item in data["items"])
        expected_delivery = (now + timedelta(days=max_lead_time + 7)).strftime("%Y-%m-%d")
        
        # Generate draft PO code
        po_count = await db.mrp_draft_pos.count_documents({})
        draft_po_code = f"WPO-{now.strftime('%Y%m%d')}-{po_count + 1:04d}"
        
        draft_po = {
            "id": str(uuid.uuid4()),
            "draft_po_code": draft_po_code,
            "mrp_run_id": run_id,
            "mrp_run_code": run.get("run_code"),
            "vendor_id": vendor_id,
            "vendor_name": data["vendor_name"],
            "weeks_covered": weeks_list,
            "total_items": len(data["items"]),
            "total_amount": data["total_amount"],
            "currency": "INR",
            "expected_delivery_date": expected_delivery,
            "suggested_order_date": now.strftime("%Y-%m-%d"),
            "status": "DRAFT",
            "source": "WEEKLY_UPLOAD",
            "lines": data["items"],
            "created_at": now.isoformat(),
            "created_by": current_user.id
        }
        
        await db.mrp_draft_pos.insert_one(draft_po)
        created_pos.append({
            "id": draft_po["id"],
            "draft_po_code": draft_po_code,
            "vendor_name": data["vendor_name"],
            "total_items": len(data["items"]),
            "total_amount": data["total_amount"],
            "weeks_covered": weeks_list
        })
    
    return {
        "message": f"Created {len(created_pos)} Draft POs",
        "created_pos": created_pos,
        "errors": errors if errors else None,
        "summary": {
            "total_pos_created": len(created_pos),
            "total_items": sum(p["total_items"] for p in created_pos),
            "total_amount": sum(p["total_amount"] for p in created_pos),
            "rows_with_errors": len(errors)
        }
    }


@router.get("/weekly-draft-pos")
async def get_weekly_draft_pos(
    run_id: Optional[str] = None,
    weeks: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user)
):
    """Get Weekly Draft POs with optional filters"""
    query = {"source": "WEEKLY_UPLOAD"}
    
    if run_id:
        query["mrp_run_id"] = run_id
    if status:
        query["status"] = status
    if weeks:
        week_list = weeks.split(",")
        query["weeks_covered"] = {"$in": week_list}
    
    draft_pos = await db.mrp_draft_pos.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    
    return draft_pos


@router.put("/weekly-draft-pos/{draft_po_id}")
async def update_weekly_draft_po(
    draft_po_id: str,
    updates: dict = Body(...),
    current_user: User = Depends(get_current_user)
):
    """
    Update a Weekly Draft PO - can modify vendor, quantities, or lines.
    """
    draft_po = await db.mrp_draft_pos.find_one({"id": draft_po_id}, {"_id": 0})
    if not draft_po:
        raise HTTPException(status_code=404, detail="Draft PO not found")
    
    if draft_po.get("status") not in ["DRAFT"]:
        raise HTTPException(status_code=400, detail="Can only edit DRAFT status POs")
    
    update_fields = {}
    
    # Update vendor
    if "vendor_id" in updates:
        vendor = await db.vendors.find_one({"id": updates["vendor_id"]}, {"_id": 0, "name": 1})
        if not vendor:
            raise HTTPException(status_code=400, detail="Invalid vendor ID")
        update_fields["vendor_id"] = updates["vendor_id"]
        update_fields["vendor_name"] = vendor["name"]
    
    # Update lines
    if "lines" in updates:
        new_lines = updates["lines"]
        total_amount = sum(line.get("line_total", line.get("quantity", 0) * line.get("unit_price", 0)) for line in new_lines)
        update_fields["lines"] = new_lines
        update_fields["total_items"] = len(new_lines)
        update_fields["total_amount"] = total_amount
    
    # Update expected delivery
    if "expected_delivery_date" in updates:
        update_fields["expected_delivery_date"] = updates["expected_delivery_date"]
    
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        update_fields["updated_by"] = current_user.id
        
        await db.mrp_draft_pos.update_one(
            {"id": draft_po_id},
            {"$set": update_fields}
        )
    
    updated_po = await db.mrp_draft_pos.find_one({"id": draft_po_id}, {"_id": 0})
    return updated_po


@router.put("/weekly-draft-pos/{draft_po_id}/line/{rm_id}")
async def update_draft_po_line(
    draft_po_id: str,
    rm_id: str,
    quantity: float = Body(..., embed=True),
    current_user: User = Depends(get_current_user)
):
    """Update quantity for a specific line item"""
    draft_po = await db.mrp_draft_pos.find_one({"id": draft_po_id}, {"_id": 0})
    if not draft_po:
        raise HTTPException(status_code=404, detail="Draft PO not found")
    
    if draft_po.get("status") != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only edit DRAFT status POs")
    
    # Find and update line
    lines = draft_po.get("lines", [])
    line_found = False
    total_amount = 0
    
    for line in lines:
        if line.get("rm_id") == rm_id:
            line["quantity"] = quantity
            line["line_total"] = quantity * line.get("unit_price", 0)
            line_found = True
        total_amount += line.get("line_total", 0)
    
    if not line_found:
        raise HTTPException(status_code=404, detail=f"Line item {rm_id} not found")
    
    await db.mrp_draft_pos.update_one(
        {"id": draft_po_id},
        {"$set": {
            "lines": lines,
            "total_amount": total_amount,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Line updated", "rm_id": rm_id, "new_quantity": quantity}


# ============ Seed Data ============

@router.post("/seed-data")
async def seed_mrp_test_data(
    current_user: User = Depends(get_current_user)
):
    """
    Seed test data for MRP module:
    - Model-level forecasts for next 11 months
    - RM procurement parameters with default values
    """
    # Allow master_admin, cpc_planner, and procurement_officer to seed data
    allowed_roles = ["master_admin", "cpc_planner", "procurement_officer"]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail=f"Access denied. Requires one of: {', '.join(allowed_roles)}")
    
    created_forecasts = 0
    created_params = 0
    
    # Get all active models
    models = await db.models.find(
        {"status": "ACTIVE"},
        {"_id": 0, "id": 1, "code": 1, "name": 1, "vertical_id": 1}
    ).to_list(1000)
    
    # Generate forecasts for months 2-12 (starting from next month)
    base_date = datetime.now(timezone.utc).replace(day=1) + relativedelta(months=1)
    
    for model in models:
        vertical = await db.verticals.find_one(
            {"id": model.get("vertical_id")},
            {"_id": 0, "code": 1}
        )
        vertical_code = vertical.get("code", "") if vertical else ""
        
        for month_offset in range(11):  # 11 months
            month_date = base_date + relativedelta(months=month_offset)
            month_year = month_date.strftime("%Y-%m")
            
            # Check if exists
            existing = await db.model_level_forecasts.find_one({
                "model_id": model["id"],
                "month_year": month_year
            })
            
            if not existing:
                # Generate random forecast qty (100-1000)
                import random
                forecast_qty = random.randint(100, 1000)
                
                forecast = {
                    "id": str(uuid.uuid4()),
                    "model_id": model["id"],
                    "model_code": model.get("code", ""),
                    "model_name": model.get("name", ""),
                    "vertical_id": model.get("vertical_id", ""),
                    "vertical_code": vertical_code,
                    "month_year": month_year,
                    "forecast_qty": forecast_qty,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": current_user.id
                }
                await db.model_level_forecasts.insert_one(forecast)
                created_forecasts += 1
    
    # Create RM procurement parameters for RMs without params
    # Get RMs that are in BOMs
    boms = await db.common_bom.find({}, {"_id": 0, "items": 1}).to_list(10000)
    rm_ids_in_bom = set()
    for bom in boms:
        for item in bom.get("items", []):
            if item.get("rm_id"):
                rm_ids_in_bom.add(item["rm_id"])
    
    for rm_id in rm_ids_in_bom:
        existing = await db.rm_procurement_parameters.find_one({"rm_id": rm_id})
        if not existing:
            rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
            if rm:
                # Find lowest price vendor
                price_doc = await db.vendor_rm_prices.find_one(
                    {"rm_id": rm_id},
                    {"_id": 0, "vendor_id": 1, "price": 1},
                    sort=[("price", 1)]
                )
                
                vendor_id = None
                vendor_name = None
                if price_doc:
                    vendor_id = price_doc["vendor_id"]
                    vendor = await db.vendors.find_one(
                        {"$or": [{"id": vendor_id}, {"vendor_id": vendor_id}]},
                        {"_id": 0, "name": 1}
                    )
                    vendor_name = vendor.get("name") if vendor else None
                
                params = {
                    "id": str(uuid.uuid4()),
                    "rm_id": rm_id,
                    "rm_name": rm.get("name", rm_id),
                    "category": rm.get("category", ""),
                    "safety_stock": 10,
                    "reorder_point": 20,
                    "moq": 50,
                    "batch_size": 10,
                    "lead_time_days": 7,
                    "preferred_vendor_id": vendor_id,
                    "preferred_vendor_name": vendor_name,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.rm_procurement_parameters.insert_one(params)
                created_params += 1
    
    return {
        "message": "Seed data created",
        "model_forecasts_created": created_forecasts,
        "rm_params_created": created_params
    }
