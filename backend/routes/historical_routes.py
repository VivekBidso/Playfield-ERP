"""Historical Data Upload Routes - Sales and Production history for reporting"""
from fastapi import APIRouter, HTTPException, File, UploadFile, Query, Depends
from datetime import datetime, timezone
from typing import Optional
import uuid
from io import BytesIO

from database import db

router = APIRouter(tags=["Historical Data"])

# Month parsing helper
MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12
}

def parse_month(month_str: str) -> dict:
    """Parse month string into {month, year, month_key}.

    Accepted formats:
      • 'Jun 2025' / 'June 2025'           (month name + 4-digit year)
      • '2025 Jun'                         (4-digit year + month name)
      • '2025-06'                          (YYYY-MM)
      • 'Jun-25' / 'Jun-2025' / 'June-25'  (Mmm[-YY|-YYYY])
      • '06-2025' / '06/2025' / '6-25'     (MM[-/]YY|YYYY)
      • '01/06/2025' / '15-06-2025'        (DD[/-]MM[/-]YYYY — day discarded)
      • Excel datetime objects             (handled by caller)
    """
    import re

    # Excel datetime cell — caller may pass datetime; handle here too
    if isinstance(month_str, datetime):
        return {
            "month": month_str.month,
            "year": month_str.year,
            "month_key": f"{month_str.year}-{month_str.month:02d}",
        }

    s = str(month_str).strip()
    if not s:
        raise ValueError("Month is empty")

    def _y4(y: int) -> int:
        # 2-digit year → 2000-prefixed (so '25' → 2025, '99' → 2099)
        return 2000 + y if y < 100 else y

    # 1) Whitespace-separated: "Jun 2025" / "June 2025" / "2025 Jun"
    parts = s.split()
    if len(parts) == 2:
        a, b = parts[0].lower().rstrip("."), parts[1].lower().rstrip(".")
        if a in MONTH_MAP and b.isdigit():
            y = _y4(int(b))
            return {"month": MONTH_MAP[a], "year": y, "month_key": f"{y}-{MONTH_MAP[a]:02d}"}
        if b in MONTH_MAP and a.isdigit():
            y = _y4(int(a))
            return {"month": MONTH_MAP[b], "year": y, "month_key": f"{y}-{MONTH_MAP[b]:02d}"}

    # 2) Hyphen / slash separated: 'Jun-25', 'Jun-2025', '2025-06', '06-2025', '06/2025', '01/06/2025'
    tokens = re.split(r"[-/\s]+", s)
    tokens = [t for t in tokens if t]

    # 2a) Two tokens
    if len(tokens) == 2:
        t1, t2 = tokens[0], tokens[1]
        # 'Jun-25' / 'Jun-2025'
        n1 = t1.lower().rstrip(".")
        if n1 in MONTH_MAP and t2.isdigit():
            y = _y4(int(t2))
            return {"month": MONTH_MAP[n1], "year": y, "month_key": f"{y}-{MONTH_MAP[n1]:02d}"}
        # '25-Jun' / '2025-Jun'
        n2 = t2.lower().rstrip(".")
        if n2 in MONTH_MAP and t1.isdigit():
            y = _y4(int(t1))
            return {"month": MONTH_MAP[n2], "year": y, "month_key": f"{y}-{MONTH_MAP[n2]:02d}"}
        # '2025-06' / '06-2025' / '6-25'
        if t1.isdigit() and t2.isdigit():
            i1, i2 = int(t1), int(t2)
            if 1 <= i1 <= 12 and i2 >= 13:        # MM-YYYY or MM-YY (with i2 >= 13 → year)
                y = _y4(i2)
                return {"month": i1, "year": y, "month_key": f"{y}-{i1:02d}"}
            if 1 <= i2 <= 12 and i1 >= 13:        # YYYY-MM or YY-MM
                y = _y4(i1)
                return {"month": i2, "year": y, "month_key": f"{y}-{i2:02d}"}
            # Both <=12 (ambiguous) — prefer YYYY-MM if first is 4-digit
            if len(t1) == 4 and 1 <= i2 <= 12:
                return {"month": i2, "year": i1, "month_key": f"{i1}-{i2:02d}"}
            if len(t2) == 4 and 1 <= i1 <= 12:
                return {"month": i1, "year": i2, "month_key": f"{i2}-{i1:02d}"}

    # 2b) Three tokens — DD-MM-YYYY (day discarded) or YYYY-MM-DD
    if len(tokens) == 3 and all(t.isdigit() for t in tokens):
        a, b, c = (int(x) for x in tokens)
        if len(tokens[0]) == 4:                   # YYYY-MM-DD
            return {"month": b, "year": a, "month_key": f"{a}-{b:02d}"}
        if len(tokens[2]) == 4 or tokens[2].isdigit():   # DD-MM-YYYY (or DD-MM-YY)
            y = _y4(c)
            return {"month": b, "year": y, "month_key": f"{y}-{b:02d}"}

    raise ValueError(
        f"Cannot parse month: '{s}'. Accepted formats: 'Jun 2025', '2025-06', "
        "'Jun-25', 'Jun-2025', '06-2025', '06/2025'."
    )


async def enrich_buyer_sku(buyer_sku_id: str) -> dict:
    """Look up buyer SKU and return enriched data (bidso, model, vertical)"""
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        return None
    
    bidso_sku_id = buyer_sku.get("bidso_sku_id", "")
    bidso = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0}) if bidso_sku_id else None
    
    return {
        "buyer_sku_name": buyer_sku.get("name", ""),
        "bidso_sku_id": bidso_sku_id,
        "brand_code": buyer_sku.get("brand_code", ""),
        "vertical_code": bidso.get("vertical_code", "") if bidso else "",
        "vertical_id": bidso.get("vertical_id", "") if bidso else "",
        "model_code": bidso.get("model_code", "") if bidso else "",
        "model_id": bidso.get("model_id", "") if bidso else "",
    }


# ==================== SALES UPLOAD ====================

@router.post("/historical-sales/upload")
async def upload_historical_sales(
    file: UploadFile = File(...),
    mode: str = Query("append", description="'append' new data or 'overwrite' existing for same months")
):
    """
    Upload historical sales data from Excel.
    Columns: Buyer SKU | Customer ID | Qty | Month | ASP
    
    Auto-enriches with: customer name, bidso SKU, model, vertical, revenue.
    """
    import openpyxl  # noqa: F401  (lazy import — keeps backend startup fast)
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    
    # Parse headers
    headers_raw = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    header_map = {}
    for idx, h in enumerate(headers_raw):
        if h in ("buyer sku", "buyer_sku", "buyer sku id", "buyer_sku_id", "sku"):
            header_map["buyer_sku_id"] = idx
        elif h in ("customer id", "customer_id", "customer", "customer name"):
            header_map["customer_id"] = idx
        elif h in ("qty", "quantity", "units", "sales qty"):
            header_map["qty"] = idx
        elif h in ("month", "period", "month year"):
            header_map["month"] = idx
        elif h in ("asp", "average selling price", "price", "selling price", "unit price"):
            header_map["asp"] = idx
    
    required = ["buyer_sku_id", "customer_id", "qty", "month", "asp"]
    missing = [k for k in required if k not in header_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {missing}. Found: {headers_raw}")
    
    # Pre-load lookups
    buyers_list = await db.buyers.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1}).to_list(1000)
    buyer_lookup = {}
    for b in buyers_list:
        buyer_lookup[b.get("id", "")] = b["name"]
        buyer_lookup[b.get("code", "")] = b["name"]
        buyer_lookup[b["name"]] = b["name"]
        buyer_lookup[b["name"].lower()] = b["name"]
    
    # Parse rows
    rows = []
    errors = []
    months_in_file = set()
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        buyer_sku_id = str(row[header_map["buyer_sku_id"]] or "").strip()
        customer_raw = str(row[header_map["customer_id"]] or "").strip()
        qty_raw = row[header_map["qty"]]
        month_raw = str(row[header_map["month"]] or "").strip()
        asp_raw = row[header_map["asp"]]

        # One structured error record per failing row — frontend renders/downloads
        # the full list as an Excel report. No truncation, no row left silent.
        def _err(reason: str):
            errors.append({
                "row": idx,
                "buyer_sku_id": buyer_sku_id,
                "customer_id": customer_raw,
                "qty": qty_raw if qty_raw is not None else "",
                "month": month_raw,
                "asp": asp_raw if asp_raw is not None else "",
                "error": reason,
            })

        if not buyer_sku_id or not month_raw:
            _err("Missing Buyer SKU or Month")
            continue

        try:
            qty = float(qty_raw) if qty_raw not in (None, "") else 0
        except (ValueError, TypeError):
            _err(f"Invalid qty '{qty_raw}' (must be a number)")
            continue

        try:
            asp = float(asp_raw) if asp_raw not in (None, "") else 0
        except (ValueError, TypeError):
            _err(f"Invalid ASP '{asp_raw}' (must be a number)")
            continue

        try:
            month_info = parse_month(month_raw)
        except ValueError as e:
            _err(str(e))
            continue

        # Resolve customer name
        customer_name = buyer_lookup.get(customer_raw) or buyer_lookup.get(customer_raw.lower()) or customer_raw

        # Enrich buyer SKU
        enrichment = await enrich_buyer_sku(buyer_sku_id)
        if not enrichment:
            _err(f"Buyer SKU '{buyer_sku_id}' not found in SKU master. Create it in SKU Management or fix the SKU id.")
            continue
        
        months_in_file.add(month_info["month_key"])
        
        rows.append({
            "id": str(uuid.uuid4()),
            "buyer_sku_id": buyer_sku_id,
            "buyer_sku_name": enrichment["buyer_sku_name"],
            "customer_id": customer_raw,
            "customer_name": customer_name,
            "qty": qty,
            "asp": asp,
            "revenue": round(qty * asp, 2),
            "month": month_info["month"],
            "year": month_info["year"],
            "month_key": month_info["month_key"],
            "bidso_sku_id": enrichment["bidso_sku_id"],
            "brand_code": enrichment["brand_code"],
            "vertical_code": enrichment["vertical_code"],
            "vertical_id": enrichment["vertical_id"],
            "model_code": enrichment["model_code"],
            "model_id": enrichment["model_id"],
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    if not rows:
        return {
            "message": f"No valid rows to upload — all {len(errors)} row(s) failed validation",
            "inserted": 0,
            "error_count": len(errors),
            "errors": errors,
        }

    # Handle overwrite mode
    deleted = 0
    if mode == "overwrite" and months_in_file:
        result = await db.historical_sales.delete_many({"month_key": {"$in": list(months_in_file)}})
        deleted = result.deleted_count

    # Insert rows
    await db.historical_sales.insert_many(rows)

    return {
        "message": f"Uploaded {len(rows)} sales records for {len(months_in_file)} month(s)",
        "inserted": len(rows),
        "deleted_previous": deleted,
        "months": sorted(list(months_in_file)),
        "error_count": len(errors),
        "errors": errors,
    }


# ==================== PRODUCTION UPLOAD ====================

@router.post("/historical-production/upload")
async def upload_historical_production(
    file: UploadFile = File(...),
    mode: str = Query("append", description="'append' new data or 'overwrite' existing for same months")
):
    """
    Upload historical production data from Excel.
    Columns: Buyer SKU | Branch ID | Qty | Month
    
    Auto-enriches with: branch name, model, vertical, production value (from avg ASP).
    """
    import openpyxl  # noqa: F401  (lazy import — keeps backend startup fast)
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    
    headers_raw = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    header_map = {}
    for idx, h in enumerate(headers_raw):
        if h in ("buyer sku", "buyer_sku", "buyer sku id", "buyer_sku_id", "sku"):
            header_map["buyer_sku_id"] = idx
        elif h in ("branch id", "branch_id", "branch", "branch name"):
            header_map["branch_id"] = idx
        elif h in ("qty", "quantity", "production qty", "units"):
            header_map["qty"] = idx
        elif h in ("month", "period", "month year"):
            header_map["month"] = idx
    
    required = ["buyer_sku_id", "branch_id", "qty", "month"]
    missing = [k for k in required if k not in header_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {missing}. Found: {headers_raw}")
    
    # Pre-load branch lookup
    branches = await db.branches.find({}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(100)
    branch_lookup = {}
    for b in branches:
        branch_lookup[b.get("branch_id", "")] = b["name"]
        branch_lookup[b.get("branch_id", "").lower()] = b["name"]
        branch_lookup[b["name"]] = b["name"]
        branch_lookup[b["name"].lower()] = b["name"]
    
    rows = []
    errors = []
    months_in_file = set()
    sku_months_for_asp = set()
    
    # First pass: parse rows and collect SKU+month pairs for ASP lookup
    parsed_rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        buyer_sku_id = str(row[header_map["buyer_sku_id"]] or "").strip()
        branch_raw = str(row[header_map["branch_id"]] or "").strip()
        qty_raw = row[header_map["qty"]]
        month_raw = str(row[header_map["month"]] or "").strip()
        
        if not buyer_sku_id or not month_raw:
            errors.append(f"Row {idx}: Missing Buyer SKU or Month")
            continue
        
        try:
            qty = float(qty_raw) if qty_raw else 0
        except (ValueError, TypeError):
            errors.append(f"Row {idx}: Invalid qty '{qty_raw}'")
            continue
        
        try:
            month_info = parse_month(month_raw)
        except ValueError as e:
            errors.append(f"Row {idx}: {str(e)}")
            continue
        
        branch_name = branch_lookup.get(branch_raw) or branch_lookup.get(branch_raw.lower())
        if not branch_name:
            errors.append(f"Row {idx}: Branch '{branch_raw}' not found")
            continue
        
        enrichment = await enrich_buyer_sku(buyer_sku_id)
        if not enrichment:
            errors.append(f"Row {idx}: Buyer SKU '{buyer_sku_id}' not found")
            continue
        
        months_in_file.add(month_info["month_key"])
        sku_months_for_asp.add((buyer_sku_id, month_info["month_key"]))
        
        parsed_rows.append({
            "buyer_sku_id": buyer_sku_id,
            "branch_name": branch_name,
            "branch_raw": branch_raw,
            "qty": qty,
            "month_info": month_info,
            "enrichment": enrichment
        })
    
    # Fetch average ASP from historical_sales for each SKU+month
    asp_map = {}
    for sku_id, month_key in sku_months_for_asp:
        pipeline = [
            {"$match": {"buyer_sku_id": sku_id, "month_key": month_key, "asp": {"$gt": 0}}},
            {"$group": {"_id": None, "avg_asp": {"$avg": "$asp"}}}
        ]
        result = await db.historical_sales.aggregate(pipeline).to_list(1)
        if result:
            asp_map[(sku_id, month_key)] = round(result[0]["avg_asp"], 2)
    
    # If no month-specific ASP, try overall average for that SKU
    for sku_id, month_key in sku_months_for_asp:
        if (sku_id, month_key) not in asp_map:
            pipeline = [
                {"$match": {"buyer_sku_id": sku_id, "asp": {"$gt": 0}}},
                {"$group": {"_id": None, "avg_asp": {"$avg": "$asp"}}}
            ]
            result = await db.historical_sales.aggregate(pipeline).to_list(1)
            if result:
                asp_map[(sku_id, month_key)] = round(result[0]["avg_asp"], 2)
    
    # Build final rows with production value
    for p in parsed_rows:
        avg_asp = asp_map.get((p["buyer_sku_id"], p["month_info"]["month_key"]), 0)
        
        rows.append({
            "id": str(uuid.uuid4()),
            "buyer_sku_id": p["buyer_sku_id"],
            "buyer_sku_name": p["enrichment"]["buyer_sku_name"],
            "branch_id": p["branch_raw"],
            "branch_name": p["branch_name"],
            "qty": p["qty"],
            "avg_asp": avg_asp,
            "production_value": round(p["qty"] * avg_asp, 2),
            "month": p["month_info"]["month"],
            "year": p["month_info"]["year"],
            "month_key": p["month_info"]["month_key"],
            "bidso_sku_id": p["enrichment"]["bidso_sku_id"],
            "brand_code": p["enrichment"]["brand_code"],
            "vertical_code": p["enrichment"]["vertical_code"],
            "vertical_id": p["enrichment"]["vertical_id"],
            "model_code": p["enrichment"]["model_code"],
            "model_id": p["enrichment"]["model_id"],
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Handle overwrite
    deleted = 0
    if mode == "overwrite" and months_in_file:
        result = await db.historical_production.delete_many({"month_key": {"$in": list(months_in_file)}})
        deleted = result.deleted_count
    
    if rows:
        await db.historical_production.insert_many(rows)
    
    skus_without_asp = [s for s, m in sku_months_for_asp if (s, m) not in asp_map]
    
    return {
        "message": f"Uploaded {len(rows)} production records for {len(months_in_file)} month(s)",
        "inserted": len(rows),
        "deleted_previous": deleted,
        "months": sorted(list(months_in_file)),
        "skus_without_asp": skus_without_asp[:20],
        "errors": errors[:50]
    }


# ==================== SUMMARY/REPORT ENDPOINTS ====================

@router.get("/historical-sales/summary")
async def get_historical_sales_summary(
    group_by: str = Query("customer", description="customer, model, vertical, bidso_sku, month"),
    from_month: Optional[str] = None,
    to_month: Optional[str] = None
):
    """Get aggregated historical sales data"""
    match = {}
    if from_month:
        match["month_key"] = {"$gte": from_month}
    if to_month:
        match.setdefault("month_key", {})["$lte"] = to_month
    
    group_field_map = {
        "customer": {"customer_name": "$customer_name"},
        "model": {"model_code": "$model_code"},
        "vertical": {"vertical_code": "$vertical_code"},
        "bidso_sku": {"bidso_sku_id": "$bidso_sku_id"},
        "month": {"month_key": "$month_key"}
    }
    
    group_id = group_field_map.get(group_by, {"customer_name": "$customer_name"})
    
    pipeline = [
        {"$match": match} if match else {"$match": {}},
        {"$group": {
            "_id": group_id,
            "total_qty": {"$sum": "$qty"},
            "total_revenue": {"$sum": "$revenue"},
            "avg_asp": {"$avg": "$asp"},
            "records": {"$sum": 1}
        }},
        {"$sort": {"total_revenue": -1}}
    ]
    
    results = await db.historical_sales.aggregate(pipeline).to_list(1000)
    
    summary = []
    for r in results:
        item = {**r["_id"], "total_qty": r["total_qty"], "total_revenue": round(r["total_revenue"], 2), "avg_asp": round(r["avg_asp"], 2), "records": r["records"]}
        summary.append(item)
    
    # Totals
    total_qty = sum(r["total_qty"] for r in summary)
    total_revenue = sum(r["total_revenue"] for r in summary)
    
    return {
        "group_by": group_by,
        "data": summary,
        "totals": {"total_qty": total_qty, "total_revenue": round(total_revenue, 2)},
        "count": len(summary)
    }


@router.get("/historical-production/summary")
async def get_historical_production_summary(
    group_by: str = Query("branch", description="branch, model, vertical, month"),
    from_month: Optional[str] = None,
    to_month: Optional[str] = None
):
    """Get aggregated historical production data"""
    match = {}
    if from_month:
        match["month_key"] = {"$gte": from_month}
    if to_month:
        match.setdefault("month_key", {})["$lte"] = to_month
    
    group_field_map = {
        "branch": {"branch_name": "$branch_name"},
        "model": {"model_code": "$model_code"},
        "vertical": {"vertical_code": "$vertical_code"},
        "month": {"month_key": "$month_key"}
    }
    
    group_id = group_field_map.get(group_by, {"branch_name": "$branch_name"})
    
    pipeline = [
        {"$match": match} if match else {"$match": {}},
        {"$group": {
            "_id": group_id,
            "total_qty": {"$sum": "$qty"},
            "total_value": {"$sum": "$production_value"},
            "avg_asp": {"$avg": "$avg_asp"},
            "records": {"$sum": 1}
        }},
        {"$sort": {"total_value": -1}}
    ]
    
    results = await db.historical_production.aggregate(pipeline).to_list(1000)
    
    summary = []
    for r in results:
        item = {**r["_id"], "total_qty": r["total_qty"], "total_value": round(r["total_value"], 2), "avg_asp": round(r.get("avg_asp", 0), 2), "records": r["records"]}
        summary.append(item)
    
    total_qty = sum(r["total_qty"] for r in summary)
    total_value = sum(r["total_value"] for r in summary)
    
    return {
        "group_by": group_by,
        "data": summary,
        "totals": {"total_qty": total_qty, "total_value": round(total_value, 2)},
        "count": len(summary)
    }


@router.get("/historical-sales/stats")
async def get_historical_sales_stats():
    """Get quick stats about uploaded historical sales data"""
    total = await db.historical_sales.count_documents({})
    if total == 0:
        return {"total_records": 0, "months": [], "total_revenue": 0, "total_qty": 0}
    
    months = await db.historical_sales.distinct("month_key")
    pipeline = [{"$group": {"_id": None, "revenue": {"$sum": "$revenue"}, "qty": {"$sum": "$qty"}}}]
    agg = await db.historical_sales.aggregate(pipeline).to_list(1)
    
    return {
        "total_records": total,
        "months": sorted(months),
        "total_revenue": round(agg[0]["revenue"], 2) if agg else 0,
        "total_qty": agg[0]["qty"] if agg else 0
    }


@router.get("/historical-production/stats")
async def get_historical_production_stats():
    """Get quick stats about uploaded historical production data"""
    total = await db.historical_production.count_documents({})
    if total == 0:
        return {"total_records": 0, "months": [], "total_value": 0, "total_qty": 0}
    
    months = await db.historical_production.distinct("month_key")
    pipeline = [{"$group": {"_id": None, "value": {"$sum": "$production_value"}, "qty": {"$sum": "$qty"}}}]
    agg = await db.historical_production.aggregate(pipeline).to_list(1)
    
    return {
        "total_records": total,
        "months": sorted(months),
        "total_value": round(agg[0]["value"], 2) if agg else 0,
        "total_qty": agg[0]["qty"] if agg else 0
    }
