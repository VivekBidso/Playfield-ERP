"""Raw Materials routes - RM CRUD, inventory, bulk upload"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
import uuid
import openpyxl
import io

from database import db
from models import User, RawMaterial, RawMaterialCreate, ActivateItemRequest
from services.utils import (
    get_current_user, check_master_admin, check_branch_access,
    get_next_rm_sequence, serialize_doc, BRANCHES,
    generate_rm_name
)
from services.rbac_service import require_permission

router = APIRouter(tags=["Raw Materials"])

# Cache for rm_categories description columns
_category_name_format_cache = {}

async def get_category_name_format(category: str) -> list:
    """Get the name format fields for a category from database"""
    global _category_name_format_cache
    
    if category in _category_name_format_cache:
        return _category_name_format_cache[category]
    
    cat_doc = await db.rm_categories.find_one({"code": category}, {"_id": 0, "description_columns": 1})
    if cat_doc:
        desc_cols = cat_doc.get("description_columns", [])
        name_fields = [
            col["key"] for col in sorted(desc_cols, key=lambda x: x.get("order", 0))
            if col.get("include_in_name")
        ]
        _category_name_format_cache[category] = name_fields
        return name_fields
    
    return []


def compute_rm_description(rm: dict, name_format: list = None) -> str:
    """
    Compute description from category_data if description field is null/empty.
    Uses dash-separated format: field1 - field2 - field3
    Supports case-insensitive key matching.
    """
    # If description already exists, return it
    existing_desc = rm.get("description")
    if existing_desc and existing_desc.strip():
        return existing_desc
    
    # Check category_data.name
    cat_data_name = rm.get("category_data", {}).get("name")
    if cat_data_name and cat_data_name.strip():
        return cat_data_name
    
    # Compute from category_data fields
    if not name_format:
        return ""
    
    category_data = rm.get("category_data", {})
    
    # Build case-insensitive lookup map: lowercase_key -> actual_value
    case_insensitive_map = {k.lower(): v for k, v in category_data.items()}
    
    parts = []
    for key in name_format:
        # Try exact match first, then case-insensitive
        value = category_data.get(key) or case_insensitive_map.get(key.lower())
        if value:
            parts.append(str(value).strip())
    
    return " - ".join(parts) if parts else ""


async def enrich_rm_with_description(rm: dict) -> dict:
    """Add computed description to RM if missing"""
    if not rm.get("description"):
        name_format = await get_category_name_format(rm.get("category", ""))
        rm["description"] = compute_rm_description(rm, name_format)
    return rm


async def enrich_rms_with_description(rms: list) -> list:
    """Add computed description to list of RMs efficiently"""
    # Pre-load all category formats needed
    categories_needed = set(rm.get("category", "") for rm in rms if not rm.get("description"))
    category_formats = {}
    
    for cat in categories_needed:
        if cat:
            category_formats[cat] = await get_category_name_format(cat)
    
    # Apply descriptions
    for rm in rms:
        if not rm.get("description"):
            name_format = category_formats.get(rm.get("category", ""), [])
            rm["description"] = compute_rm_description(rm, name_format)
    
    return rms


@router.get("/raw-materials/category-formats")
async def get_category_formats():
    """
    Get current rm_categories description format settings.
    Shows which fields have include_in_name=True for each category.
    """
    categories = await db.rm_categories.find({}, {"_id": 0, "code": 1, "name": 1, "description_columns": 1}).to_list(100)
    
    result = []
    for cat in sorted(categories, key=lambda x: x.get("code", "")):
        code = cat.get("code", "")
        name = cat.get("name", "")
        desc_cols = cat.get("description_columns", [])
        
        # Get fields with include_in_name=True, sorted by order
        name_fields = sorted(
            [c for c in desc_cols if c.get("include_in_name")],
            key=lambda x: x.get("order", 0)
        )
        
        result.append({
            "code": code,
            "name": name,
            "format_keys": [f["key"] for f in name_fields],
            "format_labels": [f["label"] for f in name_fields],
            "format_preview": " - ".join([f"<{f['label']}>" for f in name_fields]) if name_fields else "⚠️ No fields configured"
        })
    
    return {"categories": result}


@router.post("/raw-materials/sync-category-formats")
async def sync_category_formats():
    """
    Sync rm_categories description_columns with the correct format configurations.
    Run this once after deployment to ensure production has correct settings.
    """
    # Define correct description formats
    FORMATS = {
        "LB": {"fields": ["type", "buyer_sku"], "labels": ["Type", "Buyer SKU"]},
        "PM": {"fields": ["model", "type", "specs", "brand"], "labels": ["Model", "Type", "Specs", "Brand"]},
        "BS": {"fields": ["position", "type", "brand", "buyer_sku"], "labels": ["Position", "Type", "Brand", "Buyer SKU"]},
        "INP": {"fields": ["mould_code", "model_name", "part_name", "colour", "mb"], "labels": ["Mould Code", "Model Name", "Part Name", "Colour", "Masterbatch"]},
        "ACC": {"fields": ["type", "model_name", "specs", "colour"], "labels": ["Type", "Model Name", "Specs", "Colour"]},
        "INM": {"fields": ["model_name", "part_name", "colour", "mb"], "labels": ["Model Name", "Part Name", "Colour", "Masterbatch"]},
        "SP": {"fields": ["type", "specs"], "labels": ["Type", "Specs"]},
        "ELC": {"fields": ["model", "type", "specs"], "labels": ["Model", "Type", "Specs"]}
    }
    
    updated = []
    for code, config in FORMATS.items():
        cat = await db.rm_categories.find_one({"code": code})
        if not cat:
            continue
        
        existing_cols = cat.get("description_columns", [])
        new_cols = []
        
        # Add name format fields with correct order and include_in_name
        for order, key in enumerate(config["fields"]):
            existing = next((c for c in existing_cols if c.get("key") == key), None)
            col = existing.copy() if existing else {
                "key": key, "label": config["labels"][order], "type": "text", "required": False, "options": []
            }
            col["include_in_name"] = True
            col["order"] = order
            new_cols.append(col)
        
        # Add remaining columns
        used_keys = set(config["fields"])
        for col in existing_cols:
            if col.get("key") not in used_keys:
                col_copy = col.copy()
                col_copy["include_in_name"] = False
                col_copy["order"] = len(new_cols)
                new_cols.append(col_copy)
        
        await db.rm_categories.update_one({"code": code}, {"$set": {"description_columns": new_cols}})
        updated.append(f"{code}: {' - '.join(config['labels'])}")
    
    # Clear cache
    global _category_name_format_cache
    _category_name_format_cache = {}
    
    return {"success": True, "updated": updated}


@router.post("/raw-materials/backfill-descriptions")
async def backfill_rm_descriptions(force: bool = False, categories: str = None):
    """
    Backfill endpoint to compute and permanently store descriptions for raw materials.
    
    Args:
        force: If True, regenerate ALL descriptions (even existing ones)
        categories: Comma-separated list of categories to backfill (e.g., "INP,ACC,ELC")
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Clear the category format cache to pick up new settings
    global _category_name_format_cache
    _category_name_format_cache = {}
    
    # Build query
    if force:
        # Regenerate all descriptions
        query = {}
        if categories:
            cat_list = [c.strip().upper() for c in categories.split(",")]
            query["category"] = {"$in": cat_list}
    else:
        # Only RMs without description
        query = {
            "$or": [
                {"description": {"$exists": False}},
                {"description": None},
                {"description": ""}
            ]
        }
        if categories:
            cat_list = [c.strip().upper() for c in categories.split(",")]
            query["category"] = {"$in": cat_list}
    
    rms_to_update = await db.raw_materials.find(query, {"_id": 0, "rm_id": 1, "category": 1, "category_data": 1}).to_list(50000)
    logger.info(f"Backfill: Found {len(rms_to_update)} RMs to process (force={force})")
    
    if not rms_to_update:
        return {
            "success": True,
            "message": "No RMs to update",
            "updated": 0,
            "skipped": 0,
            "skipped_details": []
        }
    
    # Pre-load all category formats
    categories_needed = set(rm.get("category", "") for rm in rms_to_update)
    category_formats = {}
    for cat in categories_needed:
        if cat:
            category_formats[cat] = await get_category_name_format(cat)
    
    logger.info(f"Backfill: Category formats loaded: {category_formats}")
    
    updated = 0
    skipped = 0
    skipped_details = []
    
    for rm in rms_to_update:
        rm_id = rm.get("rm_id")
        category = rm.get("category", "")
        name_format = category_formats.get(category, [])
        category_data = rm.get("category_data", {})
        
        # For force mode, we need to compute without checking existing description
        if force:
            # Compute fresh - ignore existing description, use case-insensitive matching
            if not name_format:
                description = ""
            else:
                # Build case-insensitive lookup map
                case_insensitive_map = {k.lower(): v for k, v in category_data.items()}
                parts = []
                for key in name_format:
                    value = category_data.get(key) or case_insensitive_map.get(key.lower())
                    if value:
                        parts.append(str(value).strip())
                description = " - ".join(parts) if parts else ""
        else:
            description = compute_rm_description(rm, name_format)
        
        if description:
            await db.raw_materials.update_one(
                {"rm_id": rm_id},
                {"$set": {"description": description}}
            )
            updated += 1
        else:
            skipped += 1
            if len(skipped_details) < 20:
                skipped_details.append({
                    "rm_id": rm_id,
                    "category": category,
                    "name_format": name_format,
                    "category_data_keys": list(category_data.keys()) if category_data else [],
                    "reason": "No name_format" if not name_format else "Empty category_data fields"
                })
    
    logger.info(f"Backfill complete: {updated} updated, {skipped} skipped")
    
    return {
        "success": True,
        "message": f"Backfill complete. Updated {updated} RMs, skipped {skipped} (no data to generate description)",
        "updated": updated,
        "skipped": skipped,
        "skipped_details": skipped_details,
        "category_formats_used": {k: v for k, v in category_formats.items()},
        "force_mode": force
    }


@router.get("/raw-materials/filter-options")
async def get_filter_options():
    """Get distinct values for filter dropdowns"""
    # Get all RMs to extract unique values
    rms = await db.raw_materials.find({"status": {"$ne": "INACTIVE"}}, {"_id": 0, "category": 1, "category_data": 1}).to_list(20000)
    
    categories = set()
    types = set()
    models = set()
    colours = set()
    brands = set()
    
    for rm in rms:
        if rm.get("category"):
            categories.add(rm["category"])
        
        cat_data = rm.get("category_data", {})
        if cat_data.get("type"):
            types.add(str(cat_data["type"]))
        if cat_data.get("model_name"):
            models.add(str(cat_data["model_name"]))
        if cat_data.get("model"):
            models.add(str(cat_data["model"]))
        if cat_data.get("colour"):
            colours.add(str(cat_data["colour"]))
        if cat_data.get("brand"):
            brands.add(str(cat_data["brand"]))
    
    return {
        "categories": sorted(list(categories)),
        "types": sorted(list(types)),
        "models": sorted(list(models)),
        "colours": sorted(list(colours)),
        "brands": sorted(list(brands))
    }


@router.get("/rm-categories")
async def get_rm_categories():
    """Get all active RM categories from database (Tech Ops is single source of truth)"""
    from services.utils import get_all_rm_categories
    return await get_all_rm_categories()


@router.get("/raw-materials/search")
async def search_raw_materials(
    q: str = "",
    search: str = "",
    source_type: Optional[str] = None,
    max_bom_level: Optional[int] = None,
    limit: int = 20
):
    """
    Search raw materials by rm_id or description.
    Accepts both 'q' and 'search' as query param names.
    """
    query_text = q or search
    if len(query_text) < 2:
        return {"items": [], "total": 0}
    
    q_regex = {"$regex": query_text, "$options": "i"}
    query = {"$or": [{"rm_id": q_regex}, {"description": q_regex}, {"name": q_regex}]}
    
    if source_type == "NOT_PURCHASED":
        query["source_type"] = {"$ne": "PURCHASED"}
    elif source_type:
        query["source_type"] = source_type
    
    if max_bom_level is not None:
        # bom_level could be stored as "L1", "L2" or as int 1, 2
        level_values = []
        for i in range(1, max_bom_level + 1):
            level_values.extend([i, f"L{i}"])
        query["$and"] = query.get("$and", [])
        query.setdefault("$or", [])
        # Need to restructure query since $or is already used
        base_or = query.pop("$or")
        query = {
            "$and": [
                {"$or": base_or},
                {"$or": [{"bom_level": {"$in": level_values}}, {"bom_level": {"$exists": False}}]}
            ]
        }
        if source_type == "NOT_PURCHASED":
            query["source_type"] = {"$ne": "PURCHASED"}
        elif source_type:
            query["source_type"] = source_type
    
    cursor = db.raw_materials.find(query, {"_id": 0, "rm_id": 1, "description": 1, "name": 1, "category": 1, "bom_level": 1, "source_type": 1, "uom": 1}).limit(limit).sort("rm_id", 1)
    results = await cursor.to_list(limit)
    
    items = [{
        "item_id": r["rm_id"],
        "name": r.get("description") or r.get("name") or r["rm_id"],
        "category": r.get("category", ""),
        "bom_level": r.get("bom_level"),
        "source_type": r.get("source_type"),
        "uom": r.get("uom")
    } for r in results]
    
    return {"items": items, "total": len(items)}



@router.post("/raw-materials", response_model=RawMaterial)
@require_permission("RawMaterial", "CREATE")
async def create_raw_material(input: RawMaterialCreate, current_user: User = Depends(get_current_user)):
    """Create a new raw material (MASTER_ADMIN, TECH_OPS_ENGINEER)"""
    from services.utils import get_all_rm_categories
    all_categories = await get_all_rm_categories()
    if input.category not in all_categories:
        raise HTTPException(status_code=400, detail=f"Invalid category: {input.category}. Valid: {list(all_categories.keys())}")
    
    # Generate RM ID
    seq = await get_next_rm_sequence(input.category)
    rm_id = f"{input.category}_{seq:05d}"
    
    # Set UOM and source_type from input or category default
    cat_config = all_categories.get(input.category, {})
    
    rm = RawMaterial(
        rm_id=rm_id,
        category=input.category,
        category_data=input.category_data,
        low_stock_threshold=input.low_stock_threshold,
        rm_level=input.rm_level,
        parent_rm_id=input.parent_rm_id,
        unit_weight_grams=input.unit_weight_grams,
        scrap_factor=input.scrap_factor,
        processing_cost=input.processing_cost,
        secondary_l1_rm_id=input.secondary_l1_rm_id,
        powder_qty_grams=input.powder_qty_grams,
        coating_scrap_factor=input.coating_scrap_factor,
        uom=input.uom or cat_config.get("default_uom") or "PCS",
        source_type=input.source_type or cat_config.get("default_source_type") or "PURCHASED",
        bom_level=cat_config.get("bom_level") or 1
    )
    
    doc = rm.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['created_by'] = current_user.id
    
    # Compute and store description at creation time
    name_format = await get_category_name_format(input.category)
    computed_desc = compute_rm_description(doc, name_format)
    if computed_desc:
        doc['description'] = computed_desc
    
    await db.raw_materials.insert_one(doc)
    
    return rm


@router.post("/raw-materials/bulk-upload")
async def bulk_upload_raw_materials(file: UploadFile = File(...)):
    """
    Smart bulk upload - auto-detects file format:
    - If file has 'RM Code' column: imports with existing IDs (category auto-detected from prefix)
    - If file has 'Category' column: creates new RMs with auto-generated IDs
    """
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Get headers and normalize to lowercase for case-insensitive matching
    raw_headers = [cell.value for cell in ws[1]]
    headers = [str(h).lower().strip().replace(' ', '_') if h else '' for h in raw_headers]
    
    # Auto-detect mode: RM Code present means import with IDs
    has_rm_code = any(h in ['rm_code', 'rm_id', 'raw_material_id', 'code'] for h in headers)
    
    created = 0
    skipped = 0
    skipped_duplicates = []
    errors = []
    created_rms = []  # Track created RM details
    
    # Fetch all active categories from DB (single source of truth)
    from services.utils import get_all_rm_categories
    all_categories = await get_all_rm_categories()
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Skip empty rows
            if not any(row):
                continue
                
            row_data = dict(zip(headers, row))
            
            # MODE 1: Import with existing RM Code
            if has_rm_code:
                rm_id = (
                    row_data.get('rm_code') or 
                    row_data.get('rm_id') or 
                    row_data.get('raw_material_id') or
                    row_data.get('code') or
                    ''
                )
                if rm_id:
                    rm_id = str(rm_id).strip()
                
                if not rm_id:
                    errors.append(f"Row {idx}: No RM Code found")
                    skipped += 1
                    continue
                
                # Auto-detect category from RM code prefix (e.g., SP_197 -> SP)
                category = ''
                if '_' in rm_id:
                    prefix = rm_id.split('_')[0].upper()
                    if prefix in all_categories:
                        category = prefix
                
                # Fallback to category column if present
                if not category:
                    category = str(row_data.get('category', '')).upper().strip()
                
                if not category or category not in all_categories:
                    errors.append(f"Row {idx}: Cannot detect category for '{rm_id}'. Valid: {list(all_categories.keys())}")
                    skipped += 1
                    continue
                
                # Check for duplicate
                existing = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0, "rm_id": 1})
                if existing:
                    skipped_duplicates.append({"row": idx, "rm_id": rm_id})
                    skipped += 1
                    continue
                    
            # MODE 2: Create new with auto-generated ID
            else:
                category = (
                    row_data.get('category') or 
                    row_data.get('cat') or 
                    ''
                )
                if isinstance(category, str):
                    category = category.upper().strip()
                else:
                    category = str(category).upper().strip() if category else ''
                
                if not category or category not in all_categories:
                    errors.append(f"Row {idx}: Invalid category '{category}'. Valid: {list(all_categories.keys())}")
                    skipped += 1
                    continue
                
                # Generate new RM ID
                seq = await get_next_rm_sequence(category)
                rm_id = f"{category}_{seq:05d}"
            
            # Build category data from row - match fields case-insensitively
            category_fields = all_categories[category].get("fields", [])
            category_data = {}
            for field in category_fields:
                # Try various naming conventions
                value = (
                    row_data.get(field) or 
                    row_data.get(field.lower()) or 
                    row_data.get(field.replace('_', ' ')) or
                    row_data.get(field.replace('_', ''))
                )
                if value is not None:
                    category_data[field] = value
            
            # Generate RM name from category_data
            rm_name = generate_rm_name(category, category_data)
            if rm_name:
                category_data["name"] = rm_name
            
            # Handle threshold
            threshold = (
                row_data.get('low_stock_threshold') or 
                row_data.get('low stock threshold') or 
                row_data.get('threshold') or 
                10.0
            )
            try:
                threshold = float(threshold) if threshold else 10.0
            except (ValueError, TypeError):
                threshold = 10.0
            
            # Create RM
            rm = RawMaterial(
                rm_id=rm_id,
                category=category,
                category_data=category_data,
                low_stock_threshold=threshold
            )
            
            doc = rm.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            
            # Compute and store description at creation time
            name_format = await get_category_name_format(category)
            computed_desc = compute_rm_description(doc, name_format)
            if computed_desc:
                doc['description'] = computed_desc
            
            await db.raw_materials.insert_one(doc)
            created += 1
            
            # Track created RM for response
            created_rms.append({
                "rm_id": rm_id,
                "category": category,
                "name": computed_desc or category_data.get("name") or category_data.get("part_name") or category_data.get("type") or "-"
            })
            
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
            skipped += 1
    
    # Build response
    response = {
        "created": created,
        "skipped": skipped,
        "mode": "import_with_ids" if has_rm_code else "create_new",
        "errors": errors[:20],
        "total_errors": len(errors),
        "message": f"Created {created} RMs" + (f", skipped {skipped}" if skipped else ""),
        "created_rms": created_rms  # Include all created RMs
    }
    
    if skipped_duplicates:
        response["duplicates"] = skipped_duplicates[:20]
        response["total_duplicates"] = len(skipped_duplicates)
        response["message"] += f". {len(skipped_duplicates)} duplicate RM IDs skipped."
    
    if errors and created == 0:
        response["message"] = f"0 RMs created. {len(errors)} errors: " + "; ".join(errors[:3])
    
    return response


@router.post("/raw-materials/import-with-ids")
async def import_raw_materials_with_ids(file: UploadFile = File(...), category: str = ""):
    """Import raw materials with specific RM IDs from Excel.
    
    Supports files with RM Code column (e.g., SP_197, ACC_280).
    Auto-detects category from RM code prefix if not specified.
    """
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Get headers and normalize to lowercase
    raw_headers = [cell.value for cell in ws[1]]
    headers = [str(h).lower().strip().replace(' ', '_') if h else '' for h in raw_headers]
    
    created = 0
    skipped_duplicates = []
    errors = []
    
    # Fetch all active categories from DB (single source of truth)
    from services.utils import get_all_rm_categories
    all_categories = await get_all_rm_categories()
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Skip empty rows
            if not any(row):
                continue
                
            row_data = dict(zip(headers, row))
            
            # Get RM ID
            rm_id = (
                row_data.get('rm_code') or 
                row_data.get('rm_id') or 
                row_data.get('raw_material_id') or
                row_data.get('code') or
                ''
            )
            if rm_id:
                rm_id = str(rm_id).strip()
            
            if not rm_id:
                errors.append(f"Row {idx}: No RM Code/ID found")
                continue
            
            # Auto-detect category from RM code prefix
            cat = category.upper() if category else ''
            if not cat and '_' in rm_id:
                prefix = rm_id.split('_')[0].upper()
                if prefix in all_categories:
                    cat = prefix
            
            if not cat:
                cat = str(row_data.get('category', '')).upper().strip()
            
            if not cat or cat not in all_categories:
                errors.append(f"Row {idx}: Invalid or missing category '{cat}'. Valid: {list(all_categories.keys())}")
                continue
            
            # Check for existing RM - PREVENT OVERWRITE
            if rm_id:
                existing = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0, "rm_id": 1})
                if existing:
                    skipped_duplicates.append({"row": idx, "rm_id": rm_id, "reason": "RM ID already exists"})
                    continue
            else:
                seq = await get_next_rm_sequence(cat)
                rm_id = f"{cat}_{seq:05d}"
            
            # Build category data
            category_fields = all_categories[cat].get("fields", [])
            category_data = {}
            for field in category_fields:
                if field in row_data and row_data[field]:
                    category_data[field] = row_data[field]
            
            # Generate RM name from category_data based on nomenclature
            rm_name = generate_rm_name(cat, category_data)
            if rm_name:
                category_data["name"] = rm_name
            
            # Create new RM
            rm = RawMaterial(
                rm_id=rm_id,
                category=cat,
                category_data=category_data,
                low_stock_threshold=row_data.get('low_stock_threshold', 10.0) or 10.0,
                rm_level=row_data.get('rm_level', 'DIRECT') or 'DIRECT',
                parent_rm_id=row_data.get('parent_rm_id'),
                unit_weight_grams=row_data.get('unit_weight_grams'),
                scrap_factor=row_data.get('scrap_factor', 0.02) or 0.02,
                secondary_l1_rm_id=row_data.get('secondary_l1_rm_id'),
                powder_qty_grams=row_data.get('powder_qty_grams')
            )
            
            doc = rm.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.raw_materials.insert_one(doc)
            created += 1
                
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    # If there are duplicates, return error response
    if skipped_duplicates:
        return {
            "success": False,
            "created": created,
            "duplicates": skipped_duplicates,
            "errors": errors,
            "message": f"Upload blocked: {len(skipped_duplicates)} duplicate RM IDs found. No overwrites allowed."
        }
    
    return {
        "success": True,
        "created": created,
        "errors": errors,
        "message": f"Successfully created {created} raw materials"
    }


@router.get("/raw-materials")
async def get_raw_materials(
    branch: Optional[str] = None,
    search: Optional[str] = None,
    include_inactive: bool = False
):
    """Get all raw materials with optional filters"""
    query = {}
    if not include_inactive:
        query["status"] = {"$ne": "INACTIVE"}
    
    rms = await db.raw_materials.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by search
    if search:
        search_lower = search.lower()
        filtered = []
        for rm in rms:
            if search_lower in rm.get("rm_id", "").lower():
                filtered.append(rm)
                continue
            # Search in category_data
            for val in rm.get("category_data", {}).values():
                if val and search_lower in str(val).lower():
                    filtered.append(rm)
                    break
        rms = filtered
    
    # Add branch inventory data if branch specified
    if branch:
        for rm in rms:
            inv = await db.branch_rm_inventory.find_one(
                {"rm_id": rm["rm_id"], "branch": branch},
                {"_id": 0}
            )
            rm["branch_stock"] = inv.get("current_stock", 0.0) if inv else 0.0
            rm["is_active_in_branch"] = inv.get("is_active", False) if inv else False
    
    return rms


@router.get("/raw-materials/filtered")
async def get_raw_materials_filtered(
    page: int = 1,
    page_size: int = 100,
    branch: Optional[str] = None,
    search: Optional[str] = None,
    category: Optional[str] = None,
    type_filter: Optional[str] = None,
    model_filter: Optional[str] = None,
    colour_filter: Optional[str] = None,
    brand_filter: Optional[str] = None
):
    """Get raw materials with pagination and filters"""
    query = {"status": {"$ne": "INACTIVE"}}
    
    # Apply category filter
    if category:
        query["category"] = category
    
    # Get all matching RMs first
    rms = await db.raw_materials.find(query, {"_id": 0}).to_list(15000)
    
    # Apply text search
    if search:
        search_lower = search.lower()
        filtered = []
        for rm in rms:
            if search_lower in rm.get("rm_id", "").lower():
                filtered.append(rm)
                continue
            if search_lower in rm.get("description", "").lower():
                filtered.append(rm)
                continue
            # Search in category_data
            for val in rm.get("category_data", {}).values():
                if val and search_lower in str(val).lower():
                    filtered.append(rm)
                    break
        rms = filtered
    
    # Apply category_data filters
    if type_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("type") == type_filter]
    if model_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("model_name") == model_filter or rm.get("category_data", {}).get("model") == model_filter]
    if colour_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("colour") == colour_filter]
    if brand_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("brand") == brand_filter]
    
    # Add branch inventory data if branch specified
    if branch:
        branch_inv = await db.branch_rm_inventory.find(
            {"branch": branch},
            {"_id": 0, "rm_id": 1, "current_stock": 1, "is_active": 1}
        ).to_list(15000)
        inv_map = {inv["rm_id"]: inv for inv in branch_inv}
        
        for rm in rms:
            inv = inv_map.get(rm["rm_id"])
            rm["branch_stock"] = inv.get("current_stock", 0.0) if inv else 0.0
            rm["is_active_in_branch"] = inv.get("is_active", False) if inv else False
    
    # Description is now stored in DB - no on-the-fly computation needed
    
    # Calculate pagination
    total = len(rms)
    total_pages = (total + page_size - 1) // page_size
    start = (page - 1) * page_size
    end = start + page_size
    
    return {
        "items": rms[start:end],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }


@router.get("/raw-materials/export")
async def export_raw_materials(
    branch: Optional[str] = None,
    search: Optional[str] = None,
    category: Optional[str] = None,
    type_filter: Optional[str] = None,
    model_filter: Optional[str] = None,
    colour_filter: Optional[str] = None,
    brand_filter: Optional[str] = None
):
    """
    Export ALL raw materials matching filters (no pagination) as Excel.
    - If branch is specified: Export only that branch's stock in separate columns
    - If NO branch filter: Export ALL branches, each row = RM + Branch combo with Branch ID column
    """
    from fastapi.responses import StreamingResponse
    
    query = {"status": {"$ne": "INACTIVE"}}
    
    if category:
        query["category"] = category
    
    rms = await db.raw_materials.find(query, {"_id": 0}).to_list(15000)
    
    # Apply text search
    if search:
        search_lower = search.lower()
        filtered = []
        for rm in rms:
            if search_lower in rm.get("rm_id", "").lower():
                filtered.append(rm)
                continue
            if search_lower in rm.get("description", "").lower():
                filtered.append(rm)
                continue
            for val in rm.get("category_data", {}).values():
                if val and search_lower in str(val).lower():
                    filtered.append(rm)
                    break
        rms = filtered
    
    # Apply category_data filters
    if type_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("type") == type_filter]
    if model_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("model_name") == model_filter or rm.get("category_data", {}).get("model") == model_filter]
    if colour_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("colour") == colour_filter]
    if brand_filter:
        rms = [rm for rm in rms if rm.get("category_data", {}).get("brand") == brand_filter]
    
    # Description is now stored in DB - no on-the-fly computation needed
    
    # Get category fields for column headers
    all_cat_fields = set()
    for rm in rms:
        all_cat_fields.update(rm.get("category_data", {}).keys())
    cat_field_list = sorted(list(all_cat_fields))
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Materials"
    
    if branch:
        # SINGLE BRANCH EXPORT: Add branch inventory columns
        branch_inv = await db.branch_rm_inventory.find(
            {"branch": branch},
            {"_id": 0, "rm_id": 1, "current_stock": 1, "is_active": 1}
        ).to_list(15000)
        inv_map = {inv["rm_id"]: inv for inv in branch_inv}
        
        # Headers for single branch export
        headers = ["RM ID", "Category", "Description", "Unit", "Threshold"]
        headers.extend([f.replace("_", " ").title() for f in cat_field_list])
        headers.extend(["Branch Stock", "Active in Branch"])
        ws.append(headers)
        
        # Data rows
        for rm in rms:
            inv = inv_map.get(rm["rm_id"])
            row = [
                rm.get("rm_id", ""),
                rm.get("category", ""),
                rm.get("description", ""),
                rm.get("unit", ""),
                rm.get("low_stock_threshold", "")
            ]
            cat_data = rm.get("category_data", {})
            for field in cat_field_list:
                row.append(cat_data.get(field, ""))
            row.append(inv.get("current_stock", 0.0) if inv else 0.0)
            row.append("Yes" if (inv and inv.get("is_active")) else "No")
            ws.append(row)
        
        filename = f"rm_stock_{branch.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    else:
        # ALL BRANCHES EXPORT: Each row = RM + Branch combo with Branch ID column
        # Get all branches and their branch_ids
        branches = await db.branches.find({"is_active": True}, {"_id": 0, "name": 1, "branch_id": 1}).to_list(100)
        branch_to_id = {b["name"]: b.get("branch_id", "") for b in branches}
        
        # Get ALL branch inventory data
        all_branch_inv = await db.branch_rm_inventory.find(
            {},
            {"_id": 0, "rm_id": 1, "branch": 1, "current_stock": 1, "is_active": 1}
        ).to_list(100000)
        
        # Build inventory lookup: {rm_id: {branch: {current_stock, is_active}}}
        inv_by_rm_branch = {}
        for inv in all_branch_inv:
            rm_id = inv.get("rm_id")
            b = inv.get("branch")
            if rm_id not in inv_by_rm_branch:
                inv_by_rm_branch[rm_id] = {}
            inv_by_rm_branch[rm_id][b] = {
                "current_stock": inv.get("current_stock", 0.0),
                "is_active": inv.get("is_active", False)
            }
        
        # Headers for all-branches export (includes Branch ID)
        headers = ["Branch ID", "Branch", "RM ID", "Category", "Description", "Unit", "Threshold"]
        headers.extend([f.replace("_", " ").title() for f in cat_field_list])
        headers.extend(["Current Stock", "Active"])
        ws.append(headers)
        
        # Data rows: One row per RM + Branch combination
        for rm in rms:
            rm_id = rm.get("rm_id", "")
            rm_branch_inv = inv_by_rm_branch.get(rm_id, {})
            
            for branch_name in [b["name"] for b in branches]:
                inv = rm_branch_inv.get(branch_name, {})
                branch_id = branch_to_id.get(branch_name, "")
                
                row = [
                    branch_id,
                    branch_name,
                    rm_id,
                    rm.get("category", ""),
                    rm.get("description", ""),
                    rm.get("unit", ""),
                    rm.get("low_stock_threshold", "")
                ]
                cat_data = rm.get("category_data", {})
                for field in cat_field_list:
                    row.append(cat_data.get(field, ""))
                row.append(inv.get("current_stock", 0.0))
                row.append("Yes" if inv.get("is_active") else "No")
                ws.append(row)
        
        filename = f"rm_stock_all_branches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/raw-materials/activate")
async def activate_rm_in_branch(request: ActivateItemRequest):
    """Activate a raw material in a specific branch"""
    rm = await db.raw_materials.find_one({"rm_id": request.item_id}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    await db.branch_rm_inventory.update_one(
        {"rm_id": request.item_id, "branch": request.branch},
        {
            "$set": {"is_active": True},
            "$setOnInsert": {"id": str(uuid.uuid4()), "current_stock": 0.0, "created_at": datetime.now(timezone.utc).isoformat()}
        },
        upsert=True
    )
    
    return {"message": f"Raw material {request.item_id} activated in {request.branch}"}


@router.delete("/raw-materials/{rm_id}")
@require_permission("RawMaterial", "DELETE")
async def delete_raw_material(rm_id: str, current_user: User = Depends(get_current_user)):
    """Delete a raw material (MASTER_ADMIN only, TECH_OPS can soft-delete)"""
    result = await db.raw_materials.delete_one({"rm_id": rm_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    # Also delete branch inventory records
    await db.branch_rm_inventory.delete_many({"rm_id": rm_id})
    
    return {"message": f"Raw material {rm_id} deleted"}


# ============ RM Tagging Endpoints ============

@router.put("/raw-materials/{rm_id}")
async def update_raw_material(rm_id: str, data: dict):
    """Update a raw material - especially for brand/model/vertical tagging"""
    from models import RawMaterialUpdate
    
    existing = await db.raw_materials.find_one({"rm_id": rm_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    update_fields = {}
    
    # Handle optional fields
    if "category_data" in data and data["category_data"] is not None:
        update_fields["category_data"] = data["category_data"]
    if "low_stock_threshold" in data and data["low_stock_threshold"] is not None:
        update_fields["low_stock_threshold"] = data["low_stock_threshold"]
    if "brand_ids" in data:
        update_fields["brand_ids"] = data["brand_ids"] or []
    if "vertical_ids" in data:
        update_fields["vertical_ids"] = data["vertical_ids"] or []
    if "model_ids" in data:
        update_fields["model_ids"] = data["model_ids"] or []
    if "is_brand_specific" in data:
        update_fields["is_brand_specific"] = data["is_brand_specific"]
    if "status" in data and data["status"]:
        update_fields["status"] = data["status"]
    if "uom" in data and data["uom"]:
        update_fields["uom"] = data["uom"]
    if "source_type" in data and data["source_type"]:
        update_fields["source_type"] = data["source_type"]
    
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.raw_materials.update_one(
            {"rm_id": rm_id},
            {"$set": update_fields}
        )
    
    return {"message": f"Raw material {rm_id} updated"}


@router.get("/raw-materials/by-tags")
async def get_raw_materials_by_tags(
    brand_id: Optional[str] = None,
    vertical_id: Optional[str] = None,
    model_id: Optional[str] = None,
    is_brand_specific: Optional[bool] = None,
    category: Optional[str] = None,
    source_type: Optional[str] = None,
    bom_level: Optional[int] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1, description="Page number (1-indexed)"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page")
):
    """Get RMs filtered by brand/vertical/model tags with pagination"""
    query = {"status": {"$ne": "INACTIVE"}}
    
    if brand_id:
        query["brand_ids"] = brand_id
    if vertical_id:
        query["vertical_ids"] = vertical_id
    if model_id:
        query["model_ids"] = model_id
    if is_brand_specific is not None:
        query["is_brand_specific"] = is_brand_specific
    if category:
        query["category"] = category
    if source_type:
        query["source_type"] = source_type
    if bom_level is not None:
        query["bom_level"] = bom_level
    
    # Handle search with pagination
    if search:
        search_lower = search.lower()
        # First get all matching IDs for search
        all_rms = await db.raw_materials.find(query, {"_id": 0, "rm_id": 1, "category": 1}).to_list(10000)
        matching_ids = [rm["rm_id"] for rm in all_rms if
                       search_lower in rm.get("rm_id", "").lower() or
                       search_lower in rm.get("category", "").lower()]
        total = len(matching_ids)
        
        # Paginate
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_ids = matching_ids[start_idx:end_idx]
        
        if paginated_ids:
            rms = await db.raw_materials.find(
                {"rm_id": {"$in": paginated_ids}},
                {"_id": 0}
            ).to_list(page_size)
        else:
            rms = []
    else:
        total = await db.raw_materials.count_documents(query)
        skip = (page - 1) * page_size
        rms = await db.raw_materials.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    # Enrich with names
    for rm in rms:
        # Get brand names
        if rm.get("brand_ids"):
            brands = await db.brands.find({"id": {"$in": rm["brand_ids"]}}, {"_id": 0, "code": 1, "name": 1}).to_list(100)
            rm["brands"] = brands
        else:
            rm["brands"] = []
        
        # Get vertical names
        if rm.get("vertical_ids"):
            verticals = await db.verticals.find({"id": {"$in": rm["vertical_ids"]}}, {"_id": 0, "code": 1, "name": 1}).to_list(100)
            rm["verticals"] = verticals
        else:
            rm["verticals"] = []
        
        # Get model names
        if rm.get("model_ids"):
            models = await db.models.find({"id": {"$in": rm["model_ids"]}}, {"_id": 0, "code": 1, "name": 1}).to_list(100)
            rm["models"] = models
        else:
            rm["models"] = []
    
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    return {
        "items": rms,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }


@router.post("/raw-materials/{rm_id}/tag")
async def tag_raw_material(rm_id: str, data: dict):
    """Add tags to a raw material (brands, verticals, models)"""
    existing = await db.raw_materials.find_one({"rm_id": rm_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    update_ops = {}
    
    # Add to arrays (avoid duplicates)
    if "brand_ids" in data and data["brand_ids"]:
        update_ops["$addToSet"] = update_ops.get("$addToSet", {})
        update_ops["$addToSet"]["brand_ids"] = {"$each": data["brand_ids"]}
    
    if "vertical_ids" in data and data["vertical_ids"]:
        update_ops["$addToSet"] = update_ops.get("$addToSet", {})
        update_ops["$addToSet"]["vertical_ids"] = {"$each": data["vertical_ids"]}
    
    if "model_ids" in data and data["model_ids"]:
        update_ops["$addToSet"] = update_ops.get("$addToSet", {})
        update_ops["$addToSet"]["model_ids"] = {"$each": data["model_ids"]}
    
    if "is_brand_specific" in data:
        update_ops["$set"] = update_ops.get("$set", {})
        update_ops["$set"]["is_brand_specific"] = data["is_brand_specific"]
    
    if update_ops:
        update_ops["$set"] = update_ops.get("$set", {})
        update_ops["$set"]["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.raw_materials.update_one({"rm_id": rm_id}, update_ops)
    
    return {"message": f"Tags added to {rm_id}"}


@router.post("/raw-materials/{rm_id}/untag")
async def untag_raw_material(rm_id: str, data: dict):
    """Remove tags from a raw material"""
    existing = await db.raw_materials.find_one({"rm_id": rm_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    update_ops = {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    
    if "brand_ids" in data and data["brand_ids"]:
        update_ops["$pull"] = update_ops.get("$pull", {})
        update_ops["$pullAll"] = {"brand_ids": data["brand_ids"]}
    
    if "vertical_ids" in data and data["vertical_ids"]:
        update_ops["$pullAll"] = update_ops.get("$pullAll", {})
        update_ops["$pullAll"]["vertical_ids"] = data["vertical_ids"]
    
    if "model_ids" in data and data["model_ids"]:
        update_ops["$pullAll"] = update_ops.get("$pullAll", {})
        update_ops["$pullAll"]["model_ids"] = data["model_ids"]
    
    await db.raw_materials.update_one({"rm_id": rm_id}, update_ops)
    
    return {"message": f"Tags removed from {rm_id}"}


# ============ RM Request Workflow ============

@router.get("/rm-requests")
async def get_rm_requests(
    status: Optional[str] = None,
    requested_by: Optional[str] = None
):
    """Get all RM requests"""
    query = {}
    if status:
        query["status"] = status
    if requested_by:
        query["requested_by"] = requested_by
    
    requests = await db.rm_requests.find(query, {"_id": 0}).sort("requested_at", -1).to_list(1000)
    
    # Enrich with user names and tag names
    for req in requests:
        # Requester name
        if req.get("requested_by"):
            user = await db.users.find_one({"id": req["requested_by"]}, {"_id": 0, "name": 1})
            req["requester_name"] = user["name"] if user else None
        
        # Reviewer name
        if req.get("reviewed_by"):
            user = await db.users.find_one({"id": req["reviewed_by"]}, {"_id": 0, "name": 1})
            req["reviewer_name"] = user["name"] if user else None
        
        # Brand names
        if req.get("brand_ids"):
            brands = await db.brands.find({"id": {"$in": req["brand_ids"]}}, {"_id": 0, "code": 1, "name": 1}).to_list(100)
            req["brands"] = brands
        
        # Buyer SKU info
        if req.get("buyer_sku_id"):
            buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": req["buyer_sku_id"]}, {"_id": 0, "name": 1})
            req["buyer_sku_name"] = buyer_sku["name"] if buyer_sku else None
    
    return requests


@router.post("/rm-requests")
async def create_rm_request(data: dict, current_user: User = Depends(get_current_user)):
    """Create a new RM request (Demand team)"""
    from models import RMRequest
    
    request = RMRequest(
        category=data.get("category", "LB"),
        requested_name=data.get("requested_name", ""),
        description=data.get("description", ""),
        category_data=data.get("category_data", {}),
        artwork_files=data.get("artwork_files", []),
        brand_ids=data.get("brand_ids", []),
        vertical_ids=data.get("vertical_ids", []),
        model_ids=data.get("model_ids", []),
        buyer_sku_id=data.get("buyer_sku_id"),
        requested_by=current_user.id,
        requester_name=current_user.name
    )
    
    doc = request.model_dump()
    doc["requested_at"] = doc["requested_at"].isoformat()
    await db.rm_requests.insert_one(doc)
    
    return {"message": "RM request created", "id": request.id}


@router.post("/rm-requests/{request_id}/review")
async def review_rm_request(request_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Review (approve/reject) an RM request (Tech Ops)"""
    request = await db.rm_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="RM request not found")
    
    if request.get("status") != "PENDING":
        raise HTTPException(status_code=400, detail="Request already reviewed")
    
    action = data.get("action", "").upper()
    review_notes = data.get("review_notes", "")
    
    if action not in ["APPROVE", "REJECT"]:
        raise HTTPException(status_code=400, detail="Invalid action. Use APPROVE or REJECT")
    
    update_data = {
        "reviewed_by": current_user.id,
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
        "review_notes": review_notes
    }
    
    if action == "REJECT":
        update_data["status"] = "REJECTED"
    else:
        # Approve and create RM
        category = request.get("category", "LB")
        seq = await get_next_rm_sequence(category)
        rm_id = f"{category}_{seq:05d}"
        
        # Use category_data from the request (or override from review data)
        request_category_data = request.get("category_data", {})
        review_category_data = data.get("category_data")
        final_category_data = review_category_data if review_category_data else request_category_data
        
        # Generate or use the requested name based on nomenclature
        rm_name = generate_rm_name(category, final_category_data)
        if not rm_name:
            # Fallback to requested_name if nomenclature-based name couldn't be generated
            rm_name = request.get("requested_name", "")
        final_category_data["name"] = rm_name
        
        # Create the RM
        rm = RawMaterial(
            rm_id=rm_id,
            category=category,
            category_data=final_category_data,
            brand_ids=request.get("brand_ids", []),
            vertical_ids=request.get("vertical_ids", []),
            model_ids=request.get("model_ids", []),
            is_brand_specific=True
        )
        
        rm_doc = rm.model_dump()
        rm_doc["created_at"] = rm_doc["created_at"].isoformat()
        await db.raw_materials.insert_one(rm_doc)
        
        update_data["status"] = "APPROVED"
        update_data["created_rm_id"] = rm_id
    
    await db.rm_requests.update_one(
        {"id": request_id},
        {"$set": update_data}
    )
    
    return {
        "message": f"Request {'approved' if action == 'APPROVE' else 'rejected'}",
        "created_rm_id": update_data.get("created_rm_id")
    }


@router.get("/rm-requests/pending-count")
async def get_pending_rm_requests_count():
    """Get count of pending RM requests (for notifications)"""
    count = await db.rm_requests.count_documents({"status": "PENDING"})
    return {"pending_count": count}


# =============================================================================
# RM DATA MIGRATION (Export/Import for Preview → Production)
# =============================================================================

@router.get("/raw-materials/migrate/export")
async def export_all_rm_data():
    """
    Export ALL raw materials as JSON for migration to production.
    Downloads complete RM database for environment transfer.
    """
    import json
    from fastapi.responses import StreamingResponse
    
    # Fetch all RMs
    rms = await db.raw_materials.find({}, {"_id": 0}).to_list(20000)
    
    # Also fetch branch inventory data
    branch_inventory = await db.branch_rm_inventory.find({}, {"_id": 0}).to_list(50000)
    
    export_data = {
        "export_date": datetime.now(timezone.utc).isoformat(),
        "raw_materials_count": len(rms),
        "branch_inventory_count": len(branch_inventory),
        "raw_materials": rms,
        "branch_rm_inventory": branch_inventory
    }
    
    # Convert to JSON
    json_str = json.dumps(export_data, indent=2, default=str)
    buffer = io.BytesIO(json_str.encode('utf-8'))
    buffer.seek(0)
    
    filename = f"rm_migration_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return StreamingResponse(
        buffer,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/raw-materials/migrate/import")
async def import_rm_data(file: UploadFile = File(...)):
    """
    Import raw materials from JSON export file.
    Use this in production to restore data from preview export.
    Handles duplicates by skipping existing RM IDs.
    """
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Read file content
    try:
        content = await file.read()
        logger.info(f"Migration import: Read {len(content)} bytes from {file.filename}")
    except Exception as e:
        logger.error(f"Migration import: Failed to read file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")
    
    # Parse JSON
    try:
        data = json.loads(content.decode('utf-8'))
        logger.info("Migration import: Parsed JSON successfully")
    except json.JSONDecodeError as e:
        logger.error(f"Migration import: Invalid JSON: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid JSON file: {str(e)}")
    except UnicodeDecodeError as e:
        logger.error(f"Migration import: Unicode decode error: {e}")
        raise HTTPException(status_code=400, detail=f"File encoding error: {str(e)}")
    
    results = {
        "raw_materials": {"imported": 0, "skipped": 0, "errors": []},
        "branch_inventory": {"imported": 0, "skipped": 0, "errors": []}
    }
    
    # Import raw materials
    rms = data.get("raw_materials", [])
    logger.info(f"Migration import: Found {len(rms)} raw materials in file")
    
    if not rms:
        return {
            "success": False,
            "message": "No raw_materials found in the JSON file. Ensure you're uploading a valid migration export.",
            "results": results,
            "totals": {"raw_materials": 0, "branch_inventory": 0}
        }
    
    # Get existing RM IDs
    existing_rm_ids = set()
    async for doc in db.raw_materials.find({}, {"rm_id": 1}):
        existing_rm_ids.add(doc["rm_id"])
    
    logger.info(f"Migration import: {len(existing_rm_ids)} existing RMs in database")
    
    new_rms = []
    for rm in rms:
        rm_id = rm.get("rm_id")
        if not rm_id:
            results["raw_materials"]["errors"].append("RM missing rm_id field")
            continue
        if rm_id in existing_rm_ids:
            results["raw_materials"]["skipped"] += 1
        else:
            # Clean up any ObjectId or datetime issues
            if "_id" in rm:
                del rm["_id"]
            new_rms.append(rm)
    
    # Insert new RMs in batches to avoid timeout
    if new_rms:
        try:
            batch_size = 500
            for i in range(0, len(new_rms), batch_size):
                batch = new_rms[i:i+batch_size]
                await db.raw_materials.insert_many(batch)
                logger.info(f"Migration import: Inserted batch {i//batch_size + 1}")
            results["raw_materials"]["imported"] = len(new_rms)
        except Exception as e:
            logger.error(f"Migration import: Failed to insert RMs: {e}")
            results["raw_materials"]["errors"].append(f"Insert failed: {str(e)}")
    
    # Import branch inventory
    branch_inv = data.get("branch_rm_inventory", [])
    logger.info(f"Migration import: Found {len(branch_inv)} branch inventory records")
    
    if branch_inv:
        existing_keys = set()
        async for doc in db.branch_rm_inventory.find({}, {"rm_id": 1, "branch": 1}):
            existing_keys.add(f"{doc['rm_id']}_{doc['branch']}")
        
        new_inv = []
        for inv in branch_inv:
            if "_id" in inv:
                del inv["_id"]
            key = f"{inv.get('rm_id')}_{inv.get('branch')}"
            if key in existing_keys:
                results["branch_inventory"]["skipped"] += 1
            else:
                new_inv.append(inv)
        
        if new_inv:
            try:
                batch_size = 500
                for i in range(0, len(new_inv), batch_size):
                    batch = new_inv[i:i+batch_size]
                    await db.branch_rm_inventory.insert_many(batch)
                results["branch_inventory"]["imported"] = len(new_inv)
            except Exception as e:
                logger.error(f"Migration import: Failed to insert inventory: {e}")
                results["branch_inventory"]["errors"].append(f"Insert failed: {str(e)}")
    
    # Get final counts
    total_rms = await db.raw_materials.count_documents({})
    total_inv = await db.branch_rm_inventory.count_documents({})
    
    logger.info(f"Migration import complete: {results['raw_materials']['imported']} RMs imported")
    
    return {
        "success": True,
        "message": f"Imported {results['raw_materials']['imported']} RMs, skipped {results['raw_materials']['skipped']} duplicates",
        "results": results,
        "totals": {
            "raw_materials": total_rms,
            "branch_inventory": total_inv
        }
    }
