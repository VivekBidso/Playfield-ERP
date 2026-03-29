"""
Pantone Shade Management Routes
- Pantone shade CRUD
- Vendor-Master Batch mapping
- QC approval workflow
- Color development requests
"""

from fastapi import APIRouter, Depends, HTTPException, Query, Body, File, UploadFile
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from typing import List, Optional
import uuid
import logging
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from models.auth import User
from models.pantone import (
    PantoneShadeCreate, PantoneShadeUpdate, PantoneShade,
    VendorMasterbatchCreate, VendorMasterbatchUpdate, VendorMasterbatch,
    ColorDevelopmentRequestCreate, ColorDevelopmentRequest,
    PantoneStatus, ApprovalStatus, ColorFamily
)
from services.auth_service import get_current_user
from database import db

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/pantone", tags=["Pantone Management"])


# ============ Pantone Shades ============

@router.get("/shades")
async def get_pantone_shades(
    category: Optional[str] = None,
    color_family: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    skip: int = Query(0, ge=0),
    current_user: User = Depends(get_current_user)
):
    """Get all Pantone shades with optional filters"""
    query = {}
    
    if category:
        query["applicable_categories"] = category
    if color_family:
        query["color_family"] = color_family
    if status:
        query["status"] = status
    else:
        query["status"] = {"$ne": "DEPRECATED"}
    if search:
        query["$or"] = [
            {"pantone_code": {"$regex": search, "$options": "i"}},
            {"pantone_name": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.pantone_shades.count_documents(query)
    
    shades = await db.pantone_shades.find(
        query,
        {"_id": 0}
    ).sort("pantone_code", 1).skip(skip).limit(limit).to_list(limit)
    
    # Get vendor counts for each shade
    for shade in shades:
        vendor_count = await db.pantone_vendor_masterbatch.count_documents({
            "pantone_id": shade["id"],
            "approval_status": "APPROVED"
        })
        shade["approved_vendor_count"] = vendor_count
    
    return {
        "items": shades,
        "total": total,
        "limit": limit,
        "skip": skip
    }


# NOTE: Static routes must be defined BEFORE dynamic routes like /shades/{shade_id}
@router.get("/shades/export")
async def export_pantone_shades(
    current_user: User = Depends(get_current_user)
):
    """Export all Pantone data to Excel"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel support not available")
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Pantone Shades
    ws1 = wb.active
    ws1.title = "Pantone_Shades"
    headers1 = ["pantone_code", "pantone_name", "color_hex", "color_family", "categories", "notes", "status"]
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    shades = await db.pantone_shades.find({}, {"_id": 0}).to_list(1000)
    for row, shade in enumerate(shades, 2):
        ws1.cell(row=row, column=1, value=shade.get("pantone_code"))
        ws1.cell(row=row, column=2, value=shade.get("pantone_name"))
        ws1.cell(row=row, column=3, value=shade.get("color_hex"))
        ws1.cell(row=row, column=4, value=shade.get("color_family"))
        ws1.cell(row=row, column=5, value=",".join(shade.get("applicable_categories", [])))
        ws1.cell(row=row, column=6, value=shade.get("notes"))
        ws1.cell(row=row, column=7, value=shade.get("status"))
    
    # Sheet 2: Vendor Mapping
    ws2 = wb.create_sheet("Vendor_Mapping")
    headers2 = ["pantone_code", "vendor_code", "master_batch_code", "is_preferred", "delta_e", "lead_time", "moq", "status"]
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    mappings = await db.pantone_vendor_masterbatch.find({}, {"_id": 0}).to_list(5000)
    
    # Get vendor codes
    vendors = await db.vendors.find({}, {"_id": 0, "id": 1, "code": 1}).to_list(1000)
    vendor_codes = {v["id"]: v.get("code", "") for v in vendors}
    
    for row, m in enumerate(mappings, 2):
        ws2.cell(row=row, column=1, value=m.get("pantone_code"))
        ws2.cell(row=row, column=2, value=vendor_codes.get(m.get("vendor_id"), ""))
        ws2.cell(row=row, column=3, value=m.get("master_batch_code"))
        ws2.cell(row=row, column=4, value=str(m.get("is_preferred", False)))
        ws2.cell(row=row, column=5, value=m.get("delta_e_value"))
        ws2.cell(row=row, column=6, value=m.get("lead_time_days"))
        ws2.cell(row=row, column=7, value=m.get("moq"))
        ws2.cell(row=row, column=8, value=m.get("approval_status"))
    
    # Sheet 3: RM Mapping
    ws3 = wb.create_sheet("RM_Mapping")
    headers3 = ["rm_id", "pantone_code", "rm_name", "category"]
    
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    rms = await db.raw_materials.find(
        {"pantone_id": {"$exists": True, "$ne": None}},
        {"_id": 0, "rm_id": 1, "pantone_code": 1, "name": 1, "category": 1}
    ).to_list(5000)
    
    for row, rm in enumerate(rms, 2):
        ws3.cell(row=row, column=1, value=rm.get("rm_id"))
        ws3.cell(row=row, column=2, value=rm.get("pantone_code"))
        ws3.cell(row=row, column=3, value=rm.get("name"))
        ws3.cell(row=row, column=4, value=rm.get("category"))
    
    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Pantone_Export.xlsx"}
    )


@router.get("/shades/download-template")
async def download_import_template(
    current_user: User = Depends(get_current_user)
):
    """Download blank import template"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel support not available")
    
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Sheet 1
    ws1 = wb.active
    ws1.title = "Pantone_Shades"
    headers1 = ["pantone_code", "pantone_name", "color_hex", "color_family", "categories", "notes"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    # Sample row
    ws1.cell(row=2, column=1, value="485 C")
    ws1.cell(row=2, column=2, value="Bright Red")
    ws1.cell(row=2, column=3, value="#DA291C")
    ws1.cell(row=2, column=4, value="RED")
    ws1.cell(row=2, column=5, value="INP,INM,ACC")
    ws1.cell(row=2, column=6, value="Primary red")
    
    # Sheet 2
    ws2 = wb.create_sheet("Vendor_Mapping")
    headers2 = ["pantone_code", "vendor_code", "master_batch_code", "is_preferred", "delta_e", "lead_time_days", "moq"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    ws2.cell(row=2, column=1, value="485 C")
    ws2.cell(row=2, column=2, value="VND001")
    ws2.cell(row=2, column=3, value="CT-RED-485")
    ws2.cell(row=2, column=4, value="TRUE")
    ws2.cell(row=2, column=5, value="0.8")
    ws2.cell(row=2, column=6, value="14")
    ws2.cell(row=2, column=7, value="100")
    
    # Sheet 3
    ws3 = wb.create_sheet("RM_Mapping")
    headers3 = ["rm_id", "pantone_code"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    ws3.cell(row=2, column=1, value="INP_001")
    ws3.cell(row=2, column=2, value="485 C")
    
    # Instructions
    ws4 = wb.create_sheet("Instructions")
    instructions = [
        ["Pantone Import Template Instructions"],
        [""],
        ["Sheet: Pantone_Shades"],
        ["- pantone_code: Official Pantone code (e.g., '485 C')"],
        ["- pantone_name: Descriptive name"],
        ["- color_hex: Hex color code with # (e.g., '#DA291C')"],
        ["- color_family: RED, BLUE, GREEN, YELLOW, ORANGE, PURPLE, PINK, BROWN, BLACK, WHITE, GREY, METALLIC, OTHER"],
        ["- categories: Comma-separated list (INP, INM, ACC)"],
        [""],
        ["Sheet: Vendor_Mapping"],
        ["- pantone_code: Must match a code from Pantone_Shades"],
        ["- vendor_code: Vendor code from your system"],
        ["- master_batch_code: Vendor's master batch code"],
        ["- is_preferred: TRUE or FALSE (only one preferred per Pantone)"],
        ["- delta_e: Color difference measurement (optional)"],
        [""],
        ["Sheet: RM_Mapping"],
        ["- rm_id: Raw Material ID (must exist)"],
        ["- pantone_code: Pantone code to assign"],
    ]
    for row, data in enumerate(instructions, 1):
        ws4.cell(row=row, column=1, value=data[0] if data else "")
    ws4.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Pantone_Import_Template.xlsx"}
    )


@router.get("/shades/{shade_id}")
async def get_pantone_shade(
    shade_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a single Pantone shade with vendor mappings"""
    shade = await db.pantone_shades.find_one({"id": shade_id}, {"_id": 0})
    if not shade:
        raise HTTPException(status_code=404, detail="Pantone shade not found")
    
    # Get vendor mappings
    vendors = await db.pantone_vendor_masterbatch.find(
        {"pantone_id": shade_id},
        {"_id": 0}
    ).sort([("is_preferred", -1), ("approval_status", 1)]).to_list(100)
    
    shade["vendor_mappings"] = vendors
    
    return shade


@router.post("/shades")
async def create_pantone_shade(
    data: PantoneShadeCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new Pantone shade"""
    # Check for duplicate
    existing = await db.pantone_shades.find_one({"pantone_code": data.pantone_code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Pantone code {data.pantone_code} already exists")
    
    now = datetime.now(timezone.utc)
    
    shade = {
        "id": str(uuid.uuid4()),
        "pantone_code": data.pantone_code.upper().strip(),
        "pantone_name": data.pantone_name.strip(),
        "color_hex": data.color_hex.upper() if data.color_hex.startswith("#") else f"#{data.color_hex.upper()}",
        "color_family": data.color_family.value if hasattr(data.color_family, 'value') else data.color_family,
        "applicable_categories": data.applicable_categories,
        "status": "ACTIVE",
        "notes": data.notes,
        "created_by": current_user.id,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.pantone_shades.insert_one(shade)
    
    logger.info(f"Created Pantone shade: {shade['pantone_code']} by {current_user.email}")
    
    return {k: v for k, v in shade.items() if k != "_id"}


@router.put("/shades/{shade_id}")
async def update_pantone_shade(
    shade_id: str,
    data: PantoneShadeUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update a Pantone shade"""
    shade = await db.pantone_shades.find_one({"id": shade_id})
    if not shade:
        raise HTTPException(status_code=404, detail="Pantone shade not found")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    if "color_family" in update_data and hasattr(update_data["color_family"], 'value'):
        update_data["color_family"] = update_data["color_family"].value
    if "status" in update_data and hasattr(update_data["status"], 'value'):
        update_data["status"] = update_data["status"].value
    
    if update_data:
        update_data["updated_by"] = current_user.id
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.pantone_shades.update_one(
            {"id": shade_id},
            {"$set": update_data}
        )
    
    updated = await db.pantone_shades.find_one({"id": shade_id}, {"_id": 0})
    return updated


@router.delete("/shades/{shade_id}")
async def delete_pantone_shade(
    shade_id: str,
    current_user: User = Depends(get_current_user)
):
    """Soft delete (deprecate) a Pantone shade"""
    shade = await db.pantone_shades.find_one({"id": shade_id})
    if not shade:
        raise HTTPException(status_code=404, detail="Pantone shade not found")
    
    # Check if used in any RMs
    rm_count = await db.raw_materials.count_documents({"pantone_id": shade_id})
    if rm_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete: Pantone is used by {rm_count} raw materials. Set to DEPRECATED instead."
        )
    
    await db.pantone_shades.update_one(
        {"id": shade_id},
        {"$set": {
            "status": "DEPRECATED",
            "updated_by": current_user.id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Pantone shade deprecated"}


# ============ Vendor Master Batch Mapping ============

@router.get("/shades/{shade_id}/vendors")
async def get_shade_vendors(
    shade_id: str,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all vendor mappings for a Pantone shade"""
    query = {"pantone_id": shade_id}
    if status:
        query["approval_status"] = status
    
    vendors = await db.pantone_vendor_masterbatch.find(
        query,
        {"_id": 0}
    ).sort([("is_preferred", -1), ("vendor_name", 1)]).to_list(100)
    
    return vendors


@router.post("/vendor-masterbatch")
async def create_vendor_masterbatch(
    data: VendorMasterbatchCreate,
    current_user: User = Depends(get_current_user)
):
    """Add a vendor master batch mapping to a Pantone shade"""
    # Verify Pantone exists
    pantone = await db.pantone_shades.find_one({"id": data.pantone_id}, {"_id": 0})
    if not pantone:
        raise HTTPException(status_code=404, detail="Pantone shade not found")
    
    # Verify vendor exists
    vendor = await db.vendors.find_one({"id": data.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Check for duplicate
    existing = await db.pantone_vendor_masterbatch.find_one({
        "pantone_id": data.pantone_id,
        "vendor_id": data.vendor_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="This vendor already has a mapping for this Pantone")
    
    now = datetime.now(timezone.utc)
    
    mapping = {
        "id": str(uuid.uuid4()),
        "pantone_id": data.pantone_id,
        "pantone_code": pantone["pantone_code"],
        "vendor_id": data.vendor_id,
        "vendor_name": vendor.get("name", "Unknown"),
        "master_batch_code": data.master_batch_code.strip(),
        
        "approval_status": "PENDING",
        "submitted_by": current_user.id,
        "submitted_at": now.isoformat(),
        
        "delta_e_value": data.delta_e_value,
        "lab_report_url": data.lab_report_url,
        "sample_batch_number": data.sample_batch_number,
        
        "is_preferred": False,
        "is_active": True,
        "lead_time_days": data.lead_time_days or 14,
        "moq": data.moq or 100,
        "batch_size": data.batch_size or 25,
        
        "notes": data.notes,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.pantone_vendor_masterbatch.insert_one(mapping)
    
    logger.info(f"Created vendor masterbatch mapping: {pantone['pantone_code']} -> {vendor.get('name')} ({data.master_batch_code})")
    
    return {k: v for k, v in mapping.items() if k != "_id"}


@router.put("/vendor-masterbatch/{mapping_id}")
async def update_vendor_masterbatch(
    mapping_id: str,
    data: VendorMasterbatchUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update a vendor master batch mapping"""
    mapping = await db.pantone_vendor_masterbatch.find_one({"id": mapping_id})
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.pantone_vendor_masterbatch.update_one(
            {"id": mapping_id},
            {"$set": update_data}
        )
    
    updated = await db.pantone_vendor_masterbatch.find_one({"id": mapping_id}, {"_id": 0})
    return updated


@router.put("/vendor-masterbatch/{mapping_id}/approve")
async def approve_vendor_masterbatch(
    mapping_id: str,
    delta_e_value: Optional[float] = Body(None),
    notes: Optional[str] = Body(None),
    current_user: User = Depends(get_current_user)
):
    """QC approve a vendor master batch (requires quality_inspector or master_admin role)"""
    if current_user.role not in ["master_admin", "quality_inspector"]:
        raise HTTPException(status_code=403, detail="Only QC team can approve master batches")
    
    mapping = await db.pantone_vendor_masterbatch.find_one({"id": mapping_id})
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    
    now = datetime.now(timezone.utc)
    
    update_data = {
        "approval_status": "APPROVED",
        "reviewed_by": current_user.id,
        "reviewed_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    if delta_e_value is not None:
        update_data["delta_e_value"] = delta_e_value
    if notes:
        update_data["notes"] = notes
    
    await db.pantone_vendor_masterbatch.update_one(
        {"id": mapping_id},
        {"$set": update_data}
    )
    
    logger.info(f"Approved vendor masterbatch: {mapping['master_batch_code']} by {current_user.email}")
    
    return {"message": "Vendor master batch approved", "id": mapping_id}


@router.put("/vendor-masterbatch/{mapping_id}/reject")
async def reject_vendor_masterbatch(
    mapping_id: str,
    rejection_reason: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user)
):
    """QC reject a vendor master batch"""
    if current_user.role not in ["master_admin", "quality_inspector"]:
        raise HTTPException(status_code=403, detail="Only QC team can reject master batches")
    
    mapping = await db.pantone_vendor_masterbatch.find_one({"id": mapping_id})
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    
    now = datetime.now(timezone.utc)
    
    await db.pantone_vendor_masterbatch.update_one(
        {"id": mapping_id},
        {"$set": {
            "approval_status": "REJECTED",
            "reviewed_by": current_user.id,
            "reviewed_at": now.isoformat(),
            "rejection_reason": rejection_reason,
            "updated_at": now.isoformat()
        }}
    )
    
    logger.info(f"Rejected vendor masterbatch: {mapping['master_batch_code']} - Reason: {rejection_reason}")
    
    return {"message": "Vendor master batch rejected", "id": mapping_id}


@router.put("/vendor-masterbatch/{mapping_id}/set-preferred")
async def set_preferred_vendor(
    mapping_id: str,
    current_user: User = Depends(get_current_user)
):
    """Set a vendor as the preferred vendor for a Pantone shade"""
    mapping = await db.pantone_vendor_masterbatch.find_one({"id": mapping_id})
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    
    if mapping.get("approval_status") != "APPROVED":
        raise HTTPException(status_code=400, detail="Only approved vendors can be set as preferred")
    
    # Remove preferred from other vendors for this Pantone
    await db.pantone_vendor_masterbatch.update_many(
        {"pantone_id": mapping["pantone_id"]},
        {"$set": {"is_preferred": False}}
    )
    
    # Set this one as preferred
    await db.pantone_vendor_masterbatch.update_one(
        {"id": mapping_id},
        {"$set": {
            "is_preferred": True,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Vendor set as preferred", "id": mapping_id}


@router.get("/vendor-masterbatch/pending")
async def get_pending_approvals(
    current_user: User = Depends(get_current_user)
):
    """Get all pending master batch approvals (for QC dashboard)"""
    pending = await db.pantone_vendor_masterbatch.find(
        {"approval_status": "PENDING"},
        {"_id": 0}
    ).sort("submitted_at", 1).to_list(100)
    
    return pending


@router.delete("/vendor-masterbatch/{mapping_id}")
async def delete_vendor_masterbatch(
    mapping_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a vendor master batch mapping"""
    mapping = await db.pantone_vendor_masterbatch.find_one({"id": mapping_id})
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    
    # Check if used in any POs
    po_count = await db.purchase_order_lines.count_documents({
        "vendor_masterbatch_id": mapping_id
    })
    if po_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete: Used in {po_count} purchase orders. Set to DEPRECATED instead."
        )
    
    await db.pantone_vendor_masterbatch.delete_one({"id": mapping_id})
    
    return {"message": "Mapping deleted"}


# ============ Color Development Requests ============

@router.get("/color-requests")
async def get_color_requests(
    status: Optional[str] = None,
    my_requests: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get color development requests"""
    query = {}
    if status:
        query["status"] = status
    if my_requests:
        query["requested_by"] = current_user.id
    
    requests = await db.color_development_requests.find(
        query,
        {"_id": 0}
    ).sort("requested_at", -1).to_list(100)
    
    return requests


@router.post("/color-requests")
async def create_color_request(
    data: ColorDevelopmentRequestCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new color development request (Design Team / Demand Planner)"""
    # Check if Pantone already exists
    existing = await db.pantone_shades.find_one({"pantone_code": data.pantone_code.upper()})
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"Pantone {data.pantone_code} already exists in the system"
        )
    
    now = datetime.now(timezone.utc)
    
    request = {
        "id": str(uuid.uuid4()),
        "pantone_code": data.pantone_code.upper().strip(),
        "pantone_name": data.pantone_name.strip(),
        "color_hex": data.color_hex,
        "color_family": data.color_family.value if data.color_family and hasattr(data.color_family, 'value') else data.color_family,
        "applicable_categories": data.applicable_categories,
        "target_models": data.target_models,
        "priority": data.priority,
        "status": "REQUESTED",
        "notes": data.notes,
        "requested_by": current_user.id,
        "requested_by_name": current_user.name or current_user.email,
        "requested_at": now.isoformat(),
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.color_development_requests.insert_one(request)
    
    logger.info(f"Color development request created: {data.pantone_code} by {current_user.email}")
    
    return {k: v for k, v in request.items() if k != "_id"}


@router.put("/color-requests/{request_id}/status")
async def update_color_request_status(
    request_id: str,
    status: str = Body(..., embed=True),
    notes: Optional[str] = Body(None),
    current_user: User = Depends(get_current_user)
):
    """Update color request status"""
    valid_statuses = ["REQUESTED", "VENDOR_DEVELOPMENT", "QC_PENDING", "APPROVED", "REJECTED"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    request = await db.color_development_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    update_data = {
        "status": status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if notes:
        update_data["notes"] = notes
    
    await db.color_development_requests.update_one(
        {"id": request_id},
        {"$set": update_data}
    )
    
    return {"message": f"Status updated to {status}"}


# ============ Bulk Import ============

@router.post("/shades/bulk-import")
async def bulk_import_pantone(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Bulk import Pantone shades and vendor mappings from Excel"""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel support not available")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    results = {
        "shades_created": 0,
        "shades_updated": 0,
        "vendor_mappings_created": 0,
        "rm_mappings_updated": 0,
        "errors": []
    }
    
    now = datetime.now(timezone.utc)
    
    # Sheet 1: Pantone Shades
    if "Pantone_Shades" in wb.sheetnames or "Sheet1" in wb.sheetnames:
        ws = wb["Pantone_Shades"] if "Pantone_Shades" in wb.sheetnames else wb["Sheet1"]
        
        for row_num in range(2, ws.max_row + 1):
            pantone_code = ws.cell(row=row_num, column=1).value
            if not pantone_code:
                continue
            
            pantone_code = str(pantone_code).upper().strip()
            pantone_name = ws.cell(row=row_num, column=2).value or pantone_code
            color_hex = ws.cell(row=row_num, column=3).value or "#808080"
            color_family = ws.cell(row=row_num, column=4).value or "OTHER"
            categories = ws.cell(row=row_num, column=5).value or "INP,INM,ACC"
            notes = ws.cell(row=row_num, column=6).value
            
            # Parse categories
            if isinstance(categories, str):
                categories = [c.strip().upper() for c in categories.split(",")]
            
            # Check if exists
            existing = await db.pantone_shades.find_one({"pantone_code": pantone_code})
            
            if existing:
                # Update
                await db.pantone_shades.update_one(
                    {"pantone_code": pantone_code},
                    {"$set": {
                        "pantone_name": pantone_name,
                        "color_hex": color_hex if color_hex.startswith("#") else f"#{color_hex}",
                        "color_family": color_family.upper(),
                        "applicable_categories": categories,
                        "notes": notes,
                        "updated_at": now.isoformat()
                    }}
                )
                results["shades_updated"] += 1
            else:
                # Create
                shade = {
                    "id": str(uuid.uuid4()),
                    "pantone_code": pantone_code,
                    "pantone_name": pantone_name,
                    "color_hex": color_hex if color_hex.startswith("#") else f"#{color_hex}",
                    "color_family": color_family.upper(),
                    "applicable_categories": categories,
                    "status": "ACTIVE",
                    "notes": notes,
                    "created_by": current_user.id,
                    "created_at": now.isoformat(),
                    "updated_at": now.isoformat()
                }
                await db.pantone_shades.insert_one(shade)
                results["shades_created"] += 1
    
    # Sheet 2: Vendor Master Batch Mapping
    if "Vendor_Mapping" in wb.sheetnames:
        ws = wb["Vendor_Mapping"]
        
        # Get vendor lookup
        vendors = await db.vendors.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1}).to_list(1000)
        vendor_by_code = {v.get("code", "").upper(): v for v in vendors if v.get("code")}
        
        for row_num in range(2, ws.max_row + 1):
            pantone_code = ws.cell(row=row_num, column=1).value
            vendor_code = ws.cell(row=row_num, column=2).value
            master_batch = ws.cell(row=row_num, column=3).value
            
            if not pantone_code or not vendor_code or not master_batch:
                continue
            
            pantone_code = str(pantone_code).upper().strip()
            vendor_code = str(vendor_code).upper().strip()
            
            # Find pantone
            pantone = await db.pantone_shades.find_one({"pantone_code": pantone_code})
            if not pantone:
                results["errors"].append(f"Row {row_num}: Pantone {pantone_code} not found")
                continue
            
            # Find vendor
            vendor = vendor_by_code.get(vendor_code)
            if not vendor:
                results["errors"].append(f"Row {row_num}: Vendor {vendor_code} not found")
                continue
            
            # Check if mapping exists
            existing = await db.pantone_vendor_masterbatch.find_one({
                "pantone_id": pantone["id"],
                "vendor_id": vendor["id"]
            })
            
            if not existing:
                is_preferred = str(ws.cell(row=row_num, column=4).value or "").upper() == "TRUE"
                delta_e = ws.cell(row=row_num, column=5).value
                lead_time = ws.cell(row=row_num, column=6).value or 14
                moq = ws.cell(row=row_num, column=7).value or 100
                
                mapping = {
                    "id": str(uuid.uuid4()),
                    "pantone_id": pantone["id"],
                    "pantone_code": pantone_code,
                    "vendor_id": vendor["id"],
                    "vendor_name": vendor.get("name", ""),
                    "master_batch_code": str(master_batch).strip(),
                    "approval_status": "APPROVED",  # Bulk import = pre-approved
                    "submitted_by": current_user.id,
                    "submitted_at": now.isoformat(),
                    "reviewed_by": current_user.id,
                    "reviewed_at": now.isoformat(),
                    "delta_e_value": float(delta_e) if delta_e else None,
                    "is_preferred": is_preferred,
                    "is_active": True,
                    "lead_time_days": int(lead_time),
                    "moq": int(moq),
                    "batch_size": 25,
                    "created_at": now.isoformat(),
                    "updated_at": now.isoformat()
                }
                await db.pantone_vendor_masterbatch.insert_one(mapping)
                results["vendor_mappings_created"] += 1
    
    # Sheet 3: RM to Pantone Mapping
    if "RM_Mapping" in wb.sheetnames:
        ws = wb["RM_Mapping"]
        
        for row_num in range(2, ws.max_row + 1):
            rm_id = ws.cell(row=row_num, column=1).value
            pantone_code = ws.cell(row=row_num, column=2).value
            
            if not rm_id or not pantone_code:
                continue
            
            rm_id = str(rm_id).upper().strip()
            pantone_code = str(pantone_code).upper().strip()
            
            # Find pantone
            pantone = await db.pantone_shades.find_one({"pantone_code": pantone_code})
            if not pantone:
                results["errors"].append(f"Row {row_num}: Pantone {pantone_code} not found for RM {rm_id}")
                continue
            
            # Update RM
            result = await db.raw_materials.update_one(
                {"rm_id": rm_id},
                {"$set": {
                    "pantone_id": pantone["id"],
                    "pantone_code": pantone_code,
                    "updated_at": now.isoformat()
                }}
            )
            
            if result.modified_count > 0:
                results["rm_mappings_updated"] += 1
    
    logger.info(f"Bulk import completed: {results}")
    
    return results


# ============ Integration Helpers ============

@router.get("/shades/{shade_id}/approved-vendors")
async def get_approved_vendors_for_shade(
    shade_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get approved vendors for a Pantone shade (for PO generation)"""
    vendors = await db.pantone_vendor_masterbatch.find(
        {
            "pantone_id": shade_id,
            "approval_status": "APPROVED",
            "is_active": True
        },
        {"_id": 0}
    ).sort([("is_preferred", -1), ("vendor_name", 1)]).to_list(50)
    
    return vendors


@router.get("/by-category/{category}")
async def get_pantone_by_category(
    category: str,
    current_user: User = Depends(get_current_user)
):
    """Get all active Pantone shades for a category"""
    shades = await db.pantone_shades.find(
        {
            "applicable_categories": category.upper(),
            "status": "ACTIVE"
        },
        {"_id": 0}
    ).sort("pantone_code", 1).to_list(500)
    
    return shades
