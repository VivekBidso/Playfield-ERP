"""
Dispatch Lots V2 - Two-Stage Workflow

Stage 1: Demand Team creates simple dispatch lots (Customer + SKU + Qty)
Stage 2: Finance Team converts to full invoice with billing details

Maintained by: Demand Team (lots), Finance Team (invoices)
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query
from pydantic import BaseModel
from datetime import datetime, timezone
from typing import Optional, List
import uuid
import io
import openpyxl

from database import db
from services.utils import get_current_user, serialize_doc

router = APIRouter(prefix="/dispatch-lots-v2", tags=["Dispatch Lots V2"])


# ============ Pydantic Models ============

class DispatchLineSimple(BaseModel):
    """Simple line item for demand team"""
    buyer_sku_id: str
    quantity: int


class DispatchLotCreateSimple(BaseModel):
    """Create dispatch lot - Demand Team (simple)"""
    customer_id: str
    lines: List[DispatchLineSimple]
    notes: Optional[str] = None


class InvoiceLineItem(BaseModel):
    """Full line item for finance invoice"""
    buyer_sku_id: str
    sku_name: Optional[str] = None
    quantity: int
    rate: float
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = None
    tax_amount: Optional[float] = None
    amount: float


class InvoiceTotals(BaseModel):
    """Invoice totals"""
    sub_total: float
    discount_type: str = "percentage"  # "percentage" or "amount"
    discount_value: float = 0
    discount_amount: float = 0
    tds_tcs_type: Optional[str] = None  # "TDS" or "TCS" or None
    tds_tcs_rate: Optional[float] = None
    tds_tcs_amount: float = 0
    adjustment: float = 0
    grand_total: float


class CreateInvoiceRequest(BaseModel):
    """Create invoice from lot - Finance Team"""
    branch_id: str
    source_of_supply: str  # State name
    order_number: Optional[str] = None
    invoice_date: str  # ISO date
    payment_terms: str = "NET_30"
    due_date: str  # ISO date
    accounts_receivable: str = "Accounts Receivable"
    salesperson: Optional[str] = None
    subject: Optional[str] = None
    line_items: List[InvoiceLineItem]
    totals: InvoiceTotals
    customer_notes: Optional[str] = None
    terms_conditions: Optional[str] = None


# ============ Status Constants ============

LOT_STATUS = {
    "DRAFT": "Created by demand, not yet sent",
    "PENDING_FINANCE": "Sent to finance for invoicing",
    "INVOICED": "Invoice created by finance",
    "DISPATCHED": "Shipment completed",
    "CANCELLED": "Cancelled"
}


# ============ Helper Functions ============

async def get_next_lot_number():
    """Generate next lot number: DL-YYYY-NNNN"""
    year = datetime.now().year
    prefix = f"DL-{year}-"
    
    # Find last lot number for this year
    last_lot = await db.dispatch_lots_v2.find_one(
        {"lot_number": {"$regex": f"^{prefix}"}},
        sort=[("lot_number", -1)]
    )
    
    if last_lot:
        last_num = int(last_lot["lot_number"].split("-")[-1])
        next_num = last_num + 1
    else:
        next_num = 1
    
    return f"{prefix}{next_num:04d}"


async def check_branch_inventory(branch_id: str, lines: List[dict]):
    """Check if branch has sufficient inventory for all line items"""
    results = []
    can_proceed = True
    
    for line in lines:
        buyer_sku_id = line.get("buyer_sku_id")
        required_qty = line.get("quantity", 0)
        
        # Get the bidso_sku_id from buyer_sku
        buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
        if not buyer_sku:
            results.append({
                "buyer_sku_id": buyer_sku_id,
                "required": required_qty,
                "available": 0,
                "sufficient": False,
                "error": "SKU not found"
            })
            can_proceed = False
            continue
        
        bidso_sku_id = buyer_sku.get("bidso_sku_id")
        
        # Check branch inventory
        inventory = await db.branch_inventory.find_one({
            "branch_id": branch_id,
            "sku_id": bidso_sku_id
        })
        
        available = inventory.get("quantity", 0) if inventory else 0
        sufficient = available >= required_qty
        
        if not sufficient:
            can_proceed = False
        
        results.append({
            "buyer_sku_id": buyer_sku_id,
            "bidso_sku_id": bidso_sku_id,
            "required": required_qty,
            "available": available,
            "sufficient": sufficient
        })
    
    return {"can_proceed": can_proceed, "items": results}


# ============ Demand Team Endpoints ============

@router.get("")
async def list_dispatch_lots(
    status: Optional[str] = None,
    customer_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user = Depends(get_current_user)
):
    """List all dispatch lots with filters"""
    query = {}
    
    if status:
        query["status"] = status
    if customer_id:
        query["customer_id"] = customer_id
    
    total = await db.dispatch_lots_v2.count_documents(query)
    skip = (page - 1) * page_size
    
    cursor = db.dispatch_lots_v2.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size)
    lots = await cursor.to_list(length=page_size)
    
    return {
        "lots": lots,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/summary")
async def get_lots_summary(current_user = Depends(get_current_user)):
    """Get summary counts by status"""
    pipeline = [
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    
    results = await db.dispatch_lots_v2.aggregate(pipeline).to_list(100)
    
    summary = {s: 0 for s in LOT_STATUS.keys()}
    for r in results:
        if r["_id"] in summary:
            summary[r["_id"]] = r["count"]
    
    summary["total"] = sum(summary.values())
    
    return summary


@router.post("")
async def create_dispatch_lot(
    data: DispatchLotCreateSimple,
    current_user = Depends(get_current_user)
):
    """Create a simple dispatch lot - Demand Team"""
    # Validate customer
    customer = await db.buyers.find_one({
        "$or": [
            {"customer_code": data.customer_id},
            {"id": data.customer_id}
        ]
    })
    if not customer:
        raise HTTPException(status_code=400, detail=f"Customer {data.customer_id} not found")
    
    # Validate SKUs and build lines
    validated_lines = []
    for line in data.lines:
        sku = await db.buyer_skus.find_one({"buyer_sku_id": line.buyer_sku_id})
        if not sku:
            raise HTTPException(status_code=400, detail=f"Buyer SKU {line.buyer_sku_id} not found")
        
        validated_lines.append({
            "buyer_sku_id": line.buyer_sku_id,
            "sku_name": sku.get("name", line.buyer_sku_id),
            "bidso_sku_id": sku.get("bidso_sku_id"),
            "quantity": line.quantity
        })
    
    # Generate lot number
    lot_number = await get_next_lot_number()
    lot_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    lot = {
        "id": lot_id,
        "lot_number": lot_number,
        "customer_id": data.customer_id,
        "customer_name": customer.get("name", ""),
        "status": "DRAFT",
        "lines": validated_lines,
        "total_quantity": sum(l.quantity for l in data.lines),
        "notes": data.notes,
        "created_by": current_user.id,
        "created_by_role": current_user.role,
        "created_at": now,
        "invoice_data": None,  # Filled by finance
        "invoiced_by": None,
        "invoiced_at": None,
        "dispatched_at": None
    }
    
    await db.dispatch_lots_v2.insert_one(lot)
    
    return {"message": "Dispatch lot created", "lot_number": lot_number, "id": lot_id}


@router.put("/{lot_id}")
async def update_dispatch_lot(
    lot_id: str,
    data: DispatchLotCreateSimple,
    current_user = Depends(get_current_user)
):
    """Update a draft dispatch lot - Demand Team"""
    lot = await db.dispatch_lots_v2.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["status"] != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only edit DRAFT lots")
    
    # Validate and build lines
    validated_lines = []
    for line in data.lines:
        sku = await db.buyer_skus.find_one({"buyer_sku_id": line.buyer_sku_id})
        if not sku:
            raise HTTPException(status_code=400, detail=f"Buyer SKU {line.buyer_sku_id} not found")
        
        validated_lines.append({
            "buyer_sku_id": line.buyer_sku_id,
            "sku_name": sku.get("name", line.buyer_sku_id),
            "bidso_sku_id": sku.get("bidso_sku_id"),
            "quantity": line.quantity
        })
    
    await db.dispatch_lots_v2.update_one(
        {"id": lot_id},
        {"$set": {
            "lines": validated_lines,
            "total_quantity": sum(l.quantity for l in data.lines),
            "notes": data.notes,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.id
        }}
    )
    
    return {"message": "Lot updated"}


@router.delete("/{lot_id}")
async def delete_dispatch_lot(
    lot_id: str,
    current_user = Depends(get_current_user)
):
    """Delete a draft dispatch lot"""
    lot = await db.dispatch_lots_v2.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["status"] != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only delete DRAFT lots")
    
    await db.dispatch_lots_v2.delete_one({"id": lot_id})
    
    return {"message": "Lot deleted"}


@router.post("/{lot_id}/send-to-finance")
async def send_to_finance(
    lot_id: str,
    current_user = Depends(get_current_user)
):
    """Send lot to finance for invoicing"""
    lot = await db.dispatch_lots_v2.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["status"] != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only send DRAFT lots")
    
    await db.dispatch_lots_v2.update_one(
        {"id": lot_id},
        {"$set": {
            "status": "PENDING_FINANCE",
            "sent_to_finance_at": datetime.now(timezone.utc).isoformat(),
            "sent_to_finance_by": current_user.id
        }}
    )
    
    return {"message": "Lot sent to finance"}


@router.post("/bulk-upload")
async def bulk_upload_lots(
    file: UploadFile = File(...),
    current_user = Depends(get_current_user)
):
    """
    Bulk upload dispatch lots from Excel.
    
    Expected columns: customer_id, buyer_sku_id, quantity
    Groups by customer_id to create multiple lots
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
    
    required = ["customer_id", "buyer_sku_id", "quantity"]
    for req in required:
        if req not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required column: {req}")
    
    # Group rows by customer
    customer_lines = {}
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        row_data = dict(zip(headers, row))
        customer_id = str(row_data.get("customer_id", "")).strip()
        buyer_sku_id = str(row_data.get("buyer_sku_id", "")).strip()
        quantity = row_data.get("quantity")
        
        if not customer_id or not buyer_sku_id or quantity is None:
            errors.append({"row": row_idx, "error": "Missing required fields"})
            continue
        
        try:
            quantity = int(quantity)
        except:
            errors.append({"row": row_idx, "error": "Invalid quantity"})
            continue
        
        if customer_id not in customer_lines:
            customer_lines[customer_id] = []
        
        customer_lines[customer_id].append({
            "buyer_sku_id": buyer_sku_id,
            "quantity": quantity
        })
    
    # Create lots for each customer
    created_lots = []
    now = datetime.now(timezone.utc).isoformat()
    
    for customer_id, lines in customer_lines.items():
        # Validate customer
        customer = await db.buyers.find_one({
            "$or": [
                {"customer_code": customer_id},
                {"id": customer_id}
            ]
        })
        if not customer:
            errors.append({"customer_id": customer_id, "error": "Customer not found"})
            continue
        
        # Validate SKUs
        validated_lines = []
        has_error = False
        
        for line in lines:
            sku = await db.buyer_skus.find_one({"buyer_sku_id": line["buyer_sku_id"]})
            if not sku:
                errors.append({"buyer_sku_id": line["buyer_sku_id"], "error": "SKU not found"})
                has_error = True
                continue
            
            validated_lines.append({
                "buyer_sku_id": line["buyer_sku_id"],
                "sku_name": sku.get("name", line["buyer_sku_id"]),
                "bidso_sku_id": sku.get("bidso_sku_id"),
                "quantity": line["quantity"]
            })
        
        if has_error or not validated_lines:
            continue
        
        # Create lot
        lot_number = await get_next_lot_number()
        lot_id = str(uuid.uuid4())
        
        lot = {
            "id": lot_id,
            "lot_number": lot_number,
            "customer_id": customer_id,
            "customer_name": customer.get("name", ""),
            "status": "DRAFT",
            "lines": validated_lines,
            "total_quantity": sum(l["quantity"] for l in validated_lines),
            "notes": "Bulk upload",
            "created_by": current_user.id,
            "created_by_role": current_user.role,
            "created_at": now,
            "invoice_data": None,
            "invoiced_by": None,
            "invoiced_at": None,
            "dispatched_at": None
        }
        
        await db.dispatch_lots_v2.insert_one(lot)
        created_lots.append({"lot_number": lot_number, "customer": customer.get("name")})
    
    return {
        "message": f"Created {len(created_lots)} dispatch lots",
        "lots": created_lots,
        "errors": errors
    }


# ============ Finance Team Endpoints ============

@router.get("/{lot_id}")
async def get_lot_details(
    lot_id: str,
    current_user = Depends(get_current_user)
):
    """Get full lot details including invoice data if available"""
    lot = await db.dispatch_lots_v2.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    # Enrich line items with current pricing and HSN/GST
    enriched_lines = []
    for line in lot.get("lines", []):
        enriched = {**line}
        
        # Get HSN/GST from SKU
        sku = await db.buyer_skus.find_one({"buyer_sku_id": line["buyer_sku_id"]})
        if sku:
            enriched["hsn_code"] = sku.get("hsn_code")
            enriched["gst_rate"] = sku.get("gst_rate")
        
        # Get price from price master
        price = await db.price_master.find_one({
            "customer_id": lot["customer_id"],
            "buyer_sku_id": line["buyer_sku_id"],
            "effective_to": None
        })
        if price:
            enriched["unit_price"] = price.get("unit_price")
        
        enriched_lines.append(enriched)
    
    lot["lines"] = enriched_lines
    
    return lot


@router.get("/{lot_id}/inventory-check")
async def check_lot_inventory(
    lot_id: str,
    branch_id: str,
    current_user = Depends(get_current_user)
):
    """Check if branch has sufficient inventory for a lot"""
    lot = await db.dispatch_lots_v2.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    result = await check_branch_inventory(branch_id, lot.get("lines", []))
    
    return result


@router.post("/{lot_id}/create-invoice")
async def create_invoice(
    lot_id: str,
    data: CreateInvoiceRequest,
    current_user = Depends(get_current_user)
):
    """Create invoice from lot - Finance Team"""
    lot = await db.dispatch_lots_v2.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["status"] not in ["DRAFT", "PENDING_FINANCE"]:
        raise HTTPException(status_code=400, detail="Lot already invoiced or dispatched")
    
    # Check inventory before allowing invoice
    inventory_check = await check_branch_inventory(data.branch_id, [
        {"buyer_sku_id": li.buyer_sku_id, "quantity": li.quantity}
        for li in data.line_items
    ])
    
    if not inventory_check["can_proceed"]:
        insufficient = [i for i in inventory_check["items"] if not i["sufficient"]]
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Insufficient inventory at selected branch",
                "items": insufficient
            }
        )
    
    # Get branch details
    branch = await db.branches.find_one({"branch_id": data.branch_id})
    if not branch:
        raise HTTPException(status_code=400, detail="Branch not found")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Build invoice data with manufacturing origin
    invoice_lines = []
    for li in data.line_items:
        line_data = {
            "buyer_sku_id": li.buyer_sku_id,
            "sku_name": li.sku_name,
            "quantity": li.quantity,
            "rate": li.rate,
            "hsn_code": li.hsn_code,
            "gst_rate": li.gst_rate,
            "tax_amount": li.tax_amount or (li.quantity * li.rate * (li.gst_rate or 0) / 100),
            "amount": li.amount,
            "manufacturing_origin": {
                "branch_id": data.branch_id,  # For now, same as dispatch branch
                "branch_name": branch.get("name")
            }
        }
        invoice_lines.append(line_data)
    
    invoice_data = {
        "branch_id": data.branch_id,
        "branch_name": branch.get("name"),
        "source_of_supply": data.source_of_supply,
        "invoice_number": None,  # To be assigned later
        "order_number": data.order_number,
        "invoice_date": data.invoice_date,
        "payment_terms": data.payment_terms,
        "due_date": data.due_date,
        "accounts_receivable": data.accounts_receivable,
        "salesperson": data.salesperson,
        "subject": data.subject,
        "line_items": invoice_lines,
        "totals": {
            "sub_total": data.totals.sub_total,
            "discount_type": data.totals.discount_type,
            "discount_value": data.totals.discount_value,
            "discount_amount": data.totals.discount_amount,
            "tds_tcs_type": data.totals.tds_tcs_type,
            "tds_tcs_rate": data.totals.tds_tcs_rate,
            "tds_tcs_amount": data.totals.tds_tcs_amount,
            "adjustment": data.totals.adjustment,
            "grand_total": data.totals.grand_total
        },
        "customer_notes": data.customer_notes,
        "terms_conditions": data.terms_conditions,
        "attachments": []
    }
    
    # Update lot with invoice data
    await db.dispatch_lots_v2.update_one(
        {"id": lot_id},
        {"$set": {
            "status": "INVOICED",
            "invoice_data": invoice_data,
            "invoiced_by": current_user.id,
            "invoiced_at": now
        }}
    )
    
    # Deduct inventory (simplified - real implementation would use FIFO)
    for li in data.line_items:
        buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": li.buyer_sku_id})
        if buyer_sku:
            await db.branch_inventory.update_one(
                {"branch_id": data.branch_id, "sku_id": buyer_sku.get("bidso_sku_id")},
                {"$inc": {"quantity": -li.quantity}}
            )
    
    return {"message": "Invoice created successfully", "lot_number": lot["lot_number"]}


@router.post("/{lot_id}/dispatch")
async def mark_dispatched(
    lot_id: str,
    current_user = Depends(get_current_user)
):
    """Mark lot as dispatched"""
    lot = await db.dispatch_lots_v2.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["status"] != "INVOICED":
        raise HTTPException(status_code=400, detail="Lot must be invoiced before dispatch")
    
    await db.dispatch_lots_v2.update_one(
        {"id": lot_id},
        {"$set": {
            "status": "DISPATCHED",
            "dispatched_at": datetime.now(timezone.utc).isoformat(),
            "dispatched_by": current_user.id
        }}
    )
    
    return {"message": "Lot marked as dispatched"}


@router.post("/new-invoice")
async def create_new_invoice(
    data: CreateInvoiceRequest,
    customer_id: str,
    current_user = Depends(get_current_user)
):
    """Create a new invoice directly without a pre-existing lot - Finance Team"""
    # Validate customer
    customer = await db.buyers.find_one({
        "$or": [
            {"customer_code": customer_id},
            {"id": customer_id}
        ]
    })
    if not customer:
        raise HTTPException(status_code=400, detail="Customer not found")
    
    # Check inventory
    inventory_check = await check_branch_inventory(data.branch_id, [
        {"buyer_sku_id": li.buyer_sku_id, "quantity": li.quantity}
        for li in data.line_items
    ])
    
    if not inventory_check["can_proceed"]:
        insufficient = [i for i in inventory_check["items"] if not i["sufficient"]]
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Insufficient inventory at selected branch",
                "items": insufficient
            }
        )
    
    # Get branch details
    branch = await db.branches.find_one({"branch_id": data.branch_id})
    if not branch:
        raise HTTPException(status_code=400, detail="Branch not found")
    
    now = datetime.now(timezone.utc).isoformat()
    lot_number = await get_next_lot_number()
    lot_id = str(uuid.uuid4())
    
    # Build lines
    lines = []
    invoice_lines = []
    
    for li in data.line_items:
        sku = await db.buyer_skus.find_one({"buyer_sku_id": li.buyer_sku_id})
        
        lines.append({
            "buyer_sku_id": li.buyer_sku_id,
            "sku_name": li.sku_name or (sku.get("name") if sku else li.buyer_sku_id),
            "bidso_sku_id": sku.get("bidso_sku_id") if sku else None,
            "quantity": li.quantity
        })
        
        invoice_lines.append({
            "buyer_sku_id": li.buyer_sku_id,
            "sku_name": li.sku_name,
            "quantity": li.quantity,
            "rate": li.rate,
            "hsn_code": li.hsn_code,
            "gst_rate": li.gst_rate,
            "tax_amount": li.tax_amount or (li.quantity * li.rate * (li.gst_rate or 0) / 100),
            "amount": li.amount,
            "manufacturing_origin": {
                "branch_id": data.branch_id,
                "branch_name": branch.get("name")
            }
        })
    
    invoice_data = {
        "branch_id": data.branch_id,
        "branch_name": branch.get("name"),
        "source_of_supply": data.source_of_supply,
        "invoice_number": None,
        "order_number": data.order_number,
        "invoice_date": data.invoice_date,
        "payment_terms": data.payment_terms,
        "due_date": data.due_date,
        "accounts_receivable": data.accounts_receivable,
        "salesperson": data.salesperson,
        "subject": data.subject,
        "line_items": invoice_lines,
        "totals": {
            "sub_total": data.totals.sub_total,
            "discount_type": data.totals.discount_type,
            "discount_value": data.totals.discount_value,
            "discount_amount": data.totals.discount_amount,
            "tds_tcs_type": data.totals.tds_tcs_type,
            "tds_tcs_rate": data.totals.tds_tcs_rate,
            "tds_tcs_amount": data.totals.tds_tcs_amount,
            "adjustment": data.totals.adjustment,
            "grand_total": data.totals.grand_total
        },
        "customer_notes": data.customer_notes,
        "terms_conditions": data.terms_conditions,
        "attachments": []
    }
    
    lot = {
        "id": lot_id,
        "lot_number": lot_number,
        "customer_id": customer_id,
        "customer_name": customer.get("name", ""),
        "status": "INVOICED",
        "lines": lines,
        "total_quantity": sum(li.quantity for li in data.line_items),
        "notes": "Created directly by finance",
        "created_by": current_user.id,
        "created_by_role": current_user.role,
        "created_at": now,
        "invoice_data": invoice_data,
        "invoiced_by": current_user.id,
        "invoiced_at": now,
        "dispatched_at": None
    }
    
    await db.dispatch_lots_v2.insert_one(lot)
    
    # Deduct inventory
    for li in data.line_items:
        buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": li.buyer_sku_id})
        if buyer_sku:
            await db.branch_inventory.update_one(
                {"branch_id": data.branch_id, "sku_id": buyer_sku.get("bidso_sku_id")},
                {"$inc": {"quantity": -li.quantity}}
            )
    
    return {"message": "Invoice created", "lot_number": lot_number, "id": lot_id}



# ============ Finance Direct Create & Line Item Management ============

class FinanceDispatchLine(BaseModel):
    """Line item for finance-created dispatch lot"""
    buyer_sku_id: str
    quantity: int
    rate: Optional[float] = None  # Will auto-populate from price master
    hsn_code: Optional[str] = None  # Will auto-populate from SKU
    gst_rate: Optional[float] = None  # Will auto-populate from SKU


class FinanceCreateLotRequest(BaseModel):
    """Create dispatch lot directly - Finance Team"""
    customer_id: str
    branch_id: str  # Must specify branch for inventory check
    lines: List[FinanceDispatchLine]
    order_number: Optional[str] = None
    notes: Optional[str] = None


@router.post("/finance/create-lot")
async def finance_create_dispatch_lot(
    data: FinanceCreateLotRequest,
    current_user = Depends(get_current_user)
):
    """
    Finance creates dispatch lot directly with inventory validation.
    Auto-populates HSN, GST, and Price from master data.
    """
    
    # Validate customer
    customer = await db.buyers.find_one({"id": data.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Validate branch
    branch = await db.branches.find_one({"id": data.branch_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    branch_name = branch.get("name")
    
    # Validate SKUs and check inventory
    enriched_lines = []
    inventory_errors = []
    
    for idx, line in enumerate(data.lines):
        # Get SKU details
        sku = await db.buyer_skus.find_one({"buyer_sku_id": line.buyer_sku_id})
        if not sku:
            raise HTTPException(status_code=400, detail=f"SKU {line.buyer_sku_id} not found")
        
        # Check FG inventory at branch
        fg_inv = await db.branch_sku_inventory.find_one({
            "branch": branch_name,
            "buyer_sku_id": line.buyer_sku_id
        })
        available_qty = fg_inv.get("current_stock", 0) if fg_inv else 0
        
        if available_qty < line.quantity:
            inventory_errors.append({
                "line": idx + 1,
                "sku_id": line.buyer_sku_id,
                "sku_name": sku.get("name", ""),
                "requested": line.quantity,
                "available": available_qty,
                "shortage": line.quantity - available_qty
            })
            continue
        
        # Get price from price master
        price_entry = await db.price_master.find_one({
            "buyer_id": data.customer_id,
            "buyer_sku_id": line.buyer_sku_id
        })
        rate = line.rate or (price_entry.get("unit_price", 0) if price_entry else 0)
        
        # Get HSN/GST from SKU
        hsn_code = line.hsn_code or sku.get("hsn_code", "")
        gst_rate = line.gst_rate if line.gst_rate is not None else sku.get("gst_rate", 18)
        
        enriched_lines.append({
            "id": str(uuid.uuid4()),
            "buyer_sku_id": line.buyer_sku_id,
            "sku_name": sku.get("name", ""),
            "quantity": line.quantity,
            "rate": rate,
            "hsn_code": hsn_code,
            "gst_rate": gst_rate,
            "amount": rate * line.quantity
        })
    
    # If inventory errors, return them
    if inventory_errors:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "INSUFFICIENT_INVENTORY",
                "message": f"{len(inventory_errors)} item(s) have insufficient inventory at {branch_name}",
                "shortages": inventory_errors
            }
        )
    
    if not enriched_lines:
        raise HTTPException(status_code=400, detail="No valid line items")
    
    # Generate lot number
    today = datetime.now(timezone.utc)
    lot_prefix = f"DL-{today.strftime('%Y')}-{today.strftime('%m%d')}"
    last_lot = await db.dispatch_lots.find_one(
        {"lot_number": {"$regex": f"^{lot_prefix}"}},
        sort=[("lot_number", -1)]
    )
    if last_lot:
        try:
            last_num = int(last_lot["lot_number"].split("-")[-1])
            next_num = last_num + 1
        except:
            next_num = 1
    else:
        next_num = 1
    
    lot_number = f"{lot_prefix}-{next_num:03d}"
    
    # Calculate totals
    sub_total = sum(line["amount"] for line in enriched_lines)
    total_qty = sum(line["quantity"] for line in enriched_lines)
    
    # Create dispatch lot (Finance-created starts as PENDING_FINANCE)
    dispatch_lot = {
        "id": str(uuid.uuid4()),
        "lot_number": lot_number,
        "buyer_id": data.customer_id,
        "buyer_name": customer.get("name"),
        "buyer_code": customer.get("customer_code"),
        "branch_id": data.branch_id,
        "branch_name": branch_name,
        "status": "PENDING_FINANCE",
        "lines": enriched_lines,
        "total_quantity": total_qty,
        "sub_total": sub_total,
        "order_number": data.order_number,
        "notes": data.notes,
        "created_by_finance": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id,
        "created_by_name": current_user.name
    }
    
    await db.dispatch_lots.insert_one(dispatch_lot)
    
    return {
        "message": f"Dispatch lot {lot_number} created successfully",
        "lot_number": lot_number,
        "lot_id": dispatch_lot["id"],
        "branch": branch_name,
        "customer": customer.get("name"),
        "lines": len(enriched_lines),
        "total_quantity": total_qty,
        "sub_total": sub_total
    }


class AddLineItemRequest(BaseModel):
    """Add line item to existing dispatch lot"""
    buyer_sku_id: str
    quantity: int
    rate: Optional[float] = None
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = None


@router.post("/{lot_id}/add-line")
async def add_line_to_dispatch_lot(
    lot_id: str,
    data: AddLineItemRequest,
    current_user = Depends(get_current_user)
):
    """
    Add a line item to existing dispatch lot.
    Auto-populates HSN, GST, and Price from master data.
    Validates inventory if branch is assigned.
    """
    
    # Get lot
    lot = await db.dispatch_lots.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Dispatch lot not found")
    
    # Can only add lines to DRAFT or PENDING_FINANCE lots
    if lot.get("status") not in ["DRAFT", "PENDING_FINANCE"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot add lines to lot with status {lot.get('status')}"
        )
    
    # Get SKU details
    sku = await db.buyer_skus.find_one({"buyer_sku_id": data.buyer_sku_id})
    if not sku:
        raise HTTPException(status_code=404, detail=f"SKU {data.buyer_sku_id} not found")
    
    # If branch is assigned, validate inventory
    branch_name = lot.get("branch_name")
    if branch_name:
        fg_inv = await db.branch_sku_inventory.find_one({
            "branch": branch_name,
            "buyer_sku_id": data.buyer_sku_id
        })
        available_qty = fg_inv.get("current_stock", 0) if fg_inv else 0
        
        # Check existing lines for same SKU
        existing_qty = sum(
            line.get("quantity", 0) 
            for line in lot.get("lines", []) 
            if line.get("buyer_sku_id") == data.buyer_sku_id
        )
        
        total_requested = existing_qty + data.quantity
        if available_qty < total_requested:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "INSUFFICIENT_INVENTORY",
                    "sku_id": data.buyer_sku_id,
                    "sku_name": sku.get("name", ""),
                    "requested": data.quantity,
                    "existing_in_lot": existing_qty,
                    "available": available_qty,
                    "shortage": total_requested - available_qty
                }
            )
    
    # Get price from price master
    buyer_id = lot.get("buyer_id")
    price_entry = await db.price_master.find_one({
        "buyer_id": buyer_id,
        "buyer_sku_id": data.buyer_sku_id
    })
    rate = data.rate or (price_entry.get("unit_price", 0) if price_entry else 0)
    
    # Get HSN/GST from SKU
    hsn_code = data.hsn_code or sku.get("hsn_code", "")
    gst_rate = data.gst_rate if data.gst_rate is not None else sku.get("gst_rate", 18)
    
    # Create new line item
    new_line = {
        "id": str(uuid.uuid4()),
        "buyer_sku_id": data.buyer_sku_id,
        "sku_name": sku.get("name", ""),
        "quantity": data.quantity,
        "rate": rate,
        "hsn_code": hsn_code,
        "gst_rate": gst_rate,
        "amount": rate * data.quantity,
        "added_at": datetime.now(timezone.utc).isoformat(),
        "added_by": current_user.id,
        "added_by_name": current_user.name
    }
    
    # Update lot
    lines = lot.get("lines", [])
    lines.append(new_line)
    
    total_quantity = sum(l.get("quantity", 0) for l in lines)
    sub_total = sum(l.get("amount", 0) for l in lines)
    
    await db.dispatch_lots.update_one(
        {"id": lot_id},
        {"$set": {
            "lines": lines,
            "total_quantity": total_quantity,
            "sub_total": sub_total,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.id
        }}
    )
    
    return {
        "message": f"Line item added to {lot.get('lot_number')}",
        "line_id": new_line["id"],
        "sku_id": data.buyer_sku_id,
        "sku_name": sku.get("name", ""),
        "quantity": data.quantity,
        "rate": rate,
        "hsn_code": hsn_code,
        "gst_rate": gst_rate,
        "amount": new_line["amount"],
        "lot_total_quantity": total_quantity,
        "lot_sub_total": sub_total
    }


@router.get("/sku-lookup/{buyer_sku_id}")
async def lookup_sku_for_dispatch(
    buyer_sku_id: str,
    customer_id: Optional[str] = Query(None),
    current_user = Depends(get_current_user)
):
    """
    Lookup SKU details for adding to dispatch lot.
    Returns HSN, GST, and Price (if customer specified).
    """
    
    sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
    if not sku:
        raise HTTPException(status_code=404, detail="SKU not found")
    
    result = {
        "buyer_sku_id": buyer_sku_id,
        "name": sku.get("name", ""),
        "hsn_code": sku.get("hsn_code", ""),
        "gst_rate": sku.get("gst_rate", 18),
        "rate": 0
    }
    
    # Get price if customer specified
    if customer_id:
        price_entry = await db.price_master.find_one({
            "buyer_id": customer_id,
            "buyer_sku_id": buyer_sku_id
        })
        if price_entry:
            result["rate"] = price_entry.get("unit_price", 0)
    
    return result
