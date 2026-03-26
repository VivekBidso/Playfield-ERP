"""Tech Ops routes - Verticals, Models, Brands, Buyers, BOM"""
from fastapi import APIRouter, HTTPException, UploadFile, File
from datetime import datetime, timezone
from typing import Optional, List
import uuid
import io

from database import db
from models.master_data import (
    Vertical, VerticalCreate,
    Model, ModelCreate,
    Brand, BrandCreate,
    Buyer, BuyerCreate, BuyerUpdate, BuyerBulkImport
)

router = APIRouter(tags=["Tech Ops"])

def serialize_doc(doc):
    if doc and 'created_at' in doc and isinstance(doc['created_at'], str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    return doc

# --- Verticals CRUD ---
@router.get("/verticals")
async def get_verticals():
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(1000)
    return [serialize_doc(v) for v in verticals]

@router.post("/verticals")
async def create_vertical(data: VerticalCreate):
    existing = await db.verticals.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Vertical with code {data.code} already exists")
    
    vertical = {
        "id": str(uuid.uuid4()),
        "code": data.code.upper(),
        "name": data.name,
        "description": data.description,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.verticals.insert_one(vertical)
    del vertical["_id"]
    return serialize_doc(vertical)

@router.put("/verticals/{vertical_id}")
async def update_vertical(vertical_id: str, data: VerticalCreate):
    result = await db.verticals.update_one(
        {"id": vertical_id},
        {"$set": {"code": data.code.upper(), "name": data.name, "description": data.description}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vertical not found")
    return {"message": "Vertical updated"}

@router.delete("/verticals/{vertical_id}")
async def delete_vertical(vertical_id: str):
    models_count = await db.models.count_documents({"vertical_id": vertical_id, "status": "ACTIVE"})
    if models_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {models_count} active models use this vertical")
    
    skus_count = await db.skus.count_documents({"vertical_id": vertical_id})
    if skus_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {skus_count} SKUs use this vertical")
    
    result = await db.verticals.update_one({"id": vertical_id}, {"$set": {"status": "INACTIVE"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vertical not found")
    return {"message": "Vertical deleted"}

# --- Models CRUD ---
@router.get("/models")
async def get_models(vertical_id: Optional[str] = None):
    query = {}
    if vertical_id:
        query["vertical_id"] = vertical_id
    models = await db.models.find(query, {"_id": 0}).to_list(1000)
    return [serialize_doc(m) for m in models]

@router.post("/models")
async def create_model(data: ModelCreate):
    existing = await db.models.find_one({"vertical_id": data.vertical_id, "code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Model with code {data.code} already exists for this vertical")
    
    model = {
        "id": str(uuid.uuid4()),
        "vertical_id": data.vertical_id,
        "code": data.code.upper(),
        "name": data.name,
        "description": data.description,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.models.insert_one(model)
    del model["_id"]
    return serialize_doc(model)

@router.put("/models/{model_id}")
async def update_model(model_id: str, data: ModelCreate):
    result = await db.models.update_one(
        {"id": model_id},
        {"$set": {"vertical_id": data.vertical_id, "code": data.code.upper(), "name": data.name, "description": data.description}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Model not found")
    return {"message": "Model updated"}

@router.delete("/models/{model_id}")
async def delete_model(model_id: str):
    skus_count = await db.skus.count_documents({"model_id": model_id})
    if skus_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {skus_count} SKUs use this model")
    
    result = await db.models.update_one({"id": model_id}, {"$set": {"status": "INACTIVE"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Model not found")
    return {"message": "Model deleted"}


@router.post("/models/bulk-import")
async def bulk_import_models(file: UploadFile = File(...)):
    """
    Bulk import models from Excel file.
    Expected columns: Vertical, Model Name, Model Code
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Get all verticals for mapping
    verticals = await db.verticals.find({"status": "ACTIVE"}, {"_id": 0}).to_list(100)
    
    # Create mapping from vertical name to id
    vertical_map = {}
    for v in verticals:
        name_lower = v["name"].lower().strip()
        vertical_map[name_lower] = v["id"]
        # Also map by code and common variations
        vertical_map[v["code"].lower()] = v["id"]
    
    # Add common name mappings
    vertical_map["scooter"] = next((v["id"] for v in verticals if v["code"] == "KS"), None)
    vertical_map["tricycle"] = next((v["id"] for v in verticals if v["code"] in ["KTC", "KT"]), None)
    vertical_map["rideon"] = next((v["id"] for v in verticals if v["code"] == "SC"), None)
    vertical_map["push rideon"] = next((v["id"] for v in verticals if v["code"] == "PR"), None)
    vertical_map["walker"] = next((v["id"] for v in verticals if v["code"] == "BW"), None)
    vertical_map["shakers"] = next((v["id"] for v in verticals if v["code"] == "SH"), None)
    vertical_map["electric scooter"] = next((v["id"] for v in verticals if v["code"] == "EKS"), None)
    vertical_map["electric rideon"] = next((v["id"] for v in verticals if v["code"] == "EV"), None)
    
    # Get headers
    headers = [cell.value for cell in ws[1] if cell.value]
    headers_lower = [h.lower().strip() if h else "" for h in headers]
    
    col_map = {}
    for idx, h in enumerate(headers_lower):
        if h == "vertical":
            col_map["vertical"] = idx
        elif "model name" in h or h == "name":
            col_map["name"] = idx
        elif "model code" in h or h == "code":
            col_map["code"] = idx
    
    if "name" not in col_map or "code" not in col_map:
        raise HTTPException(status_code=400, detail="Missing required columns: 'Model Name' and 'Model Code'")
    
    created = 0
    skipped = 0
    errors = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < max(col_map.values()) + 1:
            continue
            
        name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
        code = str(row[col_map["code"]]).strip().upper() if row[col_map["code"]] else ""
        
        if not name or not code:
            continue
        
        # Determine vertical
        vertical_id = None
        if "vertical" in col_map and row[col_map["vertical"]]:
            vertical_name = str(row[col_map["vertical"]]).strip().lower()
            vertical_id = vertical_map.get(vertical_name)
        
        if not vertical_id:
            errors.append(f"Row {row_num}: Could not find vertical for '{row[col_map.get('vertical', 0)] if 'vertical' in col_map else 'unknown'}'")
            continue
        
        # Check if model already exists (by code in same vertical)
        existing = await db.models.find_one({"code": code, "vertical_id": vertical_id})
        if existing:
            skipped += 1
            continue
        
        model = {
            "id": str(uuid.uuid4()),
            "vertical_id": vertical_id,
            "code": code,
            "name": name,
            "description": "",
            "status": "ACTIVE",
            "created_at": datetime.now(timezone.utc)
        }
        
        try:
            await db.models.insert_one(model)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    return {
        "message": f"Import complete: {created} created, {skipped} skipped (duplicates)",
        "created": created,
        "skipped": skipped,
        "errors": errors[:10]
    }


@router.delete("/models/clear-all")
async def clear_all_models():
    """Delete all models (admin only) - use with caution"""
    result = await db.models.delete_many({})
    return {"message": f"Deleted {result.deleted_count} models"}

# --- Brands CRUD ---
@router.get("/brands")
async def get_brands():
    brands = await db.brands.find({}, {"_id": 0}).to_list(1000)
    return [serialize_doc(b) for b in brands]

@router.post("/brands")
async def create_brand(data: BrandCreate):
    existing = await db.brands.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Brand with code {data.code} already exists")
    
    brand = {
        "id": str(uuid.uuid4()),
        "code": data.code.upper(),
        "name": data.name,
        "buyer_id": data.buyer_id,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.brands.insert_one(brand)
    del brand["_id"]
    return serialize_doc(brand)

@router.put("/brands/{brand_id}")
async def update_brand(brand_id: str, data: BrandCreate):
    result = await db.brands.update_one(
        {"id": brand_id},
        {"$set": {"code": data.code.upper(), "name": data.name, "buyer_id": data.buyer_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    return {"message": "Brand updated"}

@router.delete("/brands/{brand_id}")
async def delete_brand(brand_id: str):
    skus_count = await db.skus.count_documents({"brand_id": brand_id})
    if skus_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {skus_count} SKUs use this brand")
    
    result = await db.brands.update_one({"id": brand_id}, {"$set": {"status": "INACTIVE"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    return {"message": "Brand deleted"}

# --- Buyers CRUD ---
async def generate_customer_code() -> str:
    """Generate unique customer code like CUST001, CUST002, etc."""
    # Find the highest existing customer code
    buyers = await db.buyers.find({}, {"customer_code": 1, "_id": 0}).to_list(10000)
    max_num = 0
    for buyer in buyers:
        code = buyer.get("customer_code", "")
        if code and code.startswith("CUST"):
            try:
                num = int(code[4:])
                if num > max_num:
                    max_num = num
            except ValueError:
                continue
    return f"CUST{str(max_num + 1).zfill(3)}"


@router.get("/buyers")
async def get_buyers():
    """Get all buyers with brands dispatched to them"""
    buyers = await db.buyers.find({}, {"_id": 0}).to_list(1000)
    
    # Get brands associated with buyers from dispatch data
    # Aggregate from dispatch_lots to find which brands have been shipped to each buyer
    pipeline = [
        {"$match": {"status": {"$ne": "CANCELLED"}}},
        {"$lookup": {
            "from": "skus",
            "localField": "sku_id",
            "foreignField": "id",
            "as": "sku_info"
        }},
        {"$unwind": {"path": "$sku_info", "preserveNullAndEmptyArrays": True}},
        {"$group": {
            "_id": "$buyer_id",
            "brand_ids": {"$addToSet": "$sku_info.brand_id"}
        }}
    ]
    
    buyer_brands = {}
    try:
        results = await db.dispatch_lots.aggregate(pipeline).to_list(1000)
        for r in results:
            buyer_brands[r["_id"]] = [b for b in r.get("brand_ids", []) if b]
    except Exception:
        pass  # If aggregation fails, continue without brand info
    
    # Get all brands for lookup
    brands = await db.brands.find({}, {"_id": 0}).to_list(1000)
    brand_map = {b["id"]: b["name"] for b in brands}
    
    # Enrich buyers with brand names
    for buyer in buyers:
        brand_ids = buyer_brands.get(buyer.get("id"), [])
        buyer["brands_dispatched"] = [brand_map.get(bid, "") for bid in brand_ids if bid in brand_map]
    
    return [serialize_doc(b) for b in buyers]


@router.post("/buyers")
async def create_buyer(data: BuyerCreate):
    """Create a new buyer with auto-generated customer code"""
    # Check if buyer with same name exists
    existing = await db.buyers.find_one({"name": data.name, "status": "ACTIVE"})
    if existing:
        raise HTTPException(status_code=400, detail=f"Buyer with name '{data.name}' already exists")
    
    # Generate unique customer code
    customer_code = await generate_customer_code()
    
    buyer = {
        "id": str(uuid.uuid4()),
        "customer_code": customer_code,
        "name": data.name,
        "gst": data.gst,
        "email": data.email,
        "phone_no": data.phone_no,
        "poc_name": data.poc_name,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.buyers.insert_one(buyer)
    del buyer["_id"]
    return serialize_doc(buyer)


@router.get("/buyers/{buyer_id}")
async def get_buyer(buyer_id: str):
    buyer = await db.buyers.find_one({"id": buyer_id}, {"_id": 0})
    if not buyer:
        raise HTTPException(status_code=404, detail="Buyer not found")
    return serialize_doc(buyer)


@router.put("/buyers/{buyer_id}")
async def update_buyer(buyer_id: str, data: BuyerUpdate):
    """Update buyer information (customer_code cannot be changed)"""
    update_data = {}
    if data.name is not None:
        update_data["name"] = data.name
    if data.gst is not None:
        update_data["gst"] = data.gst
    if data.email is not None:
        update_data["email"] = data.email
    if data.phone_no is not None:
        update_data["phone_no"] = data.phone_no
    if data.poc_name is not None:
        update_data["poc_name"] = data.poc_name
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.buyers.update_one(
        {"id": buyer_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Buyer not found")
    return {"message": "Buyer updated"}


@router.delete("/buyers/{buyer_id}")
async def delete_buyer(buyer_id: str):
    # Check for dependencies - forecasts, dispatch_lots
    forecasts_count = await db.forecasts.count_documents({"buyer_id": buyer_id})
    if forecasts_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {forecasts_count} forecasts reference this buyer")
    
    lots_count = await db.dispatch_lots.count_documents({"buyer_id": buyer_id})
    if lots_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {lots_count} dispatch lots reference this buyer")
    
    result = await db.buyers.update_one({"id": buyer_id}, {"$set": {"status": "INACTIVE"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Buyer not found")
    return {"message": "Buyer deleted"}


@router.post("/buyers/bulk-import")
async def bulk_import_buyers(file: UploadFile = File(...)):
    """
    Bulk import buyers from Excel file.
    Expected columns: Customer Name, GST, Email, Phone No, POC Name
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1] if cell.value]
    headers_lower = [h.lower().strip() if h else "" for h in headers]
    
    # Map expected columns
    col_map = {}
    for idx, h in enumerate(headers_lower):
        if "customer name" in h or h == "name":
            col_map["name"] = idx
        elif "gst" in h:
            col_map["gst"] = idx
        elif "email" in h:
            col_map["email"] = idx
        elif "phone" in h:
            col_map["phone_no"] = idx
        elif "poc" in h or "contact" in h:
            col_map["poc_name"] = idx
    
    if "name" not in col_map:
        raise HTTPException(status_code=400, detail="Missing required column: 'Customer Name' or 'Name'")
    
    # Process rows
    created = 0
    skipped = 0
    errors = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[col_map["name"]]:
            continue
        
        name = str(row[col_map["name"]]).strip()
        if not name:
            continue
        
        # Check if buyer exists
        existing = await db.buyers.find_one({"name": name})
        if existing:
            skipped += 1
            continue
        
        # Generate customer code
        customer_code = await generate_customer_code()
        
        buyer = {
            "id": str(uuid.uuid4()),
            "customer_code": customer_code,
            "name": name,
            "gst": str(row[col_map.get("gst", -1)] or "").strip() if "gst" in col_map and col_map["gst"] < len(row) else "",
            "email": str(row[col_map.get("email", -1)] or "").strip() if "email" in col_map and col_map["email"] < len(row) else "",
            "phone_no": str(row[col_map.get("phone_no", -1)] or "").strip() if "phone_no" in col_map and col_map["phone_no"] < len(row) else "",
            "poc_name": str(row[col_map.get("poc_name", -1)] or "").strip() if "poc_name" in col_map and col_map["poc_name"] < len(row) else "",
            "status": "ACTIVE",
            "created_at": datetime.now(timezone.utc)
        }
        
        try:
            await db.buyers.insert_one(buyer)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    return {
        "message": f"Import complete: {created} created, {skipped} skipped (duplicates)",
        "created": created,
        "skipped": skipped,
        "errors": errors[:10]  # Limit error messages
    }


@router.delete("/buyers/clear-all")
async def clear_all_buyers():
    """Delete all buyers (admin only) - use with caution"""
    result = await db.buyers.delete_many({})
    return {"message": f"Deleted {result.deleted_count} buyers"}



# --- Branches ---
BRANCHES = [
    "Unit 1 Vedica",
    "Unit 2 Trikes",
    "Unit 3 TM",
    "Unit 4 Goa",
    "Unit 5 Baabus",
    "Unit 6 Emox",
    "BHDG WH"
]

RM_CATEGORIES = {
    "INP": {
        "name": "In-house Plastic", 
        "fields": ["mould_code", "model_name", "part_name", "colour", "mb", "per_unit_weight", "unit"],
        "nameFormat": ["mould_code", "model_name", "part_name", "colour", "mb"]
    },
    "INM": {
        "name": "In-house Metal", 
        "fields": ["model_name", "part_name", "colour", "mb", "per_unit_weight", "unit"],
        "nameFormat": ["model_name", "part_name", "colour", "mb"]
    },
    "ACC": {
        "name": "Accessories", 
        "fields": ["type", "model_name", "specs", "colour", "per_unit_weight", "unit"],
        "nameFormat": ["type", "model_name", "specs", "colour"]
    },
    "ELC": {
        "name": "Electric Components", 
        "fields": ["model", "type", "specs", "per_unit_weight", "unit"],
        "nameFormat": ["model", "type", "specs"]
    },
    "SP": {
        "name": "Spares", 
        "fields": ["type", "specs", "per_unit_weight", "unit"],
        "nameFormat": ["type", "specs"]
    },
    "BS": {
        "name": "Brand Assets", 
        "fields": ["position", "type", "brand", "buyer_sku", "per_unit_weight", "unit"],
        "nameFormat": ["position", "type", "brand", "buyer_sku"]
    },
    "PM": {
        "name": "Packaging", 
        "fields": ["model", "type", "specs", "brand", "per_unit_weight", "unit"],
        "nameFormat": ["model", "type", "specs", "brand"]
    },
    "LB": {
        "name": "Labels", 
        "fields": ["type", "buyer_sku", "per_unit_weight", "unit"],
        "nameFormat": ["type", "buyer_sku"]
    }
}

@router.get("/branches")
async def get_branches():
    """Get all branches from database"""
    branches = await db.branches.find({}, {"_id": 0}).to_list(1000)
    return branches

@router.get("/branches/names")
async def get_branch_names():
    """Get branch names only"""
    return {"branches": BRANCHES}

@router.get("/rm-categories")
async def get_rm_categories():
    """Get all RM categories"""
    return RM_CATEGORIES
