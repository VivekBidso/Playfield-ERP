"""Demand routes - Forecasts, Dispatch Lots"""
from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from datetime import datetime, timezone, timedelta
from typing import Optional, List
import uuid
import io

from database import db

router = APIRouter(tags=["Demand"])

def serialize_doc(doc):
    if doc and 'created_at' in doc and isinstance(doc['created_at'], str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    if doc and 'forecast_month' in doc and isinstance(doc['forecast_month'], str):
        doc['forecast_month'] = datetime.fromisoformat(doc['forecast_month'])
    if doc and 'target_date' in doc and isinstance(doc['target_date'], str):
        doc['target_date'] = datetime.fromisoformat(doc['target_date'])
    return doc

class ForecastCreate(BaseModel):
    buyer_id: str  # REQUIRED - no "All Buyers" option
    vertical_id: Optional[str] = None
    sku_id: Optional[str] = None
    forecast_month: datetime
    quantity: int
    priority: str = "MEDIUM"
    notes: str = ""


class BulkConfirmRequest(BaseModel):
    forecast_ids: List[str]

class DispatchLotCreate(BaseModel):
    forecast_id: Optional[str] = None
    sku_id: str
    buyer_id: Optional[str] = None
    required_quantity: int
    target_date: datetime
    priority: str = "MEDIUM"
    notes: Optional[str] = ""

# New model for multi-line dispatch lots
class DispatchLotLineInput(BaseModel):
    sku_id: str
    brand_id: Optional[str] = None
    vertical_id: Optional[str] = None
    quantity: int
    forecast_id: Optional[str] = None  # Track source forecast for each line

class DispatchLotMultiCreate(BaseModel):
    buyer_id: str
    forecast_id: Optional[str] = None  # Main forecast_id for the lot (if all from same forecast)
    target_date: datetime
    priority: str = "MEDIUM"
    notes: Optional[str] = ""
    lines: List[DispatchLotLineInput]


# Model for updating dispatch lot line
class DispatchLotLineUpdate(BaseModel):
    id: Optional[str] = None  # None for new lines, existing id for updates
    sku_id: str
    brand_id: Optional[str] = None
    vertical_id: Optional[str] = None
    quantity: int
    forecast_id: Optional[str] = None


# Model for updating dispatch lot
class DispatchLotUpdate(BaseModel):
    target_date: Optional[datetime] = None
    priority: Optional[str] = None
    notes: Optional[str] = None
    lines: Optional[List[DispatchLotLineUpdate]] = None  # If provided, replaces all lines

# --- Forecasts ---
@router.get("/forecasts")
async def get_forecasts(
    buyer_id: Optional[str] = None,
    vertical_id: Optional[str] = None,
    status: Optional[str] = None
):
    query = {}
    if buyer_id:
        query["buyer_id"] = buyer_id
    if vertical_id:
        query["vertical_id"] = vertical_id
    if status:
        query["status"] = status
    forecasts = await db.forecasts.find(query, {"_id": 0}).to_list(1000)
    
    if not forecasts:
        return []
    
    # Get all forecast IDs
    forecast_ids = [f["id"] for f in forecasts]
    
    # Get production schedules for these forecasts
    schedules = await db.production_schedules.find(
        {"forecast_id": {"$in": forecast_ids}, "status": {"$ne": "CANCELLED"}},
        {"_id": 0, "forecast_id": 1, "target_quantity": 1}
    ).to_list(5000)
    
    # Group scheduled quantities by forecast_id
    scheduled_by_forecast = {}
    for s in schedules:
        fid = s.get("forecast_id")
        if fid:
            scheduled_by_forecast[fid] = scheduled_by_forecast.get(fid, 0) + s.get("target_quantity", 0)
    
    # Get dispatch lot lines linked to these forecasts
    lot_lines = await db.dispatch_lot_lines.find(
        {"forecast_id": {"$in": forecast_ids}},
        {"_id": 0, "forecast_id": 1, "quantity": 1}
    ).to_list(5000)
    
    # Group dispatch quantities by forecast_id
    dispatch_by_forecast = {}
    for line in lot_lines:
        fid = line.get("forecast_id")
        if fid:
            dispatch_by_forecast[fid] = dispatch_by_forecast.get(fid, 0) + line.get("quantity", 0)
    
    # Also check dispatch_lots with forecast_id (for lots created before line-level tracking)
    lots_with_forecast = await db.dispatch_lots.find(
        {"forecast_id": {"$in": forecast_ids}},
        {"_id": 0, "forecast_id": 1, "total_quantity": 1}
    ).to_list(5000)
    
    for lot in lots_with_forecast:
        fid = lot.get("forecast_id")
        if fid:
            # Only add if not already counted in lines
            if fid not in dispatch_by_forecast:
                dispatch_by_forecast[fid] = lot.get("total_quantity", 0)
    
    # Enrich forecasts with calculated fields
    result = []
    for f in forecasts:
        fid = f["id"]
        forecast_qty = f.get("quantity", 0)
        dispatch_allocated = dispatch_by_forecast.get(fid, 0)
        production_scheduled = scheduled_by_forecast.get(fid, 0)
        schedule_pending = max(0, forecast_qty - production_scheduled)
        
        enriched = {
            **f,
            "dispatch_allocated": dispatch_allocated,
            "production_scheduled": production_scheduled,
            "schedule_pending": schedule_pending
        }
        result.append(serialize_doc(enriched))
    
    return result

@router.post("/forecasts")
async def create_forecast(data: ForecastCreate):
    # Validate quantity is greater than 0
    if data.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
    
    # Validate buyer_id is provided
    if not data.buyer_id:
        raise HTTPException(status_code=400, detail="Buyer is required for creating a forecast")
    
    # Verify buyer exists
    buyer = await db.buyers.find_one({"id": data.buyer_id}, {"_id": 0})
    if not buyer:
        raise HTTPException(status_code=404, detail="Buyer not found")
    
    # Auto-derive vertical_id from SKU if not provided
    vertical_id = data.vertical_id
    if data.sku_id and not vertical_id:
        sku = await db.skus.find_one({"sku_id": data.sku_id}, {"_id": 0})
        if sku and sku.get("vertical_id"):
            vertical_id = sku["vertical_id"]
    
    count = await db.forecasts.count_documents({})
    forecast_code = f"FC_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    forecast = {
        "id": str(uuid.uuid4()),
        "forecast_code": forecast_code,
        "buyer_id": data.buyer_id,
        "vertical_id": vertical_id,
        "sku_id": data.sku_id,
        "forecast_month": data.forecast_month,
        "quantity": data.quantity,
        "planned_quantity": 0,  # Track production planned against this forecast
        "dispatched_quantity": 0,  # Track dispatch lots created
        "priority": data.priority,
        "status": "DRAFT",
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc)
    }
    await db.forecasts.insert_one(forecast)
    del forecast["_id"]
    return serialize_doc(forecast)


# Pydantic model for updating forecast
class ForecastUpdate(BaseModel):
    buyer_id: Optional[str] = None
    vertical_id: Optional[str] = None
    sku_id: Optional[str] = None
    forecast_month: Optional[datetime] = None
    quantity: Optional[int] = None
    priority: Optional[str] = None
    notes: Optional[str] = None


@router.put("/forecasts/{forecast_id}")
async def update_forecast(forecast_id: str, data: ForecastUpdate):
    """Update a forecast - only allowed while in DRAFT status"""
    # Check if forecast exists and is in DRAFT status
    forecast = await db.forecasts.find_one({"id": forecast_id}, {"_id": 0})
    if not forecast:
        raise HTTPException(status_code=404, detail="Forecast not found")
    
    if forecast.get("status") != "DRAFT":
        raise HTTPException(status_code=400, detail="Cannot edit forecast after confirmation. Only DRAFT forecasts can be edited.")
    
    # Build update dict
    update_data = {}
    if data.buyer_id is not None:
        # Verify buyer exists
        buyer = await db.buyers.find_one({"id": data.buyer_id}, {"_id": 0})
        if not buyer:
            raise HTTPException(status_code=404, detail="Buyer not found")
        update_data["buyer_id"] = data.buyer_id
    
    if data.vertical_id is not None:
        update_data["vertical_id"] = data.vertical_id
    
    if data.sku_id is not None:
        update_data["sku_id"] = data.sku_id
        # Auto-derive vertical_id from SKU if not explicitly provided
        if data.vertical_id is None:
            sku = await db.skus.find_one({"sku_id": data.sku_id}, {"_id": 0})
            if sku and sku.get("vertical_id"):
                update_data["vertical_id"] = sku["vertical_id"]
    
    if data.forecast_month is not None:
        update_data["forecast_month"] = data.forecast_month
    
    if data.quantity is not None:
        if data.quantity <= 0:
            raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
        update_data["quantity"] = data.quantity
    
    if data.priority is not None:
        update_data["priority"] = data.priority
    
    if data.notes is not None:
        update_data["notes"] = data.notes
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.forecasts.update_one(
        {"id": forecast_id},
        {"$set": update_data}
    )
    
    # Return updated forecast
    updated = await db.forecasts.find_one({"id": forecast_id}, {"_id": 0})
    return serialize_doc(updated)


@router.delete("/forecasts/{forecast_id}")
async def delete_forecast(forecast_id: str):
    """Delete a forecast - only allowed while in DRAFT status"""
    # Check if forecast exists and is in DRAFT status
    forecast = await db.forecasts.find_one({"id": forecast_id}, {"_id": 0})
    if not forecast:
        raise HTTPException(status_code=404, detail="Forecast not found")
    
    if forecast.get("status") != "DRAFT":
        raise HTTPException(status_code=400, detail="Cannot delete forecast after confirmation. Only DRAFT forecasts can be deleted.")
    
    # Delete the forecast
    await db.forecasts.delete_one({"id": forecast_id})
    
    return {"message": "Forecast deleted successfully", "forecast_code": forecast.get("forecast_code")}

@router.put("/forecasts/{forecast_id}/confirm")
async def confirm_forecast(forecast_id: str):
    result = await db.forecasts.update_one(
        {"id": forecast_id, "status": "DRAFT"},
        {"$set": {"status": "CONFIRMED", "confirmed_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=400, detail="Forecast not found or not in DRAFT status")
    return {"message": "Forecast confirmed"}


@router.post("/forecasts/bulk-confirm")
async def bulk_confirm_forecasts(data: BulkConfirmRequest):
    """Bulk confirm multiple forecasts (Master Admin action)"""
    if not data.forecast_ids:
        raise HTTPException(status_code=400, detail="No forecast IDs provided")
    
    result = await db.forecasts.update_many(
        {"id": {"$in": data.forecast_ids}, "status": "DRAFT"},
        {"$set": {"status": "CONFIRMED", "confirmed_at": datetime.now(timezone.utc)}}
    )
    
    return {
        "message": f"Confirmed {result.modified_count} forecasts",
        "confirmed_count": result.modified_count
    }


@router.post("/forecasts/confirm-month")
async def confirm_month_forecasts(month: str, buyer_id: Optional[str] = None):
    """Confirm all draft forecasts for a specific month"""
    # Parse month (format: YYYY-MM)
    try:
        year, mon = month.split("-")
        start_date = datetime(int(year), int(mon), 1, tzinfo=timezone.utc)
        if int(mon) == 12:
            end_date = datetime(int(year) + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(int(year), int(mon) + 1, 1, tzinfo=timezone.utc)
    except:
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    
    query = {
        "status": "DRAFT",
        "forecast_month": {"$gte": start_date, "$lt": end_date}
    }
    if buyer_id:
        query["buyer_id"] = buyer_id
    
    result = await db.forecasts.update_many(
        query,
        {"$set": {"status": "CONFIRMED", "confirmed_at": datetime.now(timezone.utc)}}
    )
    
    return {
        "message": f"Confirmed {result.modified_count} forecasts for {month}",
        "confirmed_count": result.modified_count
    }


@router.get("/forecasts/{forecast_id}/dispatch-lots")
async def get_forecast_dispatch_lots(forecast_id: str):
    """Get dispatch lots linked to a specific forecast via dispatch_lot_lines"""
    # Find all dispatch_lot_lines that reference this forecast
    lines = await db.dispatch_lot_lines.find(
        {"forecast_id": forecast_id},
        {"_id": 0}
    ).to_list(1000)
    
    if not lines:
        # Also check if forecast_code is stored in lines
        forecast = await db.forecasts.find_one({"id": forecast_id}, {"_id": 0})
        if forecast and forecast.get("forecast_code"):
            lines = await db.dispatch_lot_lines.find(
                {"forecast_code": forecast.get("forecast_code")},
                {"_id": 0}
            ).to_list(1000)
    
    # Get unique lot IDs
    lot_ids = list(set(l.get("lot_id") for l in lines if l.get("lot_id")))
    
    if not lot_ids:
        return []
    
    # Get the lots
    lots = await db.dispatch_lots.find(
        {"id": {"$in": lot_ids}},
        {"_id": 0}
    ).to_list(100)
    
    # Get line quantities per lot
    line_qty_by_lot = {}
    for line in lines:
        lot_id = line.get("lot_id")
        if lot_id:
            if lot_id not in line_qty_by_lot:
                line_qty_by_lot[lot_id] = {"lines": [], "total_qty": 0}
            line_qty_by_lot[lot_id]["lines"].append({
                "line_number": line.get("line_number"),
                "sku_id": line.get("sku_id"),
                "quantity": line.get("quantity", 0),
                "status": line.get("status")
            })
            line_qty_by_lot[lot_id]["total_qty"] += line.get("quantity", 0)
    
    # Enrich lots with line info
    result = []
    for lot in lots:
        lot_id = lot.get("id")
        lot_info = line_qty_by_lot.get(lot_id, {"lines": [], "total_qty": 0})
        result.append({
            **serialize_doc(lot),
            "forecast_lines": lot_info["lines"],
            "forecast_qty_in_lot": lot_info["total_qty"]
        })
    
    return result


@router.get("/forecasts/{forecast_id}/production-plans")
async def get_forecast_production_plans(forecast_id: str):
    """Get production plans linked to a specific forecast"""
    plans = await db.production_plans.find(
        {"forecast_id": forecast_id},
        {"_id": 0}
    ).to_list(1000)
    
    total_planned = sum(p.get("planned_quantity", 0) for p in plans)
    
    return {
        "forecast_id": forecast_id,
        "production_plans": [serialize_doc(p) for p in plans],
        "total_planned_quantity": total_planned
    }

# --- Dispatch Lots ---
@router.get("/dispatch-lots")
async def get_dispatch_lots(
    buyer_id: Optional[str] = None,
    sku_id: Optional[str] = None,
    status: Optional[str] = None
):
    query = {}
    if buyer_id:
        query["buyer_id"] = buyer_id
    if sku_id:
        query["sku_id"] = sku_id
    if status:
        query["status"] = status
    lots = await db.dispatch_lots.find(query, {"_id": 0}).to_list(1000)
    return [serialize_doc(l) for l in lots]

@router.post("/dispatch-lots")
async def create_dispatch_lot(data: DispatchLotCreate):
    count = await db.dispatch_lots.count_documents({})
    lot_code = f"DL_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    lot = {
        "id": str(uuid.uuid4()),
        "lot_code": lot_code,
        "forecast_id": data.forecast_id,
        "sku_id": data.sku_id,
        "buyer_id": data.buyer_id,
        "required_quantity": data.required_quantity,
        "produced_quantity": 0,
        "qc_passed_quantity": 0,
        "dispatched_quantity": 0,
        "target_date": data.target_date,
        "status": "CREATED",
        "priority": data.priority,
        "created_at": datetime.now(timezone.utc)
    }
    await db.dispatch_lots.insert_one(lot)
    del lot["_id"]
    return serialize_doc(lot)

@router.put("/dispatch-lots/{lot_id}/status")
async def update_dispatch_lot_status(lot_id: str, status: str):
    valid_statuses = ["CREATED", "PRODUCTION_ASSIGNED", "PARTIALLY_PRODUCED", "FULLY_PRODUCED", 
                     "QC_CLEARED", "DISPATCH_READY", "DISPATCHED", "DELIVERED"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Valid: {valid_statuses}")
    
    result = await db.dispatch_lots.update_one(
        {"id": lot_id},
        {"$set": {"status": status}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch lot not found")
    return {"message": f"Status updated to {status}"}


# --- Dispatch Lot Cascade Filter Endpoints ---

@router.get("/dispatch-lots/buyers-with-forecasts")
async def get_buyers_with_forecasts():
    """Get only buyers who have confirmed forecasts"""
    # Get unique buyer_ids from confirmed forecasts
    forecasts = await db.forecasts.find(
        {"status": {"$in": ["CONFIRMED", "CONVERTED"]}, "buyer_id": {"$ne": None}},
        {"buyer_id": 1, "_id": 0}
    ).to_list(1000)
    
    buyer_ids = list(set(f["buyer_id"] for f in forecasts if f.get("buyer_id")))
    
    if not buyer_ids:
        return []
    
    # Get buyer details
    buyers = await db.buyers.find(
        {"id": {"$in": buyer_ids}},
        {"_id": 0}
    ).to_list(100)
    
    return buyers


@router.get("/dispatch-lots/brands-by-buyer")
async def get_brands_by_buyer(buyer_id: str):
    """Get brands linked to the buyer"""
    # First get brands directly linked to buyer
    brands = await db.brands.find(
        {"buyer_id": buyer_id, "status": "ACTIVE"},
        {"_id": 0}
    ).to_list(100)
    
    # Also get brands from SKUs that have forecasts for this buyer
    forecasts = await db.forecasts.find(
        {"buyer_id": buyer_id, "status": {"$in": ["CONFIRMED", "CONVERTED"]}},
        {"sku_id": 1, "_id": 0}
    ).to_list(1000)
    
    sku_ids = [f["sku_id"] for f in forecasts if f.get("sku_id")]
    if sku_ids:
        skus = await db.skus.find(
            {"sku_id": {"$in": sku_ids}},
            {"brand_id": 1, "brand": 1, "_id": 0}
        ).to_list(1000)
        
        # Get unique brand_ids from forecasted SKUs
        brand_ids_from_skus = list(set(s.get("brand_id") for s in skus if s.get("brand_id")))
        if brand_ids_from_skus:
            additional_brands = await db.brands.find(
                {"id": {"$in": brand_ids_from_skus}, "status": "ACTIVE"},
                {"_id": 0}
            ).to_list(100)
            
            # Merge without duplicates
            existing_ids = {b["id"] for b in brands}
            for ab in additional_brands:
                if ab["id"] not in existing_ids:
                    brands.append(ab)
    
    return brands


@router.get("/dispatch-lots/verticals-by-buyer")
async def get_verticals_by_buyer(buyer_id: str, brand_id: Optional[str] = None):
    """Get verticals that have forecasted SKUs for this buyer (optionally filtered by brand)"""
    # Get forecasts for this buyer
    forecast_query = {"buyer_id": buyer_id, "status": {"$in": ["CONFIRMED", "CONVERTED"]}}
    forecasts = await db.forecasts.find(forecast_query, {"sku_id": 1, "vertical_id": 1, "_id": 0}).to_list(1000)
    
    vertical_ids = set()
    
    # Collect vertical IDs from forecasts
    for f in forecasts:
        if f.get("vertical_id"):
            vertical_ids.add(f["vertical_id"])
    
    # Also get verticals from forecasted SKUs
    sku_ids = [f["sku_id"] for f in forecasts if f.get("sku_id")]
    if sku_ids:
        sku_query = {"sku_id": {"$in": sku_ids}}
        if brand_id:
            sku_query["brand_id"] = brand_id
        
        skus = await db.skus.find(sku_query, {"vertical_id": 1, "_id": 0}).to_list(1000)
        for s in skus:
            if s.get("vertical_id"):
                vertical_ids.add(s["vertical_id"])
    
    if not vertical_ids:
        return []
    
    # Get vertical details
    verticals = await db.verticals.find(
        {"id": {"$in": list(vertical_ids)}, "status": "ACTIVE"},
        {"_id": 0}
    ).to_list(100)
    
    return verticals


@router.get("/dispatch-lots/forecasted-skus")
async def get_forecasted_skus(buyer_id: str, vertical_id: Optional[str] = None, brand_id: Optional[str] = None):
    """Get SKUs with confirmed forecasts for buyer, with available quantity"""
    # Get confirmed forecasts for this buyer
    forecast_query = {"buyer_id": buyer_id, "status": {"$in": ["CONFIRMED", "CONVERTED"]}}
    if vertical_id:
        forecast_query["vertical_id"] = vertical_id
    
    forecasts = await db.forecasts.find(forecast_query, {"_id": 0}).to_list(1000)
    
    # Get SKU-level forecasts
    sku_forecasts = {}
    vertical_forecasts = {}
    
    for f in forecasts:
        if f.get("sku_id"):
            if f["sku_id"] not in sku_forecasts:
                sku_forecasts[f["sku_id"]] = {"forecast_qty": 0, "forecast_ids": []}
            sku_forecasts[f["sku_id"]]["forecast_qty"] += f.get("quantity", 0)
            sku_forecasts[f["sku_id"]]["forecast_ids"].append(f.get("id"))
        elif f.get("vertical_id"):
            # Vertical-level forecast - will be distributed to SKUs
            if f["vertical_id"] not in vertical_forecasts:
                vertical_forecasts[f["vertical_id"]] = {"forecast_qty": 0, "forecast_ids": []}
            vertical_forecasts[f["vertical_id"]]["forecast_qty"] += f.get("quantity", 0)
            vertical_forecasts[f["vertical_id"]]["forecast_ids"].append(f.get("id"))
    
    # Get SKUs that match the criteria
    sku_query = {}
    if vertical_id:
        sku_query["vertical_id"] = vertical_id
    if brand_id:
        sku_query["brand_id"] = brand_id
    
    # Get SKUs either directly forecasted or in forecasted verticals
    forecasted_sku_ids = list(sku_forecasts.keys())
    forecasted_vertical_ids = list(vertical_forecasts.keys())
    
    if forecasted_sku_ids or forecasted_vertical_ids:
        or_conditions = []
        if forecasted_sku_ids:
            or_conditions.append({"sku_id": {"$in": forecasted_sku_ids}})
        if forecasted_vertical_ids:
            or_conditions.append({"vertical_id": {"$in": forecasted_vertical_ids}})
        
        if sku_query:
            sku_query["$and"] = [{"$or": or_conditions}]
        else:
            sku_query = {"$or": or_conditions}
    
    skus = await db.skus.find(sku_query, {"_id": 0}).to_list(1000)
    
    # Apply brand filter if specified
    if brand_id:
        skus = [s for s in skus if s.get("brand_id") == brand_id]
    
    # Get existing dispatch lot quantities
    dispatch_lots = await db.dispatch_lots.find(
        {"buyer_id": buyer_id, "status": {"$ne": "CANCELLED"}},
        {"sku_id": 1, "required_quantity": 1, "_id": 0}
    ).to_list(1000)
    
    # Also get from dispatch_lot_lines
    lot_lines = await db.dispatch_lot_lines.find(
        {},
        {"sku_id": 1, "quantity": 1, "_id": 0}
    ).to_list(5000)
    
    dispatched_by_sku = {}
    for lot in dispatch_lots:
        sku = lot.get("sku_id")
        if sku:
            dispatched_by_sku[sku] = dispatched_by_sku.get(sku, 0) + lot.get("required_quantity", 0)
    
    for line in lot_lines:
        sku = line.get("sku_id")
        if sku:
            dispatched_by_sku[sku] = dispatched_by_sku.get(sku, 0) + line.get("quantity", 0)
    
    # Build result with available quantities
    result = []
    for sku in skus:
        sku_id = sku.get("sku_id")
        
        # Calculate forecast qty (direct SKU forecast + vertical forecast share)
        forecast_qty = 0
        if sku_id in sku_forecasts:
            forecast_qty = sku_forecasts[sku_id]["forecast_qty"]
        elif sku.get("vertical_id") in vertical_forecasts:
            # For vertical-level forecasts, show the vertical total
            forecast_qty = vertical_forecasts[sku["vertical_id"]]["forecast_qty"]
        
        if forecast_qty == 0:
            continue
        
        dispatched_qty = dispatched_by_sku.get(sku_id, 0)
        available_qty = max(0, forecast_qty - dispatched_qty)
        
        result.append({
            "sku_id": sku_id,
            "description": sku.get("description", ""),
            "brand": sku.get("brand", ""),
            "brand_id": sku.get("brand_id"),
            "vertical": sku.get("vertical", ""),
            "vertical_id": sku.get("vertical_id"),
            "model": sku.get("model", ""),
            "forecast_qty": forecast_qty,
            "dispatched_qty": dispatched_qty,
            "available_qty": available_qty
        })
    
    return result


@router.post("/dispatch-lots/multi")
async def create_dispatch_lot_multi(data: DispatchLotMultiCreate):
    """Create a dispatch lot with multiple SKU lines"""
    if not data.lines:
        raise HTTPException(status_code=400, detail="At least one line item is required")
    
    if not data.buyer_id:
        raise HTTPException(status_code=400, detail="Buyer is required")
    
    # Verify buyer exists
    buyer = await db.buyers.find_one({"id": data.buyer_id})
    if not buyer:
        raise HTTPException(status_code=404, detail="Buyer not found")
    
    # Generate lot code
    count = await db.dispatch_lots.count_documents({})
    lot_code = f"DL_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    # Calculate total quantity
    total_quantity = sum(line.quantity for line in data.lines)
    
    # Determine forecast_id: use the main one if provided, else try to derive from first line
    forecast_id = data.forecast_id
    if not forecast_id:
        # Try to get forecast_id from the first line that has one
        for line in data.lines:
            if line.forecast_id:
                forecast_id = line.forecast_id
                break
    
    # Create main dispatch lot record
    lot_id = str(uuid.uuid4())
    lot = {
        "id": lot_id,
        "lot_code": lot_code,
        "buyer_id": data.buyer_id,
        "forecast_id": forecast_id,  # Link to source forecast
        "target_date": data.target_date,
        "priority": data.priority,
        "notes": data.notes or "",
        "status": "CREATED",
        "total_quantity": total_quantity,
        "total_produced": 0,
        "total_dispatched": 0,
        "line_count": len(data.lines),
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.dispatch_lots.insert_one(lot)
    
    # Create line items
    lines_created = []
    for idx, line in enumerate(data.lines):
        line_record = {
            "id": str(uuid.uuid4()),
            "lot_id": lot_id,
            "lot_code": lot_code,
            "line_number": idx + 1,
            "sku_id": line.sku_id,
            "brand_id": line.brand_id,
            "vertical_id": line.vertical_id,
            "forecast_id": line.forecast_id,  # Track per-line forecast source
            "quantity": line.quantity,
            "produced_qty": 0,
            "dispatched_qty": 0,
            "status": "PENDING",
            "created_at": datetime.now(timezone.utc)
        }
        await db.dispatch_lot_lines.insert_one(line_record)
        del line_record["_id"]
        lines_created.append(line_record)
    
    del lot["_id"]
    lot["lines"] = lines_created
    
    return serialize_doc(lot)


@router.get("/dispatch-lots/{lot_id}/lines")
async def get_dispatch_lot_lines(lot_id: str):
    """Get line items for a dispatch lot"""
    lines = await db.dispatch_lot_lines.find(
        {"lot_id": lot_id},
        {"_id": 0}
    ).to_list(1000)
    return [serialize_doc(l) for l in lines]


# Model for adding line to existing lot
class AddLineToLotRequest(BaseModel):
    sku_id: str
    quantity: int
    forecast_id: Optional[str] = None
    brand_id: Optional[str] = None
    vertical_id: Optional[str] = None


@router.post("/dispatch-lots/{lot_id}/add-line")
async def add_line_to_dispatch_lot(lot_id: str, data: AddLineToLotRequest):
    """Add a new line item to an existing dispatch lot"""
    # Get the lot
    lot = await db.dispatch_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Dispatch lot not found")
    
    # Don't allow adding to dispatched/delivered lots
    if lot.get("status") in ["DISPATCHED", "DELIVERED"]:
        raise HTTPException(status_code=400, detail="Cannot add lines to dispatched or delivered lots")
    
    # Get current line count for this lot
    existing_lines = await db.dispatch_lot_lines.count_documents({"lot_id": lot_id})
    
    # Create new line
    line_record = {
        "id": str(uuid.uuid4()),
        "lot_id": lot_id,
        "lot_code": lot.get("lot_code"),
        "line_number": existing_lines + 1,
        "sku_id": data.sku_id,
        "brand_id": data.brand_id,
        "vertical_id": data.vertical_id,
        "forecast_id": data.forecast_id,
        "quantity": data.quantity,
        "produced_qty": 0,
        "dispatched_qty": 0,
        "status": "PENDING",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.dispatch_lot_lines.insert_one(line_record)
    
    # Update lot totals
    current_total = lot.get("total_quantity", 0)
    current_line_count = lot.get("line_count", existing_lines)
    
    await db.dispatch_lots.update_one(
        {"id": lot_id},
        {
            "$set": {
                "total_quantity": current_total + data.quantity,
                "line_count": current_line_count + 1,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    del line_record["_id"]
    return {
        "message": "Line added to dispatch lot",
        "line": serialize_doc(line_record),
        "new_lot_total": current_total + data.quantity
    }


@router.get("/dispatch-lots/by-buyer/{buyer_id}")
async def get_dispatch_lots_by_buyer(buyer_id: str, exclude_completed: bool = True):
    """Get all dispatch lots for a specific buyer"""
    query = {"buyer_id": buyer_id}
    if exclude_completed:
        query["status"] = {"$nin": ["DISPATCHED", "DELIVERED"]}
    
    lots = await db.dispatch_lots.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    return [serialize_doc(l) for l in lots]


@router.get("/dispatch-lots/{lot_id}/details")
async def get_dispatch_lot_details(lot_id: str):
    """Get dispatch lot with full details including readiness status"""
    # Get the lot
    lot = await db.dispatch_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Dispatch lot not found")
    
    # Get buyer info
    buyer = None
    if lot.get("buyer_id"):
        buyer = await db.buyers.find_one({"id": lot["buyer_id"]}, {"_id": 0, "name": 1, "code": 1})
    
    # Get lines for this lot
    lines = await db.dispatch_lot_lines.find(
        {"lot_id": lot_id},
        {"_id": 0}
    ).to_list(1000)
    
    # If no lines collection, check if it's old format (single SKU per lot)
    if not lines and lot.get("sku_id"):
        lines = [{
            "id": str(uuid.uuid4()),
            "lot_id": lot_id,
            "line_number": 1,
            "sku_id": lot["sku_id"],
            "quantity": lot.get("required_quantity", 0),
            "produced_qty": lot.get("produced_quantity", 0),
            "dispatched_qty": lot.get("dispatched_quantity", 0),
            "status": "PENDING"
        }]
    
    # Get SKU details and inventory for readiness calculation
    sku_ids = [line.get("sku_id") for line in lines if line.get("sku_id")]
    skus = await db.skus.find({"sku_id": {"$in": sku_ids}}, {"_id": 0}).to_list(1000)
    sku_map = {s["sku_id"]: s for s in skus}
    
    # Get FG inventory for these SKUs (sum across all branches)
    fg_inventory = await db.fg_inventory.find(
        {"sku_id": {"$in": sku_ids}, "status": "AVAILABLE"},
        {"sku_id": 1, "quantity": 1, "_id": 0}
    ).to_list(5000)
    
    # Also check production batches for produced quantity
    production_batches = await db.production_batches.find(
        {"sku_id": {"$in": sku_ids}, "status": {"$in": ["COMPLETED", "QC_PASSED", "FG_READY"]}},
        {"sku_id": 1, "good_quantity": 1, "_id": 0}
    ).to_list(5000)
    
    # Sum inventory by SKU
    inventory_by_sku = {}
    for inv in fg_inventory:
        sku = inv.get("sku_id")
        inventory_by_sku[sku] = inventory_by_sku.get(sku, 0) + inv.get("quantity", 0)
    
    # Add produced quantity from batches
    for batch in production_batches:
        sku = batch.get("sku_id")
        inventory_by_sku[sku] = inventory_by_sku.get(sku, 0) + batch.get("good_quantity", 0)
    
    # Also check SKU current_stock if no FG inventory
    for sku_id in sku_ids:
        if sku_id not in inventory_by_sku or inventory_by_sku[sku_id] == 0:
            sku_data = sku_map.get(sku_id, {})
            inventory_by_sku[sku_id] = sku_data.get("current_stock", 0)
    
    # Calculate readiness for each line
    lines_with_readiness = []
    total_ready = 0
    total_pending = 0
    
    for line in lines:
        sku_id = line.get("sku_id")
        sku_data = sku_map.get(sku_id, {})
        required_qty = line.get("quantity", 0)
        available_qty = inventory_by_sku.get(sku_id, 0)
        
        # Calculate readiness
        ready_qty = min(required_qty, available_qty)
        pending_qty = max(0, required_qty - available_qty)
        
        readiness_pct = (ready_qty / required_qty * 100) if required_qty > 0 else 0
        
        if readiness_pct >= 100:
            readiness_status = "READY"
            total_ready += 1
        elif readiness_pct > 0:
            readiness_status = "PARTIAL"
        else:
            readiness_status = "PENDING"
            total_pending += 1
        
        lines_with_readiness.append({
            **line,
            "sku_description": sku_data.get("description", ""),
            "brand": sku_data.get("brand", ""),
            "vertical": sku_data.get("vertical", ""),
            "model": sku_data.get("model", ""),
            "available_qty": available_qty,
            "ready_qty": ready_qty,
            "pending_qty": pending_qty,
            "readiness_pct": round(readiness_pct, 1),
            "readiness_status": readiness_status
        })
    
    # Calculate overall lot readiness
    total_lines = len(lines_with_readiness)
    if total_lines == 0:
        lot_readiness = "EMPTY"
        lot_readiness_pct = 0
    elif total_ready == total_lines:
        lot_readiness = "READY"
        lot_readiness_pct = 100
    elif total_ready > 0:
        lot_readiness = "PARTIAL"
        lot_readiness_pct = round(total_ready / total_lines * 100, 1)
    else:
        lot_readiness = "PENDING"
        lot_readiness_pct = 0
    
    return {
        **serialize_doc(lot),
        "buyer_name": buyer.get("name") if buyer else None,
        "buyer_code": buyer.get("code") if buyer else None,
        "lines": lines_with_readiness,
        "readiness_status": lot_readiness,
        "readiness_pct": lot_readiness_pct,
        "ready_lines": total_ready,
        "pending_lines": total_pending,
        "total_lines": total_lines
    }


@router.put("/dispatch-lots/{lot_id}")
async def update_dispatch_lot(lot_id: str, data: DispatchLotUpdate):
    """Update a dispatch lot and optionally its line items"""
    # Get existing lot
    lot = await db.dispatch_lots.find_one({"id": lot_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Dispatch lot not found")
    
    # Don't allow editing lots that are already dispatched or delivered
    if lot.get("status") in ["DISPATCHED", "DELIVERED"]:
        raise HTTPException(status_code=400, detail="Cannot edit dispatched or delivered lots")
    
    # Build update dict for lot-level fields
    update_fields = {"updated_at": datetime.now(timezone.utc)}
    
    if data.target_date is not None:
        update_fields["target_date"] = data.target_date
    if data.priority is not None:
        update_fields["priority"] = data.priority
    if data.notes is not None:
        update_fields["notes"] = data.notes
    
    # Handle line items update
    if data.lines is not None:
        if len(data.lines) == 0:
            raise HTTPException(status_code=400, detail="At least one line item is required")
        
        # Delete existing lines
        await db.dispatch_lot_lines.delete_many({"lot_id": lot_id})
        
        # Create new lines
        total_quantity = 0
        lines_created = []
        for idx, line in enumerate(data.lines):
            line_record = {
                "id": line.id if line.id else str(uuid.uuid4()),
                "lot_id": lot_id,
                "lot_code": lot.get("lot_code"),
                "line_number": idx + 1,
                "sku_id": line.sku_id,
                "brand_id": line.brand_id,
                "vertical_id": line.vertical_id,
                "forecast_id": line.forecast_id,
                "quantity": line.quantity,
                "produced_qty": 0,
                "dispatched_qty": 0,
                "status": "PENDING",
                "created_at": datetime.now(timezone.utc)
            }
            await db.dispatch_lot_lines.insert_one(line_record)
            del line_record["_id"]
            lines_created.append(line_record)
            total_quantity += line.quantity
        
        # Update lot totals
        update_fields["total_quantity"] = total_quantity
        update_fields["line_count"] = len(data.lines)
    
    # Apply updates
    await db.dispatch_lots.update_one(
        {"id": lot_id},
        {"$set": update_fields}
    )
    
    # Fetch updated lot
    updated_lot = await db.dispatch_lots.find_one({"id": lot_id}, {"_id": 0})
    
    # Get lines
    lines = await db.dispatch_lot_lines.find(
        {"lot_id": lot_id},
        {"_id": 0}
    ).to_list(1000)
    
    updated_lot["lines"] = [serialize_doc(l) for l in lines]
    
    return serialize_doc(updated_lot)


@router.get("/dispatch-lots/with-readiness")
async def get_dispatch_lots_with_readiness(
    buyer_id: Optional[str] = None,
    status: Optional[str] = None
):
    """Get all dispatch lots with readiness status"""
    query = {}
    if buyer_id:
        query["buyer_id"] = buyer_id
    if status:
        query["status"] = status
    
    lots = await db.dispatch_lots.find(query, {"_id": 0}).to_list(1000)
    
    # Get all lot IDs
    lot_ids = [lot.get("id") for lot in lots]
    
    # Get all lines for these lots
    all_lines = await db.dispatch_lot_lines.find(
        {"lot_id": {"$in": lot_ids}},
        {"_id": 0}
    ).to_list(10000)
    
    # Group lines by lot_id
    lines_by_lot = {}
    for line in all_lines:
        lot_id = line.get("lot_id")
        if lot_id not in lines_by_lot:
            lines_by_lot[lot_id] = []
        lines_by_lot[lot_id].append(line)
    
    # Get all unique SKU IDs
    sku_ids = set()
    for lot in lots:
        if lot.get("sku_id"):
            sku_ids.add(lot["sku_id"])
    for line in all_lines:
        if line.get("sku_id"):
            sku_ids.add(line["sku_id"])
    
    sku_ids = list(sku_ids)
    
    # Get SKU current stock
    skus = await db.skus.find({"sku_id": {"$in": sku_ids}}, {"_id": 0, "sku_id": 1, "current_stock": 1}).to_list(5000)
    stock_by_sku = {s["sku_id"]: s.get("current_stock", 0) for s in skus}
    
    # Get FG inventory
    fg_inventory = await db.fg_inventory.find(
        {"sku_id": {"$in": sku_ids}, "status": "AVAILABLE"},
        {"sku_id": 1, "quantity": 1, "_id": 0}
    ).to_list(10000)
    
    for inv in fg_inventory:
        sku = inv.get("sku_id")
        stock_by_sku[sku] = stock_by_sku.get(sku, 0) + inv.get("quantity", 0)
    
    # Get buyer names
    buyer_ids = list(set(lot.get("buyer_id") for lot in lots if lot.get("buyer_id")))
    buyers = await db.buyers.find({"id": {"$in": buyer_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    buyer_map = {b["id"]: b["name"] for b in buyers}
    
    # Calculate readiness for each lot
    result = []
    for lot in lots:
        lot_id = lot.get("id")
        lines = lines_by_lot.get(lot_id, [])
        
        # For old format lots
        if not lines and lot.get("sku_id"):
            lines = [{
                "sku_id": lot["sku_id"],
                "quantity": lot.get("required_quantity", 0)
            }]
        
        ready_count = 0
        total_count = len(lines)
        
        for line in lines:
            sku_id = line.get("sku_id")
            required = line.get("quantity", 0)
            available = stock_by_sku.get(sku_id, 0)
            
            if available >= required and required > 0:
                ready_count += 1
        
        if total_count == 0:
            readiness = "EMPTY"
            readiness_pct = 0
        elif ready_count == total_count:
            readiness = "READY"
            readiness_pct = 100
        elif ready_count > 0:
            readiness = "PARTIAL"
            readiness_pct = round(ready_count / total_count * 100, 1)
        else:
            readiness = "PENDING"
            readiness_pct = 0
        
        result.append({
            **serialize_doc(lot),
            "buyer_name": buyer_map.get(lot.get("buyer_id")),
            "readiness_status": readiness,
            "readiness_pct": readiness_pct,
            "ready_lines": ready_count,
            "total_lines": total_count
        })
    
    return result


# --- Bulk Upload ---
@router.post("/forecasts/parse-excel")
async def parse_forecast_excel(file: UploadFile = File(...)):
    """
    Parse Excel/CSV file for bulk forecast upload.
    Required columns: Month, SKU (SKU ID), Customer Code, Qty (Quantity)
    Vertical, Brand, Model are auto-filled from SKU master data.
    Returns validated forecasts and errors for invalid entries.
    """
    import openpyxl
    import csv
    
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    contents = await file.read()
    raw_rows = []
    
    try:
        if file.filename.endswith('.csv'):
            # Parse CSV
            text = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            for row_num, row in enumerate(reader, start=2):
                raw_rows.append({
                    "row_num": row_num,
                    "month": row.get('Month', row.get('month', '')),
                    "sku_id": row.get('SKU', row.get('sku', row.get('SKU ID', row.get('sku_id', '')))),
                    "quantity": row.get('Qty', row.get('qty', row.get('Quantity', row.get('quantity', 0)))),
                    "customer_code": row.get('Customer Code', row.get('customer_code', row.get('Customer_Code', '')))
                })
        else:
            # Parse Excel - only read the first sheet (Forecast Upload)
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
            ws = wb.active
            
            headers = []
            header_row_idx = None
            
            # Find the header row (first row with data that contains 'month' or 'sku')
            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                if not any(row):
                    continue
                row_lower = [str(cell).lower().strip() if cell else '' for cell in row]
                if any('month' in cell or 'sku' in cell for cell in row_lower):
                    headers = [str(h).lower().strip() if h else '' for h in row]
                    header_row_idx = row_idx
                    break
            
            if not headers:
                raise HTTPException(status_code=400, detail="Could not find header row. Expected columns: Month, SKU, Customer Code, Qty")
            
            # Parse data rows (starting after header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 2, values_only=True)):
                if not any(row):
                    continue
                
                row_data = dict(zip(headers, row))
                
                # Handle month - could be datetime or string
                month_val = row_data.get('month', '')
                if hasattr(month_val, 'strftime'):
                    month_val = month_val.strftime('%Y-%m')
                elif month_val:
                    month_val = str(month_val)
                
                raw_rows.append({
                    "row_num": header_row_idx + 2 + row_idx,
                    "month": month_val,
                    "sku_id": str(row_data.get('sku', row_data.get('sku id', row_data.get('sku_id', ''))) or ''),
                    "quantity": row_data.get('qty', row_data.get('quantity', 0)),
                    "customer_code": str(row_data.get('customer code', row_data.get('customer_code', '')) or '')
                })
            wb.close()
        
        # Load SKU master data for validation and auto-fill
        skus = await db.skus.find({}, {"_id": 0}).to_list(50000)
        sku_map = {s.get("sku_id", "").upper(): s for s in skus}
        
        # Load verticals for mapping
        verticals_list = await db.verticals.find({}, {"_id": 0}).to_list(100)
        vertical_by_id = {v['id']: v for v in verticals_list}
        
        # Load brands for mapping
        brands_list = await db.brands.find({}, {"_id": 0}).to_list(1000)
        brand_by_id = {b['id']: b for b in brands_list}
        
        # Load models for mapping
        models_list = await db.models.find({}, {"_id": 0}).to_list(5000)
        model_by_id = {m['id']: m for m in models_list}
        
        # Load buyers for mapping by customer_code
        buyers_list = await db.buyers.find({"status": "ACTIVE"}, {"_id": 0}).to_list(1000)
        buyer_by_code = {b.get('customer_code', '').upper(): b for b in buyers_list if b.get('customer_code')}
        
        # Process rows - validate and enrich
        valid_forecasts = []
        errors = []
        
        for row in raw_rows:
            sku_id = str(row.get('sku_id', '')).strip().upper()
            
            # Validate required fields
            if not sku_id:
                errors.append({
                    "row_num": row['row_num'],
                    "sku_id": sku_id or "(empty)",
                    "reason": "SKU ID is required"
                })
                continue
            
            if not row.get('month'):
                errors.append({
                    "row_num": row['row_num'],
                    "sku_id": sku_id,
                    "reason": "Month is required"
                })
                continue
            
            try:
                quantity = int(row.get('quantity', 0) or 0)
                if quantity <= 0:
                    errors.append({
                        "row_num": row['row_num'],
                        "sku_id": sku_id,
                        "reason": "Quantity must be greater than 0"
                    })
                    continue
            except (ValueError, TypeError):
                errors.append({
                    "row_num": row['row_num'],
                    "sku_id": sku_id,
                    "reason": f"Invalid quantity: {row.get('quantity')}"
                })
                continue
            
            # Validate SKU exists in system
            sku = sku_map.get(sku_id)
            if not sku:
                errors.append({
                    "row_num": row['row_num'],
                    "sku_id": sku_id,
                    "reason": "SKU not found in system"
                })
                continue
            
            # Auto-fill from SKU master data
            vertical_id = sku.get('vertical_id')
            brand_id = sku.get('brand_id')
            model_id = sku.get('model_id')
            
            vertical_name = vertical_by_id.get(vertical_id, {}).get('name', '')
            brand_name = brand_by_id.get(brand_id, {}).get('name', '')
            model_name = model_by_id.get(model_id, {}).get('name', '')
            
            # Handle customer code (REQUIRED for forecast creation)
            customer_code = str(row.get('customer_code', '')).strip().upper()
            if not customer_code:
                errors.append({
                    "row_num": row['row_num'],
                    "sku_id": sku_id,
                    "reason": "Customer Code is required. Use codes from the 'Buyer Master' sheet."
                })
                continue
            
            buyer = buyer_by_code.get(customer_code)
            if not buyer:
                errors.append({
                    "row_num": row['row_num'],
                    "sku_id": sku_id,
                    "reason": f"Customer Code '{customer_code}' not found. Check the 'Buyer Master' sheet for valid codes."
                })
                continue
            
            buyer_id = buyer.get('id')
            buyer_name = buyer.get('name', '')
            buyer_name = buyer.get('name')
            
            # Build valid forecast
            valid_forecasts.append({
                "row_num": row['row_num'],
                "month": str(row['month']),
                "vertical": vertical_name,
                "vertical_id": vertical_id,
                "model": model_name,
                "model_id": model_id,
                "brand": brand_name,
                "brand_id": brand_id,
                "sku_id": sku_id,
                "quantity": quantity,
                "buyer": buyer_name,
                "buyer_id": buyer_id
            })
        
        return {
            "forecasts": valid_forecasts,
            "count": len(valid_forecasts),
            "errors": errors,
            "error_count": len(errors)
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")


@router.post("/forecasts/generate-error-report")
async def generate_forecast_error_report(errors: List[dict]):
    """
    Generate Excel error report for failed forecast uploads.
    Returns downloadable Excel file with error details.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload Errors"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    
    # Headers
    headers = ["Row #", "SKU ID", "Error Reason"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, error in enumerate(errors, 2):
        ws.cell(row=row_idx, column=1, value=error.get("row_num", ""))
        ws.cell(row=row_idx, column=2, value=error.get("sku_id", ""))
        ws.cell(row=row_idx, column=3, value=error.get("reason", ""))
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=forecast_upload_errors.xlsx"}
    )


# =============================================================================
# DISPATCH LOT BULK UPLOAD
# =============================================================================

@router.get("/forecasts/export")
async def export_forecasts(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    buyer_id: Optional[str] = None,
    brand: Optional[str] = None,
    model: Optional[str] = None,
    status: Optional[str] = None
):
    """
    Export forecasts as Excel with filters.
    Used for creating dispatch lot bulk upload data.
    Filters: date range, buyer, brand, model, status
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    
    # Build query
    query = {}
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            if "forecast_month" not in query:
                query["forecast_month"] = {}
            query["forecast_month"]["$gte"] = start
        except:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if "forecast_month" not in query:
                query["forecast_month"] = {}
            query["forecast_month"]["$lte"] = end
        except:
            pass
    
    if buyer_id:
        query["buyer_id"] = buyer_id
    
    if status:
        query["status"] = status
    
    # Fetch forecasts
    forecasts = await db.forecasts.find(query, {"_id": 0}).to_list(10000)
    
    if not forecasts:
        raise HTTPException(status_code=404, detail="No forecasts found matching filters")
    
    # Load reference data
    buyers = await db.buyers.find({}, {"_id": 0}).to_list(1000)
    buyer_map = {b["id"]: b for b in buyers}
    
    skus = await db.skus.find({}, {"_id": 0}).to_list(10000)
    sku_map = {s["sku_id"]: s for s in skus}
    
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(100)
    vertical_map = {v["id"]: v for v in verticals}
    
    models_list = await db.models.find({}, {"_id": 0}).to_list(1000)
    model_map = {m["id"]: m for m in models_list}
    
    # Filter by brand/model if specified (need to filter in memory after SKU lookup)
    filtered_forecasts = []
    for f in forecasts:
        sku = sku_map.get(f.get("sku_id"), {})
        sku_brand = sku.get("brand", "")
        sku_model = sku.get("model", "")
        
        # Apply brand filter
        if brand and brand.lower() not in sku_brand.lower():
            continue
        
        # Apply model filter
        if model and model.lower() not in sku_model.lower():
            continue
        
        filtered_forecasts.append(f)
    
    if not filtered_forecasts:
        raise HTTPException(status_code=404, detail="No forecasts found matching brand/model filters")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast Export"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    # Headers matching dispatch lot upload format
    headers = ["Forecast No", "Month", "Buyer Name", "Vertical", "Model", "Brand", "SKU ID", "SKU Description", "Forecast Qty", "Dispatched Qty", "Available Qty", "Status", "Priority"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, f in enumerate(filtered_forecasts, 2):
        buyer = buyer_map.get(f.get("buyer_id"), {})
        sku = sku_map.get(f.get("sku_id"), {})
        vertical = vertical_map.get(f.get("vertical_id"), {})
        
        forecast_month = f.get("forecast_month")
        if isinstance(forecast_month, datetime):
            month_str = forecast_month.strftime("%Y-%m")
        elif isinstance(forecast_month, str):
            month_str = forecast_month[:7]
        else:
            month_str = ""
        
        dispatched = f.get("dispatched_quantity", 0) or 0
        forecast_qty = f.get("quantity", 0)
        available = max(0, forecast_qty - dispatched)
        
        ws.cell(row=row_idx, column=1, value=f.get("forecast_code", ""))
        ws.cell(row=row_idx, column=2, value=month_str)
        ws.cell(row=row_idx, column=3, value=buyer.get("name", ""))
        ws.cell(row=row_idx, column=4, value=vertical.get("name", "") or sku.get("vertical", ""))
        ws.cell(row=row_idx, column=5, value=sku.get("model", ""))
        ws.cell(row=row_idx, column=6, value=sku.get("brand", ""))
        ws.cell(row=row_idx, column=7, value=f.get("sku_id", ""))
        ws.cell(row=row_idx, column=8, value=sku.get("description", ""))
        ws.cell(row=row_idx, column=9, value=forecast_qty)
        ws.cell(row=row_idx, column=10, value=dispatched)
        ws.cell(row=row_idx, column=11, value=available)
        ws.cell(row=row_idx, column=12, value=f.get("status", ""))
        ws.cell(row=row_idx, column=13, value=f.get("priority", ""))
    
    # Set column widths
    col_widths = [18, 10, 25, 15, 15, 15, 20, 35, 12, 12, 12, 12, 10]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"forecast_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/dispatch-lots/template")
async def download_dispatch_lot_template():
    """
    Download Excel template for dispatch lot bulk upload.
    Columns: Buyer Name | Forecast No | SKU ID | Qty | Serial No
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dispatch Lot Upload"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    
    # Headers
    headers = ["Buyer Name", "Forecast No", "SKU ID", "Qty", "Serial No"]
    comments = [
        "Exact buyer name as in system",
        "Forecast code (e.g., FC_202603_0001)",
        "SKU ID (e.g., CC_KS_BE_188)",
        "Quantity for this line",
        "Lot grouping number - same Serial No = same lot"
    ]
    
    for col, (header, comment) in enumerate(zip(headers, comments), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.comment = Comment(comment, "System")
    
    # Sample data rows
    sample_data = [
        ["Test Buyer Inc", "FC_202603_0001", "CC_KS_BE_188", 500, 1],
        ["Test Buyer Inc", "FC_202603_0001", "CC_KS_BE_189", 300, 1],
        ["Test Buyer Inc", "FC_202603_0002", "CC_KS_BE_002", 200, 2],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Instructions sheet
    ws_help = wb.create_sheet("Instructions")
    instructions = [
        ["DISPATCH LOT BULK UPLOAD INSTRUCTIONS"],
        [""],
        ["COLUMNS:"],
        ["Buyer Name - Must match exactly with buyer name in system (case-insensitive)"],
        ["Forecast No - The forecast code to link this dispatch to (e.g., FC_202603_0001)"],
        ["SKU ID - Valid SKU ID from the system"],
        ["Qty - Quantity for this line item (must be > 0)"],
        ["Serial No - Temporary grouping number. Rows with SAME Serial No become ONE dispatch lot"],
        [""],
        ["EXAMPLE:"],
        ["Serial No 1: Creates lot with 2 lines (CC_KS_BE_188: 500, CC_KS_BE_189: 300)"],
        ["Serial No 2: Creates separate lot with 1 line (CC_KS_BE_002: 200)"],
        [""],
        ["TIP: Use the Forecast Export feature to get valid Forecast Numbers and SKU IDs"],
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        cell = ws_help.cell(row=row_idx, column=1, value=row[0] if row else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
    
    ws_help.column_dimensions['A'].width = 80
    
    # Set column widths on main sheet
    col_widths = [25, 20, 20, 10, 12]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dispatch_lot_template.xlsx"}
    )


class DispatchLotBulkUploadLine(BaseModel):
    """Single line from bulk upload"""
    buyer_name: str
    forecast_no: str
    sku_id: str
    quantity: int
    serial_no: int  # Temporary lot grouping identifier


@router.post("/dispatch-lots/bulk-upload")
async def bulk_upload_dispatch_lots(file: UploadFile = File(...)):
    """
    Bulk upload dispatch lots from Excel.
    Expected columns: Buyer Name | Forecast No | SKU ID | Qty | Serial No
    Rows with same Serial No become one dispatch lot with multiple lines.
    """
    import openpyxl
    
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    contents = await file.read()
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Get headers
        headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
        
        # Map columns
        col_map = {}
        for idx, h in enumerate(headers):
            if "buyer" in h and "name" in h:
                col_map["buyer_name"] = idx
            elif "forecast" in h:
                col_map["forecast_no"] = idx
            elif "sku" in h:
                col_map["sku_id"] = idx
            elif "qty" in h or "quantity" in h:
                col_map["quantity"] = idx
            elif "serial" in h:
                col_map["serial_no"] = idx
        
        required = ["buyer_name", "forecast_no", "sku_id", "quantity", "serial_no"]
        missing = [r for r in required if r not in col_map]
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing columns: {missing}. Expected: Buyer Name, Forecast No, SKU ID, Qty, Serial No")
        
        # Parse rows and group by serial_no
        lots_data = {}  # serial_no -> {buyer_name, lines: [...]}
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            try:
                buyer_name = str(row[col_map["buyer_name"]]).strip() if row[col_map["buyer_name"]] else ""
                forecast_no = str(row[col_map["forecast_no"]]).strip() if row[col_map["forecast_no"]] else ""
                sku_id = str(row[col_map["sku_id"]]).strip() if row[col_map["sku_id"]] else ""
                quantity = int(row[col_map["quantity"]] or 0)
                serial_no = int(row[col_map["serial_no"]] or 0)
                
                if not buyer_name or not sku_id or quantity <= 0 or serial_no <= 0:
                    errors.append(f"Row {row_num}: Invalid data (missing buyer/sku/qty/serial)")
                    continue
                
                if serial_no not in lots_data:
                    lots_data[serial_no] = {"buyer_name": buyer_name, "lines": []}
                
                lots_data[serial_no]["lines"].append({
                    "forecast_no": forecast_no,
                    "sku_id": sku_id,
                    "quantity": quantity
                })
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        wb.close()
        
        # Get buyers map
        buyers = await db.buyers.find({"status": "ACTIVE"}, {"_id": 0}).to_list(1000)
        buyer_map = {b["name"].lower(): b for b in buyers}
        
        # Get forecasts map
        forecasts = await db.forecasts.find({}, {"_id": 0}).to_list(10000)
        forecast_map = {f["forecast_code"]: f for f in forecasts}
        
        # Get SKUs map
        skus = await db.skus.find({}, {"_id": 0}).to_list(10000)
        sku_map = {s["sku_id"]: s for s in skus}
        
        # Create dispatch lots
        lots_created = 0
        lines_created = 0
        
        for serial_no, lot_data in sorted(lots_data.items()):
            buyer_name = lot_data["buyer_name"]
            buyer = buyer_map.get(buyer_name.lower())
            
            if not buyer:
                errors.append(f"Serial {serial_no}: Buyer '{buyer_name}' not found")
                continue
            
            # Generate lot code
            count = await db.dispatch_lots.count_documents({})
            lot_code = f"DL_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
            lot_id = str(uuid.uuid4())
            
            # Calculate total quantity and process lines
            total_quantity = 0
            lines_to_create = []
            
            for idx, line in enumerate(lot_data["lines"]):
                sku_id = line["sku_id"]
                sku = sku_map.get(sku_id)
                
                if not sku:
                    errors.append(f"Serial {serial_no}, Line {idx+1}: SKU '{sku_id}' not found")
                    continue
                
                # Find forecast
                forecast = forecast_map.get(line["forecast_no"])
                forecast_id = forecast["id"] if forecast else None
                
                line_record = {
                    "id": str(uuid.uuid4()),
                    "lot_id": lot_id,
                    "lot_code": lot_code,
                    "line_number": idx + 1,
                    "sku_id": sku_id,
                    "brand_id": sku.get("brand_id"),
                    "vertical_id": sku.get("vertical_id"),
                    "forecast_id": forecast_id,
                    "forecast_code": line["forecast_no"],
                    "quantity": line["quantity"],
                    "allocated_inventory": 0,
                    "produced_qty": 0,
                    "dispatched_qty": 0,
                    "scheduled_date": None,
                    "actual_completion_date": None,
                    "status": "PENDING",
                    "created_at": datetime.now(timezone.utc)
                }
                lines_to_create.append(line_record)
                total_quantity += line["quantity"]
            
            if not lines_to_create:
                continue
            
            # Create lot
            lot = {
                "id": lot_id,
                "lot_code": lot_code,
                "buyer_id": buyer["id"],
                "buyer_name": buyer["name"],
                "target_date": None,  # Will be computed from production schedules
                "priority": "MEDIUM",
                "notes": f"Bulk upload - Serial #{serial_no}",
                "status": "CREATED",
                "total_quantity": total_quantity,
                "total_allocated": 0,
                "total_produced": 0,
                "total_dispatched": 0,
                "line_count": len(lines_to_create),
                "estimated_completion_date": None,
                "created_at": datetime.now(timezone.utc)
            }
            
            await db.dispatch_lots.insert_one(lot)
            lots_created += 1
            
            # Create lines
            if lines_to_create:
                await db.dispatch_lot_lines.insert_many(lines_to_create)
                lines_created += len(lines_to_create)
        
        # Run FIFO allocation after creating lots
        await run_fifo_allocation()
        
        # Update lot production dates
        await update_lot_production_dates()
        
        return {
            "message": f"Bulk upload complete: {lots_created} lots created with {lines_created} lines",
            "lots_created": lots_created,
            "lines_created": lines_created,
            "errors": errors[:20] if errors else []
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")


# =============================================================================
# FIFO INVENTORY ALLOCATION
# =============================================================================

async def run_fifo_allocation():
    """
    Auto-allocate FG inventory to dispatch lots on FIFO basis.
    Oldest lots (by created_at) get inventory first.
    """
    # Get all active dispatch lots ordered by creation date (FIFO)
    lots = await db.dispatch_lots.find(
        {"status": {"$nin": ["DISPATCHED", "DELIVERED", "CANCELLED"]}},
        {"_id": 0}
    ).sort("created_at", 1).to_list(10000)
    
    if not lots:
        return
    
    lot_ids = [lot["id"] for lot in lots]
    
    # Get all lines for these lots
    all_lines = await db.dispatch_lot_lines.find(
        {"lot_id": {"$in": lot_ids}},
        {"_id": 0}
    ).to_list(50000)
    
    # Group lines by lot_id and maintain lot order
    lines_by_lot = {}
    for lot in lots:
        lines_by_lot[lot["id"]] = []
    
    for line in all_lines:
        lot_id = line.get("lot_id")
        if lot_id in lines_by_lot:
            lines_by_lot[lot_id].append(line)
    
    # Get all unique SKU IDs
    sku_ids = list(set(line.get("sku_id") for line in all_lines if line.get("sku_id")))
    
    # Get current inventory by SKU
    # From FG inventory
    fg_inventory = await db.fg_inventory.find(
        {"sku_id": {"$in": sku_ids}, "status": "AVAILABLE"},
        {"sku_id": 1, "quantity": 1, "_id": 0}
    ).to_list(50000)
    
    inventory_by_sku = {}
    for inv in fg_inventory:
        sku = inv.get("sku_id")
        inventory_by_sku[sku] = inventory_by_sku.get(sku, 0) + inv.get("quantity", 0)
    
    # Also add from SKU current_stock if no FG inventory
    skus = await db.skus.find({"sku_id": {"$in": sku_ids}}, {"_id": 0, "sku_id": 1, "current_stock": 1}).to_list(10000)
    for sku in skus:
        sid = sku.get("sku_id")
        if sid not in inventory_by_sku or inventory_by_sku[sid] == 0:
            inventory_by_sku[sid] = sku.get("current_stock", 0)
    
    # Track remaining inventory as we allocate
    remaining_inventory = dict(inventory_by_sku)
    
    # Allocate inventory to lots in FIFO order
    for lot in lots:
        lot_id = lot["id"]
        lines = lines_by_lot.get(lot_id, [])
        lot_total_allocated = 0
        
        for line in lines:
            sku_id = line.get("sku_id")
            required = line.get("quantity", 0)
            available = remaining_inventory.get(sku_id, 0)
            
            # Allocate what we can
            allocated = min(required, available)
            
            # Update line allocation
            await db.dispatch_lot_lines.update_one(
                {"id": line["id"]},
                {"$set": {
                    "allocated_inventory": allocated,
                    "status": "READY" if allocated >= required else ("PARTIAL" if allocated > 0 else "PENDING")
                }}
            )
            
            # Reduce remaining inventory
            remaining_inventory[sku_id] = available - allocated
            lot_total_allocated += allocated
        
        # Update lot total allocation
        await db.dispatch_lots.update_one(
            {"id": lot_id},
            {"$set": {"total_allocated": lot_total_allocated}}
        )


@router.post("/dispatch-lots/run-fifo-allocation")
async def trigger_fifo_allocation():
    """Manually trigger FIFO inventory allocation"""
    await run_fifo_allocation()
    return {"message": "FIFO allocation completed"}


# =============================================================================
# PRODUCTION DATE TRACKING
# =============================================================================

async def update_lot_production_dates():
    """
    Update dispatch lot lines with production dates from schedules.
    Lot completion = last production date across all lines.
    """
    # Get all active lots
    lots = await db.dispatch_lots.find(
        {"status": {"$nin": ["DISPATCHED", "DELIVERED", "CANCELLED"]}},
        {"_id": 0}
    ).to_list(10000)
    
    if not lots:
        return
    
    lot_ids = [lot["id"] for lot in lots]
    
    # Get all lines
    all_lines = await db.dispatch_lot_lines.find(
        {"lot_id": {"$in": lot_ids}},
        {"_id": 0}
    ).to_list(50000)
    
    # Get forecast IDs from lines
    forecast_ids = list(set(line.get("forecast_id") for line in all_lines if line.get("forecast_id")))
    
    # Get production schedules for these forecasts
    schedules = await db.production_schedules.find(
        {"forecast_id": {"$in": forecast_ids}, "status": {"$ne": "CANCELLED"}},
        {"_id": 0}
    ).to_list(50000)
    
    # Group schedules by forecast_id
    schedules_by_forecast = {}
    for s in schedules:
        fid = s.get("forecast_id")
        if fid:
            if fid not in schedules_by_forecast:
                schedules_by_forecast[fid] = []
            schedules_by_forecast[fid].append(s)
    
    # Update each line with production dates
    lines_by_lot = {}
    for line in all_lines:
        lot_id = line.get("lot_id")
        if lot_id not in lines_by_lot:
            lines_by_lot[lot_id] = []
        
        forecast_id = line.get("forecast_id")
        schedules_for_line = schedules_by_forecast.get(forecast_id, [])
        
        # Get scheduled date (earliest) and actual completion (latest completed)
        scheduled_date = None
        actual_completion = None
        
        for s in schedules_for_line:
            sched_date = s.get("scheduled_date") or s.get("production_date")
            if sched_date:
                if isinstance(sched_date, str):
                    try:
                        sched_date = datetime.fromisoformat(sched_date.replace('Z', '+00:00'))
                    except:
                        pass
                if isinstance(sched_date, datetime):
                    if scheduled_date is None or sched_date < scheduled_date:
                        scheduled_date = sched_date
            
            # Check for completion
            if s.get("status") in ["COMPLETED", "DONE"]:
                comp_date = s.get("completed_at") or s.get("scheduled_date")
                if comp_date:
                    if isinstance(comp_date, str):
                        try:
                            comp_date = datetime.fromisoformat(comp_date.replace('Z', '+00:00'))
                        except:
                            pass
                    if isinstance(comp_date, datetime):
                        if actual_completion is None or comp_date > actual_completion:
                            actual_completion = comp_date
        
        # Update line
        update_data = {}
        if scheduled_date:
            update_data["scheduled_date"] = scheduled_date
        if actual_completion:
            update_data["actual_completion_date"] = actual_completion
        
        if update_data:
            await db.dispatch_lot_lines.update_one(
                {"id": line["id"]},
                {"$set": update_data}
            )
        
        # Store for lot-level calculation
        lines_by_lot[lot_id].append({
            **line,
            "scheduled_date": scheduled_date,
            "actual_completion_date": actual_completion
        })
    
    # Update lot-level estimated completion (last date among all lines)
    for lot_id, lines in lines_by_lot.items():
        latest_date = None
        for line in lines:
            # Use actual if available, else scheduled
            line_date = line.get("actual_completion_date") or line.get("scheduled_date")
            if line_date:
                if latest_date is None or line_date > latest_date:
                    latest_date = line_date
        
        if latest_date:
            await db.dispatch_lots.update_one(
                {"id": lot_id},
                {"$set": {"estimated_completion_date": latest_date}}
            )


@router.post("/dispatch-lots/update-production-dates")
async def trigger_update_production_dates():
    """Manually trigger production date update for all lots"""
    await update_lot_production_dates()
    return {"message": "Production dates updated for all lots"}


# =============================================================================
# DISPATCH LOT DETAILS WITH PRODUCTION INFO
# =============================================================================

@router.get("/dispatch-lots/{lot_id}/full-details")
async def get_dispatch_lot_full_details(lot_id: str):
    """
    Get dispatch lot with full details including:
    - Production dates (scheduled vs actual) per line
    - FIFO allocated inventory per line
    - Visual indicator if current inventory can complete lot
    - Lot completion timeline
    """
    # Get the lot
    lot = await db.dispatch_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Dispatch lot not found")
    
    # Get buyer info
    buyer = None
    if lot.get("buyer_id"):
        buyer = await db.buyers.find_one({"id": lot["buyer_id"]}, {"_id": 0})
    
    # Get lines
    lines = await db.dispatch_lot_lines.find({"lot_id": lot_id}, {"_id": 0}).to_list(1000)
    
    # Get SKU details
    sku_ids = [line.get("sku_id") for line in lines if line.get("sku_id")]
    skus = await db.skus.find({"sku_id": {"$in": sku_ids}}, {"_id": 0}).to_list(1000)
    sku_map = {s["sku_id"]: s for s in skus}
    
    # Get total inventory (not allocated, just total available)
    fg_inventory = await db.fg_inventory.find(
        {"sku_id": {"$in": sku_ids}, "status": "AVAILABLE"},
        {"sku_id": 1, "quantity": 1, "_id": 0}
    ).to_list(5000)
    
    total_inventory_by_sku = {}
    for inv in fg_inventory:
        sku = inv.get("sku_id")
        total_inventory_by_sku[sku] = total_inventory_by_sku.get(sku, 0) + inv.get("quantity", 0)
    
    # Also add SKU current_stock
    for sku in skus:
        sid = sku.get("sku_id")
        if sid not in total_inventory_by_sku or total_inventory_by_sku[sid] == 0:
            total_inventory_by_sku[sid] = sku.get("current_stock", 0)
    
    # Get production schedules for forecasts in this lot
    forecast_ids = [line.get("forecast_id") for line in lines if line.get("forecast_id")]
    schedules = await db.production_schedules.find(
        {"forecast_id": {"$in": forecast_ids}},
        {"_id": 0}
    ).to_list(5000)
    
    schedules_by_forecast = {}
    for s in schedules:
        fid = s.get("forecast_id")
        if fid:
            if fid not in schedules_by_forecast:
                schedules_by_forecast[fid] = []
            schedules_by_forecast[fid].append(s)
    
    # Process lines
    lines_enriched = []
    latest_completion = None
    can_complete_with_current_inventory = True
    
    for line in lines:
        sku_id = line.get("sku_id")
        sku = sku_map.get(sku_id, {})
        required_qty = line.get("quantity", 0)
        allocated_qty = line.get("allocated_inventory", 0)
        total_available = total_inventory_by_sku.get(sku_id, 0)
        
        # Check if current inventory can complete this line (visual indicator)
        can_complete_line = total_available >= required_qty
        if not can_complete_line:
            can_complete_with_current_inventory = False
        
        # Get production schedule info
        forecast_id = line.get("forecast_id")
        line_schedules = schedules_by_forecast.get(forecast_id, [])
        
        scheduled_date = line.get("scheduled_date")
        actual_date = line.get("actual_completion_date")
        
        # Determine line status
        line_status = line.get("status", "PENDING")
        if allocated_qty >= required_qty:
            line_status = "READY"
        elif actual_date:
            line_status = "PRODUCED"
        elif scheduled_date:
            line_status = "SCHEDULED"
        
        # Check for delay
        is_delayed = False
        if scheduled_date and not actual_date:
            if isinstance(scheduled_date, str):
                try:
                    scheduled_date = datetime.fromisoformat(scheduled_date.replace('Z', '+00:00'))
                except:
                    pass
            if isinstance(scheduled_date, datetime):
                if scheduled_date < datetime.now(timezone.utc):
                    is_delayed = True
        
        # Track latest completion for lot timeline
        line_completion = actual_date or scheduled_date
        if line_completion:
            if isinstance(line_completion, str):
                try:
                    line_completion = datetime.fromisoformat(line_completion.replace('Z', '+00:00'))
                except:
                    pass
            if isinstance(line_completion, datetime):
                if latest_completion is None or line_completion > latest_completion:
                    latest_completion = line_completion
        
        lines_enriched.append({
            **line,
            "sku_description": sku.get("description", ""),
            "brand": sku.get("brand", ""),
            "vertical": sku.get("vertical", ""),
            "model": sku.get("model", ""),
            "allocated_inventory": allocated_qty,
            "total_available_inventory": total_available,
            "can_complete_with_current_inventory": can_complete_line,
            "readiness_pct": round((allocated_qty / required_qty * 100) if required_qty > 0 else 0, 1),
            "scheduled_date": line.get("scheduled_date"),
            "actual_completion_date": line.get("actual_completion_date"),
            "production_schedules": line_schedules,
            "status": line_status,
            "is_delayed": is_delayed
        })
    
    # Calculate lot-level stats
    ready_lines = sum(1 for l in lines_enriched if l["status"] == "READY")
    delayed_lines = sum(1 for l in lines_enriched if l.get("is_delayed"))
    
    return {
        **lot,
        "buyer": buyer,
        "lines": lines_enriched,
        "estimated_completion_date": latest_completion,
        "can_complete_with_current_inventory": can_complete_with_current_inventory,
        "ready_lines": ready_lines,
        "delayed_lines": delayed_lines,
        "total_lines": len(lines_enriched),
        "lot_readiness_pct": round((ready_lines / len(lines_enriched) * 100) if lines_enriched else 0, 1)
    }


# =============================================================================
# NOTIFICATIONS SYSTEM
# =============================================================================

@router.get("/notifications")
async def get_notifications(
    user_role: Optional[str] = None,
    unread_only: bool = True,
    limit: int = 50
):
    """Get notifications for dashboard"""
    query = {}
    if user_role:
        query["target_roles"] = {"$in": [user_role, "all"]}
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return notifications


@router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str):
    """Mark a notification as read"""
    result = await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification marked as read"}


@router.put("/notifications/mark-all-read")
async def mark_all_notifications_read(user_role: Optional[str] = None):
    """Mark all notifications as read"""
    query = {}
    if user_role:
        query["target_roles"] = {"$in": [user_role, "all"]}
    
    result = await db.notifications.update_many(
        query,
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc)}}
    )
    return {"message": f"Marked {result.modified_count} notifications as read"}


async def create_notification(
    notification_type: str,
    title: str,
    message: str,
    target_roles: List[str],
    reference_type: Optional[str] = None,
    reference_id: Optional[str] = None,
    priority: str = "NORMAL"
):
    """Create a new notification"""
    notification = {
        "id": str(uuid.uuid4()),
        "type": notification_type,
        "title": title,
        "message": message,
        "target_roles": target_roles,
        "reference_type": reference_type,
        "reference_id": reference_id,
        "priority": priority,
        "is_read": False,
        "created_at": datetime.now(timezone.utc)
    }
    await db.notifications.insert_one(notification)
    return notification


@router.post("/dispatch-lots/check-delays-and-completions")
async def check_delays_and_completions():
    """
    Check for:
    1. Lines delayed (scheduled date passed, not completed) -> Notify demand team
    2. Lots completing in 2 days -> Notify teams for QC/Dispatch prep
    """
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    in_two_days = today + timedelta(days=2)
    
    notifications_created = 0
    
    # 1. Check for delayed lines
    delayed_lines = await db.dispatch_lot_lines.find({
        "scheduled_date": {"$lt": today},
        "actual_completion_date": None,
        "status": {"$nin": ["READY", "COMPLETED"]}
    }, {"_id": 0}).to_list(1000)
    
    # Group by lot
    delayed_by_lot = {}
    for line in delayed_lines:
        lot_id = line.get("lot_id")
        if lot_id not in delayed_by_lot:
            delayed_by_lot[lot_id] = []
        delayed_by_lot[lot_id].append(line)
    
    for lot_id, lines in delayed_by_lot.items():
        lot = await db.dispatch_lots.find_one({"id": lot_id}, {"_id": 0})
        if not lot:
            continue
        
        # Check if notification already exists for this lot today
        existing = await db.notifications.find_one({
            "reference_type": "dispatch_lot_delay",
            "reference_id": lot_id,
            "created_at": {"$gte": today}
        })
        
        if not existing:
            await create_notification(
                notification_type="DELAY_ALERT",
                title=f"Dispatch Lot {lot.get('lot_code')} Running Behind Schedule",
                message=f"{len(lines)} line(s) delayed. Lot readiness timeline may be affected.",
                target_roles=["demand_planner", "master_admin"],
                reference_type="dispatch_lot_delay",
                reference_id=lot_id,
                priority="HIGH"
            )
            notifications_created += 1
    
    # 2. Check for lots completing in 2 days
    lots_completing_soon = await db.dispatch_lots.find({
        "estimated_completion_date": {"$gte": today, "$lte": in_two_days},
        "status": {"$nin": ["DISPATCHED", "DELIVERED", "CANCELLED"]}
    }, {"_id": 0}).to_list(500)
    
    for lot in lots_completing_soon:
        lot_id = lot.get("id")
        
        # Check if notification already exists
        existing = await db.notifications.find_one({
            "reference_type": "dispatch_lot_completion",
            "reference_id": lot_id,
            "created_at": {"$gte": today - timedelta(days=2)}
        })
        
        if not existing:
            completion_date = lot.get("estimated_completion_date")
            if isinstance(completion_date, datetime):
                date_str = completion_date.strftime("%d %b %Y")
            else:
                date_str = str(completion_date)
            
            await create_notification(
                notification_type="COMPLETION_ALERT",
                title=f"Dispatch Lot {lot.get('lot_code')} Completing Soon",
                message=f"Lot will be ready for QC and Dispatch on {date_str}. Prepare for processing.",
                target_roles=["demand_planner", "quality_inspector", "logistics_coordinator", "master_admin"],
                reference_type="dispatch_lot_completion",
                reference_id=lot_id,
                priority="NORMAL"
            )
            notifications_created += 1
    
    return {
        "message": f"Check complete. {notifications_created} notifications created.",
        "delayed_lots": len(delayed_by_lot),
        "lots_completing_soon": len(lots_completing_soon)
    }


# =============================================================================
# DASHBOARD SUMMARY
# =============================================================================

@router.get("/dispatch-lots/dashboard-summary")
async def get_dispatch_lots_dashboard_summary():
    """Get summary for dashboard: lots by status, delays, upcoming completions"""
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    in_seven_days = today + timedelta(days=7)
    
    # Get all active lots
    lots = await db.dispatch_lots.find(
        {"status": {"$nin": ["DISPATCHED", "DELIVERED", "CANCELLED"]}},
        {"_id": 0}
    ).to_list(5000)
    
    # Counts by status
    status_counts = {}
    for lot in lots:
        status = lot.get("status", "UNKNOWN")
        status_counts[status] = status_counts.get(status, 0) + 1
    
    # Delayed lots
    delayed_count = 0
    for lot in lots:
        comp_date = lot.get("estimated_completion_date")
        if comp_date:
            if isinstance(comp_date, str):
                try:
                    comp_date = datetime.fromisoformat(comp_date.replace('Z', '+00:00'))
                except:
                    continue
            if isinstance(comp_date, datetime) and comp_date < today:
                delayed_count += 1
    
    # Upcoming completions (next 7 days)
    upcoming = []
    for lot in lots:
        comp_date = lot.get("estimated_completion_date")
        if comp_date:
            if isinstance(comp_date, str):
                try:
                    comp_date = datetime.fromisoformat(comp_date.replace('Z', '+00:00'))
                except:
                    continue
            if isinstance(comp_date, datetime) and today <= comp_date <= in_seven_days:
                upcoming.append({
                    "lot_code": lot.get("lot_code"),
                    "lot_id": lot.get("id"),
                    "buyer_name": lot.get("buyer_name"),
                    "estimated_completion_date": comp_date,
                    "total_quantity": lot.get("total_quantity", 0)
                })
    
    # Sort upcoming by date
    upcoming.sort(key=lambda x: x.get("estimated_completion_date") or datetime.max.replace(tzinfo=timezone.utc))
    
    # Unread notifications count
    unread_notifications = await db.notifications.count_documents({"is_read": False})
    
    return {
        "total_active_lots": len(lots),
        "status_counts": status_counts,
        "delayed_lots": delayed_count,
        "upcoming_completions": upcoming[:10],
        "unread_notifications": unread_notifications
    }
