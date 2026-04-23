"""Vendor routes - Vendor management, pricing, purchase entries"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from pydantic import BaseModel
import uuid
import io
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import db
from models import User, Vendor, VendorCreate, VendorRMPrice, VendorRMPriceCreate, PurchaseEntry, PurchaseEntryCreate
from services.utils import get_current_user, get_next_vendor_id, serialize_doc, update_branch_rm_inventory, generate_movement_code, get_branch_rm_stock
from services.rbac_service import require_permission

router = APIRouter(tags=["Vendors"])


# ============ Zoho Integration ============

@router.get("/zoho/accounts")
async def get_zoho_accounts(account_type: Optional[str] = None):
    """
    Fetch Chart of Accounts from Zoho Books.
    Used for selecting the expense/asset account during bill creation.
    
    Args:
        account_type: Optional filter - 'expense', 'asset', 'liability', 'equity', 'income'
    """
    from services.zoho_service import zoho_client
    
    if not zoho_client.is_configured():
        return {"accounts": [], "configured": False, "message": "Zoho Books not configured"}
    
    try:
        accounts = await zoho_client.get_chart_of_accounts(account_type)
        if not accounts:
            return {
                "accounts": [],
                "configured": True,
                "message": "No accounts returned. The OAuth token may be missing 'ZohoBooks.accountants.READ' scope. Please re-authorize with this scope."
            }
        return {"accounts": accounts, "configured": True, "message": "OK"}
    except Exception as e:
        error_msg = str(e)
        if "401" in error_msg or "not authorized" in error_msg.lower():
            return {
                "accounts": [],
                "configured": True,
                "auth_error": True,
                "message": "Authorization failed for Chart of Accounts. The OAuth token needs 'ZohoBooks.accountants.READ' scope. Please generate a new Grant Token with this scope."
            }
        raise HTTPException(status_code=502, detail=f"Failed to fetch Zoho accounts: {error_msg}")


@router.get("/zoho/taxes")
async def get_zoho_taxes():
    """Fetch all taxes configured in Zoho Books."""
    from services.zoho_service import zoho_client
    
    if not zoho_client.is_configured():
        return {"taxes": [], "configured": False}
    
    try:
        taxes = await zoho_client.get_taxes()
        return {"taxes": taxes, "configured": True}
    except Exception as e:
        error_msg = str(e)
        if "401" in error_msg or "not authorized" in error_msg.lower():
            return {"taxes": [], "configured": True, "auth_error": True, "message": "Missing scope for taxes"}
        raise HTTPException(status_code=502, detail=f"Failed to fetch Zoho taxes: {error_msg}")



@router.get("/zoho/status")
async def get_zoho_status():
    """Check if Zoho Books integration is configured and working."""
    from services.zoho_service import zoho_client
    
    if not zoho_client.is_configured():
        return {"configured": False, "message": "Zoho credentials not set"}
    
    try:
        # Try to get access token to verify credentials work
        await zoho_client.token_manager.get_access_token()
        return {"configured": True, "message": "Zoho Books connected"}
    except Exception as e:
        return {"configured": False, "message": f"Zoho connection failed: {str(e)}"}


# ============ Vendor Management ============

@router.get("/vendors")
async def get_vendors():
    """Get all vendors"""
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(1000)
    return vendors


@router.get("/vendors/{vendor_id}")
async def get_vendor_detail(vendor_id: str):
    """Get a single vendor + their tagged RM prices (enriched with RM description)."""
    # Accept either vendor_id code (VND_001) or internal id
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        vendor = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail=f"Vendor {vendor_id} not found")

    prices = await db.vendor_rm_prices.find(
        {"vendor_id": vendor["vendor_id"]},
        {"_id": 0}
    ).sort("updated_at", -1).to_list(5000)

    # Enrich with RM description/category
    rm_ids = [p["rm_id"] for p in prices if p.get("rm_id")]
    rm_info = {}
    if rm_ids:
        async for rm in db.raw_materials.find(
            {"rm_id": {"$in": rm_ids}},
            {"_id": 0, "rm_id": 1, "description": 1, "category": 1, "category_data": 1}
        ):
            rm_info[rm["rm_id"]] = rm
    for p in prices:
        rm = rm_info.get(p.get("rm_id"), {})
        p["rm_description"] = rm.get("description") or rm.get("category_data", {}).get("name")
        p["rm_category"] = rm.get("category")

    return {"vendor": vendor, "rm_prices": prices}


@router.post("/vendors", response_model=Vendor)
@require_permission("Vendor", "CREATE")
async def create_vendor(input: VendorCreate, current_user: User = Depends(get_current_user)):
    """Create a new vendor (MASTER_ADMIN, PROCUREMENT_OFFICER)"""
    vendor_id = await get_next_vendor_id()
    
    vendor = Vendor(
        vendor_id=vendor_id,
        name=input.name,
        poc=input.poc,
        phone=input.phone,
        email=input.email,
        address=input.address,
        gst=input.gst
    )
    
    doc = vendor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['created_by'] = current_user.id
    await db.vendors.insert_one(doc)
    
    return vendor


@router.put("/vendors/{vendor_id}", response_model=Vendor)
@require_permission("Vendor", "UPDATE")
async def update_vendor(vendor_id: str, input: VendorCreate, current_user: User = Depends(get_current_user)):
    """Update a vendor (MASTER_ADMIN, PROCUREMENT_OFFICER)"""
    existing = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    update_data = input.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user.id
    
    await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": update_data})
    
    updated = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    return Vendor(**updated)


@router.delete("/vendors/{vendor_id}")
@require_permission("Vendor", "DELETE")
async def delete_vendor(vendor_id: str, current_user: User = Depends(get_current_user)):
    """Delete a vendor (MASTER_ADMIN only)"""
    result = await db.vendors.delete_one({"vendor_id": vendor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": f"Vendor {vendor_id} deleted"}


# ============ Vendor RM Prices ============

@router.get("/vendor-rm-prices")
async def get_vendor_rm_prices(
    vendor_id: Optional[str] = None,
    rm_id: Optional[str] = None
):
    """Get vendor RM prices"""
    query = {}
    if vendor_id:
        query["vendor_id"] = vendor_id
    if rm_id:
        query["rm_id"] = rm_id
    
    prices = await db.vendor_rm_prices.find(query, {"_id": 0}).to_list(10000)
    
    # Enrich with vendor and RM names
    for price in prices:
        vendor = await db.vendors.find_one({"vendor_id": price["vendor_id"]}, {"_id": 0, "name": 1})
        price["vendor_name"] = vendor.get("name") if vendor else None
        
        rm = await db.raw_materials.find_one({"rm_id": price["rm_id"]}, {"_id": 0, "category_data": 1})
        price["rm_description"] = rm.get("category_data", {}).get("part_name") or rm.get("category_data", {}).get("type") if rm else None
    
    return prices


@router.post("/vendor-rm-prices", response_model=VendorRMPrice)
@require_permission("VendorRMPrice", "CREATE")
async def create_vendor_rm_price(input: VendorRMPriceCreate, current_user: User = Depends(get_current_user)):
    """Create a vendor RM price entry (MASTER_ADMIN, PROCUREMENT_OFFICER)"""
    # Verify vendor exists
    vendor = await db.vendors.find_one({"vendor_id": input.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail=f"Vendor {input.vendor_id} not found")
    
    # Verify RM exists
    rm = await db.raw_materials.find_one({"rm_id": input.rm_id}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail=f"RM {input.rm_id} not found")
    
    # Deactivate existing prices for this vendor-RM combo
    await db.vendor_rm_prices.update_many(
        {"vendor_id": input.vendor_id, "rm_id": input.rm_id},
        {"$set": {"is_active": False}}
    )
    
    price = VendorRMPrice(
        vendor_id=input.vendor_id,
        rm_id=input.rm_id,
        price=input.price,
        min_order_qty=input.min_order_qty if hasattr(input, 'min_order_qty') else 1,
        lead_time_days=input.lead_time_days if hasattr(input, 'lead_time_days') else 7,
        is_active=True
    )
    
    doc = price.model_dump()
    doc['effective_date'] = doc['effective_date'].isoformat()
    doc['created_by'] = current_user.id
    await db.vendor_rm_prices.insert_one(doc)
    
    # Record price history
    await db.price_history.insert_one({
        "id": str(uuid.uuid4()),
        "rm_id": input.rm_id,
        "vendor_id": input.vendor_id,
        "price": input.price,
        "effective_date": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return price


@router.delete("/vendor-rm-prices/{price_id}")
async def delete_vendor_rm_price(price_id: str):
    """Delete a vendor RM price"""
    result = await db.vendor_rm_prices.delete_one({"id": price_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Price not found")
    return {"message": "Price deleted"}


# ============ Bulk Upload: Vendor Master ============

@router.post("/vendors/bulk-upload")
async def bulk_upload_vendors(file: UploadFile = File(...)):
    """Bulk upload vendors from Excel. Columns: Name, GST, Address, POC, Email, Phone.
    Skips vendors that already exist by name (case-insensitive)."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    contents = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel: {e}")
    sheet = workbook.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
    header_map = {
        'name': 'name', 'vendor name': 'name',
        'gst': 'gst', 'gstin': 'gst', 'gst number': 'gst',
        'address': 'address',
        'poc': 'poc', 'point of contact': 'poc', 'contact person': 'poc',
        'email': 'email', 'email address': 'email',
        'phone': 'phone', 'phone number': 'phone', 'mobile': 'phone',
    }
    field_indices = {}
    for i, h in enumerate(headers):
        if h in header_map:
            field_indices[header_map[h]] = i

    if 'name' not in field_indices:
        raise HTTPException(status_code=400, detail="'Name' column is required")

    created = 0
    skipped = 0
    errors: List[str] = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = row[field_indices['name']] if field_indices.get('name') is not None else None
            if not name:
                continue
            name = str(name).strip()

            existing = await db.vendors.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}})
            if existing:
                skipped += 1
                continue

            vendor_data = {'name': name}
            for field in ['gst', 'address', 'poc', 'email', 'phone']:
                i = field_indices.get(field)
                if i is not None and row[i]:
                    vendor_data[field] = str(row[i]).strip()

            vendor_id = await get_next_vendor_id()
            vendor = Vendor(**vendor_data, vendor_id=vendor_id)
            doc = vendor.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.vendors.insert_one(doc)
            created += 1
        except Exception as e:
            errors.append(f"Row {idx}: {e}")

    return {"created": created, "skipped": skipped, "errors": errors[:20], "total_errors": len(errors)}


# ============ Bulk Upload: Vendor RM Prices ============

@router.get("/vendor-rm-prices/template")
async def download_vendor_rm_price_template():
    """Excel template with Vendors and RM IDs reference tabs for bulk price upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendor RM Prices"

    headers = ["Vendor ID", "RM ID", "Price", "Currency", "Min Order Qty", "Lead Time Days", "Notes"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    samples = [
        ["VND_001", "INP_1001", 125.50, "INR", 100, 7, "Preferred vendor"],
        ["VND_002", "INP_1001", 128.00, "INR", 50, 10, ""],
        ["VND_001", "ACC_001", 42.75, "INR", 500, 5, ""],
    ]
    for r, row in enumerate(samples, 2):
        for cidx, val in enumerate(row, 1):
            ws.cell(row=r, column=cidx, value=val)
    for col_letter, width in zip("ABCDEFG", [14, 14, 10, 10, 14, 14, 30]):
        ws.column_dimensions[col_letter].width = width

    ws_info = wb.create_sheet("Instructions")
    lines = [
        "Vendor RM Price Bulk Upload",
        "",
        "Required columns (first row is header):",
        "  - Vendor ID: Existing Vendor ID (see 'Vendors' tab)",
        "  - RM ID: Existing Raw Material ID (see 'RM IDs' tab)",
        "  - Price: Numeric (> 0), price per unit",
        "",
        "Optional columns:",
        "  - Currency: INR (default) / USD",
        "  - Min Order Qty: integer (default 1)",
        "  - Lead Time Days: integer (default 7)",
        "  - Notes: free text",
        "",
        "Modes (upload parameter):",
        "  - upsert: update existing vendor+RM pair, or insert new (DEFAULT)",
        "  - replace-vendor: wipe ALL existing prices for each Vendor ID in the file first, then insert",
        "",
        "Rows with unknown Vendor ID or RM ID are rejected with an error.",
    ]
    for i, line in enumerate(lines, 1):
        ws_info.cell(row=i, column=1, value=line)
    ws_info.column_dimensions["A"].width = 90

    # Vendors reference tab
    ws_v = wb.create_sheet("Vendors")
    ws_v.cell(row=1, column=1, value="Vendor ID").font = Font(bold=True)
    ws_v.cell(row=1, column=2, value="Vendor Name").font = Font(bold=True)
    ws_v.cell(row=1, column=1).fill = header_fill
    ws_v.cell(row=1, column=2).fill = header_fill
    ws_v.cell(row=1, column=1).font = header_font
    ws_v.cell(row=1, column=2).font = header_font
    vendors = await db.vendors.find({}, {"_id": 0, "vendor_id": 1, "name": 1}).sort("vendor_id", 1).to_list(5000)
    for idx, v in enumerate(vendors, 2):
        ws_v.cell(row=idx, column=1, value=v.get("vendor_id", ""))
        ws_v.cell(row=idx, column=2, value=v.get("name", ""))
    ws_v.column_dimensions["A"].width = 16
    ws_v.column_dimensions["B"].width = 50

    # RM IDs reference tab
    ws_rm = wb.create_sheet("RM IDs")
    for i, h in enumerate(["RM ID", "Category", "Description"], 1):
        c = ws_rm.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
    rms = await db.raw_materials.find(
        {"status": {"$ne": "INACTIVE"}},
        {"_id": 0, "rm_id": 1, "category": 1, "description": 1}
    ).sort("rm_id", 1).to_list(20000)
    for idx, rm in enumerate(rms, 2):
        ws_rm.cell(row=idx, column=1, value=rm.get("rm_id", ""))
        ws_rm.cell(row=idx, column=2, value=rm.get("category", ""))
        ws_rm.cell(row=idx, column=3, value=rm.get("description", ""))
    ws_rm.column_dimensions["A"].width = 16
    ws_rm.column_dimensions["B"].width = 12
    ws_rm.column_dimensions["C"].width = 50

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vendor_rm_prices_template.xlsx"},
    )


@router.post("/vendor-rm-prices/bulk-upload")
async def bulk_upload_vendor_rm_prices(
    file: UploadFile = File(...),
    mode: str = Query("upsert", description="'upsert' or 'replace-vendor'"),
):
    """Bulk upload Vendor-RM-Price mappings.

    Columns (header row 1): Vendor ID | RM ID | Price | Currency | Min Order Qty | Lead Time Days | Notes
    Unknown Vendor IDs and RM IDs are rejected per row.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel: {e}")
    sheet = wb.active

    headers_raw = [str(c.value or "").strip().lower() for c in sheet[1]]
    col = {}
    for idx, h in enumerate(headers_raw):
        if h in ("vendor id", "vendor_id"):
            col["vendor_id"] = idx
        elif h in ("rm id", "rm_id"):
            col["rm_id"] = idx
        elif h == "price":
            col["price"] = idx
        elif h == "currency":
            col["currency"] = idx
        elif h in ("min order qty", "min_order_qty"):
            col["min_order_qty"] = idx
        elif h in ("lead time days", "lead_time_days"):
            col["lead_time_days"] = idx
        elif h == "notes":
            col["notes"] = idx

    required = ["vendor_id", "rm_id", "price"]
    missing = [k for k in required if k not in col]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {missing}. Found: {headers_raw}")

    # Pre-load valid vendor and rm IDs for validation speed
    vendor_ids = set(await db.vendors.distinct("vendor_id"))
    rm_ids_known = set(await db.raw_materials.distinct("rm_id"))

    # Collect rows
    valid_rows = []
    errors: List[str] = []
    vendors_in_file: set = set()

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            vendor_id = str(row[col["vendor_id"]] or "").strip()
            rm_id = str(row[col["rm_id"]] or "").strip()
            price_raw = row[col["price"]]

            if not vendor_id or not rm_id:
                errors.append(f"Row {idx}: Missing Vendor ID or RM ID")
                continue
            if vendor_id not in vendor_ids:
                errors.append(f"Row {idx}: Vendor '{vendor_id}' not found")
                continue
            if rm_id not in rm_ids_known:
                errors.append(f"Row {idx}: RM '{rm_id}' not found")
                continue
            try:
                price = float(price_raw)
            except (TypeError, ValueError):
                errors.append(f"Row {idx}: Invalid price '{price_raw}'")
                continue
            if price <= 0:
                errors.append(f"Row {idx}: Price must be > 0")
                continue

            currency = (str(row[col["currency"]]).strip() if "currency" in col and row[col["currency"]] else "INR")
            try:
                moq = int(row[col["min_order_qty"]]) if "min_order_qty" in col and row[col["min_order_qty"]] else 1
            except (TypeError, ValueError):
                moq = 1
            try:
                lead = int(row[col["lead_time_days"]]) if "lead_time_days" in col and row[col["lead_time_days"]] else 7
            except (TypeError, ValueError):
                lead = 7
            notes = str(row[col["notes"]]).strip() if "notes" in col and row[col["notes"]] else ""

            valid_rows.append({
                "vendor_id": vendor_id,
                "rm_id": rm_id,
                "price": round(price, 4),
                "currency": currency,
                "min_order_qty": moq,
                "lead_time_days": lead,
                "notes": notes,
            })
            vendors_in_file.add(vendor_id)
        except Exception as e:
            errors.append(f"Row {idx}: {e}")

    # Replace-vendor mode wipes all existing prices for vendors in the file
    deleted = 0
    if mode == "replace-vendor" and vendors_in_file:
        res = await db.vendor_rm_prices.delete_many({"vendor_id": {"$in": list(vendors_in_file)}})
        deleted = res.deleted_count

    # Upsert each row
    now_iso = datetime.now(timezone.utc).isoformat()
    created = 0
    updated = 0
    for r in valid_rows:
        existing = await db.vendor_rm_prices.find_one(
            {"vendor_id": r["vendor_id"], "rm_id": r["rm_id"]},
            {"_id": 0, "id": 1}
        )
        if existing and mode != "replace-vendor":
            await db.vendor_rm_prices.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "price": r["price"],
                    "currency": r["currency"],
                    "min_order_qty": r["min_order_qty"],
                    "lead_time_days": r["lead_time_days"],
                    "notes": r["notes"],
                    "updated_at": now_iso,
                }}
            )
            updated += 1
        else:
            doc = {
                "id": str(uuid.uuid4()),
                **r,
                "is_active": True,
                "updated_at": now_iso,
                "effective_date": now_iso,
            }
            await db.vendor_rm_prices.insert_one(doc)
            created += 1

    return {
        "message": f"Processed {len(valid_rows)} rows",
        "created": created,
        "updated": updated,
        "deleted_previous": deleted,
        "vendors_in_file": len(vendors_in_file),
        "errors": errors[:50],
        "error_count": len(errors),
        "mode": mode,
    }


# ============ Purchase Entries (RM Inward) ============

@router.get("/purchase-entries")
async def get_purchase_entries(
    branch: Optional[str] = None,
    rm_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get purchase entries"""
    query = {}
    if branch:
        query["branch"] = branch
    if rm_id:
        query["rm_id"] = rm_id
    
    entries = await db.purchase_entries.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    return entries


@router.post("/purchase-entries")
@require_permission("RMStockMovement", "CREATE")
async def create_purchase_entry(
    input: PurchaseEntryCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a purchase entry (RM Inward) - MASTER_ADMIN, PROCUREMENT_OFFICER, BRANCH_OPS_USER"""
    # Verify RM exists
    rm = await db.raw_materials.find_one({"rm_id": input.rm_id}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail=f"RM {input.rm_id} not found")
    
    entry = PurchaseEntry(
        branch=input.branch,
        rm_id=input.rm_id,
        quantity=input.quantity,
        vendor_id=input.vendor_id if hasattr(input, 'vendor_id') else None,
        unit_price=input.unit_price if hasattr(input, 'unit_price') else 0,
        invoice_number=input.invoice_number if hasattr(input, 'invoice_number') else None
    )
    
    doc = entry.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['purchase_date'] = doc['purchase_date'].isoformat() if doc.get('purchase_date') else None
    doc['created_by'] = current_user.id
    
    await db.purchase_entries.insert_one(doc)
    
    # Update branch inventory
    current_stock = await get_branch_rm_stock(input.branch, input.rm_id)
    await update_branch_rm_inventory(input.branch, input.rm_id, input.quantity)
    
    # Record stock movement
    movement_code = await generate_movement_code()
    await db.rm_stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "movement_code": movement_code,
        "rm_id": input.rm_id,
        "branch": input.branch,
        "movement_type": "PURCHASE",
        "quantity": input.quantity,
        "reference_type": "PURCHASE_ENTRY",
        "reference_id": entry.id,
        "balance_after": current_stock + input.quantity,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Purchase entry created", "entry": doc}


# ============ RM Inward Bills ============

class BillLineItem(BaseModel):
    rm_id: str
    quantity: float
    rate: float = 0
    tax: str = "NONE"
    tax_amount: float = 0
    amount: float = 0

class BillTotals(BaseModel):
    sub_total: float = 0
    discount_type: str = "percentage"
    discount_value: float = 0
    discount_amount: float = 0
    tds_tcs: str = "NONE"
    tds_tcs_amount: float = 0
    tax_total: float = 0
    grand_total: float = 0

class RMInwardBillCreate(BaseModel):
    vendor_id: str
    vendor_name: str
    branch: str
    branch_id: Optional[str] = None
    bill_number: str
    order_number: Optional[str] = None
    bill_date: Optional[str] = None
    due_date: Optional[str] = None
    payment_terms: str = "NET_30"
    accounts_payable: str = "Trade Payables"
    reverse_charge: bool = False
    notes: Optional[str] = None
    line_items: List[dict]
    totals: dict
    date: Optional[str] = None


@router.post("/rm-inward/bills")
@require_permission("RMStockMovement", "CREATE")
async def create_rm_inward_bill(
    input: RMInwardBillCreate,
    current_user: User = Depends(get_current_user)
):
    """
    Create a full RM Inward bill with multiple line items.
    Also creates a corresponding bill in Zoho Books.
    If Zoho bill creation fails, the entire operation is rolled back.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Import Zoho service
    from services.zoho_service import zoho_client
    
    # Generate bill ID
    bill_count = await db.rm_inward_bills.count_documents({})
    bill_id = f"BILL_{datetime.now(timezone.utc).strftime('%Y%m')}_{bill_count + 1:05d}"
    
    # Validate all RMs exist
    for item in input.line_items:
        rm = await db.raw_materials.find_one({"rm_id": item["rm_id"]}, {"_id": 0, "rm_id": 1})
        if not rm:
            raise HTTPException(status_code=404, detail=f"RM {item['rm_id']} not found")
    
    # ========== ZOHO BOOKS INTEGRATION (OPTIONAL) ==========
    zoho_result = None
    zoho_error = None
    
    if zoho_client.is_configured():
        try:
            logger.info(f"Creating Zoho bill for {input.bill_number}...")
            
            # Look up vendor GST from our DB
            vendor_doc = await db.vendors.find_one({"vendor_id": input.vendor_id}, {"_id": 0, "gst": 1})
            vendor_gst = vendor_doc.get("gst") if vendor_doc else None
            
            # Get or create vendor in Zoho
            zoho_vendor_id = await zoho_client.get_or_create_vendor(input.vendor_name, gst=vendor_gst)
            
            # Auto-apply reverse charge for unregistered vendors (Zoho requires it)
            apply_reverse_charge = input.reverse_charge or not vendor_gst
            
            # Create bill in Zoho Books
            zoho_result = await zoho_client.create_bill(
                vendor_id=zoho_vendor_id,
                vendor_name=input.vendor_name,
                bill_number=input.bill_number,
                bill_date=input.bill_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                line_items=input.line_items,
                reference_number=input.order_number,
                notes=input.notes,
                due_date=input.due_date,
                is_reverse_charge=apply_reverse_charge
            )
            
            logger.info(f"Zoho bill created: {zoho_result.get('zoho_bill_id')}")
            
        except Exception as e:
            zoho_error = str(e)
            logger.error(f"Zoho bill creation failed: {zoho_error}")
            # Fail the entire operation if Zoho fails
            raise HTTPException(
                status_code=502, 
                detail=f"Failed to create bill in Zoho Books: {zoho_error}. RM Inward not recorded."
            )
    else:
        logger.warning("Zoho Books integration not configured - bill will be created locally only")
    
    # ========== CREATE LOCAL BILL ==========
    # Create bill document
    bill = {
        "id": str(uuid.uuid4()),
        "bill_id": bill_id,
        "vendor_id": input.vendor_id,
        "vendor_name": input.vendor_name,
        "branch": input.branch,
        "branch_id": input.branch_id,
        "bill_number": input.bill_number,
        "order_number": input.order_number,
        "bill_date": input.bill_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "due_date": input.due_date,
        "payment_terms": input.payment_terms,
        "accounts_payable": input.accounts_payable,
        "reverse_charge": input.reverse_charge,
        "notes": input.notes,
        "line_items": input.line_items,
        "totals": input.totals,
        "status": "POSTED",
        "created_by": current_user.id,
        "created_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        # Zoho integration fields
        "zoho_bill_id": zoho_result.get("zoho_bill_id") if zoho_result else None,
        "zoho_bill_number": zoho_result.get("zoho_bill_number") if zoho_result else None,
        "zoho_synced": zoho_result is not None
    }
    
    await db.rm_inward_bills.insert_one(bill)
    
    # Create individual purchase entries and update inventory for each line item
    entries_created = []
    for item in input.line_items:
        entry_id = str(uuid.uuid4())
        entry = {
            "id": entry_id,
            "bill_id": bill_id,
            "bill_number": input.bill_number,
            "vendor_id": input.vendor_id,
            "vendor_name": input.vendor_name,
            "branch": input.branch,
            "rm_id": item["rm_id"],
            "description": item.get("description", ""),
            "hsn": item.get("hsn", ""),
            "quantity": item["quantity"],
            "rate": item.get("rate", 0),
            "tax": item.get("tax", "NONE"),
            "tax_amount": item.get("tax_amount", 0),
            "amount": item.get("amount", 0),
            "date": input.date or datetime.now(timezone.utc).isoformat(),
            "notes": input.notes,
            "created_by": current_user.id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.purchase_entries.insert_one(entry)
        entries_created.append(entry)
        
        # Update branch inventory
        current_stock = await get_branch_rm_stock(input.branch, item["rm_id"])
        await update_branch_rm_inventory(input.branch, item["rm_id"], item["quantity"])
        
        # Record stock movement
        movement_code = await generate_movement_code()
        await db.rm_stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "movement_code": movement_code,
            "rm_id": item["rm_id"],
            "branch": input.branch,
            "movement_type": "PURCHASE",
            "quantity": item["quantity"],
            "reference_type": "RM_INWARD_BILL",
            "reference_id": bill_id,
            "balance_after": current_stock + item["quantity"],
            "notes": f"Bill: {input.bill_number}, Vendor: {input.vendor_name}",
            "created_by": current_user.id,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    response = {
        "message": f"Bill {input.bill_number} recorded successfully",
        "bill_id": bill_id,
        "entries_count": len(entries_created),
        "grand_total": input.totals.get("grand_total", 0)
    }
    
    # Add Zoho info to response
    if zoho_result:
        response["zoho_synced"] = True
        response["zoho_bill_id"] = zoho_result.get("zoho_bill_id")
        response["zoho_message"] = "Bill also created in Zoho Books"
    
    # Emit event
    try:
        from services.event_system import event_bus, EventType
        await event_bus.publish(
            EventType.RM_INWARD_RECEIVED,
            {
                "bill_id": bill_id,
                "bill_number": input.bill_number,
                "vendor_id": input.vendor_id,
                "vendor_name": input.vendor_name,
                "branch": input.branch,
                "branch_id": input.branch_id,
                "line_item_count": len(input.line_items),
                "rm_ids": [it.get("rm_id") for it in input.line_items],
                "total_quantity": sum(float(it.get("quantity", 0)) for it in input.line_items),
                "grand_total": input.totals.get("grand_total", 0),
                "zoho_synced": zoho_result is not None,
            },
            source_module="rm_inward",
        )
    except Exception as e:
        logger.warning(f"Event publish failed (RM_INWARD_RECEIVED): {e}")

    return response


@router.get("/rm-inward/bills")
async def get_rm_inward_bills(
    branch: Optional[str] = None,
    vendor_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get RM Inward bills"""
    query = {}
    if branch:
        query["branch"] = branch
    if vendor_id:
        query["vendor_id"] = vendor_id
    
    bills = await db.rm_inward_bills.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"bills": bills, "total": len(bills)}
