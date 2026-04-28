"""SKU routes - SKU CRUD, mappings, subscriptions"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
import uuid
import io

from database import db
from models import User, SKU, SKUCreate, SKUMapping, SKUMappingCreate, ActivateItemRequest
from models.core import SKUBranchAssignment, BranchSKUInventory, BranchRMInventory
from services.utils import get_current_user, serialize_doc
from services.rbac_service import require_permission, check_user_permission
from services.sku_service import (
    get_sku_by_buyer_id, 
    get_skus_by_buyer_ids, 
    get_all_skus,
    search_skus,
    sku_exists
)

router = APIRouter(tags=["SKUs"])


@router.post("/skus", response_model=SKU)
@require_permission("BuyerSKU", "CREATE")
async def create_sku(input: SKUCreate, current_user: User = Depends(get_current_user)):
    """Create a new SKU (MASTER_ADMIN, DEMAND_PLANNER)"""
    # Check if SKU already exists in buyer_skus
    existing = await db.buyer_skus.find_one({"buyer_sku_id": input.sku_id}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail=f"SKU {input.sku_id} already exists")
    
    sku = SKU(**input.model_dump())
    doc = sku.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['created_by'] = current_user.id
    
    # Insert into buyer_skus collection (new model)
    buyer_sku_doc = {
        "id": doc.get("id") or str(uuid.uuid4()),
        "buyer_sku_id": input.sku_id,
        "bidso_sku_id": input.bidso_sku or "",
        "brand_id": input.brand_id or "",
        "name": input.description or "",
        "description": input.description or "",
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user.id
    }
    await db.buyer_skus.insert_one(buyer_sku_doc)
    
    return sku


@router.get("/skus")
async def get_skus(
    branch: Optional[str] = None,
    search: Optional[str] = None,
    include_inactive: bool = False
):
    """Get all SKUs with optional filters. When branch is provided, returns only SKUs active at that branch."""
    # Use SKU service to get from new model
    if search:
        skus = await search_skus(search, limit=10000)
    else:
        skus = await get_all_skus(include_inactive=include_inactive)
    
    # Filter and enrich by branch
    if branch:
        # Get SKUs active at this branch
        branch_inventory = await db.branch_sku_inventory.find(
            {"branch": branch, "is_active": True},
            {"_id": 0, "buyer_sku_id": 1, "current_stock": 1}
        ).to_list(10000)
        
        active_sku_ids = {inv["buyer_sku_id"] for inv in branch_inventory if inv.get("buyer_sku_id")}
        branch_stock_map = {inv["buyer_sku_id"]: inv.get("current_stock", 0) for inv in branch_inventory if inv.get("buyer_sku_id")}
        
        # Filter SKUs to only those active at this branch
        skus = [sku for sku in skus if sku.get("sku_id") in active_sku_ids]
        
        # Enrich with branch-specific data
        for sku in skus:
            sku["branch_stock"] = branch_stock_map.get(sku["sku_id"], 0)
            sku["is_active_in_branch"] = True  # Already filtered to active only
            
            # FG stock is same as branch_stock from branch_sku_inventory
            sku["fg_stock"] = branch_stock_map.get(sku["sku_id"], 0)
    
    return skus


@router.get("/skus/filter-options")
async def get_sku_filter_options():
    """Get all distinct verticals, models, and brands for filter dropdowns"""
    # Get from reference collections directly (more accurate)
    verticals_docs = await db.verticals.find({}, {"_id": 0, "name": 1}).to_list(100)
    models_docs = await db.models.find({}, {"_id": 0, "name": 1}).to_list(500)
    brands_docs = await db.brands.find({}, {"_id": 0, "name": 1}).to_list(200)
    
    verticals = sorted([v["name"] for v in verticals_docs if v.get("name")])
    models = sorted([m["name"] for m in models_docs if m.get("name")])
    brands = sorted([b["name"] for b in brands_docs if b.get("name")])
    
    return {
        "verticals": verticals,
        "models": models,
        "brands": brands
    }


@router.get("/skus/models-by-vertical")
async def get_models_by_vertical(vertical: str):
    """Get distinct models for a specific vertical"""
    # Find vertical by name
    vertical_doc = await db.verticals.find_one({"name": vertical}, {"_id": 0, "id": 1})
    if not vertical_doc:
        return {"models": []}
    
    # Get models for this vertical
    models_docs = await db.models.find(
        {"vertical_id": vertical_doc["id"]},
        {"_id": 0, "name": 1}
    ).to_list(500)
    
    models = sorted([m["name"] for m in models_docs if m.get("name")])
    return {"models": models}


@router.get("/skus/brands-by-vertical-model")
async def get_brands_by_vertical_model(vertical: str, model: Optional[str] = None):
    """Get distinct brands for a specific vertical and optionally model"""
    # Build query to find bidso_skus
    vertical_doc = await db.verticals.find_one({"name": vertical}, {"_id": 0, "id": 1})
    if not vertical_doc:
        return {"brands": []}
    
    bidso_query = {"vertical_id": vertical_doc["id"]}
    
    if model:
        model_doc = await db.models.find_one({"name": model}, {"_id": 0, "id": 1})
        if model_doc:
            bidso_query["model_id"] = model_doc["id"]
    
    # Get bidso_sku_ids matching the criteria
    bidso_skus = await db.bidso_skus.find(bidso_query, {"_id": 0, "bidso_sku_id": 1}).to_list(5000)
    bidso_ids = [b["bidso_sku_id"] for b in bidso_skus]
    
    if not bidso_ids:
        return {"brands": []}
    
    # Get brand_ids from buyer_skus linked to these bidso_skus
    buyer_skus = await db.buyer_skus.find(
        {"bidso_sku_id": {"$in": bidso_ids}},
        {"_id": 0, "brand_id": 1}
    ).to_list(10000)
    
    brand_ids = list(set(b["brand_id"] for b in buyer_skus if b.get("brand_id")))
    
    # Get brand names
    brands_docs = await db.brands.find(
        {"id": {"$in": brand_ids}},
        {"_id": 0, "name": 1}
    ).to_list(200)
    
    brands = sorted([b["name"] for b in brands_docs if b.get("name")])
    return {"brands": brands}


@router.get("/skus/unmapped")
async def get_skus_without_rm_mapping():
    """Get SKUs that don't have RM mappings"""
    all_buyer_skus = await db.buyer_skus.find(
        {"status": {"$ne": "INACTIVE"}},
        {"buyer_sku_id": 1, "_id": 0}
    ).to_list(10000)
    
    # Get all mapped SKUs
    mappings = await db.sku_mappings.find({}, {"sku_id": 1, "_id": 0}).to_list(10000)
    bom_mappings = await db.bill_of_materials.find({}, {"sku_id": 1, "_id": 0}).to_list(10000)
    
    mapped_ids = set(m["sku_id"] for m in mappings) | set(m["sku_id"] for m in bom_mappings)
    
    unmapped = [s["buyer_sku_id"] for s in all_buyer_skus if s["buyer_sku_id"] not in mapped_ids]
    
    return {
        "total": len(unmapped),
        "unmapped_skus": unmapped
    }


@router.post("/skus/activate")
async def activate_sku_in_branch(request: ActivateItemRequest):
    """Activate an SKU in a specific branch - also activates all mapped RMs"""
    sku = await get_sku_by_buyer_id(request.item_id)
    if not sku:
        raise HTTPException(status_code=404, detail="SKU not found")
    
    # Activate SKU (atomic upsert - prevents duplicates)
    await db.branch_sku_inventory.update_one(
        {"buyer_sku_id": request.item_id, "branch": request.branch},
        {
            "$set": {"is_active": True},
            "$setOnInsert": {"id": str(uuid.uuid4()), "current_stock": 0, "created_at": datetime.now(timezone.utc).isoformat()}
        },
        upsert=True
    )
    
    # Get RM mappings and activate them too
    mappings = await db.sku_mappings.find_one({"sku_id": request.item_id}, {"_id": 0})
    bom = await db.bill_of_materials.find_one({"sku_id": request.item_id}, {"_id": 0})
    
    activated_rms = []
    rm_list = []
    
    if mappings and mappings.get("rm_mappings"):
        rm_list = [m["rm_id"] for m in mappings["rm_mappings"]]
    elif bom and bom.get("rm_mappings"):
        rm_list = [m["rm_id"] for m in bom["rm_mappings"]]
    
    for rm_id in rm_list:
        await db.branch_rm_inventory.update_one(
            {"rm_id": rm_id, "branch": request.branch},
            {
                "$set": {"is_active": True},
                "$setOnInsert": {"id": str(uuid.uuid4()), "current_stock": 0.0, "created_at": datetime.now(timezone.utc).isoformat()}
            },
            upsert=True
        )
        activated_rms.append(rm_id)
    
    return {
        "message": f"SKU {request.item_id} and {len(activated_rms)} RMs activated in {request.branch}",
        "activated_rms": activated_rms
    }


@router.put("/skus/{sku_id}", response_model=SKU)
@require_permission("BuyerSKU", "UPDATE")
async def update_sku(sku_id: str, input: SKUCreate, current_user: User = Depends(get_current_user)):
    """Update an existing SKU (MASTER_ADMIN, DEMAND_PLANNER, TECH_OPS_ENGINEER)"""
    existing = await db.buyer_skus.find_one({"buyer_sku_id": sku_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="SKU not found")
    
    update_data = {
        "name": input.description,
        "description": input.description,
        "bidso_sku_id": input.bidso_sku,
        "brand_id": input.brand_id,
        "updated_at": datetime.now(timezone.utc),
        "updated_by": current_user.id
    }
    
    await db.buyer_skus.update_one({"buyer_sku_id": sku_id}, {"$set": update_data})
    
    # Return in legacy format
    sku = await get_sku_by_buyer_id(sku_id)
    return SKU(**sku)


@router.delete("/skus/{sku_id}")
@require_permission("BuyerSKU", "DELETE")
async def delete_sku(sku_id: str, current_user: User = Depends(get_current_user)):
    """Delete an SKU (MASTER_ADMIN, DEMAND_PLANNER with constraints)"""
    result = await db.buyer_skus.delete_one({"buyer_sku_id": sku_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="SKU not found")
    return {"message": f"SKU {sku_id} deleted"}


# ============ SKU Mappings ============

@router.post("/sku-mappings", response_model=SKUMapping)
async def create_sku_mapping(input: SKUMappingCreate):
    """Create SKU to RM mapping"""
    # Verify SKU exists using new model
    sku = await get_sku_by_buyer_id(input.sku_id)
    if not sku:
        raise HTTPException(status_code=404, detail=f"SKU {input.sku_id} not found")
    
    # Create or update mapping
    existing = await db.sku_mappings.find_one({"sku_id": input.sku_id})
    
    if existing:
        await db.sku_mappings.update_one(
            {"sku_id": input.sku_id},
            {"$set": {"rm_mappings": [m.model_dump() for m in input.rm_mappings]}}
        )
    else:
        mapping = SKUMapping(
            sku_id=input.sku_id,
            rm_mappings=[m.model_dump() for m in input.rm_mappings]
        )
        await db.sku_mappings.insert_one(mapping.model_dump())
    
    result = await db.sku_mappings.find_one({"sku_id": input.sku_id}, {"_id": 0})
    return SKUMapping(**result)


@router.post("/sku-mappings/bulk-upload")
async def bulk_upload_sku_mappings(file: UploadFile = File(...)):
    """Bulk upload SKU mappings from Excel - NO OVERWRITES ALLOWED"""
    import openpyxl  # noqa: F401  (lazy import — keeps backend startup fast)
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    created = 0
    skipped_duplicates = []
    errors = []
    
    # Group by SKU
    sku_mappings = {}
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            row_data = dict(zip(headers, row))
            
            sku_id = str(row_data.get('sku_id', '')).strip()
            rm_id = str(row_data.get('rm_id', '')).strip()
            qty = float(row_data.get('quantity', 1))
            
            if not sku_id or not rm_id:
                errors.append(f"Row {idx}: Missing sku_id or rm_id")
                continue
            
            if sku_id not in sku_mappings:
                sku_mappings[sku_id] = {"rows": [], "mappings": []}
            
            sku_mappings[sku_id]["rows"].append(idx)
            sku_mappings[sku_id]["mappings"].append({
                "rm_id": rm_id,
                "quantity": qty
            })
            
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    # Check for existing and save only new mappings
    for sku_id, data in sku_mappings.items():
        existing = await db.sku_mappings.find_one({"sku_id": sku_id})
        
        if existing:
            # PREVENT OVERWRITE
            skipped_duplicates.append({
                "rows": data["rows"],
                "sku_id": sku_id,
                "reason": "SKU mapping already exists"
            })
        else:
            await db.sku_mappings.insert_one({
                "id": str(uuid.uuid4()),
                "sku_id": sku_id,
                "rm_mappings": data["mappings"]
            })
            created += 1
    
    if skipped_duplicates:
        return {
            "success": False if created == 0 else True,
            "created": created,
            "duplicates": skipped_duplicates,
            "errors": errors,
            "message": f"Upload complete: {created} created, {len(skipped_duplicates)} skipped (duplicates - no overwrites allowed)"
        }
    
    return {
        "success": True,
        "created": created,
        "duplicates": [],
        "errors": errors,
        "message": f"Successfully created {created} SKU mappings"
    }


@router.get("/sku-mappings/{sku_id}", response_model=SKUMapping)
async def get_sku_mapping(sku_id: str):
    """Get SKU mapping by SKU ID"""
    # Try sku_mappings first
    mapping = await db.sku_mappings.find_one({"sku_id": sku_id}, {"_id": 0})
    if mapping:
        return SKUMapping(**mapping)
    
    # Try bill_of_materials
    bom = await db.bill_of_materials.find_one({"sku_id": sku_id}, {"_id": 0})
    if bom:
        return SKUMapping(
            id=bom.get("id", str(uuid.uuid4())),
            sku_id=bom["sku_id"],
            rm_mappings=bom.get("rm_mappings", [])
        )
    
    raise HTTPException(status_code=404, detail=f"Mapping not found for SKU {sku_id}")


@router.get("/sku-mappings", response_model=List[SKUMapping])
async def get_all_sku_mappings():
    """Get all SKU mappings"""
    # Combine both sources
    mappings = await db.sku_mappings.find({}, {"_id": 0}).to_list(10000)
    boms = await db.bill_of_materials.find({}, {"_id": 0}).to_list(10000)
    
    # Dedupe by sku_id
    seen = set()
    result = []
    
    for m in mappings:
        if m["sku_id"] not in seen:
            seen.add(m["sku_id"])
            result.append(SKUMapping(**m))
    
    for b in boms:
        if b["sku_id"] not in seen:
            seen.add(b["sku_id"])
            result.append(SKUMapping(
                id=b.get("id", str(uuid.uuid4())),
                sku_id=b["sku_id"],
                rm_mappings=b.get("rm_mappings", [])
            ))
    
    return result


@router.get("/skus/filtered")
async def get_filtered_skus(
    vertical_id: Optional[str] = None,
    model_id: Optional[str] = None,
    brand_id: Optional[str] = None,
    buyer_id: Optional[str] = None,
    search: Optional[str] = None,
    branch: Optional[str] = None,
    include_inactive: bool = False,
    page: int = 1,
    page_size: int = 50
):
    """
    Get SKUs with relational filters and pagination.
    When branch is provided, returns only SKUs active at that branch.
    Now queries consolidated db.buyer_skus collection (single source of truth).
    Enriches with vertical/model from parent Bidso SKU.
    """
    query = {}
    if not include_inactive:
        query["status"] = {"$ne": "INACTIVE"}
    
    if brand_id:
        query["brand_id"] = brand_id
    if buyer_id:
        query["buyer_id"] = buyer_id
    
    # Apply search filter at DB level
    if search:
        query["$or"] = [
            {"buyer_sku_id": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    # If filtering by vertical or model, we need to find matching bidso_skus first
    bidso_filter = None
    if vertical_id or model_id:
        bidso_query = {}
        if vertical_id:
            bidso_query["vertical_id"] = vertical_id
        if model_id:
            bidso_query["model_id"] = model_id
        
        matching_bidso = await db.bidso_skus.find(bidso_query, {"_id": 0, "bidso_sku_id": 1}).to_list(5000)
        bidso_filter = [b["bidso_sku_id"] for b in matching_bidso]
        
        if not bidso_filter:
            return {"items": [], "total": 0, "page": page, "page_size": page_size, "total_pages": 1}
        
        query["bidso_sku_id"] = {"$in": bidso_filter}
    
    # If branch is specified, filter to only SKUs active at that branch
    branch_stock_map = {}
    if branch:
        branch_inventory = await db.branch_sku_inventory.find(
            {"branch": branch, "is_active": True},
            {"_id": 0, "buyer_sku_id": 1, "current_stock": 1}
        ).to_list(10000)
        
        active_sku_ids = [inv["buyer_sku_id"] for inv in branch_inventory if inv.get("buyer_sku_id")]
        branch_stock_map = {inv["buyer_sku_id"]: inv.get("current_stock", 0) for inv in branch_inventory if inv.get("buyer_sku_id")}
        
        if not active_sku_ids:
            return {"items": [], "total": 0, "page": page, "page_size": page_size, "total_pages": 1}
        
        query["buyer_sku_id"] = {"$in": active_sku_ids}
    
    # Get total count for pagination
    total = await db.buyer_skus.count_documents(query)
    
    # Calculate pagination
    skip = (page - 1) * page_size
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    # Query from consolidated buyer_skus collection with pagination
    skus = await db.buyer_skus.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    # Cache for bidso_sku lookups to avoid repeated queries
    bidso_cache = {}
    
    # Batch fetch all verticals and models for enrichment
    vertical_ids = set()
    model_ids = set()
    brand_ids = set()
    buyer_ids = set()
    
    # First pass: get bidso data and collect IDs
    for sku in skus:
        sku["sku_id"] = sku.get("buyer_sku_id", "")
        
        bidso_sku_id = sku.get("bidso_sku_id")
        if bidso_sku_id:
            if bidso_sku_id not in bidso_cache:
                bidso = await db.bidso_skus.find_one(
                    {"bidso_sku_id": bidso_sku_id},
                    {"_id": 0, "vertical_id": 1, "model_id": 1}
                )
                bidso_cache[bidso_sku_id] = bidso or {}
            
            bidso_data = bidso_cache[bidso_sku_id]
            sku["vertical_id"] = bidso_data.get("vertical_id")
            sku["model_id"] = bidso_data.get("model_id")
        
        if sku.get("vertical_id"):
            vertical_ids.add(sku["vertical_id"])
        if sku.get("model_id"):
            model_ids.add(sku["model_id"])
        if sku.get("brand_id"):
            brand_ids.add(sku["brand_id"])
        if sku.get("buyer_id"):
            buyer_ids.add(sku["buyer_id"])
    
    # Batch fetch reference data
    verticals_map = {}
    models_map = {}
    brands_map = {}
    buyers_map = {}
    
    if vertical_ids:
        v_docs = await db.verticals.find({"id": {"$in": list(vertical_ids)}}, {"_id": 0}).to_list(100)
        verticals_map = {v["id"]: v for v in v_docs}
    
    if model_ids:
        m_docs = await db.models.find({"id": {"$in": list(model_ids)}}, {"_id": 0}).to_list(200)
        models_map = {m["id"]: m for m in m_docs}
    
    if brand_ids:
        b_docs = await db.brands.find({"id": {"$in": list(brand_ids)}}, {"_id": 0}).to_list(100)
        brands_map = {b["id"]: b for b in b_docs}
    
    if buyer_ids:
        bu_docs = await db.buyers.find({"id": {"$in": list(buyer_ids)}}, {"_id": 0}).to_list(500)
        buyers_map = {bu["id"]: bu for bu in bu_docs}
    
    # Second pass: enrich with names
    for sku in skus:
        if sku.get("vertical_id") and sku["vertical_id"] in verticals_map:
            v = verticals_map[sku["vertical_id"]]
            sku["vertical_name"] = v.get("name")
            sku["vertical_code"] = v.get("code")
            sku["vertical"] = {"id": sku["vertical_id"], "name": v.get("name"), "code": v.get("code")}
        
        if sku.get("model_id") and sku["model_id"] in models_map:
            m = models_map[sku["model_id"]]
            sku["model_name"] = m.get("name")
            sku["model_code"] = m.get("code")
            sku["model"] = {"id": sku["model_id"], "name": m.get("name"), "code": m.get("code")}
        
        if sku.get("brand_id") and not sku.get("brand_name") and sku["brand_id"] in brands_map:
            sku["brand_name"] = brands_map[sku["brand_id"]].get("name")
        
        if sku.get("buyer_id") and sku["buyer_id"] in buyers_map:
            sku["buyer_name"] = buyers_map[sku["buyer_id"]].get("name")
        
        # Branch inventory data (use cached data if branch was specified)
        if branch:
            sku_id = sku.get("buyer_sku_id")
            sku["branch_stock"] = branch_stock_map.get(sku_id, 0)
            sku["is_active_in_branch"] = True  # Already filtered to active only
            
            # FG stock is same as branch_stock from branch_sku_inventory
            sku["fg_stock"] = branch_stock_map.get(sku_id, 0)
    
    return {
        "items": skus,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }



@router.get("/sku-branch-assignments/all")
async def get_all_sku_branch_assignments():
    """Get all SKU branch assignments for demand planning"""
    assignments = await db.branch_sku_inventory.find(
        {"is_active": True},
        {"_id": 0, "buyer_sku_id": 1, "branch": 1}
    ).to_list(50000)
    
    return assignments


# ============ SKU Branch Assignment Routes ============

async def activate_rms_for_sku(sku_id: str, branch: str) -> int:
    """
    Activate all RMs in the BOM for a given SKU in a branch.
    Returns the number of RMs activated.
    """
    activated_count = 0
    
    # Get RM mappings from sku_rm_mapping collection (bulk uploaded)
    rm_mappings = await db.sku_rm_mapping.find({"sku_id": sku_id}, {"_id": 0, "rm_id": 1}).to_list(1000)
    
    # Also check legacy sku_mappings collection
    legacy_mapping = await db.sku_mappings.find_one({"sku_id": sku_id}, {"_id": 0})
    if legacy_mapping and legacy_mapping.get('rm_mappings'):
        for rm in legacy_mapping['rm_mappings']:
            rm_mappings.append({"rm_id": rm['rm_id']})
    
    # Also check bill_of_materials collection
    bom_mappings = await db.bill_of_materials.find({"sku_id": sku_id}, {"_id": 0, "rm_id": 1}).to_list(1000)
    for bom in bom_mappings:
        if bom.get("rm_id"):
            rm_mappings.append({"rm_id": bom['rm_id']})
    
    # Activate each RM in the branch
    for mapping in rm_mappings:
        rm_id = mapping.get('rm_id')
        if not rm_id:
            continue
        
        # Check if RM exists in the system
        rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
        if not rm:
            continue
        
        # Atomic upsert - prevents duplicates
        result = await db.branch_rm_inventory.update_one(
            {"rm_id": rm_id, "branch": branch},
            {
                "$set": {"is_active": True},
                "$setOnInsert": {"id": str(uuid.uuid4()), "current_stock": 0.0, "activated_at": datetime.now(timezone.utc).isoformat()}
            },
            upsert=True
        )
        if result.upserted_id or result.modified_count > 0:
            activated_count += 1
    
    return activated_count


@router.post("/sku-branch-assignments/upload")
async def upload_sku_branch_assignments(file: UploadFile = File(...), branch: str = ""):
    """Upload SKU IDs to assign to a branch. Also activates corresponding RMs."""
    import openpyxl  # noqa: F401  (lazy import — keeps backend startup fast)
    if not branch:
        raise HTTPException(status_code=400, detail="Branch is required")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        assigned_count = 0
        skipped_count = 0
        not_found = []
        total_rms_activated = 0
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            
            sku_id = str(row[0]).strip()
            
            # Check if SKU exists using new model
            sku = await get_sku_by_buyer_id(sku_id)
            
            if not sku:
                not_found.append(sku_id)
                continue
            
            actual_sku_id = sku['sku_id']
            
            # Check if already assigned
            existing = await db.sku_branch_assignments.find_one(
                {"sku_id": actual_sku_id, "branch": branch},
                {"_id": 0}
            )
            
            if existing:
                skipped_count += 1
                continue
            
            # Create assignment
            assignment = SKUBranchAssignment(sku_id=actual_sku_id, branch=branch)
            doc = assignment.model_dump()
            doc['assigned_at'] = doc['assigned_at'].isoformat()
            await db.sku_branch_assignments.insert_one(doc)
            
            # Also activate SKU in branch inventory (atomic upsert)
            await db.branch_sku_inventory.update_one(
                {"buyer_sku_id": actual_sku_id, "branch": branch},
                {
                    "$set": {"is_active": True},
                    "$setOnInsert": {"id": str(uuid.uuid4()), "current_stock": 0, "activated_at": datetime.now(timezone.utc).isoformat()}
                },
                upsert=True
            )
            
            # Activate corresponding RMs for this SKU
            rms_activated = await activate_rms_for_sku(actual_sku_id, branch)
            total_rms_activated += rms_activated
            
            assigned_count += 1
        
        return {
            "assigned": assigned_count,
            "skipped": skipped_count,
            "not_found": not_found[:20],
            "total_not_found": len(not_found),
            "rms_activated": total_rms_activated
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.get("/sku-branch-assignments")
async def get_sku_branch_assignments(branch: Optional[str] = None):
    """Get SKU assignments, optionally filtered by branch"""
    query = {}
    if branch:
        query["branch"] = branch
    
    assignments = await db.sku_branch_assignments.find(query, {"_id": 0}).to_list(5000)
    
    # Enrich with SKU details using SKU service
    sku_ids = [a['sku_id'] for a in assignments]
    skus_list = await get_skus_by_buyer_ids(sku_ids)
    skus_map = {s['sku_id']: s for s in skus_list}
    
    result = []
    for a in assignments:
        sku = skus_map.get(a['sku_id'])
        if sku:
            result.append({
                **serialize_doc(a),
                "buyer_sku_id": sku.get('buyer_sku_id', ''),
                "bidso_sku": sku.get('bidso_sku', ''),
                "description": sku.get('description', ''),
                "brand": sku.get('brand', ''),
                "vertical": sku.get('vertical', ''),
                "model": sku.get('model', '')
            })
    
    return result


@router.delete("/sku-branch-assignments/{sku_id}/{branch}")
async def delete_sku_branch_assignment(sku_id: str, branch: str):
    """Remove SKU assignment from a branch"""
    await db.sku_branch_assignments.delete_one({"sku_id": sku_id, "branch": branch})
    return {"message": "Assignment removed"}


@router.post("/sku-branch-assignments/bulk-subscribe")
async def bulk_subscribe_skus(
    branch: str,
    vertical: Optional[str] = None,
    model: Optional[str] = None
):
    """Bulk subscribe all SKUs matching vertical and/or model to a branch. Also activates corresponding RMs."""
    if not branch:
        raise HTTPException(status_code=400, detail="Branch is required")
    
    if not vertical and not model:
        raise HTTPException(status_code=400, detail="At least vertical or model must be specified")
    
    # Build query for matching SKUs using new model
    # Need to find via bidso_skus since vertical/model are there
    bidso_query = {}
    if vertical:
        vertical_doc = await db.verticals.find_one({"name": vertical}, {"_id": 0, "id": 1})
        if vertical_doc:
            bidso_query["vertical_id"] = vertical_doc["id"]
    if model:
        model_doc = await db.models.find_one({"name": model}, {"_id": 0, "id": 1})
        if model_doc:
            bidso_query["model_id"] = model_doc["id"]
    
    # Find matching bidso_skus
    bidso_skus = await db.bidso_skus.find(bidso_query, {"_id": 0, "bidso_sku_id": 1}).to_list(5000)
    bidso_ids = [b["bidso_sku_id"] for b in bidso_skus]
    
    if not bidso_ids:
        return {
            "assigned": 0,
            "skipped": 0,
            "total_matching": 0,
            "rms_activated": 0,
            "message": "No SKUs found matching the criteria"
        }
    
    # Find buyer_skus linked to these bidso_skus
    matching_buyer_skus = await db.buyer_skus.find(
        {"bidso_sku_id": {"$in": bidso_ids}},
        {"_id": 0, "buyer_sku_id": 1}
    ).to_list(10000)
    
    if not matching_buyer_skus:
        return {
            "assigned": 0,
            "skipped": 0,
            "total_matching": 0,
            "rms_activated": 0,
            "message": "No SKUs found matching the criteria"
        }
    
    assigned_count = 0
    skipped_count = 0
    total_rms_activated = 0
    
    for sku in matching_buyer_skus:
        sku_id = sku['buyer_sku_id']
        
        # Check if already assigned
        existing = await db.sku_branch_assignments.find_one(
            {"sku_id": sku_id, "branch": branch},
            {"_id": 0}
        )
        
        if existing:
            skipped_count += 1
            continue
        
        # Create assignment
        assignment = SKUBranchAssignment(sku_id=sku_id, branch=branch)
        doc = assignment.model_dump()
        doc['assigned_at'] = doc['assigned_at'].isoformat()
        await db.sku_branch_assignments.insert_one(doc)
        
        # Also activate SKU in branch inventory (atomic upsert)
        await db.branch_sku_inventory.update_one(
            {"buyer_sku_id": sku_id, "branch": branch},
            {
                "$set": {"is_active": True},
                "$setOnInsert": {"id": str(uuid.uuid4()), "current_stock": 0, "activated_at": datetime.now(timezone.utc).isoformat()}
            },
            upsert=True
        )
        
        # Activate corresponding RMs for this SKU
        rms_activated = await activate_rms_for_sku(sku_id, branch)
        total_rms_activated += rms_activated
        
        assigned_count += 1
    
    return {
        "assigned": assigned_count,
        "skipped": skipped_count,
        "total_matching": len(matching_skus),
        "rms_activated": total_rms_activated,
        "message": f"Subscribed {assigned_count} SKUs to {branch}, activated {total_rms_activated} RMs"
    }


@router.delete("/sku-branch-assignments/bulk-unsubscribe")
async def bulk_unsubscribe_skus(
    branch: str,
    vertical: Optional[str] = None,
    model: Optional[str] = None
):
    """Bulk unsubscribe all SKUs matching vertical and/or model from a branch"""
    if not branch:
        raise HTTPException(status_code=400, detail="Branch is required")
    
    if not vertical and not model:
        raise HTTPException(status_code=400, detail="At least vertical or model must be specified")
    
    # Build query for matching SKUs using new model
    bidso_query = {}
    if vertical:
        vertical_doc = await db.verticals.find_one({"name": vertical}, {"_id": 0, "id": 1})
        if vertical_doc:
            bidso_query["vertical_id"] = vertical_doc["id"]
    if model:
        model_doc = await db.models.find_one({"name": model}, {"_id": 0, "id": 1})
        if model_doc:
            bidso_query["model_id"] = model_doc["id"]
    
    # Find matching bidso_skus
    bidso_skus = await db.bidso_skus.find(bidso_query, {"_id": 0, "bidso_sku_id": 1}).to_list(5000)
    bidso_ids = [b["bidso_sku_id"] for b in bidso_skus]
    
    if not bidso_ids:
        return {"removed": 0, "message": "No matching SKUs found"}
    
    # Find buyer_skus linked to these bidso_skus
    matching_buyer_skus = await db.buyer_skus.find(
        {"bidso_sku_id": {"$in": bidso_ids}},
        {"_id": 0, "buyer_sku_id": 1}
    ).to_list(10000)
    sku_ids = [s['buyer_sku_id'] for s in matching_buyer_skus]
    
    if not sku_ids:
        return {"removed": 0, "message": "No matching SKUs found"}
    
    # Remove assignments
    result = await db.sku_branch_assignments.delete_many({
        "sku_id": {"$in": sku_ids},
        "branch": branch
    })
    
    return {
        "removed": result.deleted_count,
        "message": f"Removed {result.deleted_count} SKU assignments from {branch}"
    }
