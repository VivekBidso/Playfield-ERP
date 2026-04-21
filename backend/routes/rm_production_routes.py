"""
In-House RM Production Routes
- RM Categories CRUD
- RM BOM CRUD  
- RM Production Inward (Preview & Confirm)
- Production Reports
"""
from fastapi import APIRouter, HTTPException, Depends, Query, File, UploadFile
from typing import Optional, List
from datetime import datetime, timezone
import uuid

from models.production import (
    RMCategory, RMCategoryCreate, RMCategoryUpdate, DescriptionColumn,
    RMBOM, RMBOMCreate, RMBOMUpdate, BOMComponent,
    ProductionLog, ProductionLogCreate,
    ProductionPreviewRequest, ProductionPreviewResponse, ComponentConsumptionPreview,
    ProductionConfirmRequest, ProductionConfirmResponse, ConsumedComponent,
    SourceType
)
from models.auth import User
from services.auth_service import get_current_user
from database import db

router = APIRouter(tags=["In-House RM Production"])


# ============ RM Categories (Production) ============

@router.get("/production/rm-categories")
async def get_rm_categories(
    is_active: Optional[bool] = None,
    source_type: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all RM categories with production settings"""
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    if source_type:
        query["default_source_type"] = source_type
    
    categories = await db.rm_categories.find(query, {"_id": 0}).sort("code", 1).to_list(100)
    return categories


@router.get("/production/rm-categories/{code}")
async def get_rm_category(code: str, current_user: User = Depends(get_current_user)):
    """Get a single RM category by code"""
    category = await db.rm_categories.find_one({"code": code.upper()}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return category


@router.post("/production/rm-categories")
async def create_rm_category(
    category: RMCategoryCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new RM category"""
    existing = await db.rm_categories.find_one({"code": category.code.upper()})
    if existing:
        raise HTTPException(status_code=400, detail="Category code already exists")
    
    now = datetime.now(timezone.utc)
    cat_doc = {
        "id": str(uuid.uuid4()),
        "code": category.code.upper(),
        "name": category.name,
        "description": category.description,
        "default_source_type": category.default_source_type,
        "default_bom_level": category.default_bom_level,
        "default_uom": category.default_uom,
        "rm_id_prefix": category.rm_id_prefix or category.code.upper(),
        "description_columns": [col.dict() for col in category.description_columns] if category.description_columns else [],
        "next_sequence": category.next_sequence,
        "is_active": category.is_active,
        "created_at": now,
        "updated_at": now
    }
    
    await db.rm_categories.insert_one(cat_doc)
    del cat_doc["_id"]
    return {"message": "Category created", "category": cat_doc}


@router.put("/production/rm-categories/{code}")
async def update_rm_category(
    code: str,
    update: RMCategoryUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update an RM category"""
    existing = await db.rm_categories.find_one({"code": code.upper()})
    if not existing:
        raise HTTPException(status_code=404, detail="Category not found")
    
    update_data = {}
    update_dict = update.dict()
    
    for k, v in update_dict.items():
        if v is not None:
            if k == "description_columns" and v:
                # Convert DescriptionColumn objects to dicts
                update_data[k] = [col if isinstance(col, dict) else col.dict() for col in v]
            else:
                update_data[k] = v
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.rm_categories.update_one({"code": code.upper()}, {"$set": update_data})
    
    updated = await db.rm_categories.find_one({"code": code.upper()}, {"_id": 0})
    return {"message": "Category updated", "category": updated}


@router.post("/production/rm-categories/{code}/generate-rm-id")
async def generate_rm_id(
    code: str,
    current_user: User = Depends(get_current_user)
):
    """Generate the next RM ID for a category"""
    category = await db.rm_categories.find_one({"code": code.upper()})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    prefix = category.get("rm_id_prefix", code.upper())
    seq = category.get("next_sequence", 1)
    
    # Generate RM ID
    rm_id = f"{prefix}_{seq:03d}"
    
    # Increment sequence
    await db.rm_categories.update_one(
        {"code": code.upper()},
        {"$inc": {"next_sequence": 1}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"rm_id": rm_id, "next_sequence": seq + 1}


# ============ RM BOM ============

@router.get("/rm-bom")
async def get_all_boms(
    category: Optional[str] = None,
    bom_level: Optional[int] = None,
    is_active: Optional[bool] = True,
    current_user: User = Depends(get_current_user)
):
    """Get all RM BOMs"""
    query = {}
    if category:
        query["category"] = category.upper()
    if bom_level:
        query["bom_level"] = bom_level
    if is_active is not None:
        query["is_active"] = is_active
    
    boms = await db.rm_bom.find(query, {"_id": 0}).sort("rm_id", 1).to_list(1000)
    return boms


@router.get("/rm-bom/{rm_id}")
async def get_bom(rm_id: str, current_user: User = Depends(get_current_user)):
    """Get BOM for a specific RM"""
    bom = await db.rm_bom.find_one({"rm_id": rm_id.upper()}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found for this RM")
    return bom


@router.post("/rm-bom")
async def create_bom(
    bom: RMBOMCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new RM BOM"""
    existing = await db.rm_bom.find_one({"rm_id": bom.rm_id.upper()})
    if existing:
        raise HTTPException(status_code=400, detail="BOM already exists for this RM. Use PUT to update.")
    
    rm = await db.raw_materials.find_one({"rm_id": bom.rm_id.upper()})
    if not rm:
        raise HTTPException(status_code=404, detail=f"RM {bom.rm_id} not found")
    
    components_with_names = []
    for comp in bom.components:
        comp_rm = await db.raw_materials.find_one({"rm_id": comp.component_rm_id.upper()})
        if not comp_rm:
            raise HTTPException(status_code=404, detail=f"Component RM {comp.component_rm_id} not found")
        comp_dict = comp.dict()
        comp_dict["component_name"] = comp_rm.get("description", comp.component_rm_id)
        comp_dict["component_rm_id"] = comp.component_rm_id.upper()
        components_with_names.append(comp_dict)
    
    now = datetime.now(timezone.utc)
    bom_doc = {
        "id": str(uuid.uuid4()),
        "rm_id": bom.rm_id.upper(),
        "rm_name": rm.get("description", bom.rm_id),
        "category": bom.category.upper(),
        "bom_level": bom.bom_level,
        "output_qty": bom.output_qty,
        "output_uom": bom.output_uom,
        "components": components_with_names,
        "total_weight_per_unit": bom.total_weight_per_unit,
        "yield_factor": bom.yield_factor,
        "is_active": bom.is_active,
        "created_at": now,
        "updated_at": now
    }
    
    await db.rm_bom.insert_one(bom_doc)
    
    await db.raw_materials.update_one(
        {"rm_id": bom.rm_id.upper()},
        {"$set": {"has_bom": True, "bom_level": bom.bom_level}}
    )
    
    del bom_doc["_id"]
    return {"message": "BOM created", "bom": bom_doc}


@router.put("/rm-bom/{rm_id}")
async def update_bom(
    rm_id: str,
    update: RMBOMUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update an RM BOM"""
    existing = await db.rm_bom.find_one({"rm_id": rm_id.upper()})
    if not existing:
        raise HTTPException(status_code=404, detail="BOM not found for this RM")
    
    update_data = {}
    
    if update.components is not None:
        components_with_names = []
        for comp in update.components:
            comp_rm = await db.raw_materials.find_one({"rm_id": comp.component_rm_id.upper()})
            if not comp_rm:
                raise HTTPException(status_code=404, detail=f"Component RM {comp.component_rm_id} not found")
            comp_dict = comp.dict()
            comp_dict["component_name"] = comp_rm.get("description", comp.component_rm_id)
            comp_dict["component_rm_id"] = comp.component_rm_id.upper()
            components_with_names.append(comp_dict)
        update_data["components"] = components_with_names
    
    for field in ["rm_name", "output_qty", "output_uom", "total_weight_per_unit", "yield_factor", "is_active"]:
        val = getattr(update, field)
        if val is not None:
            update_data[field] = val
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.rm_bom.update_one({"rm_id": rm_id.upper()}, {"$set": update_data})
    
    updated = await db.rm_bom.find_one({"rm_id": rm_id.upper()}, {"_id": 0})
    return {"message": "BOM updated", "bom": updated}


@router.delete("/rm-bom/{rm_id}")
async def delete_bom(rm_id: str, current_user: User = Depends(get_current_user)):
    """Delete an RM BOM"""
    result = await db.rm_bom.delete_one({"rm_id": rm_id.upper()})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    await db.raw_materials.update_one(
        {"rm_id": rm_id.upper()},
        {"$set": {"has_bom": False}}
    )
    
    return {"message": "BOM deleted"}



@router.post("/rm-bom/bulk-upload")
async def bulk_upload_bom(
    file: UploadFile = File(...),
    mode: str = Query("skip", description="'skip' existing or 'replace' existing BOMs"),
    current_user: User = Depends(get_current_user)
):
    """
    Bulk upload RM BOMs from Excel.
    
    Expected columns: RM ID, BOM RM ID, Weight in gm / Pc, Wastage %
    
    Multiple rows with same RM ID are grouped into one BOM.
    UOM is auto-detected from the component RM's category default_uom.
    BOM level is auto-detected from the output RM's category default_bom_level.
    
    Args:
        mode: 'skip' = skip if BOM already exists, 'replace' = overwrite existing BOM
    """
    import openpyxl
    from io import BytesIO
    
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    
    # Parse headers
    headers_raw = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    # Map flexible header names
    header_map = {}
    for idx, h in enumerate(headers_raw):
        if h in ("rm id", "rm_id", "output_rm_id", "output rm id"):
            header_map["rm_id"] = idx
        elif h in ("bom rm id", "bom_rm_id", "component_rm_id", "component rm id", "bom rm", "component"):
            header_map["bom_rm_id"] = idx
        elif h in ("weight in gm / pc", "weight in gm/pc", "weight", "quantity", "qty", "consumption", "weight_gm"):
            header_map["quantity"] = idx
        elif h in ("wastage %", "wastage%", "wastage", "wastage_pct", "waste %", "waste"):
            header_map["wastage"] = idx
    
    required = ["rm_id", "bom_rm_id", "quantity"]
    missing = [k for k in required if k not in header_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}. Found: {headers_raw}")
    
    # Pre-load category configs for UOM and BOM level lookup
    cat_docs = await db.rm_categories.find({}, {"_id": 0, "code": 1, "default_uom": 1, "default_bom_level": 1}).to_list(100)
    cat_config = {c["code"]: c for c in cat_docs}
    
    # Parse all rows and group by RM ID
    grouped = {}  # rm_id -> [components]
    row_errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        rm_id = str(row[header_map["rm_id"]] or "").strip().upper()
        bom_rm_id = str(row[header_map["bom_rm_id"]] or "").strip().upper()
        quantity_raw = row[header_map["quantity"]]
        wastage_raw = row[header_map.get("wastage", -1)] if "wastage" in header_map else 0
        
        if not rm_id or not bom_rm_id:
            row_errors.append(f"Row {idx}: Missing RM ID or BOM RM ID")
            continue
        
        try:
            quantity = float(quantity_raw) if quantity_raw else 0
        except (ValueError, TypeError):
            row_errors.append(f"Row {idx}: Invalid quantity '{quantity_raw}'")
            continue
        
        # Parse wastage: handle "2%", "0.02", "2", etc.
        try:
            wastage_str = str(wastage_raw or "0").replace("%", "").strip()
            wastage_pct = float(wastage_str) if wastage_str else 0
            # If value > 1, assume it's a percentage (e.g., 2 means 2%)
            if wastage_pct > 1:
                wastage_factor = 1 + (wastage_pct / 100)
            elif wastage_pct > 0:
                wastage_factor = 1 + wastage_pct
            else:
                wastage_factor = 1.0
        except (ValueError, TypeError):
            wastage_factor = 1.0
        
        if rm_id not in grouped:
            grouped[rm_id] = []
        
        grouped[rm_id].append({
            "component_rm_id": bom_rm_id,
            "quantity": quantity,
            "wastage_factor": round(wastage_factor, 4),
            "row": idx
        })
    
    # Process each grouped BOM
    created = 0
    replaced = 0
    skipped = 0
    errors = []
    processed_boms = []
    
    for rm_id, components in grouped.items():
        # Validate output RM exists
        rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0, "rm_id": 1, "category": 1, "description": 1})
        if not rm:
            errors.append(f"{rm_id}: Output RM not found in database")
            continue
        
        category = rm.get("category", "")
        cat_cfg = cat_config.get(category, {})
        bom_level = cat_cfg.get("default_bom_level", 2)
        
        # Check if BOM already exists
        existing = await db.rm_bom.find_one({"rm_id": rm_id})
        if existing and mode == "skip":
            skipped += 1
            continue
        
        # Validate and enrich each component
        enriched_components = []
        has_error = False
        total_weight = 0
        
        for comp in components:
            comp_rm = await db.raw_materials.find_one(
                {"rm_id": comp["component_rm_id"]},
                {"_id": 0, "rm_id": 1, "category": 1, "description": 1}
            )
            if not comp_rm:
                errors.append(f"{rm_id}: Component RM {comp['component_rm_id']} not found (row {comp['row']})")
                has_error = True
                break
            
            # Auto-detect UOM from component's category
            comp_cat = comp_rm.get("category", "")
            comp_cat_cfg = cat_config.get(comp_cat, {})
            comp_uom = comp_cat_cfg.get("default_uom", "PCS") or "PCS"
            
            # If UOM is KG, the quantity column is grams — convert display
            # Store as-is (grams for weight-based, pieces for piece-based)
            enriched_components.append({
                "component_rm_id": comp["component_rm_id"],
                "component_name": comp_rm.get("description", comp["component_rm_id"]),
                "quantity": comp["quantity"],
                "uom": comp_uom,
                "wastage_factor": comp["wastage_factor"],
                "percentage": None
            })
            
            total_weight += comp["quantity"]
        
        if has_error:
            continue
        
        # Build BOM document
        now = datetime.now(timezone.utc)
        bom_doc = {
            "id": str(uuid.uuid4()),
            "rm_id": rm_id,
            "rm_name": rm.get("description", rm_id),
            "category": category,
            "bom_level": bom_level,
            "output_qty": 1.0,
            "output_uom": "PCS",
            "components": enriched_components,
            "total_weight_per_unit": round(total_weight, 2),
            "yield_factor": 1.0,
            "is_active": True,
            "created_at": now,
            "updated_at": now
        }
        
        if existing and mode == "replace":
            await db.rm_bom.replace_one({"rm_id": rm_id}, bom_doc)
            replaced += 1
        else:
            await db.rm_bom.insert_one(bom_doc)
            created += 1
        
        # Mark RM as having a BOM
        await db.raw_materials.update_one(
            {"rm_id": rm_id},
            {"$set": {"has_bom": True, "bom_level": bom_level}}
        )
        
        processed_boms.append({
            "rm_id": rm_id,
            "components": len(enriched_components),
            "action": "replaced" if (existing and mode == "replace") else "created"
        })
    
    return {
        "message": f"BOM bulk upload complete. Created: {created}, Replaced: {replaced}, Skipped: {skipped}",
        "created": created,
        "replaced": replaced,
        "skipped": skipped,
        "total_in_file": len(grouped),
        "errors": errors[:50],
        "row_errors": row_errors[:50],
        "processed": processed_boms[:50]
    }



# ============ RM Production Inward ============

@router.post("/rm-production/preview")
async def preview_rm_production(
    request: ProductionPreviewRequest,
    current_user: User = Depends(get_current_user)
):
    """Preview RM production - check BOM and stock availability"""
    bom = await db.rm_bom.find_one({"rm_id": request.rm_id.upper(), "is_active": True}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail=f"No active BOM found for {request.rm_id}")
    
    rm = await db.raw_materials.find_one({"rm_id": request.rm_id.upper()}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail=f"RM {request.rm_id} not found")
    
    source_type = rm.get("source_type", "PURCHASED")
    if source_type == "PURCHASED":
        raise HTTPException(
            status_code=400, 
            detail=f"RM {request.rm_id} is marked as PURCHASED only. Cannot produce."
        )
    
    components_preview = []
    blocking_components = []
    can_produce = True
    
    units_to_produce = request.quantity_to_produce / bom.get("output_qty", 1)
    
    for comp in bom.get("components", []):
        comp_rm_id = comp["component_rm_id"]
        base_qty = comp["quantity"] * units_to_produce
        wastage = comp.get("wastage_factor", 1.0)
        required_qty = round(base_qty * wastage, 4)
        
        inventory = await db.branch_rm_inventory.find_one({
            "rm_id": comp_rm_id,
            "branch": request.branch
        })
        available_stock = inventory.get("current_stock", 0) if inventory else 0
        
        is_sufficient = available_stock >= required_qty
        shortage = max(0, round(required_qty - available_stock, 4))
        
        if not is_sufficient:
            can_produce = False
            blocking_components.append(comp_rm_id)
        
        components_preview.append(ComponentConsumptionPreview(
            rm_id=comp_rm_id,
            name=comp.get("component_name", comp_rm_id),
            required_qty=required_qty,
            uom=comp.get("uom", "PCS"),
            available_stock=round(available_stock, 4),
            is_sufficient=is_sufficient,
            shortage=shortage
        ))
    
    return ProductionPreviewResponse(
        rm_id=request.rm_id.upper(),
        rm_name=rm.get("description", request.rm_id),
        category=bom.get("category", ""),
        bom_level=bom.get("bom_level", 2),
        quantity_to_produce=request.quantity_to_produce,
        output_uom=bom.get("output_uom", "PCS"),
        components=components_preview,
        can_produce=can_produce,
        blocking_components=blocking_components
    )


@router.post("/rm-production/confirm")
async def confirm_rm_production(
    request: ProductionConfirmRequest,
    current_user: User = Depends(get_current_user)
):
    """Confirm RM production - consume components and add produced RM to inventory"""
    preview = await preview_rm_production(
        ProductionPreviewRequest(
            branch=request.branch,
            rm_id=request.rm_id,
            quantity_to_produce=request.quantity_produced
        ),
        current_user
    )
    
    if not preview.can_produce:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot produce: Insufficient stock for {', '.join(preview.blocking_components)}"
        )
    
    bom = await db.rm_bom.find_one({"rm_id": request.rm_id.upper()}, {"_id": 0})
    rm = await db.raw_materials.find_one({"rm_id": request.rm_id.upper()}, {"_id": 0})
    
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    count = await db.rm_production_log.count_documents({"production_code": {"$regex": f"^RMPROD_{date_str}"}})
    production_code = f"RMPROD_{date_str}_{count + 1:04d}"
    
    components_consumed = []
    
    for comp in preview.components:
        inv_before = await db.branch_rm_inventory.find_one({
            "rm_id": comp.rm_id,
            "branch": request.branch
        })
        stock_before = inv_before.get("current_stock", 0) if inv_before else 0
        stock_after = round(stock_before - comp.required_qty, 4)
        
        await db.branch_rm_inventory.update_one(
            {"rm_id": comp.rm_id, "branch": request.branch},
            {
                "$inc": {"current_stock": -comp.required_qty},
                "$set": {"updated_at": today}
            },
            upsert=True
        )
        
        components_consumed.append(ConsumedComponent(
            rm_id=comp.rm_id,
            name=comp.name,
            quantity_consumed=comp.required_qty,
            uom=comp.uom,
            stock_before=stock_before,
            stock_after=stock_after
        ))
    
    inv_before = await db.branch_rm_inventory.find_one({
        "rm_id": request.rm_id.upper(),
        "branch": request.branch
    })
    produced_stock_before = inv_before.get("current_stock", 0) if inv_before else 0
    
    await db.branch_rm_inventory.update_one(
        {"rm_id": request.rm_id.upper(), "branch": request.branch},
        {
            "$inc": {"current_stock": request.quantity_produced},
            "$set": {"updated_at": today}
        },
        upsert=True
    )
    
    new_inv = await db.branch_rm_inventory.find_one({
        "rm_id": request.rm_id.upper(),
        "branch": request.branch
    })
    new_stock = new_inv.get("current_stock", 0) if new_inv else request.quantity_produced
    
    production_date = request.production_date or today.strftime("%Y-%m-%d")
    log_doc = {
        "id": str(uuid.uuid4()),
        "production_code": production_code,
        "branch": request.branch,
        "rm_id": request.rm_id.upper(),
        "rm_name": rm.get("description", request.rm_id),
        "category": bom.get("category", ""),
        "bom_level": bom.get("bom_level", 2),
        "quantity_produced": request.quantity_produced,
        "uom": bom.get("output_uom", "PCS"),
        "stock_before": produced_stock_before,
        "stock_after": new_stock,
        "components_consumed": [c.dict() for c in components_consumed],
        "notes": request.notes,
        "production_date": production_date,
        "produced_by": current_user.id,
        "produced_by_name": current_user.name or current_user.email,
        "created_at": today
    }
    
    await db.rm_production_log.insert_one(log_doc)
    del log_doc["_id"]
    
    return ProductionConfirmResponse(
        success=True,
        production_code=production_code,
        rm_id=request.rm_id.upper(),
        quantity_produced=request.quantity_produced,
        components_consumed=components_consumed,
        message=f"Produced {request.quantity_produced} {bom.get('output_uom', 'PCS')} of {rm.get('description', request.rm_id)}. New stock: {new_stock}"
    )


# ============ Production Log & Reports ============

@router.get("/rm-production/log")
async def get_rm_production_log(
    branch: Optional[str] = None,
    category: Optional[str] = None,
    rm_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user)
):
    """Get RM production log with filters"""
    query = {}
    
    if branch:
        query["branch"] = branch
    if category:
        query["category"] = category.upper()
    if rm_id:
        query["rm_id"] = rm_id.upper()
    if start_date:
        query["production_date"] = query.get("production_date", {})
        query["production_date"]["$gte"] = start_date
    if end_date:
        query["production_date"] = query.get("production_date", {})
        query["production_date"]["$lte"] = end_date
    
    total = await db.rm_production_log.count_documents(query)
    skip = (page - 1) * page_size
    
    logs = await db.rm_production_log.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": logs,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/rm-production/summary")
async def get_rm_production_summary(
    branch: str,
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Get RM production summary by category for a branch"""
    pipeline = [
        {
            "$match": {
                "branch": branch,
                "production_date": {"$gte": start_date, "$lte": end_date}
            }
        },
        {
            "$group": {
                "_id": "$category",
                "total_produced": {"$sum": "$quantity_produced"},
                "production_count": {"$sum": 1},
                "items": {"$addToSet": "$rm_id"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "category": "$_id",
                "total_produced": {"$round": ["$total_produced", 2]},
                "production_count": 1,
                "unique_items": {"$size": "$items"}
            }
        },
        {"$sort": {"category": 1}}
    ]
    
    summary = await db.rm_production_log.aggregate(pipeline).to_list(100)
    
    total_produced = sum(s["total_produced"] for s in summary)
    total_entries = sum(s["production_count"] for s in summary)
    
    return {
        "branch": branch,
        "period": {"start": start_date, "end": end_date},
        "categories": summary,
        "totals": {
            "total_produced": round(total_produced, 2),
            "total_entries": total_entries
        }
    }


@router.get("/rm-production/consumption-report")
async def get_rm_consumption_report(
    branch: str,
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Get L1 material consumption report from RM production"""
    pipeline = [
        {
            "$match": {
                "branch": branch,
                "production_date": {"$gte": start_date, "$lte": end_date}
            }
        },
        {"$unwind": "$components_consumed"},
        {
            "$group": {
                "_id": {
                    "rm_id": "$components_consumed.rm_id",
                    "name": "$components_consumed.name",
                    "uom": "$components_consumed.uom"
                },
                "total_consumed": {"$sum": "$components_consumed.quantity_consumed"},
                "used_in_count": {"$sum": 1}
            }
        },
        {
            "$project": {
                "_id": 0,
                "rm_id": "$_id.rm_id",
                "name": "$_id.name",
                "uom": "$_id.uom",
                "total_consumed": {"$round": ["$total_consumed", 2]},
                "used_in_count": 1
            }
        },
        {"$sort": {"total_consumed": -1}}
    ]
    
    consumption = await db.rm_production_log.aggregate(pipeline).to_list(500)
    
    return {
        "branch": branch,
        "period": {"start": start_date, "end": end_date},
        "consumption": consumption
    }


# ============ Manufacturable RMs ============

@router.get("/rm-production/manufacturable-rms")
async def get_manufacturable_rms(
    branch: str,
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get list of RMs that can be manufactured (have BOM and source_type allows)"""
    bom_query = {"is_active": True}
    if category:
        bom_query["category"] = category.upper()
    
    boms = await db.rm_bom.find(bom_query, {"_id": 0}).to_list(1000)
    
    result = []
    for bom in boms:
        rm = await db.raw_materials.find_one({"rm_id": bom["rm_id"]}, {"_id": 0})
        if rm:
            source_type = rm.get("source_type", "PURCHASED")
            if source_type in ["MANUFACTURED", "BOTH"]:
                inv = await db.branch_rm_inventory.find_one({
                    "rm_id": bom["rm_id"],
                    "branch": branch
                })
                current_stock = inv.get("current_stock", 0) if inv else 0
                
                result.append({
                    "rm_id": bom["rm_id"],
                    "rm_name": bom.get("rm_name", rm.get("description", "")),
                    "category": bom["category"],
                    "bom_level": bom["bom_level"],
                    "source_type": source_type,
                    "current_stock": round(current_stock, 2),
                    "uom": rm.get("uom", "PCS")
                })
    
    result.sort(key=lambda x: (x["category"], x["rm_id"]))
    return result


@router.get("/rm-production/active-categories")
async def get_active_production_categories(
    branch: str,
    current_user: User = Depends(get_current_user)
):
    """Get categories that have manufacturable RMs with BOMs"""
    boms = await db.rm_bom.find({"is_active": True}, {"_id": 0, "rm_id": 1, "category": 1}).to_list(1000)
    
    active_categories = set()
    for bom in boms:
        rm = await db.raw_materials.find_one({"rm_id": bom["rm_id"]}, {"_id": 0, "source_type": 1})
        if rm and rm.get("source_type") in ["MANUFACTURED", "BOTH"]:
            active_categories.add(bom["category"])
    
    categories = []
    for cat_code in active_categories:
        cat = await db.rm_categories.find_one({"code": cat_code}, {"_id": 0})
        if cat:
            categories.append(cat)
        else:
            categories.append({"code": cat_code, "name": cat_code})
    
    categories.sort(key=lambda x: x["code"])
    return categories
