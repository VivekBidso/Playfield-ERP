"""
Bidso SKU and Buyer SKU Routes

Endpoints for managing the two-level SKU architecture:
- Bidso SKU: Base product with common BOM
- Buyer SKU: Branded variants with brand-specific additions

BOM Management:
- Common BOM: Locked at Bidso SKU level
- Brand-specific BOM: Per-brand additions (labels, packaging)
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query
from pydantic import BaseModel
from datetime import datetime, timezone
from typing import Optional, List
import uuid
import io
import re
import openpyxl

from database import db
from models.sku_models import (
    BidsoSKU, BidsoSKUCreate, BidsoSKUUpdate,
    BuyerSKU, BuyerSKUCreate, BuyerSKUUpdate,
    CommonBOM, CommonBOMCreate, BOMItem,
    BrandSpecificBOM, BrandSpecificBOMCreate,
    FullBOM, SKUMigrationResult
)
from services.utils import get_current_user, serialize_doc, get_next_rm_sequence, generate_rm_name

router = APIRouter(prefix="/sku-management", tags=["SKU Management"])


# ============ Helper Functions ============

def generate_rm_description(category: str, cat_data: dict, fallback_name: str = "") -> str:
    """
    Generate RM description based on category-specific naming conventions.
    
    Labels (LB): {Type}_{Buyer SKU}
    Packaging (PM): {Model}_{Type}_{Specs}_{Brand}
    Brand Assets (BS): {Position}_{Type}_{Brand}_{Buyer SKU}
    In House Plastic (INP): {Mould Code}_{Model Name}_{Part Name}_{Colour}_{Masterbatch}
    Accessories (ACC): {Type}_{Model Name}_{Specs}_{Colour}
    In House Metal (INM): {Model Name}_{Part Name}_{Colour}_{Masterbatch}
    Spares (SP): {Type}_{Specs}
    Electronic Components (ELC): {Model}_{Type}_{Specs}
    """
    parts = []
    
    if category == "LB":
        # Labels: {Type}_{Buyer SKU}
        if cat_data.get("type"):
            parts.append(cat_data["type"])
        if cat_data.get("buyer_sku"):
            parts.append(cat_data["buyer_sku"])
            
    elif category == "PM":
        # Packaging: {Model}_{Type}_{Specs}_{Brand}
        if cat_data.get("model"):
            parts.append(cat_data["model"])
        if cat_data.get("type"):
            parts.append(cat_data["type"])
        if cat_data.get("specs"):
            parts.append(cat_data["specs"])
        if cat_data.get("brand"):
            parts.append(cat_data["brand"])
            
    elif category == "BS":
        # Brand Assets: {Position}_{Type}_{Brand}_{Buyer SKU}
        if cat_data.get("position"):
            parts.append(cat_data["position"])
        if cat_data.get("type"):
            parts.append(cat_data["type"])
        if cat_data.get("brand"):
            parts.append(cat_data["brand"])
        if cat_data.get("buyer_sku"):
            parts.append(cat_data["buyer_sku"])
            
    elif category == "INP":
        # In House Plastic: {Mould Code}_{Model Name}_{Part Name}_{Colour}_{Masterbatch}
        if cat_data.get("mould_code"):
            parts.append(cat_data["mould_code"])
        if cat_data.get("model_name"):
            parts.append(cat_data["model_name"])
        if cat_data.get("part_name"):
            parts.append(cat_data["part_name"])
        if cat_data.get("colour"):
            parts.append(cat_data["colour"])
        if cat_data.get("mb"):
            parts.append(cat_data["mb"])
            
    elif category == "ACC":
        # Accessories: {Type}_{Model Name}_{Specs}_{Colour}
        if cat_data.get("type"):
            parts.append(cat_data["type"])
        if cat_data.get("model_name"):
            parts.append(cat_data["model_name"])
        if cat_data.get("specs"):
            parts.append(cat_data["specs"])
        if cat_data.get("colour"):
            parts.append(cat_data["colour"])
            
    elif category == "INM":
        # In House Metal: {Model Name}_{Part Name}_{Colour}_{Masterbatch}
        if cat_data.get("model_name"):
            parts.append(cat_data["model_name"])
        if cat_data.get("part_name"):
            parts.append(cat_data["part_name"])
        if cat_data.get("colour") or cat_data.get("color"):
            parts.append(cat_data.get("colour") or cat_data.get("color"))
        if cat_data.get("mb"):
            parts.append(cat_data["mb"])
            
    elif category == "SP":
        # Spares: {Type}_{Specs}
        if cat_data.get("type"):
            parts.append(cat_data["type"])
        if cat_data.get("specs"):
            parts.append(cat_data["specs"])
            
    elif category == "ELC":
        # Electronic Components: {Model}_{Type}_{Specs}
        if cat_data.get("model"):
            parts.append(cat_data["model"])
        if cat_data.get("type"):
            parts.append(cat_data["type"])
        if cat_data.get("specs"):
            parts.append(cat_data["specs"])
    
    # Join parts with underscore, filter out empty strings and convert to string
    description = "_".join(str(p) for p in parts if p is not None and str(p).strip())
    
    # If no description generated, use fallback
    if not description and fallback_name:
        description = fallback_name
    
    return description


async def get_next_numeric_code(vertical_code: str, model_code: str) -> str:
    """Generate next available numeric code for a Bidso SKU"""
    prefix = f"{vertical_code}_{model_code}_"
    
    # Find all existing Bidso SKUs with this prefix
    existing = await db.bidso_skus.find(
        {"bidso_sku_id": {"$regex": f"^{prefix}"}},
        {"bidso_sku_id": 1, "_id": 0}
    ).to_list(10000)
    
    max_num = 0
    for item in existing:
        sku_id = item.get("bidso_sku_id", "")
        parts = sku_id.split("_")
        if len(parts) >= 3:
            try:
                num = int(parts[-1])
                if num > max_num:
                    max_num = num
            except ValueError:
                continue
    
    return str(max_num + 1).zfill(3)


async def generate_bidso_sku_id(vertical_code: str, model_code: str, numeric_code: str) -> str:
    """Generate Bidso SKU ID: {VerticalCode}_{ModelCode}_{NumericCode}"""
    return f"{vertical_code}_{model_code}_{numeric_code}"


async def generate_buyer_sku_id(brand_code: str, bidso_sku_id: str) -> str:
    """Generate Buyer SKU ID: {BrandCode}_{BidsoSKU}"""
    return f"{brand_code}_{bidso_sku_id}"


# ============ Bidso SKU CRUD ============

@router.get("/bidso-skus")
async def get_bidso_skus(
    vertical_id: Optional[str] = None,
    model_id: Optional[str] = None,
    search: Optional[str] = None,
    include_inactive: bool = False,
    page: int = Query(1, ge=1, description="Page number (1-indexed)"),
    page_size: int = Query(50, ge=1, le=10000, description="Items per page (max 10000 for downloads)")
):
    """Get all Bidso SKUs with optional filters and pagination"""
    query = {}
    if not include_inactive:
        query["status"] = "ACTIVE"
    if vertical_id:
        query["vertical_id"] = vertical_id
    if model_id:
        query["model_id"] = model_id
    
    # If search is provided, we need to search in text fields
    # For MongoDB text search, we need to fetch and filter in Python
    # since the search spans multiple fields
    if search:
        search_lower = search.lower()
        # First get total matching count efficiently
        all_skus = await db.bidso_skus.find(query, {"_id": 0, "bidso_sku_id": 1, "name": 1, "description": 1}).to_list(10000)
        matching_ids = [s["bidso_sku_id"] for s in all_skus if
                       search_lower in s.get("bidso_sku_id", "").lower() or
                       search_lower in s.get("name", "").lower() or
                       search_lower in s.get("description", "").lower()]
        total = len(matching_ids)
        
        # Paginate the matching IDs
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_ids = matching_ids[start_idx:end_idx]
        
        # Fetch full documents for paginated IDs
        if paginated_ids:
            bidso_skus = await db.bidso_skus.find(
                {"bidso_sku_id": {"$in": paginated_ids}},
                {"_id": 0}
            ).to_list(page_size)
        else:
            bidso_skus = []
    else:
        # Get total count for pagination
        total = await db.bidso_skus.count_documents(query)
        
        # Calculate skip value
        skip = (page - 1) * page_size
        
        # Fetch paginated results
        bidso_skus = await db.bidso_skus.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    # Enrich with related data
    for sku in bidso_skus:
        # Get vertical name
        if sku.get("vertical_id"):
            v = await db.verticals.find_one({"id": sku["vertical_id"]}, {"_id": 0, "name": 1, "code": 1})
            sku["vertical_name"] = v["name"] if v else None
        
        # Get model name
        if sku.get("model_id"):
            m = await db.models.find_one({"id": sku["model_id"]}, {"_id": 0, "name": 1, "code": 1})
            sku["model_name"] = m["name"] if m else None
        
        # Get count of buyer SKUs
        buyer_count = await db.buyer_skus.count_documents({"bidso_sku_id": sku["bidso_sku_id"]})
        sku["buyer_sku_count"] = buyer_count
        
        # Check if BOM is locked
        common_bom = await db.common_bom.find_one({"bidso_sku_id": sku["bidso_sku_id"]}, {"_id": 0, "is_locked": 1})
        sku["has_bom"] = common_bom is not None
        sku["bom_locked"] = common_bom.get("is_locked", False) if common_bom else False
    
    # Calculate total pages
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    return {
        "items": bidso_skus,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }


@router.get("/bidso-skus/{bidso_sku_id}")
async def get_bidso_sku(bidso_sku_id: str):
    """Get a single Bidso SKU by ID"""
    sku = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
    if not sku:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    # Enrich with related data
    if sku.get("vertical_id"):
        v = await db.verticals.find_one({"id": sku["vertical_id"]}, {"_id": 0, "name": 1})
        sku["vertical_name"] = v["name"] if v else None
    
    if sku.get("model_id"):
        m = await db.models.find_one({"id": sku["model_id"]}, {"_id": 0, "name": 1})
        sku["model_name"] = m["name"] if m else None
    
    # Get buyer SKUs
    buyer_skus = await db.buyer_skus.find(
        {"bidso_sku_id": bidso_sku_id},
        {"_id": 0}
    ).to_list(1000)
    sku["buyer_skus"] = buyer_skus
    
    # Get common BOM
    common_bom = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
    sku["common_bom"] = common_bom
    
    return sku


@router.post("/bidso-skus")
async def create_bidso_sku(data: BidsoSKUCreate):
    """Create a new Bidso SKU"""
    # Get vertical info
    vertical = await db.verticals.find_one({"id": data.vertical_id}, {"_id": 0})
    if not vertical:
        raise HTTPException(status_code=404, detail="Vertical not found")
    
    # Get model info
    model = await db.models.find_one({"id": data.model_id}, {"_id": 0})
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    
    vertical_code = vertical["code"]
    model_code = model["code"]
    
    # Generate or validate numeric code
    if data.numeric_code:
        numeric_code = data.numeric_code.zfill(3)
    else:
        numeric_code = await get_next_numeric_code(vertical_code, model_code)
    
    # Generate Bidso SKU ID
    bidso_sku_id = await generate_bidso_sku_id(vertical_code, model_code, numeric_code)
    
    # Check if already exists
    existing = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id})
    if existing:
        raise HTTPException(status_code=400, detail=f"Bidso SKU {bidso_sku_id} already exists")
    
    # Create Bidso SKU
    bidso_sku = {
        "id": str(uuid.uuid4()),
        "bidso_sku_id": bidso_sku_id,
        "vertical_id": data.vertical_id,
        "vertical_code": vertical_code,
        "model_id": data.model_id,
        "model_code": model_code,
        "numeric_code": numeric_code,
        "name": data.name or f"{vertical['name']} - {model['name']}",
        "description": data.description,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.bidso_skus.insert_one(bidso_sku)
    del bidso_sku["_id"]
    
    return bidso_sku


@router.put("/bidso-skus/{bidso_sku_id}")
async def update_bidso_sku(bidso_sku_id: str, data: BidsoSKUUpdate):
    """Update a Bidso SKU"""
    existing = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    update_data = {}
    if data.name is not None:
        update_data["name"] = data.name
    if data.description is not None:
        update_data["description"] = data.description
    if data.status is not None:
        update_data["status"] = data.status
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        await db.bidso_skus.update_one(
            {"bidso_sku_id": bidso_sku_id},
            {"$set": update_data}
        )
    
    return {"message": "Bidso SKU updated"}


@router.delete("/bidso-skus/{bidso_sku_id}")
async def delete_bidso_sku(bidso_sku_id: str):
    """Soft delete a Bidso SKU (sets status to INACTIVE)"""
    # Check for active buyer SKUs
    buyer_count = await db.buyer_skus.count_documents({
        "bidso_sku_id": bidso_sku_id,
        "status": "ACTIVE"
    })
    if buyer_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete: {buyer_count} active Buyer SKUs reference this Bidso SKU"
        )
    
    result = await db.bidso_skus.update_one(
        {"bidso_sku_id": bidso_sku_id},
        {"$set": {"status": "INACTIVE", "updated_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    return {"message": "Bidso SKU deleted"}



@router.post("/bidso-skus/bulk-import")
async def bulk_import_bidso_skus(file: UploadFile = File(...)):
    """
    Bulk import Bidso SKUs from Excel file.
    
    Expected columns: bidso_sku_id, vertical_code, model_code, numeric_code, name, description
    
    This endpoint will:
    - Create new Bidso SKUs if they don't exist
    - Skip existing ones (no update)
    """
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    ws = wb.active
    results = {
        "processed": 0,
        "created": 0,
        "skipped": 0,
        "errors": []
    }
    
    # Load reference data
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(100)
    models = await db.models.find({}, {"_id": 0}).to_list(500)
    
    vertical_map = {}
    for v in verticals:
        vertical_map[v.get("code", "").upper()] = v
        vertical_map[v.get("name", "").lower()] = v
    
    model_map = {}
    for m in models:
        model_map[m.get("code", "").upper()] = m
        model_map[m.get("name", "").lower()] = m
    
    # Detect headers
    headers = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]
    
    # Find column indices
    bidso_col = None
    vertical_col = None
    model_col = None
    numeric_col = None
    name_col = None
    desc_col = None
    
    for idx, h in enumerate(headers):
        if h in ["BIDSO_SKU_ID", "BIDSO SKU ID", "BIDSO SKU", "BIDSO_SKU"]:
            bidso_col = idx
        elif h in ["VERTICAL_CODE", "VERTICAL CODE", "VERTICAL"]:
            vertical_col = idx
        elif h in ["MODEL_CODE", "MODEL CODE", "MODEL"]:
            model_col = idx
        elif h in ["NUMERIC_CODE", "NUMERIC CODE", "CODE", "NUMBER"]:
            numeric_col = idx
        elif h in ["NAME", "SKU_NAME", "SKU NAME"]:
            name_col = idx
        elif h in ["DESCRIPTION", "DESC"]:
            desc_col = idx
    
    if bidso_col is None:
        raise HTTPException(status_code=400, detail="Missing required column: BIDSO_SKU_ID")
    
    # Process rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[bidso_col]:
            continue
        
        try:
            bidso_sku_id = str(row[bidso_col]).strip()
            results["processed"] += 1
            
            # Check if already exists
            existing = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id})
            if existing:
                results["skipped"] += 1
                continue
            
            # Try to parse vertical and model from bidso_sku_id if not provided
            vertical_code = None
            model_code = None
            numeric_code = None
            
            if vertical_col is not None and row[vertical_col]:
                v_input = str(row[vertical_col]).strip()
                v = vertical_map.get(v_input.upper()) or vertical_map.get(v_input.lower())
                if v:
                    vertical_code = v.get("code")
            
            if model_col is not None and row[model_col]:
                m_input = str(row[model_col]).strip()
                m = model_map.get(m_input.upper()) or model_map.get(m_input.lower())
                if m:
                    model_code = m.get("code")
            
            if numeric_col is not None and row[numeric_col]:
                numeric_code = str(row[numeric_col]).strip().zfill(3)
            
            # Parse from bidso_sku_id if not provided (format: VERTICAL_MODEL_NUMBER)
            if not vertical_code or not model_code:
                parts = bidso_sku_id.split("_")
                if len(parts) >= 2:
                    potential_vertical = parts[0]
                    potential_model = parts[1] if len(parts) > 1 else None
                    potential_numeric = parts[2] if len(parts) > 2 else None
                    
                    if not vertical_code:
                        v = vertical_map.get(potential_vertical.upper())
                        if v:
                            vertical_code = v.get("code")
                    
                    if not model_code and potential_model:
                        m = model_map.get(potential_model.upper())
                        if m:
                            model_code = m.get("code")
                    
                    if not numeric_code and potential_numeric:
                        numeric_code = potential_numeric.zfill(3)
            
            if not vertical_code:
                results["errors"].append(f"Row {row_idx}: Cannot determine vertical for {bidso_sku_id}")
                continue
            
            if not model_code:
                results["errors"].append(f"Row {row_idx}: Cannot determine model for {bidso_sku_id}")
                continue
            
            # Get full vertical and model info
            vertical = vertical_map.get(vertical_code.upper())
            model = model_map.get(model_code.upper())
            
            if not vertical or not model:
                results["errors"].append(f"Row {row_idx}: Invalid vertical/model for {bidso_sku_id}")
                continue
            
            # Create Bidso SKU
            name = str(row[name_col]).strip() if name_col is not None and row[name_col] else f"{vertical.get('name', '')} - {model.get('name', '')}"
            description = str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else None
            
            bidso_doc = {
                "id": str(uuid.uuid4()),
                "bidso_sku_id": bidso_sku_id,
                "vertical_id": vertical.get("id"),
                "vertical_code": vertical_code,
                "model_id": model.get("id"),
                "model_code": model_code,
                "numeric_code": numeric_code or "001",
                "name": name,
                "description": description,
                "status": "ACTIVE",
                "created_at": datetime.now(timezone.utc)
            }
            
            await db.bidso_skus.insert_one(bidso_doc)
            results["created"] += 1
            
        except Exception as e:
            results["errors"].append(f"Row {row_idx}: {str(e)}")
    
    return results


@router.get("/bidso-skus/missing-from-buyer-skus")
async def get_missing_bidso_skus():
    """
    Find Bidso SKU IDs referenced by Buyer SKUs that don't exist in bidso_skus collection.
    Useful for data recovery when Bidso SKUs were accidentally deleted.
    """
    # Get all unique bidso_sku_ids referenced by buyer_skus
    referenced_ids = await db.buyer_skus.distinct("bidso_sku_id")
    referenced_ids = [b for b in referenced_ids if b]  # Filter nulls
    
    # Get existing bidso_sku_ids
    existing_ids = await db.bidso_skus.distinct("bidso_sku_id")
    
    # Find missing ones
    missing_ids = [b for b in referenced_ids if b not in existing_ids]
    
    # For each missing, try to extract vertical/model info
    missing_details = []
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(100)
    models = await db.models.find({}, {"_id": 0}).to_list(500)
    
    vertical_map = {v.get("code", "").upper(): v for v in verticals}
    model_map = {m.get("code", "").upper(): m for m in models}
    
    for bidso_id in sorted(missing_ids):
        parts = bidso_id.split("_")
        detail = {
            "bidso_sku_id": bidso_id,
            "vertical_code": parts[0] if len(parts) > 0 else None,
            "model_code": parts[1] if len(parts) > 1 else None,
            "numeric_code": parts[2] if len(parts) > 2 else None,
            "vertical_found": parts[0].upper() in vertical_map if len(parts) > 0 else False,
            "model_found": parts[1].upper() in model_map if len(parts) > 1 else False
        }
        
        # Count buyer SKUs referencing this
        buyer_count = await db.buyer_skus.count_documents({"bidso_sku_id": bidso_id})
        detail["buyer_sku_count"] = buyer_count
        
        missing_details.append(detail)
    
    return {
        "total_referenced": len(referenced_ids),
        "total_existing": len(existing_ids),
        "total_missing": len(missing_ids),
        "missing_bidso_skus": missing_details
    }


@router.post("/bidso-skus/auto-create-from-buyer-skus")
async def auto_create_missing_bidso_skus():
    """
    Automatically create missing Bidso SKUs based on references from Buyer SKUs.
    This parses the bidso_sku_id format (VERTICAL_MODEL_NUMBER) to create the records.
    """
    # Get all unique bidso_sku_ids referenced by buyer_skus
    referenced_ids = await db.buyer_skus.distinct("bidso_sku_id")
    referenced_ids = [b for b in referenced_ids if b]
    
    # Get existing bidso_sku_ids
    existing_ids = set(await db.bidso_skus.distinct("bidso_sku_id"))
    
    # Find missing ones
    missing_ids = [b for b in referenced_ids if b not in existing_ids]
    
    if not missing_ids:
        return {"message": "No missing Bidso SKUs found", "created": 0}
    
    # Load reference data
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(100)
    models = await db.models.find({}, {"_id": 0}).to_list(500)
    
    vertical_map = {v.get("code", "").upper(): v for v in verticals}
    model_map = {m.get("code", "").upper(): m for m in models}
    
    results = {
        "processed": 0,
        "created": 0,
        "errors": []
    }
    
    for bidso_id in sorted(missing_ids):
        results["processed"] += 1
        
        parts = bidso_id.split("_")
        if len(parts) < 2:
            results["errors"].append(f"{bidso_id}: Cannot parse - need at least VERTICAL_MODEL format")
            continue
        
        vertical_code = parts[0].upper()
        model_code = parts[1].upper()
        numeric_code = parts[2] if len(parts) > 2 else "001"
        
        vertical = vertical_map.get(vertical_code)
        model = model_map.get(model_code)
        
        if not vertical:
            results["errors"].append(f"{bidso_id}: Vertical '{vertical_code}' not found in system")
            continue
        
        if not model:
            results["errors"].append(f"{bidso_id}: Model '{model_code}' not found in system")
            continue
        
        # Create Bidso SKU
        bidso_doc = {
            "id": str(uuid.uuid4()),
            "bidso_sku_id": bidso_id,
            "vertical_id": vertical.get("id"),
            "vertical_code": vertical_code,
            "model_id": model.get("id"),
            "model_code": model_code,
            "numeric_code": numeric_code.zfill(3) if numeric_code.isdigit() else numeric_code,
            "name": f"{vertical.get('name', '')} - {model.get('name', '')}",
            "status": "ACTIVE",
            "created_at": datetime.now(timezone.utc)
        }
        
        try:
            await db.bidso_skus.insert_one(bidso_doc)
            results["created"] += 1
        except Exception as e:
            results["errors"].append(f"{bidso_id}: {str(e)}")
    
    return results



@router.get("/bidso-skus/next-code")
async def suggest_numeric_code(vertical_id: str, model_id: str):
    """Get next available numeric code for a vertical/model combination"""
    vertical = await db.verticals.find_one({"id": vertical_id}, {"_id": 0, "code": 1})
    model = await db.models.find_one({"id": model_id}, {"_id": 0, "code": 1})
    
    if not vertical or not model:
        raise HTTPException(status_code=404, detail="Vertical or Model not found")
    
    next_code = await get_next_numeric_code(vertical["code"], model["code"])
    
    return {
        "suggested_code": next_code,
        "preview_sku_id": f"{vertical['code']}_{model['code']}_{next_code}"
    }


# ============ SKU Bulk Import ============

@router.get("/skus/bulk-import/template")
async def download_sku_bulk_import_template():
    """Download Excel template for bulk SKU import"""
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SKUs"
    
    # Headers
    headers = ["Bidso SKU", "Buyer SKU ID", "Description", "Brand", "Vertical", "Model"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Example rows
    examples = [
        ["KS_BE_001", "FC_KS_BE_001", "Kids Scooter LED Blue", "Firstcry", "Scooter", "Blaze"],
        ["KS_BE_002", "BB_KS_BE_002", "Kids Scooter LED Green", "Blush Baby", "Scooter", "Blaze"],
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "SKU Bulk Import Instructions",
        "",
        "Required Columns:",
        "- Bidso SKU: The base product SKU code (e.g., KS_BE_001)",
        "- Buyer SKU ID: The branded variant SKU code (e.g., FC_KS_BE_001)",
        "- Description: Product description",
        "- Brand: Brand name (must exist in system)",
        "- Vertical: Product vertical (must exist in system)",
        "- Model: Product model (must exist in system)",
        "",
        "Notes:",
        "- Brand, Vertical, and Model names must match exactly what exists in the system",
        "- Duplicate Buyer SKU IDs will be skipped",
        "- Invalid references will be reported in the result file"
    ]
    for row_idx, text in enumerate(instructions, 1):
        ws2.cell(row=row_idx, column=1, value=text)
    ws2.column_dimensions['A'].width = 80
    
    # Add reference sheets
    # Brands reference
    brands = await db.brands.find({}, {"_id": 0, "name": 1, "code": 1}).to_list(100)
    ws3 = wb.create_sheet("Brands_Reference")
    ws3.cell(row=1, column=1, value="Brand Name").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Code").font = Font(bold=True)
    for idx, b in enumerate(brands, 2):
        ws3.cell(row=idx, column=1, value=b.get("name", ""))
        ws3.cell(row=idx, column=2, value=b.get("code", ""))
    
    # Verticals reference
    verticals = await db.verticals.find({}, {"_id": 0, "name": 1, "code": 1}).to_list(50)
    ws4 = wb.create_sheet("Verticals_Reference")
    ws4.cell(row=1, column=1, value="Vertical Name").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="Code").font = Font(bold=True)
    for idx, v in enumerate(verticals, 2):
        ws4.cell(row=idx, column=1, value=v.get("name", ""))
        ws4.cell(row=idx, column=2, value=v.get("code", ""))
    
    # Models reference
    models = await db.models.find({}, {"_id": 0, "name": 1, "code": 1}).to_list(200)
    ws5 = wb.create_sheet("Models_Reference")
    ws5.cell(row=1, column=1, value="Model Name").font = Font(bold=True)
    ws5.cell(row=1, column=2, value="Code").font = Font(bold=True)
    for idx, m in enumerate(models, 2):
        ws5.cell(row=idx, column=1, value=m.get("name", ""))
        ws5.cell(row=idx, column=2, value=m.get("code", ""))
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sku_bulk_import_template.xlsx"}
    )


@router.post("/skus/bulk-import")
async def bulk_import_skus(file: UploadFile = File(...)):
    """
    Bulk import SKUs from Excel file.
    
    Creates records in both `skus` collection (legacy) and `buyer_skus` collection.
    Handles brand/vertical/model resolution automatically.
    
    Returns Excel file with import results.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill
    import pandas as pd
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    
    # Normalize column names
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    
    # Map columns
    col_map = {}
    bidso_opts = ["bidso_sku", "bidso_sku_id", "bidso"]
    buyer_opts = ["buyer_sku_id", "buyer_sku", "sku_id"]
    desc_opts = ["description", "desc", "name"]
    brand_opts = ["brand", "brand_name"]
    vertical_opts = ["vertical", "vertical_name", "category"]
    model_opts = ["model", "model_name"]
    
    for opt in bidso_opts:
        if opt in df.columns:
            col_map["bidso"] = opt
            break
    for opt in buyer_opts:
        if opt in df.columns:
            col_map["buyer"] = opt
            break
    for opt in desc_opts:
        if opt in df.columns:
            col_map["desc"] = opt
            break
    for opt in brand_opts:
        if opt in df.columns:
            col_map["brand"] = opt
            break
    for opt in vertical_opts:
        if opt in df.columns:
            col_map["vertical"] = opt
            break
    for opt in model_opts:
        if opt in df.columns:
            col_map["model"] = opt
            break
    
    required = ["bidso", "buyer"]
    missing = [r for r in required if r not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}. Found: {list(df.columns)}")
    
    # Load reference data
    brands = await db.brands.find({}, {"_id": 0}).to_list(100)
    brand_map = {b["name"].lower(): b for b in brands}
    brand_code_map = {b.get("code", "").lower(): b for b in brands if b.get("code")}
    
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(50)
    vertical_map = {v["name"].lower(): v for v in verticals}
    vertical_code_map = {v.get("code", "").lower(): v for v in verticals if v.get("code")}
    
    models = await db.models.find({}, {"_id": 0}).to_list(200)
    model_map = {m["name"].lower(): m for m in models}
    model_code_map = {m.get("code", "").lower(): m for m in models if m.get("code")}
    
    # Process rows
    results = []
    created_count = 0
    skipped_count = 0
    error_count = 0
    
    for idx, row in df.iterrows():
        row_num = idx + 2
        bidso_sku = str(row[col_map["bidso"]]).strip() if pd.notna(row[col_map["bidso"]]) else ""
        buyer_sku_id = str(row[col_map["buyer"]]).strip() if pd.notna(row[col_map["buyer"]]) else ""
        description = str(row[col_map.get("desc", "")]).strip() if col_map.get("desc") and pd.notna(row.get(col_map.get("desc"))) else ""
        brand_name = str(row[col_map.get("brand", "")]).strip() if col_map.get("brand") and pd.notna(row.get(col_map.get("brand"))) else ""
        vertical_name = str(row[col_map.get("vertical", "")]).strip() if col_map.get("vertical") and pd.notna(row.get(col_map.get("vertical"))) else ""
        model_name = str(row[col_map.get("model", "")]).strip() if col_map.get("model") and pd.notna(row.get(col_map.get("model"))) else ""
        
        result = {
            "row": row_num,
            "bidso_sku": bidso_sku,
            "buyer_sku_id": buyer_sku_id,
            "description": description,
            "brand": brand_name,
            "vertical": vertical_name,
            "model": model_name,
            "status": "PENDING",
            "remarks": ""
        }
        
        # Validate required fields
        if not bidso_sku or not buyer_sku_id:
            result["status"] = "ERROR"
            result["remarks"] = "Bidso SKU and Buyer SKU ID are required"
            error_count += 1
            results.append(result)
            continue
        
        # Check if already exists in buyer_skus (new model)
        existing = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
        if existing:
            result["status"] = "SKIPPED"
            result["remarks"] = "Buyer SKU ID already exists"
            skipped_count += 1
            results.append(result)
            continue
        
        # Resolve brand
        brand = None
        brand_id = None
        if brand_name:
            brand = brand_map.get(brand_name.lower()) or brand_code_map.get(brand_name.lower())
            if brand:
                brand_id = brand.get("id")
                brand_name = brand.get("name", brand_name)
        
        # Resolve vertical
        vertical = None
        vertical_id = None
        if vertical_name:
            vertical = vertical_map.get(vertical_name.lower()) or vertical_code_map.get(vertical_name.lower())
            if vertical:
                vertical_id = vertical.get("id")
                vertical_name = vertical.get("name", vertical_name)
        
        # Resolve model
        model = None
        model_id = None
        if model_name:
            model = model_map.get(model_name.lower()) or model_code_map.get(model_name.lower())
            if model:
                model_id = model.get("id")
                model_name = model.get("name", model_name)
        
        # Insert into buyer_skus collection (new model - primary)
        try:
            buyer_sku_doc = {
                "id": str(uuid.uuid4()),
                "buyer_sku_id": buyer_sku_id,
                "bidso_sku_id": bidso_sku,
                "brand_id": brand_id,
                "brand_code": brand.get("code") if brand else None,
                "name": description,
                "description": description,
                "status": "ACTIVE",
                "gst_rate": 18,
                "hsn_code": "87141090",
                "created_at": datetime.now(timezone.utc)
            }
            await db.buyer_skus.insert_one(buyer_sku_doc)
            
            result["status"] = "CREATED"
            result["remarks"] = "Successfully imported to buyer_skus"
            created_count += 1
        except Exception as e:
            result["status"] = "ERROR"
            result["remarks"] = str(e)
            error_count += 1
        
        results.append(result)
    
    # Generate result Excel
    result_data = [{
        "Row": r["row"],
        "Bidso SKU": r["bidso_sku"],
        "Buyer SKU ID": r["buyer_sku_id"],
        "Description": r["description"],
        "Brand": r["brand"],
        "Vertical": r["vertical"],
        "Model": r["model"],
        "Status": r["status"],
        "Remarks": r["remarks"]
    } for r in results]
    
    result_df = pd.DataFrame(result_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='Import Results')
        
        ws = writer.sheets['Import Results']
        
        # Header styling
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Status color coding
        status_colors = {
            "CREATED": "C6EFCE",
            "SKIPPED": "FFEB9C",
            "ERROR": "FFC7CE"
        }
        for row_num in range(2, len(result_data) + 2):
            status_cell = ws.cell(row=row_num, column=8)
            if status_cell.value in status_colors:
                status_cell.fill = PatternFill(start_color=status_colors[status_cell.value], end_color=status_colors[status_cell.value], fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 40
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=sku_import_result.xlsx",
            "X-Import-Summary": f'{{"total": {len(df)}, "created": {created_count}, "skipped": {skipped_count}, "errors": {error_count}}}'
        }
    )


# ============ Buyer SKU CRUD ============

@router.get("/buyer-skus")
async def get_buyer_skus(
    bidso_sku_id: Optional[str] = None,
    brand_id: Optional[str] = None,
    buyer_id: Optional[str] = None,
    search: Optional[str] = None,
    include_inactive: bool = False,
    page: int = Query(1, ge=1, description="Page number (1-indexed)"),
    page_size: int = Query(50, ge=1, le=10000, description="Items per page (max 10000 for downloads)")
):
    """Get all Buyer SKUs with optional filters and pagination"""
    query = {}
    if not include_inactive:
        query["status"] = "ACTIVE"
    if bidso_sku_id:
        query["bidso_sku_id"] = bidso_sku_id
    if brand_id:
        query["brand_id"] = brand_id
    if buyer_id:
        query["buyer_id"] = buyer_id
    
    # Handle search with pagination
    if search:
        search_lower = search.lower()
        all_skus = await db.buyer_skus.find(query, {"_id": 0, "buyer_sku_id": 1, "bidso_sku_id": 1, "name": 1}).to_list(10000)
        matching_ids = [s["buyer_sku_id"] for s in all_skus if
                       search_lower in s.get("buyer_sku_id", "").lower() or
                       search_lower in s.get("bidso_sku_id", "").lower() or
                       search_lower in s.get("name", "").lower()]
        total = len(matching_ids)
        
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_ids = matching_ids[start_idx:end_idx]
        
        if paginated_ids:
            buyer_skus = await db.buyer_skus.find(
                {"buyer_sku_id": {"$in": paginated_ids}},
                {"_id": 0}
            ).to_list(page_size)
        else:
            buyer_skus = []
    else:
        total = await db.buyer_skus.count_documents(query)
        skip = (page - 1) * page_size
        buyer_skus = await db.buyer_skus.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    # Enrich with related data
    for sku in buyer_skus:
        # Get brand name
        if sku.get("brand_id"):
            b = await db.brands.find_one({"id": sku["brand_id"]}, {"_id": 0, "name": 1})
            sku["brand_name"] = b["name"] if b else None
        
        # Get buyer name
        if sku.get("buyer_id"):
            bu = await db.buyers.find_one({"id": sku["buyer_id"]}, {"_id": 0, "name": 1})
            sku["buyer_name"] = bu["name"] if bu else None
        
        # Get parent Bidso SKU info including vertical and model
        if sku.get("bidso_sku_id"):
            bidso = await db.bidso_skus.find_one(
                {"bidso_sku_id": sku["bidso_sku_id"]},
                {"_id": 0, "name": 1, "vertical_id": 1, "vertical_code": 1, "model_id": 1, "model_code": 1}
            )
            if bidso:
                sku["bidso_name"] = bidso.get("name")
                sku["vertical_id"] = bidso.get("vertical_id")
                sku["vertical_code"] = bidso.get("vertical_code")
                sku["model_id"] = bidso.get("model_id")
                sku["model_code"] = bidso.get("model_code")
                
                # Get vertical name
                if bidso.get("vertical_id"):
                    v = await db.verticals.find_one({"id": bidso["vertical_id"]}, {"_id": 0, "name": 1})
                    sku["vertical_name"] = v["name"] if v else None
                
                # Get model name
                if bidso.get("model_id"):
                    m = await db.models.find_one({"id": bidso["model_id"]}, {"_id": 0, "name": 1})
                    sku["model_name"] = m["name"] if m else None
    
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    return {
        "items": buyer_skus,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }


# ============ HSN/GST Stats (must be before dynamic routes) ============

@router.get("/buyer-skus/hsn-gst-stats")
async def get_hsn_gst_stats(current_user = Depends(get_current_user)):
    """Get statistics on HSN/GST data coverage"""
    
    total = await db.buyer_skus.count_documents({"status": "ACTIVE"})
    with_hsn = await db.buyer_skus.count_documents({
        "status": "ACTIVE",
        "hsn_code": {"$exists": True, "$ne": ""}
    })
    with_gst = await db.buyer_skus.count_documents({
        "status": "ACTIVE",
        "gst_rate": {"$exists": True, "$gt": 0}
    })
    
    return {
        "total_active_skus": total,
        "with_hsn_code": with_hsn,
        "with_gst_rate": with_gst,
        "missing_hsn": total - with_hsn,
        "coverage_percent": round((with_hsn / total) * 100, 1) if total > 0 else 0
    }


@router.get("/buyer-skus/{buyer_sku_id}")
async def get_buyer_sku(buyer_sku_id: str):
    """Get a single Buyer SKU by ID"""
    sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not sku:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    # Enrich with related data
    if sku.get("brand_id"):
        b = await db.brands.find_one({"id": sku["brand_id"]}, {"_id": 0, "name": 1})
        sku["brand_name"] = b["name"] if b else None
    
    if sku.get("buyer_id"):
        bu = await db.buyers.find_one({"id": sku["buyer_id"]}, {"_id": 0, "name": 1})
        sku["buyer_name"] = bu["name"] if bu else None
    
    # Get parent Bidso SKU
    if sku.get("bidso_sku_id"):
        bidso = await db.bidso_skus.find_one({"bidso_sku_id": sku["bidso_sku_id"]}, {"_id": 0})
        sku["bidso_sku"] = bidso
    
    # Get full BOM
    full_bom = await get_full_bom_for_buyer_sku(buyer_sku_id)
    sku["full_bom"] = full_bom
    
    return sku


@router.post("/buyer-skus")
async def create_buyer_sku(data: BuyerSKUCreate):
    """Create a new Buyer SKU"""
    # Validate Bidso SKU exists
    bidso = await db.bidso_skus.find_one({"bidso_sku_id": data.bidso_sku_id}, {"_id": 0})
    if not bidso:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    # Get brand info
    brand = await db.brands.find_one({"id": data.brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    brand_code = brand["code"]
    
    # Generate Buyer SKU ID: {BrandCode}_{BidsoSKU}
    buyer_sku_id = await generate_buyer_sku_id(brand_code, data.bidso_sku_id)
    
    # Check if already exists
    existing = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
    if existing:
        raise HTTPException(status_code=400, detail=f"Buyer SKU {buyer_sku_id} already exists")
    
    # Create Buyer SKU
    buyer_sku = {
        "id": str(uuid.uuid4()),
        "buyer_sku_id": buyer_sku_id,
        "bidso_sku_id": data.bidso_sku_id,
        "brand_id": data.brand_id,
        "brand_code": brand_code,
        "buyer_id": data.buyer_id,
        "name": data.name or f"{brand['name']} - {bidso.get('name', data.bidso_sku_id)}",
        "description": data.description,
        "mrp": data.mrp,
        "selling_price": data.selling_price,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.buyer_skus.insert_one(buyer_sku)
    del buyer_sku["_id"]
    
    return buyer_sku


@router.post("/buyer-skus/bulk-create")
async def bulk_create_buyer_skus(bidso_sku_id: str, brand_ids: List[str]):
    """Create Buyer SKUs for multiple brands from a single Bidso SKU"""
    # Validate Bidso SKU
    bidso = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
    if not bidso:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    created = []
    skipped = []
    
    for brand_id in brand_ids:
        brand = await db.brands.find_one({"id": brand_id}, {"_id": 0})
        if not brand:
            skipped.append({"brand_id": brand_id, "reason": "Brand not found"})
            continue
        
        brand_code = brand["code"]
        buyer_sku_id = await generate_buyer_sku_id(brand_code, bidso_sku_id)
        
        # Check if already exists
        existing = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
        if existing:
            skipped.append({"brand_id": brand_id, "reason": "Already exists"})
            continue
        
        # Create
        buyer_sku = {
            "id": str(uuid.uuid4()),
            "buyer_sku_id": buyer_sku_id,
            "bidso_sku_id": bidso_sku_id,
            "brand_id": brand_id,
            "brand_code": brand_code,
            "name": f"{brand['name']} - {bidso.get('name', bidso_sku_id)}",
            "description": "",
            "status": "ACTIVE",
            "created_at": datetime.now(timezone.utc)
        }
        
        await db.buyer_skus.insert_one(buyer_sku)
        created.append(buyer_sku_id)
    
    return {
        "created": len(created),
        "skipped": len(skipped),
        "created_ids": created,
        "skipped_details": skipped
    }


@router.put("/buyer-skus/{buyer_sku_id}")
async def update_buyer_sku(buyer_sku_id: str, data: BuyerSKUUpdate):
    """Update a Buyer SKU"""
    existing = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    update_data = {}
    if data.buyer_id is not None:
        update_data["buyer_id"] = data.buyer_id
    if data.name is not None:
        update_data["name"] = data.name
    if data.description is not None:
        update_data["description"] = data.description
    if data.mrp is not None:
        update_data["mrp"] = data.mrp
    if data.selling_price is not None:
        update_data["selling_price"] = data.selling_price
    if data.status is not None:
        update_data["status"] = data.status
    if data.hsn_code is not None:
        update_data["hsn_code"] = data.hsn_code
    if data.gst_rate is not None:
        update_data["gst_rate"] = data.gst_rate
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        await db.buyer_skus.update_one(
            {"buyer_sku_id": buyer_sku_id},
            {"$set": update_data}
        )
    
    return {"message": "Buyer SKU updated"}


@router.delete("/buyer-skus/{buyer_sku_id}")
async def delete_buyer_sku(buyer_sku_id: str):
    """Soft delete a Buyer SKU"""
    result = await db.buyer_skus.update_one(
        {"buyer_sku_id": buyer_sku_id},
        {"$set": {"status": "INACTIVE", "updated_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    return {"message": "Buyer SKU deleted"}


@router.post("/buyer-skus/bulk-delete/preview")
async def preview_bulk_delete_buyer_skus(file: UploadFile = File(...)):
    """
    Preview bulk delete - shows which Buyer SKUs will be deleted.
    Accepts Excel/CSV with buyer_sku_id column or plain text with one ID per line.
    """
    content = await file.read()
    skus_to_delete = []
    
    # Try to parse as Excel first
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).lower().strip().replace(' ', '_') if cell.value else '' for cell in ws[1]]
        
        # Find the buyer_sku_id column
        id_col_idx = None
        for idx, h in enumerate(headers):
            if h in ['buyer_sku_id', 'sku_id', 'sku_code', 'id', 'buyer_sku']:
                id_col_idx = idx
                break
        
        if id_col_idx is None:
            # Try first column if no header match
            id_col_idx = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[id_col_idx] if id_col_idx < len(row) else None
            if val:
                skus_to_delete.append(str(val).strip())
    except Exception:
        # Try as plain text (one ID per line)
        try:
            text = content.decode('utf-8')
            for line in text.strip().split('\n'):
                line = line.strip()
                if line and not line.startswith('#'):
                    skus_to_delete.append(line)
        except Exception:
            raise HTTPException(status_code=400, detail="Could not parse file. Use Excel with 'buyer_sku_id' column or text file with one ID per line.")
    
    if not skus_to_delete:
        raise HTTPException(status_code=400, detail="No SKU IDs found in file")
    
    # Find matching SKUs in database
    found_skus = []
    not_found = []
    
    for sku_id in skus_to_delete:
        doc = await db.buyer_skus.find_one(
            {"buyer_sku_id": sku_id, "status": {"$ne": "INACTIVE"}},
            {"_id": 0, "buyer_sku_id": 1, "name": 1, "brand_code": 1, "bidso_sku_id": 1}
        )
        if doc:
            found_skus.append(doc)
        else:
            not_found.append(sku_id)
    
    return {
        "total_in_file": len(skus_to_delete),
        "found": len(found_skus),
        "not_found": len(not_found),
        "skus_to_delete": found_skus,
        "not_found_ids": not_found[:20],  # Limit for display
        "message": f"Found {len(found_skus)} Buyer SKUs to delete. {len(not_found)} not found or already inactive."
    }


@router.post("/buyer-skus/bulk-delete/confirm")
async def confirm_bulk_delete_buyer_skus(file: UploadFile = File(...)):
    """
    Execute bulk delete - permanently deletes the Buyer SKUs.
    Same file format as preview endpoint.
    """
    content = await file.read()
    skus_to_delete = []
    
    # Parse file (same logic as preview)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).lower().strip().replace(' ', '_') if cell.value else '' for cell in ws[1]]
        
        id_col_idx = None
        for idx, h in enumerate(headers):
            if h in ['buyer_sku_id', 'sku_id', 'sku_code', 'id', 'buyer_sku']:
                id_col_idx = idx
                break
        
        if id_col_idx is None:
            id_col_idx = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[id_col_idx] if id_col_idx < len(row) else None
            if val:
                skus_to_delete.append(str(val).strip())
    except Exception:
        try:
            text = content.decode('utf-8')
            for line in text.strip().split('\n'):
                line = line.strip()
                if line and not line.startswith('#'):
                    skus_to_delete.append(line)
        except Exception:
            raise HTTPException(status_code=400, detail="Could not parse file")
    
    if not skus_to_delete:
        raise HTTPException(status_code=400, detail="No SKU IDs found in file")
    
    # Delete (hard delete, not soft delete)
    result = await db.buyer_skus.delete_many({"buyer_sku_id": {"$in": skus_to_delete}})
    
    # Get remaining count
    remaining = await db.buyer_skus.count_documents({"status": {"$ne": "INACTIVE"}})
    
    return {
        "deleted": result.deleted_count,
        "requested": len(skus_to_delete),
        "remaining_total": remaining,
        "message": f"Successfully deleted {result.deleted_count} Buyer SKUs"
    }


# ============ Common BOM Management ============

@router.get("/bom/common/{bidso_sku_id}")
async def get_common_bom(bidso_sku_id: str):
    """Get common BOM for a Bidso SKU"""
    bom = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
    if not bom:
        return {
            "bidso_sku_id": bidso_sku_id,
            "items": [],
            "is_locked": False
        }
    
    # Enrich with RM names
    for item in bom.get("items", []):
        rm = await db.raw_materials.find_one({"rm_id": item["rm_id"]}, {"_id": 0, "name": 1})
        item["rm_name"] = rm["name"] if rm else None
    
    return bom


@router.post("/bom/common")
async def create_or_update_common_bom(data: CommonBOMCreate):
    """Create or update common BOM for a Bidso SKU"""
    # Check if Bidso SKU exists
    bidso = await db.bidso_skus.find_one({"bidso_sku_id": data.bidso_sku_id})
    if not bidso:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    # Check if BOM is locked
    existing = await db.common_bom.find_one({"bidso_sku_id": data.bidso_sku_id})
    if existing and existing.get("is_locked"):
        raise HTTPException(status_code=400, detail="BOM is locked and cannot be modified")
    
    # Validate all RM IDs
    items_data = []
    for item in data.items:
        rm = await db.raw_materials.find_one({"rm_id": item.rm_id}, {"_id": 0, "name": 1})
        if not rm:
            raise HTTPException(status_code=404, detail=f"Raw Material {item.rm_id} not found")
        items_data.append({
            "rm_id": item.rm_id,
            "rm_name": rm["name"],
            "quantity": item.quantity,
            "unit": item.unit
        })
    
    if existing:
        # Update existing
        await db.common_bom.update_one(
            {"bidso_sku_id": data.bidso_sku_id},
            {"$set": {
                "items": items_data,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
    else:
        # Create new
        bom = {
            "id": str(uuid.uuid4()),
            "bidso_sku_id": data.bidso_sku_id,
            "items": items_data,
            "is_locked": False,
            "created_at": datetime.now(timezone.utc)
        }
        await db.common_bom.insert_one(bom)
    
    return {"message": "Common BOM saved"}


@router.post("/bom/common/{bidso_sku_id}/lock")
async def lock_common_bom(bidso_sku_id: str):
    """Lock the common BOM for a Bidso SKU"""
    result = await db.common_bom.update_one(
        {"bidso_sku_id": bidso_sku_id},
        {"$set": {
            "is_locked": True,
            "locked_at": datetime.now(timezone.utc)
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Common BOM not found")
    
    return {"message": "BOM locked successfully"}


@router.post("/bom/common/{bidso_sku_id}/unlock")
async def unlock_common_bom(bidso_sku_id: str):
    """Unlock the common BOM for a Bidso SKU (requires admin)"""
    result = await db.common_bom.update_one(
        {"bidso_sku_id": bidso_sku_id},
        {"$set": {
            "is_locked": False,
            "locked_at": None
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Common BOM not found")
    
    return {"message": "BOM unlocked successfully"}


# ============ Brand-specific BOM Management ============

@router.get("/bom/brand-specific/{bidso_sku_id}/{brand_id}")
async def get_brand_specific_bom(bidso_sku_id: str, brand_id: str):
    """Get brand-specific BOM additions"""
    bom = await db.brand_specific_bom.find_one(
        {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
        {"_id": 0}
    )
    
    if not bom:
        brand = await db.brands.find_one({"id": brand_id}, {"_id": 0, "code": 1})
        return {
            "bidso_sku_id": bidso_sku_id,
            "brand_id": brand_id,
            "brand_code": brand["code"] if brand else None,
            "items": []
        }
    
    # Enrich with RM names
    for item in bom.get("items", []):
        rm = await db.raw_materials.find_one({"rm_id": item["rm_id"]}, {"_id": 0, "name": 1})
        item["rm_name"] = rm["name"] if rm else None
    
    return bom


@router.post("/bom/brand-specific")
async def create_or_update_brand_specific_bom(data: BrandSpecificBOMCreate):
    """Create or update brand-specific BOM additions"""
    # Validate Bidso SKU
    bidso = await db.bidso_skus.find_one({"bidso_sku_id": data.bidso_sku_id})
    if not bidso:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    # Validate Brand
    brand = await db.brands.find_one({"id": data.brand_id}, {"_id": 0, "code": 1})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    # Validate all RM IDs
    items_data = []
    for item in data.items:
        rm = await db.raw_materials.find_one({"rm_id": item.rm_id}, {"_id": 0, "name": 1})
        if not rm:
            raise HTTPException(status_code=404, detail=f"Raw Material {item.rm_id} not found")
        items_data.append({
            "rm_id": item.rm_id,
            "rm_name": rm["name"],
            "quantity": item.quantity,
            "unit": item.unit
        })
    
    existing = await db.brand_specific_bom.find_one({
        "bidso_sku_id": data.bidso_sku_id,
        "brand_id": data.brand_id
    })
    
    if existing:
        await db.brand_specific_bom.update_one(
            {"bidso_sku_id": data.bidso_sku_id, "brand_id": data.brand_id},
            {"$set": {
                "items": items_data,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
    else:
        bom = {
            "id": str(uuid.uuid4()),
            "bidso_sku_id": data.bidso_sku_id,
            "brand_id": data.brand_id,
            "brand_code": brand["code"],
            "items": items_data,
            "created_at": datetime.now(timezone.utc)
        }
        await db.brand_specific_bom.insert_one(bom)
    
    return {"message": "Brand-specific BOM saved"}


@router.delete("/bom/brand-specific/{bidso_sku_id}/{brand_id}")
async def delete_brand_specific_bom(bidso_sku_id: str, brand_id: str):
    """Delete brand-specific BOM additions"""
    result = await db.brand_specific_bom.delete_one({
        "bidso_sku_id": bidso_sku_id,
        "brand_id": brand_id
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Brand-specific BOM not found")
    
    return {"message": "Brand-specific BOM deleted"}


# ============ Bulk BOM Upload ============

@router.get("/bom/bulk-upload/template")
async def download_bom_template():
    """Download Excel template for bulk BOM upload"""
    import openpyxl
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    
    # Instructions sheet
    ws_info = wb.active
    ws_info.title = "Instructions"
    instructions = [
        ["BOM Bulk Upload Template"],
        [""],
        ["IMPORTANT: Read before uploading"],
        [""],
        ["For Buyer SKU BOM Upload:"],
        ["1. Use the 'BuyerSKU_BOM' sheet"],
        ["2. BUYER_SKU_ID is required - system will find linked Bidso SKU automatically"],
        ["3. RM_ID and QUANTITY are required for each row"],
        ["4. BRAND_SPECIFIC: Set to 'Y' for items specific to the brand (labels, packaging)"],
        ["5. Items marked as NOT brand-specific will go to the Common BOM of the linked Bidso SKU"],
        [""],
        ["For Bidso SKU BOM Upload:"],
        ["1. Use the 'BidsoSKU_BOM' sheet"],
        ["2. BIDSO_SKU_ID, RM_ID, QUANTITY are required"],
        ["3. All items will be added to the Common BOM"],
        [""],
        ["Column Descriptions:"],
        ["- BUYER_SKU_ID / BIDSO_SKU_ID: The SKU ID (must exist in system)"],
        ["- RM_ID: Raw Material ID (must exist in system)"],
        ["- QUANTITY: Quantity required per unit"],
        ["- UNIT: Unit of measure (default: PCS)"],
        ["- BRAND_SPECIFIC: Y/N - whether item is brand-specific (for Buyer SKU only)"],
    ]
    for r_idx, row in enumerate(instructions, 1):
        for c_idx, val in enumerate(row, 1):
            ws_info.cell(row=r_idx, column=c_idx, value=val)
    
    # Buyer SKU BOM sheet
    ws_buyer = wb.create_sheet("BuyerSKU_BOM")
    headers_buyer = ["BUYER_SKU_ID", "RM_ID", "QUANTITY", "UNIT", "BRAND_SPECIFIC"]
    for c_idx, header in enumerate(headers_buyer, 1):
        ws_buyer.cell(row=1, column=c_idx, value=header)
    # Sample row
    ws_buyer.cell(row=2, column=1, value="BAYBEE_KS_PE_001")
    ws_buyer.cell(row=2, column=2, value="INP_001")
    ws_buyer.cell(row=2, column=3, value=2)
    ws_buyer.cell(row=2, column=4, value="PCS")
    ws_buyer.cell(row=2, column=5, value="N")  # Common BOM
    ws_buyer.cell(row=3, column=1, value="BAYBEE_KS_PE_001")
    ws_buyer.cell(row=3, column=2, value="LBL_001")
    ws_buyer.cell(row=3, column=3, value=1)
    ws_buyer.cell(row=3, column=4, value="PCS")
    ws_buyer.cell(row=3, column=5, value="Y")  # Brand-specific
    
    # Bidso SKU BOM sheet
    ws_bidso = wb.create_sheet("BidsoSKU_BOM")
    headers_bidso = ["BIDSO_SKU_ID", "RM_ID", "QUANTITY", "UNIT"]
    for c_idx, header in enumerate(headers_bidso, 1):
        ws_bidso.cell(row=1, column=c_idx, value=header)
    # Sample row
    ws_bidso.cell(row=2, column=1, value="KS_PE_001")
    ws_bidso.cell(row=2, column=2, value="INP_001")
    ws_bidso.cell(row=2, column=3, value=2)
    ws_bidso.cell(row=2, column=4, value="PCS")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=BOM_Upload_Template.xlsx"}
    )


@router.post("/bom/bulk-upload")
async def bulk_upload_bom(
    file: UploadFile = File(...),
    mode: str = Query(default="merge", description="'merge' to add/update items, 'overwrite' to replace entire BOM")
):
    """
    Bulk upload BOM data from Excel.
    
    For Buyer SKU BOM: Items marked as NOT brand-specific will automatically
    populate the Common BOM of the linked Bidso SKU.
    
    Only requires: BUYER_SKU_ID, RM_ID, QUANTITY
    RM names are NOT required - only RM ID is used for mapping.
    
    Modes:
    - 'merge' (default): Add new items and update existing quantities. Existing items not in upload are kept.
    - 'overwrite': Replace the entire BOM with uploaded items. Existing items not in upload are removed.
    """
    import openpyxl
    import traceback
    
    # Validate mode
    if mode not in ["merge", "overwrite"]:
        raise HTTPException(status_code=400, detail="Mode must be 'merge' or 'overwrite'")
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    results = {
        "buyer_sku_processed": 0,
        "bidso_sku_processed": 0,
        "common_bom_updated": 0,
        "brand_bom_updated": 0,
        "mode": mode,
        "errors": [],
        "warnings": [],
        "success": True
    }
    
    try:
        # Process BuyerSKU_BOM sheet
        if "BuyerSKU_BOM" in wb.sheetnames:
            ws = wb["BuyerSKU_BOM"]
            
            # First, collect all unique buyer_sku_ids and rm_ids for validation
            all_buyer_skus = set()
            all_rm_ids = set()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    all_buyer_skus.add(str(row[0]).strip())
                if row[1]:
                    all_rm_ids.add(str(row[1]).strip())
            
            # Pre-validate buyer SKUs
            existing_buyer_skus = {}
            for sku_id in all_buyer_skus:
                buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": sku_id}, {"_id": 0})
                if buyer_sku:
                    existing_buyer_skus[sku_id] = buyer_sku
            
            missing_buyer_skus = all_buyer_skus - set(existing_buyer_skus.keys())
            if missing_buyer_skus:
                results["errors"].append(f"Missing Buyer SKUs: {', '.join(sorted(missing_buyer_skus)[:10])}" + 
                                        (f" (and {len(missing_buyer_skus) - 10} more)" if len(missing_buyer_skus) > 10 else ""))
            
            # Pre-validate RMs - only check if rm_id exists, don't need name
            existing_rm_ids_set = set()
            for rm_id in all_rm_ids:
                # Exact match first
                rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0, "rm_id": 1})
                if not rm:
                    # Case-insensitive match
                    escaped_rm_id = re.escape(rm_id)
                    rm = await db.raw_materials.find_one(
                        {"rm_id": {"$regex": f"^{escaped_rm_id}$", "$options": "i"}},
                        {"_id": 0, "rm_id": 1}
                    )
                if rm:
                    existing_rm_ids_set.add(rm_id)
                    existing_rm_ids_set.add(rm_id.lower())
                    existing_rm_ids_set.add(rm.get("rm_id", rm_id))  # Add actual DB rm_id
            
            missing_rms = [rm for rm in all_rm_ids if rm not in existing_rm_ids_set and rm.lower() not in existing_rm_ids_set]
            if missing_rms:
                results["errors"].append(f"Missing RM IDs: {', '.join(sorted(missing_rms)[:10])}" + 
                                        (f" (and {len(missing_rms) - 10} more)" if len(missing_rms) > 10 else ""))
            
            # If critical validations fail, return early with detailed errors
            if missing_buyer_skus or missing_rms:
                results["success"] = False
                return results
            
            # Group items by buyer_sku_id
            buyer_sku_boms = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                buyer_sku_id = str(row[0]).strip()
                rm_id = str(row[1]).strip() if row[1] else None
                quantity = float(row[2]) if row[2] else 1.0
                unit = str(row[3]).strip() if row[3] else "PCS"
                brand_specific = str(row[4]).strip().upper() if row[4] else "N"
                
                if not rm_id:
                    results["errors"].append(f"Row {row_idx}: RM_ID is required")
                    continue
                
                if buyer_sku_id not in buyer_sku_boms:
                    buyer_sku_boms[buyer_sku_id] = {"common": [], "brand_specific": []}
                
                # Store only rm_id, quantity, unit - NO name needed
                item = {"rm_id": rm_id, "quantity": quantity, "unit": unit}
                
                if brand_specific == "Y":
                    buyer_sku_boms[buyer_sku_id]["brand_specific"].append(item)
                else:
                    buyer_sku_boms[buyer_sku_id]["common"].append(item)
            
            # Process each buyer SKU
            for buyer_sku_id, bom_data in buyer_sku_boms.items():
                # Find buyer SKU and linked Bidso SKU
                buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
                if not buyer_sku:
                    results["errors"].append(f"Buyer SKU {buyer_sku_id} not found")
                    continue
                
                # Safely get required fields with validation
                bidso_sku_id = buyer_sku.get("bidso_sku_id")
                brand_id = buyer_sku.get("brand_id")
                brand_code = buyer_sku.get("brand_code", "")
                
                if not bidso_sku_id:
                    results["errors"].append(f"Buyer SKU {buyer_sku_id} has no linked Bidso SKU")
                    continue
                
                if not brand_id:
                    results["errors"].append(f"Buyer SKU {buyer_sku_id} has no brand_id")
                    continue
                
                # Build BOM items - only rm_id, quantity, unit (no name lookup needed)
                common_items = []
                for item in bom_data["common"]:
                    # Verify RM exists (already validated above, but double-check)
                    rm = await db.raw_materials.find_one({"rm_id": item["rm_id"]}, {"_id": 0, "rm_id": 1})
                    if not rm:
                        escaped_rm_id = re.escape(item["rm_id"])
                        rm = await db.raw_materials.find_one(
                            {"rm_id": {"$regex": f"^{escaped_rm_id}$", "$options": "i"}},
                            {"_id": 0, "rm_id": 1}
                        )
                    if not rm:
                        continue  # Skip - already reported in validation
                    common_items.append({
                        "rm_id": rm.get("rm_id", item["rm_id"]),
                        "quantity": item["quantity"],
                        "unit": item["unit"]
                    })
                
                brand_items = []
                for item in bom_data["brand_specific"]:
                    rm = await db.raw_materials.find_one({"rm_id": item["rm_id"]}, {"_id": 0, "rm_id": 1})
                    if not rm:
                        escaped_rm_id = re.escape(item["rm_id"])
                        rm = await db.raw_materials.find_one(
                            {"rm_id": {"$regex": f"^{escaped_rm_id}$", "$options": "i"}},
                            {"_id": 0, "rm_id": 1}
                        )
                    if not rm:
                        continue
                    brand_items.append({
                        "rm_id": rm.get("rm_id", item["rm_id"]),
                        "quantity": item["quantity"],
                        "unit": item["unit"]
                    })
                
                # Update Common BOM (for the linked Bidso SKU)
                if common_items:
                    existing_common = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id})
                    if existing_common and existing_common.get("is_locked"):
                        results["warnings"].append(f"Common BOM for {bidso_sku_id} is locked - skipped common items")
                    else:
                        if existing_common:
                            if mode == "overwrite":
                                # Overwrite mode: Replace entire BOM with uploaded items
                                await db.common_bom.update_one(
                                    {"bidso_sku_id": bidso_sku_id},
                                    {"$set": {"items": common_items, "updated_at": datetime.now(timezone.utc)}}
                                )
                            else:
                                # Merge mode: Add new RMs, update existing quantities
                                existing_rm_ids = {item["rm_id"] for item in existing_common.get("items", [])}
                                merged_items = list(existing_common.get("items", []))
                                for item in common_items:
                                    if item["rm_id"] not in existing_rm_ids:
                                        merged_items.append(item)
                                    else:
                                        # Update quantity
                                        for i, ei in enumerate(merged_items):
                                            if ei["rm_id"] == item["rm_id"]:
                                                merged_items[i] = item
                                                break
                                
                                await db.common_bom.update_one(
                                    {"bidso_sku_id": bidso_sku_id},
                                    {"$set": {"items": merged_items, "updated_at": datetime.now(timezone.utc)}}
                                )
                        else:
                            await db.common_bom.insert_one({
                                "id": str(uuid.uuid4()),
                                "bidso_sku_id": bidso_sku_id,
                                "items": common_items,
                                "is_locked": False,
                                "created_at": datetime.now(timezone.utc)
                            })
                        results["common_bom_updated"] += 1
                
                # Update Brand-specific BOM
                if brand_items:
                    existing_brand = await db.brand_specific_bom.find_one({
                        "bidso_sku_id": bidso_sku_id,
                        "brand_id": brand_id
                    })
                    
                    if existing_brand:
                        if mode == "overwrite":
                            # Overwrite mode: Replace entire brand BOM with uploaded items
                            await db.brand_specific_bom.update_one(
                                {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
                                {"$set": {"items": brand_items, "updated_at": datetime.now(timezone.utc)}}
                            )
                        else:
                            # Merge mode: Add new RMs, update existing quantities
                            existing_rm_ids = {item["rm_id"] for item in existing_brand.get("items", [])}
                            merged_items = list(existing_brand.get("items", []))
                            for item in brand_items:
                                if item["rm_id"] not in existing_rm_ids:
                                    merged_items.append(item)
                                else:
                                    for i, ei in enumerate(merged_items):
                                        if ei["rm_id"] == item["rm_id"]:
                                            merged_items[i] = item
                                            break
                            
                            await db.brand_specific_bom.update_one(
                                {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
                                {"$set": {"items": merged_items, "updated_at": datetime.now(timezone.utc)}}
                            )
                    else:
                        await db.brand_specific_bom.insert_one({
                            "id": str(uuid.uuid4()),
                            "bidso_sku_id": bidso_sku_id,
                            "brand_id": brand_id,
                            "brand_code": brand_code,
                            "items": brand_items,
                            "created_at": datetime.now(timezone.utc)
                        })
                    results["brand_bom_updated"] += 1
                
                results["buyer_sku_processed"] += 1
        
        # Process BidsoSKU_BOM sheet (direct common BOM)
        if "BidsoSKU_BOM" in wb.sheetnames:
            ws = wb["BidsoSKU_BOM"]
            
            # Group items by bidso_sku_id
            bidso_sku_boms = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                bidso_sku_id = str(row[0]).strip()
                rm_id = str(row[1]).strip() if row[1] else None
                quantity = float(row[2]) if row[2] else 1.0
                unit = str(row[3]).strip() if row[3] else "PCS"
                
                if not rm_id:
                    results["errors"].append(f"BidsoSKU_BOM Row {row_idx}: RM_ID is required")
                    continue
                
                if bidso_sku_id not in bidso_sku_boms:
                    bidso_sku_boms[bidso_sku_id] = []
                
                bidso_sku_boms[bidso_sku_id].append({
                    "rm_id": rm_id,
                    "quantity": quantity,
                    "unit": unit
                })
            
            # Process each Bidso SKU
            for bidso_sku_id, items in bidso_sku_boms.items():
                # Verify Bidso SKU exists
                bidso = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id})
                if not bidso:
                    results["errors"].append(f"Bidso SKU {bidso_sku_id} not found")
                    continue
                
                # Check if locked
                existing = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id})
                if existing and existing.get("is_locked"):
                    results["warnings"].append(f"Common BOM for {bidso_sku_id} is locked - skipped")
                    continue
                
                # Validate RMs - only check existence, no name needed
                valid_items = []
                for item in items:
                    rm = await db.raw_materials.find_one({"rm_id": item["rm_id"]}, {"_id": 0, "rm_id": 1})
                    if not rm:
                        escaped_rm_id = re.escape(item["rm_id"])
                        rm = await db.raw_materials.find_one(
                            {"rm_id": {"$regex": f"^{escaped_rm_id}$", "$options": "i"}},
                            {"_id": 0, "rm_id": 1}
                        )
                    if not rm:
                        results["errors"].append(f"RM {item['rm_id']} not found (for {bidso_sku_id})")
                        continue
                    valid_items.append({
                        "rm_id": rm.get("rm_id", item["rm_id"]),
                        "quantity": item["quantity"],
                        "unit": item["unit"]
                    })
                
                if not valid_items:
                    continue
                
                # Update or create Common BOM
                if existing:
                    if mode == "overwrite":
                        # Overwrite mode: Replace entire BOM
                        await db.common_bom.update_one(
                            {"bidso_sku_id": bidso_sku_id},
                            {"$set": {"items": valid_items, "updated_at": datetime.now(timezone.utc)}}
                        )
                    else:
                        # Merge mode
                        existing_rm_ids = {item["rm_id"] for item in existing.get("items", [])}
                        merged_items = list(existing.get("items", []))
                        for item in valid_items:
                            if item["rm_id"] not in existing_rm_ids:
                                merged_items.append(item)
                            else:
                                for i, ei in enumerate(merged_items):
                                    if ei["rm_id"] == item["rm_id"]:
                                        merged_items[i] = item
                                        break
                        
                        await db.common_bom.update_one(
                            {"bidso_sku_id": bidso_sku_id},
                            {"$set": {"items": merged_items, "updated_at": datetime.now(timezone.utc)}}
                        )
                else:
                    await db.common_bom.insert_one({
                        "id": str(uuid.uuid4()),
                        "bidso_sku_id": bidso_sku_id,
                        "items": valid_items,
                        "is_locked": False,
                        "created_at": datetime.now(timezone.utc)
                    })
                
                results["bidso_sku_processed"] += 1
                results["common_bom_updated"] += 1
        
        # Mark as failed if there are errors and no successful processing
        if results["errors"] and results["buyer_sku_processed"] == 0 and results["bidso_sku_processed"] == 0:
            results["success"] = False
        
        return results
    
    except Exception as e:
        # Log the full traceback for debugging
        error_trace = traceback.format_exc()
        print(f"BOM Upload Error: {str(e)}\n{error_trace}")
        results["errors"].append(f"Server error: {str(e)}")
        results["success"] = False
        return results


@router.get("/bom/export")
async def export_all_bom():
    """Export all BOM data to Excel"""
    import openpyxl
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    
    # Common BOM sheet
    ws_common = wb.active
    ws_common.title = "Common_BOM"
    headers = ["BIDSO_SKU_ID", "BIDSO_SKU_NAME", "RM_ID", "RM_NAME", "QUANTITY", "UNIT"]
    for c_idx, h in enumerate(headers, 1):
        ws_common.cell(row=1, column=c_idx, value=h)
    
    row_idx = 2
    async for bom in db.common_bom.find({}, {"_id": 0}):
        bidso = await db.bidso_skus.find_one({"bidso_sku_id": bom["bidso_sku_id"]}, {"_id": 0, "name": 1})
        bidso_name = bidso["name"] if bidso else ""
        
        for item in bom.get("items", []):
            ws_common.cell(row=row_idx, column=1, value=bom["bidso_sku_id"])
            ws_common.cell(row=row_idx, column=2, value=bidso_name)
            ws_common.cell(row=row_idx, column=3, value=item.get("rm_id", ""))
            ws_common.cell(row=row_idx, column=4, value=item.get("rm_name", ""))
            ws_common.cell(row=row_idx, column=5, value=item.get("quantity", 1))
            ws_common.cell(row=row_idx, column=6, value=item.get("unit", "PCS"))
            row_idx += 1
    
    # Brand-specific BOM sheet
    ws_brand = wb.create_sheet("Brand_Specific_BOM")
    headers_brand = ["BIDSO_SKU_ID", "BRAND_CODE", "RM_ID", "RM_NAME", "QUANTITY", "UNIT"]
    for c_idx, h in enumerate(headers_brand, 1):
        ws_brand.cell(row=1, column=c_idx, value=h)
    
    row_idx = 2
    async for bom in db.brand_specific_bom.find({}, {"_id": 0}):
        for item in bom.get("items", []):
            ws_brand.cell(row=row_idx, column=1, value=bom.get("bidso_sku_id", ""))
            ws_brand.cell(row=row_idx, column=2, value=bom.get("brand_code", ""))
            ws_brand.cell(row=row_idx, column=3, value=item.get("rm_id", ""))
            ws_brand.cell(row=row_idx, column=4, value=item.get("rm_name", ""))
            ws_brand.cell(row=row_idx, column=5, value=item.get("quantity", 1))
            ws_brand.cell(row=row_idx, column=6, value=item.get("unit", "PCS"))
            row_idx += 1
    
    # Full Buyer SKU BOM sheet (combined view)
    ws_full = wb.create_sheet("Full_BuyerSKU_BOM")
    headers_full = ["BUYER_SKU_ID", "BIDSO_SKU_ID", "BRAND_CODE", "RM_ID", "RM_NAME", "QUANTITY", "UNIT", "BOM_TYPE"]
    for c_idx, h in enumerate(headers_full, 1):
        ws_full.cell(row=1, column=c_idx, value=h)
    
    row_idx = 2
    async for buyer_sku in db.buyer_skus.find({"status": "ACTIVE"}, {"_id": 0}):
        buyer_sku_id = buyer_sku["buyer_sku_id"]
        bidso_sku_id = buyer_sku["bidso_sku_id"]
        brand_code = buyer_sku.get("brand_code", "")
        brand_id = buyer_sku.get("brand_id", "")
        
        # Get common BOM
        common_bom = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
        if common_bom:
            for item in common_bom.get("items", []):
                ws_full.cell(row=row_idx, column=1, value=buyer_sku_id)
                ws_full.cell(row=row_idx, column=2, value=bidso_sku_id)
                ws_full.cell(row=row_idx, column=3, value=brand_code)
                ws_full.cell(row=row_idx, column=4, value=item.get("rm_id", ""))
                ws_full.cell(row=row_idx, column=5, value=item.get("rm_name", ""))
                ws_full.cell(row=row_idx, column=6, value=item.get("quantity", 1))
                ws_full.cell(row=row_idx, column=7, value=item.get("unit", "PCS"))
                ws_full.cell(row=row_idx, column=8, value="Common")
                row_idx += 1
        
        # Get brand-specific BOM
        brand_bom = await db.brand_specific_bom.find_one(
            {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
            {"_id": 0}
        )
        if brand_bom:
            for item in brand_bom.get("items", []):
                ws_full.cell(row=row_idx, column=1, value=buyer_sku_id)
                ws_full.cell(row=row_idx, column=2, value=bidso_sku_id)
                ws_full.cell(row=row_idx, column=3, value=brand_code)
                ws_full.cell(row=row_idx, column=4, value=item.get("rm_id", ""))
                ws_full.cell(row=row_idx, column=5, value=item.get("rm_name", ""))
                ws_full.cell(row=row_idx, column=6, value=item.get("quantity", 1))
                ws_full.cell(row=row_idx, column=7, value=item.get("unit", "PCS"))
                ws_full.cell(row=row_idx, column=8, value="Brand-Specific")
                row_idx += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=BOM_Export.xlsx"}
    )


# ============ Full BOM View ============

async def get_full_bom_for_buyer_sku(buyer_sku_id: str) -> dict:
    """Get combined BOM (common + brand-specific) for a Buyer SKU"""
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        return None
    
    bidso_sku_id = buyer_sku["bidso_sku_id"]
    brand_id = buyer_sku["brand_id"]
    brand_code = buyer_sku["brand_code"]
    
    # Get common BOM
    common_bom = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
    common_items_raw = common_bom.get("items", []) if common_bom else []
    
    # Get brand-specific BOM
    brand_bom = await db.brand_specific_bom.find_one(
        {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
        {"_id": 0}
    )
    brand_items_raw = brand_bom.get("items", []) if brand_bom else []
    
    # Collect all RM IDs to fetch descriptions
    all_rm_ids = set()
    for item in common_items_raw + brand_items_raw:
        if item.get("rm_id"):
            all_rm_ids.add(item["rm_id"])
    
    # Fetch RM details from raw_materials collection
    rm_details = {}
    if all_rm_ids:
        rm_cursor = db.raw_materials.find(
            {"rm_id": {"$in": list(all_rm_ids)}},
            {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1}
        )
        async for rm in rm_cursor:
            rm_id = rm.get("rm_id")
            category = rm.get("category", "")
            cat_data = rm.get("category_data", {}) or {}
            
            # Generate description based on category-specific format
            description = generate_rm_description(category, cat_data, rm.get("name", ""))
            rm_details[rm_id] = description
    
    # Enrich common items with descriptions
    common_items = []
    for item in common_items_raw:
        rm_id = item.get("rm_id", "")
        common_items.append({
            "rm_id": rm_id,
            "rm_name": rm_details.get(rm_id, item.get("rm_name", "")),
            "quantity": item.get("quantity", 1),
            "unit": item.get("unit", "PCS")
        })
    
    # Enrich brand-specific items with descriptions
    brand_items = []
    for item in brand_items_raw:
        rm_id = item.get("rm_id", "")
        brand_items.append({
            "rm_id": rm_id,
            "rm_name": rm_details.get(rm_id, item.get("rm_name", "")),
            "quantity": item.get("quantity", 1),
            "unit": item.get("unit", "PCS")
        })
    
    # Combine
    total_items = common_items + brand_items
    
    return {
        "buyer_sku_id": buyer_sku_id,
        "bidso_sku_id": bidso_sku_id,
        "brand_code": brand_code,
        "common_items": common_items,
        "brand_specific_items": brand_items,
        "total_items": total_items,
        "is_common_bom_locked": common_bom.get("is_locked", False) if common_bom else False
    }


@router.get("/bom/full/{buyer_sku_id}")
async def get_full_bom(buyer_sku_id: str):
    """Get full BOM for a Buyer SKU (common + brand-specific)"""
    full_bom = await get_full_bom_for_buyer_sku(buyer_sku_id)
    if not full_bom:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    return full_bom


# ============ Brand-Specific BOM Edit with Approval Workflow ============

class BOMItemEdit(BaseModel):
    rm_id: str
    new_rm_id: Optional[str] = None  # If changing the RM
    quantity: float
    unit: str

class BOMChangeRequest(BaseModel):
    buyer_sku_id: str
    original_item: dict
    new_item: dict
    change_type: str  # "MODIFY", "ADD", "REMOVE"
    reason: Optional[str] = None


async def check_production_schedule_next_10_days(buyer_sku_id: str) -> dict:
    """Check if buyer SKU is scheduled for production in the next 10 days"""
    from datetime import timedelta
    
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    ten_days_later = today + timedelta(days=10)
    
    # Get the buyer SKU to find the bidso_sku_id
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        return {"is_scheduled": False, "schedules": [], "branches": []}
    
    # Check production_schedules for this SKU
    schedules = await db.production_schedules.find({
        "$or": [
            {"buyer_sku_id": buyer_sku_id},
            {"bidso_sku_id": buyer_sku.get("bidso_sku_id")},
            {"sku_id": buyer_sku_id}
        ],
        "target_date": {"$gte": today, "$lte": ten_days_later},
        "status": {"$nin": ["COMPLETED", "CANCELLED"]}
    }, {"_id": 0, "id": 1, "branch_id": 1, "branch_name": 1, "target_date": 1, "quantity": 1}).to_list(100)
    
    # Also check dispatch_lots
    dispatch_lots = await db.dispatch_lots.find({
        "buyer_sku_id": buyer_sku_id,
        "scheduled_date": {"$gte": today, "$lte": ten_days_later},
        "status": {"$nin": ["DISPATCHED", "CANCELLED"]}
    }, {"_id": 0, "id": 1, "branch_id": 1, "branch_name": 1, "scheduled_date": 1}).to_list(100)
    
    all_schedules = schedules + dispatch_lots
    branches = list(set([s.get("branch_name") or s.get("branch_id") for s in all_schedules if s.get("branch_name") or s.get("branch_id")]))
    
    return {
        "is_scheduled": len(all_schedules) > 0,
        "schedules": all_schedules,
        "branches": branches,
        "schedule_count": len(all_schedules)
    }


async def create_bom_change_notification(
    change_request_id: str,
    buyer_sku_id: str,
    change_type: str,
    original_item: dict,
    new_item: dict,
    branches: list,
    requires_approval: bool
):
    """Create notifications for BOM change"""
    notification_base = {
        "id": str(uuid.uuid4()),
        "change_request_id": change_request_id,
        "buyer_sku_id": buyer_sku_id,
        "change_type": change_type,
        "original_rm_id": original_item.get("rm_id"),
        "new_rm_id": new_item.get("rm_id") if new_item else None,
        "message": f"BOM change for {buyer_sku_id}: {change_type} - RM {original_item.get('rm_id')}",
        "requires_approval": requires_approval,
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    notifications = []
    
    # Notify Master Admin
    notifications.append({
        **notification_base,
        "id": str(uuid.uuid4()),
        "recipient_role": "master_admin",
        "notification_type": "BOM_CHANGE_APPROVAL" if requires_approval else "BOM_CHANGE_INFO"
    })
    
    # Notify CPC Planner
    notifications.append({
        **notification_base,
        "id": str(uuid.uuid4()),
        "recipient_role": "cpc_planner",
        "notification_type": "BOM_CHANGE_INFO"
    })
    
    # Notify relevant Branch Ops
    for branch in branches:
        notifications.append({
            **notification_base,
            "id": str(uuid.uuid4()),
            "recipient_role": "branch_ops_user",
            "recipient_branch": branch,
            "notification_type": "BOM_CHANGE_INFO"
        })
    
    if notifications:
        await db.notifications.insert_many(notifications)
    
    return notifications


@router.get("/bom/buyer-sku/{buyer_sku_id}/check-schedule")
async def check_sku_production_schedule(buyer_sku_id: str):
    """Check if buyer SKU is scheduled for production in the next 10 days"""
    result = await check_production_schedule_next_10_days(buyer_sku_id)
    return result


@router.put("/bom/buyer-sku/{buyer_sku_id}/item")
async def edit_brand_specific_bom_item(
    buyer_sku_id: str,
    edit_data: BOMItemEdit,
    current_user: dict = Depends(get_current_user)
):
    """
    Edit a brand-specific BOM item.
    If SKU is scheduled for production in next 10 days, requires Master Admin approval.
    """
    # Get buyer SKU
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    bidso_sku_id = buyer_sku["bidso_sku_id"]
    brand_id = buyer_sku["brand_id"]
    
    # Get brand-specific BOM
    brand_bom = await db.brand_specific_bom.find_one(
        {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
        {"_id": 0}
    )
    if not brand_bom:
        raise HTTPException(status_code=404, detail="Brand-specific BOM not found")
    
    # Find the item to edit
    items = brand_bom.get("items", [])
    item_index = next((i for i, item in enumerate(items) if item.get("rm_id") == edit_data.rm_id), None)
    if item_index is None:
        raise HTTPException(status_code=404, detail=f"RM {edit_data.rm_id} not found in brand-specific BOM")
    
    original_item = items[item_index].copy()
    
    # Check production schedule
    schedule_check = await check_production_schedule_next_10_days(buyer_sku_id)
    requires_approval = schedule_check["is_scheduled"]
    
    # Validate new RM if changing
    new_rm_id = edit_data.new_rm_id or edit_data.rm_id
    rm = await db.raw_materials.find_one({"rm_id": new_rm_id}, {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1})
    if not rm:
        # Try case-insensitive
        escaped_rm_id = re.escape(new_rm_id)
        rm = await db.raw_materials.find_one(
            {"rm_id": {"$regex": f"^{escaped_rm_id}$", "$options": "i"}},
            {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1}
        )
    if not rm:
        raise HTTPException(status_code=404, detail=f"RM {new_rm_id} not found in RM Repository")
    
    # Prepare new item
    new_item = {
        "rm_id": rm["rm_id"],
        "rm_name": generate_rm_description(rm.get("category", ""), rm.get("category_data", {}), rm.get("name", "")),
        "quantity": edit_data.quantity,
        "unit": edit_data.unit
    }
    
    if requires_approval:
        # Create change request for approval
        change_request = {
            "id": str(uuid.uuid4()),
            "buyer_sku_id": buyer_sku_id,
            "bidso_sku_id": bidso_sku_id,
            "brand_id": brand_id,
            "change_type": "MODIFY",
            "original_item": original_item,
            "new_item": new_item,
            "item_index": item_index,
            "status": "PENDING",
            "requested_by": current_user.get("id"),
            "requested_by_name": current_user.get("name"),
            "scheduled_branches": schedule_check["branches"],
            "schedule_count": schedule_check["schedule_count"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "approved_by": None,
            "approved_at": None
        }
        await db.bom_change_requests.insert_one(change_request)
        
        # Create notifications
        await create_bom_change_notification(
            change_request["id"],
            buyer_sku_id,
            "MODIFY",
            original_item,
            new_item,
            schedule_check["branches"],
            requires_approval=True
        )
        
        return {
            "status": "PENDING_APPROVAL",
            "message": f"BOM change requires Master Admin approval. SKU is scheduled for production at {schedule_check['schedule_count']} location(s) in the next 10 days.",
            "change_request_id": change_request["id"],
            "scheduled_branches": schedule_check["branches"],
            "notifications_sent": True
        }
    else:
        # Apply change immediately
        items[item_index] = new_item
        await db.brand_specific_bom.update_one(
            {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
            {"$set": {"items": items, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {
            "status": "APPLIED",
            "message": "BOM change applied successfully.",
            "original_item": original_item,
            "new_item": new_item
        }


@router.post("/bom/buyer-sku/{buyer_sku_id}/item")
async def add_brand_specific_bom_item(
    buyer_sku_id: str,
    edit_data: BOMItemEdit,
    current_user: dict = Depends(get_current_user)
):
    """Add a new item to brand-specific BOM"""
    # Get buyer SKU
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    bidso_sku_id = buyer_sku["bidso_sku_id"]
    brand_id = buyer_sku["brand_id"]
    
    # Validate RM
    rm = await db.raw_materials.find_one({"rm_id": edit_data.rm_id}, {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1})
    if not rm:
        escaped_rm_id = re.escape(edit_data.rm_id)
        rm = await db.raw_materials.find_one(
            {"rm_id": {"$regex": f"^{escaped_rm_id}$", "$options": "i"}},
            {"_id": 0, "rm_id": 1, "name": 1, "category": 1, "category_data": 1}
        )
    if not rm:
        raise HTTPException(status_code=404, detail=f"RM {edit_data.rm_id} not found in RM Repository")
    
    new_item = {
        "rm_id": rm["rm_id"],
        "rm_name": generate_rm_description(rm.get("category", ""), rm.get("category_data", {}), rm.get("name", "")),
        "quantity": edit_data.quantity,
        "unit": edit_data.unit
    }
    
    # Check production schedule
    schedule_check = await check_production_schedule_next_10_days(buyer_sku_id)
    requires_approval = schedule_check["is_scheduled"]
    
    if requires_approval:
        change_request = {
            "id": str(uuid.uuid4()),
            "buyer_sku_id": buyer_sku_id,
            "bidso_sku_id": bidso_sku_id,
            "brand_id": brand_id,
            "change_type": "ADD",
            "original_item": None,
            "new_item": new_item,
            "item_index": None,
            "status": "PENDING",
            "requested_by": current_user.get("id"),
            "requested_by_name": current_user.get("name"),
            "scheduled_branches": schedule_check["branches"],
            "schedule_count": schedule_check["schedule_count"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.bom_change_requests.insert_one(change_request)
        
        await create_bom_change_notification(
            change_request["id"],
            buyer_sku_id,
            "ADD",
            {},
            new_item,
            schedule_check["branches"],
            requires_approval=True
        )
        
        return {
            "status": "PENDING_APPROVAL",
            "message": f"BOM change requires Master Admin approval.",
            "change_request_id": change_request["id"]
        }
    else:
        # Add immediately
        await db.brand_specific_bom.update_one(
            {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
            {"$push": {"items": new_item}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        return {"status": "APPLIED", "message": "Item added to BOM", "new_item": new_item}


@router.delete("/bom/buyer-sku/{buyer_sku_id}/item/{rm_id}")
async def remove_brand_specific_bom_item(
    buyer_sku_id: str,
    rm_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove an item from brand-specific BOM"""
    buyer_sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id}, {"_id": 0})
    if not buyer_sku:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    bidso_sku_id = buyer_sku["bidso_sku_id"]
    brand_id = buyer_sku["brand_id"]
    
    brand_bom = await db.brand_specific_bom.find_one(
        {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
        {"_id": 0}
    )
    if not brand_bom:
        raise HTTPException(status_code=404, detail="Brand-specific BOM not found")
    
    items = brand_bom.get("items", [])
    item_to_remove = next((item for item in items if item.get("rm_id") == rm_id), None)
    if not item_to_remove:
        raise HTTPException(status_code=404, detail=f"RM {rm_id} not found in brand-specific BOM")
    
    schedule_check = await check_production_schedule_next_10_days(buyer_sku_id)
    requires_approval = schedule_check["is_scheduled"]
    
    if requires_approval:
        change_request = {
            "id": str(uuid.uuid4()),
            "buyer_sku_id": buyer_sku_id,
            "bidso_sku_id": bidso_sku_id,
            "brand_id": brand_id,
            "change_type": "REMOVE",
            "original_item": item_to_remove,
            "new_item": None,
            "status": "PENDING",
            "requested_by": current_user.get("id"),
            "requested_by_name": current_user.get("name"),
            "scheduled_branches": schedule_check["branches"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.bom_change_requests.insert_one(change_request)
        
        await create_bom_change_notification(
            change_request["id"],
            buyer_sku_id,
            "REMOVE",
            item_to_remove,
            {},
            schedule_check["branches"],
            requires_approval=True
        )
        
        return {"status": "PENDING_APPROVAL", "change_request_id": change_request["id"]}
    else:
        new_items = [item for item in items if item.get("rm_id") != rm_id]
        await db.brand_specific_bom.update_one(
            {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
            {"$set": {"items": new_items, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"status": "APPLIED", "message": "Item removed from BOM"}


@router.get("/bom/change-requests")
async def list_bom_change_requests(
    status: Optional[str] = Query(None, description="Filter by status: PENDING, APPROVED, REJECTED"),
    current_user: dict = Depends(get_current_user)
):
    """List BOM change requests (for Master Admin approval)"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.bom_change_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"requests": requests, "total": len(requests)}


@router.put("/bom/change-request/{request_id}/approve")
async def approve_bom_change_request(
    request_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Approve a pending BOM change request (Master Admin only)"""
    if current_user.get("role") != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can approve BOM changes")
    
    change_request = await db.bom_change_requests.find_one({"id": request_id}, {"_id": 0})
    if not change_request:
        raise HTTPException(status_code=404, detail="Change request not found")
    
    if change_request["status"] != "PENDING":
        raise HTTPException(status_code=400, detail=f"Change request is already {change_request['status']}")
    
    bidso_sku_id = change_request["bidso_sku_id"]
    brand_id = change_request["brand_id"]
    change_type = change_request["change_type"]
    
    # Apply the change
    brand_bom = await db.brand_specific_bom.find_one(
        {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
        {"_id": 0}
    )
    items = brand_bom.get("items", []) if brand_bom else []
    
    if change_type == "MODIFY":
        item_index = change_request.get("item_index")
        if item_index is not None and item_index < len(items):
            items[item_index] = change_request["new_item"]
    elif change_type == "ADD":
        items.append(change_request["new_item"])
    elif change_type == "REMOVE":
        rm_id_to_remove = change_request["original_item"].get("rm_id")
        items = [item for item in items if item.get("rm_id") != rm_id_to_remove]
    
    # Update BOM
    if brand_bom:
        await db.brand_specific_bom.update_one(
            {"bidso_sku_id": bidso_sku_id, "brand_id": brand_id},
            {"$set": {"items": items, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        await db.brand_specific_bom.insert_one({
            "id": str(uuid.uuid4()),
            "bidso_sku_id": bidso_sku_id,
            "brand_id": brand_id,
            "items": items,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Update change request status
    await db.bom_change_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "APPROVED",
            "approved_by": current_user.get("id"),
            "approved_by_name": current_user.get("name"),
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notify requester
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "change_request_id": request_id,
        "recipient_id": change_request.get("requested_by"),
        "notification_type": "BOM_CHANGE_APPROVED",
        "message": f"Your BOM change request for {change_request['buyer_sku_id']} has been approved",
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"status": "APPROVED", "message": "BOM change applied successfully"}


@router.put("/bom/change-request/{request_id}/reject")
async def reject_bom_change_request(
    request_id: str,
    reason: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """Reject a pending BOM change request (Master Admin only)"""
    if current_user.get("role") != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can reject BOM changes")
    
    change_request = await db.bom_change_requests.find_one({"id": request_id}, {"_id": 0})
    if not change_request:
        raise HTTPException(status_code=404, detail="Change request not found")
    
    if change_request["status"] != "PENDING":
        raise HTTPException(status_code=400, detail=f"Change request is already {change_request['status']}")
    
    await db.bom_change_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "REJECTED",
            "rejected_by": current_user.get("id"),
            "rejected_by_name": current_user.get("name"),
            "rejection_reason": reason,
            "rejected_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notify requester
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "change_request_id": request_id,
        "recipient_id": change_request.get("requested_by"),
        "notification_type": "BOM_CHANGE_REJECTED",
        "message": f"Your BOM change request for {change_request['buyer_sku_id']} has been rejected. Reason: {reason or 'Not specified'}",
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"status": "REJECTED", "message": "BOM change request rejected"}


@router.get("/notifications")
async def get_user_notifications(
    current_user: dict = Depends(get_current_user),
    unread_only: bool = Query(False)
):
    """Get notifications for the current user based on role"""
    user_role = current_user.get("role")
    user_id = current_user.get("id")
    user_branches = current_user.get("assigned_branches", [])
    
    query = {"$or": [
        {"recipient_id": user_id},
        {"recipient_role": user_role}
    ]}
    
    # For branch ops, also filter by branch
    if user_role == "branch_ops_user" and user_branches:
        query["$or"].append({"recipient_branch": {"$in": user_branches}})
    
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(query, {"_id": 0}).sort("created_at", -1).to_list(50)
    unread_count = await db.notifications.count_documents({**query, "is_read": False})
    
    return {"notifications": notifications, "unread_count": unread_count}


@router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str):
    """Mark a notification as read"""
    await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"status": "ok"}


# ============ Migration from Old SKU Structure ============

@router.post("/migrate-skus")
async def migrate_skus_from_excel(file: UploadFile = File(...)):
    """
    Migrate SKUs from Excel file (SKU_RM_Mapping.xlsx format).
    Creates Bidso SKUs and Buyer SKUs from existing data.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Parse SKUs
    sku_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sku_id = str(row[0]).strip()
        rm_id = str(row[1]).strip() if row[1] else None
        qty = float(row[2]) if row[2] else 1.0
        
        if sku_id not in sku_data:
            sku_data[sku_id] = {"bom": []}
        
        if rm_id:
            sku_data[sku_id]["bom"].append({"rm_id": rm_id, "quantity": qty})
    
    result = SKUMigrationResult()
    result.total_processed = len(sku_data)
    
    # Get all brands
    brands = await db.brands.find({}, {"_id": 0}).to_list(100)
    brand_map = {b["code"]: b for b in brands}
    
    # Get all verticals and models
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(100)
    vertical_code_map = {v["code"]: v for v in verticals}
    
    models = await db.models.find({}, {"_id": 0}).to_list(500)
    model_code_map = {m["code"]: m for m in models}
    
    for sku_id, data in sku_data.items():
        try:
            parts = sku_id.split("_")
            if len(parts) < 4:
                result.errors.append(f"{sku_id}: Invalid format (less than 4 parts)")
                continue
            
            brand_code = parts[0]
            vertical_code = parts[1]
            model_code = parts[2]
            numeric_code = parts[-1]
            
            # Handle compound model codes (e.g., BE_P becomes BE_P)
            if len(parts) > 4:
                model_code = "_".join(parts[2:-1])
            
            # Generate Bidso SKU ID
            bidso_sku_id = f"{vertical_code}_{model_code}_{numeric_code}"
            
            # Check/create Bidso SKU
            existing_bidso = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id})
            
            if not existing_bidso:
                # Find vertical and model
                vertical = vertical_code_map.get(vertical_code)
                model = model_code_map.get(model_code)
                
                if not vertical or not model:
                    result.errors.append(f"{sku_id}: Vertical {vertical_code} or Model {model_code} not found")
                    continue
                
                bidso_sku = {
                    "id": str(uuid.uuid4()),
                    "bidso_sku_id": bidso_sku_id,
                    "vertical_id": vertical["id"],
                    "vertical_code": vertical_code,
                    "model_id": model["id"],
                    "model_code": model_code,
                    "numeric_code": numeric_code,
                    "name": f"{vertical['name']} - {model['name']}",
                    "description": "",
                    "status": "ACTIVE",
                    "created_at": datetime.now(timezone.utc)
                }
                await db.bidso_skus.insert_one(bidso_sku)
                result.bidso_skus_created += 1
                
                # Create common BOM
                if data["bom"]:
                    bom = {
                        "id": str(uuid.uuid4()),
                        "bidso_sku_id": bidso_sku_id,
                        "items": data["bom"],
                        "is_locked": False,
                        "created_at": datetime.now(timezone.utc)
                    }
                    await db.common_bom.insert_one(bom)
                    result.bom_migrated += 1
            
            # Check/create Buyer SKU
            brand = brand_map.get(brand_code)
            if brand:
                buyer_sku_id = f"{brand_code}_{bidso_sku_id}"
                existing_buyer = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
                
                if not existing_buyer:
                    buyer_sku = {
                        "id": str(uuid.uuid4()),
                        "buyer_sku_id": buyer_sku_id,
                        "bidso_sku_id": bidso_sku_id,
                        "brand_id": brand["id"],
                        "brand_code": brand_code,
                        "name": f"{brand['name']} - {bidso_sku_id}",
                        "description": "",
                        "status": "ACTIVE",
                        "created_at": datetime.now(timezone.utc)
                    }
                    await db.buyer_skus.insert_one(buyer_sku)
                    result.buyer_skus_created += 1
            else:
                result.errors.append(f"{sku_id}: Brand {brand_code} not found")
        
        except Exception as e:
            result.errors.append(f"{sku_id}: {str(e)}")
    
    return result


@router.get("/migration-stats")
async def get_migration_stats():
    """Get current state of SKU migration"""
    bidso_count = await db.bidso_skus.count_documents({})
    buyer_count = await db.buyer_skus.count_documents({})
    common_bom_count = await db.common_bom.count_documents({})
    brand_bom_count = await db.brand_specific_bom.count_documents({})
    
    # Legacy collection count (for migration monitoring only)
    old_sku_count = await db.skus.count_documents({})
    
    return {
        "bidso_skus": bidso_count,
        "buyer_skus": buyer_count,
        "common_boms": common_bom_count,
        "brand_specific_boms": brand_bom_count,
        "legacy_skus_deprecated": old_sku_count,
        "migration_status": "COMPLETE" if old_sku_count == 0 else "LEGACY_DATA_EXISTS",
        "collections_created": ["bidso_skus", "buyer_skus", "common_bom", "brand_specific_bom"],
        "note": "Legacy 'skus' collection is deprecated. All new operations use bidso_skus + buyer_skus."
    }


@router.delete("/legacy-skus/drop")
async def drop_legacy_skus_collection():
    """
    DANGER: Permanently drops the legacy 'skus' collection.
    Only call this after confirming all data is migrated to buyer_skus.
    """
    # Safety check: ensure buyer_skus has data
    buyer_count = await db.buyer_skus.count_documents({})
    if buyer_count == 0:
        raise HTTPException(
            status_code=400, 
            detail="Cannot drop legacy collection: buyer_skus is empty. Migration may not be complete."
        )
    
    # Get count before drop
    legacy_count = await db.skus.count_documents({})
    
    # Drop the collection
    await db.skus.drop()
    
    return {
        "success": True,
        "message": f"Legacy 'skus' collection dropped. {legacy_count} records removed.",
        "buyer_skus_count": buyer_count,
        "migration_status": "COMPLETE"
    }



# ============ Clone & Customize Bidso SKU ============

@router.get("/bidso-skus/{bidso_sku_id}/bom-for-clone")
async def get_bom_for_clone(bidso_sku_id: str):
    """
    Get Common BOM of a Bidso SKU formatted for cloning.
    Returns BOM items with edit permissions based on category:
    - INP, INM: Editable (colour change only)
    - ACC: Editable (colour change or complete swap)
    - Others (ELC, SP, etc.): Locked
    """
    # Get the Bidso SKU
    bidso_sku = await db.bidso_skus.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
    if not bidso_sku:
        raise HTTPException(status_code=404, detail="Bidso SKU not found")
    
    # Get the Common BOM
    common_bom = await db.common_bom.find_one({"bidso_sku_id": bidso_sku_id}, {"_id": 0})
    if not common_bom or not common_bom.get("items"):
        raise HTTPException(status_code=404, detail="No BOM found for this Bidso SKU")
    
    # Enrich BOM items with RM details and edit permissions
    enriched_items = []
    for item in common_bom.get("items", []):
        rm_id = item.get("rm_id")
        rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
        
        if rm:
            category = rm.get("category", "")
            category_data = rm.get("category_data", {})
            
            # Determine edit type based on category
            if category in ["INP", "INM"]:
                edit_type = "COLOUR_ONLY"  # Can only change colour variant
            elif category == "ACC":
                edit_type = "COLOUR_OR_SWAP"  # Can change colour or swap entirely
            else:
                edit_type = "LOCKED"  # Cannot edit
            
            enriched_items.append({
                "rm_id": rm_id,
                "rm_name": category_data.get("name", ""),
                "category": category,
                "category_data": category_data,
                "quantity": item.get("quantity", 1),
                "unit": item.get("unit", "nos"),
                "edit_type": edit_type,
                "colour": category_data.get("colour", ""),
                "model_name": category_data.get("model_name", ""),
                "part_name": category_data.get("part_name", ""),
            })
    
    return {
        "source_sku": bidso_sku,
        "bom_items": enriched_items,
        "total_items": len(enriched_items),
        "editable_count": len([i for i in enriched_items if i["edit_type"] != "LOCKED"]),
        "locked_count": len([i for i in enriched_items if i["edit_type"] == "LOCKED"])
    }


@router.get("/raw-materials/colour-variants/{rm_id}")
async def get_colour_variants(rm_id: str):
    """
    Find colour variants of an RM (same base part, different colours).
    For INP/INM: Same mould_code/model_name + part_name, different colour
    For ACC: Same type + model_name + specs, different colour
    """
    # Get the source RM
    source_rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
    if not source_rm:
        raise HTTPException(status_code=404, detail="RM not found")
    
    category = source_rm.get("category", "")
    category_data = source_rm.get("category_data", {})
    
    query = {"category": category, "rm_id": {"$ne": rm_id}}
    
    if category == "INP":
        # Match by mould_code, model_name, part_name
        query["category_data.mould_code"] = category_data.get("mould_code")
        query["category_data.model_name"] = category_data.get("model_name")
        query["category_data.part_name"] = category_data.get("part_name")
    elif category == "INM":
        # Match by model_name, part_name
        query["category_data.model_name"] = category_data.get("model_name")
        query["category_data.part_name"] = category_data.get("part_name")
    elif category == "ACC":
        # Match by type, model_name, specs
        query["category_data.type"] = category_data.get("type")
        query["category_data.model_name"] = category_data.get("model_name")
        query["category_data.specs"] = category_data.get("specs")
    else:
        return {"variants": [], "message": "Category does not support colour variants"}
    
    # Find variants
    variants = await db.raw_materials.find(query, {"_id": 0}).to_list(100)
    
    # Format response
    result = []
    for v in variants:
        v_data = v.get("category_data", {})
        result.append({
            "rm_id": v.get("rm_id"),
            "category": v.get("category"),
            "colour": v_data.get("colour", "N/A"),
            "name": v_data.get("name", ""),
            "category_data": v_data
        })
    
    return {
        "source_rm": {
            "rm_id": rm_id,
            "colour": category_data.get("colour", "N/A"),
            "name": category_data.get("name", "")
        },
        "variants": result,
        "total_variants": len(result)
    }


@router.get("/raw-materials/search-for-swap")
async def search_rm_for_swap(
    category: str,
    search: Optional[str] = None,
    limit: int = 50
):
    """
    Search RMs for swapping (ACC category).
    Returns RMs in same category that can replace current RM.
    """
    query = {"category": category}
    
    rms = await db.raw_materials.find(query, {"_id": 0}).to_list(1000)
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        rms = [rm for rm in rms if 
               search_lower in rm.get("rm_id", "").lower() or
               search_lower in rm.get("category_data", {}).get("name", "").lower() or
               search_lower in rm.get("category_data", {}).get("type", "").lower() or
               search_lower in rm.get("category_data", {}).get("model_name", "").lower()]
    
    # Format response
    result = []
    for rm in rms[:limit]:
        cd = rm.get("category_data", {})
        result.append({
            "rm_id": rm.get("rm_id"),
            "category": rm.get("category"),
            "name": cd.get("name", ""),
            "type": cd.get("type", ""),
            "model_name": cd.get("model_name", ""),
            "specs": cd.get("specs", ""),
            "colour": cd.get("colour", ""),
            "category_data": cd
        })
    
    return {"results": result, "total": len(result)}


@router.post("/bidso-skus/clone")
async def clone_bidso_sku(data: dict, current_user = Depends(get_current_user)):
    """
    Create a new Bidso SKU by cloning an existing one with modifications.
    
    Request body:
    {
        "source_bidso_sku_id": "KS_PE_001",
        "name": "Kids Scooter - Red",
        "description": "Red colour variant",
        "bom_modifications": [
            {"original_rm_id": "INP_001", "new_rm_id": "INP_003"},  // Colour swap
            {"original_rm_id": "ACC_010", "new_rm_id": "ACC_015"}   // ACC replacement
        ],
        "new_rms_to_create": [  // Optional: Create new RMs on-the-fly
            {
                "category": "INP",
                "category_data": {"mould_code": "M001", "model_name": "Body", "part_name": "Shell", "colour": "Red", "mb": "Red MB"},
                "replaces_rm_id": "INP_001"
            }
        ]
    }
    """
    source_sku_id = data.get("source_bidso_sku_id")
    if not source_sku_id:
        raise HTTPException(status_code=400, detail="source_bidso_sku_id is required")
    
    # Get source Bidso SKU
    source_sku = await db.bidso_skus.find_one({"bidso_sku_id": source_sku_id}, {"_id": 0})
    if not source_sku:
        raise HTTPException(status_code=404, detail="Source Bidso SKU not found")
    
    # Get source BOM
    source_bom = await db.common_bom.find_one({"bidso_sku_id": source_sku_id}, {"_id": 0})
    if not source_bom:
        raise HTTPException(status_code=404, detail="Source BOM not found")
    
    # Generate new Bidso SKU ID
    vertical_code = source_sku.get("vertical_code")
    model_code = source_sku.get("model_code")
    numeric_code = await get_next_numeric_code(vertical_code, model_code)
    new_bidso_sku_id = f"{vertical_code}_{model_code}_{numeric_code}"
    
    # Check if already exists
    existing = await db.bidso_skus.find_one({"bidso_sku_id": new_bidso_sku_id})
    if existing:
        raise HTTPException(status_code=400, detail=f"Bidso SKU {new_bidso_sku_id} already exists")
    
    # Create new RMs on-the-fly if specified
    new_rm_mapping = {}  # original_rm_id -> new_rm_id
    new_rms_to_create = data.get("new_rms_to_create", [])
    
    for new_rm_data in new_rms_to_create:
        category = new_rm_data.get("category")
        category_data = new_rm_data.get("category_data", {})
        replaces_rm_id = new_rm_data.get("replaces_rm_id")
        
        # Generate RM ID
        seq = await get_next_rm_sequence(category)
        new_rm_id = f"{category}_{seq:05d}"
        
        # Generate name from nomenclature
        rm_name = generate_rm_name(category, category_data)
        if rm_name:
            category_data["name"] = rm_name
        
        # Create the RM
        new_rm = {
            "id": str(uuid.uuid4()),
            "rm_id": new_rm_id,
            "category": category,
            "category_data": category_data,
            "status": "ACTIVE",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user.id
        }
        await db.raw_materials.insert_one(new_rm)
        
        if replaces_rm_id:
            new_rm_mapping[replaces_rm_id] = new_rm_id
    
    # Process BOM modifications
    bom_modifications = data.get("bom_modifications", [])
    for mod in bom_modifications:
        original_rm_id = mod.get("original_rm_id")
        new_rm_id = mod.get("new_rm_id")
        if original_rm_id and new_rm_id:
            new_rm_mapping[original_rm_id] = new_rm_id
    
    # Create new BOM by copying source and applying modifications
    new_bom_items = []
    for item in source_bom.get("items", []):
        rm_id = item.get("rm_id")
        
        # Check if this RM should be swapped
        if rm_id in new_rm_mapping:
            new_rm_id = new_rm_mapping[rm_id]
            # Get new RM name
            new_rm = await db.raw_materials.find_one({"rm_id": new_rm_id}, {"_id": 0})
            rm_name = new_rm.get("category_data", {}).get("name", "") if new_rm else ""
            new_bom_items.append({
                "rm_id": new_rm_id,
                "rm_name": rm_name,
                "quantity": item.get("quantity", 1),
                "unit": item.get("unit", "nos")
            })
        else:
            # Keep original
            new_bom_items.append(item)
    
    # Create new Bidso SKU
    new_bidso_sku = {
        "id": str(uuid.uuid4()),
        "bidso_sku_id": new_bidso_sku_id,
        "vertical_id": source_sku.get("vertical_id"),
        "vertical_code": vertical_code,
        "model_id": source_sku.get("model_id"),
        "model_code": model_code,
        "numeric_code": numeric_code,
        "name": data.get("name", f"{source_sku.get('name', '')} - Variant"),
        "description": data.get("description", f"Cloned from {source_sku_id}"),
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    await db.bidso_skus.insert_one(new_bidso_sku)
    
    # Create new Common BOM
    new_common_bom = {
        "id": str(uuid.uuid4()),
        "bidso_sku_id": new_bidso_sku_id,
        "items": new_bom_items,
        "is_locked": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    await db.common_bom.insert_one(new_common_bom)
    
    return {
        "message": "Bidso SKU cloned successfully",
        "new_bidso_sku_id": new_bidso_sku_id,
        "source_bidso_sku_id": source_sku_id,
        "bom_items_count": len(new_bom_items),
        "modifications_applied": len(new_rm_mapping),
        "new_rms_created": len(new_rms_to_create)
    }




# ============ HSN/GST Management ============

class HSNGSTUpdate(BaseModel):
    hsn_code: str
    gst_rate: float = 18


@router.put("/buyer-skus/{buyer_sku_id}/hsn-gst")
async def update_buyer_sku_hsn_gst(
    buyer_sku_id: str,
    data: HSNGSTUpdate,
    current_user = Depends(get_current_user)
):
    """Update HSN code and GST rate for a Buyer SKU"""
    
    # Validate HSN code format (typically 4-8 digits)
    hsn_code = data.hsn_code.strip()
    if hsn_code and not re.match(r'^\d{4,8}$', hsn_code):
        raise HTTPException(status_code=400, detail="HSN code must be 4-8 digits")
    
    # Validate GST rate
    valid_gst_rates = [0, 5, 12, 18, 28]
    if data.gst_rate not in valid_gst_rates:
        raise HTTPException(status_code=400, detail=f"GST rate must be one of: {valid_gst_rates}")
    
    # Check if SKU exists
    sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
    if not sku:
        raise HTTPException(status_code=404, detail="Buyer SKU not found")
    
    # Update
    await db.buyer_skus.update_one(
        {"buyer_sku_id": buyer_sku_id},
        {"$set": {
            "hsn_code": hsn_code,
            "gst_rate": data.gst_rate,
            "hsn_updated_at": datetime.now(timezone.utc).isoformat(),
            "hsn_updated_by": current_user.id
        }}
    )
    
    return {
        "message": f"HSN/GST updated for {buyer_sku_id}",
        "hsn_code": hsn_code,
        "gst_rate": data.gst_rate
    }


@router.post("/buyer-skus/bulk-hsn-gst")
async def bulk_update_hsn_gst(
    file: UploadFile = File(...),
    current_user = Depends(get_current_user)
):
    """
    Bulk update HSN codes and GST rates from Excel.
    Expected columns: buyer_sku_id, hsn_code, gst_rate
    """
    
    # Read Excel file
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
    
    # Parse headers
    headers = [cell.value.lower().strip() if cell.value else "" for cell in ws[1]]
    
    required = ["buyer_sku_id", "hsn_code", "gst_rate"]
    for col in required:
        if col not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required column: {col}")
    
    sku_idx = headers.index("buyer_sku_id")
    hsn_idx = headers.index("hsn_code")
    gst_idx = headers.index("gst_rate")
    
    # Process rows
    updated = 0
    errors = []
    valid_gst_rates = [0, 5, 12, 18, 28]
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[sku_idx]:
            continue
        
        buyer_sku_id = str(row[sku_idx]).strip()
        hsn_code = str(row[hsn_idx]).strip() if row[hsn_idx] else ""
        
        try:
            gst_rate = float(row[gst_idx]) if row[gst_idx] else 18
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: Invalid GST rate")
            continue
        
        # Validate HSN
        if hsn_code and not re.match(r'^\d{4,8}$', hsn_code):
            errors.append(f"Row {row_num}: Invalid HSN code '{hsn_code}' (must be 4-8 digits)")
            continue
        
        # Validate GST
        if gst_rate not in valid_gst_rates:
            errors.append(f"Row {row_num}: Invalid GST rate {gst_rate}")
            continue
        
        # Check SKU exists
        sku = await db.buyer_skus.find_one({"buyer_sku_id": buyer_sku_id})
        if not sku:
            errors.append(f"Row {row_num}: SKU '{buyer_sku_id}' not found")
            continue
        
        # Update
        await db.buyer_skus.update_one(
            {"buyer_sku_id": buyer_sku_id},
            {"$set": {
                "hsn_code": hsn_code,
                "gst_rate": gst_rate,
                "hsn_updated_at": datetime.now(timezone.utc).isoformat(),
                "hsn_updated_by": current_user.id
            }}
        )
        updated += 1
    
    return {
        "message": "Bulk HSN/GST update complete",
        "updated": updated,
        "errors": errors,
        "error_count": len(errors)
    }
