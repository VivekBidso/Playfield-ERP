from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import openpyxl
import io
import jwt
import hashlib

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480  # 8 hours

security = HTTPBearer()

BRANCHES = [
    "Unit 1 Vedica",
    "Unit 2 Trikes",
    "Unit 3 TM",
    "Unit 4 Goa",
    "Unit 5 Baabus",
    "Unit 6 Emox",
    "BHDG WH"
]

RM_CATEGORIES = {
    "INP": {"name": "In-house Plastic", "fields": ["mould_code", "model_name", "part_name", "colour", "mb", "per_unit_weight", "unit"]},
    "ACC": {"name": "Accessories", "fields": ["type", "model_name", "specs", "colour", "per_unit_weight", "unit"]},
    "ELC": {"name": "Electric Components", "fields": ["model", "type", "specs", "per_unit_weight", "unit"]},
    "SP": {"name": "Spares", "fields": ["type", "specs", "per_unit_weight", "unit"]},
    "BS": {"name": "Brand Assets", "fields": ["position", "type", "brand", "buyer_sku", "per_unit_weight", "unit"]},
    "PM": {"name": "Packaging", "fields": ["model", "type", "specs", "brand", "per_unit_weight", "unit"]},
    "LB": {"name": "Labels", "fields": ["type", "buyer_sku", "per_unit_weight", "unit"]}
}

# ============ Models ============

class RawMaterial(BaseModel):
    """Global RM definition with L1/L2 support"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    rm_id: str
    category: str
    category_data: Dict[str, Any] = {}
    low_stock_threshold: float = 10.0
    # L1/L2 Fields
    rm_level: str = "DIRECT"  # L1, L2, DIRECT
    parent_rm_id: Optional[str] = None  # L1 parent for L2 items
    unit_weight_grams: Optional[float] = None  # Weight per unit (for INP L2)
    scrap_factor: float = 0.02  # Waste factor (2% default)
    processing_cost: float = 0.0  # Per-unit processing cost
    # INM-specific fields
    secondary_l1_rm_id: Optional[str] = None  # Powder coating RM (INM only)
    powder_qty_grams: Optional[float] = None  # Predefined powder coating qty in grams
    coating_scrap_factor: float = 0.10  # Coating waste factor (INM only)
    status: str = "ACTIVE"  # ACTIVE, INACTIVE, DISCONTINUED
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class RawMaterialCreate(BaseModel):
    category: str
    category_data: Dict[str, Any]
    low_stock_threshold: float = 10.0
    # L1/L2 Fields
    rm_level: str = "DIRECT"
    parent_rm_id: Optional[str] = None
    unit_weight_grams: Optional[float] = None
    scrap_factor: float = 0.02
    processing_cost: float = 0.0
    secondary_l1_rm_id: Optional[str] = None
    powder_qty_grams: Optional[float] = None
    coating_scrap_factor: float = 0.10

class BranchRMInventory(BaseModel):
    """Branch-specific RM inventory"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    rm_id: str
    branch: str
    current_stock: float = 0.0
    is_active: bool = True
    activated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SKU(BaseModel):
    """Global SKU definition with extended fields"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_id: str
    bidso_sku: str
    buyer_sku_id: str
    description: str
    brand: str
    vertical: str
    model: str
    low_stock_threshold: float = 5.0
    # New normalized references
    vertical_id: Optional[str] = None
    model_id: Optional[str] = None
    brand_id: Optional[str] = None
    buyer_id: Optional[str] = None
    status: str = "DRAFT"  # DRAFT, BOM_PENDING, BOM_COMPLETE, ACTIVE, DISCONTINUED
    bom_finalized_at: Optional[datetime] = None
    bom_finalized_by: Optional[str] = None
    base_price: Optional[float] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class SKUCreate(BaseModel):
    sku_id: str
    bidso_sku: str
    buyer_sku_id: str
    description: str
    brand: str
    vertical: str
    model: str
    low_stock_threshold: float = 5.0

class BranchSKUInventory(BaseModel):
    """Branch-specific SKU inventory"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_id: str
    branch: str
    current_stock: float = 0.0
    is_active: bool = True
    activated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RMMapping(BaseModel):
    rm_id: str
    quantity_required: float

class SKUMapping(BaseModel):
    """Global SKU to RM mapping (BOM)"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_id: str
    rm_mappings: List[RMMapping]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SKUMappingCreate(BaseModel):
    sku_id: str
    rm_mappings: List[RMMapping]

class PurchaseEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    rm_id: str
    branch: str
    quantity: float
    date: datetime
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PurchaseEntryCreate(BaseModel):
    rm_id: str
    branch: str
    quantity: float
    date: datetime
    notes: Optional[str] = None

class ProductionEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_id: str
    branch: str
    quantity: float
    date: datetime
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductionEntryCreate(BaseModel):
    sku_id: str
    branch: str
    quantity: float
    date: datetime
    notes: Optional[str] = None

class DispatchEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_id: str
    branch: str
    quantity: float
    date: datetime
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DispatchEntryCreate(BaseModel):
    sku_id: str
    branch: str
    quantity: float
    date: datetime
    notes: Optional[str] = None

class ProductionPlanEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch: str
    plan_month: str  # YYYY-MM format
    date: datetime
    sku_id: str
    planned_quantity: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductionPlanCreate(BaseModel):
    branch: str
    plan_month: str
    date: datetime
    sku_id: str
    planned_quantity: float

class ActivateItemRequest(BaseModel):
    item_id: str
    branch: str

# ============ User & Auth Models ============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    password_hash: str
    name: str
    role: str  # "master_admin" or "branch_user"
    assigned_branches: List[str] = []  # Empty for master_admin, specific branches for branch_user
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str
    assigned_branches: List[str] = []

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    assigned_branches: List[str]
    is_active: bool
    created_at: datetime

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class LoginResponse(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

# ============ Vendor Models ============

class Vendor(BaseModel):
    """Vendor definition with extended fields"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vendor_id: str = ""  # Auto-generated unique ID like VND_001
    name: str
    gst: str = ""
    address: str = ""
    poc: str = ""  # Point of Contact
    email: str = ""
    phone: str = ""
    # Extended fields
    legal_name: str = ""  # GST-validated legal name
    gst_validated: bool = False
    gst_validated_at: Optional[datetime] = None
    payment_terms_days: int = 30
    rating: Optional[float] = None  # 1-5
    is_active: bool = True
    status: str = "ACTIVE"  # ACTIVE, INACTIVE, BLACKLISTED
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class VendorCreate(BaseModel):
    name: str
    gst: str = ""
    address: str = ""
    poc: str = ""
    email: str = ""
    phone: str = ""

class VendorRMPrice(BaseModel):
    """Vendor RM pricing"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vendor_id: str
    rm_id: str
    price: float
    currency: str = "INR"
    notes: str = ""
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class VendorRMPriceCreate(BaseModel):
    vendor_id: str
    rm_id: str
    price: float
    currency: str = "INR"
    notes: str = ""

# ============ SKU Branch Assignment Model ============

class SKUBranchAssignment(BaseModel):
    """Track which SKUs are assigned to which branches"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_id: str
    branch: str
    assigned_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============ NEW ARCHITECTURE MODELS (PRD v2) ============

# --- Master Data Entities ---

class Vertical(BaseModel):
    """Product Vertical (e.g., SCOOTER, TRIKE)"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str
    name: str
    description: str = ""
    status: str = "ACTIVE"  # ACTIVE, INACTIVE
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Model(BaseModel):
    """Product Model under a Vertical"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vertical_id: str
    code: str
    name: str
    description: str = ""
    status: str = "ACTIVE"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Brand(BaseModel):
    """Brand (tied to Buyer)"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str
    name: str
    buyer_id: Optional[str] = None
    logo_url: str = ""
    status: str = "ACTIVE"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Buyer(BaseModel):
    """Buyer/Customer"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str
    name: str
    country: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    payment_terms_days: int = 30
    status: str = "ACTIVE"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Branch(BaseModel):
    """Branch/Unit definition"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str
    name: str
    location: str = ""
    branch_type: str = "PRODUCTION"  # PRODUCTION, WAREHOUSE, HYBRID
    capacity_units_per_day: int = 0
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# --- Transactional Entities ---

class Forecast(BaseModel):
    """Demand Forecast"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    forecast_code: str
    buyer_id: Optional[str] = None
    vertical_id: Optional[str] = None
    sku_id: Optional[str] = None
    forecast_month: datetime
    quantity: int
    priority: str = "MEDIUM"  # LOW, MEDIUM, HIGH, CRITICAL
    status: str = "DRAFT"  # DRAFT, CONFIRMED, CONVERTED
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None
    confirmed_at: Optional[datetime] = None
    confirmed_by: Optional[str] = None

class DispatchLot(BaseModel):
    """Dispatch Lot for fulfillment"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lot_code: str
    forecast_id: Optional[str] = None
    sku_id: str
    buyer_id: str
    required_quantity: int
    produced_quantity: int = 0
    qc_passed_quantity: int = 0
    dispatched_quantity: int = 0
    target_date: datetime
    status: str = "CREATED"  # CREATED, PRODUCTION_ASSIGNED, PARTIALLY_PRODUCED, FULLY_PRODUCED, QC_CLEARED, DISPATCH_READY, DISPATCHED, DELIVERED
    priority: str = "MEDIUM"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class ProductionBatch(BaseModel):
    """Production Batch for tracking"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    batch_code: str
    production_plan_id: Optional[str] = None
    dispatch_lot_id: Optional[str] = None
    branch_id: str
    branch: str  # Denormalized for easy access
    sku_id: str
    planned_quantity: int
    produced_quantity: int = 0
    good_quantity: int = 0
    rejected_quantity: int = 0
    batch_date: datetime
    shift: str = "DAY"  # DAY, NIGHT
    status: str = "PLANNED"  # PLANNED, IN_PROGRESS, COMPLETED, QC_HOLD, QC_PASSED, QC_FAILED, FG_READY
    started_at: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    completed_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class RMStockMovement(BaseModel):
    """Append-only RM stock movement log"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    movement_code: str
    rm_id: str
    branch_id: str
    branch: str  # Denormalized
    movement_type: str  # INWARD, CONSUMPTION, ADJUSTMENT, TRANSFER_OUT, TRANSFER_IN, SCRAP, PRODUCTION
    quantity: float  # Positive for in, negative for out
    unit_of_measure: str  # KG, PCS, etc.
    reference_type: Optional[str] = None  # PRODUCTION_BATCH, PURCHASE_ORDER, IBT, ADJUSTMENT
    reference_id: Optional[str] = None
    l1_rm_id: Optional[str] = None  # For L2, the L1 source
    l1_quantity_consumed: Optional[float] = None
    unit_cost: Optional[float] = None
    total_cost: Optional[float] = None
    balance_after: float
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class QCChecklist(BaseModel):
    """Quality Control Checklist"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    checklist_code: str
    name: str
    description: str = ""
    check_type: str  # VISUAL, MEASUREMENT, FUNCTIONAL, SAFETY
    vertical_id: Optional[str] = None  # NULL = all verticals
    model_id: Optional[str] = None
    brand_id: Optional[str] = None
    expected_value: str = ""
    tolerance: str = ""
    is_mandatory: bool = True
    check_priority: int = 100
    status: str = "ACTIVE"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class QCResult(BaseModel):
    """QC Inspection Result"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    result_code: str
    production_batch_id: str
    checklist_id: str
    sample_size: int
    passed_count: int
    failed_count: int
    actual_value: str = ""
    result_status: str  # PASSED, FAILED, CONDITIONAL
    defect_type: str = ""
    defect_description: str = ""
    inspector_notes: str = ""
    inspected_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    inspected_by: str

class QCApproval(BaseModel):
    """Batch-level QC Approval"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    production_batch_id: str
    total_inspected: int
    total_passed: int
    total_failed: int
    overall_status: str  # APPROVED, REJECTED, CONDITIONAL, REWORK
    approved_quantity: int = 0
    rejection_reason: str = ""
    rework_instructions: str = ""
    approved_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    approved_by: str

class FGInventory(BaseModel):
    """Finished Goods Inventory"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch_id: str
    branch: str  # Denormalized
    sku_id: str
    dispatch_lot_id: Optional[str] = None
    production_batch_id: Optional[str] = None
    quantity: int
    unit_cost: Optional[float] = None
    status: str = "AVAILABLE"  # AVAILABLE, RESERVED, DISPATCHED, DAMAGED
    qc_approval_id: Optional[str] = None
    received_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class PurchaseOrder(BaseModel):
    """Purchase Order"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    po_number: str
    vendor_id: str
    branch_id: str
    branch: str  # Denormalized
    production_plan_id: Optional[str] = None
    order_date: datetime
    expected_delivery_date: Optional[datetime] = None
    total_amount: float = 0
    currency: str = "INR"
    status: str = "DRAFT"  # DRAFT, SENT, ACKNOWLEDGED, PARTIAL, RECEIVED, CANCELLED
    payment_status: str = "PENDING"  # PENDING, PARTIAL, PAID
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    approved_by: Optional[str] = None

class PurchaseOrderLine(BaseModel):
    """PO Line Item"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    po_id: str
    rm_id: str
    quantity_ordered: float
    quantity_received: float = 0
    unit_price: float
    unit_of_measure: str
    line_total: float = 0
    status: str = "PENDING"  # PENDING, PARTIAL, RECEIVED, CANCELLED

class Dispatch(BaseModel):
    """Dispatch Record"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dispatch_code: str
    dispatch_lot_id: str
    branch_id: str
    branch: str  # Denormalized
    buyer_id: str
    sku_id: str
    quantity: int
    dispatch_date: datetime
    shipping_method: str = ""
    tracking_number: str = ""
    status: str = "PENDING"  # PENDING, SHIPPED, IN_TRANSIT, DELIVERED, RETURNED
    shipped_at: Optional[datetime] = None
    delivered_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Invoice(BaseModel):
    """Invoice"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    dispatch_id: Optional[str] = None
    buyer_id: str
    invoice_date: datetime
    due_date: Optional[datetime] = None
    subtotal: float
    tax_amount: float = 0
    total_amount: float
    currency: str = "INR"
    status: str = "DRAFT"  # DRAFT, SENT, PAID, OVERDUE, CANCELLED
    payment_received_at: Optional[datetime] = None
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class IBTTransfer(BaseModel):
    """Inter-Branch Transfer"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transfer_code: str
    transfer_type: str  # RM, FG
    source_branch_id: str
    source_branch: str  # Denormalized
    destination_branch_id: str
    destination_branch: str  # Denormalized
    item_id: str  # rm_id or sku_id
    quantity: float
    unit_of_measure: str = ""
    status: str = "INITIATED"  # INITIATED, APPROVED, IN_TRANSIT, RECEIVED, COMPLETED, REJECTED, CANCELLED
    initiated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    initiated_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    approved_by: Optional[str] = None
    shipped_at: Optional[datetime] = None
    received_at: Optional[datetime] = None
    received_by: Optional[str] = None
    rejection_reason: str = ""
    notes: str = ""

class PriceHistory(BaseModel):
    """Price change audit trail"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str  # VENDOR_RM, SKU_BUYER, RM_COST
    entity_id: str
    old_price: Optional[float] = None
    new_price: float
    currency: str = "INR"
    change_reason: str = ""
    effective_date: datetime
    approved_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class AuditLog(BaseModel):
    """System-wide audit log"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str  # Table name
    entity_id: str
    action: str  # CREATE, UPDATE, DELETE, STATUS_CHANGE
    old_values: Optional[Dict[str, Any]] = None
    new_values: Optional[Dict[str, Any]] = None
    user_id: Optional[str] = None
    user_email: str = ""
    ip_address: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# --- Request/Response Models for New Entities ---

class VerticalCreate(BaseModel):
    code: str
    name: str
    description: str = ""

class ModelCreate(BaseModel):
    vertical_id: str
    code: str
    name: str
    description: str = ""

class BrandCreate(BaseModel):
    code: str
    name: str
    buyer_id: Optional[str] = None

class BuyerCreate(BaseModel):
    code: str
    name: str
    country: str = ""
    contact_email: str = ""
    payment_terms_days: int = 30

class ForecastCreate(BaseModel):
    buyer_id: Optional[str] = None
    vertical_id: Optional[str] = None
    sku_id: Optional[str] = None
    forecast_month: datetime
    quantity: int
    priority: str = "MEDIUM"
    notes: str = ""

class DispatchLotCreate(BaseModel):
    forecast_id: Optional[str] = None
    sku_id: str
    buyer_id: str
    required_quantity: int
    target_date: datetime
    priority: str = "MEDIUM"

class ProductionBatchCreate(BaseModel):
    production_plan_id: Optional[str] = None
    dispatch_lot_id: Optional[str] = None
    branch: str
    sku_id: str
    planned_quantity: int
    batch_date: datetime
    shift: str = "DAY"

class QCChecklistCreate(BaseModel):
    name: str
    description: str = ""
    check_type: str
    vertical_id: Optional[str] = None
    model_id: Optional[str] = None
    brand_id: Optional[str] = None
    expected_value: str = ""
    tolerance: str = ""
    is_mandatory: bool = True
    check_priority: int = 100

class QCResultCreate(BaseModel):
    production_batch_id: str
    checklist_id: str
    sample_size: int
    passed_count: int
    failed_count: int
    actual_value: str = ""
    defect_type: str = ""
    defect_description: str = ""
    inspector_notes: str = ""

class PurchaseOrderCreate(BaseModel):
    vendor_id: str
    branch: str
    order_date: datetime
    expected_delivery_date: Optional[datetime] = None
    notes: str = ""

class PurchaseOrderLineCreate(BaseModel):
    rm_id: str
    quantity_ordered: float
    unit_price: float
    unit_of_measure: str

# ============ Helper Functions ============

def hash_password(password: str) -> str:
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify password against hash"""
    return hash_password(plain_password) == hashed_password

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    """Create JWT access token"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_next_vendor_id() -> str:
    """Generate next sequential vendor ID like VND_001, VND_002, etc."""
    import re
    all_vendors = await db.vendors.find({}, {"_id": 0, "vendor_id": 1}).to_list(10000)
    
    max_seq = 0
    pattern = re.compile(r'^VND_(\d+)$')
    
    for v in all_vendors:
        vendor_id = v.get('vendor_id', '')
        if vendor_id:
            match = pattern.match(vendor_id)
            if match:
                seq = int(match.group(1))
                max_seq = max(max_seq, seq)
    
    next_seq = max_seq + 1
    return f"VND_{next_seq:03d}"

async def activate_rms_for_sku(sku_id: str, branch: str) -> int:
    """
    Activate all RMs in the BOM for a given SKU in a branch.
    Returns the number of RMs activated.
    """
    activated_count = 0
    
    # Get RM mappings from sku_rm_mapping collection (bulk uploaded)
    rm_mappings = await db.sku_rm_mapping.find({"sku_id": sku_id}, {"_id": 0, "rm_id": 1}).to_list(1000)
    
    # Also check legacy sku_mappings collection
    legacy_mapping = await db.sku_mappings.find_one({"sku_id": sku_id}, {"_id": 0})
    if legacy_mapping and legacy_mapping.get('rm_mappings'):
        for rm in legacy_mapping['rm_mappings']:
            rm_mappings.append({"rm_id": rm['rm_id']})
    
    # Activate each RM in the branch
    for mapping in rm_mappings:
        rm_id = mapping['rm_id']
        
        # Check if RM exists in the system
        rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
        if not rm:
            continue
        
        # Check if already activated in branch
        existing_inv = await db.branch_rm_inventory.find_one(
            {"rm_id": rm_id, "branch": branch},
            {"_id": 0}
        )
        
        if not existing_inv:
            # Activate RM in branch inventory
            inv_obj = BranchRMInventory(rm_id=rm_id, branch=branch)
            inv_doc = inv_obj.model_dump()
            inv_doc['activated_at'] = inv_doc['activated_at'].isoformat()
            await db.branch_rm_inventory.insert_one(inv_doc)
            activated_count += 1
        elif not existing_inv.get('is_active', False):
            # Re-activate if inactive
            await db.branch_rm_inventory.update_one(
                {"rm_id": rm_id, "branch": branch},
                {"$set": {"is_active": True}}
            )
            activated_count += 1
    
    return activated_count

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    """Get current authenticated user from JWT token"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Could not validate credentials")
    
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user_doc is None:
        raise HTTPException(status_code=401, detail="User not found")
    
    if not user_doc.get("is_active", True):
        raise HTTPException(status_code=403, detail="User account is deactivated")
    
    return User(**serialize_doc(user_doc))

def check_master_admin(user: User):
    """Check if user is master admin"""
    if user.role != "master_admin":
        raise HTTPException(status_code=403, detail="Only master admin can perform this action")

def check_branch_access(user: User, branch: str):
    """Check if user has access to specific branch"""
    if user.role == "master_admin":
        return True
    if branch not in user.assigned_branches:
        raise HTTPException(status_code=403, detail=f"No access to branch: {branch}")
    return True

def serialize_doc(doc):
    if doc and 'created_at' in doc and isinstance(doc['created_at'], str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    if doc and 'date' in doc and isinstance(doc['date'], str):
        doc['date'] = datetime.fromisoformat(doc['date'])
    if doc and 'activated_at' in doc and isinstance(doc['activated_at'], str):
        doc['activated_at'] = datetime.fromisoformat(doc['activated_at'])
    return doc

async def get_next_rm_sequence(category: str) -> int:
    """Get next global sequence number for RM category by finding the highest numeric suffix"""
    import re
    all_rms = await db.raw_materials.find(
        {"category": category},
        {"_id": 0, "rm_id": 1}
    ).to_list(10000)
    
    max_seq = 0
    pattern = re.compile(rf'^{category}_(\d+)$')
    
    for rm in all_rms:
        match = pattern.match(rm['rm_id'])
        if match:
            seq = int(match.group(1))
            max_seq = max(max_seq, seq)
    
    return max_seq + 1

# ============ L1/L2 Consumption Engine ============

async def generate_movement_code() -> str:
    """Generate unique movement code"""
    import random
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    random_suffix = random.randint(1000, 9999)
    return f"MV_{timestamp}_{random_suffix}"

async def get_branch_rm_stock(branch: str, rm_id: str) -> float:
    """Get current stock of an RM in a branch"""
    inv = await db.branch_rm_inventory.find_one({"branch": branch, "rm_id": rm_id})
    if inv:
        return inv.get("current_stock", 0)
    return 0

async def get_current_rm_price(rm_id: str, branch: str = None) -> float:
    """Get current price of an RM (from vendor pricing or default)"""
    # Get lowest vendor price
    prices = await db.vendor_rm_prices.find({"rm_id": rm_id}).to_list(100)
    if prices:
        return min(p.get("price", 0) for p in prices)
    return 0

async def update_branch_rm_inventory(branch: str, rm_id: str, quantity_change: float):
    """Update branch RM inventory by adding/subtracting quantity"""
    inv = await db.branch_rm_inventory.find_one({"branch": branch, "rm_id": rm_id})
    if inv:
        new_stock = inv.get("current_stock", 0) + quantity_change
        await db.branch_rm_inventory.update_one(
            {"branch": branch, "rm_id": rm_id},
            {"$set": {"current_stock": new_stock}}
        )
    else:
        # Create new inventory record
        await db.branch_rm_inventory.insert_one({
            "id": str(uuid.uuid4()),
            "rm_id": rm_id,
            "branch": branch,
            "current_stock": quantity_change,
            "is_active": True,
            "activated_at": datetime.now(timezone.utc)
        })

async def consume_inp_l2_material(
    branch: str,
    rm_id: str,
    quantity: int,
    production_batch_id: str,
    user_id: str
) -> dict:
    """
    INP L2 consumption: Weight-based L1 deduction
    """
    l2_rm = await db.raw_materials.find_one({"rm_id": rm_id})
    
    if not l2_rm or l2_rm.get("category") != "INP" or l2_rm.get("rm_level") != "L2":
        raise HTTPException(status_code=400, detail="This function is only for INP L2 materials")
    
    l1_rm_id = l2_rm.get("parent_rm_id")
    if not l1_rm_id:
        raise HTTPException(status_code=400, detail=f"INP L2 {rm_id} missing polymer L1 reference")
    
    # Calculate L1 consumption (weight-based)
    unit_weight_kg = (l2_rm.get("unit_weight_grams") or 0) / 1000
    scrap_factor = l2_rm.get("scrap_factor") or 0.02
    l1_consumption = quantity * unit_weight_kg * (1 + scrap_factor)
    
    # Check stock
    l1_stock = await get_branch_rm_stock(branch, l1_rm_id)
    if l1_stock < l1_consumption:
        raise HTTPException(
            status_code=400,
            detail=f"Need {l1_consumption:.3f} KG of {l1_rm_id}, only {l1_stock:.3f} available"
        )
    
    # Get price and calculate cost
    l1_price = await get_current_rm_price(l1_rm_id, branch)
    processing_cost = l2_rm.get("processing_cost") or 0
    l2_unit_cost = (unit_weight_kg * l1_price * (1 + scrap_factor)) + processing_cost
    
    # Create L1 consumption movement
    movement_code = await generate_movement_code()
    await db.rm_stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "movement_code": movement_code,
        "rm_id": l1_rm_id,
        "branch_id": "",
        "branch": branch,
        "movement_type": "CONSUMPTION",
        "quantity": -l1_consumption,
        "unit_of_measure": "KG",
        "reference_type": "PRODUCTION_BATCH",
        "reference_id": production_batch_id,
        "unit_cost": l1_price,
        "total_cost": l1_consumption * l1_price,
        "balance_after": l1_stock - l1_consumption,
        "notes": f"Polymer for {rm_id} x {quantity}",
        "created_at": datetime.now(timezone.utc),
        "created_by": user_id
    })
    
    # Update inventory
    await update_branch_rm_inventory(branch, l1_rm_id, -l1_consumption)
    
    # Create L2 production movement
    l2_stock = await get_branch_rm_stock(branch, rm_id)
    movement_code2 = await generate_movement_code()
    await db.rm_stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "movement_code": movement_code2,
        "rm_id": rm_id,
        "branch_id": "",
        "branch": branch,
        "movement_type": "PRODUCTION",
        "quantity": quantity,
        "unit_of_measure": "PCS",
        "reference_type": "PRODUCTION_BATCH",
        "reference_id": production_batch_id,
        "l1_rm_id": l1_rm_id,
        "l1_quantity_consumed": l1_consumption,
        "unit_cost": l2_unit_cost,
        "total_cost": quantity * l2_unit_cost,
        "balance_after": l2_stock + quantity,
        "notes": f"Polymer consumed: {l1_consumption:.4f} KG",
        "created_at": datetime.now(timezone.utc),
        "created_by": user_id
    })
    
    await update_branch_rm_inventory(branch, rm_id, quantity)
    
    return {
        "l2_rm_id": rm_id,
        "quantity_produced": quantity,
        "polymer_consumed": {
            "rm_id": l1_rm_id,
            "quantity_kg": round(l1_consumption, 4),
            "total_cost": round(l1_consumption * l1_price, 2)
        },
        "l2_unit_cost": round(l2_unit_cost, 2),
        "total_batch_cost": round(quantity * l2_unit_cost, 2)
    }

async def consume_inm_l2_material(
    branch: str,
    rm_id: str,
    quantity: int,
    production_batch_id: str,
    user_id: str
) -> dict:
    """
    INM L2 consumption:
    1. Base metal: 1:1 ratio (L2_Qty = L1_Qty consumed)
    2. Powder coating: predefined grams per unit × qty × scrap factor
    """
    l2_rm = await db.raw_materials.find_one({"rm_id": rm_id})
    
    if not l2_rm or l2_rm.get("category") != "INM" or l2_rm.get("rm_level") != "L2":
        raise HTTPException(status_code=400, detail="This function is only for INM L2 materials")
    
    base_metal_rm_id = l2_rm.get("parent_rm_id")
    powder_coating_rm_id = l2_rm.get("secondary_l1_rm_id")
    
    if not base_metal_rm_id:
        raise HTTPException(status_code=400, detail=f"INM L2 {rm_id} missing base metal L1 reference")
    if not powder_coating_rm_id:
        raise HTTPException(status_code=400, detail=f"INM L2 {rm_id} missing powder coating L1 reference")
    
    # Calculate Base Metal consumption (1:1 ratio)
    metal_consumption = quantity
    
    # Calculate Powder Coating consumption
    powder_qty_grams = l2_rm.get("powder_qty_grams") or 0
    if powder_qty_grams <= 0:
        raise HTTPException(status_code=400, detail=f"INM L2 {rm_id} missing predefined powder_qty_grams")
    
    coating_scrap_factor = l2_rm.get("coating_scrap_factor") or 0.10
    coating_consumption_kg = (quantity * powder_qty_grams / 1000) * (1 + coating_scrap_factor)
    
    # Check stock availability
    metal_stock = await get_branch_rm_stock(branch, base_metal_rm_id)
    coating_stock = await get_branch_rm_stock(branch, powder_coating_rm_id)
    
    if metal_stock < metal_consumption:
        raise HTTPException(
            status_code=400,
            detail=f"Need {metal_consumption} units of {base_metal_rm_id}, only {metal_stock} available"
        )
    if coating_stock < coating_consumption_kg:
        raise HTTPException(
            status_code=400,
            detail=f"Need {coating_consumption_kg:.4f} KG of {powder_coating_rm_id}, only {coating_stock:.4f} available"
        )
    
    # Get costs
    l1_unit_cost = await get_current_rm_price(base_metal_rm_id, branch)
    coating_price_per_kg = await get_current_rm_price(powder_coating_rm_id, branch)
    
    # Calculate L2 Unit Cost
    coating_cost_per_unit = (powder_qty_grams / 1000) * coating_price_per_kg * (1 + coating_scrap_factor)
    processing_cost = l2_rm.get("processing_cost") or 0
    l2_unit_cost = l1_unit_cost + coating_cost_per_unit + processing_cost
    
    # Create Base Metal consumption movement
    movement_code1 = await generate_movement_code()
    await db.rm_stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "movement_code": movement_code1,
        "rm_id": base_metal_rm_id,
        "branch_id": "",
        "branch": branch,
        "movement_type": "CONSUMPTION",
        "quantity": -metal_consumption,
        "unit_of_measure": "PCS",
        "reference_type": "PRODUCTION_BATCH",
        "reference_id": production_batch_id,
        "unit_cost": l1_unit_cost,
        "total_cost": metal_consumption * l1_unit_cost,
        "balance_after": metal_stock - metal_consumption,
        "notes": f"Base metal for {rm_id} x {quantity} (1:1)",
        "created_at": datetime.now(timezone.utc),
        "created_by": user_id
    })
    
    # Create Powder Coating consumption movement
    movement_code2 = await generate_movement_code()
    await db.rm_stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "movement_code": movement_code2,
        "rm_id": powder_coating_rm_id,
        "branch_id": "",
        "branch": branch,
        "movement_type": "CONSUMPTION",
        "quantity": -coating_consumption_kg,
        "unit_of_measure": "KG",
        "reference_type": "PRODUCTION_BATCH",
        "reference_id": production_batch_id,
        "unit_cost": coating_price_per_kg,
        "total_cost": coating_consumption_kg * coating_price_per_kg,
        "balance_after": coating_stock - coating_consumption_kg,
        "notes": f"Powder coating for {rm_id} x {quantity} @ {powder_qty_grams}g each",
        "created_at": datetime.now(timezone.utc),
        "created_by": user_id
    })
    
    # Create L2 production movement
    l2_stock = await get_branch_rm_stock(branch, rm_id)
    movement_code3 = await generate_movement_code()
    await db.rm_stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "movement_code": movement_code3,
        "rm_id": rm_id,
        "branch_id": "",
        "branch": branch,
        "movement_type": "PRODUCTION",
        "quantity": quantity,
        "unit_of_measure": "PCS",
        "reference_type": "PRODUCTION_BATCH",
        "reference_id": production_batch_id,
        "l1_rm_id": base_metal_rm_id,
        "l1_quantity_consumed": metal_consumption,
        "unit_cost": l2_unit_cost,
        "total_cost": quantity * l2_unit_cost,
        "balance_after": l2_stock + quantity,
        "notes": f"L1: {metal_consumption} pcs, Coating: {coating_consumption_kg:.4f} KG",
        "created_at": datetime.now(timezone.utc),
        "created_by": user_id
    })
    
    # Update inventory balances
    await update_branch_rm_inventory(branch, base_metal_rm_id, -metal_consumption)
    await update_branch_rm_inventory(branch, powder_coating_rm_id, -coating_consumption_kg)
    await update_branch_rm_inventory(branch, rm_id, quantity)
    
    return {
        "l2_rm_id": rm_id,
        "quantity_produced": quantity,
        "base_metal_consumed": {
            "rm_id": base_metal_rm_id,
            "quantity": metal_consumption,
            "unit": "PCS",
            "unit_cost": round(l1_unit_cost, 2),
            "total_cost": round(metal_consumption * l1_unit_cost, 2)
        },
        "powder_coating_consumed": {
            "rm_id": powder_coating_rm_id,
            "quantity_kg": round(coating_consumption_kg, 4),
            "grams_per_unit": powder_qty_grams,
            "unit_cost_per_kg": round(coating_price_per_kg, 2),
            "total_cost": round(coating_consumption_kg * coating_price_per_kg, 2)
        },
        "l2_unit_cost_breakdown": {
            "l1_unit_cost": round(l1_unit_cost, 2),
            "coating_cost_per_unit": round(coating_cost_per_unit, 2),
            "processing_cost": round(processing_cost, 2),
            "total_l2_unit_cost": round(l2_unit_cost, 2)
        },
        "total_batch_cost": round(quantity * l2_unit_cost, 2)
    }

async def consume_l2_material(
    branch: str,
    rm_id: str,
    quantity: int,
    production_batch_id: str,
    user_id: str
) -> dict:
    """
    Universal L2 consumption function - routes to appropriate handler based on category
    """
    l2_rm = await db.raw_materials.find_one({"rm_id": rm_id})
    
    if not l2_rm:
        raise HTTPException(status_code=404, detail=f"RM {rm_id} not found")
    
    if l2_rm.get("rm_level") != "L2":
        raise HTTPException(status_code=400, detail=f"RM {rm_id} is not an L2 material")
    
    category = l2_rm.get("category")
    
    if category == "INP":
        return await consume_inp_l2_material(branch, rm_id, quantity, production_batch_id, user_id)
    elif category == "INM":
        return await consume_inm_l2_material(branch, rm_id, quantity, production_batch_id, user_id)
    else:
        raise HTTPException(status_code=400, detail=f"L2 consumption not supported for category {category}")

# ============ Authentication Routes ============

@api_router.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    """User login"""
    user_doc = await db.users.find_one({"email": request.email}, {"_id": 0})
    
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not verify_password(request.password, user_doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user_doc.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account is deactivated")
    
    # Create access token
    access_token = create_access_token(data={"sub": user_doc["id"]})
    
    user_response = UserResponse(
        id=user_doc["id"],
        email=user_doc["email"],
        name=user_doc["name"],
        role=user_doc["role"],
        assigned_branches=user_doc.get("assigned_branches", []),
        is_active=user_doc.get("is_active", True),
        created_at=user_doc["created_at"]
    )
    
    return LoginResponse(
        access_token=access_token,
        token_type="bearer",
        user=user_response
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    """Get current user information"""
    return UserResponse(
        id=current_user.id,
        email=current_user.email,
        name=current_user.name,
        role=current_user.role,
        assigned_branches=current_user.assigned_branches,
        is_active=current_user.is_active,
        created_at=current_user.created_at
    )

@api_router.post("/auth/change-password")
async def change_password(
    request: ChangePasswordRequest,
    current_user: User = Depends(get_current_user)
):
    """Change user password"""
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    
    if not verify_password(request.current_password, user_doc["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    new_password_hash = hash_password(request.new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"password_hash": new_password_hash}}
    )
    
    return {"message": "Password changed successfully"}

# ============ User Management Routes (Master Admin Only) ============

@api_router.post("/users", response_model=UserResponse)
async def create_user(
    user_data: UserCreate,
    current_user: User = Depends(get_current_user)
):
    """Create new user (Master Admin only)"""
    check_master_admin(current_user)
    
    # Check if user already exists
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="User with this email already exists")
    
    # Validate role
    if user_data.role not in ["master_admin", "branch_user"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Validate branches
    if user_data.role == "branch_user" and not user_data.assigned_branches:
        raise HTTPException(status_code=400, detail="Branch user must have at least one assigned branch")
    
    # Create user
    user_obj = User(
        email=user_data.email,
        password_hash=hash_password(user_data.password),
        name=user_data.name,
        role=user_data.role,
        assigned_branches=user_data.assigned_branches if user_data.role == "branch_user" else []
    )
    
    doc = user_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    return UserResponse(
        id=user_obj.id,
        email=user_obj.email,
        name=user_obj.name,
        role=user_obj.role,
        assigned_branches=user_obj.assigned_branches,
        is_active=user_obj.is_active,
        created_at=user_obj.created_at
    )

@api_router.get("/users", response_model=List[UserResponse])
async def list_users(current_user: User = Depends(get_current_user)):
    """List all users (Master Admin only)"""
    check_master_admin(current_user)
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return [UserResponse(**serialize_doc(u)) for u in users]

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: str,
    user_data: UserCreate,
    current_user: User = Depends(get_current_user)
):
    """Update user (Master Admin only)"""
    check_master_admin(current_user)
    
    existing = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {
        "name": user_data.name,
        "role": user_data.role,
        "assigned_branches": user_data.assigned_branches if user_data.role == "branch_user" else []
    }
    
    # Update password if provided
    if user_data.password:
        update_data["password_hash"] = hash_password(user_data.password)
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    updated = await db.users.find_one({"id": user_id}, {"_id": 0})
    return UserResponse(**serialize_doc(updated))

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Delete user (Master Admin only)"""
    check_master_admin(current_user)
    
    # Can't delete yourself
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

@api_router.patch("/users/{user_id}/toggle-active")
async def toggle_user_active(user_id: str, current_user: User = Depends(get_current_user)):
    """Toggle user active status (Master Admin only)"""
    check_master_admin(current_user)
    
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_status = not user_doc.get("is_active", True)
    await db.users.update_one({"id": user_id}, {"$set": {"is_active": new_status}})
    
    return {"message": f"User {'activated' if new_status else 'deactivated'} successfully"}

# ============ Initialize Default Admin ============

@app.on_event("startup")
async def create_default_admin():
    """Create default master admin if no users exist"""
    user_count = await db.users.count_documents({})
    if user_count == 0:
        default_admin = User(
            email="admin@factory.com",
            password_hash=hash_password("admin123"),
            name="Master Admin",
            role="master_admin",
            assigned_branches=[],
            is_active=True
        )
        doc = default_admin.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        print("Default admin created - Email: admin@factory.com, Password: admin123")

# ============ Branch Routes ============

@api_router.get("/branches")
async def get_branches():
    return {"branches": BRANCHES}

@api_router.get("/rm-categories")
async def get_rm_categories():
    return {"categories": RM_CATEGORIES}

# ============ Global Raw Material Routes ============

@api_router.post("/raw-materials", response_model=RawMaterial)
async def create_raw_material(input: RawMaterialCreate):
    """Create global RM with auto-generated ID"""
    # Auto-generate RM ID based on category
    seq = await get_next_rm_sequence(input.category)
    rm_id = f"{input.category}_{seq:03d}"
    
    # Check if somehow this ID exists (shouldn't happen)
    existing = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
    if existing:
        # Get next available
        seq = await get_next_rm_sequence(input.category)
        rm_id = f"{input.category}_{seq:03d}"
    
    rm_obj = RawMaterial(
        rm_id=rm_id,
        category=input.category,
        category_data=input.category_data,
        low_stock_threshold=input.low_stock_threshold
    )
    doc = rm_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.raw_materials.insert_one(doc)
    return rm_obj

@api_router.post("/raw-materials/bulk-upload")
async def bulk_upload_raw_materials(file: UploadFile = File(...)):
    """Bulk upload global RMs"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            
            try:
                category = str(row[0]).strip().upper()
                if category not in RM_CATEGORIES:
                    errors.append(f"Row {idx}: Invalid category {category}")
                    continue
                
                seq = await get_next_rm_sequence(category)
                rm_id = f"{category}_{seq:03d}"
                
                existing = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
                if existing:
                    skipped_count += 1
                    continue
                
                category_fields = RM_CATEGORIES[category]["fields"]
                category_data = {}
                for i, field in enumerate(category_fields):
                    value = row[i + 1] if len(row) > i + 1 else ""
                    category_data[field] = str(value) if value else ""
                
                threshold = float(row[len(category_fields) + 1]) if len(row) > len(category_fields) + 1 and row[len(category_fields) + 1] else 10.0
                
                rm_obj = RawMaterial(
                    rm_id=rm_id,
                    category=category,
                    category_data=category_data,
                    low_stock_threshold=threshold
                )
                doc = rm_obj.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.raw_materials.insert_one(doc)
                created_count += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        return {
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.post("/raw-materials/import-with-ids")
async def import_raw_materials_with_ids(file: UploadFile = File(...), category: str = ""):
    """Import RMs from Excel where RM IDs are provided in the first column.
    Skips duplicates, preserves the ID sequence for future uploads.
    File format: RM_ID in first column, then category-specific fields.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    if not category or category.upper() not in RM_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Invalid category. Must be one of: {list(RM_CATEGORIES.keys())}")
    
    category = category.upper()
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        # Get headers from first row
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
        
        created_count = 0
        skipped_count = 0
        updated_count = 0
        errors = []
        
        category_fields = RM_CATEGORIES[category]["fields"]
        
        # Build mapping from headers to category fields
        field_mapping = {}
        header_to_field = {
            'rm code': 'rm_id', 'raw material id': 'rm_id', 'rm_id': 'rm_id', 'rm id': 'rm_id',
            'type': 'type', 'model': 'model', 'model name': 'model_name', 'mould code': 'mould_code',
            'part name': 'part_name', 'colour': 'colour', 'color': 'colour', 'mb': 'mb',
            'specs': 'specs', 'brand': 'brand', 'buyer sku': 'buyer_sku', 'buyer_sku': 'buyer_sku',
            'per unit weight (in grams)': 'per_unit_weight', 'per unit weight': 'per_unit_weight',
            'per unit weight (g)': 'per_unit_weight', 'unit': 'unit', 'position': 'position'
        }
        
        rm_id_col = None
        for i, h in enumerate(headers):
            mapped = header_to_field.get(h)
            if mapped == 'rm_id':
                rm_id_col = i
            elif mapped in category_fields:
                field_mapping[mapped] = i
        
        if rm_id_col is None:
            raise HTTPException(status_code=400, detail="Could not find RM ID column. Expected 'RM Code' or 'Raw Material Id'.")
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[rm_id_col]:
                continue
            
            try:
                rm_id = str(row[rm_id_col]).strip()
                
                # Validate RM ID format matches category
                if not rm_id.startswith(f"{category}_"):
                    errors.append(f"Row {idx}: RM ID '{rm_id}' doesn't match category {category}")
                    continue
                
                # Build category_data
                category_data = {}
                for field in category_fields:
                    col_idx = field_mapping.get(field)
                    if col_idx is not None and col_idx < len(row):
                        val = row[col_idx]
                        category_data[field] = val if val is not None else ""
                    else:
                        category_data[field] = ""
                
                # Check if RM exists
                existing = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
                if existing:
                    skipped_count += 1
                    continue
                
                rm_obj = RawMaterial(
                    rm_id=rm_id,
                    category=category,
                    category_data=category_data,
                    low_stock_threshold=10.0
                )
                doc = rm_obj.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.raw_materials.insert_one(doc)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        return {
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors[:20] if errors else [],
            "total_errors": len(errors)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/raw-materials")
async def get_raw_materials(branch: Optional[str] = None, search: Optional[str] = None, include_inactive: bool = False):
    """Get RMs - if branch specified, return only active RMs in that branch"""
    if branch:
        # Get active RMs in branch
        inventory_query = {"branch": branch}
        if not include_inactive:
            inventory_query["is_active"] = True
        
        branch_inventories = await db.branch_rm_inventory.find(inventory_query, {"_id": 0}).to_list(1000)
        active_rm_ids = [inv['rm_id'] for inv in branch_inventories]
        
        query = {"rm_id": {"$in": active_rm_ids}}
        if search:
            query["$and"] = [{"rm_id": {"$in": active_rm_ids}}, {"rm_id": {"$regex": search, "$options": "i"}}]
        
        materials = await db.raw_materials.find(query, {"_id": 0}).to_list(1000)
        
        # Merge with inventory data
        result = []
        for mat in materials:
            inv = next((i for i in branch_inventories if i['rm_id'] == mat['rm_id']), None)
            mat['current_stock'] = inv['current_stock'] if inv else 0
            mat['branch'] = branch
            result.append(serialize_doc(mat))
        return result
    else:
        # Get all global RMs
        query = {}
        if search:
            query["rm_id"] = {"$regex": search, "$options": "i"}
        materials = await db.raw_materials.find(query, {"_id": 0}).to_list(1000)
        return [serialize_doc(m) for m in materials]

@api_router.post("/raw-materials/activate")
async def activate_rm_in_branch(request: ActivateItemRequest):
    """Activate an RM in a specific branch"""
    rm = await db.raw_materials.find_one({"rm_id": request.item_id}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail="RM not found globally")
    
    existing_inv = await db.branch_rm_inventory.find_one(
        {"rm_id": request.item_id, "branch": request.branch},
        {"_id": 0}
    )
    
    if existing_inv:
        if existing_inv['is_active']:
            return {"message": "RM already active in this branch"}
        # Reactivate
        await db.branch_rm_inventory.update_one(
            {"rm_id": request.item_id, "branch": request.branch},
            {"$set": {"is_active": True, "activated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        # Create new inventory entry
        inv_obj = BranchRMInventory(rm_id=request.item_id, branch=request.branch)
        doc = inv_obj.model_dump()
        doc['activated_at'] = doc['activated_at'].isoformat()
        await db.branch_rm_inventory.insert_one(doc)
    
    return {"message": f"RM {request.item_id} activated in {request.branch}"}

@api_router.delete("/raw-materials/{rm_id}")
async def delete_raw_material(rm_id: str):
    """Delete global RM"""
    result = await db.raw_materials.delete_one({"rm_id": rm_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    # Also delete from all branch inventories
    await db.branch_rm_inventory.delete_many({"rm_id": rm_id})
    return {"message": "Raw material deleted globally"}

# ============ Purchase Entry Routes ============

@api_router.post("/purchase-entries")
async def create_purchase_entry(input: PurchaseEntryCreate):
    rm = await db.raw_materials.find_one({"rm_id": input.rm_id}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    # Check if RM is active in branch, auto-activate if not
    inventory = await db.branch_rm_inventory.find_one(
        {"rm_id": input.rm_id, "branch": input.branch},
        {"_id": 0}
    )
    
    if not inventory:
        # Auto-activate RM in this branch
        inv_obj = BranchRMInventory(rm_id=input.rm_id, branch=input.branch)
        doc = inv_obj.model_dump()
        doc['activated_at'] = doc['activated_at'].isoformat()
        await db.branch_rm_inventory.insert_one(doc)
    elif not inventory.get('is_active', True):
        # Reactivate if deactivated
        await db.branch_rm_inventory.update_one(
            {"rm_id": input.rm_id, "branch": input.branch},
            {"$set": {"is_active": True, "activated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    entry_obj = PurchaseEntry(**input.model_dump())
    doc = entry_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['date'] = doc['date'].isoformat()
    await db.purchase_entries.insert_one(doc)
    
    await db.branch_rm_inventory.update_one(
        {"rm_id": input.rm_id, "branch": input.branch},
        {"$inc": {"current_stock": input.quantity}}
    )
    
    return entry_obj

@api_router.get("/purchase-entries")
async def get_purchase_entries(branch: Optional[str] = None, rm_id: Optional[str] = None):
    query = {}
    if branch:
        query["branch"] = branch
    if rm_id:
        query["rm_id"] = rm_id
    
    entries = await db.purchase_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return [serialize_doc(e) for e in entries]

# ============ Global SKU Routes ============

@api_router.post("/skus", response_model=SKU)
async def create_sku(input: SKUCreate):
    """Create global SKU"""
    existing = await db.skus.find_one({"sku_id": input.sku_id}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="SKU ID already exists globally")
    
    sku_obj = SKU(**input.model_dump())
    doc = sku_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.skus.insert_one(doc)
    return sku_obj

@api_router.get("/skus")
async def get_skus(branch: Optional[str] = None, search: Optional[str] = None, include_inactive: bool = False):
    """Get SKUs - if branch specified, return only SKUs subscribed to that branch"""
    if branch:
        # Get SKUs subscribed to this branch via sku_branch_assignments
        assignments = await db.sku_branch_assignments.find({"branch": branch}, {"_id": 0}).to_list(10000)
        subscribed_sku_ids = [a['sku_id'] for a in assignments]
        
        if not subscribed_sku_ids:
            return []
        
        query = {"sku_id": {"$in": subscribed_sku_ids}}
        if search:
            query = {
                "$and": [
                    {"sku_id": {"$in": subscribed_sku_ids}},
                    {"$or": [
                        {"sku_id": {"$regex": search, "$options": "i"}},
                        {"bidso_sku": {"$regex": search, "$options": "i"}},
                        {"buyer_sku_id": {"$regex": search, "$options": "i"}}
                    ]}
                ]
            }
        
        skus = await db.skus.find(query, {"_id": 0}).to_list(10000)
        
        # Get inventory data for stock levels
        branch_inventories = await db.branch_sku_inventory.find(
            {"branch": branch, "sku_id": {"$in": subscribed_sku_ids}}, 
            {"_id": 0}
        ).to_list(10000)
        inv_map = {inv['sku_id']: inv for inv in branch_inventories}
        
        result = []
        for sku in skus:
            inv = inv_map.get(sku['sku_id'])
            sku['current_stock'] = inv['current_stock'] if inv else 0
            sku['branch'] = branch
            result.append(serialize_doc(sku))
        return result
    else:
        query = {}
        if search:
            query["$or"] = [
                {"sku_id": {"$regex": search, "$options": "i"}},
                {"bidso_sku": {"$regex": search, "$options": "i"}},
                {"buyer_sku_id": {"$regex": search, "$options": "i"}}
            ]
        skus = await db.skus.find(query, {"_id": 0}).to_list(10000)
        return [serialize_doc(s) for s in skus]

@api_router.get("/skus/unmapped")
async def get_skus_without_rm_mapping():
    """Get all SKUs that don't have any RM mapping (BOM not defined)"""
    # Get all SKU IDs
    all_skus = await db.skus.find({}, {"_id": 0}).to_list(10000)
    all_sku_ids = set(s['sku_id'] for s in all_skus)
    
    # Get all SKU IDs that have mappings
    mappings = await db.sku_rm_mapping.find({}, {"_id": 0, "sku_id": 1}).to_list(50000)
    mapped_sku_ids = set(m['sku_id'] for m in mappings)
    
    # Find unmapped SKUs
    unmapped_sku_ids = all_sku_ids - mapped_sku_ids
    
    # Get full details of unmapped SKUs
    unmapped_skus = [s for s in all_skus if s['sku_id'] in unmapped_sku_ids]
    
    return {
        "count": len(unmapped_skus),
        "skus": [serialize_doc(s) for s in unmapped_skus]
    }

@api_router.post("/skus/activate")
async def activate_sku_in_branch(request: ActivateItemRequest):
    """Activate SKU in branch and auto-activate its BOM RMs"""
    sku = await db.skus.find_one({"sku_id": request.item_id}, {"_id": 0})
    if not sku:
        raise HTTPException(status_code=404, detail="SKU not found globally")
    
    # Check if already active
    existing_inv = await db.branch_sku_inventory.find_one(
        {"sku_id": request.item_id, "branch": request.branch},
        {"_id": 0}
    )
    
    if existing_inv:
        if existing_inv['is_active']:
            return {"message": "SKU already active in this branch"}
        await db.branch_sku_inventory.update_one(
            {"sku_id": request.item_id, "branch": request.branch},
            {"$set": {"is_active": True, "activated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        inv_obj = BranchSKUInventory(sku_id=request.item_id, branch=request.branch)
        doc = inv_obj.model_dump()
        doc['activated_at'] = doc['activated_at'].isoformat()
        await db.branch_sku_inventory.insert_one(doc)
    
    # Auto-activate BOM RMs
    mapping = await db.sku_mappings.find_one({"sku_id": request.item_id}, {"_id": 0})
    activated_rms = []
    if mapping:
        for rm_mapping in mapping['rm_mappings']:
            rm_id = rm_mapping['rm_id']
            # Check if RM exists globally
            rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
            if rm:
                # Activate in branch if not already
                existing_rm_inv = await db.branch_rm_inventory.find_one(
                    {"rm_id": rm_id, "branch": request.branch},
                    {"_id": 0}
                )
                if not existing_rm_inv:
                    rm_inv_obj = BranchRMInventory(rm_id=rm_id, branch=request.branch)
                    doc = rm_inv_obj.model_dump()
                    doc['activated_at'] = doc['activated_at'].isoformat()
                    await db.branch_rm_inventory.insert_one(doc)
                    activated_rms.append(rm_id)
                elif not existing_rm_inv['is_active']:
                    await db.branch_rm_inventory.update_one(
                        {"rm_id": rm_id, "branch": request.branch},
                        {"$set": {"is_active": True, "activated_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    activated_rms.append(rm_id)
    
    return {
        "message": f"SKU {request.item_id} activated in {request.branch}",
        "auto_activated_rms": activated_rms
    }

@api_router.put("/skus/{sku_id}", response_model=SKU)
async def update_sku(sku_id: str, input: SKUCreate):
    existing = await db.skus.find_one({"sku_id": sku_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="SKU not found")
    
    update_data = input.model_dump()
    await db.skus.update_one({"sku_id": sku_id}, {"$set": update_data})
    
    updated = await db.skus.find_one({"sku_id": sku_id}, {"_id": 0})
    return serialize_doc(updated)

@api_router.delete("/skus/{sku_id}")
async def delete_sku(sku_id: str):
    result = await db.skus.delete_one({"sku_id": sku_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="SKU not found")
    await db.branch_sku_inventory.delete_many({"sku_id": sku_id})
    return {"message": "SKU deleted globally"}

# ============ SKU Mapping Routes (Global BOM) ============

@api_router.post("/sku-mappings", response_model=SKUMapping)
async def create_sku_mapping(input: SKUMappingCreate):
    sku = await db.skus.find_one({"sku_id": input.sku_id}, {"_id": 0})
    if not sku:
        raise HTTPException(status_code=404, detail="SKU not found")
    
    for mapping in input.rm_mappings:
        rm = await db.raw_materials.find_one({"rm_id": mapping.rm_id}, {"_id": 0})
        if not rm:
            raise HTTPException(status_code=404, detail=f"Raw material {mapping.rm_id} not found globally")
    
    existing = await db.sku_mappings.find_one({"sku_id": input.sku_id}, {"_id": 0})
    if existing:
        await db.sku_mappings.delete_one({"sku_id": input.sku_id})
    
    mapping_obj = SKUMapping(**input.model_dump())
    doc = mapping_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.sku_mappings.insert_one(doc)
    return mapping_obj

@api_router.post("/sku-mappings/bulk-upload")
async def bulk_upload_sku_mappings(file: UploadFile = File(...)):
    """Bulk upload SKU to RM mappings via Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        # Group mappings by SKU
        sku_mappings_dict = {}  # {sku_id: [{"rm_id": "...", "quantity_required": ...}]}
        
        created_count = 0
        updated_count = 0
        errors = []
        
        # Expected format: SKU_ID, RM_ID, Qty
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not row[1]:
                continue
            
            try:
                sku_id = str(row[0]).strip()
                rm_id = str(row[1]).strip()
                qty = float(row[2]) if row[2] else 0
                
                if qty <= 0:
                    errors.append(f"Row {idx}: Invalid quantity for {sku_id} - {rm_id}")
                    continue
                
                # Verify SKU exists
                sku = await db.skus.find_one({"sku_id": sku_id}, {"_id": 0})
                if not sku:
                    errors.append(f"Row {idx}: SKU {sku_id} not found")
                    continue
                
                # Verify RM exists
                rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
                if not rm:
                    errors.append(f"Row {idx}: RM {rm_id} not found")
                    continue
                
                # Add to mapping dictionary
                if sku_id not in sku_mappings_dict:
                    sku_mappings_dict[sku_id] = []
                
                sku_mappings_dict[sku_id].append({
                    "rm_id": rm_id,
                    "quantity_required": qty
                })
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        # Create/update mappings for each SKU
        for sku_id, rm_mappings in sku_mappings_dict.items():
            try:
                existing = await db.sku_mappings.find_one({"sku_id": sku_id}, {"_id": 0})
                
                mapping_obj = SKUMapping(
                    sku_id=sku_id,
                    rm_mappings=[RMMapping(**rm) for rm in rm_mappings]
                )
                doc = mapping_obj.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                
                if existing:
                    # Replace existing mapping
                    await db.sku_mappings.delete_one({"sku_id": sku_id})
                    await db.sku_mappings.insert_one(doc)
                    updated_count += 1
                else:
                    await db.sku_mappings.insert_one(doc)
                    created_count += 1
                    
            except Exception as e:
                errors.append(f"SKU {sku_id}: {str(e)}")
        
        return {
            "created": created_count,
            "updated": updated_count,
            "total_skus": len(sku_mappings_dict),
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/sku-mappings/{sku_id}", response_model=SKUMapping)
async def get_sku_mapping(sku_id: str):
    # First check old sku_mappings collection
    mapping = await db.sku_mappings.find_one({"sku_id": sku_id}, {"_id": 0})
    if mapping:
        return serialize_doc(mapping)
    
    # Check new sku_rm_mapping collection
    mappings = await db.sku_rm_mapping.find({"sku_id": sku_id}, {"_id": 0}).to_list(100)
    if not mappings:
        raise HTTPException(status_code=404, detail="Mapping not found")
    
    # Convert to expected format
    return {
        "id": mappings[0].get('id', ''),
        "sku_id": sku_id,
        "rm_mappings": [{"rm_id": m['rm_id'], "quantity_required": m.get('quantity', 1)} for m in mappings]
    }

@api_router.get("/sku-mappings", response_model=List[SKUMapping])
async def get_all_sku_mappings():
    # Get from old sku_mappings collection
    old_mappings = await db.sku_mappings.find({}, {"_id": 0}).to_list(1000)
    
    # Get from new sku_rm_mapping collection and group by sku_id
    new_mappings = await db.sku_rm_mapping.find({}, {"_id": 0}).to_list(50000)
    
    # Group new mappings by sku_id
    grouped = {}
    for m in new_mappings:
        sku_id = m['sku_id']
        if sku_id not in grouped:
            grouped[sku_id] = {
                "id": m.get('id', ''),
                "sku_id": sku_id,
                "rm_mappings": []
            }
        grouped[sku_id]['rm_mappings'].append({
            "rm_id": m['rm_id'],
            "quantity_required": m.get('quantity', 1)
        })
    
    # Combine both sources (avoid duplicates by sku_id)
    result_dict = {m['sku_id']: serialize_doc(m) for m in old_mappings}
    for sku_id, mapping in grouped.items():
        if sku_id not in result_dict:
            result_dict[sku_id] = mapping
    
    return list(result_dict.values())

# ============ Production Entry Routes ============

@api_router.post("/production-entries")
async def create_production_entry(input: ProductionEntryCreate):
    # Check SKU is active in branch
    sku_inv = await db.branch_sku_inventory.find_one(
        {"sku_id": input.sku_id, "branch": input.branch, "is_active": True},
        {"_id": 0}
    )
    if not sku_inv:
        raise HTTPException(status_code=400, detail=f"SKU not active in {input.branch}")
    
    # Get mapping from either collection
    mapping = await db.sku_mappings.find_one({"sku_id": input.sku_id}, {"_id": 0})
    rm_mappings = []
    
    if mapping:
        rm_mappings = mapping['rm_mappings']
    else:
        # Check new sku_rm_mapping collection
        new_mappings = await db.sku_rm_mapping.find({"sku_id": input.sku_id}, {"_id": 0}).to_list(100)
        if new_mappings:
            rm_mappings = [{"rm_id": m['rm_id'], "quantity_required": m.get('quantity', 1)} for m in new_mappings]
    
    if not rm_mappings:
        raise HTTPException(status_code=400, detail="SKU mapping not found. Please map raw materials first.")
    
    # Check RM stock
    for rm_mapping in rm_mappings:
        required_qty = rm_mapping['quantity_required'] * input.quantity
        rm_inv = await db.branch_rm_inventory.find_one(
            {"rm_id": rm_mapping['rm_id'], "branch": input.branch, "is_active": True},
            {"_id": 0}
        )
        if not rm_inv:
            raise HTTPException(status_code=400, detail=f"RM {rm_mapping['rm_id']} not active in {input.branch}")
        if rm_inv['current_stock'] < required_qty:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock for {rm_mapping['rm_id']}. Required: {required_qty}, Available: {rm_inv['current_stock']}"
            )
    
    entry_obj = ProductionEntry(**input.model_dump())
    doc = entry_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['date'] = doc['date'].isoformat()
    await db.production_entries.insert_one(doc)
    
    # Deduct RM stock
    for rm_mapping in rm_mappings:
        required_qty = rm_mapping['quantity_required'] * input.quantity
        await db.branch_rm_inventory.update_one(
            {"rm_id": rm_mapping['rm_id'], "branch": input.branch},
            {"$inc": {"current_stock": -required_qty}}
        )
    
    # Add SKU stock
    await db.branch_sku_inventory.update_one(
        {"sku_id": input.sku_id, "branch": input.branch},
        {"$inc": {"current_stock": input.quantity}}
    )
    
    return entry_obj

@api_router.get("/production-entries")
async def get_production_entries(branch: Optional[str] = None, sku_id: Optional[str] = None):
    query = {}
    if branch:
        query["branch"] = branch
    if sku_id:
        query["sku_id"] = sku_id
    
    entries = await db.production_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return [serialize_doc(e) for e in entries]

# ============ Dispatch Entry Routes ============

@api_router.post("/dispatch-entries")
async def create_dispatch_entry(input: DispatchEntryCreate):
    sku_inv = await db.branch_sku_inventory.find_one(
        {"sku_id": input.sku_id, "branch": input.branch, "is_active": True},
        {"_id": 0}
    )
    if not sku_inv:
        raise HTTPException(status_code=400, detail=f"SKU not active in {input.branch}")
    
    if sku_inv['current_stock'] < input.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient SKU stock. Required: {input.quantity}, Available: {sku_inv['current_stock']}"
        )
    
    entry_obj = DispatchEntry(**input.model_dump())
    doc = entry_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['date'] = doc['date'].isoformat()
    await db.dispatch_entries.insert_one(doc)
    
    await db.branch_sku_inventory.update_one(
        {"sku_id": input.sku_id, "branch": input.branch},
        {"$inc": {"current_stock": -input.quantity}}
    )
    
    return entry_obj

@api_router.get("/dispatch-entries")
async def get_dispatch_entries(branch: Optional[str] = None, sku_id: Optional[str] = None):
    query = {}
    if branch:
        query["branch"] = branch
    if sku_id:
        query["sku_id"] = sku_id
    
    entries = await db.dispatch_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return [serialize_doc(e) for e in entries]

# ============ Inter-Branch SKU Transfer ============

class SKUTransferCreate(BaseModel):
    sku_id: str
    from_branch: str
    to_branch: str
    quantity: float
    notes: Optional[str] = ""

class SKUTransfer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_id: str
    from_branch: str
    to_branch: str
    quantity: float
    notes: Optional[str] = ""
    transferred_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.post("/sku-transfers")
async def create_sku_transfer(input: SKUTransferCreate):
    """
    Transfer SKU inventory between branches.
    This is for physical inventory movement only - no RM consumption.
    """
    if input.from_branch == input.to_branch:
        raise HTTPException(status_code=400, detail="Source and destination branches must be different")
    
    if input.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")
    
    # Check if SKU exists in source branch and has sufficient stock
    source_inv = await db.branch_sku_inventory.find_one(
        {"sku_id": input.sku_id, "branch": input.from_branch, "is_active": True},
        {"_id": 0}
    )
    
    if not source_inv:
        raise HTTPException(status_code=400, detail=f"SKU {input.sku_id} not active in {input.from_branch}")
    
    if source_inv['current_stock'] < input.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient stock. Available: {source_inv['current_stock']}, Requested: {input.quantity}"
        )
    
    # Check if SKU is activated in destination branch, if not activate it
    dest_inv = await db.branch_sku_inventory.find_one(
        {"sku_id": input.sku_id, "branch": input.to_branch},
        {"_id": 0}
    )
    
    if not dest_inv:
        # Activate SKU in destination branch
        inv_obj = BranchSKUInventory(sku_id=input.sku_id, branch=input.to_branch)
        inv_doc = inv_obj.model_dump()
        inv_doc['activated_at'] = inv_doc['activated_at'].isoformat()
        await db.branch_sku_inventory.insert_one(inv_doc)
    elif not dest_inv.get('is_active', False):
        # Reactivate if inactive
        await db.branch_sku_inventory.update_one(
            {"sku_id": input.sku_id, "branch": input.to_branch},
            {"$set": {"is_active": True}}
        )
    
    # Create transfer record
    transfer_obj = SKUTransfer(**input.model_dump())
    doc = transfer_obj.model_dump()
    doc['transferred_at'] = doc['transferred_at'].isoformat()
    await db.sku_transfers.insert_one(doc)
    
    # Deduct from source branch (no RM consumption)
    await db.branch_sku_inventory.update_one(
        {"sku_id": input.sku_id, "branch": input.from_branch},
        {"$inc": {"current_stock": -input.quantity}}
    )
    
    # Add to destination branch (no RM consumption)
    await db.branch_sku_inventory.update_one(
        {"sku_id": input.sku_id, "branch": input.to_branch},
        {"$inc": {"current_stock": input.quantity}}
    )
    
    # Return clean response without ObjectId
    return {
        "message": f"Transferred {input.quantity} units of {input.sku_id} from {input.from_branch} to {input.to_branch}",
        "transfer": {
            "id": doc['id'],
            "sku_id": doc['sku_id'],
            "from_branch": doc['from_branch'],
            "to_branch": doc['to_branch'],
            "quantity": doc['quantity'],
            "notes": doc['notes'],
            "transferred_at": doc['transferred_at']
        }
    }

@api_router.get("/sku-transfers")
async def get_sku_transfers(branch: Optional[str] = None, sku_id: Optional[str] = None):
    """Get transfer history, optionally filtered by branch (as source or destination) or SKU"""
    query = {}
    if branch:
        query["$or"] = [{"from_branch": branch}, {"to_branch": branch}]
    if sku_id:
        query["sku_id"] = sku_id
    
    transfers = await db.sku_transfers.find(query, {"_id": 0}).sort("transferred_at", -1).to_list(1000)
    return [serialize_doc(t) for t in transfers]

@api_router.get("/sku-transfers/summary")
async def get_transfer_summary(branch: str):
    """Get transfer summary for a branch - incoming vs outgoing"""
    incoming = await db.sku_transfers.find({"to_branch": branch}, {"_id": 0}).to_list(1000)
    outgoing = await db.sku_transfers.find({"from_branch": branch}, {"_id": 0}).to_list(1000)
    
    return {
        "incoming_count": len(incoming),
        "outgoing_count": len(outgoing),
        "incoming_total": sum(t['quantity'] for t in incoming),
        "outgoing_total": sum(t['quantity'] for t in outgoing)
    }

# ============ Dashboard & Reports Routes ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(branch: Optional[str] = None):
    if not branch:
        raise HTTPException(status_code=400, detail="Branch parameter required")
    
    rm_count = await db.branch_rm_inventory.count_documents({"branch": branch, "is_active": True})
    sku_count = await db.branch_sku_inventory.count_documents({"branch": branch, "is_active": True})
    
    # Low stock
    rm_invs = await db.branch_rm_inventory.find({"branch": branch, "is_active": True}, {"_id": 0}).to_list(1000)
    low_stock_rm = 0
    for rm_inv in rm_invs:
        rm = await db.raw_materials.find_one({"rm_id": rm_inv['rm_id']}, {"_id": 0})
        if rm and rm_inv['current_stock'] < rm.get('low_stock_threshold', 10):
            low_stock_rm += 1
    
    sku_invs = await db.branch_sku_inventory.find({"branch": branch, "is_active": True}, {"_id": 0}).to_list(1000)
    low_stock_sku = 0
    for sku_inv in sku_invs:
        sku = await db.skus.find_one({"sku_id": sku_inv['sku_id']}, {"_id": 0})
        if sku and sku_inv['current_stock'] < sku.get('low_stock_threshold', 5):
            low_stock_sku += 1
    
    # Today's production
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_entries = await db.production_entries.find(
        {"branch": branch, "date": {"$gte": today_start.isoformat()}},
        {"_id": 0}
    ).to_list(1000)
    today_production = sum(e.get('quantity', 0) for e in today_entries)
    
    return {
        "total_rm_value": rm_count,
        "total_sku_value": sku_count,
        "low_stock_items": low_stock_rm + low_stock_sku,
        "today_production": int(today_production)
    }

@api_router.get("/reports/master-dashboard")
async def get_master_dashboard():
    stats_by_branch = {}
    for branch in BRANCHES:
        stats = await get_dashboard_stats(branch)
        stats_by_branch[branch] = stats
    
    total_rm = sum(s['total_rm_value'] for s in stats_by_branch.values())
    total_sku = sum(s['total_sku_value'] for s in stats_by_branch.values())
    total_low_stock = sum(s['low_stock_items'] for s in stats_by_branch.values())
    total_production = sum(s['today_production'] for s in stats_by_branch.values())
    
    return {
        "overall": {
            "total_rm_value": total_rm,
            "total_sku_value": total_sku,
            "low_stock_items": total_low_stock,
            "today_production": total_production
        },
        "by_branch": stats_by_branch
    }

@api_router.get("/reports/low-stock")
async def get_low_stock_report(branch: Optional[str] = None):
    if not branch:
        raise HTTPException(status_code=400, detail="Branch parameter required")
    
    low_stock_rm = []
    rm_invs = await db.branch_rm_inventory.find({"branch": branch, "is_active": True}, {"_id": 0}).to_list(1000)
    for rm_inv in rm_invs:
        rm = await db.raw_materials.find_one({"rm_id": rm_inv['rm_id']}, {"_id": 0})
        if rm and rm_inv['current_stock'] < rm.get('low_stock_threshold', 10):
            rm_data = rm.copy()
            rm_data['current_stock'] = rm_inv['current_stock']
            low_stock_rm.append(serialize_doc(rm_data))
    
    low_stock_sku = []
    sku_invs = await db.branch_sku_inventory.find({"branch": branch, "is_active": True}, {"_id": 0}).to_list(1000)
    for sku_inv in sku_invs:
        sku = await db.skus.find_one({"sku_id": sku_inv['sku_id']}, {"_id": 0})
        if sku and sku_inv['current_stock'] < sku.get('low_stock_threshold', 5):
            sku_data = sku.copy()
            sku_data['current_stock'] = sku_inv['current_stock']
            low_stock_sku.append(serialize_doc(sku_data))
    
    return {
        "raw_materials": low_stock_rm,
        "skus": low_stock_sku
    }

@api_router.get("/reports/production-summary")
async def get_production_summary(days: int = 7, branch: Optional[str] = None):
    if not branch:
        raise HTTPException(status_code=400, detail="Branch parameter required")
    
    from datetime import timedelta
    start_date = datetime.now(timezone.utc) - timedelta(days=days)
    
    entries = await db.production_entries.find(
        {"branch": branch, "date": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    
    daily_summary = {}
    for entry in entries:
        date_str = entry['date'][:10]
        if date_str not in daily_summary:
            daily_summary[date_str] = {"total_quantity": 0, "items": []}
        daily_summary[date_str]["total_quantity"] += entry['quantity']
        daily_summary[date_str]["items"].append(entry)
    
    return {
        "entries": [serialize_doc(e) for e in entries],
        "daily_summary": daily_summary
    }

@api_router.get("/reports/inventory")
async def get_inventory_report(branch: Optional[str] = None):
    if not branch:
        raise HTTPException(status_code=400, detail="Branch parameter required")
    
    rm_invs = await db.branch_rm_inventory.find({"branch": branch, "is_active": True}, {"_id": 0}).to_list(1000)
    raw_materials = []
    for rm_inv in rm_invs:
        rm = await db.raw_materials.find_one({"rm_id": rm_inv['rm_id']}, {"_id": 0})
        if rm:
            rm_data = rm.copy()
            rm_data['current_stock'] = rm_inv['current_stock']
            raw_materials.append(serialize_doc(rm_data))
    
    sku_invs = await db.branch_sku_inventory.find({"branch": branch, "is_active": True}, {"_id": 0}).to_list(1000)
    skus = []
    for sku_inv in sku_invs:
        sku = await db.skus.find_one({"sku_id": sku_inv['sku_id']}, {"_id": 0})
        if sku:
            sku_data = sku.copy()
            sku_data['current_stock'] = sku_inv['current_stock']
            skus.append(serialize_doc(sku_data))
    
    return {
        "raw_materials": raw_materials,
        "skus": skus
    }

# ============ Production Planning Routes ============

@api_router.post("/production-plans/bulk-upload")
async def bulk_upload_production_plan(file: UploadFile = File(...), branch: str = "Unit 1 Vedica"):
    """Upload production plan for the month"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        # Expected format: Date, SKU_ID, Planned_Quantity
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not row[1]:
                continue
            
            try:
                date_val = row[0]
                if isinstance(date_val, str):
                    date_obj = datetime.strptime(date_val, "%Y-%m-%d")
                else:
                    date_obj = datetime.combine(date_val, datetime.min.time())
                
                sku_id = str(row[1]).strip()
                planned_qty = float(row[2]) if row[2] else 0
                
                if planned_qty <= 0:
                    errors.append(f"Row {idx}: Invalid quantity")
                    continue
                
                # Check if SKU exists globally
                sku = await db.skus.find_one({"sku_id": sku_id}, {"_id": 0})
                if not sku:
                    errors.append(f"Row {idx}: SKU {sku_id} not found")
                    continue
                
                plan_month = date_obj.strftime("%Y-%m")
                
                # Check if entry already exists
                existing = await db.production_plans.find_one({
                    "branch": branch,
                    "date": date_obj.isoformat(),
                    "sku_id": sku_id
                }, {"_id": 0})
                
                if existing:
                    # Update existing
                    await db.production_plans.update_one(
                        {"branch": branch, "date": date_obj.isoformat(), "sku_id": sku_id},
                        {"$set": {"planned_quantity": planned_qty, "plan_month": plan_month}}
                    )
                    skipped_count += 1
                else:
                    plan_obj = ProductionPlanEntry(
                        branch=branch,
                        plan_month=plan_month,
                        date=date_obj,
                        sku_id=sku_id,
                        planned_quantity=planned_qty
                    )
                    doc = plan_obj.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    doc['date'] = doc['date'].isoformat()
                    await db.production_plans.insert_one(doc)
                    created_count += 1
                    
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        return {
            "created": created_count,
            "updated": skipped_count,
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/production-plans")
async def get_production_plans(branch: str, plan_month: Optional[str] = None):
    """Get production plans for a branch"""
    query = {"branch": branch}
    if plan_month:
        query["plan_month"] = plan_month
    
    plans = await db.production_plans.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    return [serialize_doc(p) for p in plans]

class SingleProductionPlanCreate(BaseModel):
    sku_id: str
    branch: str
    date: datetime
    planned_quantity: float

@api_router.post("/production-plans")
async def create_production_plan(plan: SingleProductionPlanCreate):
    """Create a single production plan entry for a specific SKU and date"""
    # Validate SKU exists
    sku = await db.skus.find_one({"sku_id": plan.sku_id}, {"_id": 0})
    if not sku:
        raise HTTPException(status_code=404, detail=f"SKU {plan.sku_id} not found")
    
    # Derive plan_month from date
    plan_month = plan.date.strftime("%Y-%m")
    
    # Check if a plan entry already exists for this SKU, branch, and date
    existing = await db.production_plans.find_one({
        "sku_id": plan.sku_id,
        "branch": plan.branch,
        "date": plan.date
    })
    
    if existing:
        # Update existing plan
        await db.production_plans.update_one(
            {"_id": existing["_id"]},
            {"$set": {"planned_quantity": plan.planned_quantity, "updated_at": datetime.now(timezone.utc)}}
        )
        return {"message": "Production plan updated", "sku_id": plan.sku_id, "date": plan.date.isoformat(), "planned_quantity": plan.planned_quantity}
    
    # Create new plan entry
    plan_doc = {
        "id": str(uuid.uuid4()),
        "branch": plan.branch,
        "plan_month": plan_month,
        "date": plan.date,
        "sku_id": plan.sku_id,
        "planned_quantity": plan.planned_quantity,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.production_plans.insert_one(plan_doc)
    
    return {"message": "Production plan created", "sku_id": plan.sku_id, "date": plan.date.isoformat(), "planned_quantity": plan.planned_quantity}

@api_router.delete("/production-plans/{plan_month}")
async def delete_production_plan(plan_month: str, branch: str):
    """Delete production plan for a specific month"""
    result = await db.production_plans.delete_many({"branch": branch, "plan_month": plan_month})
    return {"message": f"Deleted {result.deleted_count} plan entries"}

@api_router.delete("/production-plans/entry/{plan_id}")
async def delete_single_plan_entry(plan_id: str):
    """Delete a single production plan entry by its ID"""
    result = await db.production_plans.delete_one({"id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan entry not found")
    return {"message": "Plan entry deleted"}

@api_router.get("/production-plans/shortage-analysis")
async def get_shortage_analysis(branch: str, plan_month: str):
    """Calculate RM shortages based on production plan"""
    # Get all plans for the month
    plans = await db.production_plans.find(
        {"branch": branch, "plan_month": plan_month},
        {"_id": 0}
    ).to_list(1000)
    
    if not plans:
        raise HTTPException(status_code=404, detail="No production plan found for this month")
    
    # Calculate total RM requirements
    rm_requirements = {}  # {rm_id: total_required}
    sku_details = {}  # {sku_id: {name, total_planned}}
    
    for plan in plans:
        sku_id = plan['sku_id']
        planned_qty = plan['planned_quantity']
        
        # Get SKU details
        if sku_id not in sku_details:
            sku = await db.skus.find_one({"sku_id": sku_id}, {"_id": 0})
            if sku:
                sku_details[sku_id] = {
                    "name": sku.get('description', sku_id),
                    "total_planned": 0
                }
        
        if sku_id in sku_details:
            sku_details[sku_id]["total_planned"] += planned_qty
        
        # Get BOM mapping - check both collections for compatibility
        # First try sku_rm_mapping (flat structure: sku_id, rm_id, quantity)
        rm_mappings = await db.sku_rm_mapping.find({"sku_id": sku_id}, {"_id": 0}).to_list(1000)
        if rm_mappings:
            for rm_mapping in rm_mappings:
                rm_id = rm_mapping['rm_id']
                qty_per_unit = rm_mapping.get('quantity', 0)
                total_required = qty_per_unit * planned_qty
                
                if rm_id in rm_requirements:
                    rm_requirements[rm_id] += total_required
                else:
                    rm_requirements[rm_id] = total_required
        else:
            # Fallback to sku_mappings (nested structure with rm_mappings array)
            mapping = await db.sku_mappings.find_one({"sku_id": sku_id}, {"_id": 0})
            if mapping and mapping.get('rm_mappings'):
                for rm_mapping in mapping['rm_mappings']:
                    rm_id = rm_mapping['rm_id']
                    qty_per_unit = rm_mapping.get('quantity_required', rm_mapping.get('quantity', 0))
                    total_required = qty_per_unit * planned_qty
                    
                    if rm_id in rm_requirements:
                        rm_requirements[rm_id] += total_required
                    else:
                        rm_requirements[rm_id] = total_required
    
    # Get current inventory levels
    shortage_report = []
    sufficient_stock = []
    
    for rm_id, total_required in rm_requirements.items():
        # Get RM details
        rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
        if not rm:
            continue
        
        # Get current stock in branch
        rm_inv = await db.branch_rm_inventory.find_one(
            {"rm_id": rm_id, "branch": branch, "is_active": True},
            {"_id": 0}
        )
        
        current_stock = rm_inv['current_stock'] if rm_inv else 0
        shortage = total_required - current_stock
        
        rm_info = {
            "rm_id": rm_id,
            "category": rm.get('category', ''),
            "category_data": rm.get('category_data', {}),
            "total_required": round(total_required, 2),
            "current_stock": round(current_stock, 2),
            "shortage": round(shortage, 2) if shortage > 0 else 0,
            "status": "shortage" if shortage > 0 else "sufficient"
        }
        
        if shortage > 0:
            shortage_report.append(rm_info)
        else:
            sufficient_stock.append(rm_info)
    
    # Calculate plan summary
    total_skus = len(sku_details)
    total_units = sum(s['total_planned'] for s in sku_details.values())
    total_rm_types = len(rm_requirements)
    rm_with_shortage = len(shortage_report)
    
    return {
        "plan_summary": {
            "branch": branch,
            "plan_month": plan_month,
            "total_skus": total_skus,
            "total_units_planned": int(total_units),
            "total_rm_types": total_rm_types,
            "rm_with_shortage": rm_with_shortage,
            "plan_entries": len(plans)
        },
        "sku_details": sku_details,
        "shortage_report": sorted(shortage_report, key=lambda x: x['shortage'], reverse=True),
        "sufficient_stock": sufficient_stock
    }

@api_router.get("/production-plans/months")
async def get_available_plan_months(branch: str):
    """Get list of months with production plans"""
    plans = await db.production_plans.find({"branch": branch}, {"_id": 0, "plan_month": 1}).to_list(1000)
    months = list(set(p['plan_month'] for p in plans))
    months.sort(reverse=True)
    return {"months": months}

# ============ Vendor Routes ============

@api_router.post("/vendors")
async def create_vendor(input: VendorCreate):
    """Create a new vendor with auto-generated vendor ID"""
    vendor_id = await get_next_vendor_id()
    vendor_obj = Vendor(**input.model_dump(), vendor_id=vendor_id)
    doc = vendor_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.vendors.insert_one(doc)
    return vendor_obj

@api_router.get("/vendors")
async def get_vendors(search: Optional[str] = None):
    """Get all vendors"""
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"gst": {"$regex": search, "$options": "i"}},
            {"poc": {"$regex": search, "$options": "i"}},
            {"vendor_id": {"$regex": search, "$options": "i"}}
        ]
    vendors = await db.vendors.find(query, {"_id": 0}).sort("vendor_id", 1).to_list(1000)
    return [serialize_doc(v) for v in vendors]

@api_router.get("/vendors/{vendor_id}")
async def get_vendor(vendor_id: str):
    """Get vendor details with RM prices"""
    # Try to find by internal id or vendor_id
    vendor = await db.vendors.find_one(
        {"$or": [{"id": vendor_id}, {"vendor_id": vendor_id}]}, 
        {"_id": 0}
    )
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Use the actual vendor_id field to query prices
    actual_vendor_id = vendor.get('vendor_id', vendor_id)
    
    # Get all RM prices for this vendor
    prices = await db.vendor_rm_prices.find({"vendor_id": actual_vendor_id}, {"_id": 0}).to_list(1000)
    
    # Enrich with RM details
    enriched_prices = []
    for p in prices:
        rm = await db.raw_materials.find_one({"rm_id": p['rm_id']}, {"_id": 0})
        enriched_prices.append({
            **serialize_doc(p),
            "rm_category": rm['category'] if rm else "",
            "rm_details": rm.get('category_data', {}) if rm else {}
        })
    
    return {
        "vendor": serialize_doc(vendor),
        "rm_prices": enriched_prices
    }

@api_router.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, input: VendorCreate):
    """Update vendor details"""
    existing = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    await db.vendors.update_one(
        {"id": vendor_id},
        {"$set": input.model_dump()}
    )
    return {"message": "Vendor updated successfully"}

@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str):
    """Delete vendor and associated prices"""
    await db.vendors.delete_one({"id": vendor_id})
    await db.vendor_rm_prices.delete_many({"vendor_id": vendor_id})
    return {"message": "Vendor deleted"}

@api_router.post("/vendors/bulk-upload")
async def bulk_upload_vendors(file: UploadFile = File(...)):
    """Bulk upload vendors from Excel file.
    Expected columns: Name, GST, Address, POC, Email, Phone
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        # Get headers
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
        
        # Map headers to fields
        header_map = {
            'name': 'name', 'vendor name': 'name',
            'gst': 'gst', 'gstin': 'gst', 'gst number': 'gst',
            'address': 'address',
            'poc': 'poc', 'point of contact': 'poc', 'contact person': 'poc',
            'email': 'email', 'email address': 'email',
            'phone': 'phone', 'phone number': 'phone', 'mobile': 'phone'
        }
        
        field_indices = {}
        for i, h in enumerate(headers):
            if h in header_map:
                field_indices[header_map[h]] = i
        
        if 'name' not in field_indices:
            raise HTTPException(status_code=400, detail="Name column is required")
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = row[field_indices['name']] if field_indices.get('name') is not None else None
                if not name:
                    continue
                
                name = str(name).strip()
                
                # Check for duplicate by name
                existing = await db.vendors.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}})
                if existing:
                    skipped_count += 1
                    continue
                
                # Build vendor data
                vendor_data = {'name': name}
                for field in ['gst', 'address', 'poc', 'email', 'phone']:
                    if field in field_indices and row[field_indices[field]]:
                        vendor_data[field] = str(row[field_indices[field]]).strip()
                
                # Generate vendor ID
                vendor_id = await get_next_vendor_id()
                
                vendor_obj = Vendor(**vendor_data, vendor_id=vendor_id)
                doc = vendor_obj.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.vendors.insert_one(doc)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        return {
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors[:20],
            "total_errors": len(errors)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

# ============ Vendor RM Pricing Routes ============

@api_router.post("/vendor-rm-prices")
async def create_vendor_rm_price(input: VendorRMPriceCreate):
    """Add or update RM price for a vendor"""
    # Verify vendor exists
    vendor = await db.vendors.find_one({"id": input.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Verify RM exists
    rm = await db.raw_materials.find_one({"rm_id": input.rm_id}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    # Check if mapping already exists
    existing = await db.vendor_rm_prices.find_one(
        {"vendor_id": input.vendor_id, "rm_id": input.rm_id},
        {"_id": 0}
    )
    
    if existing:
        # Update existing price
        await db.vendor_rm_prices.update_one(
            {"vendor_id": input.vendor_id, "rm_id": input.rm_id},
            {"$set": {"price": input.price, "currency": input.currency, "notes": input.notes, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"message": "Price updated", "action": "updated"}
    else:
        # Create new price entry
        price_obj = VendorRMPrice(**input.model_dump())
        doc = price_obj.model_dump()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.vendor_rm_prices.insert_one(doc)
        return {"message": "Price added", "action": "created"}

@api_router.get("/vendor-rm-prices/by-rm/{rm_id}")
async def get_vendors_for_rm(rm_id: str):
    """Get all vendors and their prices for a specific RM"""
    prices = await db.vendor_rm_prices.find({"rm_id": rm_id}, {"_id": 0}).to_list(1000)
    
    result = []
    for p in prices:
        # vendor_id in vendor_rm_prices is the VND_XXX format, not UUID
        vendor = await db.vendors.find_one({"vendor_id": p['vendor_id']}, {"_id": 0})
        if vendor:
            result.append({
                **serialize_doc(p),
                "vendor_name": vendor['name'],
                "vendor_gst": vendor.get('gst', ''),
                "vendor_phone": vendor.get('phone', '')
            })
    
    # Sort by price (lowest first)
    result.sort(key=lambda x: x['price'])
    return result

@api_router.get("/vendor-rm-prices/comparison")
async def get_price_comparison_report():
    """Get price comparison report - lowest price per RM across all vendors"""
    # Get all prices
    all_prices = await db.vendor_rm_prices.find({}, {"_id": 0}).to_list(10000)
    
    # Group by RM and find lowest price
    rm_prices = {}
    for p in all_prices:
        rm_id = p['rm_id']
        if rm_id not in rm_prices or p['price'] < rm_prices[rm_id]['lowest_price']:
            rm_prices[rm_id] = {
                'rm_id': rm_id,
                'lowest_price': p['price'],
                'lowest_vendor_id': p['vendor_id'],
                'currency': p.get('currency', 'INR')
            }
    
    # Enrich with vendor names and RM details
    result = []
    for rm_id, data in rm_prices.items():
        vendor = await db.vendors.find_one({"id": data['lowest_vendor_id']}, {"_id": 0})
        rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
        
        # Get all vendors for this RM for comparison
        all_vendors_for_rm = await db.vendor_rm_prices.find({"rm_id": rm_id}, {"_id": 0}).to_list(100)
        
        result.append({
            'rm_id': rm_id,
            'rm_category': rm['category'] if rm else '',
            'lowest_price': data['lowest_price'],
            'currency': data['currency'],
            'lowest_vendor_name': vendor['name'] if vendor else '',
            'lowest_vendor_id': data['lowest_vendor_id'],
            'total_vendors': len(all_vendors_for_rm)
        })
    
    # Sort by RM ID
    result.sort(key=lambda x: x['rm_id'])
    return result

@api_router.delete("/vendor-rm-prices/{vendor_id}/{rm_id}")
async def delete_vendor_rm_price(vendor_id: str, rm_id: str):
    """Delete a vendor RM price mapping"""
    await db.vendor_rm_prices.delete_one({"vendor_id": vendor_id, "rm_id": rm_id})
    return {"message": "Price mapping deleted"}

@api_router.post("/vendor-rm-prices/bulk-upload")
async def bulk_upload_vendor_rm_prices(file: UploadFile = File(...)):
    """Bulk upload Vendor RM pricing from Excel file.
    Expected columns: Vendor ID, RM ID, Price
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        added_count = 0
        updated_count = 0
        errors = []
        
        # Find the header row (look for "Vendor ID" in first few rows)
        header_row = 1
        for row_num in range(1, 5):
            row_values = [cell.value for cell in sheet[row_num]]
            if 'Vendor ID' in row_values or 'vendor_id' in [str(v).lower() if v else '' for v in row_values]:
                header_row = row_num
                break
        
        # Find column indices
        headers = [cell.value for cell in sheet[header_row]]
        vendor_col = None
        rm_col = None
        price_col = None
        
        for idx, h in enumerate(headers):
            if h and 'vendor' in str(h).lower():
                vendor_col = idx
            elif h and 'rm' in str(h).lower():
                rm_col = idx
            elif h and 'price' in str(h).lower():
                price_col = idx
        
        if vendor_col is None or rm_col is None or price_col is None:
            raise HTTPException(status_code=400, detail="Could not find required columns: Vendor ID, RM ID, Price")
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            vendor_id = row[vendor_col] if vendor_col < len(row) else None
            rm_id = row[rm_col] if rm_col < len(row) else None
            price = row[price_col] if price_col < len(row) else None
            
            if not vendor_id or not rm_id:
                continue
            
            vendor_id = str(vendor_id).strip()
            rm_id = str(rm_id).strip()
            
            try:
                price_val = float(price) if price else 0
            except (ValueError, TypeError):
                errors.append(f"Invalid price for {vendor_id}/{rm_id}: {price}")
                continue
            
            # Verify vendor exists
            vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
            if not vendor:
                errors.append(f"Vendor not found: {vendor_id}")
                continue
            
            # Verify RM exists
            rm = await db.raw_materials.find_one({"rm_id": rm_id}, {"_id": 0})
            if not rm:
                errors.append(f"RM not found: {rm_id}")
                continue
            
            # Check if mapping exists
            existing = await db.vendor_rm_prices.find_one(
                {"vendor_id": vendor_id, "rm_id": rm_id},
                {"_id": 0}
            )
            
            if existing:
                # Update
                await db.vendor_rm_prices.update_one(
                    {"vendor_id": vendor_id, "rm_id": rm_id},
                    {"$set": {"price": price_val, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                updated_count += 1
            else:
                # Insert
                price_obj = VendorRMPrice(vendor_id=vendor_id, rm_id=rm_id, price=price_val)
                doc = price_obj.model_dump()
                doc['updated_at'] = doc['updated_at'].isoformat()
                await db.vendor_rm_prices.insert_one(doc)
                added_count += 1
        
        return {
            "added": added_count,
            "updated": updated_count,
            "errors": errors[:20],
            "total_errors": len(errors)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/branch-sku-inventory/bulk-upload")
async def bulk_upload_branch_inventory(file: UploadFile = File(...)):
    """Bulk upload branch-level SKU inventory from Excel file.
    Expected columns: SKU_ID, Quantity, Branch
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        updated_count = 0
        created_count = 0
        errors = []
        branch_stats = {}
        
        # Find the header row (look for "SKU_ID" in first few rows)
        header_row = 1
        for row_num in range(1, 5):
            row_values = [cell.value for cell in sheet[row_num]]
            if 'SKU_ID' in row_values or 'sku_id' in [str(v).lower() if v else '' for v in row_values]:
                header_row = row_num
                break
        
        # Find column indices
        headers = [cell.value for cell in sheet[header_row]]
        sku_col = None
        qty_col = None
        branch_col = None
        
        for idx, h in enumerate(headers):
            h_lower = str(h).lower() if h else ''
            if 'sku' in h_lower and 'id' in h_lower:
                sku_col = idx
            elif 'quantity' in h_lower or 'qty' in h_lower:
                qty_col = idx
            elif 'branch' in h_lower:
                branch_col = idx
        
        if sku_col is None or qty_col is None or branch_col is None:
            raise HTTPException(status_code=400, detail="Could not find required columns: SKU_ID, Quantity, Branch")
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            sku_id = row[sku_col] if sku_col < len(row) else None
            quantity = row[qty_col] if qty_col < len(row) else None
            branch = row[branch_col] if branch_col < len(row) else None
            
            if not sku_id or not branch:
                continue
            
            sku_id = str(sku_id).strip()
            branch = str(branch).strip()
            
            try:
                qty_val = float(quantity) if quantity else 0
            except (ValueError, TypeError):
                errors.append(f"Invalid quantity for {sku_id}: {quantity}")
                continue
            
            # Verify SKU exists
            sku = await db.skus.find_one(
                {"$or": [{"sku_id": sku_id}, {"buyer_sku_id": sku_id}]},
                {"_id": 0, "sku_id": 1}
            )
            if not sku:
                errors.append(f"SKU not found: {sku_id}")
                continue
            
            actual_sku_id = sku['sku_id']
            
            # Check if inventory exists
            existing = await db.branch_sku_inventory.find_one(
                {"sku_id": actual_sku_id, "branch": branch},
                {"_id": 0}
            )
            
            if existing:
                # Update stock
                await db.branch_sku_inventory.update_one(
                    {"sku_id": actual_sku_id, "branch": branch},
                    {"$set": {"current_stock": qty_val, "is_active": True}}
                )
                updated_count += 1
            else:
                # Create inventory record
                inv_obj = BranchSKUInventory(sku_id=actual_sku_id, branch=branch, current_stock=qty_val)
                inv_doc = inv_obj.model_dump()
                inv_doc['activated_at'] = inv_doc['activated_at'].isoformat()
                await db.branch_sku_inventory.insert_one(inv_doc)
                created_count += 1
            
            # Track branch stats
            if branch not in branch_stats:
                branch_stats[branch] = {"count": 0, "total_qty": 0}
            branch_stats[branch]["count"] += 1
            branch_stats[branch]["total_qty"] += qty_val
        
        return {
            "created": created_count,
            "updated": updated_count,
            "total_processed": created_count + updated_count,
            "branch_stats": branch_stats,
            "errors": errors[:20],
            "total_errors": len(errors)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ SKU Branch Assignment Routes ============

@api_router.post("/sku-branch-assignments/upload")
async def upload_sku_branch_assignments(file: UploadFile = File(...), branch: str = ""):
    """Upload SKU IDs to assign to a branch. Also activates corresponding RMs."""
    if not branch:
        raise HTTPException(status_code=400, detail="Branch is required")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        assigned_count = 0
        skipped_count = 0
        not_found = []
        total_rms_activated = 0
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            
            sku_id = str(row[0]).strip()
            
            # Check if SKU exists (by buyer_sku_id or sku_id)
            sku = await db.skus.find_one(
                {"$or": [{"buyer_sku_id": sku_id}, {"sku_id": sku_id}]},
                {"_id": 0}
            )
            
            if not sku:
                not_found.append(sku_id)
                continue
            
            actual_sku_id = sku['sku_id']
            
            # Check if already assigned
            existing = await db.sku_branch_assignments.find_one(
                {"sku_id": actual_sku_id, "branch": branch},
                {"_id": 0}
            )
            
            if existing:
                skipped_count += 1
                continue
            
            # Create assignment
            assignment = SKUBranchAssignment(sku_id=actual_sku_id, branch=branch)
            doc = assignment.model_dump()
            doc['assigned_at'] = doc['assigned_at'].isoformat()
            await db.sku_branch_assignments.insert_one(doc)
            
            # Also activate SKU in branch inventory
            existing_inv = await db.branch_sku_inventory.find_one(
                {"sku_id": actual_sku_id, "branch": branch},
                {"_id": 0}
            )
            if not existing_inv:
                inv_obj = BranchSKUInventory(sku_id=actual_sku_id, branch=branch)
                inv_doc = inv_obj.model_dump()
                inv_doc['activated_at'] = inv_doc['activated_at'].isoformat()
                await db.branch_sku_inventory.insert_one(inv_doc)
            
            # Activate corresponding RMs for this SKU
            rms_activated = await activate_rms_for_sku(actual_sku_id, branch)
            total_rms_activated += rms_activated
            
            assigned_count += 1
        
        return {
            "assigned": assigned_count,
            "skipped": skipped_count,
            "not_found": not_found[:20],
            "total_not_found": len(not_found),
            "rms_activated": total_rms_activated
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/sku-branch-assignments")
async def get_sku_branch_assignments(branch: Optional[str] = None):
    """Get SKU assignments, optionally filtered by branch"""
    query = {}
    if branch:
        query["branch"] = branch
    
    assignments = await db.sku_branch_assignments.find(query, {"_id": 0}).to_list(5000)
    
    # Enrich with SKU details
    result = []
    for a in assignments:
        sku = await db.skus.find_one({"sku_id": a['sku_id']}, {"_id": 0})
        if sku:
            result.append({
                **serialize_doc(a),
                "buyer_sku_id": sku.get('buyer_sku_id', ''),
                "bidso_sku": sku.get('bidso_sku', ''),
                "description": sku.get('description', ''),
                "brand": sku.get('brand', ''),
                "vertical": sku.get('vertical', ''),
                "model": sku.get('model', '')
            })
    
    return result

@api_router.delete("/sku-branch-assignments/{sku_id}/{branch}")
async def delete_sku_branch_assignment(sku_id: str, branch: str):
    """Remove SKU assignment from a branch"""
    await db.sku_branch_assignments.delete_one({"sku_id": sku_id, "branch": branch})
    return {"message": "Assignment removed"}

@api_router.post("/sku-branch-assignments/bulk-subscribe")
async def bulk_subscribe_skus(
    branch: str,
    vertical: Optional[str] = None,
    model: Optional[str] = None
):
    """Bulk subscribe all SKUs matching vertical and/or model to a branch. Also activates corresponding RMs."""
    if not branch:
        raise HTTPException(status_code=400, detail="Branch is required")
    
    if not vertical and not model:
        raise HTTPException(status_code=400, detail="At least vertical or model must be specified")
    
    # Build query for matching SKUs
    query = {}
    if vertical:
        query["vertical"] = vertical
    if model:
        query["model"] = model
    
    # Find all matching SKUs
    matching_skus = await db.skus.find(query, {"_id": 0, "sku_id": 1}).to_list(10000)
    
    if not matching_skus:
        return {
            "assigned": 0,
            "skipped": 0,
            "total_matching": 0,
            "rms_activated": 0,
            "message": "No SKUs found matching the criteria"
        }
    
    assigned_count = 0
    skipped_count = 0
    total_rms_activated = 0
    
    for sku in matching_skus:
        sku_id = sku['sku_id']
        
        # Check if already assigned
        existing = await db.sku_branch_assignments.find_one(
            {"sku_id": sku_id, "branch": branch},
            {"_id": 0}
        )
        
        if existing:
            skipped_count += 1
            continue
        
        # Create assignment
        assignment = SKUBranchAssignment(sku_id=sku_id, branch=branch)
        doc = assignment.model_dump()
        doc['assigned_at'] = doc['assigned_at'].isoformat()
        await db.sku_branch_assignments.insert_one(doc)
        
        # Also activate SKU in branch inventory
        existing_inv = await db.branch_sku_inventory.find_one(
            {"sku_id": sku_id, "branch": branch},
            {"_id": 0}
        )
        if not existing_inv:
            inv_obj = BranchSKUInventory(sku_id=sku_id, branch=branch)
            inv_doc = inv_obj.model_dump()
            inv_doc['activated_at'] = inv_doc['activated_at'].isoformat()
            await db.branch_sku_inventory.insert_one(inv_doc)
        
        # Activate corresponding RMs for this SKU
        rms_activated = await activate_rms_for_sku(sku_id, branch)
        total_rms_activated += rms_activated
        
        assigned_count += 1
    
    return {
        "assigned": assigned_count,
        "skipped": skipped_count,
        "total_matching": len(matching_skus),
        "rms_activated": total_rms_activated,
        "message": f"Subscribed {assigned_count} SKUs to {branch}, activated {total_rms_activated} RMs"
    }

@api_router.delete("/sku-branch-assignments/bulk-unsubscribe")
async def bulk_unsubscribe_skus(
    branch: str,
    vertical: Optional[str] = None,
    model: Optional[str] = None
):
    """Bulk unsubscribe all SKUs matching vertical and/or model from a branch"""
    if not branch:
        raise HTTPException(status_code=400, detail="Branch is required")
    
    if not vertical and not model:
        raise HTTPException(status_code=400, detail="At least vertical or model must be specified")
    
    # Build query for matching SKUs
    query = {}
    if vertical:
        query["vertical"] = vertical
    if model:
        query["model"] = model
    
    # Find all matching SKUs
    matching_skus = await db.skus.find(query, {"_id": 0, "sku_id": 1}).to_list(10000)
    sku_ids = [s['sku_id'] for s in matching_skus]
    
    if not sku_ids:
        return {"removed": 0, "message": "No matching SKUs found"}
    
    # Remove assignments
    result = await db.sku_branch_assignments.delete_many({
        "sku_id": {"$in": sku_ids},
        "branch": branch
    })
    
    return {
        "removed": result.deleted_count,
        "message": f"Removed {result.deleted_count} SKU assignments from {branch}"
    }

# ============ Enhanced RM Filtering ============

@api_router.get("/raw-materials/filter-options")
async def get_rm_filter_options():
    """Get unique values for RM filters"""
    # Get all categories
    categories = list(RM_CATEGORIES.keys())
    
    # Get sample of unique values for common fields
    all_rms = await db.raw_materials.find({}, {"_id": 0}).to_list(5000)
    
    # Extract unique values
    types = set()
    models = set()
    colours = set()
    brands = set()
    
    for rm in all_rms:
        data = rm.get('category_data', {})
        if data.get('type'): types.add(data['type'])
        if data.get('model'): models.add(data['model'])
        if data.get('model_name'): models.add(data['model_name'])
        if data.get('colour'): colours.add(data['colour'])
        if data.get('brand'): brands.add(data['brand'])
    
    return {
        "categories": categories,
        "types": sorted(list(types))[:100],
        "models": sorted(list(models))[:100],
        "colours": sorted(list(colours))[:100],
        "brands": sorted(list(brands))[:50]
    }

@api_router.get("/raw-materials/filtered")
async def get_filtered_raw_materials(
    category: Optional[str] = None,
    search: Optional[str] = None,
    type_filter: Optional[str] = None,
    model_filter: Optional[str] = None,
    colour_filter: Optional[str] = None,
    brand_filter: Optional[str] = None,
    branch: Optional[str] = None,
    page: int = 1,
    page_size: int = 100
):
    """Get RMs with advanced filtering and pagination - filters by branch if specified"""
    query = {}
    
    # If branch specified, only show active RMs in that branch
    active_rm_ids = None
    branch_inv_map = {}
    if branch:
        branch_inventories = await db.branch_rm_inventory.find(
            {"branch": branch, "is_active": True}, 
            {"_id": 0}
        ).to_list(10000)
        active_rm_ids = [inv['rm_id'] for inv in branch_inventories]
        branch_inv_map = {inv['rm_id']: inv for inv in branch_inventories}
        
        if not active_rm_ids:
            return {
                "items": [],
                "total": 0,
                "page": page,
                "page_size": page_size,
                "total_pages": 0
            }
        query["rm_id"] = {"$in": active_rm_ids}
    
    if category:
        query["category"] = category
    
    if search:
        if active_rm_ids:
            query["$and"] = query.get("$and", [])
            query["$and"].append({"rm_id": {"$regex": search, "$options": "i"}})
        else:
            query["rm_id"] = {"$regex": search, "$options": "i"}
    
    # Build category_data filters
    if type_filter:
        query["$or"] = query.get("$or", [])
        query["$or"].extend([
            {"category_data.type": {"$regex": type_filter, "$options": "i"}}
        ])
    
    if model_filter:
        if "$or" not in query:
            query["$or"] = []
        query["$or"].extend([
            {"category_data.model": {"$regex": model_filter, "$options": "i"}},
            {"category_data.model_name": {"$regex": model_filter, "$options": "i"}}
        ])
    
    if colour_filter:
        query["category_data.colour"] = {"$regex": colour_filter, "$options": "i"}
    
    if brand_filter:
        query["category_data.brand"] = {"$regex": brand_filter, "$options": "i"}
    
    # Count total
    total = await db.raw_materials.count_documents(query)
    
    # Paginate
    skip = (page - 1) * page_size
    materials = await db.raw_materials.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    # Add branch inventory data if branch filter is active
    result_items = []
    for mat in materials:
        if branch and mat['rm_id'] in branch_inv_map:
            mat['current_stock'] = branch_inv_map[mat['rm_id']].get('current_stock', 0)
            mat['branch'] = branch
        result_items.append(serialize_doc(mat))
    
    return {
        "items": result_items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

# ============ SKU Cascading Filter Endpoints ============

@api_router.get("/skus/filter-options")
async def get_sku_filter_options():
    """Get all distinct verticals, models, and brands for filters"""
    all_skus = await db.skus.find({}, {"_id": 0, "vertical": 1, "model": 1, "brand": 1}).to_list(10000)
    
    verticals = sorted(list(set(s.get('vertical', '') for s in all_skus if s.get('vertical'))))
    models = sorted(list(set(s.get('model', '') for s in all_skus if s.get('model'))))
    brands = sorted(list(set(s.get('brand', '') for s in all_skus if s.get('brand'))))
    
    return {
        "verticals": verticals,
        "models": models,
        "brands": brands
    }

@api_router.get("/skus/models-by-vertical")
async def get_models_by_vertical(vertical: str):
    """Get distinct models for a specific vertical"""
    skus = await db.skus.find({"vertical": vertical}, {"_id": 0, "model": 1}).to_list(10000)
    models = sorted(list(set(s.get('model', '') for s in skus if s.get('model'))))
    return {"models": models}

@api_router.get("/skus/brands-by-vertical-model")
async def get_brands_by_vertical_model(vertical: str, model: Optional[str] = None):
    """Get distinct brands for a specific vertical and optionally model"""
    query = {"vertical": vertical}
    if model:
        query["model"] = model
    skus = await db.skus.find(query, {"_id": 0, "brand": 1}).to_list(10000)
    brands = sorted(list(set(s.get('brand', '') for s in skus if s.get('brand'))))
    return {"brands": brands}

# ============ NEW ARCHITECTURE API ENDPOINTS (PRD v2) ============

# --- Verticals CRUD ---
@api_router.get("/verticals")
async def get_verticals():
    """Get all verticals"""
    verticals = await db.verticals.find({}, {"_id": 0}).to_list(1000)
    return [serialize_doc(v) for v in verticals]

@api_router.post("/verticals")
async def create_vertical(data: VerticalCreate):
    """Create a new vertical"""
    existing = await db.verticals.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Vertical with code {data.code} already exists")
    
    vertical = {
        "id": str(uuid.uuid4()),
        "code": data.code.upper(),
        "name": data.name,
        "description": data.description,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.verticals.insert_one(vertical)
    # Return without _id
    del vertical["_id"]
    return serialize_doc(vertical)

@api_router.put("/verticals/{vertical_id}")
async def update_vertical(vertical_id: str, data: VerticalCreate):
    """Update a vertical"""
    result = await db.verticals.update_one(
        {"id": vertical_id},
        {"$set": {"code": data.code.upper(), "name": data.name, "description": data.description}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vertical not found")
    return {"message": "Vertical updated"}

# --- Models CRUD ---
@api_router.get("/models")
async def get_models(vertical_id: Optional[str] = None):
    """Get all models, optionally filtered by vertical"""
    query = {}
    if vertical_id:
        query["vertical_id"] = vertical_id
    models = await db.models.find(query, {"_id": 0}).to_list(1000)
    return [serialize_doc(m) for m in models]

@api_router.post("/models")
async def create_model(data: ModelCreate):
    """Create a new model"""
    existing = await db.models.find_one({"vertical_id": data.vertical_id, "code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Model with code {data.code} already exists for this vertical")
    
    model = {
        "id": str(uuid.uuid4()),
        "vertical_id": data.vertical_id,
        "code": data.code.upper(),
        "name": data.name,
        "description": data.description,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.models.insert_one(model)
    del model["_id"]
    return serialize_doc(model)

# --- Brands CRUD ---
@api_router.get("/brands")
async def get_brands(buyer_id: Optional[str] = None):
    """Get all brands"""
    query = {}
    if buyer_id:
        query["buyer_id"] = buyer_id
    brands = await db.brands.find(query, {"_id": 0}).to_list(1000)
    return [serialize_doc(b) for b in brands]

@api_router.post("/brands")
async def create_brand(data: BrandCreate):
    """Create a new brand"""
    existing = await db.brands.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Brand with code {data.code} already exists")
    
    brand = {
        "id": str(uuid.uuid4()),
        "code": data.code.upper(),
        "name": data.name,
        "buyer_id": data.buyer_id,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.brands.insert_one(brand)
    del brand["_id"]
    return serialize_doc(brand)

# --- Buyers CRUD ---
@api_router.get("/buyers")
async def get_buyers():
    """Get all buyers"""
    buyers = await db.buyers.find({}, {"_id": 0}).to_list(1000)
    return [serialize_doc(b) for b in buyers]

@api_router.post("/buyers")
async def create_buyer(data: BuyerCreate):
    """Create a new buyer"""
    existing = await db.buyers.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Buyer with code {data.code} already exists")
    
    buyer = {
        "id": str(uuid.uuid4()),
        "code": data.code.upper(),
        "name": data.name,
        "country": data.country,
        "contact_email": data.contact_email,
        "payment_terms_days": data.payment_terms_days,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.buyers.insert_one(buyer)
    return serialize_doc(buyer)

@api_router.get("/buyers/{buyer_id}")
async def get_buyer(buyer_id: str):
    """Get buyer by ID"""
    buyer = await db.buyers.find_one({"id": buyer_id}, {"_id": 0})
    if not buyer:
        raise HTTPException(status_code=404, detail="Buyer not found")
    return serialize_doc(buyer)

# --- Branches CRUD ---
@api_router.get("/branches")
async def get_branches_list():
    """Get all branches from database"""
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    if not branches:
        # Return default branches if none exist
        return [{"name": b, "code": b.replace(" ", "_").upper()} for b in BRANCHES]
    return [serialize_doc(b) for b in branches]

@api_router.post("/branches/initialize")
async def initialize_branches():
    """Initialize branches collection from BRANCHES constant"""
    for branch_name in BRANCHES:
        existing = await db.branches.find_one({"name": branch_name})
        if not existing:
            await db.branches.insert_one({
                "id": str(uuid.uuid4()),
                "code": branch_name.replace(" ", "_").upper(),
                "name": branch_name,
                "location": "",
                "branch_type": "PRODUCTION",
                "capacity_units_per_day": 0,
                "is_active": True,
                "created_at": datetime.now(timezone.utc)
            })
    return {"message": f"Initialized {len(BRANCHES)} branches"}

# --- Forecasts CRUD ---
@api_router.get("/forecasts")
async def get_forecasts(
    buyer_id: Optional[str] = None,
    status: Optional[str] = None,
    forecast_month: Optional[str] = None
):
    """Get forecasts with filters"""
    query = {}
    if buyer_id:
        query["buyer_id"] = buyer_id
    if status:
        query["status"] = status
    if forecast_month:
        query["forecast_month"] = {"$regex": f"^{forecast_month}"}
    
    forecasts = await db.forecasts.find(query, {"_id": 0}).sort("forecast_month", -1).to_list(1000)
    return [serialize_doc(f) for f in forecasts]

@api_router.post("/forecasts")
async def create_forecast(data: ForecastCreate):
    """Create a new forecast"""
    # Generate forecast code
    count = await db.forecasts.count_documents({})
    forecast_code = f"FC_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    forecast = {
        "id": str(uuid.uuid4()),
        "forecast_code": forecast_code,
        "buyer_id": data.buyer_id,
        "vertical_id": data.vertical_id,
        "sku_id": data.sku_id,
        "forecast_month": data.forecast_month,
        "quantity": data.quantity,
        "priority": data.priority,
        "status": "DRAFT",
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc)
    }
    await db.forecasts.insert_one(forecast)
    return serialize_doc(forecast)

@api_router.put("/forecasts/{forecast_id}/confirm")
async def confirm_forecast(forecast_id: str):
    """Confirm a forecast"""
    result = await db.forecasts.update_one(
        {"id": forecast_id},
        {"$set": {"status": "CONFIRMED", "confirmed_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Forecast not found")
    return {"message": "Forecast confirmed"}

# --- Dispatch Lots CRUD ---
@api_router.get("/dispatch-lots")
async def get_dispatch_lots(
    buyer_id: Optional[str] = None,
    sku_id: Optional[str] = None,
    status: Optional[str] = None
):
    """Get dispatch lots with filters"""
    query = {}
    if buyer_id:
        query["buyer_id"] = buyer_id
    if sku_id:
        query["sku_id"] = sku_id
    if status:
        query["status"] = status
    
    lots = await db.dispatch_lots.find(query, {"_id": 0}).sort("target_date", 1).to_list(1000)
    return [serialize_doc(l) for l in lots]

@api_router.post("/dispatch-lots")
async def create_dispatch_lot(data: DispatchLotCreate):
    """Create a new dispatch lot"""
    count = await db.dispatch_lots.count_documents({})
    lot_code = f"DL_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    lot = {
        "id": str(uuid.uuid4()),
        "lot_code": lot_code,
        "forecast_id": data.forecast_id,
        "sku_id": data.sku_id,
        "buyer_id": data.buyer_id,
        "required_quantity": data.required_quantity,
        "produced_quantity": 0,
        "qc_passed_quantity": 0,
        "dispatched_quantity": 0,
        "target_date": data.target_date,
        "status": "CREATED",
        "priority": data.priority,
        "created_at": datetime.now(timezone.utc)
    }
    await db.dispatch_lots.insert_one(lot)
    return serialize_doc(lot)

@api_router.put("/dispatch-lots/{lot_id}/status")
async def update_dispatch_lot_status(lot_id: str, status: str):
    """Update dispatch lot status"""
    valid_statuses = ["CREATED", "PRODUCTION_ASSIGNED", "PARTIALLY_PRODUCED", "FULLY_PRODUCED", 
                     "QC_CLEARED", "DISPATCH_READY", "DISPATCHED", "DELIVERED"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    result = await db.dispatch_lots.update_one(
        {"id": lot_id},
        {"$set": {"status": status}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch lot not found")
    return {"message": "Status updated"}

# --- Production Batches ---
@api_router.get("/production-batches")
async def get_production_batches(
    branch: Optional[str] = None,
    sku_id: Optional[str] = None,
    status: Optional[str] = None,
    batch_date: Optional[str] = None
):
    """Get production batches with filters"""
    query = {}
    if branch:
        query["branch"] = branch
    if sku_id:
        query["sku_id"] = sku_id
    if status:
        query["status"] = status
    if batch_date:
        query["batch_date"] = {"$regex": f"^{batch_date}"}
    
    batches = await db.production_batches.find(query, {"_id": 0}).sort("batch_date", -1).to_list(1000)
    return [serialize_doc(b) for b in batches]

@api_router.post("/production-batches")
async def create_production_batch(data: ProductionBatchCreate):
    """Create a new production batch"""
    count = await db.production_batches.count_documents({})
    branch_code = data.branch.split()[0].upper() if data.branch else "XX"
    batch_code = f"PB_{branch_code}_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{count + 1:04d}"
    
    batch = {
        "id": str(uuid.uuid4()),
        "batch_code": batch_code,
        "production_plan_id": data.production_plan_id,
        "dispatch_lot_id": data.dispatch_lot_id,
        "branch_id": "",
        "branch": data.branch,
        "sku_id": data.sku_id,
        "planned_quantity": data.planned_quantity,
        "produced_quantity": 0,
        "good_quantity": 0,
        "rejected_quantity": 0,
        "batch_date": data.batch_date,
        "shift": data.shift,
        "status": "PLANNED",
        "created_at": datetime.now(timezone.utc)
    }
    await db.production_batches.insert_one(batch)
    return serialize_doc(batch)

@api_router.put("/production-batches/{batch_id}/start")
async def start_production_batch(batch_id: str):
    """Mark batch as in progress"""
    result = await db.production_batches.update_one(
        {"id": batch_id},
        {"$set": {"status": "IN_PROGRESS", "started_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    return {"message": "Batch started"}

@api_router.put("/production-batches/{batch_id}/complete")
async def complete_production_batch(batch_id: str, produced_quantity: int):
    """Mark batch as completed with produced quantity"""
    result = await db.production_batches.update_one(
        {"id": batch_id},
        {"$set": {
            "status": "COMPLETED",
            "produced_quantity": produced_quantity,
            "completed_at": datetime.now(timezone.utc)
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    return {"message": "Batch completed", "produced_quantity": produced_quantity}

# --- L2 Production Endpoint ---
@api_router.post("/production-batches/{batch_id}/produce-l2")
async def produce_l2_in_batch(batch_id: str, rm_id: str, quantity: int):
    """Produce L2 material within a production batch (triggers L1 consumption)"""
    batch = await db.production_batches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    result = await consume_l2_material(
        branch=batch["branch"],
        rm_id=rm_id,
        quantity=quantity,
        production_batch_id=batch_id,
        user_id="system"  # Would come from auth in real implementation
    )
    
    return result

# --- QC Checklists ---
@api_router.get("/qc-checklists")
async def get_qc_checklists(
    vertical_id: Optional[str] = None,
    model_id: Optional[str] = None,
    brand_id: Optional[str] = None
):
    """Get QC checklists - supports inheritance query"""
    # Build effective checklist query
    queries = []
    
    # Vertical level (base)
    if vertical_id:
        queries.append({"vertical_id": vertical_id, "model_id": None, "brand_id": None})
    
    # Model level
    if model_id:
        queries.append({"model_id": model_id, "brand_id": None})
    
    # Brand level
    if brand_id:
        queries.append({"brand_id": brand_id})
    
    # If no filters, get all
    if not queries:
        checklists = await db.qc_checklists.find({"status": "ACTIVE"}, {"_id": 0}).to_list(1000)
    else:
        checklists = await db.qc_checklists.find(
            {"$or": queries, "status": "ACTIVE"},
            {"_id": 0}
        ).sort("check_priority", 1).to_list(1000)
    
    return [serialize_doc(c) for c in checklists]

@api_router.post("/qc-checklists")
async def create_qc_checklist(data: QCChecklistCreate):
    """Create a new QC checklist item"""
    count = await db.qc_checklists.count_documents({})
    checklist_code = f"QC_{count + 1:04d}"
    
    checklist = {
        "id": str(uuid.uuid4()),
        "checklist_code": checklist_code,
        "name": data.name,
        "description": data.description,
        "check_type": data.check_type,
        "vertical_id": data.vertical_id,
        "model_id": data.model_id,
        "brand_id": data.brand_id,
        "expected_value": data.expected_value,
        "tolerance": data.tolerance,
        "is_mandatory": data.is_mandatory,
        "check_priority": data.check_priority,
        "status": "ACTIVE",
        "created_at": datetime.now(timezone.utc)
    }
    await db.qc_checklists.insert_one(checklist)
    return serialize_doc(checklist)

# --- QC Results ---
@api_router.get("/qc-results")
async def get_qc_results(production_batch_id: str):
    """Get QC results for a production batch"""
    results = await db.qc_results.find({"production_batch_id": production_batch_id}, {"_id": 0}).to_list(1000)
    return [serialize_doc(r) for r in results]

@api_router.post("/qc-results")
async def create_qc_result(data: QCResultCreate):
    """Record a QC inspection result"""
    count = await db.qc_results.count_documents({})
    result_code = f"QCR_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{count + 1:04d}"
    
    result_status = "PASSED" if data.failed_count == 0 else ("FAILED" if data.passed_count == 0 else "CONDITIONAL")
    
    result = {
        "id": str(uuid.uuid4()),
        "result_code": result_code,
        "production_batch_id": data.production_batch_id,
        "checklist_id": data.checklist_id,
        "sample_size": data.sample_size,
        "passed_count": data.passed_count,
        "failed_count": data.failed_count,
        "actual_value": data.actual_value,
        "result_status": result_status,
        "defect_type": data.defect_type,
        "defect_description": data.defect_description,
        "inspector_notes": data.inspector_notes,
        "inspected_at": datetime.now(timezone.utc),
        "inspected_by": "system"  # Would come from auth
    }
    await db.qc_results.insert_one(result)
    
    # Update batch status to QC_HOLD if not already
    await db.production_batches.update_one(
        {"id": data.production_batch_id, "status": "COMPLETED"},
        {"$set": {"status": "QC_HOLD"}}
    )
    
    return serialize_doc(result)

# --- QC Approval ---
@api_router.post("/qc-approvals")
async def create_qc_approval(
    production_batch_id: str,
    overall_status: str,
    approved_quantity: int = 0,
    rejection_reason: str = "",
    rework_instructions: str = ""
):
    """Create QC approval for a batch"""
    # Get all QC results for batch
    results = await db.qc_results.find({"production_batch_id": production_batch_id}, {"_id": 0}).to_list(1000)
    
    total_passed = sum(r.get("passed_count", 0) for r in results)
    total_failed = sum(r.get("failed_count", 0) for r in results)
    total_inspected = total_passed + total_failed
    
    approval = {
        "id": str(uuid.uuid4()),
        "production_batch_id": production_batch_id,
        "total_inspected": total_inspected,
        "total_passed": total_passed,
        "total_failed": total_failed,
        "overall_status": overall_status,
        "approved_quantity": approved_quantity,
        "rejection_reason": rejection_reason,
        "rework_instructions": rework_instructions,
        "approved_at": datetime.now(timezone.utc),
        "approved_by": "system"
    }
    await db.qc_approvals.insert_one(approval)
    
    # Update batch status
    new_status = "QC_PASSED" if overall_status == "APPROVED" else "QC_FAILED"
    await db.production_batches.update_one(
        {"id": production_batch_id},
        {"$set": {"status": new_status, "good_quantity": approved_quantity, "rejected_quantity": total_failed}}
    )
    
    # If approved, create FG inventory
    if overall_status == "APPROVED" and approved_quantity > 0:
        batch = await db.production_batches.find_one({"id": production_batch_id})
        if batch:
            fg_entry = {
                "id": str(uuid.uuid4()),
                "branch_id": batch.get("branch_id", ""),
                "branch": batch.get("branch"),
                "sku_id": batch.get("sku_id"),
                "dispatch_lot_id": batch.get("dispatch_lot_id"),
                "production_batch_id": production_batch_id,
                "quantity": approved_quantity,
                "status": "AVAILABLE",
                "qc_approval_id": approval["id"],
                "received_at": datetime.now(timezone.utc)
            }
            await db.fg_inventory.insert_one(fg_entry)
    
    return serialize_doc(approval)

# --- FG Inventory ---
@api_router.get("/fg-inventory")
async def get_fg_inventory(
    branch: Optional[str] = None,
    sku_id: Optional[str] = None,
    status: Optional[str] = None
):
    """Get finished goods inventory"""
    query = {}
    if branch:
        query["branch"] = branch
    if sku_id:
        query["sku_id"] = sku_id
    if status:
        query["status"] = status
    
    inventory = await db.fg_inventory.find(query, {"_id": 0}).to_list(10000)
    return [serialize_doc(i) for i in inventory]

@api_router.get("/fg-inventory/summary")
async def get_fg_inventory_summary(branch: Optional[str] = None):
    """Get FG inventory summary by SKU"""
    match_stage = {"status": "AVAILABLE"}
    if branch:
        match_stage["branch"] = branch
    
    pipeline = [
        {"$match": match_stage},
        {"$group": {
            "_id": {"branch": "$branch", "sku_id": "$sku_id"},
            "total_quantity": {"$sum": "$quantity"}
        }},
        {"$sort": {"_id.branch": 1, "_id.sku_id": 1}}
    ]
    
    results = await db.fg_inventory.aggregate(pipeline).to_list(10000)
    return [{"branch": r["_id"]["branch"], "sku_id": r["_id"]["sku_id"], "total_quantity": r["total_quantity"]} for r in results]

# --- Purchase Orders ---
@api_router.get("/purchase-orders")
async def get_purchase_orders(
    vendor_id: Optional[str] = None,
    branch: Optional[str] = None,
    status: Optional[str] = None
):
    """Get purchase orders"""
    query = {}
    if vendor_id:
        query["vendor_id"] = vendor_id
    if branch:
        query["branch"] = branch
    if status:
        query["status"] = status
    
    orders = await db.purchase_orders.find(query, {"_id": 0}).sort("order_date", -1).to_list(1000)
    return [serialize_doc(o) for o in orders]

@api_router.post("/purchase-orders")
async def create_purchase_order(data: PurchaseOrderCreate):
    """Create a new purchase order"""
    count = await db.purchase_orders.count_documents({})
    po_number = f"PO_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    po = {
        "id": str(uuid.uuid4()),
        "po_number": po_number,
        "vendor_id": data.vendor_id,
        "branch_id": "",
        "branch": data.branch,
        "order_date": data.order_date,
        "expected_delivery_date": data.expected_delivery_date,
        "total_amount": 0,
        "currency": "INR",
        "status": "DRAFT",
        "payment_status": "PENDING",
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc)
    }
    await db.purchase_orders.insert_one(po)
    return serialize_doc(po)

@api_router.post("/purchase-orders/{po_id}/lines")
async def add_po_line(po_id: str, data: PurchaseOrderLineCreate):
    """Add line item to a PO"""
    line = {
        "id": str(uuid.uuid4()),
        "po_id": po_id,
        "rm_id": data.rm_id,
        "quantity_ordered": data.quantity_ordered,
        "quantity_received": 0,
        "unit_price": data.unit_price,
        "unit_of_measure": data.unit_of_measure,
        "line_total": data.quantity_ordered * data.unit_price,
        "status": "PENDING"
    }
    await db.purchase_order_lines.insert_one(line)
    
    # Update PO total
    lines = await db.purchase_order_lines.find({"po_id": po_id}, {"_id": 0}).to_list(1000)
    total = sum(l.get("line_total", 0) for l in lines)
    await db.purchase_orders.update_one({"id": po_id}, {"$set": {"total_amount": total}})
    
    return serialize_doc(line)

@api_router.get("/purchase-orders/{po_id}/lines")
async def get_po_lines(po_id: str):
    """Get all lines for a PO"""
    lines = await db.purchase_order_lines.find({"po_id": po_id}, {"_id": 0}).to_list(1000)
    return [serialize_doc(l) for l in lines]

# --- RM Stock Movements ---
@api_router.get("/rm-stock-movements")
async def get_rm_stock_movements(
    rm_id: Optional[str] = None,
    branch: Optional[str] = None,
    movement_type: Optional[str] = None,
    limit: int = 100
):
    """Get RM stock movements (audit trail)"""
    query = {}
    if rm_id:
        query["rm_id"] = rm_id
    if branch:
        query["branch"] = branch
    if movement_type:
        query["movement_type"] = movement_type
    
    movements = await db.rm_stock_movements.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return [serialize_doc(m) for m in movements]

# --- IBT Transfers ---
@api_router.get("/ibt-transfers")
async def get_ibt_transfers(
    source_branch: Optional[str] = None,
    destination_branch: Optional[str] = None,
    status: Optional[str] = None,
    transfer_type: Optional[str] = None
):
    """Get inter-branch transfers"""
    query = {}
    if source_branch:
        query["source_branch"] = source_branch
    if destination_branch:
        query["destination_branch"] = destination_branch
    if status:
        query["status"] = status
    if transfer_type:
        query["transfer_type"] = transfer_type
    
    transfers = await db.ibt_transfers.find(query, {"_id": 0}).sort("initiated_at", -1).to_list(1000)
    return [serialize_doc(t) for t in transfers]

@api_router.post("/ibt-transfers")
async def create_ibt_transfer(
    transfer_type: str,
    source_branch: str,
    destination_branch: str,
    item_id: str,
    quantity: float,
    unit_of_measure: str = "",
    notes: str = ""
):
    """Create a new inter-branch transfer"""
    if source_branch == destination_branch:
        raise HTTPException(status_code=400, detail="Source and destination branches must be different")
    
    count = await db.ibt_transfers.count_documents({})
    transfer_code = f"IBT_{datetime.now(timezone.utc).strftime('%Y%m')}_{count + 1:04d}"
    
    transfer = {
        "id": str(uuid.uuid4()),
        "transfer_code": transfer_code,
        "transfer_type": transfer_type,
        "source_branch_id": "",
        "source_branch": source_branch,
        "destination_branch_id": "",
        "destination_branch": destination_branch,
        "item_id": item_id,
        "quantity": quantity,
        "unit_of_measure": unit_of_measure,
        "status": "INITIATED",
        "initiated_at": datetime.now(timezone.utc),
        "notes": notes
    }
    await db.ibt_transfers.insert_one(transfer)
    return serialize_doc(transfer)

@api_router.put("/ibt-transfers/{transfer_id}/approve")
async def approve_ibt_transfer(transfer_id: str):
    """Approve an IBT transfer"""
    result = await db.ibt_transfers.update_one(
        {"id": transfer_id, "status": "INITIATED"},
        {"$set": {"status": "APPROVED", "approved_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transfer not found or already processed")
    return {"message": "Transfer approved"}

@api_router.put("/ibt-transfers/{transfer_id}/ship")
async def ship_ibt_transfer(transfer_id: str):
    """Mark IBT as shipped/in-transit"""
    transfer = await db.ibt_transfers.find_one({"id": transfer_id})
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    
    # Deduct from source
    if transfer["transfer_type"] == "RM":
        await update_branch_rm_inventory(transfer["source_branch"], transfer["item_id"], -transfer["quantity"])
    else:  # FG
        # Update FG inventory status
        await db.fg_inventory.update_many(
            {"branch": transfer["source_branch"], "sku_id": transfer["item_id"], "status": "AVAILABLE"},
            {"$set": {"status": "IN_TRANSIT"}}
        )
    
    await db.ibt_transfers.update_one(
        {"id": transfer_id},
        {"$set": {"status": "IN_TRANSIT", "shipped_at": datetime.now(timezone.utc)}}
    )
    return {"message": "Transfer shipped"}

@api_router.put("/ibt-transfers/{transfer_id}/receive")
async def receive_ibt_transfer(transfer_id: str):
    """Mark IBT as received"""
    transfer = await db.ibt_transfers.find_one({"id": transfer_id})
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    
    # Add to destination
    if transfer["transfer_type"] == "RM":
        await update_branch_rm_inventory(transfer["destination_branch"], transfer["item_id"], transfer["quantity"])
    else:  # FG
        # Create new FG inventory at destination
        fg_entry = {
            "id": str(uuid.uuid4()),
            "branch_id": "",
            "branch": transfer["destination_branch"],
            "sku_id": transfer["item_id"],
            "quantity": int(transfer["quantity"]),
            "status": "AVAILABLE",
            "received_at": datetime.now(timezone.utc)
        }
        await db.fg_inventory.insert_one(fg_entry)
    
    await db.ibt_transfers.update_one(
        {"id": transfer_id},
        {"$set": {"status": "COMPLETED", "received_at": datetime.now(timezone.utc)}}
    )
    return {"message": "Transfer received and completed"}

# --- Price History ---
@api_router.get("/price-history")
async def get_price_history(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    limit: int = 100
):
    """Get price change history"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    
    history = await db.price_history.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return [serialize_doc(h) for h in history]

# --- Audit Logs ---
@api_router.get("/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 100
):
    """Get audit logs"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    if action:
        query["action"] = action
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return [serialize_doc(l) for l in logs]

# --- Initialize Predefined Powder Coating RMs ---
@api_router.post("/setup/initialize-powder-coatings")
async def initialize_powder_coatings():
    """Initialize predefined powder coating L1 RMs for INM"""
    powder_coatings = [
        {"rm_id": "INM_PC_001", "category": "INM", "rm_level": "L1", "category_data": {"type": "Powder Coating", "color": "Black", "finish": "Matte", "unit": "KG"}},
        {"rm_id": "INM_PC_002", "category": "INM", "rm_level": "L1", "category_data": {"type": "Powder Coating", "color": "White", "finish": "Gloss", "unit": "KG"}},
        {"rm_id": "INM_PC_003", "category": "INM", "rm_level": "L1", "category_data": {"type": "Powder Coating", "color": "Red", "finish": "Matte", "unit": "KG"}},
        {"rm_id": "INM_PC_004", "category": "INM", "rm_level": "L1", "category_data": {"type": "Powder Coating", "color": "Blue", "finish": "Gloss", "unit": "KG"}},
        {"rm_id": "INM_PC_005", "category": "INM", "rm_level": "L1", "category_data": {"type": "Powder Coating", "color": "Silver", "finish": "Metallic", "unit": "KG"}},
        {"rm_id": "INM_PC_006", "category": "INM", "rm_level": "L1", "category_data": {"type": "Powder Coating", "color": "Custom", "finish": "Various", "unit": "KG"}},
    ]
    
    created = 0
    for pc in powder_coatings:
        existing = await db.raw_materials.find_one({"rm_id": pc["rm_id"]})
        if not existing:
            pc["id"] = str(uuid.uuid4())
            pc["low_stock_threshold"] = 10.0
            pc["status"] = "ACTIVE"
            pc["created_at"] = datetime.now(timezone.utc)
            await db.raw_materials.insert_one(pc)
            created += 1
    
    return {"message": f"Initialized {created} powder coating RMs"}

@api_router.get("/skus/filtered")
async def get_filtered_skus(
    vertical: Optional[str] = None,
    model: Optional[str] = None,
    brand: Optional[str] = None,
    search: Optional[str] = None,
    branch: Optional[str] = None
):
    """Get SKUs filtered by vertical, model, brand, and optionally branch subscription"""
    query = {}
    
    if vertical:
        query["vertical"] = vertical
    if model:
        query["model"] = model
    if brand:
        query["brand"] = brand
    if search:
        query["$or"] = [
            {"sku_id": {"$regex": search, "$options": "i"}},
            {"bidso_sku": {"$regex": search, "$options": "i"}},
            {"buyer_sku_id": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    # If branch is specified, only return SKUs subscribed to that branch
    if branch:
        # Get subscribed SKUs from sku_branch_assignments
        assignments = await db.sku_branch_assignments.find({"branch": branch}, {"_id": 0}).to_list(10000)
        subscribed_sku_ids = [a['sku_id'] for a in assignments]
        
        if not subscribed_sku_ids:
            return []
        
        if query:
            query = {"$and": [query, {"sku_id": {"$in": subscribed_sku_ids}}]}
        else:
            query = {"sku_id": {"$in": subscribed_sku_ids}}
        
        skus = await db.skus.find(query, {"_id": 0}).to_list(10000)
        
        # Get inventory data for stock levels
        branch_inventories = await db.branch_sku_inventory.find(
            {"branch": branch, "sku_id": {"$in": subscribed_sku_ids}}, 
            {"_id": 0}
        ).to_list(10000)
        inv_map = {inv['sku_id']: inv for inv in branch_inventories}
        
        result = []
        for sku in skus:
            inv = inv_map.get(sku['sku_id'])
            sku['current_stock'] = inv['current_stock'] if inv else 0
            sku['branch'] = branch
            result.append(serialize_doc(sku))
        return result
    else:
        skus = await db.skus.find(query, {"_id": 0}).to_list(10000)
        return [serialize_doc(s) for s in skus]

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()